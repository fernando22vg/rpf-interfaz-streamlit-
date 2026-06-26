"""
bloque_dsl_params.py
--------------------
Bloque 02 de la interfaz RPF — Gestión de Parámetros DSL Andritz.

Tabs:
  1. Parámetros Actuales  — visualización por familia/unidad, selección de relevantes
  2. Experimentos         — variación de valores, generación de Excel staging, registro
  3. Historial            — comparación de experimentos, exportación para IA

El Excel comparativo ({SYM}_Comparacion_Final.xlsx) es SOLO LECTURA.
Toda escritura va al Excel de staging ({SYM}_EXP_*.xlsx) y a PostgreSQL.
"""

from __future__ import annotations
import json
import os
import re
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st

import kpi_calc

#  Rutas y familias 
BASE_DATOS = Path(r"C:\Datos Cobee\03_DATOS GEN")

CONFIG_PATH = Path(__file__).parent / "dsl_config.json"

# sym → (familia, subdir)
UNIDADES_DSL: dict[str, tuple[str, str]] = {
    "ZON01": ("F1", "01_ZON"),
    "TIQ01": ("F1", "02_TIQ"),
    "CUT05": ("F1", "04_CUT/CUT05"),
    "CHU01": ("F1", "07_CHU"),
    "CHU02": ("F1", "07_CHU"),
    "HAR01": ("F1", "08_HAR"),
    "HAR02": ("F1", "08_HAR"),
    "CAH01": ("F1", "09_CAH"),
    "CAH02": ("F1", "09_CAH"),
    "HUA01": ("F1", "10_HUA"),
    "HUA02": ("F1", "10_HUA"),
    "BOT01": ("F2", "03_BOT/BOT01"),
    "BOT02": ("F2", "03_BOT/BOT02"),
    "BOT03": ("F2", "03_BOT/BOT03"),
    "CUT01": ("F2", "04_CUT/CUT01"),
    "CUT02": ("F2", "04_CUT/CUT02"),
    "CUT03": ("F2", "04_CUT/CUT03"),
    "CUT04": ("F2", "04_CUT/CUT04"),
    "ANG03": ("F2", "11_ANG03"),
    "CRB01": ("F2", "12_CRB"),   # Excel: CRB_Comparacion_Final.xlsx (stem diferente)
    "SRO01": ("F3", "05_SRO/SRO01"),
    "SRO02": ("F3", "05_SRO/SRO02"),
    "SAI01": ("F3", "06_SAI"),
}

FAMILIAS = {
    "F1": "Andritz HIPASE T/E",
    "F2": "Reivax RTVX 1000",
    "F3": "VA TECH / Andritz TM1703",
}

# Nombres de Excel comparativo no estándar (cuando no sigue el patrón {SYM}_Comparacion_Final.xlsx)
_EXCEL_NOMBRES: dict[str, str] = {
    "SRO02": "SR02_Comparativa.xlsx",
    "SAI01": "SAI01_Comparacion.xlsx",
}

# Columnas del Excel comparativo (base 1)
_COL_BLOQUE  = 1   # A
_COL_SIM_SYM = 4   # D — Símbolo
_COL_DESC    = 5   # E
_COL_UNIDAD  = 6   # F — Unidad física (s, pu, etc.)
_COL_G       = 7   # Valor Diseño MOD
_COL_H       = 8   # Fuente MOD
_COL_I       = 9   # Valor PES
_COL_J       = 10  # Fuente PES
_COL_K       = 11  # Valor HMI
_COL_L       = 12  # Fuente HMI
_COL_R       = 18  # Valor Simulación (fórmula Excel: HMI>PES>MOD — SOLO LECTURA)
_FILA_INI    = 6

#  Mapeo col A (nombre de sección Excel) → slot DSL de PowerFactory 
# Los slots corresponden exactamente a los de DSL_SLOTS en ExtractorDSL_Andritz.py
_BLOQUE_A_SLOT: dict[str, str] = {
    # Gobernador
    "Gobernador (GOV)":                          "GOV",
    "pcu_Andritz (Gobernador de velocidad)":     "GOV",
    "General":                                   "GOV",
    # Turbina hidráulica
    "Turbina (TUR)":                             "TUR",
    "Servomotor":                                "TUR",
    "Tubería forzada":                           "TUR",
    "Túnel y almenara":                          "TUR",
    "Embalse":                                   "TUR",
    # Regulador de tensión
    "Regulador de Tensión (AVR)":                "AVR",
    "Configuración AVR":                         "AVR",
    "VREG — Regulador de tensión":               "AVR",
    "FCR — Regulador corriente campo":           "AVR",
    "VAR — Regulador de reactiva":               "AVR",
    # Parte de potencia / excitador
    "Parte de Potencia (PPT)":                   "PPT",
    "Power Part (THYNE 500/600)":                "PPT",
    "Configuración modelo":                      "PPT",
    "Excitatriz Estático/Rotatorio (ROT)":       "ROT",
    "Static/Rotating Exciter":                   "ROT",
    "Normalizacion Exitatriz (REN)":             "REN",
    "Renormalization":                           "REN",
    # Estabilizador PSS4B
    "Estabilizador de Potencia (IEEE PSS4B)":    "PSS4B",
    "Banda Alta (2–4 Hz)":                  "PSS4B",
    "Banda Baja (0.1–0.5 Hz)":             "PSS4B",
    "Banda Media (0.5–2 Hz)":              "PSS4B",
    "Filtros notch":                             "PSS4B",
    "Límites salida PSS4B":                      "PSS4B",
    # Limitadores
    "Limiters":                                  "Limiters",
    "FCL — Limitador campo instantáneo":         "Limiters",
    "OCL — Limitador corriente estátor (thl_IT)":"Limiters",
    "OEL — Limitador sobreexcitación retardado (thl_ifd)": "Limiters",
    "UEL — Limitador subexcitación (curva PQ)":  "Limiters",
    "UEL — Limitador subexcitación (ángulo carga)":"Limiters",
    "Limitador V/Hz (fx)":                       "Limiters",
    "Límites salida limitadores":                "Limiters",
    # Generación de señales / medición
    "Generación de señales":                     "Signal Gen",
    "Signal Generation":                         "Signal Gen",
    # Nominales (no son parámetros de modelo, solo referencia)
    "GEN. NOMINALES":                            "GEN NOM",
    "Generador":                                 "GEN NOM",
    "Generación de señales":                     "Signal Gen",

    #  F2: Reivax RTVX 1000 
    "GOBERNADOR":                                "GOV",
    "Control velocidad":                         "GOV",
    "Control posición":                          "POS_SIMP",
    "POSICAO":                                   "POS_SIMP",
    "Límites posición":                          "POS_SIMP",
    "Límites compuerta":                         "POS_SIMP",
    "Límites integrador":                        "POS_SIMP",
    "Parámetros servo":                          "ATUADOR_SIMP",
    "ACTUADOR":                                  "ATUADOR_SIMP",
    "Tasas":                                     "ATUADOR_SIMP",
    "CONDUCTO Y TURBINA":                        "CONDUTO_TURBINA",
    "Datos conducto":                            "CONDUTO_TURBINA",
    "Curva Px":                                  "CONDUTO_TURBINA",
    "Curva YAYD":                                "CONDUTO_TURBINA",
    "Curva conjugación":                         "CONDUTO_TURBINA",
    "Curva Array K":                             "CONDUTO_TURBINA",
    "AVR":                                       "AVR",
    "Control tensión":                           "AVR",
    "Ganancias":                                 "AVR",
    "Referencias":                               "AVR",
    "Filtros":                                   "AVR",
    "Lead-lag":                                  "AVR",
    "Zona Kred":                                 "AVR",
    "Drive excitación":                          "DRIVE",
    "DRIVE":                                     "DRIVE",
    "Excitatriz":                                "EXCITATRIZ",
    "Curva de saturación":                       "EXCITATRIZ",
    "PSS slot (PSS2C)":                          "PSS_COMP",
    "Limitador UEL":                             "UEL",
    "Tabla 95 — Curva PQ":                       "UEL",
    "vco slot (MEL)":                            "MEL",
    "Limitador MEL":                             "MEL",
    "Limitador OEL":                             "OEL",
    "Limitador SCL":                             "SCL",
    "Limitador V/Hz":                            "VHZL",
    "Limitador VHZ":                             "VHZL",
    "Bloqueo/reset":                             "AVR",
    "Límites y lógica":                          "AVR",
}

def _slot_dsl(bloque_excel: str) -> str:
    """Convierte nombre de sección Excel → slot DSL de PowerFactory."""
    slot = _BLOQUE_A_SLOT.get(bloque_excel)
    if slot:
        return slot
    # Fallback: detectar por palabras clave (Andritz y Reivax)
    bu = bloque_excel.upper()
    if "GOB" in bu or "GOV" in bu or "PCU" in bu or "VELOC" in bu:  return "GOV"
    if "ACTUAD" in bu or "ATUAD" in bu or "SERVO" in bu:            return "ATUADOR_SIMP"
    if "POSIC" in bu or "POSAO" in bu:                               return "POS_SIMP"
    if "CONDUT" in bu or "TURB" in bu or "TUBA" in bu:              return "CONDUTO_TURBINA"
    if "AVR" in bu or "VREG" in bu or "TENS" in bu or "REG_T" in bu: return "AVR"
    if "DRIVE" in bu:                                                 return "DRIVE"
    if "EXCIT" in bu:                                                 return "EXCITATRIZ"
    if "PSS" in bu or "BANDA" in bu or "ESTAB" in bu:               return "PSS4B"
    if "PPT" in bu or "THYNE" in bu:                                 return "PPT"
    if "ROT" in bu:                                                   return "ROT"
    if "REN" in bu or "NORM" in bu:                                  return "REN"
    if "MEL" in bu:                                                   return "MEL"
    if "SCL" in bu:                                                   return "SCL"
    if "VHZ" in bu or "V/HZ" in bu:                                  return "VHZL"
    if "LIMIT" in bu or "OEL" in bu or "UEL" in bu or "OCL" in bu or "FCL" in bu:
        return "Limiters"
    if "SIGNAL" in bu or "GEN.SE" in bu:                             return "Signal Gen"
    return bloque_excel   # sin mapeo conocido: devolver tal cual


#  Helpers de archivos 

def _excel_path(sym: str) -> Path | None:
    """Devuelve la ruta al Excel comparativo o None si no existe."""
    if sym not in UNIDADES_DSL:
        return None
    _, subdir = UNIDADES_DSL[sym]
    # Nombre de archivo personalizado (ej. SR02_Comparativa.xlsx, SAI01_Comparacion.xlsx)
    if sym in _EXCEL_NOMBRES:
        p = BASE_DATOS / subdir / _EXCEL_NOMBRES[sym]
        return p if p.exists() else None
    # Caso normal: {SYM}_Comparacion_Final.xlsx
    p = BASE_DATOS / subdir / f"{sym}_Comparacion_Final.xlsx"
    if p.exists():
        return p
    # Caso especial: stem sin número de unidad (ej. CRB_Comparacion_Final.xlsx)
    stem = re.sub(r'\d+$', '', sym)   # CRB01 → CRB
    p2 = BASE_DATOS / subdir / f"{stem}_Comparacion_Final.xlsx"
    return p2 if p2.exists() else None


def _exp_dir(sym: str) -> Path:
    """Carpeta de staging para experimentos de la unidad."""
    _, subdir = UNIDADES_DSL[sym]
    d = BASE_DATOS / subdir / "Experimentos"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _siguiente_correlativo(sym: str) -> int:
    """Retorna el siguiente número correlativo de experimento para la unidad."""
    d = _exp_dir(sym)
    nums = []
    for f in d.glob(f"{sym}_EXP_*"):
        m = re.search(r"_EXP_\d{8}_(\d{3})", f.name)
        if m:
            nums.append(int(m.group(1)))
    return max(nums, default=0) + 1


def _nombre_experimento(sym: str, sufijo: str = "") -> str:
    n = _siguiente_correlativo(sym)
    hoy = datetime.now().strftime("%Y%m%d")
    base = f"{sym}_EXP_{hoy}_{n:03d}"
    return f"{base}_{sufijo.strip()}" if sufijo.strip() else base


#  Configuración persistente (dsl_config.json) 

def _cargar_config() -> dict:
    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _guardar_config(cfg: dict):
    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def _cfg_unidad(cfg: dict, sym: str) -> dict:
    return cfg.get(sym, {})


#  Lectura del Excel comparativo 

def _es_numero(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _leer_comparativo(sym: str) -> pd.DataFrame:
    """
    Lee el Excel comparativo y devuelve un DataFrame con columnas:
      bloque, simbolo, descripcion, unidad,
      val_dis, fuen_dis, val_pes, fuen_pes, val_hmi, fuen_hmi,
      val_actual (col R)
    Solo filas con símbolo en col D y al menos un valor numérico en G/I/K.
    """
    try:
        import openpyxl
    except ImportError:
        st.error("Instalar openpyxl: pip install openpyxl")
        return pd.DataFrame()

    path = _excel_path(sym)
    if path is None:
        return pd.DataFrame()

    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    filas = []
    bloque_excel  = ""   # nombre de sección tal como está en col A
    slot_actual   = ""   # slot DSL de PF correspondiente

    for row in range(_FILA_INI, ws.max_row + 1):
        a_val = ws.cell(row, _COL_BLOQUE).value
        if a_val and ws.cell(row, _COL_SIM_SYM).value is None:
            bloque_excel = str(a_val).strip()
            slot_actual  = _slot_dsl(bloque_excel)
            continue

        sym_val = ws.cell(row, _COL_SIM_SYM).value
        if sym_val is None:
            continue
        simbolo = str(sym_val).strip()
        if not simbolo:
            continue

        g = ws.cell(row, _COL_G).value
        i = ws.cell(row, _COL_I).value
        k = ws.cell(row, _COL_K).value

        g_num = g if _es_numero(g) else None
        i_num = i if _es_numero(i) else None
        k_num = k if _es_numero(k) else None

        # Ocultar filas donde ninguna fuente tiene valor numérico
        if g_num is None and i_num is None and k_num is None:
            continue

        # Col R tiene fórmula Excel (HMI > PES > MOD).
        # data_only=True devuelve el valor cacheado por Excel; si es None
        # (archivo guardado por openpyxl sin recalcular) lo calculamos aquí.
        r_cached = ws.cell(row, _COL_R).value
        if _es_numero(r_cached):
            val_actual = r_cached
        else:
            # Replicar fórmula: prioridad K > I > G
            val_actual = k_num if k_num is not None else (
                         i_num if i_num is not None else g_num)

        filas.append({
            "slot_dsl":   slot_actual,
            "bloque":     bloque_excel,
            "simbolo":    simbolo,
            "descripcion": str(ws.cell(row, _COL_DESC).value or "").strip(),
            "unidad":     str(ws.cell(row, _COL_UNIDAD).value or "—").strip(),
            "val_dis":    g_num,
            "fuen_dis":   str(ws.cell(row, _COL_H).value or "").strip(),
            "val_pes":    i_num,
            "fuen_pes":   str(ws.cell(row, _COL_J).value or "").strip(),
            "val_hmi":    k_num,
            "fuen_hmi":   str(ws.cell(row, _COL_L).value or "").strip(),
            "val_actual": val_actual,
        })

    wb.close()
    return pd.DataFrame(filas)


#  Generación del Excel de staging 

def _generar_excel_experimento(sym: str, nombre: str,
                                params_exp: list[dict]) -> Path:
    """
    Crea {sym}_EXP_*.xlsx con los parámetros del experimento.
    Columnas: Bloque | Símbolo | Descripción | Valor_Base | Valor_Experimento
    Devuelve la ruta del archivo creado.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("Instalar openpyxl")

    ruta = _exp_dir(sym) / f"{nombre}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parámetros"

    # Cabecera
    cabeceras = ["Bloque", "Símbolo", "Descripción", "Valor_Base", "Valor_Experimento"]
    ws.append(cabeceras)
    fill_hdr = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for cell in ws[1]:
        cell.fill = fill_hdr
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Datos
    fill_par = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
    fill_imp = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for i, p in enumerate(params_exp):
        ws.append([
            p.get("bloque", ""),
            p["simbolo"],
            p.get("descripcion", ""),
            p.get("valor_base"),
            p["valor"],
        ])
        f = fill_par if i % 2 == 0 else fill_imp
        for cell in ws[ws.max_row]:
            cell.fill = f

    # Ancho de columnas
    for col in ws.columns:
        ancho = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 50)

    # Hoja de metadata
    ws2 = wb.create_sheet("Metadata")
    ws2.append(["Campo", "Valor"])
    ws2.append(["Unidad", sym])
    ws2.append(["Nombre experimento", nombre])
    ws2.append(["Fecha generación", datetime.now().isoformat()])
    ws2.append(["Total parámetros", len(params_exp)])

    wb.save(str(ruta))
    return ruta


#  Lectura de curvas exportadas por DatosCurvas_v3.py 

@st.cache_data(show_spinner=False)
def _leer_curva_experimento(path: str, t_falla: float):
    """Lee un xlsx exportado por DatosCurvas_v3.py (col. tiempo/frecuencia/potencia).

    Misma lógica que _cached_sim_arrays de interfaz_analisis_RPF.py, usando
    kpi_calc._robust_col_detect para no duplicar la heurística de columnas.
    Devuelve (ts_aligned, fs_hz, ps_mw).
    """
    df = pd.read_excel(path, engine="calamine").dropna()
    tc, fc, pc = kpi_calc._robust_col_detect(df)
    ts_raw = pd.to_numeric(df[tc], errors="coerce").values
    fs_raw = pd.to_numeric(df[fc], errors="coerce").ffill().values
    fs_hz  = fs_raw * 50.0 if np.nanmax(fs_raw) < 2.0 else fs_raw
    ps_mw  = pd.to_numeric(df[pc], errors="coerce").ffill().values
    valid  = ~np.isnan(ts_raw)
    ts_raw, fs_hz, ps_mw = ts_raw[valid], fs_hz[valid], ps_mw[valid]
    return ts_raw - t_falla, fs_hz, ps_mw


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — Parámetros Actuales
# ══════════════════════════════════════════════════════════════════════════════

def _tab_parametros_actuales():
    st.caption(
        "Visualización de los parámetros DSL por familia de gobernador y unidad. "
        "Fuente: Excel comparativo (solo lectura). "
        "Define aquí cuáles son relevantes y sus rangos de variación para experimentos."
    )

    cfg = _cargar_config()

    #  Selector de familia 
    familia_sel = st.radio(
        "Familia de gobernador",
        options=list(FAMILIAS.keys()),
        format_func=lambda k: f"{k} — {FAMILIAS[k]}",
        horizontal=True,
        key="dsl_familia",
    )

    unidades_familia = [s for s, (f, _) in UNIDADES_DSL.items() if f == familia_sel]

    #  Selector de unidad 
    col_u, col_info = st.columns([3, 2])
    with col_u:
        sym = st.selectbox(
            "Unidad",
            options=unidades_familia,
            key="dsl_sym_tab1",
        )

    path = _excel_path(sym)
    with col_info:
        if path:
            st.success(f"✅ {path.name}", icon=None)
        else:
            st.error(f"⚠ Excel comparativo no encontrado para {sym}")
            return

    #  Cargar datos (invalida caché si el schema no coincide) 
    cache_key = f"dsl_df_{sym}"
    cached = st.session_state.get(cache_key)
    if cached is None or "slot_dsl" not in cached.columns:
        with st.spinner(f"Leyendo {path.name}…"):
            st.session_state[cache_key] = _leer_comparativo(sym)

    df: pd.DataFrame = st.session_state[cache_key]
    if df.empty:
        st.warning("No se encontraron parámetros con valores numéricos.")
        return

    cfg_u = _cfg_unidad(cfg, sym)

    #  Filtro por slot DSL (composite model) 
    # Orden canónico de slots DSL (igual que en ExtractorDSL / CargadorDSL)
    # F1 Andritz + F2 Reivax — slots en orden canónico
    _ORDEN_SLOTS = [
        "GOV", "TUR", "AVR", "PPT", "ROT", "REN", "PSS4B", "Limiters",
        "ATUADOR_SIMP", "POS_SIMP", "CONDUTO_TURBINA", "VHZL",
        "UEL", "OEL", "MEL", "SCL", "DRIVE", "EXCITATRIZ", "PSS_COMP",
        "Signal Gen",
    ]
    # GEN NOM son datos de referencia fijos — no se muestran ni modifican
    df = df[df["slot_dsl"] != "GEN NOM"].copy()

    slots_presentes = [s for s in _ORDEN_SLOTS
                       if s in df["slot_dsl"].values]
    slots_otros = sorted(
        set(df["slot_dsl"].dropna().unique()) - set(_ORDEN_SLOTS)
    )
    slots_opciones = ["(todos)"] + slots_presentes + slots_otros

    if st.session_state.get("dsl_bloque_filtro") not in slots_opciones:
        st.session_state.pop("dsl_bloque_filtro", None)

    slot_filtro = st.pills(
        "Slot DSL (composite model)",
        options=slots_opciones,
        default="(todos)",
        key="dsl_bloque_filtro",
    )
    df_vis = df if slot_filtro == "(todos)" else df[df["slot_dsl"] == slot_filtro]

    st.divider()

    #  Tabla interactiva 
    st.markdown(
        f"**{len(df_vis)} parámetros** — "
        f"col R (Parámetros Actuales) es el punto de partida para experimentos."
    )

    filas_edit = []
    for _, row in df_vis.iterrows():
        sym_p = row["simbolo"]
        p_cfg = cfg_u.get(sym_p, {})
        filas_edit.append({
            "Slot DSL":     row["slot_dsl"],
            "Símbolo":      sym_p,
            "Descripción":  row["descripcion"],
            "Unidad":       row["unidad"],
            "Diseño (G)":   row["val_dis"],
            "PES (I)":      row["val_pes"],
            "HMI (K)":      row["val_hmi"],
            "Actual (R)":   row["val_actual"],
            "Min":          p_cfg.get("min", None),
            "Max":          p_cfg.get("max", None),
            "Editable":     p_cfg.get("relevante", False),
        })

    df_edit = pd.DataFrame(filas_edit)

    edited = st.data_editor(
        df_edit,
        column_config={
            "Slot DSL":    st.column_config.TextColumn(disabled=True, width="small"),
            "Símbolo":     st.column_config.TextColumn(disabled=True, width="small"),
            "Descripción": st.column_config.TextColumn(disabled=True, width="medium"),
            "Unidad":      st.column_config.TextColumn(disabled=True, width="small"),
            "Diseño (G)":  st.column_config.NumberColumn(disabled=True, format="%.4g"),
            "PES (I)":     st.column_config.NumberColumn(disabled=True, format="%.4g"),
            "HMI (K)":     st.column_config.NumberColumn(disabled=True, format="%.4g"),
            "Actual (R)":  st.column_config.NumberColumn(disabled=True, format="%.4g",
                               help="Valor Simulación actual — punto de partida"),
            "Min":         st.column_config.NumberColumn(format="%.4g",
                               help="Límite inferior para variación en experimentos"),
            "Max":         st.column_config.NumberColumn(format="%.4g",
                               help="Límite superior para variación en experimentos"),
            "Editable":    st.column_config.CheckboxColumn(
                               help="Marcar para incluir en experimentos de optimización"),
        },
        hide_index=True,
        use_container_width=True,
        key=f"dsl_editor_{sym}",
        num_rows="fixed",
    )

    #  Guardar selección
    n_edit = int(edited["Editable"].sum())
    familia_sym, _ = UNIDADES_DSL[sym]
    propaga_familia = familia_sym in ("F1", "F2")

    col_g1, col_g2 = st.columns([2, 1])
    with col_g1:
        st.caption(
            f"**{n_edit}** parámetros marcados como editables"
            + (f" — se propagarán a todas las unidades {familia_sym}" if propaga_familia else "")
        )
    with col_g2:
        if st.button("💾 Guardar selección + rangos", type="primary",
                     key="dsl_guardar_cfg"):
            cfg = _cargar_config()
            cfg.setdefault(sym, {})
            for _, row in edited.iterrows():
                sp = row["Símbolo"]
                cfg[sym][sp] = {
                    "relevante":   bool(row["Editable"]),
                    "ajustable":   bool(row["Editable"]),
                    "min":         row["Min"] if pd.notna(row["Min"]) else None,
                    "max":         row["Max"] if pd.notna(row["Max"]) else None,
                }
            # Para F1 y F2: propagar flag Editable (y min/max) a todas las unidades de la familia
            if propaga_familia:
                otras = [s for s, (f, _) in UNIDADES_DSL.items() if f == familia_sym and s != sym]
                for otra in otras:
                    cfg.setdefault(otra, {})
                    for _, row in edited.iterrows():
                        sp = row["Símbolo"]
                        entry = cfg[otra].get(sp, {})
                        entry["relevante"] = bool(row["Editable"])
                        entry["ajustable"] = bool(row["Editable"])
                        if pd.notna(row["Min"]):
                            entry["min"] = row["Min"]
                        if pd.notna(row["Max"]):
                            entry["max"] = row["Max"]
                        cfg[otra][sp] = entry
            _guardar_config(cfg)
            for k in list(st.session_state.keys()):
                if k.startswith("dsl_params_relevantes_"):
                    del st.session_state[k]
            msg = "Configuración guardada."
            if propaga_familia:
                otras_names = [s for s, (f, _) in UNIDADES_DSL.items() if f == familia_sym and s != sym]
                msg += f" Propagado a: {', '.join(otras_names)}."
            st.success(msg)

    #  Carga completa a PowerFactory (todos los parámetros actuales) 
    st.divider()
    st.markdown("#### Cargar parámetros actuales completos a PowerFactory")
    st.caption(
        "Genera un Excel de staging con **todos** los parámetros de col R "
        "(sin filtro de relevancia) para cargar el estado actual completo en PF. "
        "Útil como punto de partida antes de iniciar experimentos."
    )

    params_actuales = df[df["val_actual"].notna()][
        ["slot_dsl", "simbolo", "descripcion", "val_actual"]
    ]
    n_actuales = len(params_actuales)
    st.caption(f"**{n_actuales}** parámetros con valor en col R")

    col_ca1, col_ca2 = st.columns([2, 1])
    with col_ca1:
        sufijo_base = st.text_input(
            "Sufijo del archivo (opcional)",
            value="BASELINE",
            key="dsl_sufijo_baseline",
        )
    with col_ca2:
        st.write("")
        if st.button("📊 Generar Excel completo", key="dsl_gen_baseline",
                     use_container_width=True, disabled=n_actuales == 0):
            nombre_bl = _nombre_experimento(sym, sufijo_base)
            params_bl = [
                {
                    "bloque":       row["slot_dsl"],
                    "simbolo":      row["simbolo"],
                    "descripcion":  row["descripcion"],
                    "valor_base":   row["val_actual"],
                    "valor":        row["val_actual"],
                    "es_ajustable": False,
                }
                for _, row in params_actuales.iterrows()
            ]
            try:
                ruta_bl = _generar_excel_experimento(sym, nombre_bl, params_bl)
                st.session_state[f"dsl_excel_ruta_{sym}"] = str(ruta_bl)
                st.success(f"Generado: `{ruta_bl.name}`")
                st.info(
                    f"**Paso siguiente:** En PowerFactory ejecutar "
                    f"`CargadorDSL_Andritz.py` apuntando al archivo  \n"
                    f"`{ruta_bl}`",
                    icon="ℹ️",
                )
            except Exception as e:
                st.error(f"Error: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — Experimentos
# ══════════════════════════════════════════════════════════════════════════════

def _tab_experimentos():
    st.caption(
        "Varía los parámetros relevantes respecto a los Parámetros Actuales (col R). "
        "Genera el Excel de staging, carga en PF y registra el experimento en la base de datos."
    )

    cfg = _cargar_config()
    familia_sel = st.session_state.get("dsl_familia", "F1")
    unidades_familia = [s for s, (f, _) in UNIDADES_DSL.items() if f == familia_sel]

    col_s, col_sem, col_ev = st.columns(3)
    with col_s:
        sym = st.selectbox("Unidad", options=unidades_familia, key="dsl_sym_tab2")

    _raiz_dsl = st.session_state.get("cfg_RAIZ") or r"C:\Datos del CNDC\01_INFO CNDC_RPF"
    try:
        semestres_dsl = sorted(d for d in os.listdir(_raiz_dsl)
                               if os.path.isdir(os.path.join(_raiz_dsl, d)))
    except OSError:
        semestres_dsl = []

    with col_sem:
        semestre_sel = st.selectbox(
            "Semestre", options=semestres_dsl or ["(no encontrado)"],
            key="dsl_semestre_ref", disabled=not semestres_dsl,
        )

    eventos_dsl = []
    if semestres_dsl:
        _base_ev_dsl = os.path.join(_raiz_dsl, semestre_sel, "Análisis_todos_los_eventos")
        try:
            eventos_dsl = sorted(
                (d for d in os.listdir(_base_ev_dsl) if os.path.isdir(os.path.join(_base_ev_dsl, d))),
                key=lambda d: int(m.group(1)) if (m := re.search(r"(\d+)$", d)) else -1)
        except OSError:
            eventos_dsl = []

    with col_ev:
        evento_sel = st.selectbox(
            "Evento", options=eventos_dsl or ["(no encontrado)"],
            key="dsl_evento_ref", disabled=not eventos_dsl,
        )

    # Convención: "{semestre}::{evento}" — permite a la Tab 4 (Comparar Curvas)
    # resolver de vuelta la carpeta E0/E1/SCADA del mismo evento.
    evento_ref = f"{semestre_sel}::{evento_sel}" if semestres_dsl and eventos_dsl else ""
    if not evento_ref:
        st.warning(
            f"No se encontraron semestres/eventos en `{_raiz_dsl}`. "
            "Verifique la ruta RAIZ en Configuración (Bloque 07)."
        )

    cfg_u = _cfg_unidad(cfg, sym)
    relevantes = {s: d for s, d in cfg_u.items()
                  if d.get("relevante") and d.get("ajustable")}

    if not relevantes:
        st.info(
            f"No hay parámetros relevantes/ajustables configurados para {sym}. "
            "Ve a **Parámetros Actuales** y marca los que quieres optimizar.",
            icon="ℹ️",
        )
        return

    # Cargar col R (Parámetros Actuales) como base
    cache_key = f"dsl_df_{sym}"
    cached2 = st.session_state.get(cache_key)
    if cached2 is None or "slot_dsl" not in cached2.columns:
        path = _excel_path(sym)
        if path is None:
            st.error(f"Excel comparativo no encontrado para {sym}.")
            return
        with st.spinner("Leyendo parámetros actuales…"):
            st.session_state[cache_key] = _leer_comparativo(sym)

    df_base: pd.DataFrame = st.session_state[cache_key]
    base_vals = df_base.set_index("simbolo")["val_actual"].to_dict() if not df_base.empty else {}

    #  Tabla de variación 
    st.markdown("**Ajustar valores para el experimento:**")

    valores_exp: dict[str, float] = {}
    bloques_exp: dict[str, str]   = {}
    bloque_actual = ""

    for sym_p, p_cfg in relevantes.items():
        fila = df_base[df_base["simbolo"] == sym_p]
        bloque_p = fila["slot_dsl"].iloc[0] if not fila.empty else "—"
        desc_p   = fila["descripcion"].iloc[0] if not fila.empty else sym_p
        val_base = base_vals.get(sym_p)

        if bloque_p != bloque_actual:
            st.markdown(f"**{bloque_p}**")
            bloque_actual = bloque_p

        vmin = p_cfg.get("min")
        vmax = p_cfg.get("max")
        step = round((vmax - vmin) / 100, 6) if (vmin is not None and vmax is not None) else 0.001

        col_lbl, col_inp, col_base = st.columns([2, 1, 1])
        with col_lbl:
            st.markdown(f"<small>{sym_p} — {desc_p}</small>", unsafe_allow_html=True)
        with col_base:
            st.caption(f"Base: {val_base if val_base is not None else 'N/D'}")
        with col_inp:
            v = st.number_input(
                label=sym_p,
                value=float(val_base) if val_base is not None else 0.0,
                min_value=float(vmin) if vmin is not None else None,
                max_value=float(vmax) if vmax is not None else None,
                step=step,
                label_visibility="collapsed",
                key=f"dsl_exp_input_{sym}_{sym_p}",
            )
        valores_exp[sym_p] = v
        bloques_exp[sym_p] = bloque_p

    st.divider()

    #  Nombre del experimento 
    nombre_auto = _nombre_experimento(sym)
    col_n, col_suf = st.columns([1, 2])
    with col_n:
        st.text_input("Nombre base (auto)", value=nombre_auto, disabled=True,
                      key="dsl_nombre_auto")
    with col_suf:
        sufijo = st.text_input("Sufijo descriptivo (opcional)", value="",
                               placeholder="ej. GOV_Kp_alto",
                               key="dsl_sufijo")

    nombre_final = _nombre_experimento(sym, sufijo)
    st.caption(f"📄 Nombre final: **{nombre_final}**")

    notas = st.text_area("Notas / Observaciones", key="dsl_notas", height=60)

    #  Acciones 
    col_a, col_b, col_c = st.columns(3)

    params_lista = [
        {
            "bloque":       bloques_exp.get(s, ""),          # slot DSL (GOV, AVR…)
            "simbolo":      s,
            "descripcion":  (df_base[df_base["simbolo"] == s]["descripcion"].iloc[0]
                             if not df_base[df_base["simbolo"] == s].empty else ""),
            "valor_base":   base_vals.get(s),
            "valor":        v,
            "es_ajustable": True,
        }
        for s, v in valores_exp.items()
    ]

    excel_path_gen = None

    with col_a:
        if st.button("📊 Generar Excel de experimento", use_container_width=True,
                     key="dsl_gen_excel"):
            try:
                ruta = _generar_excel_experimento(sym, nombre_final, params_lista)
                st.session_state[f"dsl_excel_ruta_{sym}"] = str(ruta)
                st.success(f"Excel generado: `{ruta.name}`")
            except Exception as e:
                st.error(f"Error al generar Excel: {e}")

    ruta_generada = st.session_state.get(f"dsl_excel_ruta_{sym}", "")

    with col_b:
        cargado_pf = st.checkbox(
            "✅ Cargado en PF",
            key="dsl_cargado_pf",
            help="Marcar tras ejecutar CargadorDSL_Andritz.py en PowerFactory",
            disabled=not bool(ruta_generada),
        )
        if ruta_generada:
            st.caption(
                f"Ejecutar en PF → `CargadorDSL_Andritz.py`  \n"
                f"apuntando a: `{Path(ruta_generada).name}`"
            )

    with col_c:
        if st.button("🗄 Registrar en base de datos", type="primary",
                     use_container_width=True, key="dsl_registrar",
                     disabled=not bool(ruta_generada)):
            try:
                import dsl_db
                exp_id = dsl_db.registrar_experimento(
                    sym=sym,
                    familia=UNIDADES_DSL[sym][0],
                    evento_ref=evento_ref,
                    nombre=nombre_final,
                    notas=notas,
                    excel_path=ruta_generada,
                )
                dsl_db.registrar_params(exp_id, params_lista)
                st.session_state[f"dsl_ultimo_exp_id_{sym}"] = exp_id
                st.success(f"Experimento registrado (ID: {exp_id})")
            except Exception as e:
                st.error(f"Error al registrar: {e}")

    #  Vincular curva extraída + KPIs automáticos (post-simulación) 
    exp_id_actual = st.session_state.get(f"dsl_ultimo_exp_id_{sym}")
    if exp_id_actual and cargado_pf:
        st.divider()
        st.markdown("#### Vincular curva extraída y calcular KPIs")
        st.caption(
            "Tras simular en PowerFactory, ejecute `DatosCurvas_v3.py` (rama "
            "*Experimento DSL*) y luego calcule los KPIs aquí con la misma "
            "metodología CNDC que usan los Bloques 3/4 para E0/E1."
        )

        _carpeta_default = str(_exp_dir(sym) / nombre_final / "Datos Curvas")
        col_tf, col_dt = st.columns(2)
        with col_tf:
            t_falla_exp = st.number_input(
                "t₀ falla [s]", min_value=0.0, max_value=300.0, step=0.5,
                value=5.0, key="dsl_exp_t_falla",
            )
        with col_dt:
            dt_exp = st.number_input(
                "Δt evaluación [s]", min_value=1, max_value=120, step=1,
                value=35, key="dsl_exp_dt",
            )

        carpeta_curvas = st.text_input(
            "Carpeta `Datos Curvas` del experimento",
            value=_carpeta_default, key="dsl_carpeta_curvas",
        )

        if st.button("📈 Calcular KPIs desde curva", key="dsl_calc_kpis", type="primary"):
            if not evento_ref or "::" not in evento_ref:
                st.error("Seleccione un Semestre/Evento válido arriba antes de calcular KPIs.")
            elif not os.path.isdir(carpeta_curvas):
                st.error(f"La carpeta `{carpeta_curvas}` no existe. Ejecute `DatosCurvas_v3.py` primero.")
            else:
                _archivo_unidad = os.path.join(carpeta_curvas, f"F.P. {sym}.xlsx")
                if not os.path.isfile(_archivo_unidad):
                    _archivo_unidad = os.path.join(carpeta_curvas, f"F.P. {sym}.ALL.xlsx")
                if not os.path.isfile(_archivo_unidad):
                    st.error(
                        f"No se encontró `F.P. {sym}.xlsx` en `{carpeta_curvas}`. "
                        "Verifique que ejecutó `DatosCurvas_v3.py` apuntando a este experimento."
                    )
                else:
                    try:
                        ts_e, fs_e, ps_e = _leer_curva_experimento(_archivo_unidad, t_falla_exp)

                        _, _evento_part = evento_ref.split("::", 1)
                        _m_ev = re.search(r"(\d+)", _evento_part)
                        _n_ev = _m_ev.group(1) if _m_ev else "0"
                        _ev_path = os.path.join(
                            _raiz_dsl, semestre_sel, "Análisis_todos_los_eventos", _evento_part
                        )
                        _loc_gen_path = st.session_state.get("cfg_LOC_NAMES_GEN_PATH") or ""

                        _pm = kpi_calc._load_pmax_cargado(_ev_path, _n_ev)
                        _tm = kpi_calc._load_tech_map(_loc_gen_path)
                        pm_v, tk, _ = kpi_calc._get_pmax_from_cargado(sym, _pm, _tm)
                        rp_v = float(kpi_calc._get_rp_default(tk, _loc_gen_path)) / 100.0

                        kpis_calc = kpi_calc._cndc_kpis(ts_e, fs_e, ps_e, pm_v, rp_v, dt_exp)
                        rocof_calc = kpi_calc._calcular_rocof(ts_e, fs_e, 3.0)

                        if kpis_calc is None:
                            st.error("No se pudieron calcular KPIs — verifique el contenido de la curva.")
                        else:
                            kpis_out = {
                                "f0": kpis_calc["f0"], "f_min": kpis_calc["f_min"],
                                "t_min": kpis_calc["t_min"], "delta_f": kpis_calc["delta_f"],
                                "f_delta_t": kpis_calc["f_dt"], "p0": kpis_calc["p0"],
                                "p_max": pm_v, "p_delta_t": kpis_calc["p_dt"],
                                "rocof": rocof_calc, "delta_p": kpis_calc["dp"],
                                "delta_p_pct": kpis_calc["dp_pct"], "aporta_rpf": kpis_calc["aporta"],
                            }
                            import dsl_db
                            dsl_db.registrar_kpis(exp_id_actual, kpis_out)
                            dsl_db.vincular_curva(exp_id_actual, carpeta_curvas)
                            st.success(
                                f"KPIs calculados y guardados — f_min={kpis_calc['f_min']:.4f} Hz, "
                                f"ROCOF={rocof_calc:.4f} Hz/s, ΔP%={kpis_calc['dp_pct']:.2f}%."
                            )
                    except Exception as e:
                        st.error(f"Error al calcular KPIs desde la curva: {e}")

        with st.expander("✏️ Editar KPIs manualmente (respaldo)"):
            st.caption("Usar solo si la curva todavía no está disponible o la detección de columnas falla.")
            c1, c2, c3, c4 = st.columns(4)
            kpis_in = {
                "f_min":       c1.number_input("f_min (Hz)",      value=49.5,  step=0.01, key="kpi_fmin"),
                "t_min":       c2.number_input("t_min (s)",       value=10.0,  step=0.1,  key="kpi_tmin"),
                "rocof":       c3.number_input("ROCOF (Hz/s)",    value=-0.4,  step=0.01, key="kpi_rocof"),
                "delta_p_pct": c4.number_input("ΔP%",            value=2.0,   step=0.1,  key="kpi_dpp"),
            }
            c5, c6, c7, c8 = st.columns(4)
            kpis_in.update({
                "delta_f":    c5.number_input("Δf (Hz)",         value=0.5,   step=0.01, key="kpi_df"),
                "f_delta_t":  c6.number_input("f_Δt (Hz)",       value=49.7,  step=0.01, key="kpi_fdt"),
                "delta_p":    c7.number_input("ΔP (MW)",         value=5.0,   step=0.1,  key="kpi_dp"),
                "aporta_rpf": c8.checkbox("✅ Aporta RPF",                               key="kpi_rpf"),
            })

            if st.button("💾 Guardar KPIs manuales", key="dsl_guardar_kpis"):
                try:
                    import dsl_db
                    dsl_db.registrar_kpis(exp_id_actual, kpis_in)
                    st.success(f"KPIs guardados para experimento {exp_id_actual}.")
                except Exception as e:
                    st.error(f"Error al guardar KPIs: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — Historial / Comparación
# ══════════════════════════════════════════════════════════════════════════════

def _tab_historial():
    st.caption(
        "Historial de experimentos registrados en la base de datos. "
        "Compara parámetros y KPIs entre experimentos para identificar la mejor configuración. "
        "Exporta en formato listo para entrenamiento de modelos IA."
    )

    #  Filtros 
    col_f1, col_f2, col_f3, col_f4 = st.columns(4)
    with col_f1:
        filtro_fam = st.selectbox("Familia", ["(todas)"] + list(FAMILIAS.keys()),
                                  key="hist_familia")
    with col_f2:
        unidades_disp = (
            [s for s, (f, _) in UNIDADES_DSL.items() if f == filtro_fam]
            if filtro_fam != "(todas)"
            else list(UNIDADES_DSL.keys())
        )
        filtro_sym = st.selectbox("Unidad", ["(todas)"] + unidades_disp,
                                  key="hist_sym")
    with col_f3:
        filtro_estado = st.selectbox("Estado",
                                     ["(todos)", "configurado", "simulado", "analizado"],
                                     key="hist_estado")
    with col_f4:
        if st.button("🔄 Recargar", key="hist_reload", use_container_width=True):
            st.session_state.pop("hist_df_cache", None)

    #  Cargar experimentos 
    if "hist_df_cache" not in st.session_state:
        try:
            import dsl_db
            df_hist = dsl_db.listar_experimentos(
                sym=None if filtro_sym == "(todas)" else filtro_sym,
                familia=None if filtro_fam == "(todas)" else filtro_fam,
                estado=None if filtro_estado == "(todos)" else filtro_estado,
            )
            st.session_state["hist_df_cache"] = df_hist
        except Exception as e:
            st.error(f"Error al leer el historial de experimentos: {e}")
            st.info("Verificar que el directorio de datos local sea accesible. "
                    "Ejecutar `dsl_db.crear_tablas()` para inicializar si es la primera vez.")
            return

    df_hist: pd.DataFrame = st.session_state.get("hist_df_cache", pd.DataFrame())

    if df_hist.empty:
        st.info("No hay experimentos registrados con los filtros seleccionados.")
        return

    #  Tabla resumen 
    st.markdown(f"**{len(df_hist)} experimentos**")

    col_map = {
        "id": "ID", "sym": "Unidad", "familia": "Familia",
        "evento_ref": "Evento", "nombre": "Nombre",
        "fecha": "Fecha", "estado": "Estado",
        "f_min": "f_min (Hz)", "rocof": "ROCOF",
        "delta_p_pct": "ΔP%", "aporta_rpf": "Aporta RPF",
    }
    df_show = df_hist[[c for c in col_map if c in df_hist.columns]].rename(columns=col_map)

    sel = st.dataframe(
        df_show,
        use_container_width=True,
        hide_index=True,
        on_select="rerun",
        selection_mode="multi-row",
        key="hist_tabla",
    )

    ids_sel = []
    if sel and sel.selection and sel.selection.get("rows"):
        ids_sel = [int(df_hist.iloc[i]["id"]) for i in sel.selection["rows"]]

    #  Comparación 
    if len(ids_sel) >= 2:
        st.divider()
        st.markdown(f"#### Comparación de {len(ids_sel)} experimentos")

        try:
            import dsl_db
            dfs_params = {eid: dsl_db.params_de_experimento(eid) for eid in ids_sel}
        except Exception as e:
            st.error(f"Error al cargar parámetros: {e}")
            dfs_params = {}

        if dfs_params:
            # Pivot: símbolo → valor por experimento
            piv = None
            for eid, dfp in dfs_params.items():
                if dfp.empty:
                    continue
                serie = dfp.set_index("simbolo")["valor"].rename(f"EXP-{eid}")
                piv = serie if piv is None else pd.concat([piv, serie], axis=1)

            if piv is not None:
                piv = piv.reset_index()
                # Agregar delta respecto al primer experimento
                base_col = f"EXP-{ids_sel[0]}"
                for c in piv.columns[2:]:
                    piv[f"Δ {c}"] = piv[c] - piv[base_col]

                st.dataframe(piv, use_container_width=True, hide_index=True)

        # KPIs comparados
        kpi_cols = ["f_min", "t_min", "delta_f", "rocof", "delta_p_pct", "aporta_rpf"]
        kpi_rows = df_hist[df_hist["id"].isin(ids_sel)][
            ["id", "nombre"] + [c for c in kpi_cols if c in df_hist.columns]
        ]
        if not kpi_rows.empty:
            st.markdown("**KPIs:**")
            st.dataframe(kpi_rows, use_container_width=True, hide_index=True)

    #  Exportar para IA 
    st.divider()
    col_exp1, col_exp2 = st.columns(2)
    with col_exp1:
        sym_ia = st.selectbox("Exportar para IA — Unidad",
                              ["(todas)"] + list(UNIDADES_DSL.keys()),
                              key="ia_sym")
    with col_exp2:
        st.write("")
        if st.button("📤 Exportar CSV para IA", key="ia_export", use_container_width=True):
            try:
                import dsl_db
                df_ia = dsl_db.exportar_para_ia(
                    sym=None if sym_ia == "(todas)" else sym_ia
                )
                if df_ia.empty:
                    st.warning("No hay experimentos simulados para exportar.")
                else:
                    csv_bytes = df_ia.to_csv(index=False, sep=";").encode("utf-8-sig")
                    fname = f"DSL_IA_{sym_ia}_{datetime.now().strftime('%Y%m%d')}.csv"
                    st.download_button(
                        "⬇ Descargar CSV",
                        data=csv_bytes,
                        file_name=fname,
                        mime="text/csv",
                        key="ia_download",
                    )
                    st.caption(
                        f"{len(df_ia)} experimentos · "
                        f"{len(df_ia.columns)} columnas (features + targets)"
                    )
            except Exception as e:
                st.error(f"Error al exportar: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# PUNTO DE ENTRADA PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def _tab_bar(tab_defs: list, block_key: str) -> str:
    """Barra de tabs con botones nativos — igual que _v4_tab_bar en el interfaz principal."""
    sk  = f"v4_tab_{block_key}"
    ids = [td["id"] for td in tab_defs]
    if sk not in st.session_state or st.session_state[sk] not in ids:
        st.session_state[sk] = ids[0]
    active = st.session_state[sk]

    cols = st.columns(len(tab_defs))
    for td, col in zip(tab_defs, cols):
        with col:
            if st.button(
                td["label"],
                key=f"{sk}_{td['id']}",
                type="primary" if td["id"] == active else "secondary",
                use_container_width=True,
            ):
                st.session_state[sk] = td["id"]
                st.rerun()

    return active


def render_bloque_dsl(session_state=None):
    """Renderiza el bloque completo. Llamar desde interfaz_analisis_RPF.py."""

    st.markdown("### ⚙ Gestión de Parámetros DSL (DigSILENT Simulation Language) — Optimización")

    active_tab = _tab_bar([
        {"id": "params",  "label": "Parámetros Actuales"},
        {"id": "exp",     "label": "Experimentos"},
        {"id": "hist",    "label": "Historial / Comparación"},
    ], "b02_dsl")

    if active_tab == "params":
        _tab_parametros_actuales()
    elif active_tab == "exp":
        _tab_experimentos()
    elif active_tab == "hist":
        _tab_historial()
