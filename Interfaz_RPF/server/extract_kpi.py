#!/usr/bin/env python3
"""
extract_kpi.py — Capa 2: Extrae KPIs de tabla_resultados_COBEE.xlsx → PostgreSQL rpf_kpi_cobee

Uso:
  python3 extract_kpi.py <ruta_al_archivo.xlsx>
  python3 extract_kpi.py <ruta> --dry-run
  python3 extract_kpi.py --pendientes        # procesa todos los archivos sin procesar en rpf_file_log
"""

import sys
import os
import re
import json
import argparse
import logging
from datetime import datetime, date, time
from pathlib import Path

import openpyxl
import psycopg2
from psycopg2.extras import execute_values
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / '.env')

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
)
log = logging.getLogger(__name__)

# Orden y tipos de métricas exactamente como aparecen en tabla_extendida (col A, filas 2-18)
METRIC_MAP = [
    ('P_max',            'p_max_mw',       'float'),
    ('P_0',              'p_0_mw',         'float'),
    ('P_35',             'p_35_mw',        'float'),
    ('R.inicial[MW]',    'r_inicial_mw',   'float'),
    ('R.inicial[%]',     'r_inicial_pct',  'float'),
    ('P.entregada[MW]',  'p_entregada_mw', 'float'),
    ('P.entregada[%]',   'p_entregada_pct','float'),
    ('Aporta_RPF',       'aporta_rpf',     'str'),
    ('droop_informado',  'droop_inf_pct',  'float'),
    ('droop_calculado',  'droop_calc_pct', 'float'),
    ('f_0',              'f_0_hz',         'float'),
    ('f_min',            'f_min_hz',       'float'),
    ('f_35',             'f_35_hz',        'float'),
    ('t_0',              't_0',            'time'),
    ('t_min',            't_min',          'time'),
    ('t_35',             't_35',           'time'),
    ('fecha',            'fecha_evento',   'date'),
]

# ─── Conversores ────────────────────────────────────────────────────────────

def safe_float(val):
    if val is None or val == '' or val == '-':
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def safe_time(val):
    if val is None:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    # Excel guarda tiempo como fracción de día (float 0–1)
    if isinstance(val, float):
        total_s = int(round(val * 86400))
        return time(total_s // 3600 % 24, (total_s % 3600) // 60, total_s % 60)
    try:
        parts = str(val).split(':')
        if len(parts) >= 2:
            return time(int(parts[0]) % 24, int(parts[1]),
                        int(parts[2]) if len(parts) > 2 else 0)
    except Exception:
        pass
    return None


def safe_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(str(val), fmt).date()
        except ValueError:
            pass
    return None


# ─── Metadatos del path ──────────────────────────────────────────────────────

def parse_path_metadata(filepath: str):
    semestre = None
    evento = None
    m = re.search(r'([12][0-9]{3})\s*sem([12])', filepath, re.IGNORECASE)
    if m:
        semestre = f"{m.group(1)}_sem{m.group(2)}"
    m2 = re.search(r'[Ee]vento\s*(\d+)', filepath)
    if m2:
        evento = f"Evento_{m2.group(1)}"
    return semestre, evento


# ─── Extracción Excel ────────────────────────────────────────────────────────

def extract_from_excel(filepath):
    """
    Lee tabla_extendida (transpuesta: filas=métricas, cols=unidades).
    Devuelve (records: list[dict], semestre: str, evento: str).
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    if 'tabla_extendida' not in wb.sheetnames:
        raise ValueError(f"Hoja 'tabla_extendida' no encontrada. Hojas disponibles: {wb.sheetnames}")
    ws = wb['tabla_extendida']

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Hoja tabla_extendida tiene menos de 2 filas")

    # Fila 1 (índice 0): encabezados de unidades (col B en adelante)
    unit_row = rows[0]
    units = [str(h).strip() for h in unit_row[1:] if h is not None]
    n_units = len(units)
    log.info(f"Unidades encontradas ({n_units}): {', '.join(units)}")

    # Filas 2-18 (índices 1-17): métricas
    # Construir dict: nombre_métrica → [val_u0, val_u1, ...]
    metrics = {}
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        vals = list(row[1: n_units + 1])
        metrics[key] = vals

    log.info(f"Métricas leídas: {list(metrics.keys())}")

    # fecha_evento: un valor global por evento (primer no-nulo de la fila 'fecha')
    fecha_evento = None
    for v in metrics.get('fecha', []):
        d = safe_date(v)
        if d:
            fecha_evento = d
            break

    semestre, evento = parse_path_metadata(str(filepath))
    log.info(f"Metadatos path → semestre={semestre}, evento={evento}")

    records = []
    for i, unidad in enumerate(units):
        rec = {
            'semestre':    semestre,
            'evento':      evento,
            'fecha_evento': fecha_evento,
            'unidad':      unidad,
            'source_file': str(filepath),
        }
        for (excel_name, pg_col, dtype) in METRIC_MAP:
            if excel_name == 'fecha':
                continue  # ya procesado arriba
            vals = metrics.get(excel_name, [])
            raw = vals[i] if i < len(vals) else None
            if dtype == 'float':
                rec[pg_col] = safe_float(raw)
            elif dtype == 'time':
                rec[pg_col] = safe_time(raw)
            elif dtype == 'str':
                rec[pg_col] = str(raw).strip() if raw is not None else None
            else:
                rec[pg_col] = raw
        records.append(rec)

    return records, semestre, evento


# ─── Conexión PostgreSQL ─────────────────────────────────────────────────────

def get_conn():
    return psycopg2.connect(
        host=os.getenv('POSTGRES_HOST', 'localhost'),
        port=int(os.getenv('POSTGRES_PORT', '5432')),
        dbname=os.getenv('POSTGRES_DB', 'rpf_intelligence'),
        user=os.getenv('POSTGRES_USER', 'n8n'),
        password=os.getenv('POSTGRES_PASSWORD', ''),
    )


# ─── Inserción ───────────────────────────────────────────────────────────────

COLS = [
    'semestre', 'evento', 'fecha_evento', 'unidad',
    'p_max_mw', 'p_0_mw', 'p_35_mw',
    'r_inicial_mw', 'r_inicial_pct',
    'p_entregada_mw', 'p_entregada_pct',
    'aporta_rpf', 'droop_inf_pct', 'droop_calc_pct',
    'f_0_hz', 'f_min_hz', 'f_35_hz',
    't_0', 't_min', 't_35', 'source_file',
]


def insert_records(records, dry_run=False):
    if dry_run:
        log.info(f"[DRY-RUN] {len(records)} registros — primeros 3:")
        for r in records[:3]:
            log.info(f"  {r['unidad']}: p_max={r.get('p_max_mw')}, "
                     f"p_0={r.get('p_0_mw')}, aporta={r.get('aporta_rpf')}, "
                     f"f_0={r.get('f_0_hz')}")
        return len(records)

    values = [tuple(r.get(c) for c in COLS) for r in records]
    with get_conn() as conn:
        with conn.cursor() as cur:
            execute_values(
                cur,
                f"INSERT INTO rpf_kpi_cobee ({','.join(COLS)}) VALUES %s",
                values,
            )
        conn.commit()
    return len(records)


def mark_processed(filepath, dry_run=False):
    if dry_run:
        log.info(f"[DRY-RUN] Marcaría processed=true para: {filepath}")
        return
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE rpf_file_log SET processed=true WHERE filepath=%s",
                (str(filepath),),
            )
        conn.commit()


# ─── Modo --pendientes ───────────────────────────────────────────────────────

def procesar_pendientes(dry_run=False):
    """Busca en rpf_file_log archivos tabla_resultados_COBEE sin procesar."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT filepath FROM rpf_file_log
                WHERE processed = false
                  AND filepath LIKE '%tabla_resultados_COBEE%'
                ORDER BY received_at
            """)
            rows = cur.fetchall()

    if not rows:
        log.info("No hay archivos pendientes en rpf_file_log.")
        return

    log.info(f"Archivos pendientes: {len(rows)}")
    results = []
    for (fp,) in rows:
        path = Path(fp)
        if not path.exists():
            log.warning(f"Archivo no existe en disco: {fp}")
            results.append({'file': fp, 'ok': False, 'error': 'not_found'})
            continue
        result = process_file(fp, dry_run)
        results.append(result)

    ok = sum(1 for r in results if r.get('ok'))
    log.info(f"Procesados: {ok}/{len(results)} OK")
    print(json.dumps({'total': len(results), 'ok': ok, 'results': results}))


# ─── Función principal de procesamiento ─────────────────────────────────────

def process_file(filepath, dry_run=False):
    try:
        records, semestre, evento = extract_from_excel(filepath)
        log.info(f"Extraídos {len(records)} registros")

        n = insert_records(records, dry_run=dry_run)
        mark_processed(str(filepath), dry_run=dry_run)

        result = {
            'ok': True,
            'file': str(filepath),
            'semestre': semestre,
            'evento': evento,
            'records_inserted': n,
        }
        log.info(f"OK — {n} registros insertados en rpf_kpi_cobee")
        return result

    except Exception as e:
        log.error(f"Error: {e}", exc_info=True)
        return {'ok': False, 'file': str(filepath), 'error': str(e)}


# ─── Entry point ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='Extrae KPIs de tabla_resultados_COBEE.xlsx → rpf_kpi_cobee'
    )
    parser.add_argument('filepath', nargs='?', help='Ruta al archivo .xlsx')
    parser.add_argument('--dry-run', action='store_true',
                        help='Muestra datos sin escribir en DB')
    parser.add_argument('--pendientes', action='store_true',
                        help='Procesa todos los archivos pendientes en rpf_file_log')
    args = parser.parse_args()

    if args.pendientes:
        procesar_pendientes(dry_run=args.dry_run)
        return

    if not args.filepath:
        parser.error('Se requiere <filepath> o --pendientes')

    fp = Path(args.filepath)
    if not fp.exists():
        log.error(f"Archivo no encontrado: {fp}")
        sys.exit(1)

    result = process_file(str(fp), dry_run=args.dry_run)
    print(json.dumps(result))
    sys.exit(0 if result['ok'] else 1)


if __name__ == '__main__':
    main()
