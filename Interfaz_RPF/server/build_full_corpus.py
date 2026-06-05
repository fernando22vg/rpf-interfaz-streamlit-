#!/usr/bin/env python3
"""
build_full_corpus.py — Construye el corpus completo para la IA de COBEE

Escanea TODAS las fuentes disponibles y crea múltiples Knowledge Bases en Open WebUI:
  KB "RPF KPIs"          ← Base de datos PostgreSQL (estadísticas y KPIs)
  KB "RPF Análisis CNDC" ← Archivos de análisis CNDC por evento
  KB "RPF Normativa CDM" ← Documentos normativos y manuales
  KB "RPF Sistema"       ← Código Python del sistema (para autocompletar/corregir)
  KB "RPF SCADA"         ← Resúmenes de señales SCADA procesadas

Uso:
  python3 build_full_corpus.py                    # todo
  python3 build_full_corpus.py --kb rpf-kpis      # solo una KB
  python3 build_full_corpus.py --dry-run          # listar sin subir
"""

import os
import sys
import re
import json
import time
import argparse
import logging
import tempfile
from pathlib import Path
from datetime import datetime, timezone

import requests
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / '.env')

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
log = logging.getLogger(__name__)

# ─── Config ───────────────────────────────────────────────────────────────────
WEBUI_URL   = os.getenv('WEBUI_URL', 'http://localhost:3000')
WEBUI_EMAIL = os.getenv('WEBUI_EMAIL', '')
WEBUI_PASS  = os.getenv('WEBUI_PASSWORD', '')

DATA_ROOT   = os.getenv('DATA_ROOT', '/home/joselozano/rpf-proyecto-datos')
RAIZ_RPF    = os.path.join(DATA_ROOT, '01_INFO_CNDC_RPF')
CODE_ROOT   = '/home/joselozano/rpf-ejecucion'

# ─── Auth Open WebUI ──────────────────────────────────────────────────────────

def get_session():
    s = requests.Session()
    r = s.post(f'{WEBUI_URL}/api/v1/auths/signin',
               json={'email': WEBUI_EMAIL, 'password': WEBUI_PASS}, timeout=10)
    if r.status_code != 200:
        raise RuntimeError(f"Login fallido: {r.text[:200]}")
    s.headers['Authorization'] = f"Bearer {r.json()['token']}"
    return s


def get_or_create_kb(session, name: str, description: str) -> str:
    r = session.get(f'{WEBUI_URL}/api/v1/knowledge/', timeout=10)
    items = r.json().get('items', []) if isinstance(r.json(), dict) else r.json()
    for kb in items:
        if isinstance(kb, dict) and kb.get('name') == name:
            return kb['id']
    r2 = session.post(f'{WEBUI_URL}/api/v1/knowledge/create',
                      json={'name': name, 'description': description}, timeout=10)
    if r2.status_code not in (200, 201):
        raise RuntimeError(f"Error creando KB '{name}': {r2.text[:200]}")
    return r2.json()['id']


def clear_kb(session, kb_id: str):
    """Elimina archivos de la KB Y borra los archivos físicos del servidor."""
    r = session.get(f'{WEBUI_URL}/api/v1/knowledge/{kb_id}', timeout=10)
    for f in (r.json().get('files') or []):
        fid = f.get('id') or f.get('file', {}).get('id')
        if fid:
            # Desasociar de la KB
            session.delete(f'{WEBUI_URL}/api/v1/knowledge/{kb_id}/file/delete',
                           json={'file_id': fid}, timeout=10)
            # Borrar el archivo físico para evitar "Duplicate content"
            session.delete(f'{WEBUI_URL}/api/v1/files/{fid}', timeout=10)


def purge_all_files(session):
    """Borra TODOS los archivos del servidor (limpieza inicial)."""
    r = session.get(f'{WEBUI_URL}/api/v1/files/', timeout=15)
    if r.status_code != 200:
        return
    files = r.json() if isinstance(r.json(), list) else r.json().get('files', [])
    deleted = 0
    for f in files:
        fid = f.get('id')
        if fid:
            session.delete(f'{WEBUI_URL}/api/v1/files/{fid}', timeout=10)
            deleted += 1
    if deleted:
        log.info(f"  Archivos físicos eliminados: {deleted}")


def upload_text(session, kb_id: str, filename: str, content: str) -> bool:
    # Añadir timestamp para evitar "Duplicate content detected" en re-subidas
    stamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%S')
    content = content + f'\n<!-- build:{stamp} -->\n'
    with tempfile.NamedTemporaryFile(mode='w', suffix='.md',
                                     encoding='utf-8', delete=False) as tf:
        tf.write(content)
        tmp = tf.name
    try:
        with open(tmp, 'rb') as f:
            r = session.post(f'{WEBUI_URL}/api/v1/files/',
                             files={'file': (filename, f, 'text/markdown')},
                             timeout=90)
        if r.status_code not in (200, 201):
            log.warning(f"Upload fallido para '{filename}': {r.text[:100]}")
            return False
        file_id = r.json().get('id')
        # Reintentos con timeout amplio — Open WebUI tarda en crear embeddings
        for attempt in range(3):
            try:
                r2 = session.post(f'{WEBUI_URL}/api/v1/knowledge/{kb_id}/file/add',
                                  json={'file_id': file_id}, timeout=120)
                if r2.status_code in (200, 201):
                    return True
                log.warning(f"  KB add intento {attempt+1} fallido ({r2.status_code}): {r2.text[:80]}")
            except Exception as e:
                log.warning(f"  KB add intento {attempt+1} timeout/error: {e}")
            if attempt < 2:
                time.sleep(5)
        return False
    finally:
        os.unlink(tmp)


# ─── Extractores de texto ─────────────────────────────────────────────────────

def extract_excel_text(path: Path, max_rows: int = 500) -> str:
    """Extrae texto legible de un archivo Excel."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        lines = [f"# {path.name}\n"]
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"## Hoja: {sheet_name}\n")
            rows_read = 0
            for row in ws.iter_rows(values_only=True):
                if rows_read >= max_rows:
                    lines.append(f"_(truncado en {max_rows} filas)_\n")
                    break
                row_vals = [str(v).strip() if v is not None else '' for v in row]
                if any(row_vals):
                    lines.append('| ' + ' | '.join(row_vals) + ' |')
                    rows_read += 1
            lines.append('')
        return '\n'.join(lines)
    except Exception as e:
        return f"# {path.name}\n_Error al leer: {e}_\n"


def extract_python_text(path: Path) -> str:
    """Extrae docstrings y estructura de un archivo Python."""
    try:
        import ast
        source = path.read_text(encoding='utf-8', errors='replace')
        lines = [f"# {path.name}\n```python\n"]
        # Incluir el código completo pero limitar a 300 líneas
        code_lines = source.splitlines()
        if len(code_lines) > 300:
            lines.extend(code_lines[:300])
            lines.append(f"\n# ... ({len(code_lines) - 300} líneas más)")
        else:
            lines.extend(code_lines)
        lines.append("\n```\n")
        return '\n'.join(lines)
    except Exception as e:
        return f"# {path.name}\n_Error: {e}_\n"


# ─── Fuente 1: PostgreSQL ─────────────────────────────────────────────────────

def build_kpi_corpus() -> list[tuple[str, str]]:
    """Genera el corpus de KPIs desde PostgreSQL (reutiliza generate_context.py)."""
    try:
        import subprocess
        result = subprocess.run(
            ['python3', os.path.join(CODE_ROOT, 'generate_context.py'),
             '--output', '/tmp/rpf_context_full.md'],
            capture_output=True, text=True, timeout=120
        )
        content = Path('/tmp/rpf_context_full.md').read_text(encoding='utf-8')
        return [('rpf_kpis_completo.md', content)]
    except Exception as e:
        log.error(f"Error generando contexto KPI: {e}")
        return []


# ─── Fuente 2: Archivos de análisis CNDC ─────────────────────────────────────

def build_cndc_corpus() -> list[tuple[str, str]]:
    """Escanea archivos ANÁLISIS RPF_*.xlsx de todos los eventos."""
    docs = []
    base = Path(RAIZ_RPF)
    if not base.exists():
        log.warning(f"No existe: {base}")
        return docs

    for sem_dir in sorted(base.iterdir()):
        if not sem_dir.is_dir():
            continue
        analisis_dir = sem_dir / 'Análisis_todos_los_eventos'
        if not analisis_dir.exists():
            analisis_dir = sem_dir / 'Analisis_todos_los_eventos'
        if not analisis_dir.exists():
            continue

        for ev_dir in sorted(analisis_dir.iterdir()):
            if not ev_dir.is_dir():
                continue
            # Buscar archivos ANÁLISIS RPF en la carpeta del evento
            for xlsx in ev_dir.rglob('ANÁLISIS RPF*.xlsx'):
                try:
                    content = extract_excel_text(xlsx, max_rows=200)
                    fname = f"analisis_{sem_dir.name}_{ev_dir.name}_{xlsx.stem}.md"
                    fname = re.sub(r'[^a-zA-Z0-9_\-.]', '_', fname)
                    docs.append((fname, content))
                    log.info(f"  CNDC: {xlsx.relative_to(base)}")
                except Exception as e:
                    log.warning(f"  Error {xlsx}: {e}")

            # Buscar archivos de resultados COBEE
            cobee_dir = ev_dir / 'Resultados_COBEE'
            if cobee_dir.exists():
                for xlsx in cobee_dir.glob('*.xlsx'):
                    try:
                        content = extract_excel_text(xlsx, max_rows=100)
                        fname = f"resultado_{sem_dir.name}_{ev_dir.name}_{xlsx.stem}.md"
                        fname = re.sub(r'[^a-zA-Z0-9_\-.]', '_', fname)
                        docs.append((fname, content))
                    except Exception as e:
                        log.warning(f"  Error {xlsx}: {e}")

    log.info(f"CNDC corpus: {len(docs)} documentos")
    return docs


# ─── Fuente 3: Documentos normativos ─────────────────────────────────────────

def build_normativa_corpus() -> list[tuple[str, str]]:
    """Busca documentos normativos (CDM, CNDC) en toda la carpeta de datos."""
    docs = []
    base = Path(DATA_ROOT)
    if not base.exists():
        return docs

    # Patrones de nombres que sugieren documentos normativos
    norm_patterns = [
        '*CDM*', '*CNDC*', '*normativa*', '*manual*', '*reglamento*',
        '*Reglamento*', '*Manual*', '*Normativa*', '*procedimiento*',
        '*Procedimiento*', '*directriz*', '*Directriz*'
    ]

    for pattern in norm_patterns:
        for f in base.rglob(pattern):
            if f.suffix.lower() in ('.xlsx', '.xls') and not f.name.startswith('~$'):
                try:
                    content = extract_excel_text(f, max_rows=300)
                    fname = re.sub(r'[^a-zA-Z0-9_\-.]', '_', f.name.replace(' ', '_')) + '.md'
                    docs.append((fname, content))
                    log.info(f"  Normativa: {f.name}")
                except Exception as e:
                    log.warning(f"  Error {f}: {e}")

    log.info(f"Normativa corpus: {len(docs)} documentos")
    return docs


# ─── Fuente 4: Código Python del sistema ─────────────────────────────────────

def build_code_corpus() -> list[tuple[str, str]]:
    """Indexa el código Python del sistema para que la IA pueda entender y corregir."""
    docs = []
    root = Path(CODE_ROOT)

    # Scripts principales a indexar
    priority_scripts = [
        'extract_kpi.py',
        'generate_context.py',
        'upload_context.py',
        'sp_sync_daemon.py',
    ]

    # Indexar scripts prioritarios primero
    for name in priority_scripts:
        p = root / name
        if p.exists():
            docs.append((f'code_{name}.md', extract_python_text(p)))
            log.info(f"  Código: {name}")

    # Buscar otros scripts Python
    for py in root.glob('*.py'):
        if py.name not in priority_scripts:
            docs.append((f'code_{py.name}.md', extract_python_text(py)))

    # Agregar README si existe
    for readme in root.glob('README*'):
        try:
            content = readme.read_text(encoding='utf-8', errors='replace')
            docs.append(('README.md', f"# README del sistema\n\n{content}"))
        except Exception:
            pass

    log.info(f"Código corpus: {len(docs)} archivos")
    return docs


# ─── Fuente 5: Resúmenes SCADA ────────────────────────────────────────────────

def build_scada_corpus() -> list[tuple[str, str]]:
    """Genera resúmenes de los datos SCADA procesados (no sube archivos crudos)."""
    docs = []
    base = Path(RAIZ_RPF)
    if not base.exists():
        return docs

    all_events = []
    for sem_dir in sorted(base.iterdir()):
        if not sem_dir.is_dir():
            continue
        analisis_dir = sem_dir / 'Análisis_todos_los_eventos'
        if not analisis_dir.exists():
            analisis_dir = sem_dir / 'Analisis_todos_los_eventos'
        if not analisis_dir.exists():
            continue
        for ev_dir in sorted(analisis_dir.iterdir()):
            scada_dir = ev_dir / 'Graficas Registro 1SEG COBEE'
            if scada_dir.exists():
                units = [f.stem for f in scada_dir.glob('*.xlsx')
                         if not f.stem[0].isdigit()]
                if units:
                    all_events.append({
                        'semestre': sem_dir.name,
                        'evento': ev_dir.name,
                        'unidades': units,
                    })

    if all_events:
        lines = ["# Datos SCADA Procesados — COBEE\n"]
        lines.append("Registro de archivos SCADA (1 segundo) procesados por evento.\n")
        lines.append("| Semestre | Evento | Unidades con datos SCADA |")
        lines.append("|----------|--------|--------------------------|")
        for ev in all_events:
            units_str = ', '.join(ev['unidades'])
            lines.append(f"| {ev['semestre']} | {ev['evento']} | {units_str} |")
        content = '\n'.join(lines)
        docs.append(('scada_disponible.md', content))
        log.info(f"SCADA corpus: {len(all_events)} eventos con datos")

    return docs


# ─── Knowledge Bases ──────────────────────────────────────────────────────────

KB_CONFIG = {
    'rpf-kpis': {
        'name':        'RPF KPIs COBEE',
        'description': 'Base de datos completa: KPIs, cumplimiento RPF, droop, frecuencias por unidad y evento',
        'builder':     build_kpi_corpus,
    },
    'rpf-cndc': {
        'name':        'RPF Análisis CNDC',
        'description': 'Archivos de análisis del CNDC por evento: métricas, cálculos, resultados de evaluación',
        'builder':     build_cndc_corpus,
    },
    'rpf-normativa': {
        'name':        'RPF Normativa CDM',
        'description': 'Documentos normativos: CDM, manuales CNDC, reglamentos de operación del SIN',
        'builder':     build_normativa_corpus,
    },
    'rpf-codigo': {
        'name':        'RPF Sistema Python',
        'description': 'Código fuente del sistema: scripts de extracción, análisis, sincronización y procesamiento',
        'builder':     build_code_corpus,
    },
    'rpf-scada': {
        'name':        'RPF SCADA Disponible',
        'description': 'Inventario de señales SCADA procesadas por evento y unidad generadora',
        'builder':     build_scada_corpus,
    },
}


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Construye corpus completo para la IA COBEE')
    parser.add_argument('--kb', choices=list(KB_CONFIG.keys()) + ['all'],
                        default='all', help='KB a construir (default: all)')
    parser.add_argument('--dry-run', action='store_true',
                        help='Listar documentos sin subir')
    parser.add_argument('--no-clear', action='store_true',
                        help='No borrar documentos previos de la KB')
    args = parser.parse_args()

    kbs_to_build = list(KB_CONFIG.keys()) if args.kb == 'all' else [args.kb]

    if not args.dry_run:
        log.info("Conectando a Open WebUI...")
        session = get_session()
        log.info("Conectado OK")
        log.info("Limpiando archivos físicos anteriores (evita duplicados)...")
        purge_all_files(session)

    # Directorio local para guardar corpus en disco (usado por expand_dataset.py)
    corpus_cache = Path(__file__).parent / 'corpus_cache'
    corpus_cache.mkdir(exist_ok=True)

    total_docs = 0
    for kb_key in kbs_to_build:
        cfg = KB_CONFIG[kb_key]
        log.info(f"\n{'='*50}")
        log.info(f"KB: {cfg['name']}")
        log.info(f"{'='*50}")

        docs = cfg['builder']()

        if args.dry_run:
            log.info(f"[DRY-RUN] {len(docs)} documentos:")
            for fname, content in docs:
                log.info(f"  {fname} ({len(content)//1024} KB)")
            total_docs += len(docs)
            continue

        if not docs:
            log.warning(f"  Sin documentos para {cfg['name']}")
            continue

        kb_id = get_or_create_kb(session, cfg['name'], cfg['description'])
        log.info(f"  KB id: {kb_id}")

        if not args.no_clear:
            clear_kb(session, kb_id)
            log.info("  Documentos anteriores eliminados")

        ok = 0
        for fname, content in docs:
            # Guardar en disco para fine-tuning (expand_dataset.py lo lee)
            (corpus_cache / fname).write_text(content, encoding='utf-8')
            if upload_text(session, kb_id, fname, content):
                ok += 1
                log.info(f"  ✓ {fname} ({len(content)//1024} KB)")
            else:
                log.warning(f"  ✗ {fname}")

        log.info(f"  Subidos: {ok}/{len(docs)}")
        total_docs += ok

    log.info(f"\nTotal documentos procesados: {total_docs}")


if __name__ == '__main__':
    main()
