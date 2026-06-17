r"""
interfaz_analisis_RPF.py (v2 - Integrada)
Interfaz Streamlit integrada para workflow completo RPF.

Módulos:
  · Extracción de datos CNDC (ExtFLujos2daO)
  · Generación de condiciones iniciales (CondInicialesPF)
  · Carga en PowerFactory (CargaCondIniciales_PF_run)

Ejecutar con:
    streamlit run ProgramasLimpio\interfaz_analisis_RPF.py
"""
import os
import sys
import glob
import re
import json
import subprocess
import threading
import time
from datetime import datetime
from plotly.subplots import make_subplots

import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go

# Detecta si la app corre en Streamlit Cloud (sin acceso a rutas Windows locales)
IS_CLOUD = not os.path.isdir(r"C:\Datos del CNDC")

# SharePoint disponible tanto en cloud como en local (para sincronización bidireccional)
try:
    import sharepoint_client as _sp
    _SP_OK = True
except Exception as _sp_err:
    _SP_OK = False
    _SP_ERR_MSG = str(_sp_err)

# Watcher de sincronización local → SharePoint (solo modo local)
if not IS_CLOUD:
    try:
        from sync_watcher import get_watcher as _get_watcher
        _WATCHER_MOD_OK = True
    except ImportError:
        _WATCHER_MOD_OK = False
else:
    _WATCHER_MOD_OK = False

# 
# MÓDULOS DE GRÁFICAS ESTÁNDARES
# 
from graph_config import DEFAULT_GRAPH_CONFIG
from graph_builders import (
    create_dual_axis_timeseries,
    create_comparison_chart,
    add_kpi_markers,
    add_pmax_marker,
    add_reference_lines,
    apply_standard_layout,
)
from kpi_calc import (
    _load_tech_map, _load_pmax_cargado, _get_pmax, _get_pmax_from_cargado, _get_rp_default,
    _rp_cfg_path, _load_rp_cfg, _cndc_kpis, _calcular_rocof,
    _is_frequency_column, _robust_col_detect, _find_pmax_time,
)

# 
# EXCEL STYLES (for formatted exports)
# 
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

_HDR_FILL = PatternFill("solid", start_color="2E4057", end_color="2E4057")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_THIN     = Side(style="thin", color="CCCCCC")
_BORDER   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CTR      = Alignment(horizontal="center", vertical="center")
_LEFT     = Alignment(horizontal="left", vertical="center")
_RIGHT    = Alignment(horizontal="right", vertical="center")

# Specific colors for KPI tables
_KPI_OK_FILL   = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE") # Light green
_KPI_WARN_FILL = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C") # Light yellow
_KPI_ERROR_FILL = PatternFill("solid", start_color="FFCCCC", end_color="FFCCCC") # Light red

def _formato_hoja_excel(ws, df, kpi_col=None, kpi_ok_val="✅ Sí", kpi_error_val="❌ No"):
    """Aplica formato visual a la hoja de Excel exportada."""
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.fill = _HDR_FILL; cell.font = _HDR_FONT; cell.alignment = _CTR; cell.border = _BORDER

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.border = _BORDER; cell.alignment = _CTR
            if kpi_col and str(ws.cell(1, c).value) == kpi_col:
                val = str(cell.value)
                if kpi_ok_val in val: cell.fill = _KPI_OK_FILL
                elif kpi_error_val in val: cell.fill = _KPI_ERROR_FILL

    for c in range(1, ws.max_column + 1):
        col_let = get_column_letter(c)
        max_w = 0
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, c).value
            if val: max_w = max(max_w, len(str(val)))
        ws.column_dimensions[col_let].width = min(max_w + 3, 50)
    ws.freeze_panes = "A2"

def _apply_excel_formatting(df, sheet_name="Sheet1", kpi_col=None, kpi_ok_val="✅ Sí", kpi_error_val="❌ No"):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]

        # Apply formatting (header, borders, auto-width)
        _formato_hoja_excel(worksheet, df, kpi_col, kpi_ok_val, kpi_error_val)

    output.seek(0)
    return output.getvalue()

def _buscar_archivo_unidad(unit_name, file_list):
    """Busca en una lista de archivos el que mejor coincida con el nombre de la unidad."""
    if not unit_name: return None
    u_norm = str(unit_name).upper().replace("SYM_", "").replace("SYM", "")
    # Coincidencia exacta (normalizada)
    for f in file_list:
        f_base = os.path.splitext(f)[0].upper().replace("SYM_", "").replace("SYM", "")
        if u_norm == f_base: return f
    # Coincidencia por sub-cadena
    for f in file_list:
        f_base = os.path.splitext(f)[0].upper().replace("SYM_", "").replace("SYM", "")
        if u_norm in f_base or f_base in u_norm: return f
    return None


# 
# CONFIGURACIÓN DE PÁGINA
# 
st.set_page_config(
    page_title="Analisis RPF",
    page_icon="⚡",  # fallback; el favicon SVG real se inyecta via components.html
    layout="wide",
    initial_sidebar_state="collapsed",
)

# 
# UI V4 — DESIGN TOKENS + HELPERS
# 

_V4_TOKENS = {
    "light": {
        "bg": "#F5F7FA", "surface": "#FFFFFF", "surfaceAlt": "#F9FAFB",
        "surfaceHover": "#F3F4F6", "border": "#E5E7EB", "borderStrong": "#D1D5DB",
        "text": "#111827", "textMuted": "#6B7280", "textSubtle": "#9CA3AF",
        "primary": "#2E5C8A", "accent": "#2563EB", "accent2": "#F97316",
        "success": "#10B981", "successBg": "#D1FAE5",
        "warning": "#F59E0B", "warningBg": "#FEF3C7",
        "danger": "#DC2626", "dangerBg": "#FEE2E2",
        "info": "#3B82F6", "infoBg": "#DBEAFE", "chartGrid": "#E5E7EB",
    },
    "dark": {
        "bg": "#0B0F19", "surface": "#141925", "surfaceAlt": "#1A2030",
        "surfaceHover": "#1F2738", "border": "#252C3D", "borderStrong": "#323A50",
        "text": "#E5E7EB", "textMuted": "#9CA3AF", "textSubtle": "#6B7280",
        "primary": "#5B8DD6", "accent": "#60A5FA", "accent2": "#FB923C",
        "success": "#34D399", "successBg": "rgba(52,211,153,0.14)",
        "warning": "#FBBF24", "warningBg": "rgba(251,191,36,0.14)",
        "danger": "#F87171", "dangerBg": "rgba(248,113,113,0.14)",
        "info": "#60A5FA", "infoBg": "rgba(96,165,250,0.14)", "chartGrid": "#252C3D",
    },
}

_V4_ICON_PATHS = {
    "bolt":     '<path d="M13 2L4 14h7l-1 8 9-12h-7l1-8z"/>',
    "activity": '<polyline points="3 12 7 12 10 4 14 20 17 12 21 12"/>',
    "chart":    '<line x1="3" y1="20" x2="21" y2="20"/><polyline points="5 16 9 11 13 14 18 7"/>',
    "sliders":  '<line x1="4" y1="6" x2="20" y2="6"/><line x1="4" y1="12" x2="20" y2="12"/><line x1="4" y1="18" x2="20" y2="18"/><circle cx="9" cy="6" r="2" fill="currentColor"/><circle cx="15" cy="12" r="2" fill="currentColor"/><circle cx="7" cy="18" r="2" fill="currentColor"/>',
    "database": '<ellipse cx="12" cy="5" rx="9" ry="3"/><path d="M3 5v6c0 1.66 4.03 3 9 3s9-1.34 9-3V5"/><path d="M3 11v6c0 1.66 4.03 3 9 3s9-1.34 9-3v-6"/>',
    "download": '<path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/>',
    "scale":    '<path d="M12 3v18"/><path d="M5 8h14"/><path d="M5 8l-3 6a3 3 0 0 0 6 0z"/><path d="M19 8l3 6a3 3 0 0 1-6 0z"/>',
    "report":   '<path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="8" y1="13" x2="16" y2="13"/><line x1="8" y1="17" x2="13" y2="17"/>',
    "palette":  '<circle cx="12" cy="12" r="9"/><circle cx="7.5" cy="10.5" r="1" fill="currentColor"/><circle cx="12" cy="7" r="1" fill="currentColor"/><circle cx="16.5" cy="10.5" r="1" fill="currentColor"/><path d="M12 21a3 3 0 0 1-3-3v-1a3 3 0 0 1 3-3h2a2 2 0 0 0 2-2v-1"/>',
    "cloud":    '<path d="M18 10a6 6 0 0 0-11.8-1A4.5 4.5 0 0 0 6 19h12a4 4 0 0 0 0-9z"/>',
    "server":   '<rect x="3" y="4" width="18" height="7" rx="1"/><rect x="3" y="13" width="18" height="7" rx="1"/>',
    "check":    '<polyline points="20 6 9 17 4 12"/>',
    "info":     '<circle cx="12" cy="12" r="9"/><line x1="12" y1="11" x2="12" y2="17"/><line x1="12" y1="7" x2="12" y2="7.01"/>',
}

def _v4_icon(name: str, size: int = 14, color: str = "currentColor", stroke: float = 1.75) -> str:
    path = _V4_ICON_PATHS.get(name, "")
    return (
        f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" '
        f'stroke="{color}" stroke-width="{stroke}" stroke-linecap="round" '
        f'stroke-linejoin="round" style="vertical-align:middle;flex-shrink:0">{path}</svg>'
    )

def _v4_t() -> dict:
    """Devuelve el dict de tokens del tema activo (light / dark)."""
    return _V4_TOKENS[st.session_state.get("ui_theme", "light")]

#  CSS template (plain str, no f-string) — avoids Python 3.12 C-tokenizer bug
# that fires when inspect.getsource() processes f-strings with subscript exprs.
# Use .format(**_v4_t()) at call time; {{ }} = literal CSS braces.
_V4_CSS_TEMPLATE = (
    "<style>"
    "@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700"
    "&family=JetBrains+Mono:wght@400;500;600&display=swap');"
    " #MainMenu, footer {{ visibility: hidden; height: 0; }}"
    " header[data-testid='stHeader'] {{"
    " background: transparent !important; height: 0 !important;"
    " overflow: visible !important; border: none !important; }}"
    " div[data-testid='stDecoration'] {{ display: none; }}"
    " div[data-testid='stToolbar'] {{"
    " position: fixed !important; top: 120px !important; right: 0 !important;"
    " left: auto !important; width: auto !important;"
    " z-index: 99999 !important; background: transparent !important;"
    " pointer-events: none !important; }}"
    " div[data-testid='stToolbar'] > * {{ pointer-events: auto !important; }}"
    " [data-testid='stSidebarCollapsedControl'] {{"
    " top: 120px !important; left: 0 !important;"
    " width: auto !important; height: auto !important;"
    " z-index: 99999 !important;"
    " position: fixed !important; visibility: visible !important;"
    " pointer-events: none !important; }}"
    " [data-testid='stSidebarCollapsedControl'] > * {{ pointer-events: auto !important; }}"
    " [data-testid='stSidebarCollapseButton'] {{"
    " visibility: visible !important; opacity: 1 !important;"
    " position: relative !important; z-index: 100 !important; }}"
    " .block-container {{ padding: 116px 20px 24px 20px !important; max-width: 1200px !important; margin-left: auto !important; margin-right: auto !important; }}"
    " html {{ overflow-y: auto !important; }}"
    " body {{ overflow-y: auto !important; overflow-x: hidden !important; }}"
    " .stApp {{ overflow: visible !important; min-height: 100vh;"
    " background: {bg} !important;"
    " font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important; }}"
    " [data-testid='stAppViewContainer'] {{ overflow: visible !important; }}"
    " section[data-testid='stMain'] {{ overflow: visible !important; background: {surface} !important; min-height: 100vh !important; }}"
    " section[data-testid='stSidebar'] {{ overflow-y: auto !important; }}"
    " .v4-topbar {{"
    " position: fixed; top: 0; left: 0; right: 0; z-index: 9999; height: 64px;"
    " background: {surface}; border-bottom: 1px solid {border};"
    " display: flex; align-items: center; justify-content: space-between;"
    " padding: 0 20px; gap: 12px; box-sizing: border-box;"
    " }}"
    " .v4-brand {{ display: flex; align-items: center; gap: 10px; }}"
    " .v4-brand-mark {{"
    " width: 38px; height: 38px; border-radius: 9px; flex-shrink: 0;"
    " background: linear-gradient(135deg, {primary}, {accent});"
    " display: flex; align-items: center; justify-content: center;"
    " }}"
    " .v4-brand-title {{ font-size: 15px; font-weight: 700; color: {text}; line-height: 1.1; }}"
    " .v4-brand-sub   {{ font-size: 11.5px; color: {textMuted}; line-height: 1.2; margin-top: 1px; }}"
    " .v4-topbar-center {{ display: flex; align-items: center; gap: 8px; flex: 1; overflow-x: auto; white-space: nowrap; min-width: 0; }}"
    " .v4-topbar-right {{ display: flex; align-items: center; gap: 8px; flex-shrink: 0; }}"
    " .v4-event-pill {{"
    " display: inline-flex; align-items: center; gap: 7px; padding: 5px 13px; height: 40px;"
    " background: {surfaceAlt}; border: 1px solid {border}; border-radius: 8px;"
    " font-size: 12px; font-weight: 500; color: {text}; flex-shrink: 0;"
    " }}"
    " .v4-event-pill.danger {{ background: {dangerBg}; border-color: {danger}; }}"
    " .v4-event-label {{ font-size: 11px; font-weight: 700; color: {textSubtle}; text-transform: uppercase; letter-spacing: .06em; }}"
    " .v4-event-val   {{ font-size: 14px; font-weight: 600; color: {text}; }}"
    " .v4-event-val.danger {{ color: {danger}; }}"
    " .v4-sep {{ width: 1px; height: 16px; background: {border}; flex-shrink: 0; }}"
    " .v4-mode-badge {{"
    " display: inline-flex; align-items: center; gap: 6px; padding: 4px 10px;"
    " border-radius: 6px; font-size: 11px; font-weight: 600;"
    " border: 1px solid {border}; color: {textMuted};"
    " }}"
    " .v4-mode-badge.local {{ border-color: {success}; color: {success}; background: {successBg}; }}"
    " .v4-mode-badge.cloud {{ border-color: {info};    color: {info};    background: {infoBg}; }}"
    " .v4-stepper {{"
    " position: fixed; top: 64px; left: 0; right: 0; z-index: 9998;"
    " background: {surface}; border-bottom: 1px solid {border};"
    " padding: 0 20px; overflow-x: auto; white-space: nowrap;"
    " }}"
    " .v4-stepper-inner {{ display: inline-flex; align-items: center; height: 52px; gap: 0; }}"
    " .v4-step {{"
    " display: inline-flex; align-items: center; gap: 7px;"
    " padding: 7px 11px; border-radius: 7px;"
    " font-size: 12px; font-weight: 500; color: {textMuted};"
    " white-space: nowrap; cursor: default;"
    " }}"
    " .v4-step.active {{ background: {surfaceHover}; color: {text}; font-weight: 600; border-bottom: 2px solid {primary}; border-radius: 7px 7px 0 0; }}"
    " .v4-step.past   {{ color: {success}; }}"
    " .v4-step.disabled {{ opacity: 0.45; }}"
    " .v4-step-num {{"
    " width: 22px; height: 22px; border-radius: 50%; flex-shrink: 0;"
    " display: inline-flex; align-items: center; justify-content: center;"
    " font-size: 11px; font-weight: 700; font-variant-numeric: tabular-nums;"
    " background: {surfaceAlt}; border: 1px solid {border}; color: {textMuted};"
    " }}"
    " .v4-step.active .v4-step-num {{ background: {primary}; border-color: {primary}; color: #FFF; }}"
    " .v4-step.past   .v4-step-num {{ background: {success}; border-color: {success}; color: #FFF; font-size: 10px; }}"
    " .v4-step-badge {{"
    " font-size: 9px; padding: 1px 4px; border-radius: 3px; font-weight: 600;"
    " background: {surfaceAlt}; border: 1px solid {border}; color: {textSubtle};"
    " text-transform: uppercase; letter-spacing: 0.04em;"
    " }}"
    " .v4-connector {{ width: 20px; height: 1px; background: {border}; display: inline-block; vertical-align: middle; flex-shrink: 0; }}"
    " .v4-connector.past {{ background: {success}; }}"
    " .v4-unit-bar {{"
    " position: fixed !important; top: 116px !important; left: 0 !important; right: 0 !important;"
    " height: 44px !important; z-index: 9997 !important;"
    " background: {surface} !important; border-bottom: 1px solid {border} !important;"
    " display: flex !important; align-items: center !important; gap: 16px !important;"
    " padding: 0 24px 0 170px !important; box-shadow: 0 1px 4px rgba(0,0,0,0.06) !important;"
    " pointer-events: none !important;"
    " }}"
    " .v4-unit-ctx {{"
    " display: flex; align-items: center; gap: 16px; flex-wrap: wrap;"
    " padding: 4px 0;"
    " }}"
    " .v4-unit-dot {{"
    " display: inline-block; width: 8px; height: 8px; border-radius: 50%;"
    " background: {success}; box-shadow: 0 0 0 3px {successBg};"
    " margin-right: 2px; vertical-align: middle; flex-shrink: 0;"
    " }}"
    " .v4-unit-name {{"
    " font-size: 14px; font-weight: 700; color: {text};"
    " font-family: 'JetBrains Mono', ui-monospace, monospace;"
    " font-variant-numeric: tabular-nums;"
    " }}"
    " .v4-stat {{ display: flex; flex-direction: column; line-height: 1.2; }}"
    " .v4-stat-label {{ font-size: 10px; font-weight: 600; color: {textSubtle}; text-transform: uppercase; letter-spacing: 0.05em; }}"
    " .v4-stat-value {{ font-size: 13px; font-weight: 700; color: {text}; font-variant-numeric: tabular-nums; }}"
    " .v4-stat-unit {{ font-size: 11px; color: {textMuted}; font-weight: 500; margin-left: 2px; }}"
    " .v4-stat-sep {{ width: 1px; height: 28px; background: {border}; flex-shrink: 0; }}"
    " .v4-block-wrap {{ padding: 10px 14px 0 14px; }}"
    " .v4-breadcrumb {{ display: flex; align-items: center; gap: 5px; font-size: 10.5px; color: {textMuted}; margin-bottom: 6px; }}"
    " .v4-bc-sep {{ color: {textSubtle}; }}"
    " .v4-bc-active {{ color: {text}; font-weight: 600; }}"
    " .v4-block-head {{ display: flex; align-items: flex-start; gap: 10px; margin-bottom: 8px; }}"
    " .v4-block-num {{"
    " width: 36px; height: 36px; border-radius: 8px; flex-shrink: 0;"
    " background: {surface}; border: 1px solid {border};"
    " display: flex; align-items: center; justify-content: center;"
    " font-size: 14px; font-weight: 700; color: {primary};"
    " font-variant-numeric: tabular-nums;"
    " font-family: 'JetBrains Mono', ui-monospace, monospace;"
    " }}"
    " .v4-block-title {{ font-size: 16px; font-weight: 700; color: {text}; line-height: 1.2; letter-spacing: -0.01em; }}"
    " .v4-block-sub   {{ font-size: 11.5px; color: {textMuted}; margin-top: 2px; max-width: 720px; line-height: 1.4; }}"
    " .v4-banner {{"
    " display: flex; align-items: flex-start; gap: 8px;"
    " padding: 6px 10px; border-radius: 7px;"
    " font-size: 11.5px; color: {text}; line-height: 1.4;"
    " margin: 0 0 8px 0;"
    " }}"
    " .v4-banner.info    {{ background: {infoBg};    border: 1px solid {info}; }}"
    " .v4-banner.warning {{ background: {warningBg}; border: 1px solid {warning}; }}"
    " section[data-testid='stSidebar'] > div:first-child {{"
    " background: {surface} !important; border-right: 1px solid {border} !important;"
    " }}"
    " .v4-nav-group-label {{"
    " font-size: 10px; font-weight: 700; color: {textSubtle};"
    " text-transform: uppercase; letter-spacing: 0.08em;"
    " padding: 10px 4px 3px 4px; display: block;"
    " }}"
    " section[data-testid='stSidebar'] .stButton > button {{"
    " width: 100%; background: transparent !important; border: none !important;"
    " color: {textMuted} !important; font-size: 11.5px !important;"
    " font-weight: 500 !important; text-align: left !important;"
    " padding: 5px 8px !important; border-radius: 6px !important;"
    " height: auto !important; line-height: 1.3 !important;"
    " justify-content: flex-start !important;"
    " }}"
    " section[data-testid='stSidebar'] .stButton > button:hover {{"
    " background: {surfaceHover} !important; color: {text} !important;"
    " }}"
    " section[data-testid='stSidebar'] .stButton > button[data-testid*='primary'] {{"
    " background: {primary} !important; color: #FFF !important; font-weight: 600 !important;"
    " }}"
    " section[data-testid='stSidebar'] .stButton > button:disabled {{"
    " opacity: 0.45 !important; cursor: not-allowed !important;"
    " }}"
    " .v4-content {{ padding: 0 14px 14px 14px; }}"
    " .stTextInput input, .stSelectbox > div > div, .stNumberInput input {{"
    " background: {surfaceAlt} !important; color: {text} !important;"
    " border-color: {border} !important;"
    " font-size: 12px !important; padding: 4px 8px !important; min-height: 32px !important;"
    " }}"
    " .stSelectbox > div > div {{ min-height: 32px !important; }}"
    " .stNumberInput > div {{ min-height: 32px !important; }}"
    " .stTextInput > div {{ min-height: 32px !important; }}"
    " .stTextInput label, .stSelectbox label, .stNumberInput label,"
    " .stSlider label, .stDateInput label, .stTextArea label {{"
    " font-size: 11.5px !important; font-weight: 600 !important;"
    " color: {textMuted} !important; margin-bottom: 2px !important;"
    " }}"
    " .stButton > button {{"
    " font-size: 12px !important; padding: 4px 12px !important;"
    " min-height: 32px !important; height: auto !important;"
    " border-radius: 6px !important; font-weight: 500 !important;"
    " }}"
    " .stButton > button[kind='primary'] {{"
    " background: {primary} !important; border-color: {primary} !important;"
    " color: #FFF !important; font-weight: 600 !important;"
    " }}"
    " .stDownloadButton > button {{"
    " font-size: 12px !important; padding: 4px 12px !important;"
    " min-height: 32px !important; height: auto !important;"
    " border-radius: 6px !important;"
    " }}"
    " .stExpander > details > summary {{"
    " font-size: 12px !important; padding: 6px 10px !important;"
    " }}"
    " .stExpander > details {{ padding: 0 !important; }}"
    " .stSlider > div {{ padding: 0 !important; }}"
    " .stCheckbox label, .stToggle label {{ font-size: 12px !important; }}"
    " .stCheckbox label span, .stToggle label span {{ color: {text} !important; }}"
    " .stMarkdown p, .stMarkdown li {{ font-size: 12.5px !important; color: {text} !important; }}"
    " .stMarkdown h1 {{ font-size: 17px !important; color: {text} !important; }}"
    " .stMarkdown h2 {{ font-size: 15px !important; font-weight: 700 !important; color: {text} !important; }}"
    " .stMarkdown h3 {{ font-size: 13px !important; color: {text} !important; }}"
    " .stMarkdown h4 {{ font-size: 12px !important; color: {text} !important; }}"
    " .stCaption {{ font-size: 11px !important; color: {textMuted} !important; }}"
    " .stExpander {{ background: {surface} !important; border-color: {border} !important; }}"
    " .stExpander summary {{ color: {text} !important; }}"
    " .stAlert {{ background: {surfaceAlt} !important; border-color: {border} !important;"
    " font-size: 12px !important; padding: 8px 12px !important; }}"
    " .stAlert p {{ font-size: 12px !important; }}"
    " .stTabs [data-baseweb='tab-list'] {{"
    " background: {surface} !important; border-bottom: 1px solid {border} !important; }}"
    " .stTabs [data-baseweb='tab'] {{ color: {textMuted} !important; font-size: 12px !important; }}"
    " .stTabs [aria-selected='true'] {{"
    " color: {primary} !important; border-bottom-color: {primary} !important; }}"
    " .stTabs [data-baseweb='tab-panel'] {{"
    " background: {surface} !important; padding-top: 8px !important; }}"
    " .stDataFrame {{ font-size: 11.5px !important; }}"
    " .stMetric {{ padding: 8px 10px !important; }}"
    " .stMetric label {{ font-size: 10.5px !important; color: {textMuted} !important; }}"
    " .stMetric [data-testid='stMetricValue'] {{ font-size: 18px !important; }}"
    " .stMetric [data-testid='stMetricDelta'] {{ font-size: 11px !important; }}"
    " div[data-testid='column'] {{ gap: 8px !important; }}"
    " .element-container {{ margin-bottom: 6px !important; }}"
    " [data-testid='stHorizontalBlock'] {{ gap: 8px !important; }}"
    " [data-testid='stVerticalBlockBorderWrapper'] {{ padding: 0 !important; }}"
    " .stRadio > label {{ font-size: 11.5px !important; font-weight: 600 !important;"
    " color: {textMuted} !important; margin-bottom: 2px !important; }}"
    " .stRadio [data-testid='stWidgetLabel'] {{ font-size: 11.5px !important; font-weight: 600 !important; }}"
    " .stRadio > div > label {{ font-size: 12px !important; padding: 2px 4px !important; }}"
    " .stMultiSelect label {{ font-size: 11.5px !important; font-weight: 600 !important;"
    " color: {textMuted} !important; margin-bottom: 2px !important; }}"
    " .stMultiSelect [data-baseweb='select'] > div {{"
    " min-height: 32px !important; font-size: 12px !important; }}"
    " .stMultiSelect [data-baseweb='tag'] {{ font-size: 11px !important;"
    " padding: 0 6px !important; height: 20px !important; }}"
    " hr {{ margin: 6px 0 !important; border-color: {border} !important; }}"
    " .stDivider {{ margin: 6px 0 !important; }}"
    " .stExpander [data-testid='stExpanderDetails'] {{ padding: 8px 10px !important; }}"
    " [data-testid='stExpanderToggleIcon'] {{ width: 14px !important; height: 14px !important; }}"
    " .stNumberInput [data-testid='stNumberInputStepDown'],"
    " .stNumberInput [data-testid='stNumberInputStepUp'] {{"
    " min-height: 16px !important; padding: 0 4px !important; font-size: 10px !important; }}"
    " .stSelectbox [data-baseweb='select'] > div {{"
    " min-height: 32px !important; font-size: 12px !important; }}"
    " .stSuccess, .stWarning, .stInfo, .stError {{"
    " padding: 6px 10px !important; font-size: 12px !important; }}"
    " [data-testid='stPopover'] button {{ font-size: 12px !important;"
    " padding: 4px 8px !important; min-height: 28px !important; }}"
    " section[data-testid='stSidebar'] {{"
    " position: fixed !important;"
    " top: 116px !important; left: 0 !important;"
    " height: calc(100vh - 116px) !important;"
    " overflow-y: auto !important; overflow-x: hidden !important;"
    " background: {surface} !important;"
    " z-index: 9997 !important; }}"
    " section[data-testid='stSidebar'] > div:first-child {{ padding-top: 12px !important; }}"
    " section[data-testid='stSidebar'] .stToggle label {{"
    " color: {textMuted} !important; font-size: 12px !important;"
    " }}"
    " section[data-testid='stSidebar'] hr {{ border-color: {border} !important; opacity: 1; }}"
    " .v4-topbar-disc-label {{"
    " font-size: 9px; font-weight: 800; letter-spacing: 0.08em; text-transform: uppercase;"
    " color: {danger}; white-space: nowrap; flex-shrink: 0;"
    " }}"
    " .v4-disc-chip {{"
    " display: inline-flex; align-items: center; gap: 5px; padding: 4px 9px; height: 30px;"
    " background: {dangerBg}; border: 1px solid {danger}; border-radius: 6px;"
    " font-size: 11.5px; font-weight: 700; color: {danger}; white-space: nowrap; flex-shrink: 0;"
    " }}"
    " .v4-disc-chip-mw {{ font-weight: 400; opacity: 0.8; margin-left: 3px; }}"
    " .v4-section-head {{"
    " padding: 6px 0 5px 0; border-bottom: 1px solid {border}; margin-bottom: 8px; }}"
    " .v4-section-title {{"
    " font-size: 13px; font-weight: 700; color: {text}; display: flex; align-items: center; gap: 6px; }}"
    " .v4-section-sub {{"
    " font-size: 11px; color: {textMuted}; margin-top: 2px; }}"
    " .v4-ctx-bar {{"
    " display: flex; align-items: center; gap: 10px; padding: 4px 10px;"
    " background: {surfaceHover}; border: 1px solid {border}; border-radius: 6px;"
    " margin-bottom: 7px; }}"
    " .v4-ctx-unit {{ font-size: 13px; font-weight: 800; font-family: 'JetBrains Mono', monospace;"
    " color: {primary}; }}"
    " .v4-ctx-sep {{ color: {border}; font-size: 12px; }}"
    " .v4-ctx-item {{ font-size: 11px; color: {textMuted}; }}"
    " .v4-ctx-item b {{ color: {text}; }}"
    " .v4-tab-bar {{"
    " display: flex; gap: 0; border-bottom: 2px solid {border}; margin-bottom: 8px;"
    " overflow-x: auto; white-space: nowrap; }}"
    " .v4-tab-btn {{"
    " display: inline-flex; align-items: center; gap: 4px; padding: 5px 12px 7px 12px;"
    " font-size: 11.5px; font-weight: 500; color: {textMuted}; background: none;"
    " border: none; border-bottom: 2px solid transparent; margin-bottom: -2px;"
    " cursor: pointer; white-space: nowrap; transition: color .15s; flex-shrink: 0;"
    " }}"
    " .v4-tab-btn.active {{"
    " color: {primary} !important; font-weight: 700 !important;"
    " border-bottom: 2px solid {primary} !important; }}"
    " .v4-tab-btn:hover {{ color: {text} !important; }}"
    " [data-testid='stDataFrame'] {{ max-width: 100% !important; overflow-x: auto !important; }}"
    " .stDataFrameResizable {{"
    " width: fit-content !important; margin-left: auto !important; margin-right: auto !important; }}"
    " [data-testid='stDataFrame'] * {{ font-size: 12px !important; }}"
    " [data-testid='stDataFrame'] [role='row'] {{"
    " min-height: 28px !important; max-height: 28px !important; }}"
    " [data-testid='stDataFrame'] [role='columnheader'] {{"
    " min-height: 30px !important; font-size: 11.5px !important;"
    " font-weight: 700 !important; }}"
    " [data-testid='stDataFrame'] [role='gridcell'] {{ padding: 2px 8px !important; }}"
    " .element-container:has(.v4-unit-select-marker) {{"
    " height: 0 !important; margin: 0 !important; padding: 0 !important; overflow: hidden !important; }}"
    " .element-container:has(.v4-unit-select-marker) + .element-container {{"
    " position: fixed !important; top: 119px !important; left: 14px !important;"
    " right: auto !important; z-index: 9998 !important;"
    " width: 148px !important; max-width: 148px !important;"
    " height: 34px !important; margin: 0 !important; padding: 0 !important;"
    " overflow: visible !important; }}"
    " .element-container:has(.v4-unit-select-marker) + .element-container .stSelectbox {{"
    " width: 148px !important; max-width: 148px !important;"
    " margin: 0 !important; padding: 0 !important; }}"
    " .element-container:has(.v4-unit-select-marker) + .element-container .stSelectbox label {{"
    " display: none !important; }}"
    " .element-container:has(.v4-unit-select-marker) + .element-container"
    " [data-baseweb='select'] {{"
    " width: 148px !important; max-width: 148px !important; }}"
    " .element-container:has(.v4-unit-select-marker) + .element-container"
    " [data-baseweb='select'] > div {{"
    " width: 148px !important; max-width: 148px !important;"
    " min-height: 32px !important; height: 32px !important; padding: 0 8px !important;"
    " border-color: {primary} !important; background: {surfaceHover} !important;"
    " font-size: 13px !important; font-weight: 700 !important;"
    " font-family: 'JetBrains Mono', monospace !important;"
    " color: {primary} !important; border-radius: 6px !important; border-width: 1.5px !important;"
    " cursor: pointer !important; pointer-events: auto !important; }}"
    " .element-container:has(.v4-unit-select-marker) + .element-container span {{"
    " color: {primary} !important; font-weight: 700 !important; font-size: 13px !important; }}"
    " .element-container:has(.v4-unit-select-marker) + .element-container * {{"
    " pointer-events: auto !important; }}"
    " .v4-card {{"
    " background: {surface}; border: 1px solid {border}; border-radius: 8px;"
    " padding: 12px 14px; box-shadow: 0 1px 3px rgba(0,0,0,0.06); }}"
    " .v4-card-sm {{"
    " background: {surface}; border: 1px solid {border}; border-radius: 6px;"
    " padding: 8px 10px; box-shadow: 0 1px 2px rgba(0,0,0,0.04); }}"
    " .v4-chip-tech {{"
    " display: inline-flex; align-items: center; padding: 2px 8px; border-radius: 4px;"
    " font-size: 11px; font-weight: 600; line-height: 1.4; }}"
    " .v4-chip-hidro  {{ background: #DBEAFE; color: #1D4ED8; border: 1px solid #BFDBFE; }}"
    " .v4-chip-termo  {{ background: #FEE2E2; color: #B91C1C; border: 1px solid #FECACA; }}"
    " .v4-chip-gas    {{ background: #FEF3C7; color: #92400E; border: 1px solid #FDE68A; }}"
    " .v4-chip-eolico {{ background: #D1FAE5; color: #065F46; border: 1px solid #A7F3D0; }}"
    " .v4-chip-solar  {{ background: #FEF9C3; color: #854D0E; border: 1px solid #FEF08A; }}"
    " .stButton > button:hover {{"
    " background: {surfaceHover} !important; border-color: {borderStrong} !important;"
    " transition: background 0.12s, border-color 0.12s; }}"
    " .stButton > button[kind='primary']:hover {{"
    " filter: brightness(1.08) !important; }}"
    " .stTextInput input:focus, .stNumberInput input:focus {{"
    " border-color: {primary} !important;"
    " box-shadow: 0 0 0 2px {infoBg} !important; outline: none !important; }}"
    " .stSelectbox [data-baseweb='select'] > div:focus-within {{"
    " border-color: {primary} !important;"
    " box-shadow: 0 0 0 2px {infoBg} !important; }}"
    " .stMetric [data-testid='stMetricValue'] {{"
    " font-size: 20px !important; font-weight: 700 !important;"
    " color: {text} !important; font-variant-numeric: tabular-nums; }}"
    " .stMetric label {{ letter-spacing: 0.04em !important; text-transform: uppercase !important; }}"
    " .stDataFrameResizable {{"
    " margin-left: auto !important; margin-right: auto !important; }}"
    " [data-testid='stDataFrame'] {{"
    " margin-left: auto !important; margin-right: auto !important; }}"
    "</style>"
)

def _inject_v4_css():
    """Inyecta el CSS del tema v4. Usa template + .format() para evitar el bug
    del tokenizador C de Python 3.12 con f-strings y expresiones de subscript."""
    st.markdown(_V4_CSS_TEMPLATE.format(**_v4_t()), unsafe_allow_html=True)


def _v4_tab_bar(tab_defs: list, block_key: str) -> str:
    """Barra de tabs persistente usando session_state.
    tab_defs: [{"id": "t1", "label": "Tab", "icon": "activity"}, ...]
    Retorna el id del tab activo.
    """
    sk  = f"v4_tab_{block_key}"
    ids = [td["id"] for td in tab_defs]
    if sk not in st.session_state or st.session_state[sk] not in ids:
        st.session_state[sk] = ids[0]
    active = st.session_state[sk]

    # Renderizar barra de tabs: botones Streamlit nativos en columnas
    _tab_cols = st.columns(len(tab_defs))
    for td, _col in zip(tab_defs, _tab_cols):
        with _col:
            if st.button(
                td["label"],
                key=f"{sk}_{td['id']}",
                type="primary" if td["id"] == active else "secondary",
                use_container_width=True,
            ):
                st.session_state[sk] = td["id"]
                st.rerun()

    return active


def _v4_section_head(title: str, description: str = "", icon: str = ""):
    """Header estandarizado dentro de un tab/sección."""
    t         = _v4_t()
    icon_html = _v4_icon(icon, 14, t["primary"]) + " " if icon else ""
    desc_html = (
        f'<div style="font-size:11px;color:{t["textMuted"]};margin-top:2px">{description}</div>'
        if description else ""
    )
    st.markdown(
        f'<div style="padding:6px 0 5px 0;border-bottom:1px solid {t["border"]};margin-bottom:8px">'
        f'<div style="font-size:13px;font-weight:700;color:{t["text"]};display:flex;align-items:center;gap:6px">{icon_html}{title}</div>'
        f'{desc_html}</div>',
        unsafe_allow_html=True,
    )


def _v4_inline_ctx(unit_name: str, evento: str, pmax_mw: float):
    """Barra de contexto inline respeta el tema (reemplaza las hardcoded dark)."""
    t = _v4_t()
    st.markdown(
        f'<div style="display:flex;align-items:center;gap:10px;padding:4px 10px;'
        f'background:{t["surfaceHover"]};border:1px solid {t["border"]};'
        f'border-radius:6px;margin-bottom:7px">'
        f'<span style="font-size:13px;font-weight:800;font-family:JetBrains Mono,monospace;color:{t["primary"]}">'
        f'{unit_name}</span>'
        f'<span style="color:{t["border"]};font-size:12px">|</span>'
        f'<span style="font-size:11px;color:{t["textMuted"]}">Evento '
        f'<b style="color:{t["text"]}">{evento}</b></span>'
        f'<span style="color:{t["border"]};font-size:12px">|</span>'
        f'<span style="font-size:11px;color:{t["textMuted"]}">Pmax '
        f'<b style="color:{t["text"]}">{pmax_mw:.1f} MW</b></span>'
        f'</div>',
        unsafe_allow_html=True,
    )


_V4_BLOQUES = [
    {"id": "modelo_base",           "num": "00", "short": "Modelo",     "label": "Datos del Modelo",     "icon": "database", "grupo": "Setup",    "pf": True},
    {"id": "carga_datos",           "num": "01", "short": "Carga",      "label": "Carga de Datos",       "icon": "download", "grupo": "Setup",    "pf": True},
    {"id": "config_unidades",       "num": "02", "short": "DSL",        "label": "Parámetros DSL",       "icon": "sliders",  "grupo": "Setup",    "pf": False},
    {"id": "analisis_datos",        "num": "03", "short": "SCADA",      "label": "Análisis SCADA/EMF",   "icon": "activity", "grupo": "Análisis", "pf": False},
    {"id": "analisis_simulacion",   "num": "04", "short": "Sim",        "label": "Análisis Simulación",  "icon": "chart",    "grupo": "Análisis", "pf": True},
    {"id": "comparativa_real_simu", "num": "05", "short": "Real vs Sim","label": "Real vs Simulación",   "icon": "scale",    "grupo": "Análisis", "pf": True},
    {"id": "kpi_historico",         "num": "06", "short": "KPI",        "label": "Histórico RPF",        "icon": "chart",    "grupo": "Análisis", "pf": False},
    {"id": "reporte_tecnico",       "num": "07", "short": "Reporte",    "label": "Reporte Técnico",      "icon": "report",   "grupo": "Salida",   "pf": False},
    {"id": "config_global",         "num": "08", "short": "Config",     "label": "Configuración",        "icon": "sliders",  "grupo": "Salida",   "pf": False},
]

@st.cache_data(ttl=120, show_spinner=False)
def _load_event_header_info(ev_path: str, n_evento: str, raiz: str, semestre: str):
    """Lee fecha, unidades disparadas y potencia desconectada para la topbar.

    Fuente primaria : datos_cargados_Ev{n}*.xlsx → Resumen_Cargado + pgini_GEN_FINAL
      - Resumen  fila "Disparo"                → nombres de unidades disparadas
      - Resumen  fila "p_desc registrado (MW)" → potencia total desconectada
      - Resumen  fila "Fecha y hora"           → fecha del evento
      - pgini_GEN_FINAL col pgini_MW           → MW individual de cada unidad disparada

    Fuente fallback: condiciones_iniciales_*.xlsx (pgini < 0) + Tabla_Eventos_*.xlsx
    """
    import glob as _glob
    import re as _re
    fecha_str: str       = ""
    disparo_units: list  = []
    p_desc_mw: float     = 0.0

    #  Fuente primaria: datos_cargados_Ev{n}*.xlsx 
    # Preferir el archivo SIN sufijo _ajustado (tiene Resumen_Cargado con Disparo)
    _found_primary = False
    if ev_path and os.path.isdir(ev_path):
        _dc_all = _glob.glob(os.path.join(ev_path, f"datos_cargados_Ev{n_evento}*.xlsx"))
        _dc_files = sorted(
            _dc_all,
            key=lambda p: (1 if "ajustado" in os.path.basename(p).lower() else 0,
                           -os.path.getmtime(p)),
        )
        for _dc_path in _dc_files:
            try:
                _xl = pd.ExcelFile(_dc_path, engine="calamine")
                if "Resumen_Cargado" not in _xl.sheet_names:
                    continue
                _df_res = _xl.parse("Resumen_Cargado", header=None)

                # Parsear filas clave del Resumen
                _disparo_str = ""
                for _, _row in _df_res.iterrows():
                    if len(_row) < 2:
                        continue
                    _k = str(_row.iloc[0]).strip()
                    _v = str(_row.iloc[1]).strip() if _row.iloc[1] is not None else ""
                    if _k in ("Fecha y hora",):
                        fecha_str = _v
                    elif _k in ("Disparo",):
                        _disparo_str = _v
                    elif "p_desc registrado" in _k.lower():
                        try:
                            p_desc_mw = float(_v)
                        except ValueError:
                            pass

                # Parsear string de disparo: "SJS01, SJS02, SJE01 y SJE02"
                if _disparo_str and _disparo_str not in ("nan", "—", ""):
                    _raw_names = [
                        n.strip()
                        for n in _re.split(r"[,;]+|\s+y\s+", _disparo_str)
                        if n.strip()
                    ]
                    # Buscar MW individual en pgini_GEN_FINAL
                    _pgini_map = {}
                    if "pgini_GEN_FINAL" in _xl.sheet_names:
                        _df_pg = _xl.parse("pgini_GEN_FINAL")
                        if "loc_name PF" in _df_pg.columns and "pgini_MW" in _df_pg.columns:
                            for _, _pr in _df_pg.iterrows():
                                _loc = str(_pr["loc_name PF"]).strip()
                                _bare = _loc.replace("sym_", "").replace("SYM_", "")
                                try:
                                    _pgini_map[_bare.upper()] = float(_pr["pgini_MW"])
                                except (ValueError, TypeError):
                                    pass

                    for _nm in _raw_names:
                        # El nombre CNDC puede traer prefijo "CC" (ej. "CCERI30")
                        # que no existe en loc_name PF (sym_ERI30 -> ERI30).
                        _nm_clean = _re.sub(r"^CC([A-Z])", r"\1", _nm)
                        _mw = _pgini_map.get(
                            _nm.upper(), _pgini_map.get(_nm_clean.upper(), 0.0)
                        )
                        disparo_units.append((_nm, _mw))

                if fecha_str or disparo_units or p_desc_mw > 0:
                    _found_primary = True
                    break
            except Exception:
                continue

    #  Fallback: condiciones_iniciales_*.xlsx + Tabla_Eventos_*.xlsx 
    if not _found_primary and ev_path and os.path.isdir(ev_path):
        _ci_files = sorted(
            _glob.glob(os.path.join(ev_path, "condiciones_iniciales_*.xlsx")),
            key=os.path.getmtime, reverse=True,
        )
        for _ci_path in _ci_files:
            try:
                _ef = pd.ExcelFile(_ci_path, engine="calamine")
                # fecha
                for _sn in _ef.sheet_names:
                    if _sn.lower() == "resumen":
                        _df_r = _ef.parse(_sn, header=None)
                        for _, _row in _df_r.iterrows():
                            if len(_row) < 2:
                                continue
                            _k = str(_row.iloc[0]).strip()
                            _v = str(_row.iloc[1]).strip()
                            if _k == "Fecha y hora":
                                fecha_str = _v
                                break
                        break
                # unidades con pgini < 0
                _sheets_pg = [s for s in _ef.sheet_names if "pgini_gen" in s.lower()]
                _sheets_pg.sort(key=lambda s: (0 if "final" in s.lower() else 1))
                for _sn2 in _sheets_pg:
                    _df_pg = _ef.parse(_sn2)
                    _cn, _cm = None, None
                    for _c in _df_pg.columns:
                        _cl = str(_c).lower()
                        if _cn is None and any(k in _cl for k in ("loc_name", "nombre", "name")):
                            _cn = _c
                        if _cm is None and any(k in _cl for k in ("pgini", "p_mw", "_mw")):
                            _cm = _c
                    if _cn and _cm:
                        for _, _pr in _df_pg.iterrows():
                            try:
                                _mw = float(_pr[_cm])
                            except (ValueError, TypeError):
                                continue
                            if _mw < 0:
                                _n = str(_pr[_cn]).strip()
                                for _pfx in ("sym_", "WT_", "wt_"):
                                    _n = _n.replace(_pfx, "")
                                disparo_units.append((_n, abs(_mw)))
                        if disparo_units:
                            break
                break
            except Exception:
                continue

    # fallback p_desc desde Tabla_Eventos
    if p_desc_mw == 0.0 and raiz and semestre and n_evento:
        try:
            _tev_files = sorted(
                _glob.glob(os.path.join(raiz, semestre, "Tabla_Eventos_*.xlsx")),
                key=os.path.getmtime, reverse=True,
            )
            if _tev_files:
                _df_tev = pd.read_excel(_tev_files[0], header=0, engine="calamine")
                _pdesc_col = None
                for _c in _df_tev.columns:
                    if any(k in str(_c).lower() for k in ("desc", "desconect", "pdesc", "delta")):
                        _pdesc_col = _c
                        break
                _ev_col = _df_tev.columns[0]
                for _, _tr in _df_tev.iterrows():
                    try:
                        if int(_tr[_ev_col]) == int(n_evento):
                            p_desc_mw = float(_tr[_pdesc_col]) if _pdesc_col else float(_tr.iloc[3])
                            break
                    except (ValueError, TypeError):
                        pass
        except Exception:
            pass

    return fecha_str, disparo_units, p_desc_mw


def _build_topbar_html() -> str:
    """Construye HTML del topbar fijo (sin llamar a st.markdown)."""
    t          = _v4_t()
    sem        = st.session_state.get("semestre_global") or "—"
    ev         = st.session_state.get("evento_global")  or "—"
    ev_path    = st.session_state.get("ev_path_global") or ""
    n_ev       = st.session_state.get("n_evento_global") or ""
    raiz       = st.session_state.get("cfg_RAIZ", _cfg.get("RAIZ", ""))
    mode_cls   = "local" if not IS_CLOUD else "cloud"
    mode_label = "Local + PF" if not IS_CLOUD else "Nube"
    mode_icon  = _v4_icon("server", 13) if not IS_CLOUD else _v4_icon("cloud", 13)

    # Cargar info del evento (cached — no bloquea el render)
    fecha_str    = ""
    disparo_units: list = []
    p_desc_mw    = 0.0
    if ev_path and n_ev:
        try:
            fecha_str, disparo_units, p_desc_mw = _load_event_header_info(
                ev_path, n_ev, raiz, sem
            )
        except Exception:
            pass

    # Formatear fecha: solo dd/mm/aaaa hh:mm
    fecha_disp = ""
    if fecha_str and fecha_str not in ("—", "nan", ""):
        # "2024-05-15 14:23:00" → "15/05/2024 14:23"
        try:
            _parts = fecha_str.strip().split(" ")
            _d = _parts[0].replace("-", "/")
            # invertir si viene yyyy/mm/dd
            _dp = _d.split("/")
            if len(_dp) == 3 and len(_dp[0]) == 4:
                _d = f"{_dp[2]}/{_dp[1]}/{_dp[0]}"
            if len(_parts) > 1:
                _t = ":".join(_parts[1].split(":")[:2])
                fecha_disp = f"{_d} {_t}"
            else:
                fecha_disp = _d
        except Exception:
            fecha_disp = fecha_str[:10]

    # Construir pill central con semestre + evento + fecha inline
    fecha_suffix = f'<div class="v4-sep"></div><span class="v4-event-val" style="opacity:0.65;font-weight:400">{fecha_disp}</span>' if fecha_disp else ""

    # Construir chips de unidades desconectadas
    disc_chips_html = ""
    if disparo_units:
        chips = ""
        for _unit_name, _unit_mw in disparo_units:
            chips += (
                f'<span class="v4-disc-chip">'
                f'{_unit_name}'
                f'<span class="v4-disc-chip-mw">{_unit_mw:.0f}&nbsp;MW</span>'
                f'</span>'
            )
        disc_chips_html = (
            f'<div class="v4-sep"></div>'
            f'<span class="v4-topbar-disc-label">DESCONECTADAS</span>'
            f'{chips}'
        )
    elif p_desc_mw > 0:
        # fallback: solo mostrar ΔP total si no hay desglose por unidad
        disc_chips_html = (
            f'<div class="v4-sep"></div>'
            f'<div class="v4-event-pill danger">'
            f'{_v4_icon("bolt", 13, t["danger"])}'
            f'<span class="v4-event-label">&#916;P</span>'
            f'<span class="v4-event-val danger">{p_desc_mw:.1f}&nbsp;MW</span>'
            f'</div>'
        )

    return (
        f'<div class="v4-topbar">'
        f'<div class="v4-brand">'
        f'<div class="v4-brand-mark">{_v4_icon("bolt", 16, "#FFF")}</div>'
        f'<div><div class="v4-brand-title">RPF Analysis</div>'
        f'<div class="v4-brand-sub">Respuesta Primaria de Frecuencia</div></div>'
        f'</div>'
        f'<div class="v4-topbar-center">'
        f'<div class="v4-event-pill">'
        f'{_v4_icon("database", 13, t["textMuted"])}'
        f'<span class="v4-event-label">Semestre</span>'
        f'<span class="v4-event-val">{sem}</span>'
        f'<div class="v4-sep"></div>'
        f'{_v4_icon("bolt", 13, t["accent"])}'
        f'<span class="v4-event-label">Evento</span>'
        f'<span class="v4-event-val">{ev}</span>'
        f'{fecha_suffix}'
        f'</div>'
        f'{disc_chips_html}'
        f'</div>'
        f'<div class="v4-topbar-right">'
        f'<span class="v4-mode-badge {mode_cls}">{mode_icon}&nbsp;{mode_label}</span>'
        f'</div>'
        f'</div>'
    )


def _build_stepper_html(active_block: str) -> str:
    """Construye HTML del stepper fijo (sin llamar a st.markdown)."""
    t          = _v4_t()
    active_idx = next((i for i, b in enumerate(_V4_BLOQUES) if b["id"] == active_block), 0)
    items_html = ""
    for i, b in enumerate(_V4_BLOQUES):
        is_active  = b["id"] == active_block
        is_past    = i < active_idx
        cls        = "v4-step" + (" active" if is_active else " past" if is_past else "")
        num_txt    = "✓" if is_past else b["num"]
        badge_html = f'<span class="v4-step-badge">local</span>' if (IS_CLOUD and b["pf"]) else ""
        icon_col   = t["success"] if is_past else (t["accent"] if is_active else t["textMuted"])
        items_html += (
            f'<span class="{cls}" title="{b["label"]}">'
            f'<span class="v4-step-num">{num_txt}</span>'
            f'{_v4_icon(b["icon"], 13, icon_col)}'
            f'<span>{b["short"]}</span>{badge_html}</span>'
        )
        if i < len(_V4_BLOQUES) - 1:
            items_html += f'<span class="v4-connector{" past" if is_past else ""}"></span>'
    return (
        f'<div class="v4-stepper">'
        f'<div class="v4-stepper-inner">{items_html}</div>'
        f'</div>'
    )

def _render_block_header(num: str, title: str, subtitle: str, grupo: str, pf_required: bool = False):
    """Breadcrumb + número de bloque + título + subtítulo (opcional banner cloud)."""
    t = _v4_t()
    crumb = f'<span>{grupo}</span><span class="v4-bc-sep"> › </span><span class="v4-bc-active">{title}</span>'
    pf_banner = (
        f'<div class="v4-banner warning">{_v4_icon("cloud", 15, t["warning"])}'
        f'<span><strong>Modo presentación:</strong> Este bloque ejecuta DIgSILENT PowerFactory. '
        f'Los botones de ejecución están deshabilitados en Streamlit Cloud — '
        f'los resultados y gráficas precargados son de solo lectura.</span></div>'
    ) if pf_required and IS_CLOUD else ""
    st.markdown(f"""
    <div class="v4-block-wrap">
      <div class="v4-breadcrumb">{crumb}</div>
      <div class="v4-block-head">
        <div class="v4-block-num">{num}</div>
        <div><div class="v4-block-title">{title}</div><div class="v4-block-sub">{subtitle}</div></div>
      </div>
    </div>{pf_banner}""", unsafe_allow_html=True)

def _build_unit_bar_html() -> str:
    """Construye HTML de la barra de unidad fija (sin llamar a st.markdown)."""
    t       = _v4_t()
    unit    = st.session_state.get("global_selected_unit") or ""
    u_clean = unit.replace("sym_", "")
    # Intentar leer valores; si falla usar caché de session_state
    _cache_key = f"_unitbar_cache_{u_clean}"
    _cached    = st.session_state.get(_cache_key, {})
    pmax_val   = _cached.get("pmax", "—")
    tech_val   = _cached.get("tech", "—")
    estat_val  = _cached.get("estat", "—")
    if u_clean:
        try:
            _pm  = _load_pmax_cargado(st.session_state.get("ev_path_global"),
                                       st.session_state.get("n_evento_global"))
            _tm  = _load_tech_map(LOC_NAMES_GEN_PATH)
            _pv, _, _ = _get_pmax_from_cargado(unit, _pm, _tm)
            if _pv:
                pmax_val = f"{float(_pv):.1f}"
                _row     = (_tm or {}).get(u_clean, (_tm or {}).get(f"sym_{u_clean}", {}))
                _tv      = str(_row.get("Tecnología", _row.get("tecnologia", "Hidroeléctrica")))
                tech_val = "Hidro" if "hidro" in _tv.lower() else _tv
        except Exception:
            pass
        try:
            _ep = _get_rp_default(u_clean, LOC_NAMES_GEN_PATH)
            estat_val = f"{_ep:.1f}"
        except Exception:
            pass
        # Actualizar caché solo cuando se obtienen valores reales
        if pmax_val != "—" or estat_val != "—":
            st.session_state[_cache_key] = {
                "pmax": pmax_val, "tech": tech_val, "estat": estat_val
            }
    # Chip de tecnología con color según tipo
    _tech_color_map = {
        "hidro": "hidro", "hidroeléctrica": "hidro", "hidroelectrica": "hidro",
        "termo": "termo", "termoeléctrica": "termo", "termoelectrica": "termo",
        "gas":   "gas",   "turbina gas": "gas",
        "eólico": "eolico", "eolico": "eolico", "eólica": "eolico", "eolica": "eolico",
        "solar": "solar", "fotovoltaica": "solar",
    }
    _tkey = _tech_color_map.get(tech_val.lower().strip(), "")
    if _tkey:
        tech_html = f'<span class="v4-chip-tech v4-chip-{_tkey}">{tech_val}</span>'
    else:
        tech_html = f'<span class="v4-stat-value">{tech_val}</span>'
    return (
        f'<div class="v4-unit-bar">'
        f'<div class="v4-stat"><span class="v4-stat-label">P_MAX</span>'
        f'<span class="v4-stat-value">{pmax_val}<span class="v4-stat-unit">MW</span></span></div>'
        f'<div class="v4-stat-sep"></div>'
        f'<div class="v4-stat"><span class="v4-stat-label">Tecnología</span>'
        f'{tech_html}</div>'
        f'<div class="v4-stat-sep"></div>'
        f'<div class="v4-stat"><span class="v4-stat-label">Estatismo</span>'
        f'<span class="v4-stat-value">{estat_val}<span class="v4-stat-unit">%</span></span></div>'
        f'</div>'
    )


def _render_unit_ctx_bar(available_units: list, loc_gen_path: str = ""):
    """Asegura que global_selected_unit sea válido. La barra visual está en el topbar fijo."""
    _cur = st.session_state.global_selected_unit
    if _cur not in available_units:
        _cur = available_units[0]
        st.session_state.global_selected_unit = _cur

# 
# CONSTANTES GLOBALÉS
# 
CARPETA_COBEE_EMF = "Resultados_COBEE" # Output folder for ExtractorResultadosCNDC.py
CARPETA_DATOS_CURVAS = "Datos Curvas" # Output folder for DatosCurvas_v3.py
CARPETA_COSTO_MARGINAL = "Costo Marginal STI" # Subcarpeta para archivos postot/td_

# 23 unidades COBEE de interés para análisis RPF (sin prefijo sym_)
COBEE_UNITS_INTERES = {
    "ZON01", "TIQ01", "BOT01", "BOT02", "BOT03",
    "CUT01", "CUT02", "CUT03", "CUT04", "CUT05",
    "SRO01", "SRO02", "SAI01", "CHU01", "CHU02",
    "HAR01", "HAR02", "CAH01", "CAH02", "HUA01", "HUA02",
    "ANG03", "CRB01",
}

# 
# CONFIGURACIÓN DE RUTAS — cargada desde archivo JSON (persistente)
# 
_CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config_rutas.json")

_DEFAULTS_CONFIG = {
    "RAIZ":               r"C:\Datos del CNDC\01_INFO CNDC_RPF",
    "RAIZ_DATOS":         r"C:\Datos del CNDC\02_DATOS CNDC_RPF",
    "PF_BASE":            r"C:\Program Files\DIgSILENT\PowerFactory 2025 SP2",
    "LOC_NAMES_GEN_PATH": r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\Designacion de loc_name\loc_names_gen.xlsx",
    "LOC_CAR_PATH":       r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\Designacion de loc_name\loc_name_cargas.xlsx",
    "LOC_XFO_PATH":       r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\Designacion de loc_name\loc_names_xfo.xlsx",
    "PF_PROYECTO":        "PMP_NOV25_OCT29_31102025(1)",
    "CASO_BASE":          "CNDC",
    "EXCLUIR_SLACK":      "sym_AGU02",
    "XFO_PF":             0.90,
}

def _cargar_config():
    if os.path.isfile(_CONFIG_PATH):
        try:
            with open(_CONFIG_PATH, encoding="utf-8") as _f:
                data = json.load(_f)
            # Combinar con defaults para tolerar claves nuevas
            return {**_DEFAULTS_CONFIG, **data}
        except Exception:
            pass
    return dict(_DEFAULTS_CONFIG)

def _guardar_config(cfg: dict):
    with open(_CONFIG_PATH, "w", encoding="utf-8") as _f:
        json.dump(cfg, _f, ensure_ascii=False, indent=2)

_cfg = _cargar_config()

#  Config vars — desde session_state (actualizado por Bloque 07) o desde archivo 
# Siempre disponibles en todos los bloques aunque el Bloque 07 no esté activo.
RAIZ               = st.session_state.get("cfg_RAIZ",               _cfg.get("RAIZ", ""))
RAIZ_DATOS         = st.session_state.get("cfg_RAIZ_DATOS",         _cfg.get("RAIZ_DATOS", ""))
PF_BASE            = st.session_state.get("cfg_PF_BASE",            _cfg.get("PF_BASE", ""))
LOC_NAMES_GEN_PATH = st.session_state.get("cfg_LOC_NAMES_GEN_PATH", _cfg.get("LOC_NAMES_GEN_PATH", ""))
LOC_CAR_PATH       = st.session_state.get("cfg_LOC_CAR_PATH",       _cfg.get("LOC_CAR_PATH", ""))
LOC_XFO_PATH       = st.session_state.get("cfg_LOC_XFO_PATH",       _cfg.get("LOC_XFO_PATH", ""))
PF_PROYECTO        = st.session_state.get("cfg_PF_PROYECTO",        _cfg.get("PF_PROYECTO", ""))
PF_PROYECTO_2      = "PMP_NOV25_OCT29_31102025(2)"   # Proyecto fijo para Tab 4 — no configurable
CASO_BASE          = st.session_state.get("cfg_CASO_BASE",          _cfg.get("CASO_BASE", ""))
EXCLUIR_SLACK      = st.session_state.get("cfg_EXCLUIR_SLACK",      _cfg.get("EXCLUIR_SLACK", ""))
XFO_PF             = float(st.session_state.get("cfg_XFO_PF",       _cfg.get("XFO_PF", 1.0)))
show_hhmmss        = st.session_state.get("global_show_hhmmss",     False)

# 
# FUNCIONES DE ANÁLISIS RPF — compartidas entre bloques 2, 3 y 4
# 

def _to_plotly_time(t_val, show_hhmmss):
    """Convierte segundos a Datetime para que Plotly los maneje correctamente."""
    if not show_hhmmss:
        return t_val
    res = pd.to_datetime(t_val, unit='s')
    # Solución TypeError (Pandas 2.x): Plotly usa sum() internamente en anotaciones de add_vline/add_hline.
    # El objeto pd.Timestamp no permite sumas con el '0' inicial de sum(). Al devolver milisegundos 
    # desde la época (float), se permite el cálculo de promedios para posicionar etiquetas.
    if isinstance(res, pd.Timestamp):
        return res.timestamp() * 1000
    return res

@st.cache_data(ttl=60)
def _listar_archivos_cache(directorio, patron, recursivo=False):
    """Cache para evitar escaneos repetitivos del sistema de archivos."""
    if not os.path.isdir(directorio):
        return []
    if recursivo:
        import glob as _glob
        return sorted(_glob.glob(os.path.join(directorio, "**", patron), recursive=True))
    return sorted([
        f for f in os.listdir(directorio) 
        if f.lower().endswith(patron.replace('*', '').lower()) and not f.startswith('~$')
    ])

def _kill_powerfactory():
    """Mata cualquier proceso PowerFactory.exe activo en segundo plano para liberar licencias."""
    try:
        # /F fuerza el cierre, /IM nombre de imagen, /T cierra procesos hijos (como la API)
        subprocess.run(["taskkill", "/F", "/IM", "PowerFactory.exe", "/T"], capture_output=True, check=False)
        return True
    except Exception:
        return False

def _leer_log_generic(path):
    """Helper ligero para leer archivos de texto."""
    try:
        return open(path, encoding="utf-8", errors="replace").read()
    except OSError:
        return ""

@st.fragment(run_every=2)
def _monitor_process_fragment(log_path, status_file):
    """Fragmento optimizado: actualiza solo el log sin recargar toda la interfaz."""
    log_txt = _leer_log_generic(log_path)
    if log_txt:
        st.text_area("📋 Log en vivo (actualización local)", value=log_txt, height=300, disabled=True)
    if os.path.exists(status_file):
        st.rerun() # Rerun global solo cuando el proceso termina

def _parse_to_seconds(series):
    """Versión optimizada y vectorizada para convertir tiempo (HH:MM:SS o float) a segundos."""
    # Convertir a string y limpiar una sola vez
    s = series.astype(str).str.strip().str.replace(',', '.')
    
    # Inicializar resultado con NaNs
    result = pd.Series(np.nan, index=series.index)
    
    # Procesar formato con dos puntos (HH:MM:SS) de forma vectorizada
    has_colon = s.str.contains(':')
    if has_colon.any():
        parts = s[has_colon].str.split(':')
        h = pd.to_numeric(parts.str[0], errors='coerce').fillna(0)
        m = pd.to_numeric(parts.str[1], errors='coerce').fillna(0)
        sec = pd.to_numeric(parts.str[2], errors='coerce').fillna(0)
        result[has_colon] = h * 3600 + m * 60 + sec
        
    # Procesar valores que ya son numéricos (como floats de segundos)
    not_colon = ~has_colon
    if not_colon.any():
        result[not_colon] = pd.to_numeric(s[not_colon], errors='coerce')
        
    return result.fillna(0.0)

_load_tech_map = st.cache_data(_load_tech_map)
_load_pmax_cargado = st.cache_data(_load_pmax_cargado)


@st.cache_data(ttl=30, show_spinner=False)
def get_event_units(ev_path=None, n_evento=None):
    """Obtiene la lista maestra de unidades disponibles para el evento seleccionado."""
    ev_path = ev_path or st.session_state.ev_path_global
    n_evento = n_evento or st.session_state.n_evento_global
    if not ev_path: return []
    
    # Patrones a ignorar (no son unidades de generación)
    BLACKLIST = [
        "Velocidades", "Ángulos", "Resumen", "Info", "tabla_resultados",
        "F.", "F.P.", "P.F", "Barras", "frecuencia", "slack", "Evento"
    ]

    def _is_valid_unit(name):
        name_up = name.upper()
        return not any(p.upper() in name_up for p in BLACKLIST)

    def _normalizar_codigo_unidad(name):
        """Algunas centrales con una sola unidad exportan el codigo sin el
        sufijo numerico (TIQ, SAI, ZON) mientras que el catalogo de interes
        (COBEE_UNITS_INTERES) usa el codigo completo (TIQ01, SAI01, ZON01).
        Sin esto, esas unidades quedan invisibles en todos los eventos."""
        return name if re.search(r"\d+$", name) else f"{name}01"

    def _clean_list(d, p):
        files = _listar_archivos_cache(d, p)
        # Extraer nombre, quitar sym_ para unificar, y filtrar por blacklist
        names = {
            _normalizar_codigo_unidad(os.path.splitext(f)[0].replace("sym_", "").replace("SYM_", ""))
            for f in files
        }
        return {n for n in names if _is_valid_unit(n)}

    def _clean_list_emf(d, p, min_power_mw=1.0):
        """Como _clean_list pero excluye unidades cuya potencia máxima < min_power_mw."""
        files = _listar_archivos_cache(d, p)
        valid = set()
        for f in files:
            name = _normalizar_codigo_unidad(os.path.splitext(f)[0].replace("sym_", "").replace("SYM_", ""))
            if not _is_valid_unit(name):
                continue
            try:
                df = pd.read_excel(os.path.join(d, f), engine="calamine")
                pot_cols = [c for c in df.columns if c.lower() not in ("tiempo_s", "frecuencia_hz", "hora")]
                if pot_cols:
                    max_val = pd.to_numeric(df[pot_cols[0]], errors="coerce").abs().max()
                    if pd.notna(max_val) and max_val >= min_power_mw:
                        valid.add(name)
                else:
                    valid.add(name)
            except Exception:
                valid.add(name)
        return valid

    # Buscar en SCADA
    u_scada = _clean_list(os.path.join(ev_path, "Graficas Registro 1SEG COBEE"), "*.xlsx")
    # Buscar en EMF — excluir unidades con potencia máxima < 1 MW
    u_emf   = _clean_list_emf(os.path.join(ev_path, CARPETA_COBEE_EMF), "*.xlsx")
    # Buscar en Simulación (E0 y E1)
    u_sim0  = _clean_list(os.path.join(ev_path, f"E{n_evento}.0", CARPETA_DATOS_CURVAS), "*.xlsx")
    u_sim1  = _clean_list(os.path.join(ev_path, f"E{n_evento}.1", CARPETA_DATOS_CURVAS), "*.xlsx")
    
    all_raw = (u_scada | u_emf | u_sim0 | u_sim1) & COBEE_UNITS_INTERES
    return sorted(list(all_raw))



def _save_rp_cfg(loc_gen_path, loc_key, droop_pct):
    p = _rp_cfg_path(loc_gen_path)
    cfg = _load_rp_cfg(loc_gen_path)
    cfg[loc_key] = round(float(droop_pct), 3)
    with open(p, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def _event_cfg_path(ev_path):
    return os.path.join(ev_path, "event_config.json")

def _global_unit_cfg_path():
    """Ruta al archivo de configuración global de unidades (independiente del evento)."""
    if IS_CLOUD:
        import tempfile
        return os.path.join(tempfile.gettempdir(), "rpf_sharepoint", "unit_global_config.json")
    # Local: junto a loc_names_gen.xlsx (definido más adelante en el script)
    try:
        return os.path.join(os.path.dirname(LOC_NAMES_GEN_PATH), "unit_global_config.json")
    except NameError:
        import tempfile
        return os.path.join(tempfile.gettempdir(), "rpf_sharepoint", "unit_global_config.json")

def _load_global_unit_cfg():
    p = _global_unit_cfg_path()
    try:
        if os.path.isfile(p):
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception: pass
    return {}

def _sync_global_cfg_from_sp():
    """Descarga unit_global_config.json desde SharePoint al caché local (solo cloud)."""
    if not (IS_CLOUD and _SP_OK):
        return
    p = _global_unit_cfg_path()
    if os.path.isfile(p):
        return
    try:
        sp_folder = _sp.sp_global_cfg_folder()
        cfg = _sp.download_json(f"{sp_folder}/unit_global_config.json")
        if cfg:
            os.makedirs(os.path.dirname(p), exist_ok=True)
            with open(p, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _save_global_unit_cfg(unit, key, value):
    p = _global_unit_cfg_path()
    cfg = _load_global_unit_cfg()
    if unit not in cfg: cfg[unit] = {}
    cfg[unit][key] = value
    try:
        os.makedirs(os.path.dirname(p), exist_ok=True)
        with open(p, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        return False
    # Subir a SharePoint siempre que esté disponible (cloud Y local)
    if _SP_OK:
        try:
            sp_folder = _sp.sp_global_cfg_folder()
            _sp.upload_json(sp_folder, "unit_global_config.json", cfg)
        except Exception:
            pass
    return True

def _load_event_cfg(ev_path):
    if ev_path:
        p = _event_cfg_path(ev_path)
        try:
            if os.path.isfile(p):
                with open(p, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
    return {}

def _sync_event_cfg_from_sp(ev_path):
    """Descarga event_config.json desde SharePoint al caché local (solo cloud)."""
    if not (IS_CLOUD and _SP_OK and ev_path):
        return
    p = _event_cfg_path(ev_path)
    if os.path.isfile(p):
        return  # ya en caché local
    try:
        sp_folder = _sp.sp_folder_from_local(ev_path)
        cfg = _sp.download_json(f"{sp_folder}/event_config.json")
        if cfg:
            os.makedirs(ev_path, exist_ok=True)
            with open(p, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _save_event_cfg(ev_path, key, value):
    p = _event_cfg_path(ev_path)
    cfg = _load_event_cfg(ev_path)
    cfg[key] = value
    try:
        os.makedirs(os.path.dirname(p), exist_ok=True)
        with open(p, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception as e:
        st.error(f"Error al guardar configuración del evento: {e}")
        return False
    # Subir a SharePoint siempre que esté disponible (cloud Y local)
    if _SP_OK:
        try:
            if IS_CLOUD:
                sp_folder = _sp.sp_folder_from_local(ev_path)
            else:
                # Modo local: mapear ruta Windows → SP usando RAIZ_RPF
                _raiz_local = st.session_state.get("raiz_rpf_local", "")
                sp_folder = (_sp.local_path_to_sp_folder(ev_path, _raiz_local)
                             if _raiz_local else _sp.sp_folder_from_local(ev_path))
            _sp.upload_json(sp_folder, "event_config.json", cfg)
        except Exception:
            pass  # no bloquear por error de upload
    return True

# Definición de categorías de configuración para sincronización inteligente
# X-axis: compartido por bloque/evento (todas las unidades ven el mismo zoom temporal)
# Frecuencia: compartida en todos los bloques para el mismo evento
# Potencia: por-unidad por-bloque en el evento (las unidades son independientes)
_SHARED_EVENT_KEYS = {
    "y_f_min", "y_f_max", "scada_t0_s", "emf_t0_s", "t_sim_falla", "delta_t_cndc", "y_auto",
    # Rangos X por bloque (compartido entre todas las unidades del mismo bloque+evento)
    "scada_xmin", "scada_xmax",
    "emf_xmin",   "emf_xmax",
    "sim0_xmin",  "sim0_xmax",
    "sim1_xmin",  "sim1_xmax",
    "simc_xmin",  "simc_xmax",
    "comp_xmin",  "comp_xmax",
    "b5_xmin",    "b5_xmax",
}
_UNIT_GLOBAL_KEYS: set = set()  # Potencia ahora es por-unidad-por-bloque en evento config

def _get_unit_cfg(ev_path, unit, key, default):
    """Recupera configuración. Tiempos y escalas de frecuencia son compartidos por evento."""
    if key in _SHARED_EVENT_KEYS:
        cfg = _load_event_cfg(ev_path)
        return cfg.get(key, default)
    
    if key in _UNIT_GLOBAL_KEYS:
        g_cfg = _load_global_unit_cfg()
        return g_cfg.get(unit, {}).get(key, default)
        
    cfg = _load_event_cfg(ev_path)
    return cfg.get("units", {}).get(unit, {}).get(key, default)

def _save_unit_cfg(ev_path, unit, key, value):
    """Guarda configuración. Escalas de frecuencia y tiempos t0 se aplican a todo el evento."""
    if key in _SHARED_EVENT_KEYS:
        return _save_event_cfg(ev_path, key, value)
        
    if key in _UNIT_GLOBAL_KEYS:
        return _save_global_unit_cfg(unit, key, value)
        
    cfg = _load_event_cfg(ev_path)
    if "units" not in cfg: cfg["units"] = {}
    if unit not in cfg["units"]: cfg["units"][unit] = {}
    cfg["units"][unit][key] = value
    return _save_event_cfg(ev_path, "units", cfg["units"])

def _sync_rpf_y_axis(key_to_update, widget_key):
    """Sincroniza frecuencia y auto-escala en todos los bloques. La potencia NO se sincroniza — es por-bloque."""
    if not st.session_state.get("global_selected_unit"): return
    val = st.session_state.get(widget_key)
    if val is None: return

    # Potencia es independiente por bloque — no se sincroniza globalmente
    if key_to_update in ("y_p_min", "y_p_max"):
        return

    st.session_state[f"b3_sync_{key_to_update}"] = val

    # Sincronizar frecuencia y auto-escala entre todos los bloques
    _freq_auto_prefixes = ["b2_sc", "b2_emf", "b3_comp", "b3_sim0", "b3_sim1", "b3_simc", "b4_val",
                           "scada", "emf", "sim0", "sim1", "simc", "comp", "b5"]
    for pfx in _freq_auto_prefixes:
        for sfx in (f"_{key_to_update}", f"_fmin" if key_to_update == "y_f_min" else "",
                    f"_fmax" if key_to_update == "y_f_max" else "",
                    f"_auto" if key_to_update == "y_auto" else ""):
            tk = f"{pfx}{sfx}" if sfx else None
            if tk and tk in st.session_state:
                st.session_state[tk] = val

    _save_unit_cfg(st.session_state.ev_path_global, st.session_state.global_selected_unit, key_to_update, val)

def _sync_session_scale_config(ev_path, unit_name):
    """Carga y sincroniza configuración de escala desde archivo. Potencia NO se sincroniza globalmente."""
    if not unit_name or not ev_path:
        return

    # Compartidos por evento (frecuencia, tiempos, auto)
    y_f_min  = _get_unit_cfg(ev_path, unit_name, "y_f_min",  49.0)
    y_f_max  = _get_unit_cfg(ev_path, unit_name, "y_f_max",  51.0)
    y_auto   = _get_unit_cfg(ev_path, unit_name, "y_auto",   True)
    t0_scada = _get_unit_cfg(ev_path, unit_name, "scada_t0_s", None)
    t0_emf   = _get_unit_cfg(ev_path, unit_name, "emf_t0_s",  None)
    t0_sim   = _get_unit_cfg(ev_path, unit_name, "t_sim_falla", 5.0)
    dt_cndc  = _get_unit_cfg(ev_path, unit_name, "delta_t_cndc", 35)

    # Sincronizar Session State (solo frecuencia y auto)
    st.session_state.b3_sync_y_f_min = float(y_f_min)
    st.session_state.b3_sync_y_f_max = float(y_f_max)
    st.session_state.b3_sync_y_auto  = bool(y_auto)

    # t0 y dt en widgets
    if t0_scada is not None: st.session_state.b2_sc_t_falla = float(t0_scada)
    if t0_emf   is not None: st.session_state.b2_emf_t_falla = float(t0_emf)
    st.session_state.b3_t_falla      = float(t0_sim)
    st.session_state.b3_dt           = int(dt_cndc)
    st.session_state.b4_delta_t_cndc = int(dt_cndc)

    # Propagar frecuencia y auto a todos los prefijos de widgets
    _all_pfx = ["b2_sc", "b2_emf", "b3_comp", "b3_sim0", "b3_sim1", "b3_simc", "b4_val",
                "scada", "emf", "sim0", "sim1", "simc", "comp", "b5"]
    for pfx in _all_pfx:
        st.session_state[f"{pfx}_y_f_min"] = st.session_state.b3_sync_y_f_min
        st.session_state[f"{pfx}_y_f_max"] = st.session_state.b3_sync_y_f_max
        st.session_state[f"{pfx}_y_auto"]  = st.session_state.b3_sync_y_auto
        st.session_state[f"{pfx}_fmin"]    = st.session_state.b3_sync_y_f_min
        st.session_state[f"{pfx}_fmax"]    = st.session_state.b3_sync_y_f_max
        st.session_state[f"{pfx}_auto"]    = st.session_state.b3_sync_y_auto
        if f"{pfx}_y1min" in st.session_state: st.session_state[f"{pfx}_y1min"] = float(y_f_min)
        if f"{pfx}_y1max" in st.session_state: st.session_state[f"{pfx}_y1max"] = float(y_f_max)


def _on_freq_range_change(widget_key):
    """Obsoleto — mantenido por compatibilidad."""
    pass


def _sync_freq_input(widget_key, which):
    """Sincroniza un number_input de frecuencia con b3_sync y propaga a todos los bloques."""
    val = st.session_state.get(widget_key)
    if val is None:
        return
    f = float(val)
    if which == "min":
        st.session_state["b3_sync_y_f_min"] = f
        for _pfx in ("scada", "emf", "sim0", "sim1", "simc", "comp", "b5"):
            _k = f"{_pfx}_fmin_inp"
            if _k != widget_key:
                st.session_state[_k] = f
    else:
        st.session_state["b3_sync_y_f_max"] = f
        for _pfx in ("scada", "emf", "sim0", "sim1", "simc", "comp", "b5"):
            _k = f"{_pfx}_fmax_inp"
            if _k != widget_key:
                st.session_state[_k] = f


def _compute_auto_p_range(traces, x_min, x_max, pad_pct=0.10, min_pad=5.0):
    """Rango Y2 ajustado a los datos de potencia visibles en [x_min, x_max].

    traces: iterable de (t_arr, p_arr).
    Retorna [p_lo, p_hi] o None si no hay datos finitos en la ventana.
    """
    all_p = []
    for t_arr, p_arr in traces:
        t_a = np.asarray(t_arr, dtype=float).ravel()
        p_a = np.asarray(p_arr, dtype=float).ravel()
        n = min(len(t_a), len(p_a))
        if n == 0:
            continue
        mask = (t_a[:n] >= x_min) & (t_a[:n] <= x_max) & np.isfinite(p_a[:n])
        if mask.any():
            all_p.append(p_a[:n][mask])
    if not all_p:
        return None
    vals = np.concatenate(all_p)
    lo, hi = float(vals.min()), float(vals.max())
    pad = max((hi - lo) * pad_pct, min_pad)
    return [max(0.0, lo - pad), hi + pad]


def _cb_p_auto_changed(block_key, ev_path, sel_unit):
    """Persiste el estado del toggle auto-escala al cambiar."""
    _save_unit_cfg(ev_path, sel_unit, f"{block_key}_p_auto",
                   bool(st.session_state.get(f"{block_key}_p_auto", True)))


def _cb_p_inp_changed(ev_path, sel_unit, cfg_key, widget_key):
    """Persiste P mín/P máx al JSON de config cada vez que el usuario cambia el valor."""
    val = st.session_state.get(widget_key)
    if val is not None:
        _save_unit_cfg(ev_path, sel_unit, cfg_key, float(val))


def _render_axis_controls(block_key, ev_path, sel_unit, x_def_min, x_def_max,
                           p_def_max=200.0, traces=None,
                           auto_col=None, popover_col=None):
    """
    Controles de escala con entradas numéricas en columna izquierda, gráfico en columna derecha.

    Crea internamente st.columns([1, 3]) y devuelve la columna derecha para el gráfico.

    Persistencia:
    - Eje X     : nivel evento/bloque — compartido para TODAS las unidades del bloque.
    - Frecuencia: nivel evento global — compartida entre TODOS los bloques.
    - Potencia  : nivel unidad/bloque en evento — independiente por unidad.

    traces: list of (t_arr, p_arr) para auto-escala de potencia.

    Retorna: (x_min, x_max, f_min, f_max, p_min, p_max, auto_p, chart_col)
    """
    _k = block_key

    # Detectar cambio de unidad: si la unidad es distinta a la del render anterior,
    # limpiar las claves de session state de P para este bloque, para que los widgets
    # carguen los valores desde el config JSON de la nueva unidad (no del anterior).
    _last_unit_key = f"{_k}_last_unit"
    if st.session_state.get(_last_unit_key) != sel_unit:
        for _sk in (f"{_k}_p_auto", f"{_k}_pmin_inp", f"{_k}_pmax_inp"):
            st.session_state.pop(_sk, None)
        st.session_state[_last_unit_key] = sel_unit

    # Valores persistidos en config
    _x_min = float(_get_unit_cfg(ev_path, sel_unit, f"{_k}_xmin", x_def_min))
    _x_max = float(_get_unit_cfg(ev_path, sel_unit, f"{_k}_xmax", x_def_max))
    _f_min = float(st.session_state.get("b3_sync_y_f_min", 49.0))
    _f_max = float(st.session_state.get("b3_sync_y_f_max", 51.0))
    _p_min = float(_get_unit_cfg(ev_path, sel_unit, f"{_k}_p_min", 0.0))
    _p_max = float(_get_unit_cfg(ev_path, sel_unit, f"{_k}_p_max", p_def_max))
    # _p_auto: leer desde config (persiste entre unidades/bloques);
    # session state toma precedencia si el widget ya fue renderizado en esta sesión.
    _p_auto_cfg = bool(_get_unit_cfg(ev_path, sel_unit, f"{_k}_p_auto", True))
    _p_auto = bool(st.session_state.get(f"{_k}_p_auto", _p_auto_cfg))

    # Valores actuales de X desde widgets (para auto-P reactivo)
    _cur_xmin = float(st.session_state.get(f"{_k}_xmin_inp", _x_min))
    _cur_xmax = float(st.session_state.get(f"{_k}_xmax_inp", _x_max))

    # Auto-P: calcular rango y pre-cargar en session state antes de renderizar inputs
    if _p_auto and traces:
        _ap = _compute_auto_p_range(traces, _cur_xmin, _cur_xmax)
        if _ap:
            st.session_state[f"{_k}_pmin_inp"] = round(_ap[0], 2)
            st.session_state[f"{_k}_pmax_inp"] = round(_ap[1], 2)
            _p_min, _p_max = _ap[0], _ap[1]

    _p_min = float(st.session_state.get(f"{_k}_pmin_inp", _p_min))
    _p_max = float(st.session_state.get(f"{_k}_pmax_inp", _p_max))

    # Columnas: [controles | gráfico]
    _ctrl_col, _chart_col = st.columns([1, 3])

    with _ctrl_col:
        st.markdown("**📐 Controles de ejes**")

        auto_w = st.toggle(
            "Auto-escala Potencia", value=_p_auto, key=f"{_k}_p_auto",
            help="Calcula el rango de potencia a partir de los datos visibles en la ventana X",
            on_change=_cb_p_auto_changed, args=(_k, ev_path, sel_unit),
        )

        # ── Eje X ────────────────────────────────────────────────────
        st.caption("**Eje X (s)** — compartido para el bloque")
        _cx1, _cx2 = st.columns(2)
        x_min_w = _cx1.number_input(
            "X Mín", value=_x_min, step=1.0, format="%.1f", key=f"{_k}_xmin_inp",
        )
        x_max_w = _cx2.number_input(
            "X Máx", value=_x_max, step=1.0, format="%.1f", key=f"{_k}_xmax_inp",
        )

        # ── Frecuencia ───────────────────────────────────────────────
        st.caption("**Frecuencia (Hz)** — compartida entre bloques")
        _cf1, _cf2 = st.columns(2)
        f_min_w = _cf1.number_input(
            "F Mín", value=_f_min, step=0.05, format="%.3f", key=f"{_k}_fmin_inp",
            on_change=_sync_freq_input, args=(f"{_k}_fmin_inp", "min"),
        )
        f_max_w = _cf2.number_input(
            "F Máx", value=_f_max, step=0.05, format="%.3f", key=f"{_k}_fmax_inp",
            on_change=_sync_freq_input, args=(f"{_k}_fmax_inp", "max"),
        )

        # ── Potencia (deshabilitada cuando auto) ─────────────────────
        _p_caption = f"**Potencia (MW)** — {sel_unit or 'unidad'}"
        st.caption(_p_caption + (" *(auto)*" if (auto_w and traces) else ""))
        _cp1, _cp2 = st.columns(2)
        p_min_w = _cp1.number_input(
            "P Mín", value=_p_min, step=0.5, format="%.2f",
            key=f"{_k}_pmin_inp", disabled=bool(auto_w and traces),
            on_change=_cb_p_inp_changed,
            args=(ev_path, sel_unit, f"{_k}_p_min", f"{_k}_pmin_inp"),
        )
        p_max_w = _cp2.number_input(
            "P Máx", value=_p_max, step=0.5, format="%.2f",
            key=f"{_k}_pmax_inp", disabled=bool(auto_w and traces),
            on_change=_cb_p_inp_changed,
            args=(ev_path, sel_unit, f"{_k}_p_max", f"{_k}_pmax_inp"),
        )

        # ── Botones ──────────────────────────────────────────────────
        st.divider()
        _b1, _b2 = st.columns(2)
        if _b1.button("🔄 Reset", key=f"{_k}_reset", use_container_width=True):
            for _rk in (f"{_k}_xmin_inp", f"{_k}_xmax_inp",
                        f"{_k}_fmin_inp", f"{_k}_fmax_inp",
                        f"{_k}_pmin_inp", f"{_k}_pmax_inp"):
                st.session_state.pop(_rk, None)
            _save_unit_cfg(ev_path, sel_unit, f"{_k}_xmin", x_def_min)
            _save_unit_cfg(ev_path, sel_unit, f"{_k}_xmax", x_def_max)
            _save_unit_cfg(ev_path, sel_unit, f"{_k}_p_min", 0.0)
            _save_unit_cfg(ev_path, sel_unit, f"{_k}_p_max", p_def_max)
            st.rerun()
        if _b2.button("💾 Guardar", key=f"{_k}_save", use_container_width=True):
            _save_unit_cfg(ev_path, sel_unit, f"{_k}_xmin", x_min_w)
            _save_unit_cfg(ev_path, sel_unit, f"{_k}_xmax", x_max_w)
            _save_unit_cfg(ev_path, sel_unit, f"{_k}_p_min", p_min_w)
            _save_unit_cfg(ev_path, sel_unit, f"{_k}_p_max", p_max_w)
            st.toast(f"✅ Escala '{_k}' guardada")

    return x_min_w, x_max_w, f_min_w, f_max_w, p_min_w, p_max_w, auto_w, _chart_col


def _widget_pmax_rp(loc_key, loc_gen_path, extra_cols=None, key_prefix=""):
    """
    Renderiza una fila con P_max (desde Excel) y Estatismo editable + botón guardar.
    Devuelve (p_max_mw, rp_decimal).
    extra_cols: lista de (label, widget_fn) para columnas adicionales antes del botón.
    """
    tmap = _load_tech_map(loc_gen_path)
    td   = tmap.get(loc_key, tmap.get(f"sym_{loc_key}", {"P_max (MW)": 100.0}))
    pmax_val = _get_pmax(td)
    rp_val   = _get_rp_default(loc_key, loc_gen_path)

    n_extra = len(extra_cols) if extra_cols else 0
    cols = st.columns([2, 2] + ([1] * n_extra) + [1])

    p_max_out = cols[0].number_input(
        "P_max [MW]", value=float(pmax_val),
        min_value=0.1, step=0.1, key=f"{key_prefix}_pmax",
        help="Cargado automáticamente desde loc_names_gen.xlsx",
    )
    rp_pct = cols[1].number_input(
        "Estatismo [%]", value=float(rp_val),
        min_value=0.1, max_value=20.0, step=0.1, key=f"{key_prefix}_rp",
    )
    if extra_cols:
        for i, (lbl, fn) in enumerate(extra_cols):
            fn(cols[2 + i])

    cols[-1].markdown("&nbsp;", unsafe_allow_html=True)
    if cols[-1].button("💾", key=f"{key_prefix}_save_rp",
                       help="Guardar Estatismo para esta unidad"):
        _save_rp_cfg(loc_gen_path, loc_key, rp_pct)
        st.toast(f"Estatismo {loc_key}: {rp_pct:.1f}% guardado", icon="✅")

    return float(p_max_out), float(rp_pct) / 100.0


def _detectar_inicio_falla(freq_array, umbral_dfdt=-0.02, ventana_suavizado=5):
    """Detección robusta del inicio de falla por df/dt sostenido sobre señal suavizada."""
    n = len(freq_array)
    if n < ventana_suavizado + 2:
        return 0
    
    # Suavizado usando convolución de NumPy (más rápido que rolling de Pandas)
    kernel = np.ones(ventana_suavizado) / ventana_suavizado
    freq_smooth = np.convolve(freq_array.astype(float), kernel, mode='same')
    
    # Corregir bordes de la convolución
    freq_smooth[:ventana_suavizado//2] = freq_smooth[ventana_suavizado//2]
    freq_smooth[-(ventana_suavizado//2):] = freq_smooth[-(ventana_suavizado//2)-1]

    dfdt = np.diff(freq_smooth)
    
    # Detección vectorizada de caída sostenida
    condicion = (dfdt[:-1] < umbral_dfdt) & (dfdt[1:] < umbral_dfdt)
    indices = np.where(condicion)[0]
    
    if len(indices) > 0:
        return int(indices[0] + 1)

    candidatos = np.where(dfdt < umbral_dfdt)[0]
    return int(candidatos[0] + 1) if len(candidatos) > 0 else 0


# 
# PIPELINES CACHEADOS — Evitan re-leer el mismo Excel en B3, B4 y B5
# @st.cache_data hashea los argumentos: mismo fichero + mismos params → hit de caché
# 

@st.cache_data(show_spinner=False)
def _cached_sim_arrays(file_path: str, t_falla: float):
    """Lee, parsea y alinea un archivo de simulación (PowerFactory xlsx).
    Devuelve (ts_aligned, fs_hz, ps_mw, df_raw).
    Cacheado por (ruta, t_falla): B3, B4 y B5 comparten el resultado si los
    argumentos son iguales → evita re-leer el disco y re-parsear."""
    df = pd.read_excel(file_path, engine="calamine").dropna()
    tc, fc, pc = _robust_col_detect(df)
    ts_raw = pd.to_numeric(df[tc], errors="coerce").values
    fs_raw = pd.to_numeric(df[fc], errors="coerce").ffill().values
    fs_hz  = fs_raw * 50.0 if np.nanmax(fs_raw) < 2.0 else fs_raw
    ps_mw  = pd.to_numeric(df[pc], errors="coerce").ffill().values
    _v     = ~np.isnan(ts_raw)
    ts_raw, fs_hz, ps_mw = ts_raw[_v], fs_hz[_v], ps_mw[_v]
    return ts_raw - t_falla, fs_hz, ps_mw, df


@st.cache_data(show_spinner=False)
def _cached_real_arrays(file_path: str, umbral_dfdt: float, ventana: int):
    """Lee, parsea y detecta la falla en datos reales (SCADA 1SEG o EMF).
    Devuelve (tr_aligned, fr_arr, pr_arr, idx_falla, t_f_auto).
    t_f_auto es el t₀ auto-detectado en segundos normalizados (desde inicio del archivo).
    Cacheado por (ruta, umbral, ventana): cualquier bloque que use el mismo fichero
    con los mismos parámetros de detección obtiene el resultado sin releer el disco."""
    df = pd.read_excel(file_path, engine="calamine").dropna()
    tr_raw  = _parse_to_seconds(df.iloc[:, 0])
    tr_norm = tr_raw - tr_raw.min()
    _fr_c   = [c for c in df.columns if any(kw in c.lower() for kw in ["frec", "hz", "freq"])]
    _fr_col = _fr_c[0] if _fr_c else df.columns[1]
    _pr_col = df.columns[2] if len(df.columns) > 2 else df.columns[1]
    fr_arr  = pd.to_numeric(df[_fr_col], errors="coerce").ffill().values
    pr_arr  = pd.to_numeric(df[_pr_col], errors="coerce").ffill().values
    idx_f   = _detectar_inicio_falla(fr_arr, umbral_dfdt, ventana)
    t_f     = float(tr_norm.iloc[idx_f])
    return (tr_norm - t_f).values, fr_arr, pr_arr, idx_f, t_f


def _add_marker(fig, t, y, label, symbol, color, yaxis='y', size=12):
    fig.add_trace(go.Scatter(
        x=[t], y=[y],
        mode='markers+text',
        marker=dict(symbol=symbol, size=size, color=color,
                    line=dict(width=1.5, color='white')),
        text=[label], textposition='top right',
        textfont=dict(size=10, color=color),
        yaxis=yaxis, showlegend=False, hoverinfo='skip',
    ))

def _get_kpi_table_data_lists(kpi, p_max, delta_t, rocof):
    """Helper para convertir el dict de KPI a listas para una Tabla de Plotly."""
    if not kpi: return [], []
    rows = [
        ("P_max [MW]", f"{p_max:.2f}"),
        ("f₀ [Hz]", f"{kpi['f0']:.4f}"),
        ("P₀ [MW]", f"{kpi['p0']:.3f}"),
        ("f_min [Hz]", f"{kpi['f_min']:.4f}"),
        ("t_min [s]", f"{kpi['t_min']:.1f}"),
        ("Δf [Hz]", f"{kpi['delta_f']:.4f}"),
        (f"f_Δt ({delta_t}s) [Hz]", f"{kpi['f_dt']:.4f}"),
        (f"P_Δt ({delta_t}s) [MW]", f"{kpi['p_dt']:.3f}"),
        ("ΔP [MW]", f"{kpi['dp']:.3f}"),
        ("ΔP% [%]", f"{kpi['dp_pct']:.2f}"),
        ("¿Aporta?", "Sí" if kpi['aporta'] else "No"),
        ("Droop Nom. [%]", f"{kpi['droop_nom']:.1f}"),
        ("Droop Calc. [%]", str(kpi['droop_calc'])),
    ]
    if rocof is not None and rocof == rocof:
        rows.append(("ROCOF [Hz/s]", f"{rocof:.4f}"))
    return [r[0] for r in rows], [r[1] for r in rows]

def _create_kpi_summary_table_fig(kpi, p_max, delta_t, rocof, unit_name, fuente):
    """Crea una figura de Plotly que contiene solo la tabla de KPIs para exportar como imagen compacta."""
    if not kpi: return None
    labels, values = _get_kpi_table_data_lists(kpi, p_max, delta_t, rocof)
    
    fig = go.Figure(data=[go.Table(
        header=dict(
            values=[['<b>Parámetro</b>'], [f'<b>{fuente}</b>']],
            fill_color='#2E4057',
            align='left',
            font=dict(color='white', size=12)
        ),
        cells=dict(
            values=[labels, values],
            fill_color='#F5F5F5',
            align='left',
            font=dict(color='#333333', size=11),
            height=22
        )
    )])
    
    fig.update_layout(
        title=dict(text=f"Resumen KPIs: {unit_name}", x=0.05, y=0.98),
        width=400,
        height=450,
        margin=dict(l=10, r=10, t=50, b=10)
    )
    return fig

def _load_sim_data_only(sim_type_suffix, sel_file, n_evento, ev_path):
    sim_dir = os.path.join(ev_path, f"E{n_evento}.{sim_type_suffix}", CARPETA_DATOS_CURVAS)
    if not sel_file or not os.path.isdir(sim_dir):
        return None, None
    df_sim = pd.read_excel(os.path.join(sim_dir, sel_file), engine="calamine").dropna()
    return df_sim, sel_file

def _mostrar_tabla_cndc(kpi, p_max, delta_t, fuente="Valor", rocof=None):
    """Renderiza la tabla CNDC siguiendo los pasos del manual."""
    if not kpi:
        st.warning("No se pudieron calcular KPIs.")
        return
    rows = [
        ("Potencia Efectiva (P_max) [MW]", f"{p_max:.2f}"),
        ("Paso 2: f₀ (Inicio evento) [Hz]", f"{kpi['f0']:.4f}"),
        ("Paso 2: P₀ (Inicio evento) [MW]", f"{kpi['p0']:.3f}"),
        ("Paso 2: f_min (Nadir) [Hz]",      f"{kpi['f_min']:.4f}"),
        ("Paso 2: t_min (Nadir) [s]",       f"{kpi['t_min']:.1f}"),
        ("Δf (f₀ − f_min) [Hz]",           f"{kpi['delta_f']:.4f}"),
        (f"Paso 2: f_Δt ({delta_t}s) [Hz]", f"{kpi['f_dt']:.4f}"),
        (f"Paso 2: P_Δt ({delta_t}s) [MW]", f"{kpi['p_dt']:.3f}"),
        ("Paso 3: Reserva Inicial (R_inic) [MW]",   f"{kpi['r_inic']:.3f}"),
        ("Paso 3: Reserva Inicial (R_inic) [%]",    f"{kpi['r_inic_pct']:.2f}"),
        ("Paso 3: Potencia entregada (ΔP) [MW]",     f"{kpi['dp']:.3f}"),
        ("Paso 3: Aporte Porcentual (ΔP%) [%]",      f"{kpi['dp_pct']:.2f}"),
        ("¿Aporta a la RPF? (ΔP% ≥ 1.5%)",           "✅ Sí" if kpi['aporta'] else "❌ No"),
        ("Estatismo (Droop) Nominal [%]",            f"{kpi['droop_nom']:.1f}"),
        ("Paso 4: Estatismo (Droop) Calculado [%]",  str(kpi['droop_calc'])),
    ]
    if rocof is not None and rocof == rocof:
        rows.append(("Paso 2: ROCOF (df/dt inicial) [Hz/s]", f"{rocof:.4f}"))
    _df_t = _df_safe(pd.DataFrame(rows, columns=["Parámetro", fuente]))

    # Tabla KPI: HTML puro — sin scrollbars, centrada en la página
    _th_style  = ("padding:7px 14px;text-align:left;font-weight:600;"
                  "background:#2E4057;color:#fff;font-size:12px;white-space:nowrap;")
    _td_p_style = ("padding:6px 14px;border-bottom:1px solid #e0e0e0;"
                   "font-size:12px;white-space:nowrap;min-width:240px;")
    _td_v_style = ("padding:6px 14px;border-bottom:1px solid #e0e0e0;"
                   "font-size:12px;text-align:right;white-space:nowrap;min-width:100px;")

    _html_rows = ""
    for _, _row in _df_t.iterrows():
        _param = str(_row["Parámetro"])
        _val   = str(_row[fuente])
        _row_bg = ""
        if "Aporta" in _param:
            _row_bg = "background-color:#d4edda;" if "✅" in _val else "background-color:#f8d7da;"
        _html_rows += (
            f'<tr style="{_row_bg}">'
            f'<td style="{_td_p_style}{_row_bg}">{_param}</td>'
            f'<td style="{_td_v_style}{_row_bg}">{_val}</td>'
            f'</tr>'
        )

    _html_table = f"""
    <div style="display:flex;justify-content:center;margin:8px 0 4px 0;">
      <table style="border-collapse:collapse;border:1px solid #d0d0d0;border-radius:6px;overflow:hidden;">
        <thead>
          <tr>
            <th style="{_th_style}">Parámetro</th>
            <th style="{_th_style}text-align:right;">{fuente}</th>
          </tr>
        </thead>
        <tbody>{_html_rows}</tbody>
      </table>
    </div>
    """
    st.markdown(_html_table, unsafe_allow_html=True)

    # Export button for KPI table
    if st.button(f"⬇️ Descargar KPIs a Excel ({fuente})", key=f"dl_kpis_{fuente.replace(' ', '_')}"):
        excel_data = _apply_excel_formatting(
            _df_t,
            sheet_name=f"KPIs_{fuente.replace(' ', '_')}",
            kpi_col=fuente,
            kpi_ok_val="✅ Sí",
            kpi_error_val="❌ No"
        )
        st.download_button(f"Descargar {fuente} KPIs", excel_data,
                           file_name=f"kpis_{fuente.replace(' ', '_')}_Ev{st.session_state.n_evento_global}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _mostrar_tabla_cndc_duo(kpi1, p_max, delta_t1, fuente1, rocof=None,
                             kpi2=None, delta_t2=None, fuente2=None):
    """Tabla CNDC con evaluación en t₀+Δt y, opcionalmente, en P_máxima — ambas columnas lado a lado."""
    if not kpi1:
        st.warning("No se pudieron calcular KPIs.")
        return

    has_col2 = kpi2 is not None and delta_t2 is not None

    def _vals_from(kpi, roc):
        v = [
            f"{p_max:.2f}",
            f"{kpi['f0']:.4f}",
            f"{kpi['p0']:.3f}",
            f"{kpi['f_min']:.4f}",
            f"{kpi['t_min']:.1f}",
            f"{kpi['delta_f']:.4f}",
            f"{kpi['f_dt']:.4f}",
            f"{kpi['p_dt']:.3f}",
            f"{kpi['r_inic']:.3f}",
            f"{kpi['r_inic_pct']:.2f}",
            f"{kpi['dp']:.3f}",
            f"{kpi['dp_pct']:.2f}",
            "✅ Sí" if kpi['aporta'] else "❌ No",
            f"{kpi['droop_nom']:.1f}",
            str(kpi['droop_calc']),
        ]
        if roc is not None and roc == roc:
            v.append(f"{roc:.4f}")
        return v

    _LABELS = [
        "Potencia Efectiva (P_max) [MW]",
        "Paso 2: f₀ (Inicio evento) [Hz]",
        "Paso 2: P₀ (Inicio evento) [MW]",
        "Paso 2: f_min (Nadir) [Hz]",
        "Paso 2: t_min (Nadir) [s]",
        "Δf (f₀ − f_min) [Hz]",
        "Paso 2: f_Δt [Hz]",
        "Paso 2: P_Δt [MW]",
        "Paso 3: Reserva Inicial (R_inic) [MW]",
        "Paso 3: Reserva Inicial (R_inic) [%]",
        "Paso 3: Potencia entregada (ΔP) [MW]",
        "Paso 3: Aporte Porcentual (ΔP%) [%]",
        "¿Aporta a la RPF? (ΔP% ≥ 1.5%)",
        "Estatismo (Droop) Nominal [%]",
        "Paso 4: Estatismo (Droop) Calculado [%]",
    ]
    if rocof is not None and rocof == rocof:
        _LABELS.append("Paso 2: ROCOF (df/dt inicial) [Hz/s]")

    vals1 = _vals_from(kpi1, rocof)
    vals2 = _vals_from(kpi2, None) if has_col2 else []

    _hdr1 = (f"{fuente1}"
             f"<br><small style='font-weight:400;opacity:0.85;'>t₀+{int(delta_t1)}s</small>")
    _hdr2 = (f"{fuente2}"
             f"<br><small style='font-weight:400;opacity:0.85;'>t = {delta_t2:.1f}s</small>") if has_col2 else ""

    _th_p = ("padding:7px 14px;text-align:left;font-weight:600;"
             "background:#2E4057;color:#fff;font-size:12px;white-space:nowrap;")
    _th_v = ("padding:7px 12px;text-align:right;font-weight:600;"
             "background:#2E4057;color:#fff;font-size:12px;white-space:nowrap;min-width:120px;")
    _td_p = "padding:6px 14px;border-bottom:1px solid #e0e0e0;font-size:12px;white-space:nowrap;"
    _td_v = ("padding:6px 12px;border-bottom:1px solid #e0e0e0;"
             "font-size:12px;text-align:right;white-space:nowrap;min-width:120px;")

    hdr = (f'<th style="{_th_p}">Parámetro</th>'
           f'<th style="{_th_v}">{_hdr1}</th>')
    if has_col2:
        hdr += f'<th style="{_th_v}">{_hdr2}</th>'

    body = ""
    for i, lbl in enumerate(_LABELS):
        v1 = vals1[i] if i < len(vals1) else "—"
        v2 = vals2[i] if has_col2 and i < len(vals2) else "—"
        is_aporta = "Aporta" in lbl
        bg1 = bg2 = ""
        if is_aporta:
            bg1 = "background:#d4edda;color:#155724;" if "✅" in v1 else "background:#f8d7da;color:#721c24;"
            if has_col2:
                bg2 = "background:#d4edda;color:#155724;" if "✅" in v2 else "background:#f8d7da;color:#721c24;"
        cells = (f'<td style="{_td_p}{bg1}">{lbl}</td>'
                 f'<td style="{_td_v}{bg1}">{v1}</td>')
        if has_col2:
            cells += f'<td style="{_td_v}{bg2}">{v2}</td>'
        body += f'<tr>{cells}</tr>'

    html = (
        '<div style="display:flex;justify-content:center;margin:8px 0 4px 0;">'
        '<table style="border-collapse:collapse;border:1px solid #d0d0d0;border-radius:6px;overflow:hidden;">'
        f'<thead><tr>{hdr}</tr></thead>'
        f'<tbody>{body}</tbody>'
        '</table></div>'
    )
    st.markdown(html, unsafe_allow_html=True)

    _df_t = pd.DataFrame({"Parámetro": _LABELS, fuente1: vals1[:len(_LABELS)]})
    if has_col2:
        _v2_pad = (vals2 + ["—"] * len(_LABELS))[:len(_LABELS)]
        _df_t[fuente2] = _v2_pad
    _key = f"dl_kpis_{fuente1.replace(' ', '_')}"
    if st.button(f"⬇️ Descargar KPIs a Excel ({fuente1})", key=_key):
        excel_data = _apply_excel_formatting(
            _df_t,
            sheet_name=f"KPIs_{fuente1.replace(' ', '_')}",
            kpi_col=fuente1,
            kpi_ok_val="✅ Sí",
            kpi_error_val="❌ No"
        )
        st.download_button(
            f"Descargar {fuente1} KPIs", excel_data,
            file_name=f"kpis_{fuente1.replace(' ', '_')}_Ev{st.session_state.n_evento_global}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# 
# INICIALIZAR SESSION STATE
# 
if "global_selected_unit" not in st.session_state:
    st.session_state.global_selected_unit = None
if "active_block" not in st.session_state:
    st.session_state.active_block = "carga_datos"
if "ui_theme" not in st.session_state:
    st.session_state.ui_theme = "light"

# Inicialización robusta de configuración de gráficas
if "graph_config" not in st.session_state:
    st.session_state.graph_config = dict(DEFAULT_GRAPH_CONFIG)
else:
    # Asegurar que todas las claves por defecto existan (evita KeyErrors en v3)
    # preservando los valores personalizados del usuario.
    for key, value in DEFAULT_GRAPH_CONFIG.items():
        if key not in st.session_state.graph_config:
            st.session_state.graph_config[key] = value
if "semestre_global" not in st.session_state:
    st.session_state.semestre_global = None
if "evento_global" not in st.session_state:
    st.session_state.evento_global = None
if "ev_path_global" not in st.session_state:
    st.session_state.ev_path_global = None
if "n_evento_global" not in st.session_state:
    st.session_state.n_evento_global = None
if "pf_running" not in st.session_state:
    st.session_state.pf_running = False
if "pf_return_code" not in st.session_state:
    st.session_state.pf_return_code = None
if "pf_waiting_close" not in st.session_state:
    st.session_state.pf_waiting_close = False
if "pf_needs_rerun" not in st.session_state:
    st.session_state.pf_needs_rerun = False
if "pf_status_file" not in st.session_state:
    st.session_state.pf_status_file = None
if "pf_log_file" not in st.session_state:
    st.session_state.pf_log_file = None
if "pf_saved_log" not in st.session_state:
    st.session_state.pf_saved_log = None
#  Tab 4: PowerFactory proyecto (2) — estados independientes
if "pf2_running" not in st.session_state:
    st.session_state.pf2_running = False
if "pf2_return_code" not in st.session_state:
    st.session_state.pf2_return_code = None
if "pf2_waiting_close" not in st.session_state:
    st.session_state.pf2_waiting_close = False
if "pf2_status_file" not in st.session_state:
    st.session_state.pf2_status_file = None
if "pf2_log_file" not in st.session_state:
    st.session_state.pf2_log_file = None
if "pf2_saved_log" not in st.session_state:
    st.session_state.pf2_saved_log = None
#  Tab 1: extracción CNDC
if "ext_running" not in st.session_state:
    st.session_state.ext_running = False
if "ext_status_file" not in st.session_state:
    st.session_state.ext_status_file = None
if "ext_return_code" not in st.session_state:
    st.session_state.ext_return_code = None
if "ext_log_file" not in st.session_state:
    st.session_state.ext_log_file = None
if "ext_saved_log" not in st.session_state:
    st.session_state.ext_saved_log = None
#  Tab 1b: CondInicialesPF 
if "ci_running" not in st.session_state:
    st.session_state.ci_running = False
if "ci_status_file" not in st.session_state:
    st.session_state.ci_status_file = None
if "ci_return_code" not in st.session_state:
    st.session_state.ci_return_code = None
if "ci_log_file" not in st.session_state:
    st.session_state.ci_log_file = None
if "ci_saved_log" not in st.session_state:
    st.session_state.ci_saved_log = None
#  Bloque 0: DatsoGENBUSLNE (extracción modelo base) 
if "mod_running" not in st.session_state:
    st.session_state.mod_running = False
if "mod_status_file" not in st.session_state:
    st.session_state.mod_status_file = None
if "mod_return_code" not in st.session_state:
    st.session_state.mod_return_code = None
if "mod_log_file" not in st.session_state:
    st.session_state.mod_log_file = None
if "mod_saved_log" not in st.session_state:
    st.session_state.mod_saved_log = None
#  Bloque 0: scripts adicionales de modelo base 
for _pfx in ("gen", "lne", "xfo", "sht", "car"):
    if f"{_pfx}_running"     not in st.session_state: st.session_state[f"{_pfx}_running"]     = False
    if f"{_pfx}_status_file" not in st.session_state: st.session_state[f"{_pfx}_status_file"] = None
    if f"{_pfx}_return_code" not in st.session_state: st.session_state[f"{_pfx}_return_code"] = None
    if f"{_pfx}_log_file"    not in st.session_state: st.session_state[f"{_pfx}_log_file"]    = None
    if f"{_pfx}_saved_log"   not in st.session_state: st.session_state[f"{_pfx}_saved_log"]   = None
#  Tab 2: OrdenadorDatosEvento (SCADA) 
if "scada_running" not in st.session_state:
    st.session_state.scada_running = False
if "scada_status_file" not in st.session_state:
    st.session_state.scada_status_file = None
if "scada_return_code" not in st.session_state:
    st.session_state.scada_return_code = None
if "scada_log_file" not in st.session_state:
    st.session_state.scada_log_file = None
if "scada_saved_log" not in st.session_state:
    st.session_state.scada_saved_log = None
#  Tab 3: ExtractorResultadosCNDC (EMF) 
if "emf_running" not in st.session_state:
    st.session_state.emf_running = False
if "emf_status_file" not in st.session_state:
    st.session_state.emf_status_file = None
if "emf_return_code" not in st.session_state:
    st.session_state.emf_return_code = None
if "emf_log_file" not in st.session_state:
    st.session_state.emf_log_file = None
if "emf_saved_log" not in st.session_state:
    st.session_state.emf_saved_log = None
#  Bloque 3: Análisis de datos 
if "b2_scada_df" not in st.session_state:
    st.session_state.b2_scada_df = None
if "b2_emf_df" not in st.session_state:
    st.session_state.b2_emf_df = None
if "b2_selected_unit" not in st.session_state:
    st.session_state.b2_selected_unit = None
# Inicialización global de variables de sincronización para el Bloque 3
# Global defaults for Block 3 Y-axis synchronization
if "b3_sync_y_f_min" not in st.session_state: st.session_state.b3_sync_y_f_min = 49.0
if "b3_sync_y_f_max" not in st.session_state: st.session_state.b3_sync_y_f_max = 51.0
if "b3_sync_y_p_min" not in st.session_state: st.session_state.b3_sync_y_p_min = 0.0
if "b3_sync_y_p_max" not in st.session_state: st.session_state.b3_sync_y_p_max = 200.0
if "b3_sync_y_p_max" not in st.session_state: st.session_state.b3_sync_y_p_max = 200.0 # Generic default
if "b3_sync_y_auto" not in st.session_state: st.session_state.b3_sync_y_auto = True
if "b3_last_unit" not in st.session_state: st.session_state.b3_last_unit = None
if "b3_last_event_path" not in st.session_state: st.session_state.b3_last_event_path = None
if "b3_selected_unit" not in st.session_state:
    st.session_state.b3_selected_unit = None

# --- Variables para persistencia de descarga de KPIs compactos ---
if "b3_kpi_zip_bytes" not in st.session_state: st.session_state.b3_kpi_zip_bytes = None
if "b3_kpi_zip_name" not in st.session_state: st.session_state.b3_kpi_zip_name = ""
if "b3_kpi_zip_count" not in st.session_state: st.session_state.b3_kpi_zip_count = 0
if "b3_kpi_excel_bytes" not in st.session_state: st.session_state.b3_kpi_excel_bytes = None
if "b3_kpi_excel_name" not in st.session_state: st.session_state.b3_kpi_excel_name = ""

# --- Variables para persistencia de descarga de gráficos ZIP ---
if "b3_plots_zip_bytes" not in st.session_state: st.session_state.b3_plots_zip_bytes = None
if "b3_plots_zip_name" not in st.session_state: st.session_state.b3_plots_zip_name = ""
if "b4_sim_zip_bytes" not in st.session_state: st.session_state.b4_sim_zip_bytes = None
if "b4_sim_zip_name" not in st.session_state: st.session_state.b4_sim_zip_name = ""


# 
# SIDEBAR — Solo apariencia y selector de evento
# Navegación: usar el stepper (clickeable) en la zona superior de la página
# 
with st.sidebar:
    st.markdown('<div style="padding:6px 4px 4px 4px;font-size:13px;font-weight:700">⚡ RPF Analysis</div>', unsafe_allow_html=True)
    st.markdown('<div class="v4-nav-group-label">Apariencia</div>', unsafe_allow_html=True)
    _dark_on = st.toggle(
        "🌙 Modo oscuro",
        value=(st.session_state.get("ui_theme", "light") == "dark"),
        key="_sidebar_dark_toggle",
        help="Alterna entre el tema claro y oscuro de la interfaz.",
    )
    _new_theme = "dark" if _dark_on else "light"
    if st.session_state.get("ui_theme", "light") != _new_theme:
        st.session_state.ui_theme = _new_theme
        st.rerun()
    st.markdown("---")

    #  NAVEGACIÓN DE BLOQUES 
    _any_running_nav = (
        st.session_state.get("pf_running") or st.session_state.get("pf2_running") or st.session_state.get("mod_running")
        or any(st.session_state.get(f"{_p}_running") for _p in ("gen", "lne", "xfo", "sht", "car"))
        or st.session_state.get("ext_running") or st.session_state.get("ci_running")
        or st.session_state.get("scada_running") or st.session_state.get("emf_running")
    )
    if _any_running_nav:
        st.warning("Proceso en ejecución — navegación bloqueada.")
    for _grp_name, _grp_ids in [
        ("Setup",    ["modelo_base", "carga_datos", "config_unidades"]),
        ("Análisis", ["analisis_datos", "analisis_simulacion", "comparativa_real_simu", "kpi_historico"]),
        ("Salida",   ["reporte_tecnico", "config_global"]),
    ]:
        st.markdown(f'<div class="v4-nav-group-label">{_grp_name}</div>', unsafe_allow_html=True)
        for _b in _V4_BLOQUES:
            if _b["id"] not in _grp_ids:
                continue
            _is_active = st.session_state.active_block == _b["id"]
            if st.button(
                f'{_b["num"]} · {_b["label"]}',
                key=f"nav_{_b['id']}",
                disabled=_any_running_nav,
                type="primary" if _is_active else "secondary",
                use_container_width=True,
            ):
                st.session_state.active_block = _b["id"]
                st.rerun()
    st.markdown("---")

bloque_trabajo = st.session_state.active_block

def _short_col_name(col):
    """Extrae el nombre corto de columnas con ruta completa de PowerFactory.
    Ejemplo: '\\user\\Prj\\SIN.ElmNet\\sym_ZON01.ElmSym' → 'sym_ZON01'
    """
    s = str(col).strip("': ")
    if '\\' in s:
        last = s.split('\\')[-1].strip("': ")
        if '.' in last:
            last = last.rsplit('.', 1)[0]
        return last or s
    return s


with st.sidebar:
    #  SELECCIÓN DE EVENTO 
    _raiz = st.session_state.get("cfg_RAIZ", _cfg.get("RAIZ", ""))
    st.markdown('<div class="v4-nav-group-label">Evento</div>', unsafe_allow_html=True)

    if IS_CLOUD:
        #  Modo nube: datos desde SharePoint 
        if not _SP_OK:
            st.error(f"❌ No se pudo conectar a SharePoint: {_SP_ERR_MSG}")
            st.session_state.semestre_global = None
        else:
            try:
                semestres = _sp.listar_semestres()
            except Exception as _e:
                st.error(f"❌ Error listando semestres en SharePoint:\n{_e}")
                semestres = []

            if semestres:
                idx_sem = 0
                if st.session_state.semestre_global in semestres:
                    idx_sem = semestres.index(st.session_state.semestre_global)
                semestre_sel = st.selectbox(
                    "Seleccione Semestre", semestres, index=idx_sem, key="sel_semestre_global"
                )
                st.session_state.semestre_global = semestre_sel
            else:
                st.warning("❌ No se encontraron semestres en SharePoint")
                st.session_state.semestre_global = None

            if st.session_state.semestre_global:
                try:
                    eventos = _sp.listar_eventos(st.session_state.semestre_global)
                except Exception as _e:
                    st.error(f"❌ Error listando eventos:\n{_e}")
                    eventos = []

                if eventos:
                    idx_ev = 0
                    if st.session_state.evento_global in eventos:
                        idx_ev = eventos.index(st.session_state.evento_global)
                    evento_sel = st.selectbox(
                        "Seleccione Evento", eventos, index=idx_ev, key="sel_evento_global"
                    )
                    st.session_state.evento_global = evento_sel

                    # Descargar evento a /tmp/ si cambia la selección
                    _ev_key = f"{st.session_state.semestre_global}/{st.session_state.evento_global}"
                    if st.session_state.get("_sp_ev_key") != _ev_key:
                        with st.spinner("⬇️ Descargando datos del evento desde SharePoint..."):
                            try:
                                _local_ev = _sp.descargar_evento(
                                    st.session_state.semestre_global,
                                    st.session_state.evento_global,
                                )
                                st.session_state._sp_ev_local = str(_local_ev)
                                st.session_state._sp_ev_key = _ev_key
                            except Exception as _e:
                                st.error(f"❌ Error descargando evento:\n{_e}")
                                st.session_state._sp_ev_local = None

                    ev_path = st.session_state.get("_sp_ev_local")
                    if ev_path:
                        m_ev = re.search(r"(\d+)$", st.session_state.evento_global.strip())
                        n_evento = m_ev.group(1) if m_ev else st.session_state.evento_global.split()[-1]
                        st.session_state.ev_path_global = ev_path
                        st.session_state.n_evento_global = n_evento
                        if st.session_state.get("last_n_evento_global") != n_evento:
                            st.session_state.b3_kpi_zip_bytes = None
                            st.session_state.b3_kpi_excel_bytes = None
                            st.session_state.b3_plots_zip_bytes = None
                            st.session_state.b4_sim_zip_bytes = None
                            for _k in ("b3_t_falla", "b3_dt"):
                                st.session_state.pop(_k, None)
                            st.session_state.last_n_evento_global = n_evento
                        # Cargar configs guardados desde SharePoint (persistencia entre sesiones)
                        _sync_event_cfg_from_sp(ev_path)
                        _sync_global_cfg_from_sp()
                        st.success(f"Evento {n_evento} listo")
                else:
                    st.warning("❌ No hay eventos en este semestre")
                    st.session_state.evento_global = None
            else:
                st.info("← Seleccione semestre primero")
                st.session_state.evento_global = None

    else:
        #  Modo local: rutas de Windows 
        if os.path.isdir(_raiz):
            semestres = sorted(
                d for d in os.listdir(_raiz)
                if os.path.isdir(os.path.join(_raiz, d))
            )
            if semestres:
                idx_sem = 0
                if st.session_state.semestre_global in semestres:
                    idx_sem = semestres.index(st.session_state.semestre_global)
                semestre_sel = st.selectbox(
                    "Seleccione Semestre", semestres, index=idx_sem, key="sel_semestre_global"
                )
                st.session_state.semestre_global = semestre_sel
            else:
                st.warning("❌ No se encontraron semestres")
                st.session_state.semestre_global = None
        else:
            st.error(f"❌ Ruta no encontrada:\n{_raiz}")
            st.session_state.semestre_global = None

        if st.session_state.semestre_global:
            base_ev = os.path.join(_raiz, st.session_state.semestre_global, "Análisis_todos_los_eventos")
            if os.path.isdir(base_ev):
                eventos = sorted(
                    (d for d in os.listdir(base_ev) if os.path.isdir(os.path.join(base_ev, d))),
                    key=lambda d: int(m.group(1)) if (m := re.search(r"(\d+)$", d)) else -1
                )
                if eventos:
                    idx_ev = 0
                    if st.session_state.evento_global in eventos:
                        idx_ev = eventos.index(st.session_state.evento_global)
                    evento_sel = st.selectbox(
                        "Seleccione Evento", eventos, index=idx_ev, key="sel_evento_global"
                    )
                    st.session_state.evento_global = evento_sel
                    ev_path = os.path.join(_raiz, st.session_state.semestre_global,
                                           "Análisis_todos_los_eventos", st.session_state.evento_global)
                    m_ev = re.search(r"(\d+)$", st.session_state.evento_global.strip())
                    n_evento = m_ev.group(1) if m_ev else st.session_state.evento_global.split()[-1]
                    st.session_state.ev_path_global = ev_path
                    st.session_state.n_evento_global = n_evento
                    st.session_state.raiz_rpf_local  = _raiz   # para _save_event_cfg
                    if st.session_state.get("last_n_evento_global") != n_evento:
                        st.session_state.b3_kpi_zip_bytes = None
                        st.session_state.b3_kpi_excel_bytes = None
                        st.session_state.b3_plots_zip_bytes = None
                        st.session_state.b4_sim_zip_bytes = None
                        # Resetear parámetros de análisis al cambiar de evento
                        for _k in ("b3_t_falla", "b3_dt"):
                            st.session_state.pop(_k, None)
                        st.session_state.last_n_evento_global = n_evento
                    st.success(f"Evento {n_evento} seleccionado")
                else:
                    st.warning("❌ No hay eventos en este semestre")
                    st.session_state.evento_global = None
            else:
                st.error("❌ Carpeta de eventos no encontrada")
                st.session_state.evento_global = None
        else:
            st.info("← Configure RAIZ en Bloque 07 → Config. Simulación")
            st.session_state.evento_global = None

    st.markdown("---")
    st.caption("⚙️ Rutas y config en Bloque 07")

# 
# TÍTULO PRINCIPAL
# 
# En modo nube, sobreescribir todas las rutas locales con las rutas temporales descargadas
if IS_CLOUD and _SP_OK:
    RAIZ       = str(_sp.TMP_RAIZ)
    RAIZ_DATOS = str(_sp.TMP_DATOS)
    # Descargar archivos estáticos de mapeo (loc_names) si aún no están en caché
    try:
        _sp.descargar_archivos_estaticos()
    except Exception:
        pass
    _loc_dir = str(_sp.TMP_LOC_FOLDER)
    LOC_NAMES_GEN_PATH = os.path.join(_loc_dir, "loc_names_gen.xlsx")
    LOC_CAR_PATH       = os.path.join(_loc_dir, "loc_name_cargas.xlsx")
    LOC_XFO_PATH       = os.path.join(_loc_dir, "loc_names_xfo.xlsx")

#  PRE-INICIALIZAR UNIDAD ANTES DEL TOPBAR 
# global_selected_unit se setea en _render_unit_ctx_bar(), que corre DESPUÉS del topbar.
# Para que el pill de unidad aparezca desde el primer render, lo inicializamos aquí.
if (bloque_trabajo in ["analisis_datos", "analisis_simulacion", "comparativa_real_simu"]
        and not st.session_state.get("global_selected_unit")):
    try:
        _pre_units = get_event_units(
            st.session_state.get("ev_path_global"),
            st.session_state.get("n_evento_global"),
        )
        if _pre_units:
            st.session_state.global_selected_unit = _pre_units[0]
    except Exception:
        pass

#  CHROME: CSS + todos los elementos fijos en un solo st.markdown 
_IN_ANALYSIS = bloque_trabajo in ["analisis_datos", "analisis_simulacion", "comparativa_real_simu", "reporte_tecnico"]

# CSS condicional inline
_extra_css = ""
if _IN_ANALYSIS:
    _extra_css = (
        "<style>"
        " .block-container { padding-top: 162px !important; }"
        " section[data-testid='stSidebar'] { top: 160px !important;"
        "   height: calc(100vh - 160px) !important; align-self: flex-start !important; }"
        " [data-testid='stSidebarCollapsedControl'] { top: 160px !important; }"
        " div[data-testid='stToolbar'] { top: 164px !important; }"
        "</style>"
    )

# Unit bar (solo en bloques de análisis, si hay unidad)
_unit_bar_html = _build_unit_bar_html() if _IN_ANALYSIS else ""

# Un solo st.markdown para CSS base + topbar + stepper + unit bar
# El wrapper height:0 colapsa el element-container en el flujo del DOM
_inject_v4_css()           # <style> único para tema
st.markdown(               # un solo element-container para todo el chrome visible
    f'<div style="height:0;overflow:visible;margin:0;padding:0;line-height:0">'
    f'{_build_topbar_html()}'
    f'{_build_stepper_html(bloque_trabajo)}'
    f'{_unit_bar_html}'
    f'</div>'
    + _extra_css,
    unsafe_allow_html=True,
)

#  JS: favicon SVG + ajuste dinámico del layout con sidebar 
# st.markdown bloquea <script> — se usa components.html (iframe mismo origen)
# window.parent accede al documento principal desde el iframe
import streamlit.components.v1 as _v1_cmp
_v1_cmp.html(
    """<script>
(function(){
  var P=window.parent;
  if(!P)return;

  //  1. Favicon SVG (rayo sobre gradiente azul/cyan) 
  // Para usar logo COBEE: reemplazar el SVG por <img> o cambiar el href por
  // la URL/data-URI del logo PNG/SVG de COBEE.
  (function(){
    var svg='<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100">'
      +'<defs><linearGradient id="g" x1="0%" y1="0%" x2="100%" y2="100%">'
      +'<stop offset="0%" stop-color="%232563EB"/>'
      +'<stop offset="100%" stop-color="%2306B6D4"/>'
      +'</linearGradient></defs>'
      +'<rect width="100" height="100" rx="18" fill="url(%23g)"/>'
      +'<path d="M60 8 L37 52 L54 52 L48 92 L78 44 L61 44 L65 8 Z" fill="white"/>'
      +'</svg>';
    var href='data:image/svg+xml,'+svg;
    var doc=P.document;
    var lk=doc.querySelector('link[rel="icon"],link[rel="shortcut icon"]');
    if(!lk){lk=doc.createElement('link');lk.rel='icon';doc.head.appendChild(lk);}
    lk.type='image/svg+xml';
    lk.href=href;
  })();

  //  2. Ajuste dinámico stMain cuando sidebar se abre/cierra 
  function adj(){
    var sb=P.document.querySelector("section[data-testid='stSidebar']");
    var mn=P.document.querySelector("section[data-testid='stMain']");
    if(!sb||!mn)return;
    var w=sb.getBoundingClientRect().width;
    if(w>60){
      mn.style.setProperty("padding-left",w+"px","important");
    }else{
      mn.style.setProperty("padding-left","0px","important");
    }
  }
  adj();
  setTimeout(adj,150);setTimeout(adj,600);setTimeout(adj,1500);
  (function at(){
    var sb=P.document.querySelector("section[data-testid='stSidebar']");
    if(!sb){setTimeout(at,300);return;}
    if(P.ResizeObserver){new P.ResizeObserver(adj).observe(sb);}
  })();
  new P.MutationObserver(function(ml){
    ml.forEach(function(m){
      if(m.attributeName==="aria-expanded"||m.type==="childList")adj();
    });
  }).observe(P.document.body,{childList:true,subtree:true,attributes:true,
    attributeFilter:["aria-expanded","data-collapsed","style"]});
})();
</script>""",
    height=0,
)

#  SELECTOR DE UNIDAD (solo Bloques 3, 4, 5) 
if _IN_ANALYSIS:
    # Validar unidad actual
    _available_units = get_event_units(st.session_state.ev_path_global, st.session_state.n_evento_global)
    if _available_units:
        _render_unit_ctx_bar(_available_units, LOC_NAMES_GEN_PATH)
        # Sincronizar config de escala si cambió evento/unidad
        if st.session_state.global_selected_unit and st.session_state.ev_path_global:
            if (st.session_state.get("b3_last_unit") != st.session_state.global_selected_unit or
                    st.session_state.get("b3_last_event_path") != st.session_state.ev_path_global):
                _sync_session_scale_config(st.session_state.ev_path_global, st.session_state.global_selected_unit)
                st.session_state.b3_last_unit = st.session_state.global_selected_unit
                st.session_state.b3_last_event_path = st.session_state.ev_path_global
        #  Selector de unidad fijo en unit-bar (marcador + JS fallback) 
        st.markdown(
            '<div class="v4-unit-select-marker"></div>'
            '<script>(function(){'
            'var W="148px";'
            'function _fix(){'
            'var m=document.querySelector(".v4-unit-select-marker");'
            'if(!m)return;'
            'var mc=m.closest(".element-container")||m.parentElement;'
            'var ns=mc?mc.nextElementSibling:null;'
            'if(!ns||ns._rpfFixed)return;'
            'var S=ns.style;'
            'S.setProperty("position","fixed","important");'
            'S.setProperty("top","119px","important");'
            'S.setProperty("left","14px","important");'
            'S.setProperty("right","auto","important");'
            'S.setProperty("width",W,"important");'
            'S.setProperty("max-width",W,"important");'
            'S.setProperty("height","34px","important");'
            'S.setProperty("z-index","9998","important");'
            'S.setProperty("margin","0","important");'
            'S.setProperty("overflow","visible","important");'
            '[".stSelectbox","[data-baseweb=select]","[data-baseweb=select]>div"]'
            '.forEach(function(sel){'
            'ns.querySelectorAll(sel).forEach(function(el){'
            'el.style.setProperty("width",W,"important");'
            'el.style.setProperty("max-width",W,"important");'
            '});});'
            'ns._rpfFixed=true;'
            '}'
            '_fix();setTimeout(_fix,80);setTimeout(_fix,400);setTimeout(_fix,1200);'
            'new MutationObserver(function(){_fix();}).observe(document.body,{childList:true,subtree:false});'
            '})();</script>',
            unsafe_allow_html=True,
        )
        _tb_cur = st.session_state.get("global_selected_unit")
        _tb_idx = _available_units.index(_tb_cur) if _tb_cur in _available_units else 0
        _tb_sel = st.selectbox(
            "Unidad activa", _available_units, index=_tb_idx,
            key="topbar_unit_sel",
            format_func=lambda u: u.replace("sym_", ""),
            label_visibility="collapsed",
        )
        if _tb_sel != st.session_state.get("global_selected_unit"):
            st.session_state.global_selected_unit = _tb_sel
            st.rerun()
    else:
        st.info("⬆️ Seleccione evento en el panel izquierdo para comenzar el análisis.")

def _df_safe(df):
    """Convierte columnas object con tipos mixtos a str para evitar ArrowTypeError."""
    out = df.copy()
    for col in out.columns:
        if out[col].dtype == object:
            out[col] = out[col].astype(str)
    return out

if bloque_trabajo == "modelo_base":
    _render_block_header("00", "Datos del Modelo",
        "Extracción completa de parámetros técnicos y topológicos desde PowerFactory. "
        "Ejecutar únicamente cuando el modelo base (.pfd) ha sido modificado.",
        "Setup", pf_required=True)

    with st.expander("📝 Descripción de Tareas", expanded=True):
        st.markdown(f"""
        Los programas en `C:\\Programas Python\\ProgramasLimpio\\Programas_1_uso_modelo` realizan:

        *   **Barras y Líneas:** Extracción de tensiones nominales, longitudes y parámetros de carga.
        *   **Generadores y Cargas:** Mapeo de potencias nominales y conectividad a terminales.
        *   **Escenarios y Variaciones:** Búsqueda recursiva para identificar cambios realizados por cada escenario de operación y variación de red.
        *   **Casos de Estudio:** Indexación de Study Cases configurados en el árbol del proyecto.

        **Objetivo:** Actualizar los archivos Excel de mapeo base para que los simuladores operen con la última versión de la red.
        """)

    _v4_section_head("Configuración de Escaneo", icon="sliders")
    c1, c2 = st.columns(2)
    with c1:
        st.text_input("Proyecto PowerFactory", value=PF_PROYECTO, disabled=True)
        st.caption("Configurado en la barra lateral")
    with c2:
        st.text_input("Directorio de Scripts", value=r"...\Programas_1_uso_modelo", disabled=True)

    st.markdown("---")

    # Derivar rutas de salida desde LOC_NAMES_GEN_PATH # type: ignore
    _datos_extraidos_dir = os.path.dirname(os.path.dirname(LOC_NAMES_GEN_PATH))
    _mod_output_path     = os.path.join(_datos_extraidos_dir, "DatosSINdigsilent.xlsx")
    _mod_pf_py           = os.path.join(PF_BASE, "Python", "3.12")
    _mod_runner_path     = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "runners", "DatsoGENBUSLNE_run.py",
    )
    _mod_status_file = os.path.join(_datos_extraidos_dir, "_mod_status.txt")
    _mod_log_file    = os.path.join(_datos_extraidos_dir, "_mod_log.txt")
    _mod_params_path = os.path.join(_datos_extraidos_dir, "_mod_params.json")

    _can_mod = os.path.isfile(_mod_runner_path)
    if not _can_mod and not IS_CLOUD:
        st.error(f"No se encontró el runner: `{_mod_runner_path}`")

    if st.button(
        "Iniciar Extracción del Modelo Base",
        type="primary",
        use_container_width=True,
        disabled=IS_CLOUD or not _can_mod or st.session_state.mod_running,
    ):
        # Limpiar archivos previos
        for _f in (_mod_status_file, _mod_log_file):
            if os.path.exists(_f):
                try:
                    os.remove(_f)
                except OSError:
                    pass

        _mod_params = {
            "PF_DIR":      PF_BASE,
            "PF_PY":       _mod_pf_py,
            "PF_PROYECTO": PF_PROYECTO,
            "output_path": _mod_output_path,
        }
        with open(_mod_params_path, "w", encoding="utf-8") as _fp:
            json.dump(_mod_params, _fp, ensure_ascii=False, indent=2)

        st.session_state.mod_running     = True
        st.session_state.mod_status_file = _mod_status_file
        st.session_state.mod_log_file    = _mod_log_file
        st.session_state.mod_return_code = None
        st.session_state.mod_saved_log   = None

        def _mod_thread_fn(runner, params_path, env_vars, status_file, log_file):
            rc = -1
            try:
                proc = subprocess.Popen(
                    [sys.executable, runner, params_path],
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    env=env_vars,
                )
                with open(log_file, "w", encoding="utf-8") as _lf:
                    for _line in proc.stdout:
                        _lf.write(_line)
                        _lf.flush()
                rc = proc.wait()
            except Exception as _exc:
                try:
                    with open(log_file, "a", encoding="utf-8") as _lf:
                        _lf.write(f"\n[ERROR] {_exc}\n")
                except OSError:
                    pass
            finally:
                with open(status_file, "w", encoding="utf-8") as _sf:
                    _sf.write(str(rc))

        _mod_env = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"}
        threading.Thread(
            target=_mod_thread_fn,
            args=(_mod_runner_path, _mod_params_path, _mod_env, _mod_status_file, _mod_log_file),
            daemon=True,
        ).start()
        st.rerun()

    def _leer_mod_log(path):
        try:
            return open(path, encoding="utf-8", errors="replace").read()
        except OSError:
            return ""

    if st.session_state.mod_running:
        _sf = st.session_state.get("mod_status_file") or _mod_status_file
        _lf = st.session_state.get("mod_log_file")    or _mod_log_file

        if _sf and os.path.exists(_sf):
            try:
                st.session_state.mod_return_code = int(open(_sf, encoding="utf-8").read().strip())
            except (OSError, ValueError):
                st.session_state.mod_return_code = -1
            st.session_state.mod_running = False
            st.rerun()
        else:
            st.info("⏳ PowerFactory extrayendo datos del modelo (Ejecución en fragmento)...")
            _monitor_process_fragment(_lf, _sf)

    elif st.session_state.mod_return_code is not None:
        _rc = st.session_state.mod_return_code
        if _rc == 0:
            st.success("Extracción completada. `DatosSINdigsilent.xlsx` actualizado.")
        else:
            st.error(f"❌ Error en la extracción (código {_rc}).")

        _saved = st.session_state.get("mod_saved_log")
        _lf    = st.session_state.get("mod_log_file") or _mod_log_file
        _log_content = _leer_mod_log(_saved) if _saved and os.path.isfile(_saved) else _leer_mod_log(_lf)
        if _log_content:
            with st.expander("📋 Log de extracción", expanded=(_rc != 0)):
                st.code(_log_content, language="")
                _col1, _col2 = st.columns(2)
                with _col1:
                    if not (_saved and os.path.isfile(_saved)):
                        if st.button("💾 Guardar log", key="mod_save_log"):
                            import datetime # type: ignore
                            _ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                            _dest = os.path.join(_datos_extraidos_dir, f"log_modelo_{_ts}.txt")
                            try:
                                open(_dest, "w", encoding="utf-8").write(_log_content)
                                st.session_state.mod_saved_log = _dest # type: ignore
                                st.success(f"Guardado en: `{_dest}`")
                            except OSError as _e:
                                st.error(str(_e))
                    else:
                        st.caption(f"💾 Guardado en: `{_saved}`")
                with _col2:
                    st.download_button(
                        "⬇️ Descargar log",
                        data=_log_content.encode("utf-8"),
                        file_name=os.path.basename(_saved) if _saved else "log_modelo.txt",
                        mime="text/plain",
                    )

    # Mostrar logs de ejecuciones anteriores si no hay una activa
    if st.session_state.mod_return_code is None and not st.session_state.mod_running:
        _prev_mod_logs = sorted(
            glob.glob(os.path.join(_datos_extraidos_dir, "log_modelo_*.txt")),
            reverse=True,
        )
        if _prev_mod_logs:
            with st.expander(f"📋 Logs anteriores ({len(_prev_mod_logs)})"):
                _sel = st.selectbox(
                    "Seleccionar log",
                    _prev_mod_logs,
                    format_func=os.path.basename,
                    key="sel_mod_log",
                )
                if _sel:
                    st.code(_leer_mod_log(_sel), language="")
                    st.download_button(
                        "⬇️ Descargar",
                        data=_leer_mod_log(_sel).encode("utf-8"),
                        file_name=os.path.basename(_sel),
                        mime="text/plain",
                        key="dl_mod_log",
                    )

    # 
    # HELPER compartido para lanzar/monitorizar scripts de modelo base
    # 
    def _bloque_script(pfx, runner_name, params_dict, log_name, any_other_running):
        """Lanza runner, muestra log en vivo y resultado final.
        pfx            : prefijo de session_state (gen, lne, xfo, sht, car)
        runner_name    : nombre del archivo .py en runners/
        params_dict    : dict con los parámetros JSON
        log_name       : prefijo para el archivo de log permanente
        any_other_running : True si otro script ya está corriendo
        """
        _runner = os.path.join( # type: ignore
            os.path.dirname(os.path.abspath(__file__)),
            "runners", runner_name,
        )
        _sf  = os.path.join(_datos_extraidos_dir, f"_{pfx}_status.txt")
        _lf  = os.path.join(_datos_extraidos_dir, f"_{pfx}_log.txt")
        _pf  = os.path.join(_datos_extraidos_dir, f"_{pfx}_params.json")

        _can = os.path.isfile(_runner)
        if not _can and not IS_CLOUD:
            st.error(f"Runner no encontrado: `{runner_name}`")

        _is_running = st.session_state.get(f"{pfx}_running", False)
        _rc_prev    = st.session_state.get(f"{pfx}_return_code") # type: ignore

        # Icono de estado en el título del botón de ejecución
        _lbl_icon = "⏳" if _is_running else ("✅" if _rc_prev == 0 else ("❌" if _rc_prev is not None else "▶️"))

        if st.button(
            f"{_lbl_icon} Ejecutar",
            key=f"btn_{pfx}",
            type="primary",
            disabled=IS_CLOUD or not _can or _is_running or any_other_running,
        ):
            for _old in (_sf, _lf):
                if os.path.exists(_old):
                    try: os.remove(_old)
                    except OSError: pass
            with open(_pf, "w", encoding="utf-8") as _fp:
                json.dump(params_dict, _fp, ensure_ascii=False, indent=2)

            st.session_state[f"{pfx}_running"]     = True
            st.session_state[f"{pfx}_status_file"] = _sf
            st.session_state[f"{pfx}_log_file"]    = _lf
            st.session_state[f"{pfx}_return_code"] = None
            st.session_state[f"{pfx}_saved_log"]   = None

            def _thread(runner, params_path, status_file, log_file):
                rc = -1
                try:
                    proc = subprocess.Popen(
                        [sys.executable, runner, params_path],
                        stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                        text=True, encoding="utf-8", errors="replace",
                        env={**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"},
                    )
                    with open(log_file, "w", encoding="utf-8") as _lf2:
                        for line in proc.stdout:
                            _lf2.write(line); _lf2.flush()
                    rc = proc.wait()
                except Exception as exc:
                    try:
                        with open(log_file, "a", encoding="utf-8") as _lf2:
                            _lf2.write(f"\n[ERROR] {exc}\n")
                    except OSError: pass
                finally:
                    with open(status_file, "w", encoding="utf-8") as _sf2:
                        _sf2.write(str(rc))

            threading.Thread(target=_thread, args=(_runner, _pf, _sf, _lf), daemon=True).start()
            st.rerun()

        def _read(path):
            try: return open(path, encoding="utf-8", errors="replace").read()
            except OSError: return ""

        if _is_running:
            _sf_cur = st.session_state.get(f"{pfx}_status_file") or _sf
            _lf_cur = st.session_state.get(f"{pfx}_log_file")    or _lf
            if _sf_cur and os.path.exists(_sf_cur):
                try: st.session_state[f"{pfx}_return_code"] = int(open(_sf_cur).read().strip())
                except (OSError, ValueError): st.session_state[f"{pfx}_return_code"] = -1
                st.session_state[f"{pfx}_running"] = False
                st.rerun()
            else:
                st.info(f"⏳ Ejecutando {log_name}...")
                _monitor_process_fragment(_lf_cur, _sf_cur)

        elif _rc_prev is not None:
            if _rc_prev == 0:
                st.success("✅ Completado correctamente.")
            else:
                st.error(f"❌ Error (código {_rc_prev}).")
            _saved  = st.session_state.get(f"{pfx}_saved_log")
            _lf_cur = st.session_state.get(f"{pfx}_log_file") or _lf
            _txt    = _read(_saved) if _saved and os.path.isfile(_saved) else _read(_lf_cur)
            if _txt:
                with st.expander("📋 Log", expanded=(_rc_prev != 0)):
                    st.code(_txt, language="")
                    _c1, _c2 = st.columns(2)
                    with _c1:
                        if not (_saved and os.path.isfile(_saved)):
                            if st.button("💾 Guardar log", key=f"save_{pfx}_log"):
                                _dest = os.path.join(_datos_extraidos_dir,
                                    f"log_{log_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
                                try:
                                    open(_dest, "w", encoding="utf-8").write(_txt)
                                    st.session_state[f"{pfx}_saved_log"] = _dest # type: ignore
                                    st.success(f"Guardado en: `{_dest}`")
                                except OSError as _e:
                                    st.error(str(_e))
                        else:
                            st.caption(f"💾 Guardado en: `{_saved}`")
                    with _c2:
                        st.download_button("⬇️ Descargar log", data=_txt.encode(),
                                           file_name=os.path.basename(_saved) if _saved else f"log_{log_name}.txt",
                                           mime="text/plain", key=f"dl_{pfx}_log")

    # 
    # SECCION: scripts adicionales del modelo base
    # 
    _loc_designacion = os.path.join(_datos_extraidos_dir, "Designacion de loc_name")
    _any_mod_running = st.session_state.mod_running or any(
        st.session_state.get(f"{_p}_running") for _p in ("gen","lne","xfo","sht","car"))

    st.markdown("---") # type: ignore
    _v4_section_head("Scripts de Mapeo y Catálogo", icon="database")
    st.caption("Ejecutar en orden después de actualizar `DatosSINdigsilent.xlsx`.")

    #  2. loc_namesGEN 
    with st.expander("2️⃣ Mapeo Generadores CNDC → PF  (`loc_namesGEN`)", expanded=False):
        st.caption("Genera `loc_names_gen.xlsx` con el mapeo de unidades CNDC a loc_names de PowerFactory.")
        _gen_datos_sin = st.text_input(
            "Datos_SIN (*.xls)",
            value=r"C:\Datos del CNDC\Datos_SIN_20251210.xls",
            key="gen_datos_sin",
        )
        # Buscar automáticamente el archivo de referencia más reciente en RAIZ
        _sim_ref_default = ""
        try:
            import glob as _glob # type: ignore
            _candidates = sorted(
                _glob.glob(os.path.join(RAIZ, "**", "datos_simulacion_*.xlsx"), recursive=True),
                key=os.path.getmtime, reverse=True,
            )
            if _candidates:
                _sim_ref_default = _candidates[0]
        except Exception:
            pass
        _gen_sim_ref = st.text_input(
            "Archivo simulación de referencia (datos_simulacion_*.xlsx)",
            value=_sim_ref_default,
            key="gen_sim_ref",
            help="Cualquier evento del semestre activo — los nombres de generadores son consistentes.",
        )
        if _gen_sim_ref and not os.path.isfile(_gen_sim_ref): # type: ignore
            st.warning(f"⚠️ Archivo no encontrado: `{_gen_sim_ref}`")
        _bloque_script(
            pfx="gen",
            runner_name="loc_namesGEN_run.py",
            params_dict={
                "DATOS_PF":       _mod_output_path,
                "DATOS_SIN_PATH": _gen_datos_sin,
                "OUTPUT_DIR":     _loc_designacion,
                "SIM_REF_PATH":   _gen_sim_ref,
            },
            log_name="loc_namesGEN",
            any_other_running=_any_mod_running,
        )

    #  3. loc_namesLineas 
    with st.expander("3️⃣ Catálogo de Líneas PF  (`loc_namesLineas`)", expanded=False):
        st.caption("Genera `loc_names_lineas.xlsx` con loc_names y nombre descriptivo de cada línea.")
        _bloque_script(
            pfx="lne",
            runner_name="loc_namesLineas_run.py",
            params_dict={
                "DATOS_PF":   _mod_output_path,
                "OUTPUT_DIR": _loc_designacion,
            },
            log_name="loc_namesLineas",
            any_other_running=_any_mod_running,
        )

    #  4. loc_names_xfo 
    with st.expander("4️⃣ Loc-names de Transformadores  (`loc_names_xfo`)", expanded=False):
        st.caption("Genera `loc_names_xfo.xlsx` con barras HV/LV de cada transformador.")
        _xfo_topologia = st.text_input(
            "Topología completa PF (topologia_completa_pf.xlsx)",
            value=os.path.join(_datos_extraidos_dir, "Topologia", "topologia_completa_pf.xlsx"),
            key="xfo_topologia",
            help="Opcional — si no existe se infieren las barras desde el nombre del transformador.",
        )
        _bloque_script(
            pfx="xfo",
            runner_name="loc_names_xfo_run.py",
            params_dict={
                "DATOS_PF":     _mod_output_path,
                "TOPOLOGIA_PF": _xfo_topologia,
                "OUTPUT_DIR":   _loc_designacion,
            },
            log_name="loc_names_xfo",
            any_other_running=_any_mod_running,
        )

    #  5. InventarioShunts 
    with st.expander("5️⃣ Inventario de Shunts y Compensadores  (`InventarioShunts_PF`)", expanded=False):
        st.caption("Conecta a PowerFactory y genera el inventario de shunts del caso base.")
        _sht_tap = st.checkbox(
            "Activar 'Tap Adjustment of Shunts' en el Load Flow",
            value=False,
            key="sht_tap",
        )
        _bloque_script(
            pfx="sht",
            runner_name="InventarioShunts_PF_run.py",
            params_dict={
                "PF_BASE":     PF_BASE,
                "PF_PROYECTO": PF_PROYECTO,
                "CASO_BASE":   CASO_BASE,
                "OUTPUT_DIR":  _loc_designacion,
                "tap_shunts":  _sht_tap,
            },
            log_name="InventarioShunts",
            any_other_running=_any_mod_running,
        )

    #  6. MapeoRetirosSTI 
    with st.expander("6️⃣ Mapeo de Cargas → Distribuidores  (`MapeoRetirosSTI_v6`)", expanded=False):
        st.caption("Genera `loc_name_cargas.xlsx` mapeando cargas PF a distribuidores/CNR usando el instructivo CNDC.")
        _car_deener = st.text_input(
            "Archivo deener_*.xlsx",
            value="",
            key="car_deener",
            help="Ruta completa al archivo de demanda de energía del evento de referencia.",
        )
        _car_postot = st.text_input(
            "Archivo postot_*.xlsx (opcional)",
            value="",
            key="car_postot",
            help="Dejar vacío si no está disponible.",
        )
        _car_hora = st.text_input(
            "Hora del evento (HH:MM)",
            value="18:45",
            key="car_hora",
        )
        _bloque_script(
            pfx="car",
            runner_name="MapeoRetirosSTI_run.py",
            params_dict={
                "DATOS_PATH":        _mod_output_path,
                "LOC_NAMES_XFO":     os.path.join(_loc_designacion, "loc_names_xfo.xlsx"),
                "DEENER_PATH":       _car_deener,
                "POSTOT_PATH":       _car_postot,
                "OUTPUT_DIR":        _loc_designacion,
                "HORA_EVENTO_LABEL": _car_hora,
            },
            log_name="MapeoRetirosSTI",
            any_other_running=_any_mod_running,
        )

elif bloque_trabajo == "config_unidades":
    _render_block_header("02", "Parámetros DSL",
        "Gestión y optimización de parámetros DSL Andritz por familia de gobernador.",
        "Setup", pf_required=False)
    from bloque_dsl_params import render_bloque_dsl
    render_bloque_dsl(st.session_state)

elif bloque_trabajo == "config_global": # type: ignore
    _render_block_header("07", "Configuración",
        "Parámetros del proyecto, unidades COBEE y visualización de gráficas.",
        "Salida", pf_required=False)

    _b7_tab = _v4_tab_bar([
        {"id": "sim",      "label": "⚙ Config. Simulación"},
        {"id": "unidades", "label": "🔧 Config. Unidades"},
        {"id": "graficas", "label": "🎨 Config. Gráficas"},
    ], "b07")

    #  TAB 1: CONFIGURACIÓN SIMULACIÓN 
    if _b7_tab == "sim":
        _v4_section_head("Rutas del Proyecto", "Directorios y parámetros usados por los módulos de análisis.", "database")
        c1, c2 = st.columns(2)
        with c1:
            RAIZ = st.text_input("Ruta base CNDC",
                value=st.session_state.get("cfg_RAIZ", _cfg.get("RAIZ", "")), key="cfg_RAIZ",
                help="Carpeta raíz donde están los semestres.")
            RAIZ_DATOS = st.text_input("Ruta origen de datos (SCADA/EMF)",
                value=st.session_state.get("cfg_RAIZ_DATOS", _cfg.get("RAIZ_DATOS", "")), key="cfg_RAIZ_DATOS",
                help="Ruta donde se encuentran los archivos fuente para procesar.")
            PF_BASE = st.text_input("PowerFactory — directorio base",
                value=st.session_state.get("cfg_PF_BASE", _cfg.get("PF_BASE", "")), key="cfg_PF_BASE")
            PF_PROYECTO = st.text_input("Proyecto PowerFactory",
                value=st.session_state.get("cfg_PF_PROYECTO", _cfg.get("PF_PROYECTO", "")), key="cfg_PF_PROYECTO")
            CASO_BASE = st.text_input("Caso base",
                value=st.session_state.get("cfg_CASO_BASE", _cfg.get("CASO_BASE", "")), key="cfg_CASO_BASE")
        with c2:
            LOC_NAMES_GEN_PATH = st.text_input("loc_names_gen.xlsx",
                value=st.session_state.get("cfg_LOC_NAMES_GEN_PATH", _cfg.get("LOC_NAMES_GEN_PATH", "")), key="cfg_LOC_NAMES_GEN_PATH")
            LOC_CAR_PATH = st.text_input("loc_name_cargas.xlsx",
                value=st.session_state.get("cfg_LOC_CAR_PATH", _cfg.get("LOC_CAR_PATH", "")), key="cfg_LOC_CAR_PATH")
            LOC_XFO_PATH = st.text_input("loc_names_xfo.xlsx",
                value=st.session_state.get("cfg_LOC_XFO_PATH", _cfg.get("LOC_XFO_PATH", "")), key="cfg_LOC_XFO_PATH")
            EXCLUIR_SLACK = st.text_input("Generadores excluidos de slack",
                value=st.session_state.get("cfg_EXCLUIR_SLACK", _cfg.get("EXCLUIR_SLACK", "")), key="cfg_EXCLUIR_SLACK")
            XFO_PF = st.number_input("Factor XFO_PF",
                value=float(st.session_state.get("cfg_XFO_PF", _cfg.get("XFOa_PF", 1.0))), key="cfg_XFO_PF")
        st.markdown("---")
        show_hhmmss = st.checkbox("Mostrar tiempo en HH:MM:SS",
            value=st.session_state.get("global_show_hhmmss", False), key="global_show_hhmmss",
            help="Muestra el eje de tiempo en formato HH:MM:SS en todas las gráficas.")
        st.markdown("---")
        if st.button("💾 Guardar configuración", type="primary", help="Guarda las rutas actuales en el archivo de config."):
            _guardar_config({
                "RAIZ":               st.session_state.get("cfg_RAIZ", ""),
                "RAIZ_DATOS":         st.session_state.get("cfg_RAIZ_DATOS", ""),
                "PF_BASE":            st.session_state.get("cfg_PF_BASE", ""),
                "LOC_NAMES_GEN_PATH": st.session_state.get("cfg_LOC_NAMES_GEN_PATH", ""),
                "LOC_CAR_PATH":       st.session_state.get("cfg_LOC_CAR_PATH", ""),
                "LOC_XFO_PATH":       st.session_state.get("cfg_LOC_XFO_PATH", ""),
                "PF_PROYECTO":        st.session_state.get("cfg_PF_PROYECTO", ""),
                "CASO_BASE":          st.session_state.get("cfg_CASO_BASE", ""),
                "EXCLUIR_SLACK":      st.session_state.get("cfg_EXCLUIR_SLACK", ""),
                "XFO_PF":             st.session_state.get("cfg_XFO_PF", 1.0),
            })
            st.success("✅ Configuración guardada.")
        # SharePoint sync — solo modo local
        if not IS_CLOUD:
            _raiz_b7 = st.session_state.get("cfg_RAIZ", "")
            if _SP_OK and _WATCHER_MOD_OK and os.path.isdir(_raiz_b7):
                st.markdown("---")
                _v4_section_head("Sincronización SharePoint", "Sync automático de archivos locales a SharePoint.", "cloud")
                _w = _get_watcher()
                if not _w.is_running:
                    if st.button("▶ Iniciar sync automático", key="btn_start_sync"):
                        ok = _w.start(_raiz_b7)
                        if ok:
                            st.toast("✅ Sync iniciado", icon="🔄")
                        else:
                            st.warning("⚠️ Instale: `pip install watchdog`")
                else:
                    _ws = _w.stats
                    _pending_lbl = f" ({_ws['pending']} pendientes)" if _ws["pending"] else ""
                    st.success(f"🔄 Sync activo{_pending_lbl}")
                    if _ws["last_file"]:
                        st.caption(f"Último: {_ws['last_file']} a las {_ws['last_ts']}")
                    st.caption(f"↑ {_ws['uploaded']} archivos  |  ⚠ {_ws['errors']} errores")
                    if st.button("⏹ Detener sync", key="btn_stop_sync"):
                        _w.stop()
                        st.rerun()
            elif not _SP_OK:
                st.caption("☁ SharePoint no disponible — sync desactivado")

    #  TAB 2: CONFIGURACIÓN UNIDADES 
    elif _b7_tab == "unidades":
        _v4_section_head("Inventario de Unidades COBEE",
            "P_max, droop y tecnología de las 23 unidades. Independiente del evento.", "sliders")
        _tmap = _load_tech_map(LOC_NAMES_GEN_PATH)
        if not _tmap:
            st.error(f"No se pudo cargar `loc_names_gen.xlsx` desde:\n`{LOC_NAMES_GEN_PATH}`")
        else:
            _pmax_map = {}
            if st.session_state.get("ev_path_global") and st.session_state.get("n_evento_global"):
                _pmax_map = _load_pmax_cargado(
                    st.session_state.ev_path_global, st.session_state.n_evento_global
                )
            _cfg_rows = []
            for _tk in sorted(_tmap.keys()):
                if _tk.replace("sym_", "").replace("SYM_", "") not in COBEE_UNITS_INTERES:
                    continue
                _rp_v = _get_rp_default(_tk, LOC_NAMES_GEN_PATH)
                _pm_v = _pmax_map.get(_tk, _tmap[_tk].get("P_max (MW)", 0.0))
                _src  = "datos_cargados" if _tk in _pmax_map else "loc_names_gen"
                _cfg_rows.append({
                    "ID PowerFactory": _tk,
                    "P_max [MW]": _pm_v,
                    "Estatismo (Rp) [%]": _rp_v,
                    "Fuente Pmax": _src,
                })
            if _cfg_rows:
                st.dataframe(_df_safe(pd.DataFrame(_cfg_rows)), use_container_width=False, hide_index=True)
                st.markdown("---")
                _v4_section_head("Importar / Exportar Configuración", icon="download")
                ci1, ci2 = st.columns(2)
                with ci1:
                    st.markdown("**Cargar parámetros:**")
                    _up_csv = st.file_uploader("Subir archivo CSV:", type=["csv"], key="config_uploader")
                    _csv_path_input = st.text_input("O ingresar ruta absoluta:",
                        value=r"C:\Users\jose.lozano\Downloads\2026-05-07T00-20_export.csv")
                    if st.button("📥 Procesar e Importar"):
                        _source_df = None
                        if _up_csv: _source_df = pd.read_csv(_up_csv)
                        elif os.path.isfile(_csv_path_input): _source_df = pd.read_csv(_csv_path_input)
                        if _source_df is not None:
                            if "ID PowerFactory" in _source_df.columns and "Estatismo (Rp) [%]" in _source_df.columns:
                                for _, row in _source_df.iterrows():
                                    _save_rp_cfg(LOC_NAMES_GEN_PATH, str(row["ID PowerFactory"]), float(row["Estatismo (Rp) [%]"]))
                                st.success("✅ Parámetros importados correctamente.")
                                st.rerun()
                            else:
                                st.error("Columnas requeridas: 'ID PowerFactory' y 'Estatismo (Rp) [%]'")
                        else:
                            st.error("No se pudo acceder al archivo.")
                with ci2:
                    st.markdown("**Guardar estado actual:**")
                    _export_csv = pd.DataFrame(_cfg_rows).to_csv(index=False).encode('utf-8')
                    st.download_button("⬇️ Descargar Config. Actual (CSV)", _export_csv,
                                       file_name="config_rpf_unidades.csv", mime="text/csv")
                st.markdown("---")
                _v4_section_head("Edición de Parámetros", icon="sliders")
                _master_ids = [r["ID PowerFactory"] for r in _cfg_rows]
                _u_to_edit = st.selectbox("Seleccione unidad para modificar:", _master_ids)
                if _u_to_edit:
                    st.markdown(f"**Editando:** `{_u_to_edit}`")
                    _widget_pmax_rp(_u_to_edit, LOC_NAMES_GEN_PATH, key_prefix="cfg_edit")

    #  TAB 3: CONFIGURACIÓN GRÁFICAS 
    elif _b7_tab == "graficas":
        _v4_section_head("Colores y Estilo",
            "Personalice paleta, grosor de línea y plantillas Plotly.", "palette")
        c1, c2, c3 = st.columns(3)
        with c1:
            st.session_state.graph_config["freq_color_real"] = st.color_picker("Frecuencia Real", st.session_state.graph_config["freq_color_real"])
            st.session_state.graph_config["freq_color_sim0"] = st.color_picker("Frecuencia Sim E.0", st.session_state.graph_config["freq_color_sim0"])
            st.session_state.graph_config["freq_color_sim1"] = st.color_picker("Frecuencia Sim E.1", st.session_state.graph_config["freq_color_sim1"])
        with c2:
            st.session_state.graph_config["pot_color_real"] = st.color_picker("Potencia Real", st.session_state.graph_config["pot_color_real"])
            st.session_state.graph_config["pot_color_sim0"] = st.color_picker("Potencia Sim E.0", st.session_state.graph_config["pot_color_sim0"])
            st.session_state.graph_config["pot_color_sim1"] = st.color_picker("Potencia Sim E.1", st.session_state.graph_config["pot_color_sim1"])
        with c3:
            st.session_state.graph_config["line_width"] = st.slider("Grosor de línea", 1.0, 5.0, float(st.session_state.graph_config["line_width"]), 0.5)
            st.session_state.graph_config["marker_size"] = st.slider("Tamaño de marcadores", 5, 25, int(st.session_state.graph_config["marker_size"]))
            st.session_state.graph_config["show_grid"] = st.checkbox("Mostrar cuadrícula", value=st.session_state.graph_config["show_grid"])
            st.session_state.graph_config["plot_height"] = st.slider("Altura del gráfico (px)", 400, 1000, int(st.session_state.graph_config["plot_height"]), 20)
            st.session_state.graph_config["template"] = st.selectbox("Plantilla de color",
                ["plotly_white", "plotly", "ggplot2", "seaborn", "simple_white", "none"],
                index=["plotly_white", "plotly", "ggplot2", "seaborn", "simple_white", "none"].index(st.session_state.graph_config["template"]),
                help="Plantilla de colores para los gráficos.")
        st.markdown("---")
        _v4_section_head("Visibilidad de Marcadores CNDC", icon="activity")
        mc1, mc2 = st.columns(2)
        with mc1:
            st.session_state.graph_config["show_initial"] = st.toggle("Mostrar Iniciales (f₀, P₀)", value=st.session_state.graph_config["show_initial"])
            st.session_state.graph_config["show_nadir"] = st.toggle("Mostrar Nadir (f_min)", value=st.session_state.graph_config["show_nadir"])
        with mc2:
            st.session_state.graph_config["show_dt_eval"] = st.toggle("Mostrar t₀+35s (f_Δt, P_Δt)", value=st.session_state.graph_config["show_dt_eval"])
            st.session_state.graph_config["show_deadband"] = st.toggle("Mostrar Banda Muerta (±25mHz)", value=st.session_state.graph_config["show_deadband"])
            st.session_state.graph_config.setdefault("show_pmax_marker", True)
            st.session_state.graph_config["show_pmax_marker"] = st.toggle(
                "Mostrar P_máxima en gráfico (×/○)",
                value=st.session_state.graph_config["show_pmax_marker"],
                help="Marca con × la potencia máxima en [t_nadir, t₀+Δt] y con ○ la frecuencia en ese instante",
            )

elif bloque_trabajo == "carga_datos":
    _render_block_header("01", "Carga de Datos",
        "Workflow lineal para extracción CNDC, generación de condiciones iniciales y carga en PowerFactory.",
        "Setup", pf_required=True)

    semestre = st.session_state.semestre_global
    evento = st.session_state.evento_global
    ev_path = st.session_state.ev_path_global
    n_evento = st.session_state.n_evento_global

    #  Tab bar persistente 
    _b1_tab = _v4_tab_bar([
        {"id": "ext",  "icon": "download", "label": "1 · Extracción CNDC"},
        {"id": "cond", "icon": "database", "label": "2 · Condiciones Iniciales"},
        {"id": "pf",   "icon": "server",   "label": "3 · PF Proyecto (1)"},
        {"id": "pf2",  "icon": "server",   "label": "4 · PF Proyecto (2)"},
    ], "b01")

    # ═════════════════════════════════════════════════════════════════════════════
    # TAB 1: EXTRACCIÓN DE DATOS CNDC
    # ═════════════════════════════════════════════════════════════════════════════
    if _b1_tab == "ext":
        _v4_section_head("Extracción de Datos CNDC",
            "📥 Este módulo extrae datos de despacho y demanda CNDC, combinando "
            "información de archivos DC, DCDR, DEENER y tabla_resultados para generar "
            "`datos_simulacion_*_2daopcion.xlsx`"
        )

        if semestre and evento:
            dc_files       = glob.glob(os.path.join(ev_path, "Despacho", "dc_*.xls*"))
            dcdr_files     = glob.glob(os.path.join(ev_path, "Despacho", "dcdr_*.xls*"))
            deener_files   = glob.glob(os.path.join(ev_path, "Demanda de Energia y Potencia", "deener_*.xlsx"))
            tabla_files    = glob.glob(os.path.join(RAIZ, semestre, "Tabla_Eventos_*.xlsx"))
            result_files   = glob.glob(os.path.join(ev_path, "Resultados_*", "tabla_resultados_*.xlsx"))

            todos_ok = True
            status_archivos = []
            for nombre, archivos in [
                ("Despacho/dc_*.xls", dc_files),
                ("Despacho/dcdr_*.xls", dcdr_files),
                ("Demanda/deener_*.xlsx", deener_files),
                ("Tabla_Eventos_*.xlsx (semestre)", tabla_files),
                ("tabla_resultados_*.xlsx (evento)", result_files),
            ]:
                existe = len(archivos) > 0
                todos_ok = todos_ok and existe
                status_archivos.append({
                    "Archivo": nombre,
                    "Estado": "OK" if existe else "Falta",
                    "Cantidad": len(archivos),
                })

            st.dataframe(
                pd.DataFrame(status_archivos), hide_index=True, use_container_width=False,
                column_config={
                    "Archivo":   st.column_config.TextColumn("Archivo",   width=220),
                    "Estado":    st.column_config.TextColumn("Estado",    width=70),
                    "Cantidad":  st.column_config.NumberColumn("Cantidad", width=70),
                },
            )

            # Vista previa de Entradas: DC y DCDR
            if dc_files or dcdr_files:
                with st.expander("📋 Vista previa: Datos de Despacho (Entradas DC/DCDR)"):
                    col_a, col_b = st.columns(2)

                    with col_a:
                        if dc_files:
                            st.caption(f"Archivo DC: `{os.path.basename(dc_files[0])}`")
                            try:
                                xl_dc = pd.ExcelFile(dc_files[0], engine="calamine")
                                tabs_dc = st.tabs(xl_dc.sheet_names)
                                for i, sheet in enumerate(xl_dc.sheet_names):
                                    with tabs_dc[i]:
                                        st.dataframe(_df_safe(xl_dc.parse(sheet).head(20)), use_container_width=False)
                            except Exception as e:
                                st.error(f"Error al leer DC: {e}")

                    with col_b:
                        if dcdr_files:
                            st.caption(f"Archivo DCDR: `{os.path.basename(dcdr_files[0])}`")
                            try:
                                xl_dcdr = pd.ExcelFile(dcdr_files[0], engine="calamine")
                                tabs_dcdr = st.tabs(xl_dcdr.sheet_names)
                                for i, sheet in enumerate(xl_dcdr.sheet_names):
                                    with tabs_dcdr[i]:
                                        st.dataframe(_df_safe(xl_dcdr.parse(sheet).head(20)), use_container_width=False)
                            except Exception as e:
                                st.error(f"Error al leer DCDR: {e}")

            # Verificar si el archivo de salida ya existe y mostrar todas sus hojas
            output_sim = glob.glob(os.path.join(ev_path, "datos_simulacion_*_2daopcion.xlsx"))
            if output_sim:
                st.success(f"📂 **Archivo de salida detectado:** `{os.path.basename(output_sim[0])}`")
                with st.expander("📊 Vista previa completa del resultado (Todas las hojas)"):
                    try:
                        xl_out = pd.ExcelFile(output_sim[0], engine="calamine")
                        nombres_hojas = xl_out.sheet_names
                        tabs_out = st.tabs(nombres_hojas)

                        for i, nombre_hoja in enumerate(nombres_hojas):
                            with tabs_out[i]:
                                df_sheet = xl_out.parse(nombre_hoja)
                                st.caption(f"Mostrando primeras 100 filas de '{nombre_hoja}'")
                                st.dataframe(_df_safe(df_sheet.head(100)), use_container_width=False)
                    except Exception as e:
                        st.error(f"No se pudo leer la salida: {e}")
            else:
                st.info("ℹ️ No se detectó archivo de salida previo para este evento.")
            st.markdown("---") # type: ignore

            if todos_ok:
                _ext_runner = os.path.join(
                    os.path.dirname(os.path.abspath(__file__)),
                    "runners", "ExtFLujos2daO_run.py",
                )
                _can_ext = os.path.isfile(_ext_runner)
                if not _can_ext and not IS_CLOUD:
                    st.error(f"No se encontró el runner: `{_ext_runner}`")

                col_ext_btn, _ = st.columns([1, 3])
                with col_ext_btn:
                    ext_btn = st.button(
                        "Ejecutar Extracción de Datos",
                        type="primary",
                        use_container_width=True,
                        disabled=IS_CLOUD or not _can_ext or st.session_state.ext_running,
                    )

                _ext_status_file = st.session_state.get("ext_status_file") or os.path.join(ev_path, "_ext_status.txt")

                if ext_btn and _can_ext:
                    _ext_status_f = os.path.join(ev_path, "_ext_status.txt")
                    if os.path.exists(_ext_status_f):
                        try:
                            os.remove(_ext_status_f)
                        except OSError:
                            pass

                    _ext_params = {
                        "semestre":           semestre,
                        "evento":             evento,
                        "RAIZ":               RAIZ,
                        "LOC_NAMES_GEN_PATH": LOC_NAMES_GEN_PATH,
                    }
                    _ext_params_path = os.path.join(ev_path, "_ext_params.json")
                    with open(_ext_params_path, "w", encoding="utf-8") as _fp:
                        json.dump(_ext_params, _fp, ensure_ascii=False, indent=2)

                    _ext_log_f = os.path.join(ev_path, "_ext_log.txt")
                    st.session_state.ext_running    = True
                    st.session_state.ext_status_file = _ext_status_f
                    st.session_state.ext_log_file    = _ext_log_f
                    st.session_state.ext_return_code = None
                    st.session_state.ext_saved_log   = None

                    def _ext_thread_fn(runner, params_path, env_vars, status_file, log_file):
                        rc = -1
                        try:
                            proc = subprocess.Popen(
                                [sys.executable, runner, params_path],
                                stdout=subprocess.PIPE,
                                stderr=subprocess.STDOUT,
                                text=True,
                                encoding="utf-8",
                                errors="replace",
                                env=env_vars,
                            )
                            with open(log_file, "w", encoding="utf-8") as _lf:
                                for _line in proc.stdout:
                                    _lf.write(_line)
                                    _lf.flush()
                            rc = proc.wait()
                        except Exception as _exc:
                            try:
                                with open(log_file, "a", encoding="utf-8") as _lf:
                                    _lf.write(f"\n[ERROR] {_exc}\n")
                            except OSError:
                                pass
                        finally:
                            with open(status_file, "w", encoding="utf-8") as _sf:
                                _sf.write(str(rc))

                    _ext_env = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"}
                    threading.Thread(
                        target=_ext_thread_fn,
                        args=(_ext_runner, _ext_params_path, _ext_env, _ext_status_f, _ext_log_f),
                        daemon=True,
                    ).start()
                    st.rerun()

                _ext_log_live = st.session_state.get("ext_log_file") or os.path.join(ev_path, "_ext_log.txt")

                def _leer_ext_log(path):
                    try:
                        return open(path, encoding="utf-8", errors="replace").read()
                    except OSError:
                        return ""

                def _guardar_ext_log(log_path, ev_path_, n_ev_):
                    contenido = _leer_ext_log(log_path)
                    if not contenido:
                        return None
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    dest = os.path.join(ev_path_, f"log_EXT_Ev{n_ev_}_{ts}.txt")
                    try:
                        with open(dest, "w", encoding="utf-8") as _f:
                            _f.write(contenido)
                        return dest
                    except OSError:
                        return None

                if st.session_state.ext_running:
                    if _ext_status_file and os.path.exists(_ext_status_file):
                        try:
                            _ext_rc = int(open(_ext_status_file, encoding="utf-8").read().strip())
                        except (OSError, ValueError):
                            _ext_rc = -1
                        st.session_state.ext_return_code = _ext_rc
                        st.session_state.ext_running = False
                        st.rerun()
                    else:
                        st.info("⏳ Extracción CNDC en curso...")
                        _monitor_process_fragment(_ext_log_live, _ext_status_file)
                elif st.session_state.ext_return_code is not None:
                    _ext_rc = st.session_state.ext_return_code
                    if _ext_rc == 0:
                        st.success("✅ Extracción completada. Cambie a la pestaña **2. Condiciones Iniciales**.")
                    else:
                        st.error(f"❌ Error en extracción (código {_ext_rc}).")
                    _ext_saved = st.session_state.get("ext_saved_log")
                    _ext_log_content = _leer_ext_log(_ext_saved) if _ext_saved and os.path.isfile(_ext_saved) else _leer_ext_log(_ext_log_live)
                    if _ext_log_content:
                        with st.expander("📋 Log de ejecución Extracción", expanded=(_ext_rc != 0)):
                            st.code(_ext_log_content, language="")
                            _el1, _el2 = st.columns(2)
                            with _el1:
                                if not (_ext_saved and os.path.isfile(_ext_saved)):
                                    if st.button("💾 Guardar log", key="ext_save_log"):
                                        _dest = _guardar_ext_log(_ext_log_live, ev_path, n_evento) # type: ignore
                                        if _dest:
                                            st.session_state.ext_saved_log = _dest
                                            st.success(f"Guardado en: `{_dest}`")
                                        else:
                                            st.error("No se pudo guardar el log.")
                                else: # type: ignore
                                    st.caption(f"💾 Guardado en: `{_ext_saved}`")
                            with _el2:
                                st.download_button(
                                    "⬇️ Descargar log",
                                    data=_ext_log_content.encode("utf-8"),
                                    file_name=os.path.basename(_ext_saved) if _ext_saved else f"log_EXT_Ev{n_evento}.txt",
                                    mime="text/plain",
                                    key="ext_dl_log",
                                )
            else:
                st.warning("⚠️ Faltan archivos de entrada. Verifique la estructura de carpetas del evento.")
        else:
            st.warning("👈 Seleccione Semestre y Evento en la barra lateral.")

    # ═════════════════════════════════════════════════════════════════════════════
    # ═════════════════════════════════════════════════════════════════════════════
    # TAB 2: CONDICIONES INICIALES
    # ═════════════════════════════════════════════════════════════════════════════
    elif _b1_tab == "cond":
        _v4_section_head("Generación de Condiciones Iniciales",
            "📝 Genera condiciones iniciales (pgini para generadores y plini para cargas) "
            "desde archivos de datos de simulación.",
            "database")

        if semestre and evento:

            st.markdown("---") # type: ignore
            _v4_section_head("Archivos de entrada requeridos", icon="check")

            sim_files  = glob.glob(os.path.join(ev_path, "datos_simulacion_*_2daopcion.xlsx"))
            loc_gen_ok = os.path.isfile(LOC_NAMES_GEN_PATH)
            loc_car_ok = os.path.isfile(LOC_CAR_PATH)
            loc_xfo_ok = os.path.isfile(LOC_XFO_PATH)

            archivos_requeridos = [
                {
                    "Archivo": "datos_simulacion_*_2daopcion.xlsx",
                    "Estado": "OK" if sim_files else "Falta",
                    "Descripción": "Genera módulo Extracción",
                },
                {
                    "Archivo": os.path.basename(LOC_NAMES_GEN_PATH),
                    "Estado": "OK" if loc_gen_ok else "Falta",
                    "Descripción": "Mapeo generadores ↔ loc_names PF",
                },
                {
                    "Archivo": os.path.basename(LOC_CAR_PATH),
                    "Estado": "OK" if loc_car_ok else "Falta",
                    "Descripción": "Mapeo cargas ↔ loc_names PF",
                },
                {
                    "Archivo": os.path.basename(LOC_XFO_PATH),
                    "Estado": "OK" if loc_xfo_ok else "Falta",
                    "Descripción": "Parámetros transformadores",
                },
            ]

            st.dataframe(
                pd.DataFrame(archivos_requeridos), hide_index=True, use_container_width=False,
                column_config={
                    "Archivo":     st.column_config.TextColumn("Archivo",     width=230),
                    "Estado":      st.column_config.TextColumn("Estado",      width=65),
                    "Descripción": st.column_config.TextColumn("Descripción", width=200),
                },
            )

            # Vista previa de Entrada: Datos de simulación
            if sim_files:
                with st.expander("📋 Vista previa: Datos de Simulación (Entrada)"):
                    try:
                        df_sim_in = pd.read_excel(sim_files[0], engine="calamine")
                        st.dataframe(_df_safe(df_sim_in.head(20)), use_container_width=False)
                    except Exception as e:
                        st.error(f"Error al leer entrada de simulación: {e}")

            # Verificar si el archivo de salida ya existe
            output_ci = glob.glob(os.path.join(ev_path, "condiciones_iniciales_*.xlsx"))
            if output_ci:
                st.success(f"📂 **Archivo de salida detectado:** `{os.path.basename(output_ci[0])}`")
                with st.expander("📊 Vista previa: Condiciones Iniciales (Salida)"):
                    try:
                        xl_ci = pd.ExcelFile(output_ci[0], engine="calamine")
                        df_pg = xl_ci.parse("pgini_GEN")
                        df_pl = xl_ci.parse("plini_CAR")

                        c1, c2 = st.columns(2)
                        with c1:
                            st.caption("Generadores (pgini)")
                            st.dataframe(df_pg.head(15), use_container_width=False)
                        with c2:
                            st.caption("Cargas (plini)")
                            st.dataframe(df_pl.head(15), use_container_width=False)
                    except Exception as e:
                        st.info("El archivo existe pero no se pudieron leer las pestañas pgini/plini.")
                st.markdown("---")
            else:
                st.info("No se detectaron condiciones iniciales generadas para este evento.")
            st.markdown("---")

            todos_ok = len(sim_files) > 0 and loc_gen_ok and loc_car_ok and loc_xfo_ok

            if todos_ok:
                _v4_section_head("Opciones de generación", icon="sliders")

                col_o1, col_o2 = st.columns(2)
                with col_o1:
                    generar_balance = st.checkbox(
                        "Incluir hoja Balance_Plini",
                        value=True,
                        help="Diagnóstico de residuos de redondeo por distribuidor"
                    )
                with col_o2:
                    precision_decimales = st.number_input(
                        "Precisión (decimales)",
                        value=4,
                        min_value=2,
                        max_value=6,
                        help="Precisión para plini (potencia de cargas)"
                    )

                _ci_runner = os.path.join(
                    os.path.dirname(os.path.abspath(__file__)),
                    "CondInicialesPF_run.py",
                )
                _can_ci = os.path.isfile(_ci_runner)
                if not _can_ci and not IS_CLOUD:
                    st.error(f"No se encontró el runner: `{_ci_runner}`")

                col_ci_btn, _ = st.columns([1, 3])
                with col_ci_btn:
                    ci_btn = st.button(
                        "Generar Condiciones Iniciales",
                        type="primary",
                        use_container_width=True,
                        disabled=IS_CLOUD or not _can_ci or st.session_state.ci_running,
                    )

                _ci_status_file = st.session_state.get("ci_status_file") or os.path.join(ev_path, "_ci_status.txt")

                if ci_btn and _can_ci:
                    _ci_status_f = os.path.join(ev_path, "_ci_status.txt")
                    _ci_log_f = os.path.join(ev_path, "_ci_log.txt")
                    if os.path.exists(_ci_status_f):
                        try:
                            os.remove(_ci_status_f)
                        except OSError:
                            pass

                    _ci_params = {
                        "semestre":     semestre,
                        "evento":       evento,
                        "RAIZ":         RAIZ,
                        "LOC_GEN_PATH": LOC_NAMES_GEN_PATH,
                        "LOC_CAR_PATH": LOC_CAR_PATH,
                        "LOC_XFO_PATH": LOC_XFO_PATH,
                    }
                    _ci_params_path = os.path.join(ev_path, "_ci_params.json")
                    with open(_ci_params_path, "w", encoding="utf-8") as _fp:
                        json.dump(_ci_params, _fp, ensure_ascii=False, indent=2)

                    st.session_state.ci_running    = True
                    st.session_state.ci_status_file = _ci_status_f
                    st.session_state.ci_log_file    = _ci_log_f
                    st.session_state.ci_return_code = None
                    st.session_state.ci_saved_log   = None

                    # CI log en vivo: no escribir a disco (evita crear _ci_log.txt)
                    if "ci_log_buffer" not in st.session_state:
                        st.session_state.ci_log_buffer = ""
                    st.session_state.ci_log_buffer = ""

                    def _ci_thread_fn(runner, params_path, env_vars, status_file, log_file):
                        rc = -1
                        try:
                            proc = subprocess.Popen(
                                [sys.executable, runner, params_path],
                                stdout=subprocess.PIPE,
                                stderr=subprocess.STDOUT,
                                text=True,
                                encoding="utf-8",
                                errors="replace",
                                env=env_vars,
                            )
                            with open(log_file, "w", encoding="utf-8") as _lf:
                                for _line in proc.stdout:
                                    _lf.write(_line)
                                    _lf.flush()
                            rc = proc.wait()
                        except Exception as exc:
                            try:
                                with open(log_file, "a", encoding="utf-8") as _lf:
                                    _lf.write(f"\n[ERROR] {exc}\n")
                            except OSError:
                                pass
                        finally:
                            with open(status_file, "w", encoding="utf-8") as _sf:
                                _sf.write(str(rc))

                    _ci_env = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"}
                    threading.Thread(
                        target=_ci_thread_fn,
                        args=(_ci_runner, _ci_params_path, _ci_env, _ci_status_f, _ci_log_f),
                        daemon=True,
                    ).start()
                    st.rerun()

                _ci_log_live = st.session_state.get("ci_log_file") or os.path.join(ev_path, "_ci_log.txt")

                def _leer_ci_log(path):
                    try:
                        return open(path, encoding="utf-8", errors="replace").read()
                    except OSError:
                        return ""

                def _guardar_ci_log(log_path, ev_path_, n_ev_):
                    contenido = _leer_ci_log(log_path)
                    if not contenido:
                        return None
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    dest = os.path.join(ev_path_, f"log_CI_Ev{n_ev_}_{ts}.txt")
                    try:
                        with open(dest, "w", encoding="utf-8") as _f:
                            _f.write(contenido)
                        return dest
                    except OSError:
                        return None

                if st.session_state.ci_running:
                    if _ci_status_file and os.path.exists(_ci_status_file):
                        try:
                            _ci_rc = int(open(_ci_status_file, encoding="utf-8").read().strip())
                        except (OSError, ValueError):
                            _ci_rc = -1
                        st.session_state.ci_return_code = _ci_rc
                        st.session_state.ci_running = False
                        st.rerun()
                    else:
                        st.info("⏳ Generación de Condiciones Iniciales en curso...")
                        _monitor_process_fragment(_ci_log_live, _ci_status_file)
                elif st.session_state.ci_return_code is not None:
                    _ci_rc = st.session_state.ci_return_code
                    if _ci_rc == 0:
                        st.success("✅ Condiciones iniciales generadas correctamente.")
                    else:
                        st.error(f"❌ Error en la generación (código {_ci_rc}).")
                    _ci_saved = st.session_state.get("ci_saved_log")
                    _ci_log_content = _leer_ci_log(_ci_saved) if _ci_saved and os.path.isfile(_ci_saved) else _leer_ci_log(_ci_log_live)
                    if _ci_log_content:
                        with st.expander("📋 Log de ejecución Condiciones Iniciales", expanded=(_ci_rc != 0)):
                            st.code(_ci_log_content, language="")
                            _cl1, _cl2 = st.columns(2)
                            with _cl1:
                                if not (_ci_saved and os.path.isfile(_ci_saved)):
                                    if st.button("💾 Guardar log", key="ci_save_log"):
                                        _dest = _guardar_ci_log(_ci_log_live, ev_path, n_evento) # type: ignore
                                        if _dest:
                                            st.session_state.ci_saved_log = _dest
                                            st.success(f"Guardado en: `{_dest}`")
                                        else:
                                            st.error("No se pudo guardar el log.")
                                else: # type: ignore
                                    st.caption(f"💾 Guardado en: `{_ci_saved}`")
                            with _cl2:
                                st.download_button(
                                    "⬇️ Descargar log",
                                    data=_ci_log_content.encode("utf-8"),
                                    file_name=os.path.basename(_ci_saved) if _ci_saved else f"log_CI_Ev{n_evento}.txt",
                                    mime="text/plain",
                                    key="ci_dl_log",
                                )
            else:
                missing = []
                if not sim_files:
                    missing.append("• `datos_simulacion_*_2daopcion.xlsx` → ejecute **módulo Extracción**")
                if not loc_gen_ok:
                    missing.append(f"• `{os.path.basename(LOC_NAMES_GEN_PATH)}`")
                if not loc_car_ok:
                    missing.append(f"• `{os.path.basename(LOC_CAR_PATH)}`")
                if not loc_xfo_ok:
                    missing.append(f"• `{os.path.basename(LOC_XFO_PATH)}`")

                st.warning("⚠️ Faltan archivos:\n" + "\n".join(missing))
        else:
            st.warning("👈 Seleccione Semestre y Evento en la barra lateral.")

    # ═════════════════════════════════════════════════════════════════════════════
    # ═════════════════════════════════════════════════════════════════════════════
    # TAB 3: POWERFACTORY
    # ═════════════════════════════════════════════════════════════════════════════
    elif _b1_tab == "pf":
        _v4_section_head("Carga en PowerFactory — Proyecto (1)",
            "Carga condiciones iniciales en el modelo PF (proyecto configurable) y ejecuta el flujo de trabajo RMS.",
            "server")

        if semestre and evento:
            st.text_input("Proyecto PowerFactory (configurable en Bloque 8)", value=PF_PROYECTO, disabled=True, key="pf1_proyecto_display")

            #  SECCIÓN — ARCHIVOS DE ENTRADA
            _v4_section_head("Archivos de Entrada", icon="database")

            ci_files    = glob.glob(os.path.join(ev_path, "condiciones_iniciales_*.xlsx"))
            dsim_files  = glob.glob(os.path.join(ev_path, "datos_simulacion_*_2daopcion.xlsx"))
            tabla_files = glob.glob(os.path.join(RAIZ, semestre, "Tabla_Eventos_*.xlsx"))
            xfo_ok      = os.path.isfile(LOC_XFO_PATH)

            # Check for Costo Marginal STI files
            costo_marginal_dir = os.path.join(ev_path, CARPETA_COSTO_MARGINAL)
            costo_marginal_files = []
            if os.path.isdir(costo_marginal_dir):
                costo_marginal_files = (glob.glob(os.path.join(costo_marginal_dir, "postot*.xlsx")) or
                                        glob.glob(os.path.join(costo_marginal_dir, "td_*.xlsx")))
            costo_marginal_found = bool(costo_marginal_files)
            def _estado(found):
                return "✅ OK" if found else "❌ Falta"
            
            tabla_archivos = pd.DataFrame([
                {
                    "Estado": _estado(ci_files),
                    "Archivo": os.path.basename(ci_files[0]) if ci_files else "condiciones_iniciales_*.xlsx",
                    "Descripción": "Condiciones iniciales (generadores + cargas)",
                    "Carpeta": os.path.dirname(ci_files[0]) if ci_files else ev_path, # type: ignore
                },
                {
                    "Estado": _estado(dsim_files),
                    "Archivo": os.path.basename(dsim_files[0]) if dsim_files else "datos_simulacion_*_2daopcion.xlsx",
                    "Descripción": "Pdem del evento (celda B8)",
                    "Carpeta": ev_path, # type: ignore
                },
                {
                    "Estado": _estado(tabla_files), # type: ignore
                    "Archivo": os.path.basename(tabla_files[0]) if tabla_files else "Tabla_Eventos_*.xlsx",
                    "Descripción": "p_desc del disparo (potencia desconectada)",
                    "Carpeta": os.path.join(RAIZ, semestre),
                },
                {
                    "Estado": _estado(xfo_ok),
                    "Archivo": os.path.basename(LOC_XFO_PATH),
                    "Descripción": "Capacidad de transformadores (restricción de cargas)", # type: ignore
                    "Carpeta": os.path.dirname(LOC_XFO_PATH),
                },
                {
                    "Estado": _estado(costo_marginal_found),
                    "Archivo": os.path.basename(costo_marginal_files[0]) if costo_marginal_files else "postot*.xlsx o td_*.xlsx",
                    "Descripción": "Costo Marginal STI (retiros CNDC nodal)",
                    "Carpeta": os.path.dirname(LOC_XFO_PATH),
                },
            ])
            st.dataframe(tabla_archivos, use_container_width=False, hide_index=True)

            #  SECCIÓN 3 — VISTA PREVIA DE CONDICIONES INICIALES 
            if ci_files:
                st.header("3 · Vista Previa de Condiciones Iniciales")
                ci_path = ci_files[0]

                try: # type: ignore
                    xl_ci_prev = pd.ExcelFile(ci_path, engine="calamine")
                    df_res_ci  = xl_ci_prev.parse("Resumen")
                    df_pgini   = xl_ci_prev.parse("pgini_GEN")
                    df_plini   = xl_ci_prev.parse("plini_CAR")

                    info_ci = dict(
                        zip(
                            df_res_ci.iloc[:, 0].astype(str).str.strip(),
                            df_res_ci.iloc[:, 1].astype(str).str.strip(),
                        )
                    )

                    pgen_ci = df_pgini["pgini_MW"].sum() if "pgini_MW" in df_pgini.columns else 0.0
                    pdem_ci = df_plini["plini_MW"].sum() if "plini_MW" in df_plini.columns else 0.0

                    k1, k2, k3, k4 = st.columns(4)
                    k1.metric("Generadores", len(df_pgini))
                    k2.metric("Cargas", len(df_plini))
                    k3.metric("Pgen Excel (MW)", f"{pgen_ci:.1f}")
                    k4.metric("Pdem Excel (MW)", f"{pdem_ci:.1f}") # type: ignore

                    ia, ib = st.columns(2)
                    with ia:
                        st.info(f"**Fecha y hora:** {info_ci.get('Fecha y hora', '—')}")
                        st.info(f"**Disparo:** {info_ci.get('Disparo', '—')}")
                    with ib:
                        st.info(f"**Hora generación:** {info_ci.get('Hora evento (gen)', '—')}")
                        st.info(f"**Hora cargas:** {info_ci.get('Hora Po (cargas)', '—')}")

                    with st.expander("📋 Generadores (pgini_GEN) — primeras 30 filas"):
                        st.dataframe(df_pgini.head(30), use_container_width=False, hide_index=True)

                    with st.expander("📋 Cargas (plini_CAR) — primeras 30 filas"):
                        st.dataframe(df_plini.head(30), use_container_width=False, hide_index=True)

                except Exception as _e:
                    st.error(f"Error al leer condiciones_iniciales: {_e}")
                    df_pgini = None
                    info_ci  = {}
            else: # type: ignore
                st.warning("No se encontró `condiciones_iniciales_*.xlsx`. Primero genere las condiciones iniciales.")
                df_pgini = None
                info_ci  = {}
                st.info("💡 Use el **módulo 2** para generar las condiciones iniciales.")

            #  SECCIÓN 4 — OPCIONES DE EJECUCIÓN 
            if ci_files:
                st.header("4 · Opciones de Ejecución")

                col_opt_a, col_opt_b = st.columns(2)

                with col_opt_a:
                    _v4_section_head("Potencia del disparo", icon="bolt")

                    p_desc_ui = 0.0 # type: ignore
                    if tabla_files:
                        try:
                            import openpyxl as _opx
                            _wb = _opx.load_workbook(tabla_files[0], data_only=True)
                            _sh = _wb.active
                            for _fila in _sh.iter_rows(min_row=3, values_only=True):
                                if _fila[0] is None:
                                    continue
                                try:
                                    if int(_fila[0]) == int(n_evento):
                                        p_desc_ui = float(_fila[3]) if _fila[3] else 0.0
                                        break
                                except (ValueError, TypeError):
                                    pass
                        except Exception:
                            pass

                    if p_desc_ui > 0:
                        st.metric("p_desc registrado (MW)", f"{p_desc_ui:.2f}")
                    else:
                        st.caption("p_desc no encontrado en Tabla_Eventos")

                    #  Identificar unidades del disparo desde el Excel CI 
                    _disp_units = [] # type: ignore
                    if df_pgini is not None and info_ci:
                        _disparo_str = info_ci.get("Disparo", "")
                        _disp_str_clean = re.sub(r"\by\b", ",", _disparo_str, flags=re.IGNORECASE)
                        _sti_disp = {x.strip() for x in _disp_str_clean.split(",") if x.strip() and x.strip() != "nan"}

                        def _sti_de_ui(loc_name):
                            s = re.sub(r"\(\d+\)$", "", str(loc_name).strip())
                            for _pref in ("sym_", "WT_", "PV-", "PV_", "sta_"):
                                if s.lower().startswith(_pref.lower()):
                                    s = s[len(_pref):]
                                    break
                            s = re.sub(r"_EQ$", "", s, flags=re.IGNORECASE)
                            s = re.sub(r"_II$", "", s, flags=re.IGNORECASE)
                            s = re.sub(r"^LOD_", "", s, flags=re.IGNORECASE)
                            return s.strip()

                        for _, _row in df_pgini.iterrows():
                            _loc = str(_row.get("loc_name PF", "")).strip()
                            if _sti_de_ui(_loc) in _sti_disp:
                                _disp_units.append({"loc": _loc, "pgini_actual": float(_row.get("pgini_MW", 0.0))})

                    def _preview_prop_pdesc(units, p_desc): # type: ignore
                        """Vista previa opción 3: proporcional a pgini actual → suma = p_desc."""
                        suma = sum(u["pgini_actual"] for u in units)
                        if suma <= 0 or p_desc <= 0:
                            n = len(units) or 1
                            return {u["loc"]: round(p_desc / n, 2) for u in units}
                        return {u["loc"]: round(u["pgini_actual"] * p_desc / suma, 2) for u in units}

                    modo_disparo = st.radio(
                        "Modo de asignación al disparo:",
                        options=["1", "2", "3"],
                        format_func=lambda x: { # type: ignore
                            "1": "Mantener valores actuales (proporcional)  <- DEFAULT",
                            "2": "Ingreso manual por unidad",
                            "3": "Distribuir p_desc proporcional a pgini actual (respeta Pmax)",
                        }[x],
                        index=0,
                        key="modo_disparo",
                    )

                    pgini_manual = {}

                    #  Tabla de valores y entradas según opción  # type: ignore
                    if _disp_units:
                        def _dif_badge(dif):
                            if abs(dif) < 1.0:
                                st.success(f"Diferencia con p_desc: {dif:+.2f} MW ✓")
                            elif abs(dif) < 5.0:
                                st.warning(f"Diferencia con p_desc: {dif:+.2f} MW")
                            else:
                                st.error(f"Diferencia con p_desc: {dif:+.2f} MW")

                        if modo_disparo == "1":
                            _rows1 = [{"Unidad": u["loc"], "pgini asignado (MW)": round(u["pgini_actual"], 2)} for u in _disp_units]
                            _suma1 = sum(u["pgini_actual"] for u in _disp_units)
                            _rows1.append({"Unidad": "SUMA", "pgini asignado (MW)": round(_suma1, 2)})
                            st.dataframe(pd.DataFrame(_rows1), hide_index=True, use_container_width=False)
                            if p_desc_ui > 0:
                                _dif_badge(_suma1 - p_desc_ui)

                        elif modo_disparo == "2":
                            st.caption("Ingrese la potencia para cada unidad:")
                            _suma2 = 0.0
                            for _u in _disp_units:
                                # Intentar recuperar valor guardado previamente en este evento
                                _saved_manual = _get_unit_cfg(ev_path, _u['loc'], "manual_pgini", float(_u["pgini_actual"]))
                                _val2 = st.number_input(
                                    f"{_u['loc']}  (actual: {_u['pgini_actual']:.2f} MW)",
                                    value=float(_saved_manual),
                                    min_value=0.0,
                                    step=1.0,
                                    format="%.2f",
                                    key=f"pgini_manual_{_u['loc']}",
                                )
                                pgini_manual[_u["loc"]] = _val2
                                _suma2 += _val2

                            if st.button("💾 Guardar potencias manuales", key="save_manual_pgini_btn", use_container_width=True):
                                for _loc_m, _val_m in pgini_manual.items():
                                    _save_unit_cfg(ev_path, _loc_m, "manual_pgini", _val_m)
                                st.toast(f"Potencias manuales guardadas para el evento {n_evento}", icon="✅")

                            st.markdown(f"**Suma:** `{_suma2:.2f} MW`")
                            if p_desc_ui > 0:
                                _dif_badge(_suma2 - p_desc_ui)
                        elif modo_disparo == "3":
                            _prev3 = _preview_prop_pdesc(_disp_units, p_desc_ui) if p_desc_ui > 0 else {u["loc"]: u["pgini_actual"] for u in _disp_units}
                            _rows3 = [{"Unidad": loc, "pgini asignado (MW)": val} for loc, val in _prev3.items()]
                            _suma3 = sum(_prev3.values())
                            _rows3.append({"Unidad": "SUMA", "pgini asignado (MW)": round(_suma3, 2)})
                            st.dataframe(pd.DataFrame(_rows3), hide_index=True, use_container_width=False)
                            if p_desc_ui > 0:
                                _dif3 = _suma3 - p_desc_ui
                                if abs(_dif3) < 0.1:
                                    st.success(f"Diferencia con p_desc: {_dif3:+.2f} MW ✓")
                                else:
                                    st.caption(f"Diferencia: {_dif3:+.2f} MW (aproximado — sin restricción Pmax)")
                    else:
                        st.caption("No se identificaron unidades del disparo en las condiciones iniciales.")

                with col_opt_b:
                    _v4_section_head("Post Load Flow", icon="activity")

                    ajustar_post_lf = st.checkbox(
                        "Activar ajuste post-LF  (AJUSTAR_POST_LF)",
                        value=True, # type: ignore
                        help=(
                            "Si está activo, el script iterará para que la potencia real "
                            "de la slack coincida con su P0_medido, redistribuyendo el "
                            "delta entre unidades CNDC_proporcional."
                        ),
                    )

                    guardar_escenario = st.checkbox(
                        "Guardar escenario de operación al finalizar",
                        value=True,
                        help=(
                            "Llama a escenario.Save() en PowerFactory al terminar la carga, "
                            "guardando pgini/plini y el resultado del Load Flow en el "
                            "IntScenario creado para este evento."
                        ),
                    )

            #  SECCIÓN 5 — EJECUCIÓN 
            st.header("5 · Ejecución")

            _can_run = bool(ci_files)
            if not _can_run:
                st.error("❌ No se puede ejecutar: falta `condiciones_iniciales_*.xlsx`.")

            col_btn, col_reset, col_nota = st.columns([1.2, 1.2, 2.6])
            with col_btn:
                run_btn = st.button(
                    "Ejecutar en PowerFactory",
                    disabled=IS_CLOUD or not _can_run or st.session_state.pf_running,
                    type="primary",
                    use_container_width=True,
                )
            with col_reset:
                if st.button("🔄 Liberar Licencia", help="Cierra procesos colgados de PowerFactory y limpia el estado"):
                    with st.spinner("Limpiando procesos..."):
                        _kill_powerfactory()
                        st.session_state.pf_running = False
                        st.session_state.pf_return_code = None
                        st.session_state.pf_waiting_close = False
                        st.toast("Licencia liberada y procesos terminados", icon="🧹")
                        time.sleep(1)
                        st.rerun()
            with col_nota:
                st.info(
                    "PowerFactory debe estar instalado en esta máquina. "
                    "El proceso puede tardar varios minutos."
                )
            
            if run_btn and ci_files:
                _params = {
                    "semestre":        semestre,
                    "evento":          evento,
                    "RAIZ":            RAIZ,
                    "PF_BASE":         PF_BASE,
                    "LOC_XFO_PATH":    LOC_XFO_PATH,
                    "LOC_GEN_PATH":    LOC_NAMES_GEN_PATH,
                    "PF_PROYECTO":     PF_PROYECTO,
                    "CASO_BASE":       CASO_BASE,
                    "modo_disparo":    modo_disparo if 'modo_disparo' in locals() else "1",
                    "pgini_manual":    pgini_manual if 'pgini_manual' in locals() else {}, # type: ignore
                    "ajustar_post_lf":   ajustar_post_lf  if 'ajustar_post_lf'  in locals() else False,
                    "guardar_escenario": guardar_escenario if 'guardar_escenario' in locals() else True,
                    "excluir_slack":   [s.strip() for s in EXCLUIR_SLACK.split(",") if s.strip()],
                    "xfo_pf":          XFO_PF,
                    "keep_pf_open":    True,
                    "ev_suffix":       ".1",
                }

                _params_path = os.path.join(ev_path, "_streamlit_params.json")
                with open(_params_path, "w", encoding="utf-8") as _fp:
                    json.dump(_params, _fp, ensure_ascii=False, indent=2)

                _runner_path = os.path.join(
                    os.path.dirname(os.path.abspath(__file__)),
                    "runners", "CargaCondIniciales_PF_run.py",
                )
                if not os.path.isfile(_runner_path):
                    if not IS_CLOUD:
                        st.error(f"No se encontró el runner: `{_runner_path}`")
                    st.stop()

                # Limpiar flags residuales de ejecuciones anteriores
                for _old_flag in ("_pf_waiting.flag", "_pf_continue.flag"):
                    _fp_flag = os.path.join(ev_path, _old_flag)
                    if os.path.exists(_fp_flag):
                        os.remove(_fp_flag)

                _status_file = os.path.join(ev_path, "_pf_status.txt")
                if os.path.exists(_status_file):
                    try:
                        os.remove(_status_file)
                    except OSError:
                        pass

                _log_file = os.path.join(ev_path, "_pf_log.txt")
                if os.path.exists(_log_file):
                    try:
                        os.remove(_log_file)
                    except OSError:
                        pass

                st.session_state.pf_return_code   = None
                st.session_state.pf_waiting_close  = False
                st.session_state.pf_running        = True
                st.session_state.pf_status_file    = _status_file
                st.session_state.pf_log_file       = _log_file

                def _pf_thread_fn(runner, params_path, env_vars, status_file, log_file):
                    """Ejecuta el subprocess capturando stdout+stderr al log."""
                    rc = -1
                    try:
                        proc = subprocess.Popen(
                            [sys.executable, runner, params_path],
                            stdout=subprocess.PIPE,
                            stderr=subprocess.STDOUT,
                            text=True,
                            encoding="utf-8",
                            errors="replace",
                            env=env_vars,
                        )
                        with open(log_file, "w", encoding="utf-8") as _lf:
                            for _line in proc.stdout:
                                _lf.write(_line)
                                _lf.flush()
                        rc = proc.wait()
                    except Exception as _exc:
                        try:
                            with open(log_file, "a", encoding="utf-8") as _lf:
                                _lf.write(f"\n[ERROR] {_exc}\n")
                        except OSError:
                            pass
                    finally:
                        with open(status_file, "w", encoding="utf-8") as _sf:
                            _sf.write(str(rc))

                _env = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"}
                threading.Thread(
                    target=_pf_thread_fn,
                    args=(_runner_path, _params_path, _env, _status_file, _log_file),
                    daemon=True,
                ).start()
                st.rerun()

            #  Estado de ejecución + botón "Cerrar PF" 
            _flag_waiting  = os.path.join(ev_path, "_pf_waiting.flag")
            _flag_continue = os.path.join(ev_path, "_pf_continue.flag")
            _status_file   = st.session_state.get("pf_status_file") or os.path.join(ev_path, "_pf_status.txt")
            _log_live      = st.session_state.get("pf_log_file") or os.path.join(ev_path, "_pf_log.txt")

            def _leer_log(path):
                try:
                    return open(path, encoding="utf-8", errors="replace").read()
                except OSError:
                    return ""

            def _guardar_log_final(log_path, ev_path_, n_evento_):
                """Copia el log temporal a un archivo permanente con timestamp."""
                contenido = _leer_log(log_path)
                if not contenido:
                    return None
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                dest = os.path.join(ev_path_, f"log_PF_Ev{n_evento_}_{ts}.txt")
                try:
                    with open(dest, "w", encoding="utf-8") as _f:
                        _f.write(contenido)
                    return dest
                except OSError:
                    return None

            if st.session_state.pf_running:
                # ¿El hilo terminó? → existe el archivo de estado
                if _status_file and os.path.exists(_status_file):
                    try:
                        _rc_text = open(_status_file, encoding="utf-8").read().strip()
                        st.session_state.pf_return_code = int(_rc_text)
                    except (OSError, ValueError):
                        st.session_state.pf_return_code = -1
                    st.session_state.pf_running = False
                    st.session_state.pf_waiting_close = False
                    st.rerun()
                else:
                    # Detectar si el script está esperando que el usuario cierre PF
                    if os.path.exists(_flag_waiting):
                        st.session_state.pf_waiting_close = True

                    if st.session_state.pf_waiting_close:
                        st.success("Datos cargados en PowerFactory. DIgSILENT permanece abierto.")
                        if st.button("🔒 Cerrar PowerFactory", type="primary", use_container_width=True):
                            with open(_flag_continue, "w") as _fc:
                                _fc.write("continue")
                            st.info("Señal enviada — cerrando PowerFactory...")
                            st.session_state.pf_waiting_close = False
                            time.sleep(1)
                            st.rerun()
                    else:
                        st.info("⏳ Carga en PowerFactory en curso...")
                        _monitor_process_fragment(_log_live, _status_file)

            elif st.session_state.pf_return_code is not None:
                _rc = st.session_state.pf_return_code
                if _rc == 0:
                    st.success("Ejecución completada correctamente.")
                elif _rc in (-1073741819, 3221225477):
                    st.warning(
                        "⚠️ PowerFactory se cerró abruptamente (ACCESS_VIOLATION 0xC0000005). "
                        "Revise la **Sección 6** — si el archivo `datos_cargados_Ev*.xlsx` existe, "
                        "los datos **sí fueron cargados** antes del crash."
                    )
                else:
                    st.error(f"❌ Código de error {_rc} (0x{_rc & 0xFFFFFFFF:08X}).")

                _post_files = sorted(
                    glob.glob(os.path.join(ev_path, f"datos_cargados_Ev{n_evento}.1*.xlsx"))
                )
                if _post_files:
                    st.success(
                        f"`{os.path.basename(_post_files[0])}` encontrado — "
                        "datos cargados correctamente. Ver **Sección 6**."
                    )
                else:
                    st.info("ℹ️ No se encontró archivo de resultados.")

                #  Log de la ejecución 
                _saved_log   = st.session_state.get("pf_saved_log")
                _log_content = _leer_log(_saved_log) if _saved_log and os.path.isfile(_saved_log) else _leer_log(_log_live)
                if _log_content:
                    with st.expander("📋 Log de ejecución PowerFactory", expanded=(_rc != 0)):
                        st.code(_log_content, language="")
                        _lc1, _lc2 = st.columns(2)
                        with _lc1:
                            if not (_saved_log and os.path.isfile(_saved_log)):
                                if st.button("💾 Guardar log", key="pf_save_log"):
                                    _dest = _guardar_log_final(_log_live, ev_path, n_evento) # type: ignore
                                    if _dest:
                                        st.session_state.pf_saved_log = _dest
                                        st.success(f"Guardado en: `{_dest}`")
                                    else:
                                        st.error("No se pudo guardar el log.")
                            else: # type: ignore
                                st.caption(f"💾 Guardado en: `{_saved_log}`")
                        with _lc2:
                            st.download_button(
                                "⬇️ Descargar log",
                                data=_log_content.encode("utf-8"),
                                file_name=os.path.basename(_saved_log) if _saved_log else f"log_PF_Ev{n_evento}.txt",
                                mime="text/plain",
                            )

            # Mostrar logs de ejecuciones previas si no hay una activa
            if st.session_state.pf_return_code is None and not st.session_state.pf_running:
                _prev_logs = sorted(
                    glob.glob(os.path.join(ev_path, f"log_PF_Ev{n_evento}_*.txt")),
                    reverse=True,
                )
                if _prev_logs:
                    with st.expander(f"📋 Logs guardados de ejecuciones anteriores ({len(_prev_logs)})"):
                        _sel_log = st.selectbox(
                            "Seleccionar log",
                            _prev_logs,
                            format_func=os.path.basename,
                            key="sel_prev_log_pf",
                        )
                        if _sel_log:
                            st.code(_leer_log(_sel_log), language="")
                            st.download_button(
                                "⬇️ Descargar",
                                data=_leer_log(_sel_log).encode("utf-8"),
                                file_name=os.path.basename(_sel_log),
                                mime="text/plain",
                                key="dl_prev_log_pf",
                            )

            #  SECCIÓN 6 — RESULTADOS
            st.header("6 · Resultados — Proyecto (1)")
            st.caption(f"Archivos con nomenclatura `Ev{n_evento}.1` correspondientes al proyecto `{PF_PROYECTO}`")

            _result_files = sorted(
                glob.glob(os.path.join(ev_path, f"datos_cargados_Ev{n_evento}.1*.xlsx"))
            )

            if not _result_files:
                st.info("Aún no hay archivos de resultados. Ejecute el programa primero.")
            else:
                for _rf in _result_files:
                    _rf_name = os.path.basename(_rf)
                    _v4_section_head(_rf_name, icon="chart")

                    try:
                        _is_ajustado = "ajustado" in _rf_name
                        _sheet_res = "Resumen_Ajustado" if _is_ajustado else "Resumen_Cargado"
                        _sheet_gen = "pgini_GEN_AJUSTADO" if _is_ajustado else "pgini_GEN_FINAL"
                        _sheet_car = "plini_CAR_FINAL"

                        _xl = pd.ExcelFile(_rf, engine="calamine")
                        _df_gen_out = _xl.parse(_sheet_gen)

                        if _is_ajustado and "plini_CAR_AJUSTADO" in _xl.sheet_names:
                            _sheet_car = "plini_CAR_AJUSTADO"
                        elif "plini_CAR_FINAL" not in _xl.sheet_names and "plini_CAR" in _xl.sheet_names:
                            _sheet_car = "plini_CAR"
                        _df_car_out = _xl.parse(_sheet_car)
                        _df_res_out = _xl.parse(_sheet_res)

                        _res_dict = {
                            str(k).strip(): v
                            for k, v in zip(_df_res_out.iloc[:, 0], _df_res_out.iloc[:, 1])
                        }

                        def _v(*keys, default=0.0):
                            for k in keys:
                                raw = _res_dict.get(k)
                                if raw is not None:
                                    try:
                                        return float(raw)
                                    except (ValueError, TypeError):
                                        pass
                            return default

                        _pgen_out  = _df_gen_out["pgini_MW"].sum() if "pgini_MW" in _df_gen_out.columns else _v("Pgen total asignada (MW)", "Pgen total ajustada (MW)")
                        _pdem_out  = _df_car_out["plini_MW"].sum() if "plini_MW" in _df_car_out.columns else _v("Pdem total asignada (MW)", "Pdem total (MW)")
                        _bal_out   = _v("Balance Pgen-Pdem (MW)", default=_pgen_out - _pdem_out)
                        _pdem_ev   = _v("Pdem_evento (MW)")
                        _slack_out = str(_res_dict.get("Slack", "—"))

                        # Potencia asignada a la slack
                        _slack_pgini: float | None = None
                        _raw_sp = _res_dict.get("pgini slack P0_medido (MW)") \
                                  or _res_dict.get("Slack real LF (MW)")
                        if _raw_sp is not None:
                            try:
                                _slack_pgini = float(_raw_sp)
                            except (ValueError, TypeError):
                                pass
                        if _slack_pgini is None and "pgini_MW" in _df_gen_out.columns:
                            _col_loc = next(
                                (c for c in _df_gen_out.columns if "loc_name" in c.lower()), None
                            )
                            if _col_loc:
                                _mask_sl = (
                                    _df_gen_out[_col_loc].astype(str).str.strip()
                                    == _slack_out.strip()
                                )
                                if _mask_sl.any():
                                    _slack_pgini = float(
                                        _df_gen_out.loc[_mask_sl, "pgini_MW"].iloc[0]
                                    )

                        _gens_asig = len(_df_gen_out)
                        _cars_asig = len(_df_car_out)
                        _cars_miss = _v("Cargas no encontradas", "Cargas no encontradas (ajuste)")

                        kc1, kc2, kc3, kc4 = st.columns(4)
                        kc1.metric("Pgen total (MW)", f"{_pgen_out:.2f}")
                        kc2.metric("Pdem total (MW)", f"{_pdem_out:.2f}")
                        kc3.metric(
                            "Balance Pgen−Pdem (MW)",
                            f"{_bal_out:+.2f}",
                            delta=None,
                            help="Diferencia entre generación y demanda asignadas a PF.",
                        )
                        kc4.metric("Pdem evento (MW)", f"{_pdem_ev:.2f}")

                        kc5, kc6, kc7, kc8 = st.columns(4)
                        kc5.metric(
                            f"P slack — {_slack_out}",
                            f"{_slack_pgini:.2f} MW" if _slack_pgini is not None else "—",
                            help="Potencia asignada (pgini) a la unidad slack.",
                        )
                        kc6.metric("Generadores asignados", int(_gens_asig))
                        kc7.metric("Cargas asignadas", int(_cars_asig))
                        kc8.metric("Cargas no encontradas", int(_cars_miss))

                        _estado_val = str(_res_dict.get("Estado validacion", ""))
                        if _estado_val == "OK":
                            st.success("✅ Estado de validación: OK")
                        elif _estado_val:
                            st.warning(f"⚠️ Estado de validación: {_estado_val}")

                        _col_ta, _col_tb = st.columns(2)

                        with _col_ta:
                            with st.expander("📋 Generadores (pgini final)"):
                                st.dataframe(_df_safe(_df_gen_out), use_container_width=False, hide_index=True)

                        with _col_tb:
                            with st.expander(f"📋 Cargas ({_sheet_car})"):
                                st.dataframe(_df_safe(_df_car_out), use_container_width=False, hide_index=True)

                        with st.expander("📋 Resumen completo"):
                            st.dataframe(_df_safe(_df_res_out), use_container_width=False, hide_index=True)

                        with open(_rf, "rb") as _fdown:
                            st.download_button(
                                label=f"⬇️  Descargar {_rf_name}",
                                data=_fdown.read(),
                                file_name=_rf_name,
                                mime=(
                                    "application/vnd.openxmlformats-officedocument"
                                    ".spreadsheetml.sheet"
                                ),
                                key=f"dl_{_rf_name}",
                            )

                    except Exception as _e:
                        st.error(f"Error al leer `{_rf_name}`: {_e}")
        else:
            st.warning("👈 Seleccione Semestre y Evento en la barra lateral para habilitar el bloque de carga.")

    # ═════════════════════════════════════════════════════════════════════════════
    # TAB 4: POWERFACTORY — PROYECTO (2) PERMANENTE
    # ═════════════════════════════════════════════════════════════════════════════
    elif _b1_tab == "pf2":
        _v4_section_head("Carga en PowerFactory — Proyecto (2) Permanente",
            f"Carga condiciones iniciales en el proyecto fijo `{PF_PROYECTO_2}` (no configurable) y ejecuta el flujo de trabajo RMS.",
            "server")

        if semestre and evento:
            st.text_input("Proyecto PowerFactory (fijo — no modificable)", value=PF_PROYECTO_2, disabled=True, key="pf2_proyecto_display")

            #  SECCIÓN — ARCHIVOS DE ENTRADA
            _v4_section_head("Archivos de Entrada", icon="database")

            ci_files    = glob.glob(os.path.join(ev_path, "condiciones_iniciales_*.xlsx"))
            dsim_files  = glob.glob(os.path.join(ev_path, "datos_simulacion_*_2daopcion.xlsx"))
            tabla_files = glob.glob(os.path.join(RAIZ, semestre, "Tabla_Eventos_*.xlsx"))
            xfo_ok      = os.path.isfile(LOC_XFO_PATH)

            costo_marginal_dir   = os.path.join(ev_path, CARPETA_COSTO_MARGINAL)
            costo_marginal_files = []
            if os.path.isdir(costo_marginal_dir):
                costo_marginal_files = (glob.glob(os.path.join(costo_marginal_dir, "postot*.xlsx")) or
                                        glob.glob(os.path.join(costo_marginal_dir, "td_*.xlsx")))
            costo_marginal_found = bool(costo_marginal_files)

            def _estado2(found):
                return "✅ OK" if found else "❌ Falta"

            tabla_archivos2 = pd.DataFrame([
                {
                    "Estado": _estado2(ci_files),
                    "Archivo": os.path.basename(ci_files[0]) if ci_files else "condiciones_iniciales_*.xlsx",
                    "Descripción": "Condiciones iniciales (generadores + cargas)",
                    "Carpeta": os.path.dirname(ci_files[0]) if ci_files else ev_path,
                },
                {
                    "Estado": _estado2(dsim_files),
                    "Archivo": os.path.basename(dsim_files[0]) if dsim_files else "datos_simulacion_*_2daopcion.xlsx",
                    "Descripción": "Pdem del evento (celda B8)",
                    "Carpeta": ev_path,
                },
                {
                    "Estado": _estado2(tabla_files),
                    "Archivo": os.path.basename(tabla_files[0]) if tabla_files else "Tabla_Eventos_*.xlsx",
                    "Descripción": "p_desc del disparo (potencia desconectada)",
                    "Carpeta": os.path.join(RAIZ, semestre),
                },
                {
                    "Estado": _estado2(xfo_ok),
                    "Archivo": os.path.basename(LOC_XFO_PATH),
                    "Descripción": "Capacidad de transformadores (restricción de cargas)",
                    "Carpeta": os.path.dirname(LOC_XFO_PATH),
                },
                {
                    "Estado": _estado2(costo_marginal_found),
                    "Archivo": os.path.basename(costo_marginal_files[0]) if costo_marginal_files else "postot*.xlsx o td_*.xlsx",
                    "Descripción": "Costo Marginal STI (retiros CNDC nodal)",
                    "Carpeta": os.path.dirname(LOC_XFO_PATH),
                },
            ])
            st.dataframe(tabla_archivos2, use_container_width=False, hide_index=True)

            #  SECCIÓN 3 — VISTA PREVIA DE CONDICIONES INICIALES
            if ci_files:
                st.header("3 · Vista Previa de Condiciones Iniciales")
                ci_path = ci_files[0]

                try:
                    xl_ci_prev2 = pd.ExcelFile(ci_path, engine="calamine")
                    df_res_ci2  = xl_ci_prev2.parse("Resumen")
                    df_pgini2   = xl_ci_prev2.parse("pgini_GEN")
                    df_plini2   = xl_ci_prev2.parse("plini_CAR")

                    info_ci2 = dict(
                        zip(
                            df_res_ci2.iloc[:, 0].astype(str).str.strip(),
                            df_res_ci2.iloc[:, 1].astype(str).str.strip(),
                        )
                    )

                    pgen_ci2 = df_pgini2["pgini_MW"].sum() if "pgini_MW" in df_pgini2.columns else 0.0
                    pdem_ci2 = df_plini2["plini_MW"].sum() if "plini_MW" in df_plini2.columns else 0.0

                    k1b, k2b, k3b, k4b = st.columns(4)
                    k1b.metric("Generadores", len(df_pgini2))
                    k2b.metric("Cargas", len(df_plini2))
                    k3b.metric("Pgen Excel (MW)", f"{pgen_ci2:.1f}")
                    k4b.metric("Pdem Excel (MW)", f"{pdem_ci2:.1f}")

                    ia2, ib2 = st.columns(2)
                    with ia2:
                        st.info(f"**Fecha y hora:** {info_ci2.get('Fecha y hora', '—')}")
                        st.info(f"**Disparo:** {info_ci2.get('Disparo', '—')}")
                    with ib2:
                        st.info(f"**Hora generación:** {info_ci2.get('Hora evento (gen)', '—')}")
                        st.info(f"**Hora cargas:** {info_ci2.get('Hora Po (cargas)', '—')}")

                    with st.expander("📋 Generadores (pgini_GEN) — primeras 30 filas"):
                        st.dataframe(df_pgini2.head(30), use_container_width=False, hide_index=True)

                    with st.expander("📋 Cargas (plini_CAR) — primeras 30 filas"):
                        st.dataframe(df_plini2.head(30), use_container_width=False, hide_index=True)

                except Exception as _e2:
                    st.error(f"Error al leer condiciones_iniciales: {_e2}")
                    df_pgini2 = None
                    info_ci2  = {}
            else:
                st.warning("No se encontró `condiciones_iniciales_*.xlsx`. Primero genere las condiciones iniciales.")
                df_pgini2 = None
                info_ci2  = {}
                st.info("💡 Use el **módulo 2** para generar las condiciones iniciales.")

            #  SECCIÓN 4 — OPCIONES DE EJECUCIÓN
            if ci_files:
                st.header("4 · Opciones de Ejecución")

                col_opt_a2, col_opt_b2 = st.columns(2)

                with col_opt_a2:
                    _v4_section_head("Potencia del disparo", icon="bolt")

                    p_desc_ui2 = 0.0
                    if tabla_files:
                        try:
                            import openpyxl as _opx2
                            _wb2 = _opx2.load_workbook(tabla_files[0], data_only=True)
                            _sh2 = _wb2.active
                            for _fila2 in _sh2.iter_rows(min_row=3, values_only=True):
                                if _fila2[0] is None:
                                    continue
                                try:
                                    if int(_fila2[0]) == int(n_evento):
                                        p_desc_ui2 = float(_fila2[3]) if _fila2[3] else 0.0
                                        break
                                except (ValueError, TypeError):
                                    pass
                        except Exception:
                            pass

                    if p_desc_ui2 > 0:
                        st.metric("p_desc registrado (MW)", f"{p_desc_ui2:.2f}")
                    else:
                        st.caption("p_desc no encontrado en Tabla_Eventos")

                    _disp_units2 = []
                    if df_pgini2 is not None and info_ci2:
                        _disparo_str2   = info_ci2.get("Disparo", "")
                        _disp_str_clean2 = re.sub(r"\by\b", ",", _disparo_str2, flags=re.IGNORECASE)
                        _sti_disp2 = {x.strip() for x in _disp_str_clean2.split(",") if x.strip() and x.strip() != "nan"}

                        def _sti_de_ui2(loc_name):
                            s = re.sub(r"\(\d+\)$", "", str(loc_name).strip())
                            for _pref in ("sym_", "WT_", "PV-", "PV_", "sta_"):
                                if s.lower().startswith(_pref.lower()):
                                    s = s[len(_pref):]
                                    break
                            s = re.sub(r"_EQ$", "", s, flags=re.IGNORECASE)
                            s = re.sub(r"_II$", "", s, flags=re.IGNORECASE)
                            s = re.sub(r"^LOD_", "", s, flags=re.IGNORECASE)
                            return s.strip()

                        for _, _row2 in df_pgini2.iterrows():
                            _loc2 = str(_row2.get("loc_name PF", "")).strip()
                            if _sti_de_ui2(_loc2) in _sti_disp2:
                                _disp_units2.append({"loc": _loc2, "pgini_actual": float(_row2.get("pgini_MW", 0.0))})

                    def _preview_prop_pdesc2(units, p_desc):
                        suma = sum(u["pgini_actual"] for u in units)
                        if suma <= 0 or p_desc <= 0:
                            n = len(units) or 1
                            return {u["loc"]: round(p_desc / n, 2) for u in units}
                        return {u["loc"]: round(u["pgini_actual"] * p_desc / suma, 2) for u in units}

                    modo_disparo2 = st.radio(
                        "Modo de asignación al disparo:",
                        options=["1", "2", "3"],
                        format_func=lambda x: {
                            "1": "Mantener valores actuales (proporcional)  <- DEFAULT",
                            "2": "Ingreso manual por unidad",
                            "3": "Distribuir p_desc proporcional a pgini actual (respeta Pmax)",
                        }[x],
                        index=0,
                        key="modo_disparo_pf2",
                    )

                    pgini_manual2 = {}

                    if _disp_units2:
                        def _dif_badge2(dif):
                            if abs(dif) < 1.0:
                                st.success(f"Diferencia con p_desc: {dif:+.2f} MW ✓")
                            elif abs(dif) < 5.0:
                                st.warning(f"Diferencia con p_desc: {dif:+.2f} MW")
                            else:
                                st.error(f"Diferencia con p_desc: {dif:+.2f} MW")

                        if modo_disparo2 == "1":
                            _rows1b = [{"Unidad": u["loc"], "pgini asignado (MW)": round(u["pgini_actual"], 2)} for u in _disp_units2]
                            _suma1b = sum(u["pgini_actual"] for u in _disp_units2)
                            _rows1b.append({"Unidad": "SUMA", "pgini asignado (MW)": round(_suma1b, 2)})
                            st.dataframe(pd.DataFrame(_rows1b), hide_index=True, use_container_width=False)
                            if p_desc_ui2 > 0:
                                _dif_badge2(_suma1b - p_desc_ui2)

                        elif modo_disparo2 == "2":
                            st.caption("Ingrese la potencia para cada unidad:")
                            _suma2b = 0.0
                            for _u2 in _disp_units2:
                                _saved_manual2 = _get_unit_cfg(ev_path, _u2['loc'], "manual_pgini2", float(_u2["pgini_actual"]))
                                _val2b = st.number_input(
                                    f"{_u2['loc']}  (actual: {_u2['pgini_actual']:.2f} MW)",
                                    value=float(_saved_manual2),
                                    min_value=0.0,
                                    step=1.0,
                                    format="%.2f",
                                    key=f"pgini2_manual_{_u2['loc']}",
                                )
                                pgini_manual2[_u2["loc"]] = _val2b
                                _suma2b += _val2b

                            if st.button("💾 Guardar potencias manuales", key="save_manual_pgini_btn_pf2", use_container_width=True):
                                for _loc_m2, _val_m2 in pgini_manual2.items():
                                    _save_unit_cfg(ev_path, _loc_m2, "manual_pgini2", _val_m2)
                                st.toast(f"Potencias manuales guardadas para el evento {n_evento} (Proyecto 2)", icon="✅")

                            st.markdown(f"**Suma:** `{_suma2b:.2f} MW`")
                            if p_desc_ui2 > 0:
                                _dif_badge2(_suma2b - p_desc_ui2)

                        elif modo_disparo2 == "3":
                            _prev3b = _preview_prop_pdesc2(_disp_units2, p_desc_ui2) if p_desc_ui2 > 0 else {u["loc"]: u["pgini_actual"] for u in _disp_units2}
                            _rows3b = [{"Unidad": loc, "pgini asignado (MW)": val} for loc, val in _prev3b.items()]
                            _suma3b = sum(_prev3b.values())
                            _rows3b.append({"Unidad": "SUMA", "pgini asignado (MW)": round(_suma3b, 2)})
                            st.dataframe(pd.DataFrame(_rows3b), hide_index=True, use_container_width=False)
                            if p_desc_ui2 > 0:
                                _dif3b = _suma3b - p_desc_ui2
                                if abs(_dif3b) < 0.1:
                                    st.success(f"Diferencia con p_desc: {_dif3b:+.2f} MW ✓")
                                else:
                                    st.caption(f"Diferencia: {_dif3b:+.2f} MW (aproximado — sin restricción Pmax)")
                    else:
                        st.caption("No se identificaron unidades del disparo en las condiciones iniciales.")

                with col_opt_b2:
                    _v4_section_head("Post Load Flow", icon="activity")

                    ajustar_post_lf2 = st.checkbox(
                        "Activar ajuste post-LF  (AJUSTAR_POST_LF)",
                        value=True,
                        help=(
                            "Si está activo, el script iterará para que la potencia real "
                            "de la slack coincida con su P0_medido, redistribuyendo el "
                            "delta entre unidades CNDC_proporcional."
                        ),
                        key="ajustar_post_lf_pf2",
                    )

                    guardar_escenario2 = st.checkbox(
                        "Guardar escenario de operación al finalizar",
                        value=True,
                        help=(
                            "Llama a escenario.Save() en PowerFactory al terminar la carga, "
                            "guardando pgini/plini y el resultado del Load Flow en el "
                            "IntScenario creado para este evento."
                        ),
                        key="guardar_escenario_pf2",
                    )

            #  SECCIÓN 5 — EJECUCIÓN
            st.header("5 · Ejecución")

            _can_run2 = bool(ci_files)
            if not _can_run2:
                st.error("❌ No se puede ejecutar: falta `condiciones_iniciales_*.xlsx`.")

            col_btn2, col_reset2, col_nota2 = st.columns([1.2, 1.2, 2.6])
            with col_btn2:
                run_btn2 = st.button(
                    "Ejecutar en PowerFactory (2)",
                    disabled=IS_CLOUD or not _can_run2 or st.session_state.pf2_running,
                    type="primary",
                    use_container_width=True,
                    key="run_pf2_btn",
                )
            with col_reset2:
                if st.button("🔄 Liberar Licencia", help="Cierra procesos colgados de PowerFactory y limpia el estado", key="liberar_pf2_btn"):
                    with st.spinner("Limpiando procesos..."):
                        _kill_powerfactory()
                        st.session_state.pf2_running = False
                        st.session_state.pf2_return_code = None
                        st.session_state.pf2_waiting_close = False
                        st.toast("Licencia liberada y procesos terminados", icon="🧹")
                        time.sleep(1)
                        st.rerun()
            with col_nota2:
                st.info(
                    f"Cargará en el proyecto fijo `{PF_PROYECTO_2}`. "
                    "PowerFactory debe estar instalado en esta máquina. "
                    "El proceso puede tardar varios minutos."
                )

            if run_btn2 and ci_files:
                _params2 = {
                    "semestre":        semestre,
                    "evento":          evento,
                    "RAIZ":            RAIZ,
                    "PF_BASE":         PF_BASE,
                    "LOC_XFO_PATH":    LOC_XFO_PATH,
                    "LOC_GEN_PATH":    LOC_NAMES_GEN_PATH,
                    "PF_PROYECTO":     PF_PROYECTO_2,
                    "CASO_BASE":       CASO_BASE,
                    "modo_disparo":    modo_disparo2 if 'modo_disparo2' in locals() else "1",
                    "pgini_manual":    pgini_manual2 if 'pgini_manual2' in locals() else {},
                    "ajustar_post_lf":   ajustar_post_lf2  if 'ajustar_post_lf2'  in locals() else False,
                    "guardar_escenario": guardar_escenario2 if 'guardar_escenario2' in locals() else True,
                    "excluir_slack":   [s.strip() for s in EXCLUIR_SLACK.split(",") if s.strip()],
                    "xfo_pf":          XFO_PF,
                    "keep_pf_open":    True,
                    "ev_suffix":       ".2",
                }

                _params2_path = os.path.join(ev_path, "_streamlit_params2.json")
                with open(_params2_path, "w", encoding="utf-8") as _fp2:
                    json.dump(_params2, _fp2, ensure_ascii=False, indent=2)

                _runner_path2 = os.path.join(
                    os.path.dirname(os.path.abspath(__file__)),
                    "runners", "CargaCondIniciales_PF_run.py",
                )
                if not os.path.isfile(_runner_path2):
                    if not IS_CLOUD:
                        st.error(f"No se encontró el runner: `{_runner_path2}`")
                    st.stop()

                for _old_flag2 in ("_pf2_waiting.flag", "_pf2_continue.flag"):
                    _fp_flag2 = os.path.join(ev_path, _old_flag2)
                    if os.path.exists(_fp_flag2):
                        os.remove(_fp_flag2)

                _status_file2 = os.path.join(ev_path, "_pf2_status.txt")
                if os.path.exists(_status_file2):
                    try:
                        os.remove(_status_file2)
                    except OSError:
                        pass

                _log_file2 = os.path.join(ev_path, "_pf2_log.txt")
                if os.path.exists(_log_file2):
                    try:
                        os.remove(_log_file2)
                    except OSError:
                        pass

                st.session_state.pf2_return_code   = None
                st.session_state.pf2_waiting_close  = False
                st.session_state.pf2_running        = True
                st.session_state.pf2_status_file    = _status_file2
                st.session_state.pf2_log_file       = _log_file2

                def _pf2_thread_fn(runner, params_path, env_vars, status_file, log_file):
                    """Ejecuta el subprocess de PF proyecto (2) capturando stdout+stderr al log."""
                    rc = -1
                    try:
                        proc = subprocess.Popen(
                            [sys.executable, runner, params_path],
                            stdout=subprocess.PIPE,
                            stderr=subprocess.STDOUT,
                            text=True,
                            encoding="utf-8",
                            errors="replace",
                            env=env_vars,
                        )
                        with open(log_file, "w", encoding="utf-8") as _lf2:
                            for _line2 in proc.stdout:
                                _lf2.write(_line2)
                                _lf2.flush()
                        rc = proc.wait()
                    except Exception as _exc2:
                        try:
                            with open(log_file, "a", encoding="utf-8") as _lf2:
                                _lf2.write(f"\n[ERROR] {_exc2}\n")
                        except OSError:
                            pass
                    finally:
                        with open(status_file, "w", encoding="utf-8") as _sf2:
                            _sf2.write(str(rc))

                _env2 = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"}
                threading.Thread(
                    target=_pf2_thread_fn,
                    args=(_runner_path2, _params2_path, _env2, _status_file2, _log_file2),
                    daemon=True,
                ).start()
                st.rerun()

            #  Estado de ejecución + botón "Cerrar PF"
            _flag_waiting2  = os.path.join(ev_path, "_pf2_waiting.flag")
            _flag_continue2 = os.path.join(ev_path, "_pf2_continue.flag")
            _status_file2   = st.session_state.get("pf2_status_file") or os.path.join(ev_path, "_pf2_status.txt")
            _log_live2      = st.session_state.get("pf2_log_file") or os.path.join(ev_path, "_pf2_log.txt")

            def _leer_log2(path):
                try:
                    return open(path, encoding="utf-8", errors="replace").read()
                except OSError:
                    return ""

            def _guardar_log_final2(log_path, ev_path_, n_evento_):
                contenido = _leer_log2(log_path)
                if not contenido:
                    return None
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                dest = os.path.join(ev_path_, f"log_PF2_Ev{n_evento_}_{ts}.txt")
                try:
                    with open(dest, "w", encoding="utf-8") as _f2:
                        _f2.write(contenido)
                    return dest
                except OSError:
                    return None

            if st.session_state.pf2_running:
                if _status_file2 and os.path.exists(_status_file2):
                    try:
                        _rc_text2 = open(_status_file2, encoding="utf-8").read().strip()
                        st.session_state.pf2_return_code = int(_rc_text2)
                    except (OSError, ValueError):
                        st.session_state.pf2_return_code = -1
                    st.session_state.pf2_running = False
                    st.session_state.pf2_waiting_close = False
                    st.rerun()
                else:
                    if os.path.exists(_flag_waiting2):
                        st.session_state.pf2_waiting_close = True

                    if st.session_state.pf2_waiting_close:
                        st.success("Datos cargados en PowerFactory (2). DIgSILENT permanece abierto.")
                        if st.button("🔒 Cerrar PowerFactory (2)", type="primary", use_container_width=True, key="cerrar_pf2_btn"):
                            with open(_flag_continue2, "w") as _fc2:
                                _fc2.write("continue")
                            st.info("Señal enviada — cerrando PowerFactory...")
                            st.session_state.pf2_waiting_close = False
                            time.sleep(1)
                            st.rerun()
                    else:
                        st.info("⏳ Carga en PowerFactory (2) en curso...")
                        _monitor_process_fragment(_log_live2, _status_file2)

            elif st.session_state.pf2_return_code is not None:
                _rc2 = st.session_state.pf2_return_code
                if _rc2 == 0:
                    st.success("Ejecución completada correctamente.")
                elif _rc2 in (-1073741819, 3221225477):
                    st.warning(
                        "⚠️ PowerFactory se cerró abruptamente (ACCESS_VIOLATION 0xC0000005). "
                        "Revise la **Sección 6** — si el archivo `datos_cargados_Ev*.2.xlsx` existe, "
                        "los datos **sí fueron cargados** antes del crash."
                    )
                else:
                    st.error(f"❌ Código de error {_rc2} (0x{_rc2 & 0xFFFFFFFF:08X}).")

                _post_files2 = sorted(
                    glob.glob(os.path.join(ev_path, f"datos_cargados_Ev{n_evento}.2*.xlsx"))
                )
                if _post_files2:
                    st.success(
                        f"`{os.path.basename(_post_files2[0])}` encontrado — "
                        "datos cargados correctamente. Ver **Sección 6**."
                    )
                else:
                    st.info("ℹ️ No se encontró archivo de resultados.")

                _saved_log2   = st.session_state.get("pf2_saved_log")
                _log_content2 = _leer_log2(_saved_log2) if _saved_log2 and os.path.isfile(_saved_log2) else _leer_log2(_log_live2)
                if _log_content2:
                    with st.expander("📋 Log de ejecución PowerFactory (2)", expanded=(_rc2 != 0)):
                        st.code(_log_content2, language="")
                        _lc1b, _lc2b = st.columns(2)
                        with _lc1b:
                            if not (_saved_log2 and os.path.isfile(_saved_log2)):
                                if st.button("💾 Guardar log", key="pf2_save_log"):
                                    _dest2 = _guardar_log_final2(_log_live2, ev_path, n_evento)
                                    if _dest2:
                                        st.session_state.pf2_saved_log = _dest2
                                        st.success(f"Guardado en: `{_dest2}`")
                                    else:
                                        st.error("No se pudo guardar el log.")
                            else:
                                st.caption(f"💾 Guardado en: `{_saved_log2}`")
                        with _lc2b:
                            st.download_button(
                                "⬇️ Descargar log",
                                data=_log_content2.encode("utf-8"),
                                file_name=os.path.basename(_saved_log2) if _saved_log2 else f"log_PF2_Ev{n_evento}.txt",
                                mime="text/plain",
                                key="dl_log_pf2",
                            )

            if st.session_state.pf2_return_code is None and not st.session_state.pf2_running:
                _prev_logs2 = sorted(
                    glob.glob(os.path.join(ev_path, f"log_PF2_Ev{n_evento}_*.txt")),
                    reverse=True,
                )
                if _prev_logs2:
                    with st.expander(f"📋 Logs guardados de ejecuciones anteriores ({len(_prev_logs2)})"):
                        _sel_log2 = st.selectbox(
                            "Seleccionar log",
                            _prev_logs2,
                            format_func=os.path.basename,
                            key="sel_prev_log_pf2",
                        )
                        if _sel_log2:
                            st.code(_leer_log2(_sel_log2), language="")
                            st.download_button(
                                "⬇️ Descargar",
                                data=_leer_log2(_sel_log2).encode("utf-8"),
                                file_name=os.path.basename(_sel_log2),
                                mime="text/plain",
                                key="dl_prev_log_pf2",
                            )

            #  SECCIÓN 6 — RESULTADOS
            st.header("6 · Resultados — Proyecto (2)")
            st.caption(f"Archivos con nomenclatura `Ev{n_evento}.2` correspondientes al proyecto fijo `{PF_PROYECTO_2}`")

            _result_files2 = sorted(
                glob.glob(os.path.join(ev_path, f"datos_cargados_Ev{n_evento}.2*.xlsx"))
            )

            if not _result_files2:
                st.info("Aún no hay archivos de resultados. Ejecute el programa primero.")
            else:
                for _rf2 in _result_files2:
                    _rf_name2 = os.path.basename(_rf2)
                    _v4_section_head(_rf_name2, icon="chart")

                    try:
                        _is_ajustado2 = "ajustado" in _rf_name2
                        _sheet_res2 = "Resumen_Ajustado" if _is_ajustado2 else "Resumen_Cargado"
                        _sheet_gen2 = "pgini_GEN_AJUSTADO" if _is_ajustado2 else "pgini_GEN_FINAL"
                        _sheet_car2 = "plini_CAR_FINAL"

                        _xl2 = pd.ExcelFile(_rf2, engine="calamine")
                        _df_gen_out2 = _xl2.parse(_sheet_gen2)

                        if _is_ajustado2 and "plini_CAR_AJUSTADO" in _xl2.sheet_names:
                            _sheet_car2 = "plini_CAR_AJUSTADO"
                        elif "plini_CAR_FINAL" not in _xl2.sheet_names and "plini_CAR" in _xl2.sheet_names:
                            _sheet_car2 = "plini_CAR"
                        _df_car_out2 = _xl2.parse(_sheet_car2)
                        _df_res_out2 = _xl2.parse(_sheet_res2)

                        _res_dict2 = {
                            str(k).strip(): v
                            for k, v in zip(_df_res_out2.iloc[:, 0], _df_res_out2.iloc[:, 1])
                        }

                        def _v2(*keys, default=0.0):
                            for k in keys:
                                raw = _res_dict2.get(k)
                                if raw is not None:
                                    try:
                                        return float(raw)
                                    except (ValueError, TypeError):
                                        pass
                            return default

                        _pgen_out2  = _df_gen_out2["pgini_MW"].sum() if "pgini_MW" in _df_gen_out2.columns else _v2("Pgen total asignada (MW)", "Pgen total ajustada (MW)")
                        _pdem_out2  = _df_car_out2["plini_MW"].sum() if "plini_MW" in _df_car_out2.columns else _v2("Pdem total asignada (MW)", "Pdem total (MW)")
                        _bal_out2   = _v2("Balance Pgen-Pdem (MW)", default=_pgen_out2 - _pdem_out2)
                        _pdem_ev2   = _v2("Pdem_evento (MW)")
                        _slack_out2 = str(_res_dict2.get("Slack", "—"))

                        _slack_pgini2: float | None = None
                        _raw_sp2 = _res_dict2.get("pgini slack P0_medido (MW)") \
                                   or _res_dict2.get("Slack real LF (MW)")
                        if _raw_sp2 is not None:
                            try:
                                _slack_pgini2 = float(_raw_sp2)
                            except (ValueError, TypeError):
                                pass
                        if _slack_pgini2 is None and "pgini_MW" in _df_gen_out2.columns:
                            _col_loc2 = next(
                                (c for c in _df_gen_out2.columns if "loc_name" in c.lower()), None
                            )
                            if _col_loc2:
                                _mask_sl2 = (
                                    _df_gen_out2[_col_loc2].astype(str).str.strip()
                                    == _slack_out2.strip()
                                )
                                if _mask_sl2.any():
                                    _slack_pgini2 = float(
                                        _df_gen_out2.loc[_mask_sl2, "pgini_MW"].iloc[0]
                                    )

                        _gens_asig2 = len(_df_gen_out2)
                        _cars_asig2 = len(_df_car_out2)
                        _cars_miss2 = _v2("Cargas no encontradas", "Cargas no encontradas (ajuste)")

                        kc1b, kc2b, kc3b, kc4b = st.columns(4)
                        kc1b.metric("Pgen total (MW)", f"{_pgen_out2:.2f}")
                        kc2b.metric("Pdem total (MW)", f"{_pdem_out2:.2f}")
                        kc3b.metric(
                            "Balance Pgen−Pdem (MW)",
                            f"{_bal_out2:+.2f}",
                            delta=None,
                            help="Diferencia entre generación y demanda asignadas a PF.",
                        )
                        kc4b.metric("Pdem evento (MW)", f"{_pdem_ev2:.2f}")

                        kc5b, kc6b, kc7b, kc8b = st.columns(4)
                        kc5b.metric(
                            f"P slack — {_slack_out2}",
                            f"{_slack_pgini2:.2f} MW" if _slack_pgini2 is not None else "—",
                            help="Potencia asignada (pgini) a la unidad slack.",
                        )
                        kc6b.metric("Generadores asignados", int(_gens_asig2))
                        kc7b.metric("Cargas asignadas", int(_cars_asig2))
                        kc8b.metric("Cargas no encontradas", int(_cars_miss2))

                        _estado_val2 = str(_res_dict2.get("Estado validacion", ""))
                        if _estado_val2 == "OK":
                            st.success("✅ Estado de validación: OK")
                        elif _estado_val2:
                            st.warning(f"⚠️ Estado de validación: {_estado_val2}")

                        _col_ta2, _col_tb2 = st.columns(2)

                        with _col_ta2:
                            with st.expander("📋 Generadores (pgini final)"):
                                st.dataframe(_df_safe(_df_gen_out2), use_container_width=False, hide_index=True)

                        with _col_tb2:
                            with st.expander(f"📋 Cargas ({_sheet_car2})"):
                                st.dataframe(_df_safe(_df_car_out2), use_container_width=False, hide_index=True)

                        with st.expander("📋 Resumen completo"):
                            st.dataframe(_df_safe(_df_res_out2), use_container_width=False, hide_index=True)

                        with open(_rf2, "rb") as _fdown2:
                            st.download_button(
                                label=f"⬇️  Descargar {_rf_name2}",
                                data=_fdown2.read(),
                                file_name=_rf_name2,
                                mime=(
                                    "application/vnd.openxmlformats-officedocument"
                                    ".spreadsheetml.sheet"
                                ),
                                key=f"dl2_{_rf_name2}",
                            )

                    except Exception as _e2r:
                        st.error(f"Error al leer `{_rf_name2}`: {_e2r}")
        else:
            st.warning("👈 Seleccione Semestre y Evento en la barra lateral para habilitar el bloque de carga.")

elif bloque_trabajo == "analisis_datos":
    _render_block_header("03", "Análisis SCADA / EMF",
        "Procesa registros SCADA (1SEG) y curvas EMF del CNDC, detecta t₀ automáticamente y calcula KPIs CNDC.",
        "Análisis", pf_required=False)

    if not st.session_state.semestre_global or not st.session_state.evento_global:
        st.warning("👈 Seleccione Semestre y Evento en la barra lateral.")
        st.stop()

    ev_path = st.session_state.ev_path_global
    n_evento = st.session_state.n_evento_global
    _sel_unit = st.session_state.global_selected_unit

    #  1. LÓGICA DE CARGA Y SINCRONIZACIÓN POR UNIDAD 
    if _sel_unit:
        _pmax_map_b3 = _load_pmax_cargado(ev_path, n_evento)
        _tmap_b3 = _load_tech_map(LOC_NAMES_GEN_PATH)
        _pm_v_b3, _, _ = _get_pmax_from_cargado(_sel_unit, _pmax_map_b3, _tmap_b3)
        _pmax_ref = float(_pm_v_b3)
    else:
        _pmax_ref = 200.0

    semestre = st.session_state.semestre_global
    evento = st.session_state.evento_global
    _sel_unit = st.session_state.global_selected_unit
    _event_cfg = _load_event_cfg(ev_path)   # Config a nivel de evento (t₀ global SCADA/EMF)

    if not _sel_unit:
        st.warning("⬆️ Seleccione una unidad en el selector superior para ver el análisis.")
        st.stop()

    #  Tab bar persistente 
    _b3_tab = _v4_tab_bar([
        {"id": "scada", "icon": "activity", "label": "SCADA COBEE (1SEG)"},
        {"id": "emf",   "icon": "chart",    "label": "EMF CNDC"},
        {"id": "comp",  "icon": "scale",    "label": "Comparativa SCADA vs CNDC"},
    ], "b03")

    # ═════════════════════════════════════════════════════════════════════════
    # TAB 1: SCADA COBEE
    # ═════════════════════════════════════════════════════════════════════════
    if _b3_tab == "scada":
        _v4_section_head(
            "Procesamiento de Registros SCADA (1 Segundo)",
            "Busca el archivo '1 seg' en la carpeta de FALLA del CNDC y organiza "
            "la potencia y frecuencia en archivos CSV individuales por unidad.",
            "activity",
        )

        scada_dir = os.path.join(ev_path, "Graficas Registro 1SEG COBEE")
        
        col1, col2 = st.columns([1, 2])
        _scada_runner = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "runners", "OrdenadorDatosEvento_run.py",
        )
        _can_scada = os.path.isfile(_scada_runner)
        if not _can_scada and not IS_CLOUD:
            st.error(f"No se encontró el runner: `{_scada_runner}`")

        with col1:
            scada_btn = st.button(
                "Ejecutar Ordenador de Datos SCADA",
                type="primary",
                use_container_width=True,
                disabled=not _can_scada or st.session_state.scada_running,
            )

        _scada_status_file = st.session_state.get("scada_status_file") or os.path.join(ev_path, "_scada_status.txt")

        if scada_btn and _can_scada:
            if IS_CLOUD and _SP_OK:
                try:
                    import OrdenadorDatosEvento as _ode
                    _ode.RAIZ_RPF = RAIZ  # apuntar a la ruta temporal descargada
                    _fe = _ode.leer_fecha_evento(semestre, int(n_evento))
                    if _fe:
                        with st.spinner(f"⬇️ Descargando registros de falla ({_fe.strftime('%d.%m.%y')}) desde SharePoint..."):
                            _sp.descargar_scada_falla(_fe)
                            st.toast("Carpeta de FALLA descargada desde SharePoint.", icon="📂")
                    else:
                        st.error("No se pudo identificar la fecha del evento en Tabla_Eventos.")
                        st.stop()
                except Exception as _e:
                    st.error(f"Error descargando datos desde SharePoint: {_e}")
                    st.stop()

            _scada_status_f = os.path.join(ev_path, "_scada_status.txt")
            if os.path.exists(_scada_status_f):
                try:
                    os.remove(_scada_status_f)
                except OSError:
                    pass

            _scada_params = {
                "semestre":   semestre,
                "evento":     evento,
                "RAIZ_RPF":   RAIZ,
                "RAIZ_DATOS": RAIZ_DATOS,
            }
            _scada_params_path = os.path.join(ev_path, "_scada_params.json")
            with open(_scada_params_path, "w", encoding="utf-8") as _fp:
                json.dump(_scada_params, _fp, ensure_ascii=False, indent=2)

            _scada_log_f = os.path.join(ev_path, "_scada_log.txt")
            st.session_state.scada_running    = True
            st.session_state.scada_status_file = _scada_status_f
            st.session_state.scada_log_file    = _scada_log_f
            st.session_state.scada_return_code = None
            st.session_state.scada_saved_log   = None

            def _scada_thread_fn(runner, params_path, env_vars, status_file, log_file):
                rc = -1
                try:
                    proc = subprocess.Popen(
                        [sys.executable, runner, params_path],
                        stdout=subprocess.PIPE,
                        stderr=subprocess.STDOUT,
                        text=True,
                        encoding="utf-8",
                        errors="replace",
                        env=env_vars,
                    )
                    with open(log_file, "w", encoding="utf-8") as _lf:
                        for _line in proc.stdout:
                            _lf.write(_line)
                            _lf.flush()
                    rc = proc.wait()
                except Exception as _exc:
                    try:
                        with open(log_file, "a", encoding="utf-8") as _lf:
                            _lf.write(f"\n[ERROR] {_exc}\n")
                    except OSError:
                        pass
                finally:
                    with open(status_file, "w", encoding="utf-8") as _sf:
                        _sf.write(str(rc))

            _scada_env = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"}
            threading.Thread(
                target=_scada_thread_fn,
                args=(_scada_runner, _scada_params_path, _scada_env, _scada_status_f, _scada_log_f),
                daemon=True,
            ).start()
            st.rerun()

        if os.path.isdir(scada_dir):
            xlsx_files_scada = _listar_archivos_cache(scada_dir, "*.xlsx")
            if xlsx_files_scada:
                st.success(f"Se encontraron {len(xlsx_files_scada)} unidades procesadas.")

                _scada_file = _buscar_archivo_unidad(_sel_unit, xlsx_files_scada)
                if _scada_file:
                    df_scada = pd.read_excel(os.path.join(scada_dir, _scada_file), engine="calamine").dropna()
                    t_raw    = _parse_to_seconds(df_scada.iloc[:, 0])
                    # Si se pide HH:MM:SS, usamos el tiempo original (t_raw) para ver la hora del día
                    # Si no, usamos el tiempo relativo (t_norm) que empieza en 0
                    t_base   = t_raw if show_hhmmss else (t_raw - t_raw.min())
                    t_norm   = t_raw - t_raw.min() # Para cálculos internos de KPIs e índices
                    unit_col = df_scada.columns[2]
                    _freq_b2_arr = pd.to_numeric(df_scada['Frecuencia_Hz'], errors='coerce').ffill().values
                    _pot_b2_arr  = pd.to_numeric(df_scada[unit_col], errors='coerce').ffill().values

                    # Inicialización de límites de ejes X para SCADA (evita NameError)
                    xaxis_min_sc = _get_unit_cfg(ev_path, _sel_unit, "b3_tab_scada_xmin", float(t_norm.min()))
                    xaxis_max_sc = _get_unit_cfg(ev_path, _sel_unit, "b3_tab_scada_xmax", float(t_norm.max()))

                    #  Parámetros CNDC (antes del gráfico) 
                    _v4_section_head("Parámetros CNDC", icon="sliders")
                    _pmax_cargado_b2 = _load_pmax_cargado(ev_path, n_evento) # type: ignore
                    _tmap_b2         = _load_tech_map(LOC_NAMES_GEN_PATH)
                    _b2_pm_val, _tk_b2, _b2_pm_fuente = _get_pmax_from_cargado(
                        _sel_unit, _pmax_cargado_b2, _tmap_b2 # type: ignore
                    )
                    if _b2_pm_fuente:
                        st.caption(f"✅ Unidad: **{_tk_b2}** — P_max desde `{_b2_pm_fuente}`")
                    else:
                        st.warning(f"⚠️ No se encontró **{_sel_unit}** en datos_cargados ni loc_names_gen.")

                    _b2_pmax = float(_b2_pm_val)
                    _b2_rp_pct = float(_get_rp_default(_tk_b2, LOC_NAMES_GEN_PATH))
                    #  Fila 1: parámetros de detección 
                    _pp1, _pp2, _pp3 = st.columns(3)
                    _b2_dt        = _pp1.number_input("Δt CNDC [s]", value=35, min_value=20, max_value=60, step=1, key="b2_sc_dt")
                    _b2_umbral_k  = _pp2.number_input("Umbral df/dt [Hz/s]", value=-0.04, min_value=-2.0, max_value=-0.001, step=0.005, format="%.3f", key="b2_sc_umbral")
                    _b2_vent_suav = _pp3.number_input("Ventana suavizado", value=5, min_value=2, max_value=20, step=1, key="b2_sc_vsuav")

                    _idx_auto_b2 = _detectar_inicio_falla(_freq_b2_arr, float(_b2_umbral_k), int(_b2_vent_suav))
                    _t_auto_b2   = float(t_norm.iloc[_idx_auto_b2])

                    #  Fila 2: label + hint + botones en la misma línea 
                    _lf1, _lf2, _lf3 = st.columns([5, 1, 1])
                    _lf1.markdown(
                        f"**t₀ inicio de falla [s]** "
                        f"<span style='color:#888;font-size:11px;margin-left:6px'>"
                        f"Auto-detectado: {_t_auto_b2:.1f} s</span>",
                        unsafe_allow_html=True,
                    )
                    if _lf2.button("↩ Auto", key="reset_b2sc_t0",
                                   help=f"Restaurar al tiempo auto-detectado ({_t_auto_b2:.1f} s)"):
                        st.session_state["b2_sc_t_falla"] = _t_auto_b2
                        st.rerun()
                    if _lf3.button("💾 Guardar", key="save_idx_scada",
                                   help="Guardar t₀ para TODAS las unidades del evento (tab SCADA)"):
                        _t0_sv = st.session_state.get("b2_sc_t_falla", _t_auto_b2)
                        _idx_sv = int(np.argmin(np.abs(t_norm.values - _t0_sv)))
                        _save_event_cfg(ev_path, "scada_t0_s", float(t_norm.iloc[_idx_sv]))
                        _save_unit_cfg(ev_path, _sel_unit, "scada_wall_clock_t0", float(t_raw.iloc[_idx_sv]))
                        st.toast(f"t₀ SCADA = {_t0_sv:.1f} s guardado para todas las unidades", icon="✅")

                    #  Fila 3: input numérico (sin label visible) 
                    _t_input_b2 = st.number_input(
                        "t₀ inicio de falla [s]",
                        value=st.session_state.get("b2_sc_t_falla", _t_auto_b2),
                        min_value=float(t_norm.min()),
                        max_value=float(t_norm.max()),
                        step=1.0, format="%.1f",
                        key="b2_sc_t_falla",
                        label_visibility="collapsed",
                    )
                    _idx_falla_b2 = int(np.argmin(np.abs(t_norm.values - _t_input_b2)))
                    _t_falla_abs  = float(t_norm.iloc[_idx_falla_b2])
                    _t_al_b2      = (t_norm - t_norm.iloc[_idx_falla_b2]).values
                    _kpi_b2       = _cndc_kpis(_t_al_b2, _freq_b2_arr, _pot_b2_arr,
                                               _b2_pmax, _b2_rp_pct / 100.0, int(_b2_dt))
                    _rocof_b2     = _calcular_rocof(_t_al_b2, _freq_b2_arr, 3.0)

                    #  Fila de métricas + controles de escala estandarizados
                    _cm1, _cm2, _cm3 = st.columns(3)
                    _cm1.metric("t falla", f"{_t_falla_abs:.1f} s")
                    _cm2.metric("f₀",      f"{_freq_b2_arr[_idx_falla_b2]:.4f} Hz")
                    _cm3.metric("P₀",      f"{_pot_b2_arr[_idx_falla_b2]:.3f} MW")
                    xaxis_min, xaxis_max, yaxis1_min, yaxis1_max, yaxis2_min, yaxis2_max, auto_scale_sc, _sc_chart_col = \
                        _render_axis_controls("scada", ev_path, _sel_unit,
                                              float(t_norm.min()), float(t_norm.max()),
                                              float(_b2_pmax * 1.1),
                                              traces=[(t_norm.values, _pot_b2_arr)])

                    #  Gráfico con marcadores CNDC (usando funciones estándares)  # type: ignore
                    _gcfg = st.session_state.graph_config

                    # Crear gráfica dual eje usando función estándar
                    fig = create_dual_axis_timeseries(
                        t_data=t_base,
                        freq_data=_freq_b2_arr,
                        pot_data=_pot_b2_arr,
                        title=f"Registro SCADA con puntos CNDC — {_scada_file}",
                        freq_label="Frecuencia SCADA (Hz)",
                        pot_label=f"Potencia {unit_col} (MW)",
                        show_hhmmss=show_hhmmss,
                        freq_color=_gcfg["freq_color_real"],
                        pot_color=_gcfg["pot_color_real"],
                        line_width=_gcfg["line_width"],
                        template=_gcfg["template"],
                        height=_gcfg["plot_height"],
                        legend_position="bottom_center",
                        x_range=[xaxis_min, xaxis_max] if not show_hhmmss else None,
                        y1_range=[yaxis1_min, yaxis1_max],
                        y2_range=[yaxis2_min, yaxis2_max],
                    )
                    
                    # Añadir líneas de referencia (banda muerta, t₀, t₀+Δt)
                    if _kpi_b2:
                        # Con HH:MM:SS, la base del gráfico es t_raw (tiempo absoluto del día).
                        # Las referencias y marcadores deben estar en la MISMA escala que t_base.
                        _t_falla_plot = float(t_raw.iloc[_idx_falla_b2]) if show_hhmmss else _t_falla_abs

                        # Líneas de referencia: se pasa _t_falla_plot para que add_reference_lines
                        # aplique _to_plotly_time correctamente (segundos absolutos → ms de época).
                        fig = add_reference_lines(
                            fig,
                            t_fault_abs=_t_falla_plot,
                            t_eval_abs=_t_falla_plot + int(_b2_dt),
                            show_hhmmss=show_hhmmss,
                            show_deadband=_gcfg["show_deadband"],
                            show_fault_line=True,
                            show_eval_line=True,
                            eval_line_label=f"t₀+Δt ({int(_b2_dt)} s)",
                        )

                        # Marcadores KPI: convertir a la misma escala del eje X con _to_plotly_time.
                        # show_hhmmss=False → devuelve segundos sin cambio.
                        # show_hhmmss=True  → convierte segundos a ms de época (lo que espera el eje 'date').
                        _t0_plot   = _to_plotly_time(_t_falla_plot, show_hhmmss)
                        _tmin_plot = _to_plotly_time(_t_falla_plot + float(_kpi_b2['t_min']), show_hhmmss)
                        _tdt_plot  = _to_plotly_time(_t_falla_plot + int(_b2_dt), show_hhmmss)

                        fig = add_kpi_markers(
                            fig,
                            t_fault_abs=_t_falla_plot,
                            kpi_dict=_kpi_b2,
                            show_hhmmss=show_hhmmss,
                            dt_seconds=int(_b2_dt),
                            marker_size=_gcfg["marker_size"],
                            freq_color=_gcfg["freq_color_real"],
                            pot_color=_gcfg["pot_color_real"],
                            t0_plot=_t0_plot,
                            tmin_plot=_tmin_plot,
                            tdt_plot=_tdt_plot,
                        )
                        # P_máxima en ventana [t_nadir, t₀+Δt]
                        _t_nadir_b2 = float(_kpi_b2['t_min']) if _kpi_b2 else 0.0
                        _t_pmax_b2_al, _p_pmax_b2 = _find_pmax_time(
                            _t_al_b2, _pot_b2_arr, int(_b2_dt), t_min_eval=_t_nadir_b2
                        )
                        if _t_pmax_b2_al is not None and _gcfg.get("show_pmax_marker", True):
                            _idx_pm_b2  = int(np.argmin(np.abs(_t_al_b2 - _t_pmax_b2_al)))
                            _f_pmax_b2  = float(_freq_b2_arr[_idx_pm_b2])
                            _tpmax_b2_plot = _to_plotly_time(_t_falla_plot + _t_pmax_b2_al, show_hhmmss)
                            fig = add_pmax_marker(
                                fig, _tpmax_b2_plot, _p_pmax_b2, _f_pmax_b2,
                                pot_color=_gcfg["pot_color_real"],
                                freq_color=_gcfg["freq_color_real"],
                                marker_size=_gcfg["marker_size"],
                            )
                        elif _t_pmax_b2_al is None:
                            _p_pmax_b2 = None
                    else:
                        _t_pmax_b2_al = None
                        _p_pmax_b2    = None
                        # Banda muerta sin KPI
                        fig = add_reference_lines(
                            fig,
                            show_hhmmss=show_hhmmss,
                            show_deadband=_gcfg["show_deadband"],
                            show_fault_line=False,
                            show_eval_line=False,
                        )

                    _sc_chart_col.plotly_chart(fig, use_container_width=True, key="b2sc_plotly_chart")

                    with st.expander("📄 Ver tabla de datos"):
                        st.dataframe(_df_safe(df_scada), use_container_width=False)
                        if st.button("⬇️ Descargar datos SCADA a Excel", key=f"dl_scada_data_{_sel_unit}"):
                            excel_data = _apply_excel_formatting(
                                df_scada,
                                sheet_name=f"SCADA_{_sel_unit}",
                            )
                            st.download_button(
                                f"Descargar SCADA {_sel_unit}",
                                excel_data,
                                file_name=f"scada_data_{_sel_unit}_Ev{n_evento}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                    #  Tabla KPIs CNDC
                    st.markdown("---") # type: ignore
                    _v4_section_head("KPIs CNDC — Criterio RPF", "Registro Real SCADA COBEE", icon="chart")
                    _kpi_pmax_b2 = None
                    if _t_pmax_b2_al is not None:
                        _kpi_pmax_b2 = _cndc_kpis(
                            _t_al_b2, _freq_b2_arr, _pot_b2_arr,
                            _b2_pmax, _b2_rp_pct / 100.0, _t_pmax_b2_al,
                        )
                    _mostrar_tabla_cndc_duo(
                        _kpi_b2, _b2_pmax, int(_b2_dt), "SCADA COBEE (1SEG)", rocof=_rocof_b2,
                        kpi2=_kpi_pmax_b2, delta_t2=_t_pmax_b2_al, fuente2="P_máxima SCADA",
                    )

                elif _sel_unit:
                    st.warning(f"La unidad **{_sel_unit}** no tiene datos SCADA procesados para este evento.")
            else:
                st.warning("No se encontraron archivos Excel en 'Graficas Registro 1SEG COBEE'.")
        else:
            st.info("ℹ️ Presione el botón para organizar los datos del SCADA.")

        _scada_log_live = st.session_state.get("scada_log_file") or os.path.join(ev_path, "_scada_log.txt")

        def _leer_scada_log(path):
            try:
                return open(path, encoding="utf-8", errors="replace").read()
            except OSError:
                return ""

        def _guardar_scada_log(log_path, ev_path_, n_ev_):
            contenido = _leer_scada_log(log_path)
            if not contenido:
                return None
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest = os.path.join(ev_path_, f"log_SCADA_Ev{n_ev_}_{ts}.txt")
            try:
                with open(dest, "w", encoding="utf-8") as _f:
                    _f.write(contenido)
                return dest
            except OSError:
                return None

        if st.session_state.scada_running:
            if _scada_status_file and os.path.exists(_scada_status_file):
                try:
                    _scada_rc = int(open(_scada_status_file, encoding="utf-8").read().strip())
                except (OSError, ValueError):
                    _scada_rc = -1
                st.session_state.scada_return_code = _scada_rc
                st.session_state.scada_running = False
                _listar_archivos_cache.clear()
                st.rerun()
            else:
                st.info("⏳ Procesamiento SCADA en curso...")
                _monitor_process_fragment(_scada_log_live, _scada_status_file)
        elif st.session_state.scada_return_code is not None:
            _scada_rc = st.session_state.scada_return_code
            if _scada_rc == 0:
                st.success("✅ Procesamiento SCADA completado.")
            else:
                st.error(f"❌ Error en procesamiento SCADA (código {_scada_rc}).")
            _scada_saved = st.session_state.get("scada_saved_log")
            _scada_log_content = _leer_scada_log(_scada_saved) if _scada_saved and os.path.isfile(_scada_saved) else _leer_scada_log(_scada_log_live)
            if _scada_log_content:
                with st.expander("📋 Log de ejecución SCADA", expanded=(_scada_rc != 0)):
                    st.code(_scada_log_content, language="")
                    _sl1, _sl2 = st.columns(2)
                    with _sl1:
                        if not (_scada_saved and os.path.isfile(_scada_saved)):
                            if st.button("💾 Guardar log", key="scada_save_log"):
                                _dest = _guardar_scada_log(_scada_log_live, ev_path, n_evento)
                                if _dest:
                                    st.session_state.scada_saved_log = _dest
                                    st.success(f"Guardado en: `{_dest}`")
                                else:
                                    st.error("No se pudo guardar el log.")
                        else:
                            st.caption(f"💾 Guardado en: `{_scada_saved}`")
                    with _sl2:
                        st.download_button(
                            "⬇️ Descargar log",
                            data=_scada_log_content.encode("utf-8"),
                            file_name=os.path.basename(_scada_saved) if _scada_saved else f"log_SCADA_Ev{n_evento}.txt",
                            mime="text/plain",
                            key="scada_dl_log",
                        )

    # ═════════════════════════════════════════════════════════════════════════
    # TAB 2: EMF CNDC
    # ═════════════════════════════════════════════════════════════════════════
    elif _b3_tab == "emf":
        _v4_section_head(
            "Extracción de Datos desde Gráficos EMF CNDC",
            "Digitaliza archivos EMF (Enhanced Metafile) para extraer los puntos "
            "exactos de frecuencia y potencia graficados por el CNDC.",
            "chart",
        )

        emf_dir = os.path.join(ev_path, CARPETA_COBEE_EMF)
        
        col1, col2 = st.columns([1, 2])
        _emf_runner = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "runners", "ExtractorResultadosCNDC_run.py",
        )
        _can_emf = os.path.isfile(_emf_runner)
        if not _can_emf and not IS_CLOUD:
            st.error(f"No se encontró el runner: `{_emf_runner}`")

        with col1:
            emf_btn = st.button(
                "Ejecutar Extractor de Gráficos EMF",
                type="primary",
                use_container_width=True,
                disabled=not _can_emf or st.session_state.emf_running,
            )

        _emf_status_file = st.session_state.get("emf_status_file") or os.path.join(ev_path, "_emf_status.txt")

        if emf_btn and _can_emf:
            _emf_status_f = os.path.join(ev_path, "_emf_status.txt")
            if os.path.exists(_emf_status_f):
                try:
                    os.remove(_emf_status_f)
                except OSError:
                    pass

            _emf_params = {
                "semestre":      semestre,
                "evento":        evento,
                "RAIZ":          RAIZ,
                "CARPETA_COBEE": CARPETA_COBEE_EMF,
            }
            _emf_params_path = os.path.join(ev_path, "_emf_params.json")
            with open(_emf_params_path, "w", encoding="utf-8") as _fp:
                json.dump(_emf_params, _fp, ensure_ascii=False, indent=2)

            _emf_log_f = os.path.join(ev_path, "_emf_log.txt")
            st.session_state.emf_running    = True
            st.session_state.emf_status_file = _emf_status_f
            st.session_state.emf_log_file    = _emf_log_f
            st.session_state.emf_return_code = None
            st.session_state.emf_saved_log   = None

            def _emf_thread_fn(runner, params_path, env_vars, status_file, log_file):
                rc = -1
                try:
                    proc = subprocess.Popen(
                        [sys.executable, runner, params_path],
                        stdout=subprocess.PIPE,
                        stderr=subprocess.STDOUT,
                        text=True,
                        encoding="utf-8",
                        errors="replace",
                        env=env_vars,
                    )
                    with open(log_file, "w", encoding="utf-8") as _lf:
                        for _line in proc.stdout:
                            _lf.write(_line)
                            _lf.flush()
                    rc = proc.wait()
                except Exception as _exc:
                    try:
                        with open(log_file, "a", encoding="utf-8") as _lf:
                            _lf.write(f"\n[ERROR] {_exc}\n")
                    except OSError:
                        pass
                finally:
                    with open(status_file, "w", encoding="utf-8") as _sf:
                        _sf.write(str(rc))

            _emf_env = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"}
            threading.Thread(
                target=_emf_thread_fn,
                args=(_emf_runner, _emf_params_path, _emf_env, _emf_status_f, _emf_log_f),
                daemon=True,
            ).start()
            st.rerun()

        if os.path.isdir(emf_dir):
            # Buscar los Excels generados por el extractor
            xlsx_files = _listar_archivos_cache(emf_dir, "*.xlsx")
            if xlsx_files:
                st.success(f"Se encontraron {len(xlsx_files)} gráficos digitalizados.")

                _emf_file = _buscar_archivo_unidad(_sel_unit, xlsx_files)
                if _emf_file:
                    df_emf = pd.read_excel(os.path.join(emf_dir, _emf_file), engine="calamine").dropna()

                    # Detectar columnas
                    col_tiempo = 'tiempo_s' if 'tiempo_s' in df_emf.columns else df_emf.columns[0]
                    col_freq = 'frecuencia_hz' if 'frecuencia_hz' in df_emf.columns else None

                    # Ordenar por tiempo ascendente (el extractor EMF puede producir orden descendente)
                    _t_sort = _parse_to_seconds(df_emf[col_tiempo])
                    df_emf = df_emf.assign(_t_sort=_t_sort).sort_values('_t_sort').drop(columns=['_t_sort']).reset_index(drop=True)

                    # Normalizar tiempo (hh:mm:ss -> segundos -> relativo a 0)
                    t_raw = _parse_to_seconds(df_emf[col_tiempo])
                    t_norm = t_raw - t_raw.min()

                    # La columna de potencia es la que no es tiempo ni frecuencia
                    cols_pot = [c for c in df_emf.columns if c not in [col_tiempo, col_freq, 'hora']]

                    # Guard: columnas mínimas requeridas para graficar
                    if col_freq is None:
                        st.warning(
                            f"⚠️ El archivo **{_emf_file}** no contiene columna `frecuencia_hz`. "
                            "El extractor EMF no encontró curva de frecuencia en el archivo .emf. "
                            "Re-ejecute el extractor para regenerar el Excel."
                        )
                        st.stop()
                    if not cols_pot:
                        st.warning(
                            f"⚠️ El archivo **{_emf_file}** no contiene columna de potencia reconocida. "
                            "Re-ejecute el extractor EMF."
                        )
                        st.stop()

                    # Recuperar el clock de SCADA para sincronizar el formato HH:MM:SS
                    _scada_wall_t0 = _get_unit_cfg(ev_path, _sel_unit, "scada_wall_clock_t0", 0.0)

                    #  Integración de Metodología CNDC en pestaña EMF 
                    _v4_section_head("Parámetros de Análisis", "Metodología CNDC", icon="sliders")
                    _pmax_cargado_emf = _load_pmax_cargado(ev_path, n_evento)
                    _tmap_emf         = _load_tech_map(LOC_NAMES_GEN_PATH)
                    _emf_pm_val, _tk_emf, _emf_pm_fuente = _get_pmax_from_cargado(
                        _sel_unit, _pmax_cargado_emf, _tmap_emf
                    )

                    _gcfg = st.session_state.graph_config
                    _emf_pmax = float(_emf_pm_val)
                    _emf_rp_pct = float(_get_rp_default(_tk_emf, LOC_NAMES_GEN_PATH))

                    _ep1, _ep2 = st.columns(2)
                    _emf_dt       = _ep1.number_input("Δt CNDC [s]", value=35, min_value=20, max_value=60, key="b2_emf_dt")
                    _emf_umbral_k = _ep2.number_input("Umbral df/dt [Hz/s]", value=-0.04, format="%.3f", key="b2_emf_um")

                    # Detección y Análisis
                    _freq_emf_arr = pd.to_numeric(df_emf[col_freq], errors='coerce').ffill().values
                    _pot_emf_arr = pd.to_numeric(df_emf[cols_pot[0]], errors='coerce').ffill().values
                    _initial_auto_idx_emf = _detectar_inicio_falla(_freq_emf_arr, _emf_umbral_k)
                    _t_auto_emf = float(t_norm.iloc[_initial_auto_idx_emf])

                    _cemf1, _cemf2, _cemf_btn = st.columns([3, 1, 1])
                    _t_input_emf = _cemf1.number_input(
                        "t₀ inicio de falla [s]",
                        value=st.session_state.get("b2_emf_t_falla", _t_auto_emf),
                        min_value=float(t_norm.min()),
                        max_value=float(t_norm.max()),
                        step=1.0, format="%.1f",
                        key="b2_emf_t_falla",
                        help=f"Escriba el tiempo exacto en segundos. Auto-detectado: {_t_auto_emf:.1f} s",
                    )
                    _idx_falla_emf = int(np.argmin(np.abs(t_norm.values - _t_input_emf)))
                    if _cemf2.button("↩ Auto", key="reset_b2emf_t0",
                                     help=f"Restaurar al tiempo auto-detectado ({_t_auto_emf:.1f} s)"):
                        st.session_state["b2_emf_t_falla"] = _t_auto_emf
                        st.rerun()
                    if _cemf_btn.button("💾 Guardar t₀ EMF", key="save_idx_emf",
                                        help="Guardar t₀ para TODAS las unidades del evento (tab EMF)"):
                        _save_event_cfg(ev_path, "emf_t0_s", float(t_norm.iloc[_idx_falla_emf]))
                        st.toast(f"t₀ EMF = {_t_input_emf:.1f} s guardado para todas las unidades", icon="✅")

                    #  Controles de escala EMF
                    xaxis_min, xaxis_max, yaxis1_min, yaxis1_max, yaxis2_min, yaxis2_max, auto_scale_emf, _emf_chart_col = \
                        _render_axis_controls("emf", ev_path, _sel_unit,
                                              float(t_norm.min()), float(t_norm.max()),
                                              float(_emf_pmax * 1.1),
                                              traces=[(t_norm.values, _pot_emf_arr)])

                    _t_falla_emf = float(t_norm.iloc[_idx_falla_emf])
                    _t_al_emf = (t_norm - _t_falla_emf).values
                    
                    _kpi_emf = _cndc_kpis(_t_al_emf, _freq_emf_arr, _pot_emf_arr, _emf_pmax, _emf_rp_pct/100.0, _emf_dt)
                    
                    #  Gráfico EMF con metodología CNDC (usando funciones estándares) 
                    # Con HH:MM:SS: sincronizar EMF al reloj de pared del SCADA.
                    # Sin HH:MM:SS: usar tiempo relativo normalizado (t_norm).
                    if show_hhmmss and _scada_wall_t0 > 0:
                        # Proyectar tiempo relativo EMF sobre el reloj absoluto del SCADA:
                        # t=0 (falla EMF) → _scada_wall_t0
                        _t_base_emf = t_norm - _t_falla_emf + _scada_wall_t0
                    else:
                        _t_base_emf = t_raw if show_hhmmss else t_norm
                    fig_emf = create_dual_axis_timeseries(
                        t_data=_t_base_emf,
                        freq_data=_freq_emf_arr,
                        pot_data=_pot_emf_arr,
                        title=f"Análisis Metodológico EMF - {_emf_file}",
                        freq_label="Frecuencia CNDC (Hz)",
                        pot_label=f"Potencia {cols_pot[0]} (MW)",
                        show_hhmmss=show_hhmmss,
                        freq_color="cyan",
                        pot_color=_gcfg["pot_color_sim0"],
                        line_width=_gcfg["line_width"],
                        template=_gcfg["template"],
                        height=_gcfg["plot_height"],
                        legend_position="bottom_center",
                        x_range=[xaxis_min, xaxis_max] if not show_hhmmss else None,
                        y1_range=[yaxis1_min, yaxis1_max],
                        y2_range=[yaxis2_min, yaxis2_max],
                    )

                    # Añadir líneas de referencia y marcadores KPI
                    if _kpi_emf:
                        # Con HH:MM:SS: tiempo de falla en la escala de _t_base_emf (sincronizado con SCADA).
                        # Sin HH:MM:SS: _t_falla_emf (segundos relativos).
                        _t_falla_emf_plot = float(_t_base_emf.iloc[_idx_falla_emf]) if show_hhmmss else _t_falla_emf
                        fig_emf = add_reference_lines(
                            fig_emf,
                            t_fault_abs=_t_falla_emf_plot,
                            t_eval_abs=_t_falla_emf_plot + _emf_dt,
                            show_hhmmss=show_hhmmss,
                            show_deadband=_gcfg["show_deadband"],
                            show_fault_line=True,
                            show_eval_line=True,
                            eval_line_label=f"t₀+Δt ({_emf_dt}s)",
                        )

                        _t0_plot_emf   = _to_plotly_time(_t_falla_emf_plot, show_hhmmss)
                        _tmin_plot_emf = _to_plotly_time(_t_falla_emf_plot + float(_kpi_emf['t_min']), show_hhmmss)
                        _tdt_plot_emf  = _to_plotly_time(_t_falla_emf_plot + int(_emf_dt), show_hhmmss)

                        fig_emf = add_kpi_markers(
                            fig_emf,
                            t_fault_abs=_t_falla_emf_plot,
                            kpi_dict=_kpi_emf,
                            show_hhmmss=show_hhmmss,
                            dt_seconds=_emf_dt,
                            marker_size=_gcfg["marker_size"],
                            freq_color="cyan",
                            pot_color=_gcfg["pot_color_sim0"],
                            t0_plot=_t0_plot_emf,
                            tmin_plot=_tmin_plot_emf,
                            tdt_plot=_tdt_plot_emf,
                        )
                        # P_máxima en ventana [t_nadir, t₀+Δt]
                        _t_nadir_emf = float(_kpi_emf['t_min']) if _kpi_emf else 0.0
                        _t_pmax_emf_al, _p_pmax_emf = _find_pmax_time(
                            _t_al_emf, _pot_emf_arr, _emf_dt, t_min_eval=_t_nadir_emf
                        )
                        if _t_pmax_emf_al is not None and _gcfg.get("show_pmax_marker", True):
                            _idx_pm_emf  = int(np.argmin(np.abs(_t_al_emf - _t_pmax_emf_al)))
                            _f_pmax_emf  = float(_freq_emf_arr[_idx_pm_emf])
                            _tpmax_emf_plot = _to_plotly_time(_t_falla_emf_plot + _t_pmax_emf_al, show_hhmmss)
                            fig_emf = add_pmax_marker(
                                fig_emf, _tpmax_emf_plot, _p_pmax_emf, _f_pmax_emf,
                                pot_color=_gcfg["pot_color_sim0"],
                                freq_color="cyan",
                                marker_size=_gcfg["marker_size"],
                            )
                        elif _t_pmax_emf_al is None:
                            _p_pmax_emf = None
                    else:
                        _t_pmax_emf_al = None
                        _p_pmax_emf    = None
                        fig_emf = add_reference_lines(
                            fig_emf,
                            show_hhmmss=show_hhmmss,
                            show_deadband=_gcfg["show_deadband"],
                            show_fault_line=False,
                            show_eval_line=False,
                        )

                    _emf_chart_col.plotly_chart(fig_emf, use_container_width=True, key="b2emf_plotly_chart")

                    _v4_section_head("KPIs CNDC — Criterio RPF", "Registro EMF CNDC", icon="chart")
                    if st.button("Descargar datos EMF a Excel", key=f"dl_emf_data_{_sel_unit}"): # type: ignore
                        excel_data = _apply_excel_formatting(
                            df_emf,
                            sheet_name=f"EMF_{_sel_unit}",
                        )
                        st.download_button(
                            f"Descargar EMF {_sel_unit}",
                            excel_data,
                            file_name=f"emf_data_{_sel_unit}_Ev{n_evento}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    _kpi_pmax_emf = None
                    if _t_pmax_emf_al is not None:
                        _kpi_pmax_emf = _cndc_kpis(
                            _t_al_emf, _freq_emf_arr, _pot_emf_arr,
                            _emf_pmax, _emf_rp_pct / 100.0, _t_pmax_emf_al,
                        )
                    _mostrar_tabla_cndc_duo(
                        _kpi_emf, _emf_pmax, _emf_dt, "Gráfico EMF CNDC",
                        kpi2=_kpi_pmax_emf, delta_t2=_t_pmax_emf_al, fuente2="P_máxima EMF",
                    )
                elif _sel_unit:
                    st.warning(f"La unidad **{_sel_unit}** no tiene datos EMF procesados para este evento.")
            else:
                # Mostrar archivos EMF disponibles pero no procesados # type: ignore
                emfs = [f for f in os.listdir(emf_dir) if f.lower().endswith('.emf')]
                if emfs:
                    st.info(f"Se detectaron {len(emfs)} archivos .emf listos para extraer. Presione el botón superior.")
                    for e in emfs[:5]:
                        st.caption(f"• {e}")
                    if len(emfs) > 5: st.caption(f"... y {len(emfs)-5} más.")
                else:
                    st.warning("No se encontraron archivos .emf en la carpeta 'Resultados_COBEE'.")
        else:
            st.info(f"ℹ️ La carpeta '{CARPETA_COBEE_EMF}' no existe en este evento.")

        _emf_log_live = st.session_state.get("emf_log_file") or os.path.join(ev_path, "_emf_log.txt")

        def _leer_emf_log(path):
            try:
                return open(path, encoding="utf-8", errors="replace").read()
            except OSError:
                return ""

        def _guardar_emf_log(log_path, ev_path_, n_ev_):
            contenido = _leer_emf_log(log_path)
            if not contenido:
                return None
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest = os.path.join(ev_path_, f"log_EMF_Ev{n_ev_}_{ts}.txt")
            try:
                with open(dest, "w", encoding="utf-8") as _f:
                    _f.write(contenido)
                return dest
            except OSError:
                return None

        if st.session_state.emf_running:
            if _emf_status_file and os.path.exists(_emf_status_file):
                try:
                    _emf_rc = int(open(_emf_status_file, encoding="utf-8").read().strip())
                except (OSError, ValueError):
                    _emf_rc = -1
                st.session_state.emf_return_code = _emf_rc
                st.session_state.emf_running = False
                _listar_archivos_cache.clear()
                st.rerun()
            else:
                st.info("⏳ Extracción EMF en curso...")
                _monitor_process_fragment(_emf_log_live, _emf_status_file)
        elif st.session_state.emf_return_code is not None:
            _emf_rc = st.session_state.emf_return_code
            if _emf_rc == 0:
                st.success("✅ Extracción EMF completada.")
            else:
                st.error(f"❌ Error en extracción EMF (código {_emf_rc}).")
            _emf_saved = st.session_state.get("emf_saved_log")
            _emf_log_content = _leer_emf_log(_emf_saved) if _emf_saved and os.path.isfile(_emf_saved) else _leer_emf_log(_emf_log_live)
            if _emf_log_content:
                with st.expander("📋 Log de ejecución EMF", expanded=(_emf_rc != 0)):
                    st.code(_emf_log_content, language="")
                    _eml1, _eml2 = st.columns(2)
                    with _eml1:
                        if not (_emf_saved and os.path.isfile(_emf_saved)):
                            if st.button("💾 Guardar log", key="emf_save_log"):
                                _dest = _guardar_emf_log(_emf_log_live, ev_path, n_evento)
                                if _dest:
                                    st.session_state.emf_saved_log = _dest
                                    st.success(f"Guardado en: `{_dest}`")
                                else:
                                    st.error("No se pudo guardar el log.")
                        else:
                            st.caption(f"💾 Guardado en: `{_emf_saved}`")
                    with _eml2:
                        st.download_button(
                            "⬇️ Descargar log",
                            data=_emf_log_content.encode("utf-8"),
                            file_name=os.path.basename(_emf_saved) if _emf_saved else f"log_EMF_Ev{n_evento}.txt",
                            mime="text/plain",
                            key="emf_dl_log",
                        )

    # ═════════════════════════════════════════════════════════════════════════
    # TAB 3: COMPARATIVA
    # ═════════════════════════════════════════════════════════════════════════
    elif _b3_tab == "comp":
        _v4_section_head("Comparativa Dinámica: SCADA vs CNDC", "", "scale")
        
        scada_dir = os.path.join(ev_path, "Graficas Registro 1SEG COBEE")
        emf_dir = os.path.join(ev_path, "Resultados_COBEE") # type: ignore

        if not _sel_unit:
            st.info("ℹ️ Seleccione una unidad en el selector superior para ver la comparativa.")
        elif os.path.isdir(scada_dir) and os.path.isdir(emf_dir):
            _s_file = _buscar_archivo_unidad(_sel_unit, _listar_archivos_cache(scada_dir, "*.xlsx"))
            _e_file = _buscar_archivo_unidad(_sel_unit, _listar_archivos_cache(emf_dir, "*.xlsx"))
            _has_s = _s_file is not None
            _has_e = _e_file is not None

            # Cargar tiempos de falla guardados en el evento (segundos relativos)
            t0_scada_s = _get_unit_cfg(ev_path, _sel_unit, "scada_t0_s", 0.0)
            t0_emf_s   = _get_unit_cfg(ev_path, _sel_unit, "emf_t0_s", 0.0)

            if _has_s and _has_e:
                # --- CARGA Y ALINEACIÓN DE DATOS ---
                df_s = pd.read_excel(os.path.join(scada_dir, _s_file), engine="calamine").dropna()
                df_e = pd.read_excel(os.path.join(emf_dir, _e_file), engine="calamine").dropna()

                ts_raw = _parse_to_seconds(df_s.iloc[:, 0])
                df_s['t_norm'] = ts_raw - ts_raw.min() 
                t_scada_aligned = df_s['t_norm'] - t0_scada_s

                te_raw = _parse_to_seconds(df_e['tiempo_s'])
                df_e['t_norm'] = te_raw - te_raw.min() 
                t_emf_aligned = df_e['t_norm'] - t0_emf_s

                # --- CONSTRUCCIÓN DEL GRÁFICO ESTANDARIZADO ---
                _gcfg = st.session_state.graph_config
                p_col_s = df_s.columns[2]
                p_col_e = [c for c in df_e.columns if c not in ['tiempo_s', 'frecuencia_hz', 'hora', 't_norm']][0]

                # Recuperar hora real del SCADA
                _scada_wall_t0 = _get_unit_cfg(ev_path, _sel_unit, "scada_wall_clock_t0", 0.0)
                
                # Preparar ejes X (Relativo vs Reloj)
                if show_hhmmss and _scada_wall_t0 > 0:
                    t_scada_plot = t_scada_aligned + _scada_wall_t0
                    t_emf_plot = t_emf_aligned + _scada_wall_t0
                else:
                    t_scada_plot = t_scada_aligned
                    t_emf_plot = t_emf_aligned

                # 1. Crear base con SCADA (igual que Tab 1)
                fig_c = create_dual_axis_timeseries(
                    t_data=t_scada_plot,
                    freq_data=df_s['Frecuencia_Hz'],
                    pot_data=df_s[p_col_s],
                    title=f"Comparativa Registro Real vs CNDC — {_sel_unit}",
                    freq_label="Frec. SCADA",
                    pot_label="Pot. SCADA",
                    show_hhmmss=show_hhmmss,
                    freq_color=_gcfg["freq_color_real"],
                    pot_color=_gcfg["pot_color_real"],
                    line_width=_gcfg["line_width"],
                    template=_gcfg["template"],
                    height=_gcfg["plot_height"]
                )

                # 2. Añadir capas de CNDC (EMF) usando colores distintos para diferenciar fuente
                #    (SCADA = tonos "real", EMF = tonos "sim" para que se vean claramente)
                fig_c.add_trace(go.Scatter(
                    x=_to_plotly_time(t_emf_plot, show_hhmmss),
                    y=df_e['frecuencia_hz'],
                    name="Frec. CNDC (EMF)",
                    line=dict(color="cyan", width=_gcfg["line_width"]),
                    yaxis="y",
                ))
                fig_c.add_trace(go.Scatter(
                    x=_to_plotly_time(t_emf_plot, show_hhmmss),
                    y=df_e[p_col_e],
                    name="Pot. CNDC (EMF)",
                    line=dict(color=_gcfg["pot_color_sim0"], width=_gcfg["line_width"]),
                    yaxis="y2",
                ))

                # 3. Aplicar líneas de referencia segmentadas (t0 y t0+35s)
                # Con HH:MM:SS: t0 está en _scada_wall_t0 (segundos absolutos del reloj).
                # Sin HH:MM:SS: t0 está en 0.0 (datos ya alineados → falla = t=0).
                _t_ref_fault = _scada_wall_t0 if (show_hhmmss and _scada_wall_t0 > 0) else 0.0
                fig_c = add_reference_lines(
                    fig_c,
                    t_fault_abs=_t_ref_fault,
                    t_eval_abs=_t_ref_fault + 35.0,
                    show_hhmmss=show_hhmmss,
                    show_deadband=_gcfg["show_deadband"],
                    show_fault_line=True,
                    show_eval_line=True,
                    eval_line_label="t₀+35s (CNDC)"
                )

                # 4. Finalizar con el layout estándar de la aplicación
                fig_c = apply_standard_layout(
                    fig_c,
                    title=f"Comparativa Registro Real vs CNDC — {_sel_unit}",
                    xaxis_title="Segundos desde inicio falla (Alineado)",
                    yaxis_title="Frecuencia (Hz)",
                    yaxis2_title="Potencia (MW)",
                    template=_gcfg["template"],
                    height=_gcfg["plot_height"]
                )

                # Mini análisis de error
                with st.expander("Análisis de desviación"):
                    st.write("Diferencia promedio en Frecuencia: " +
                             f"{abs(df_s['Frecuencia_Hz'].mean() - df_e['frecuencia_hz'].mean()):.4f} Hz")
                    st.write(f"Diferencia promedio en Potencia: " +
                             f"{abs(df_s[p_col_s].mean() - df_e[p_col_e].mean()):.2f} MW") # type: ignore

                _t_comb_aligned = pd.concat([t_scada_aligned, t_emf_aligned]).dropna()
                _p_comb_max = max(df_s[p_col_s].max(), df_e[p_col_e].max()) * 1.1 if not _t_comb_aligned.empty else 200.0
                xaxis_min, xaxis_max, yaxis1_min, yaxis1_max, yaxis2_min, yaxis2_max, auto_scale_comp, _comp_chart_col = \
                    _render_axis_controls("comp", ev_path, _sel_unit,
                                          float(_t_comb_aligned.min()) if not _t_comb_aligned.empty else -10.0,
                                          float(_t_comb_aligned.max()) if not _t_comb_aligned.empty else 100.0,
                                          float(_p_comb_max),
                                          traces=[(t_scada_aligned.values, df_s[p_col_s].values),
                                                  (t_emf_aligned.values, df_e[p_col_e].values)])

                fig_c.update_layout(
                    xaxis=dict(range=[_to_plotly_time(xaxis_min, show_hhmmss), _to_plotly_time(xaxis_max, show_hhmmss)]),
                    yaxis=dict(range=[yaxis1_min, yaxis1_max]),
                    yaxis2=dict(range=[yaxis2_min, yaxis2_max]),
                )
                _comp_chart_col.plotly_chart(fig_c, use_container_width=True)

            else:
                _missing = []
                if not _has_s: _missing.append("SCADA")
                if not _has_e: _missing.append("EMF CNDC")
                st.warning(f"La unidad **{_sel_unit}** no tiene datos procesados en: {', '.join(_missing)}.")
        else:
            st.info("Asegúrese de haber ejecutado los procesadores en las pestañas anteriores.")

    #  Exportación masiva Bloque 3 
    st.markdown("---") # type: ignore
    _v4_section_head("Exportar todos los gráficos de Bloque 3", icon="download")
    st.caption("Genera capturas PNG de SCADA y EMF para todas las unidades disponibles.")

    if st.button("🗂️ Generar ZIP de gráficos registrados (SCADA/EMF)", key="btn_zip_b2"):
        import io, zipfile
        from plotly.io import to_image

        # Resetear descarga previa
        st.session_state.b3_plots_zip_bytes = None

        _zip_buf = io.BytesIO()
        _n_ok = 0
        _available = get_event_units(ev_path, n_evento)
        _pmax_map_exp = _load_pmax_cargado(ev_path, n_evento)
        _tmap_exp = _load_tech_map(LOC_NAMES_GEN_PATH)
        _prog = st.progress(0, text="Iniciando exportación masiva...")
        _gcfg = st.session_state.graph_config
        
        with zipfile.ZipFile(_zip_buf, 'w', zipfile.ZIP_DEFLATED) as _zf:
            for _idx, _uname in enumerate(_available):
                _prog.progress((_idx + 1) / len(_available), text=f"Procesando {_uname}...")
                
                # --- EXPORTAR SCADA ---
                _s_file = os.path.join(ev_path, "Graficas Registro 1SEG COBEE", f"{_uname}.xlsx")
                if not os.path.isfile(_s_file):
                    _s_file = _buscar_archivo_unidad(_uname, _listar_archivos_cache(os.path.join(ev_path, "Graficas Registro 1SEG COBEE"), "*.xlsx"))
                    if _s_file: _s_file = os.path.join(ev_path, "Graficas Registro 1SEG COBEE", _s_file)

                if os.path.isfile(_s_file):
                    try: # type: ignore
                        _df_s = pd.read_excel(_s_file, engine="calamine").dropna()
                        _tr_s = _parse_to_seconds(_df_s.iloc[:, 0])
                        _t_norm_s = _tr_s - _tr_s.min()
                        _fr_s = pd.to_numeric(_df_s['Frecuencia_Hz'], errors='coerce').ffill().values
                        _pt_s = pd.to_numeric(_df_s.iloc[:, 2], errors='coerce').ffill().values
                        
                        _t0_s = float(_get_unit_cfg(ev_path, _uname, "scada_t0_s", 0.0))
                        _idx_f = int(np.argmin(np.abs(_t_norm_s.values - _t0_s))) if _t0_s > 0 else _detectar_inicio_falla(_fr_s)
                        _t_f_abs = float(_t_norm_s.iloc[_idx_f])
                        
                        _pm_v, _tk, _ = _get_pmax_from_cargado(_uname, _pmax_map_exp, _tmap_exp)
                        _rp_v = _get_rp_default(_tk, LOC_NAMES_GEN_PATH) / 100.0
                        _dt_v = int(_get_unit_cfg(ev_path, _uname, "b2_sc_dt", 35))
                        _t_al_s = (_t_norm_s - _t_f_abs).values
                        _kpi = _cndc_kpis(_t_al_s, _fr_s, _pt_s, _pm_v, _rp_v, _dt_v)
                        _rocof_s = _calcular_rocof(_t_al_s, _fr_s, 3.0)

                        _y_auto = _get_unit_cfg(ev_path, _uname, "y_auto", True)
                        
                        _fig = create_dual_axis_timeseries(
                            t_data=_t_norm_s if not show_hhmmss else _tr_s,
                            freq_data=_fr_s, pot_data=_pt_s,
                            title=f"Registro SCADA - {_uname}",
                            show_hhmmss=show_hhmmss,
                            x_range=None if _y_auto else [_get_unit_cfg(ev_path, _uname, "b3_tab_scada_xmin", None), _get_unit_cfg(ev_path, _uname, "b3_tab_scada_xmax", None)],
                            y1_range=None if _y_auto else [_get_unit_cfg(ev_path, _uname, "y_f_min", None), _get_unit_cfg(ev_path, _uname, "y_f_max", None)],
                            y2_range=None if _y_auto else [_get_unit_cfg(ev_path, _uname, "y_p_min", None), _get_unit_cfg(ev_path, _uname, "y_p_max", None)],
                        )
                        _fig = add_reference_lines(_fig, t_fault_abs=_t_f_abs if not show_hhmmss else _tr_s.iloc[_idx_f],
                                                  t_eval_abs=(_t_f_abs + _dt_v) if not show_hhmmss else (_tr_s.iloc[_idx_f] + _dt_v),
                                                  show_hhmmss=show_hhmmss)
                        _fig = add_kpi_markers(
                            _fig,
                            t_fault_abs=_t_f_abs if not show_hhmmss else float(_tr_s.iloc[_idx_f]),
                            kpi_dict=_kpi,
                            show_hhmmss=show_hhmmss,
                            dt_seconds=_dt_v,
                        )




                        
                        _img = to_image(_fig, format="png", width=1200, height=600, scale=2)
                        _zf.writestr(f"SCADA_{_uname}_Ev{n_evento}.png", _img)
                        _n_ok += 1
                    except: pass

                # --- EXPORTAR EMF ---
                _e_file = os.path.join(ev_path, CARPETA_COBEE_EMF, f"{_uname}.xlsx")
                if not os.path.isfile(_e_file):
                    _e_file = _buscar_archivo_unidad(_uname, _listar_archivos_cache(os.path.join(ev_path, CARPETA_COBEE_EMF), "*.xlsx"))
                    if _e_file: _e_file = os.path.join(ev_path, CARPETA_COBEE_EMF, _e_file)

                if os.path.isfile(_e_file):
                    try: # type: ignore
                        _df_e = pd.read_excel(_e_file, engine="calamine").dropna()
                        _tr_e = _parse_to_seconds(_df_e['tiempo_s'])
                        _t_norm_e = _tr_e - _tr_e.min()
                        _fr_e = pd.to_numeric(_df_e['frecuencia_hz'], errors='coerce').ffill().values
                        _pcol = [c for c in _df_e.columns if c not in ['tiempo_s', 'frecuencia_hz', 'hora', 't_norm']][0]
                        _pt_e = pd.to_numeric(_df_e[_pcol], errors='coerce').ffill().values
                        
                        _t0_e = float(_get_unit_cfg(ev_path, _uname, "emf_t0_s", 0.0))
                        _idx_e = int(np.argmin(np.abs(_t_norm_e.values - _t0_e))) if _t0_e > 0 else _detectar_inicio_falla(_fr_e)
                        _t_fe_abs = float(_t_norm_e.iloc[_idx_e])
                        
                        _pm_v, _tk, _ = _get_pmax_from_cargado(_uname, _pmax_map_exp, _tmap_exp)
                        _rp_v = _get_rp_default(_tk, LOC_NAMES_GEN_PATH) / 100.0
                        _dt_v = int(_get_unit_cfg(ev_path, _uname, "b2_emf_dt", 35))
                        _t_al_e = (_t_norm_e - _t_fe_abs).values
                        _kpi_e = _cndc_kpis(_t_al_e, _fr_e, _pt_e, _pm_v, _rp_v, _dt_v)
                        _rocof_e = _calcular_rocof(_t_al_e, _fr_e, 3.0)

                        _y_auto_e = _get_unit_cfg(ev_path, _uname, "y_auto", True)
                        _xmin_e = _get_unit_cfg(ev_path, _uname, "b3_tab_emf_xmin", 0.0)
                        _xmax_e = _get_unit_cfg(ev_path, _uname, "b3_tab_emf_xmax", 100.0)
                        
                        _fig_e = create_dual_axis_timeseries(
                            t_data=_t_norm_e if not show_hhmmss else _tr_e,
                            freq_data=_fr_e, pot_data=_pt_e,
                            title=f"Gráfico CNDC (EMF) - {_uname}",
                            show_hhmmss=show_hhmmss,
                            freq_color="cyan",
                            pot_color=_gcfg["pot_color_sim0"],
                            x_range=None if _y_auto_e else [
                                _to_plotly_time((_tr_e.min() + _xmin_e) if show_hhmmss else _xmin_e, show_hhmmss),
                                _to_plotly_time((_tr_e.min() + _xmax_e) if show_hhmmss else _xmax_e, show_hhmmss)
                            ],
                            y1_range=None if _y_auto_e else [_get_unit_cfg(ev_path, _uname, "y_f_min", None), _get_unit_cfg(ev_path, _uname, "y_f_max", None)],
                            y2_range=None if _y_auto_e else [_get_unit_cfg(ev_path, _uname, "y_p_min", None), _get_unit_cfg(ev_path, _uname, "y_p_max", None)],
                        )
                        _t_falla_e_plot = float(_tr_e.iloc[_idx_e]) if show_hhmmss else _t_fe_abs
                        _fig_e = add_reference_lines(_fig_e, t_fault_abs=_t_falla_e_plot, 
                                                   t_eval_abs=(_t_falla_e_plot + _dt_v),
                                                   show_hhmmss=show_hhmmss, show_deadband=_gcfg["show_deadband"])
                        if _kpi_e:
                            _t0_e_p = _to_plotly_time(_t_falla_e_plot, show_hhmmss)
                            _tmin_e_p = _to_plotly_time(_t_falla_e_plot + float(_kpi_e['t_min']), show_hhmmss)
                            _tdt_e_p = _to_plotly_time(_t_falla_e_plot + _dt_v, show_hhmmss)
                            _fig_e = add_kpi_markers(_fig_e, t_fault_abs=_t_falla_e_plot,
                                                   kpi_dict=_kpi_e, show_hhmmss=show_hhmmss,
                                                   dt_seconds=_dt_v, marker_size=_gcfg["marker_size"],
                                                   freq_color="cyan",
                                                   pot_color=_gcfg["pot_color_sim0"],
                                                   t0_plot=_t0_e_p, tmin_plot=_tmin_e_p, tdt_plot=_tdt_e_p)
                        
                        _img_e = to_image(_fig_e, format="png", width=1200, height=600, scale=2)
                        _zf.writestr(f"EMF_{_uname}_Ev{n_evento}.png", _img_e)
                        _n_ok += 1
                    except: pass

        _prog.empty()
        if _n_ok > 0:
            st.session_state.b3_plots_zip_bytes = _zip_buf.getvalue()
            st.session_state.b3_plots_zip_name = f"graficos_registrados_Ev{n_evento}_{datetime.now().strftime('%H%M%S')}.zip"
            st.success(f"✅ Se generaron {_n_ok} gráficos con éxito.")
        else:
            st.error("No se pudieron generar imágenes. Verifique que existan archivos procesados.")

    if st.session_state.get("b3_plots_zip_bytes"):
        st.download_button(
            label=f"⬇️ Descargar ZIP de Gráficos Registrados",
            data=st.session_state.b3_plots_zip_bytes,
            file_name=st.session_state.b3_plots_zip_name,
            mime="application/zip",
            type="primary"
        )

elif bloque_trabajo == "analisis_simulacion":
    _render_block_header("04", "Análisis Simulación",
        "Lee resultados de escenarios RMS (E{N}.0 y E{N}.1) generados por PowerFactory y calcula KPIs CNDC.",
        "Análisis", pf_required=True)
    if not IS_CLOUD:
        st.info(
            "⚠️ Primero ejecute `DatosCurvas_v3.py` **dentro de PowerFactory** para generar los archivos Excel "
            "en las carpetas `E{N}.0/Datos Curvas/` y `E{N}.1/Datos Curvas/`."
        )

    if not st.session_state.semestre_global or not st.session_state.evento_global:
        st.warning("👈 Seleccione Semestre y Evento en la barra lateral.")
        st.stop()

    ev_path = st.session_state.ev_path_global
    n_evento = st.session_state.n_evento_global
    _sel_unit = st.session_state.global_selected_unit
    _event_cfg = _load_event_cfg(ev_path)

    _dir0_b3 = os.path.join(ev_path, f"E{n_evento}.0", CARPETA_DATOS_CURVAS)
    _dir1_b3 = os.path.join(ev_path, f"E{n_evento}.1", CARPETA_DATOS_CURVAS) # type: ignore

    # Resolver archivo de simulación basado en la unidad global seleccionada
    _sel_file_b3 = None
    if st.session_state.global_selected_unit:
        _target = st.session_state.global_selected_unit.replace("sym_", "")
        # Buscar en las carpetas de simulación
        for _d in [_dir0_b3, _dir1_b3]:
            if os.path.isdir(_d):
                for _f in os.listdir(_d):
                    if _target in _f and _f.endswith('.xlsx'):
                        _sel_file_b3 = _f
                        break # type: ignore
                if _sel_file_b3: break

    #  Auto-detección de t₀ desde el archivo de simulación 
    # El event_id NO incluye el archivo de simulación para que cambiar de unidad
    # no resetee t₀ si ya fue guardado para este evento.
    _b3_event_id = f"{st.session_state.semestre_global}|{st.session_state.evento_global}"
    _t0_saved = _event_cfg.get("t_sim_falla")   # None si nunca se guardó
    _t0_autodet = float(_t0_saved) if _t0_saved is not None else 5.0
    if _sel_file_b3 and os.path.isdir(_dir0_b3):
        _autodet_path = os.path.join(_dir0_b3, _sel_file_b3)
        if os.path.isfile(_autodet_path):
            try:
                _df_ad = pd.read_excel(_autodet_path, engine="calamine").dropna()
                _t_ad = pd.to_numeric(_df_ad.iloc[:, 0], errors='coerce').values
                _fc_ad = [c for c in _df_ad.columns[1:] if _is_frequency_column(c, _df_ad[c])]
                if _fc_ad:
                    _fq_ad = pd.to_numeric(_df_ad[_fc_ad[0]], errors='coerce').ffill().bfill().values
                    _fq_ad = _fq_ad * 50.0 if np.nanmax(_fq_ad) < 2.0 else _fq_ad
                    # Filtrar filas con tiempo NaN antes de detectar inicio
                    _valid_ad = ~np.isnan(_t_ad)
                    _t_ad_v  = _t_ad[_valid_ad]
                    _fq_ad_v = _fq_ad[_valid_ad]
                    _idx_ad = _detectar_inicio_falla(_fq_ad_v)
                    if _idx_ad > 0 and _idx_ad < len(_t_ad_v):
                        # Solo sobreescribir el auto-detectado si NO hay valor guardado
                        if _t0_saved is None:
                            _t0_autodet = float(_t_ad_v[_idx_ad])
            except Exception:
                pass
    _t_falla_sim_val = st.session_state.get("b3_t_falla", _t0_autodet)


    #  Indicador de contexto compacto 
    if _sel_unit:
        _u_clean_b4 = _sel_unit.replace("sym_", "")
        _pm_b4 = _load_pmax_cargado(ev_path, n_evento)
        _tm_b4 = _load_tech_map(LOC_NAMES_GEN_PATH)
        _pmax_b4, _, _ = _get_pmax_from_cargado(_sel_unit, _pm_b4, _tm_b4)
    else:
        st.warning("⬆️ Seleccione una unidad en el selector superior para ver el análisis.")
        st.stop()

    #  Parámetros compartidos entre tabs 
    _v4_section_head("⚙️ Parámetros de Análisis CNDC",
        f"t₀ auto-detectado: {_t0_autodet:.1f} s  ·  Ajuste si difiere del evento real.",
        "sliders")
    _bp1, _bp2 = st.columns(2)
    # Inicializar session_state antes del widget para evitar conflicto value= vs key=
    if "b3_t_falla" not in st.session_state:
        st.session_state["b3_t_falla"] = float(_t_falla_sim_val)
    if "b3_dt" not in st.session_state:
        st.session_state["b3_dt"] = int(_event_cfg.get("delta_t_cndc", 35))
    _b3_t_falla = _bp1.number_input(
        "t₀ falla sim. [s]",
        min_value=0.0, max_value=300.0, step=0.5,
        help=f"Instante t₀ del evento (auto-detectado: {_t0_autodet:.1f} s). Ajuste si es necesario.",
        key="b3_t_falla",
    )
    _b3_dt = int(_bp2.number_input( # type: ignore
        "Δt CNDC [s]",
        min_value=20, max_value=60, step=1,
        help="Tiempo desde t₀ para evaluar f_Δt y P_Δt. CNDC usa 30–50 s (típicamente 35 s).",
        key="b3_dt",
    ))
    _bsave1, _bsave2 = st.columns(2)
    if _bsave1.button("💾 Guardar t₀ y Δt", key="save_b3_params",
                      help="Guarda t₀ y Δt en la configuración del evento para próximas sesiones."):
        _save_event_cfg(ev_path, "t_sim_falla", _b3_t_falla)
        _save_event_cfg(ev_path, "delta_t_cndc", _b3_dt)
        st.toast(f"Guardado t₀={_b3_t_falla:.1f} s, Δt={_b3_dt} s", icon="✅")
    if _bsave2.button("↩ Usar t₀ auto-detectado", key="reset_b3_t0",
                      help=f"Restaurar t₀ al valor auto-detectado ({_t0_autodet:.1f} s)."):
        st.session_state.b3_t_falla = _t0_autodet
        st.rerun()

    #  Tab bar persistente 
    _b4_tab = _v4_tab_bar([
        {"id": "cndc",  "icon": "database", "label": f"Simulación E{n_evento}.0 (CNDC)"},
        {"id": "cobee", "icon": "chart",    "label": f"Simulación E{n_evento}.1 (COBEE)"},
        {"id": "comp",  "icon": "scale",    "label": "Comparativa de Simulaciones"},
    ], "b04")

    def load_and_display_simulation_data(sim_type_suffix, sel_file):
        """Loads simulation data and displays basic info and dataframe."""
        sim_dir = os.path.join(st.session_state.ev_path_global, f"E{n_evento}.{sim_type_suffix}", CARPETA_DATOS_CURVAS) # type: ignore

        if not sel_file:
            st.info("ℹ️ Seleccione una unidad en el selector superior.")
            return None, None

        if not os.path.isdir(sim_dir):
            st.info(f"ℹ️ La carpeta '{sim_dir}' no existe. Asegúrese de haber ejecutado `DatosCurvas_v3.py` en PowerFactory.")
            return None, None

        xlsx_files = sorted([f for f in os.listdir(sim_dir) if f.endswith('.xlsx') and not f.startswith('~$')])

        if not xlsx_files:
            st.warning(f"No se encontraron archivos Excel en '{sim_dir}'.")
            return None, None

        if sel_file not in xlsx_files: # type: ignore
            st.warning(f"El archivo **{sel_file}** no existe en E{n_evento}.{sim_type_suffix}.")
            return None, None

        st.success(f"✅ {len(xlsx_files)} archivos disponibles — mostrando: **{sel_file}**")
        df_sim = pd.read_excel(os.path.join(sim_dir, sel_file), engine="calamine").dropna() # type: ignore

        time_col = df_sim.columns[0]
        data_cols = [col for col in df_sim.columns if col != time_col]

        with st.expander("📄 Ver tabla de datos"):
            st.dataframe(df_sim, use_container_width=False)
        return df_sim, sel_file

    def _load_sim_tab_data(sim_dir, sel_file):
        """Carga y normaliza un archivo de simulación. Devuelve (ts_aligned, fs_hz, ps_mw, df) o None.
        Delega a _cached_sim_arrays para evitar re-leer el mismo fichero en B4/B5."""
        if not sel_file or not os.path.isdir(sim_dir):
            return None
        fpath = os.path.join(sim_dir, sel_file)
        if not os.path.isfile(fpath):
            return None
        return _cached_sim_arrays(fpath, _b3_t_falla)  # (ts_aligned, fs_hz, ps_mw, df)

    def _sim_kpis_and_pmax(sel_unit, ev_path, n_evento, rp_fallback=0.05, pmax_fallback=200.0):
        """Recupera Pmax y Rp para el análisis CNDC de simulaciones."""
        _pm = _load_pmax_cargado(ev_path, n_evento)
        _tm = _load_tech_map(LOC_NAMES_GEN_PATH)
        if sel_unit:
            pm_v, tk, _ = _get_pmax_from_cargado(sel_unit, _pm, _tm)
            rp = float(_get_rp_default(tk, LOC_NAMES_GEN_PATH)) / 100.0
        else:
            pm_v, rp = pmax_fallback, rp_fallback
        return float(pm_v), rp

    def _render_sim_tab(sim_ver, sim_dir, sim_color_f, sim_color_p, pfx):
        """Renderiza una pestaña de simulación completa (gráfico + escalado + KPI + descarga)."""
        _gcfg = st.session_state.graph_config
        if not os.path.isdir(sim_dir):
            st.info(f"ℹ️ La carpeta `{sim_dir}` no existe. Ejecute `DatosCurvas_v3.py` en PowerFactory.")
            return None

        xlsx_files = sorted([f for f in os.listdir(sim_dir) if f.endswith('.xlsx') and not f.startswith('~$')])
        if not xlsx_files:
            st.warning(f"No se encontraron archivos Excel en `{sim_ver}/Datos Curvas`.")
            return None

        st.success(f"✅ {len(xlsx_files)} archivos disponibles.")
        if not _sel_file_b3:
            st.info("ℹ️ Seleccione una unidad en la barra lateral.")
            return None
        if _sel_file_b3 not in xlsx_files:
            st.warning(f"El archivo **{_sel_file_b3}** no está en `{sim_ver}`. Ejecute `DatosCurvas_v3.py`.")
            return None

        result = _load_sim_tab_data(sim_dir, _sel_file_b3)
        if result is None:
            st.error(f"No se pudo cargar `{_sel_file_b3}` desde `{sim_ver}`.")
            return None
        ts_aligned, fs_hz, ps_mw, df_raw = result

        pm_v, rp_v = _sim_kpis_and_pmax(_sel_unit, ev_path, n_evento)
        _kpi = _cndc_kpis(ts_aligned, fs_hz, ps_mw, pm_v, rp_v, _b3_dt)
        _rocof = _calcular_rocof(ts_aligned, fs_hz, 3.0)

        _xdef_min = float(ts_aligned.min()) if len(ts_aligned) else 0.0
        _xdef_max = float(ts_aligned.max()) if len(ts_aligned) else 100.0
        _xmin_cfg = f"{pfx}_xmin"
        _xmax_cfg = f"{pfx}_xmax"

        _xmin_w, _xmax_w, _y1min_w, _y1max_w, _y2min_w, _y2max_w, auto_s, _sim_chart_col = \
            _render_axis_controls(pfx, ev_path, _sel_unit or "",
                                  _xdef_min, _xdef_max, float(pm_v * 1.1),
                                  traces=[(ts_aligned, ps_mw)])

        fig_s = create_dual_axis_timeseries(
            t_data=ts_aligned, freq_data=fs_hz, pot_data=ps_mw,
            title=f"Simulación {sim_ver} — {_sel_file_b3}",
            freq_label=f"Frecuencia {sim_ver} (Hz)",
            pot_label=f"Potencia {sim_ver} (MW)",
            freq_color=sim_color_f, pot_color=sim_color_p,
            line_width=_gcfg["line_width"], template=_gcfg["template"],
            height=_gcfg["plot_height"], legend_position="bottom_center",
            x_range=[_xmin_w, _xmax_w],
            y1_range=[_y1min_w, _y1max_w],
            y2_range=[_y2min_w, _y2max_w],
        )
        # P_máxima en ventana [t_nadir, t₀+Δt]
        # Si el nadir cae fuera del período de evaluación (t_nadir > _b3_dt),
        # se busca desde t=0 para no obtener una ventana vacía.
        _t_nadir_s3 = float(_kpi['t_min']) if _kpi else 0.0
        _t_pmax_start = _t_nadir_s3 if _t_nadir_s3 < float(_b3_dt) else 0.0
        _t_pmax_s3, _p_pmax_s3 = _find_pmax_time(ts_aligned, ps_mw, _b3_dt, t_min_eval=_t_pmax_start)
        _f_pmax_s3 = None
        if _t_pmax_s3 is not None:
            _idx_pm_s3 = int(np.argmin(np.abs(ts_aligned - _t_pmax_s3)))
            _f_pmax_s3 = float(fs_hz[_idx_pm_s3])

        if _kpi:
            fig_s = add_reference_lines(
                fig_s, t_fault_abs=0.0, t_eval_abs=_b3_dt, show_hhmmss=False,
                show_deadband=_gcfg["show_deadband"],
                eval_line_label=f"t₀+Δt ({_b3_dt} s)",
            )
            fig_s = add_kpi_markers(
                fig_s, t_fault_abs=0.0, kpi_dict=_kpi, show_hhmmss=False,
                dt_seconds=_b3_dt, marker_size=_gcfg["marker_size"],
                freq_color=sim_color_f, pot_color=sim_color_p,
            )
        else:
            fig_s = add_reference_lines(fig_s, show_hhmmss=False, show_deadband=_gcfg["show_deadband"],
                                         show_fault_line=False, show_eval_line=False)
        if _t_pmax_s3 is not None and _gcfg.get("show_pmax_marker", True):
            fig_s = add_pmax_marker(
                fig_s, _t_pmax_s3, _p_pmax_s3, _f_pmax_s3,
                pot_color=sim_color_p, freq_color=sim_color_f,
                marker_size=_gcfg["marker_size"],
            )

        _sim_chart_col.plotly_chart(fig_s, use_container_width=True)

        with st.expander("📄 Ver tabla de datos"):
            st.dataframe(_df_safe(df_raw), use_container_width=False)
            if st.button("⬇️ Descargar datos a Excel", key=f"dl_{pfx}_data"):
                _unit_label = (_sel_unit or "unidad").replace("sym_", "")
                st.download_button(
                    f"Descargar {sim_ver}",
                    _apply_excel_formatting(df_raw, sheet_name=f"{sim_ver}_{_unit_label}"),
                    file_name=f"sim_{sim_ver.replace('.', '_')}_{_unit_label}_Ev{n_evento}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_{pfx}_btn",
                )

        st.markdown("---")
        _v4_section_head(f"KPIs CNDC — {sim_ver}", "Resultado simulación PowerFactory", icon="chart")
        _kpi_pmax_s3 = None
        if _t_pmax_s3 is not None:
            _kpi_pmax_s3 = _cndc_kpis(ts_aligned, fs_hz, ps_mw, pm_v, rp_v, _t_pmax_s3)
        _mostrar_tabla_cndc_duo(
            _kpi, pm_v, _b3_dt, f"Simulación {sim_ver}", rocof=_rocof,
            kpi2=_kpi_pmax_s3, delta_t2=_t_pmax_s3, fuente2=f"P_máxima {sim_ver}",
        )
        return _kpi, ts_aligned, fs_hz, ps_mw, pm_v, rp_v

    # Pestaña 1: Simulación E{N}.0 (CNDC)
    if _b4_tab == "cndc":
        _v4_section_head(f"Simulación E{n_evento}.0 — Escenario CNDC", icon="database")
        _gcfg = st.session_state.graph_config
        _r0 = _render_sim_tab(
            sim_ver=f"E{n_evento}.0",
            sim_dir=_dir0_b3,
            sim_color_f=_gcfg["freq_color_sim0"],
            sim_color_p=_gcfg["pot_color_sim0"],
            pfx="b3_sim0",
        )

    elif _b4_tab == "cobee":
        _v4_section_head(f"Simulación E{n_evento}.1 — Escenario COBEE", icon="chart")
        _gcfg = st.session_state.graph_config
        _r1 = _render_sim_tab(
            sim_ver=f"E{n_evento}.1",
            sim_dir=_dir1_b3,
            sim_color_f=_gcfg["freq_color_sim1"],
            sim_color_p=_gcfg["pot_color_sim1"],
            pfx="b3_sim1",
        )

    elif _b4_tab == "comp":
        _v4_section_head(f"Comparativa E{n_evento}.0 vs E{n_evento}.1", icon="scale")
        _gcfg = st.session_state.graph_config
        ok0, ok1 = os.path.isdir(_dir0_b3), os.path.isdir(_dir1_b3)
        if not ok0 or not ok1:
            missing = []
            if not ok0: missing.append(f"`E{n_evento}.0/Datos Curvas`")
            if not ok1: missing.append(f"`E{n_evento}.1/Datos Curvas`")
            st.warning(f"Faltan carpetas: {', '.join(missing)}. Ejecute `DatosCurvas_v3.py` en PowerFactory.")
        elif not _sel_file_b3:
            st.info("ℹ️ Seleccione una unidad en la barra lateral.")
        else:
            _d0 = _load_sim_tab_data(_dir0_b3, _sel_file_b3)
            _d1 = _load_sim_tab_data(_dir1_b3, _sel_file_b3)
            _any = _d0 or _d1
            if not _any:
                st.warning(f"No se encontró `{_sel_file_b3}` en ninguna carpeta de simulación.")
            else:
                pm_vc, rp_vc = _sim_kpis_and_pmax(_sel_unit, ev_path, n_evento)

                # Extraer datos antes de controles (para auto-P con ambas trazas)
                _base = _d0 or _d1
                _base_ver = f"E{n_evento}.0" if _d0 else f"E{n_evento}.1"
                _base_cf = _gcfg["freq_color_sim0"] if _d0 else _gcfg["freq_color_sim1"]
                _base_cp = _gcfg["pot_color_sim0"] if _d0 else _gcfg["pot_color_sim1"]
                _ts_b, _fs_b, _ps_b, _ = _base
                _simc_traces = [(_ts_b, _ps_b)]
                if _d0 and _d1:
                    _ts_ov, _fs_ov, _ps_ov, _ = _d1
                    _simc_traces.append((_ts_ov, _ps_ov))

                _sc_xmin, _sc_xmax, _sc_y1min, _sc_y1max, _sc_y2min, _sc_y2max, auto_sc, _simc_chart_col = \
                    _render_axis_controls("simc", ev_path, _sel_unit or "",
                                          -10.0, 100.0, float(pm_vc * 1.1),
                                          traces=_simc_traces)

                fig_sc = create_dual_axis_timeseries(
                    t_data=_ts_b, freq_data=_fs_b, pot_data=_ps_b,
                    title=f"Comparativa Simulaciones — {_sel_file_b3}",
                    freq_label=f"Frecuencia {_base_ver} (Hz)",
                    pot_label=f"Potencia {_base_ver} (MW)",
                    freq_color=_base_cf, pot_color=_base_cp,
                    line_width=_gcfg["line_width"], template=_gcfg["template"],
                    height=_gcfg["plot_height"], legend_position="bottom_center",
                    x_range=[_sc_xmin, _sc_xmax],
                    y1_range=[_sc_y1min, _sc_y1max],
                    y2_range=[_sc_y2min, _sc_y2max],
                )

                # Overlay de la otra simulación
                if _d0 and _d1:
                    fig_sc.add_trace(go.Scatter(
                        x=_ts_ov, y=_fs_ov, name=f"Frecuencia E{n_evento}.1 (Hz)",
                        line=dict(color=_gcfg["freq_color_sim1"], dash="dash", width=_gcfg["line_width"]), yaxis="y",
                    ))
                    fig_sc.add_trace(go.Scatter(
                        x=_ts_ov, y=_ps_ov, name=f"Potencia E{n_evento}.1 (MW)",
                        line=dict(color=_gcfg["pot_color_sim1"], dash="dash", width=_gcfg["line_width"]), yaxis="y2",
                    ))

                fig_sc = add_reference_lines(
                    fig_sc, t_fault_abs=0.0, t_eval_abs=_b3_dt, show_hhmmss=False,
                    show_deadband=_gcfg["show_deadband"], eval_line_label=f"t₀+Δt ({_b3_dt} s)",
                )
                _simc_chart_col.plotly_chart(fig_sc, use_container_width=True)

                # Tabla comparativa de KPIs
                _kpi_rows_sc = []
                if _d0:
                    _ts0, _fs0, _ps0, _ = _d0
                    _k0 = _cndc_kpis(_ts0, _fs0, _ps0, pm_vc, rp_vc, _b3_dt)
                    if _k0: _kpi_rows_sc.append({"Fuente": f"E{n_evento}.0 (CNDC)", **_k0})
                if _d1:
                    _ts1, _fs1, _ps1, _ = _d1
                    _k1 = _cndc_kpis(_ts1, _fs1, _ps1, pm_vc, rp_vc, _b3_dt)
                    if _k1: _kpi_rows_sc.append({"Fuente": f"E{n_evento}.1 (COBEE)", **_k1})
                if _kpi_rows_sc:
                    _v4_section_head("Comparativa de KPIs", icon="chart")
                    _df_ksc = _df_safe(pd.DataFrame(_kpi_rows_sc))
                    for _c in _df_ksc.columns:
                        if _df_ksc[_c].dtype == object:
                            _df_ksc[_c] = _df_ksc[_c].astype(str)
                    st.dataframe(_df_ksc, hide_index=True, use_container_width=False)


elif bloque_trabajo == "comparativa_real_simu":
    _render_block_header("05", "Real vs Simulación",
        "Compara registros SCADA con simulaciones RMS alineadas en t₀ y evalúa cumplimiento contra criterios CNDC.",
        "Análisis", pf_required=True)

    if not st.session_state.semestre_global or not st.session_state.evento_global:
        st.warning("👈 Seleccione Semestre y Evento en la barra lateral.")
        st.stop()

    ev_path = st.session_state.ev_path_global
    n_evento = st.session_state.n_evento_global

    _event_cfg = _load_event_cfg(ev_path)

    _sel_unit_b5 = st.session_state.global_selected_unit
    if not _sel_unit_b5:
        st.warning("⬆️ Seleccione una unidad en el selector superior para ver el análisis.")
        st.stop()

    #  Funciones auxiliares 
    def _local_parse_sec(series):
        def _to_sec(val):
            if pd.isna(val): return 0.0
            s = str(val).strip()
            if ':' in s:
                parts = s.split(':')
                try: return int(parts[0]) * 3600 + int(parts[1]) * 60 + (int(parts[2]) if len(parts) > 2 else 0)
                except: pass
            try: return float(s.replace(',', '.'))
            except: return 0.0
        return series.apply(_to_sec)

    #  Panel de configuración — mismo patrón que B04 
    # Orden: t₀ (col1) · Δt (col2) · Fuente (col3) · Simulaciones (col4) · Guardar (col5)
    _v4_section_head("⚙️ Parámetros de Análisis CNDC", icon="sliders")
    _event_cfg = _load_event_cfg(ev_path)
    _pc1, _pc2, _pc3, _pc4, _pc5 = st.columns([1, 1, 1.2, 1.6, 0.9])
    with _pc1:
        t_sim_falla = st.number_input(
            "t₀ falla sim. [s]",
            value=_event_cfg.get("t_sim_falla", 5.0), min_value=0.0, max_value=300.0, step=0.5,
            help="Instante t en la simulación RMS donde ocurre el evento (PowerFactory).",
            key="b4_t_sim_falla",
        )
    with _pc2:
        delta_t_cndc = st.number_input(
            "Δt CNDC [s]",
            value=35, min_value=20, max_value=60, step=1,
            help="Tiempo desde t₀ para leer f_Δt y P_Δt. CNDC usa entre 30–50 s (típicamente 35 s).",
            key="b4_delta_t_cndc",
        )
    with _pc3:
        src_real = st.radio("Fuente Real:", ["SCADA COBEE (1SEG)", "EMF CNDC"], key="b4_src_real")
    with _pc4:
        src_sim = st.multiselect(
            "Simulaciones:",
            [f"E{n_evento}.0", f"E{n_evento}.1"],
            default=[f"E{n_evento}.0", f"E{n_evento}.1"],
        )
    with _pc5:
        st.markdown('<div style="height:20px"></div>', unsafe_allow_html=True)
        if st.button("💾 Guardar t₀", key="save_t_sim_falla", help="Guardar tiempo de falla en config del evento"):
            if _save_event_cfg(ev_path, "t_sim_falla", t_sim_falla):
                st.toast("t₀ guardado.")
        try:
            df_tech = pd.read_excel(LOC_NAMES_GEN_PATH, sheet_name="Detalle_PF", engine="calamine")
            _pcol = 'P_max (MW)' if 'P_max (MW)' in df_tech.columns else 'P nom. (MW)'
            tech_map = (df_tech.set_index('loc_name PF')[[_pcol]]
                        .rename(columns={_pcol: 'P_max (MW)'})
                        .to_dict('index'))
            st.caption(f"✅ P_max: {len(tech_map)} uds.")
        except Exception as _e:
            tech_map = {}
            st.caption(f"⚠️ Sin P_max")

    # Parámetros avanzados de detección (colapsados — opcionales)
    with st.expander("⚙ Detección avanzada de t₀", expanded=False):
        _da1, _da2, _da3 = st.columns(3)
        with _da1:
            umbral_dfdt = st.number_input(
                "Umbral df/dt (Hz/s)",
                value=-0.04, min_value=-2.0, max_value=-0.001, step=0.005, format="%.3f",
                help="Caída sostenida de frecuencia (suavizada) para detectar inicio de falla.",
                key="b4_umbral_dfdt",
            )
        with _da2:
            ventana_suavizado = st.number_input(
                "Ventana suavizado (muestras)",
                value=5, min_value=2, max_value=20, step=1,
                help="Número de muestras para rolling mean antes de calcular df/dt.",
                key="b4_ventana_suav",
            )
        with _da3:
            ventana_pre = st.number_input(
                "Ventana pre-falla (s)",
                value=10, min_value=0, max_value=120, step=5,
                help="Segundos antes de t=0 a incluir en la gráfica.",
                key="b4_ventana_pre",
            )

    _b5_sel_unit = st.session_state.get("global_selected_unit", "")
    _xmin_v, _xmax_v, _y1min_v, _y1max_v, _y2min_v, _y2max_v, auto_v, _b5_chart_col = \
        _render_axis_controls("b5", ev_path, _b5_sel_unit, -10.0, 100.0, 200.0)

    #  Carga y Alineación Real 
    _sel_unit = st.session_state.global_selected_unit
    real_subdir = "Graficas Registro 1SEG COBEE" if "SCADA" in src_real else "Resultados_COBEE"
    _r_dir = os.path.join(ev_path, real_subdir)
    
    if _sel_unit and os.path.isdir(_r_dir):
        _rf_match = _buscar_archivo_unidad(_sel_unit, os.listdir(_r_dir))
        if _rf_match:
            # _cached_real_arrays: caché compartido con B3 → hit si ya fue leído
            _r_fpath = os.path.join(_r_dir, _rf_match)
            tr_aligned, _fr_arr, _pr_arr, idx_f_r, _t_f_auto = _cached_real_arrays(
                _r_fpath, umbral_dfdt, int(ventana_suavizado)
            )

            # Aplicar t₀ guardado en B3 (si el usuario lo ajustó manualmente)
            _src_key = "scada_t0_s" if "SCADA" in src_real else "emf_t0_s"
            _t0_manual = _event_cfg.get(_src_key)
            if _t0_manual is not None:
                # tr_aligned = tr_norm - t_f_auto → recentrar en el t₀ manual
                tr_aligned = tr_aligned + (_t_f_auto - float(_t0_manual))

            # Auto-P para B5: los widgets están deshabilitados en modo auto,
            # así que solo actualizamos las variables locales para el gráfico.
            if auto_v:
                _ap_b5 = _compute_auto_p_range([(tr_aligned, _pr_arr)], _xmin_v, _xmax_v)
                if _ap_b5:
                    _y2min_v, _y2max_v = _ap_b5[0], _ap_b5[1]

            # --- Construcción del Gráfico de Validación ---
            _gcfg = st.session_state.graph_config
            fig = create_dual_axis_timeseries(
                t_data=tr_aligned, freq_data=_fr_arr, pot_data=_pr_arr,
                title=f"Validación Real vs Simulación — {_sel_unit}",
                freq_label=f"Frec. Real ({src_real})", pot_label=f"Pot. Real ({src_real})",
                freq_color=_gcfg["freq_color_real"], pot_color=_gcfg["pot_color_real"],
                show_hhmmss=False,
                x_range=[_xmin_v, _xmax_v],
                y1_range=[_y1min_v, _y1max_v],
                y2_range=[_y2min_v, _y2max_v],
            )
            
            # Marcadores y KPIs — Datos Reales
            _pmax_map_v = _load_pmax_cargado(ev_path, n_evento)
            _pm_v, _tk, _pm_fuente = _get_pmax_from_cargado(_sel_unit, _pmax_map_v, _load_tech_map(LOC_NAMES_GEN_PATH))
            _rp_v = _get_rp_default(_tk, LOC_NAMES_GEN_PATH) / 100.0
            _kr    = _cndc_kpis(tr_aligned, _fr_arr, _pr_arr, _pm_v, _rp_v, delta_t_cndc)
            _rocof_r = _calcular_rocof(tr_aligned, _fr_arr, 3.0)

            if _kr:
                fig = add_kpi_markers(
                    fig, t_fault_abs=0.0, kpi_dict=_kr, show_hhmmss=False,
                    dt_seconds=delta_t_cndc,
                    freq_color=_gcfg["freq_color_real"], pot_color=_gcfg["pot_color_real"],
                )
            fig = add_reference_lines(
                fig, t_fault_abs=0.0, t_eval_abs=delta_t_cndc, show_hhmmss=False,
                show_deadband=_gcfg["show_deadband"],
                eval_line_label=f"t₀+Δt ({delta_t_cndc} s)",
            )
            # P_máxima — datos reales en [t_nadir, t₀+Δt]
            _t_nadir_r5 = float(_kr['t_min']) if _kr else 0.0
            _t_pmax_r5, _p_pmax_r5 = _find_pmax_time(
                tr_aligned, _pr_arr, delta_t_cndc, t_min_eval=_t_nadir_r5
            )
            _kpi_pmax_per_src5 = {}
            if _t_pmax_r5 is not None:
                _idx_pm_r5  = int(np.argmin(np.abs(tr_aligned - _t_pmax_r5)))
                _f_pmax_r5  = float(_fr_arr[_idx_pm_r5])
                if _gcfg.get("show_pmax_marker", True):
                    fig = add_pmax_marker(
                        fig, _t_pmax_r5, _p_pmax_r5, _f_pmax_r5,
                        pot_color=_gcfg["pot_color_real"], freq_color=_gcfg["freq_color_real"],
                        marker_size=_gcfg["marker_size"],
                    )
                if _kr:
                    _kpi_pm_r5 = _cndc_kpis(tr_aligned, _fr_arr, _pr_arr, float(_pm_v), _rp_v, _t_pmax_r5)
                    _kpi_pmax_per_src5[f"Real ({src_real})"] = (_kpi_pm_r5, float(_pm_v), _t_pmax_r5, _p_pmax_r5)

            #  Carga y Alineación Simulaciones
            _kpi_rows   = [{"Fuente": f"REAL ({src_real})", **_kr}] if _kr else []
            _rocof_rows = {f"REAL ({src_real})": _rocof_r}
            _sim_for_error = []
            _kpi_per_src  = {}   # fuente → (kpi_dict, rocof, pm_v)

            for s_ver in src_sim:
                _s_dir = os.path.join(ev_path, s_ver, CARPETA_DATOS_CURVAS)
                if not os.path.isdir(_s_dir):
                    continue
                _sf_match = _buscar_archivo_unidad(_sel_unit, os.listdir(_s_dir))
                if not _sf_match:
                    continue

                # _cached_sim_arrays: mismo caché que B3/B4 → hit si ya fue leído
                _s_cached = _cached_sim_arrays(os.path.join(_s_dir, _sf_match), t_sim_falla)
                ts_al, fs_hz, ps_mw, _ = _s_cached

                _color_f = _gcfg["freq_color_sim0"] if "0" in s_ver else _gcfg["freq_color_sim1"]
                _color_p = _gcfg["pot_color_sim0"]  if "0" in s_ver else _gcfg["pot_color_sim1"]

                fig.add_trace(go.Scatter(
                    x=ts_al, y=fs_hz, name=f"Frec. {s_ver}",
                    line=dict(color=_color_f, dash="dash", width=_gcfg["line_width"]), yaxis="y",
                ))
                fig.add_trace(go.Scatter(
                    x=ts_al, y=ps_mw, name=f"Pot. {s_ver}",
                    line=dict(color=_color_p, dash="dash", width=_gcfg["line_width"]), yaxis="y2",
                ))

                _ks    = _cndc_kpis(ts_al, fs_hz, ps_mw, _pm_v, _rp_v, delta_t_cndc)
                _roc_s = _calcular_rocof(ts_al, fs_hz, 3.0)
                if _ks:
                    _kpi_rows.append({"Fuente": s_ver, **_ks})
                    _kpi_per_src[s_ver] = (_ks, _roc_s, float(_pm_v))
                _rocof_rows[s_ver] = _roc_s
                _sim_for_error.append({"ver": s_ver, "t": ts_al, "f": fs_hz, "color": _color_f})

                # P_máxima por fuente de simulación en [t_nadir, t₀+Δt]
                _t_nadir_s5 = float(_ks['t_min']) if _ks else 0.0
                _t_pmax_s5, _p_pmax_s5 = _find_pmax_time(
                    ts_al, ps_mw, delta_t_cndc, t_min_eval=_t_nadir_s5
                )
                if _t_pmax_s5 is not None:
                    _idx_pm_s5 = int(np.argmin(np.abs(ts_al - _t_pmax_s5)))
                    _f_pmax_s5 = float(fs_hz[_idx_pm_s5])
                    if _gcfg.get("show_pmax_marker", True):
                        fig = add_pmax_marker(
                            fig, _t_pmax_s5, _p_pmax_s5, _f_pmax_s5,
                            pot_color=_color_p, freq_color=_color_f,
                            marker_size=_gcfg["marker_size"],
                        )
                    if _ks:
                        _kpi_pm_s5 = _cndc_kpis(ts_al, fs_hz, ps_mw, float(_pm_v), _rp_v, _t_pmax_s5)
                        _kpi_pmax_per_src5[f"Sim {s_ver}"] = (_kpi_pm_s5, float(_pm_v), _t_pmax_s5, _p_pmax_s5)

            _b5_chart_col.plotly_chart(fig, use_container_width=True)

            #  Tabla comparativa KPIs (todas las fuentes en una sola vista) 
            if _pm_fuente:
                st.caption(f"✅ P_max desde `{_pm_fuente}` → **{float(_pm_v):.2f} MW** | Rp = {_rp_v*100:.1f}%")
            else:
                st.warning(f"⚠️ No se encontró Pmax para **{_sel_unit}** en datos_cargados ni loc_names_gen.")

            st.markdown("---")
            _v4_section_head("KPIs CNDC — Comparativa Real vs. Simulación", icon="chart")

            # Definición de filas: (etiqueta, función(kpi, p_max, rocof) → str)
            _dt = delta_t_cndc  # captura local para lambdas
            _KPI_FILAS_B5 = [
                ("P_max [MW]",                   lambda k, pm, roc: f"{pm:.2f}"),
                ("f₀ — Inicio evento [Hz]",      lambda k, pm, roc: f"{k['f0']:.4f}"),
                ("P₀ — Inicio evento [MW]",      lambda k, pm, roc: f"{k['p0']:.3f}"),
                ("f_min — Nadir [Hz]",           lambda k, pm, roc: f"{k['f_min']:.4f}"),
                ("t_min — Nadir [s]",            lambda k, pm, roc: f"{k['t_min']:.1f}"),
                ("Δf = f₀ − f_min [Hz]",        lambda k, pm, roc: f"{k['delta_f']:.4f}"),
                (f"f_Δt ({_dt}s) [Hz]",         lambda k, pm, roc: f"{k['f_dt']:.4f}"),
                (f"P_Δt ({_dt}s) [MW]",         lambda k, pm, roc: f"{k['p_dt']:.3f}"),
                ("R_inic [MW]",                  lambda k, pm, roc: f"{k['r_inic']:.3f}"),
                ("R_inic [%]",                   lambda k, pm, roc: f"{k['r_inic_pct']:.2f}"),
                ("ΔP entregada [MW]",            lambda k, pm, roc: f"{k['dp']:.3f}"),
                ("ΔP% aporte [%]",               lambda k, pm, roc: f"{k['dp_pct']:.2f}"),
                ("¿Aporta RPF? (ΔP% ≥ 1.5%)",   lambda k, pm, roc: "✅ Sí" if k['aporta'] else "❌ No"),
                ("Droop Nominal [%]",            lambda k, pm, roc: f"{k['droop_nom']:.1f}"),
                ("Droop Calculado [%]",          lambda k, pm, roc: str(k['droop_calc'])),
                ("ROCOF [Hz/s]",                 lambda k, pm, roc: f"{roc:.4f}" if (roc is not None and roc == roc) else "—"),
            ]

            # Construir dict ordenado: nombre fuente → (kpi, p_max, rocof)
            _comp_srcs = {}
            if _kr:
                _comp_srcs[f"Real ({src_real})"] = (_kr, float(_pm_v), _rocof_r)
            for _sv, (_ks_i, _roc_i, _pm_i) in _kpi_per_src.items():
                _comp_srcs[f"Sim {_sv}"] = (_ks_i, float(_pm_i), _roc_i)

            if _comp_srcs:
                _tabla_b5 = []
                for _lbl, _fn in _KPI_FILAS_B5:
                    _row = {"KPI": _lbl}
                    for _sname, (_kpi_s, _pm_s, _roc_s) in _comp_srcs.items():
                        try:
                            _row[_sname] = _fn(_kpi_s, _pm_s, _roc_s)
                        except Exception:
                            _row[_sname] = "—"
                    _tabla_b5.append(_row)

                _df_comp_b5 = pd.DataFrame(_tabla_b5)

                # Tabla comparativa: HTML puro — centrada, sin scrollbars
                _src_names_b5 = list(_comp_srcs.keys())
                _th_k5 = ("padding:7px 12px;text-align:left;font-weight:600;"
                           "background:#2E4057;color:#fff;font-size:12px;white-space:nowrap;")
                _th_v5 = ("padding:7px 12px;text-align:right;font-weight:600;"
                           "background:#2E4057;color:#fff;font-size:12px;white-space:nowrap;")
                _td_k5 = "padding:5px 12px;border-bottom:1px solid #e0e0e0;font-size:12px;white-space:nowrap;"
                _td_v5 = ("padding:5px 12px;border-bottom:1px solid #e0e0e0;"
                           "font-size:12px;text-align:right;white-space:nowrap;min-width:110px;")
                _hdr5 = f'<th style="{_th_k5}">KPI</th>' + "".join(
                    f'<th style="{_th_v5}">{_sn}</th>' for _sn in _src_names_b5
                )
                _body5 = ""
                for _row5 in _tabla_b5:
                    _lbl5    = _row5["KPI"]
                    _aporta5 = "Aporta" in _lbl5
                    _ispmax5 = "P_max"  in _lbl5
                    _bgrow5  = "background:#f2f2f2;" if _ispmax5 else ""
                    _cells5  = f'<td style="{_td_k5}{_bgrow5}">{_lbl5}</td>'
                    for _sn5 in _src_names_b5:
                        _val5 = _row5.get(_sn5, "—")
                        _cs5  = _td_v5
                        if _aporta5:
                            if "✅" in str(_val5):   _cs5 += "background:#d4edda;color:#155724;"
                            elif "❌" in str(_val5): _cs5 += "background:#f8d7da;color:#721c24;"
                        elif _ispmax5:
                            _cs5 += _bgrow5
                        _cells5 += f'<td style="{_cs5}">{_val5}</td>'
                    _body5 += f'<tr>{_cells5}</tr>'
                _html_b5 = (
                    '<div style="display:flex;justify-content:center;margin:8px 0 4px 0;">'
                    '<table style="border-collapse:collapse;border:1px solid #d0d0d0;border-radius:6px;overflow:hidden;">'
                    f'<thead><tr>{_hdr5}</tr></thead>'
                    f'<tbody>{_body5}</tbody>'
                    '</table></div>'
                )
                st.markdown(_html_b5, unsafe_allow_html=True)

                # Exportar tabla comparativa
                if st.button("⬇️ Exportar comparativa a Excel", key="dl_b5_comp"):
                    try:
                        _excel_comp = _apply_excel_formatting(
                            _df_comp_b5,
                            sheet_name="Comparativa_KPIs",
                            kpi_col="¿Aporta RPF? (ΔP% ≥ 1.5%)",
                            kpi_ok_val="✅ Sí",
                            kpi_error_val="❌ No",
                        )
                        st.download_button(
                            "📥 Descargar",
                            _excel_comp,
                            file_name=f"kpis_comparativa_Ev{n_evento}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl_b5_comp_file",
                        )
                    except Exception as _ex:
                        st.error(f"Error exportando: {_ex}")

                # KPIs evaluados en P_máxima por fuente — tabla multi-columna
                if _kpi_pmax_per_src5:
                    st.markdown("---")
                    _v4_section_head("KPIs CNDC — P_máxima (por fuente)", icon="chart")
                    _src_pm5 = list(_kpi_pmax_per_src5.keys())
                    _KPI_FILAS_PM5 = [
                        ("P_max [MW]",                    lambda k, pm, t, p: f"{pm:.2f}"),
                        ("t_Pmáx [s]",                    lambda k, pm, t, p: f"{t:.1f}"),
                        ("f₀ — Inicio evento [Hz]",       lambda k, pm, t, p: f"{k['f0']:.4f}" if k else "—"),
                        ("P₀ — Inicio evento [MW]",       lambda k, pm, t, p: f"{k['p0']:.3f}" if k else "—"),
                        ("f_min — Nadir [Hz]",            lambda k, pm, t, p: f"{k['f_min']:.4f}" if k else "—"),
                        ("t_min — Nadir [s]",             lambda k, pm, t, p: f"{k['t_min']:.1f}" if k else "—"),
                        ("f en t_Pmáx [Hz]",              lambda k, pm, t, p: f"{k['f_dt']:.4f}" if k else "—"),
                        ("P en t_Pmáx [MW]",              lambda k, pm, t, p: f"{p:.3f}"),
                        ("R_inic [MW]",                   lambda k, pm, t, p: f"{k['r_inic']:.3f}" if k else "—"),
                        ("R_inic [%]",                    lambda k, pm, t, p: f"{k['r_inic_pct']:.2f}" if k else "—"),
                        ("ΔP entregada [MW]",             lambda k, pm, t, p: f"{k['dp']:.3f}" if k else "—"),
                        ("ΔP% aporte [%]",                lambda k, pm, t, p: f"{k['dp_pct']:.2f}" if k else "—"),
                        ("¿Aporta RPF? (ΔP% ≥ 1.5%)",    lambda k, pm, t, p: ("✅ Sí" if k['aporta'] else "❌ No") if k else "—"),
                        ("Droop Nominal [%]",             lambda k, pm, t, p: f"{k['droop_nom']:.1f}" if k else "—"),
                        ("Droop Calculado [%]",           lambda k, pm, t, p: str(k['droop_calc']) if k else "—"),
                    ]
                    _th_kpm = ("padding:7px 12px;text-align:left;font-weight:600;"
                               "background:#2E4057;color:#fff;font-size:12px;white-space:nowrap;")
                    _th_vpm = ("padding:7px 12px;text-align:right;font-weight:600;"
                               "background:#2E4057;color:#fff;font-size:12px;white-space:nowrap;")
                    _td_kpm = "padding:5px 12px;border-bottom:1px solid #e0e0e0;font-size:12px;white-space:nowrap;"
                    _td_vpm = ("padding:5px 12px;border-bottom:1px solid #e0e0e0;"
                               "font-size:12px;text-align:right;white-space:nowrap;min-width:110px;")
                    _hdrpm = f'<th style="{_th_kpm}">KPI</th>' + "".join(
                        f'<th style="{_th_vpm}">{_sn}</th>' for _sn in _src_pm5
                    )
                    _bodypm = ""
                    for _lblpm, _fnpm in _KPI_FILAS_PM5:
                        _aporta_rowpm = "Aporta" in _lblpm
                        _ispmax_rowpm = "P_max"  in _lblpm
                        _bgrow_pm     = "background:#f2f2f2;" if _ispmax_rowpm else ""
                        _cells_pm     = f'<td style="{_td_kpm}{_bgrow_pm}">{_lblpm}</td>'
                        for _snpm in _src_pm5:
                            _kpm5, _pm5p, _t5p, _p5p = _kpi_pmax_per_src5[_snpm]
                            try:
                                _val_pm = _fnpm(_kpm5, _pm5p, _t5p, _p5p)
                            except Exception:
                                _val_pm = "—"
                            _cspm = _td_vpm
                            if _aporta_rowpm:
                                if "✅" in str(_val_pm):   _cspm += "background:#d4edda;color:#155724;"
                                elif "❌" in str(_val_pm): _cspm += "background:#f8d7da;color:#721c24;"
                            elif _ispmax_rowpm:
                                _cspm += _bgrow_pm
                            _cells_pm += f'<td style="{_cspm}">{_val_pm}</td>'
                        _bodypm += f'<tr>{_cells_pm}</tr>'
                    _html_pm5 = (
                        '<div style="display:flex;justify-content:center;margin:8px 0 4px 0;">'
                        '<table style="border-collapse:collapse;border:1px solid #d0d0d0;'
                        'border-radius:6px;overflow:hidden;">'
                        f'<thead><tr>{_hdrpm}</tr></thead>'
                        f'<tbody>{_bodypm}</tbody>'
                        '</table></div>'
                    )
                    st.markdown(_html_pm5, unsafe_allow_html=True)

                # --- Curva de Error de Seguimiento y Barras KPI ---------------
                if _sim_for_error:
                    ce_col, bar_col = st.columns([1, 1])
                    with ce_col:
                        fig_err = go.Figure()
                        for s in _sim_for_error:
                            f_real_interp = np.interp(s["t"], tr_aligned, _fr_arr)
                            err = f_real_interp - s["f"]
                            fig_err.add_trace(go.Scatter(
                                x=s["t"], y=err, name=f"Err {s['ver']}",
                                line=dict(color=s["color"]),
                            ))
                        fig_err.update_layout(
                            title="Error de Seguimiento de Frecuencia (Hz)",
                            xaxis_title="Tiempo desde t₀ (s)",
                            yaxis_title="Error (Hz)",
                            height=350, template=_gcfg["template"],
                        )
                        st.plotly_chart(fig_err, use_container_width=True)

                    with bar_col:
                        fig_bar = go.Figure()
                        _fuentes_bar = [r["Fuente"] for r in _kpi_rows]
                        _dps_bar     = [r["dp_pct"] for r in _kpi_rows]
                        _bar_colors  = [
                            _gcfg["pot_color_real"] if "REAL" in f
                            else (_gcfg["pot_color_sim0"] if "0" in f else _gcfg["pot_color_sim1"])
                            for f in _fuentes_bar
                        ]
                        fig_bar.add_trace(go.Bar(
                            x=_fuentes_bar, y=_dps_bar,
                            marker_color=_bar_colors,
                            text=[f"{v:.2f}%" for v in _dps_bar],
                            textposition="outside",
                        ))
                        fig_bar.add_hline(
                            y=1.5, line_dash="dash", line_color="red",
                            annotation_text="Mínimo 1.5% (CNDC)",
                        )
                        fig_bar.update_layout(
                            title="Aporte Porcentual ΔP (%) por Fuente",
                            yaxis_title="ΔP%",
                            height=350, template=_gcfg["template"],
                        )
                        st.plotly_chart(fig_bar, use_container_width=True)
else:
        st.info("ℹ️ Seleccione una unidad y verifique los archivos en las carpetas correspondientes.")

if bloque_trabajo == "kpi_historico":
    _render_block_header("06", "Histórico RPF",
        "Análisis histórico de cumplimiento RPF por unidad y evento. Datos desde el servidor PostgreSQL.",
        "Análisis", pf_required=False)
    from bloque_kpi_historico import render_bloque_kpi
    render_bloque_kpi(st.session_state)

if bloque_trabajo == "reporte_tecnico":
    _render_block_header("07", "Reporte Técnico",
        "Consolida KPIs de SCADA, EMF y simulación. Exporta informe para entrega al CNDC.",
        "Salida", pf_required=False)
    # Inicializar estado del generador de reporte
    for _k7 in ("b7_docx_bytes", "b7_docx_ctx", "b7_show_ctx_pack"):
        if _k7 not in st.session_state:
            st.session_state[_k7] = None
    _v4_section_head("Configuración del Reporte", icon="sliders")
    if not st.session_state.semestre_global or not st.session_state.evento_global:
        st.warning("👈 Seleccione Semestre y Evento en la barra lateral para ver la auditoría del proyecto.")
        st.stop()

    ev_path = st.session_state.ev_path_global
    n_evento = st.session_state.n_evento_global

    #  SECCIÓN 1: AUDITORÍA DE ARCHIVOS DEL EVENTO 
    _v4_section_head("Auditoría de Archivos del Evento", icon="server")
    st.caption("Estado actual de los archivos generados y requeridos para el evento seleccionado.")
    
    def _check_file(path_glob):
        hits = glob.glob(path_glob)
        return (True, os.path.basename(hits[0])) if hits else (False, "Faltante")

    audit_data = [
        {"Bloque": "1. Carga", "Concepto": "Datos Simulación (Extracción)", "Estado": _check_file(os.path.join(ev_path, "datos_simulacion_*_2daopcion.xlsx"))},
        {"Bloque": "1. Carga", "Concepto": "Condiciones Iniciales", "Estado": _check_file(os.path.join(ev_path, "condiciones_iniciales_*.xlsx"))}, # type: ignore
        {"Bloque": "1. Carga", "Concepto": "Resultados Carga PowerFactory", "Estado": _check_file(os.path.join(ev_path, f"datos_cargados_Ev{n_evento}.xlsx"))}, # type: ignore
        {"Bloque": "2. Real", "Concepto": "Registros SCADA (1 SEG)", "Estado": (os.path.isdir(os.path.join(ev_path, "Graficas Registro 1SEG COBEE")), "Carpeta de Unidades")}, # type: ignore
        {"Bloque": "2. Real", "Concepto": "Digitalización EMF CNDC", "Estado": (os.path.isdir(os.path.join(ev_path, "Resultados_COBEE")), "Carpeta de Unidades")}, # type: ignore
        {"Bloque": "3. Simu", "Concepto": "Curvas Simulación CNDC (E.0)", "Estado": (os.path.isdir(os.path.join(ev_path, f"E{n_evento}.0", "Datos Curvas")), "Disponible")},
        {"Bloque": "3. Simu", "Concepto": "Curvas Simulación COBEE (E.1)", "Estado": (os.path.isdir(os.path.join(ev_path, f"E{n_evento}.1", "Datos Curvas")), "Disponible")},
    ]

    df_audit = pd.DataFrame([
        {
            "Bloque": d["Bloque"],
            "Documento/Proceso": d["Concepto"],
            "Estado": "✅" if d["Estado"][0] else "❌",
            "Detalle": d["Estado"][1]
        } for d in audit_data
    ])
    st.table(_df_safe(df_audit))

    if st.button("⬇️ Exportar Auditoría a Excel"):
        _df_audit_exp = df_audit.copy()
        _df_audit_exp.columns = ["Bloque", "Documento/Proceso", "Estado (Icono)", "Detalle"]
        excel_audit = _apply_excel_formatting(
            _df_audit_exp,
            sheet_name="Auditoria_Proyecto",
            kpi_col="Estado (Icono)",
            kpi_ok_val="✅",
            kpi_error_val="❌"
        )
        st.download_button("Descargar Reporte de Auditoría", excel_audit,
                           file_name=f"auditoria_Ev{n_evento}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    #  SECCIÓN 2: DOCUMENTACIÓN DE MEMORIA Y PROCESOS 
    st.markdown("---") # type: ignore
    _v4_section_head("Manual Técnico y Memoria del Sistema", icon="report")

    with st.expander("🧠 Archivos de Memoria y Contexto", expanded=True):
        st.markdown(r"""
        La interfaz mantiene la continuidad del trabajo mediante archivos JSON que actúan como la 'memoria' del sistema:

        | Archivo | Función | Ubicación |
        | :--- | :--- | :--- |
        | `config_rutas.json` | Almacena las rutas raíz (CNDC, PowerFactory, Mapeos) definidas en el Sidebar. | Directorio del script principal |
        | `estatismo_config.json` | Guarda los valores de Estatismo (Rp) ajustados manualmente para cada unidad generadora. | `...\Designacion de loc_name\` |
        | `_streamlit_params.json` | Puente de parámetros entre la UI y los scripts de ejecución no interactiva. | Carpeta del Evento |
        """)

    with st.expander(" Bloque 0: Datos del Modelo (Gestión de Mapeo)", expanded=False):
        st.markdown(r"""
        **Objetivo:** Sincronizar los catálogos técnicos con el modelo `.pfd` de PowerFactory.
        
        | Script | Entrada | Ubicación Entrada | Salida | Ubicación Salida |
        | :--- | :--- | :--- | :--- | :--- |
        | `DatsoGENBUSLNE.py` | Red PowerFactory | Proyecto PF Activo | `DatosSINdigsilent.xlsx` | `...\DATOS EXTRAIDOS DE DIGSILENT\` |
        | `loc_namesGEN.py` | `DatosSINdigsilent.xlsx`, Despacho CNDC | `...\DATOS EXTRAIDOS...`, `RAIZ_DATOS` | `loc_names_gen.xlsx` | `...\Designacion de loc_name\` |
        | `loc_namesLineas.py` | `DatosSINdigsilent.xlsx` | `...\DATOS EXTRAIDOS...` | `loc_names_lineas.xlsx` | `...\Designacion de loc_name\` |
        | `loc_names_xfo.py` | `DatosSINdigsilent.xlsx` | `...\DATOS EXTRAIDOS...` | `loc_names_xfo.xlsx` | `...\Designacion de loc_name\` |
        | `MapeoRetirosSTI.py`| `deener_*.xlsx`, `DatosSINdigsilent.xlsx` | Carpeta Evento / `...\DATOS...` | `loc_name_cargas.xlsx` | `...\Designacion de loc_name\` |
        """)

    with st.expander("📦 Bloque 1: Carga de Datos y PowerFactory", expanded=False):
        st.markdown(fr"""
        **Objetivo:** Definir el despacho y demanda (Snapshot) para la simulación RMS.

        | Script | Entrada | Ubicación Entrada | Salida | Ubicación Salida |
        | :--- | :--- | :--- | :--- | :--- |
        | `ExtFLujos2daO.py` | `dc_*, dcdr_*, deener_*, Tabla_Eventos` | Carpeta Evento / `RAIZ` | `datos_simulacion_*.xlsx` | Carpeta del Evento |
        | `CondInicialesPF.py`| `datos_simulacion_*.xlsx`, Mapeos (GEN/CAR/XFO) | Carpeta Evento / `...\Designacion...` | `condiciones_iniciales_*.xlsx` | Carpeta del Evento |
        | `CargaCondIniciales_PF.py` | `condiciones_iniciales_*.xlsx`, `loc_names_xfo.xlsx` | Carpeta Evento / `...\Designacion...` | `datos_cargados_Ev{n_evento}.xlsx` | Carpeta del Evento |
        """)

    with st.expander("📊 Bloque 2: Análisis de Datos Registrados", expanded=False):
        st.markdown(r"""
        **Objetivo:** Procesar registros de campo (SCADA) y gráficas oficiales CNDC.

        | Script | Entrada | Ubicación Entrada | Salida | Ubicación Salida |
        | :--- | :--- | :--- | :--- | :--- |
        | `OrdenadorDatosEvento.py` | `1 seg.*.xls` (Falla) | `RAIZ_DATOS` / Año / Carpeta FALLA | `Unidad.xlsx` (Excel 1seg) | `...\Graficas Registro 1SEG COBEE\` |
        | `ExtractorResultadosCNDC.py`| Archivos `*.emf` del CNDC | Carpeta del Evento | `Unidad.xlsx` (Digitalizado) | `...\Resultados_COBEE\` |
        """)

    with st.expander("📈 Bloque 3: Análisis de Simulación", expanded=False):
        st.markdown(r"""
        **Objetivo:** Exportar los resultados de las simulaciones RMS de PowerFactory.

        | Script | Ejecución | Salida | Ubicación Salida |
        | :--- | :--- | :--- | :--- |
        | `DatosCurvas_v3.py` | Script Python en PowerFactory | Archivos `.xlsx` por página de gráficos | `...\E{N}.x\Datos Curvas\` |
        """)

    with st.expander("⚖️ Bloque 4: Validación y Comparativa", expanded=False):
        st.markdown("""
        **Integración de Datos:**
        
        | Comparación | Fuente A | Fuente B | t = 0 (Referencia) |
        | :--- | :--- | :--- | :--- |
        | **Alineación** | Registro Real (SCADA/EMF) | Simulación (PowerFactory) | Instante de falla (df/dt < Umbral) |

        **Salidas:**
        *   **KPIs:** f₀, f_min, ΔP, Droop calculado, ROCOF, RMSE.
        *   **Exportación:** Imágenes `.png` individuales y archivos `.zip` consolidados por evento.
        """)

    #  SECCIÓN 3: RESUMEN DE RUTAS 
    st.markdown("---")
    _v4_section_head("Rutas de Memoria Activas", icon="server")
    
    c_r1, c_r2 = st.columns(2)
    with c_r1:
        st.write("**📁 Almacenamiento CNDC:**")
        st.code(f"RAIZ RPF: {RAIZ}\nRAIZ DATOS: {RAIZ_DATOS}")
        st.write("**⚙️ PowerFactory:**")
        st.code(f"Proyecto: {PF_PROYECTO}\nCaso Base: {CASO_BASE}")
    
    with c_r2:
        st.write("**📄 Archivos de Mapeo:**")
        st.caption(f"Generadores: `{os.path.basename(LOC_NAMES_GEN_PATH)}`")
        st.caption(f"Cargas: `{os.path.basename(LOC_CAR_PATH)}`")
        st.caption(f"Trafos: `{os.path.basename(LOC_XFO_PATH)}`")

    # ── SECCIÓN 4: GENERADOR DE REPORTE TÉCNICO ──────────────────────────────
    st.markdown("---")
    _v4_section_head(
        "Generador de Reporte Técnico",
        "Consolida KPIs de SCADA, EMF y simulación en un documento Word listo para entregar al CNDC.",
        "report",
    )

    _sel_unit_b7 = st.session_state.global_selected_unit
    _ev_cfg_b7   = _load_event_cfg(ev_path)
    _b7_dt       = int(_ev_cfg_b7.get("delta_t_cndc", 35))
    _b7_t_falla  = float(_ev_cfg_b7.get("t_sim_falla", 5.0))

    if not _sel_unit_b7:
        st.info("ℹ️ Seleccione una unidad generadora en el selector superior para generar el reporte.")
    else:
        _unit_clean_b7 = _sel_unit_b7.replace("sym_", "").replace("SYM_", "")
        st.caption(f"Unidad seleccionada: **{_unit_clean_b7}** | Δt={_b7_dt} s | t₀={_b7_t_falla:.1f} s")

        _cb1, _cb2, _cb3 = st.columns(3)
        with _cb1:
            _b7_incluir_hist  = st.checkbox("Incluir histórico de la unidad", value=True,  key="b7_hist")
        with _cb2:
            _b7_incluir_datos = st.checkbox("Incluir apéndice de datos crudos", value=False, key="b7_raw")
        with _cb3:
            _b7_subir_sp = st.checkbox(
                "Subir a SharePoint automáticamente", value=False, key="b7_sp",
                help="Requiere conexión activa a SharePoint (contraseña configurada en secrets.toml)"
            )

        _btn_col1, _btn_col2, _ = st.columns([1, 1, 2])
        with _btn_col1:
            _btn_generar_b7  = st.button("📄 Generar Reporte Word", type="primary", key="b7_gen_word")
        with _btn_col2:
            _btn_ctx_pack_b7 = st.button("🤖 Copilot Context Pack", key="b7_ctx_pack",
                                          help="Genera texto estructurado para copiar/pegar en Teams Copilot Chat")

        # ── Copilot Context Pack (operación ligera, sin generar Word) ────────
        if _btn_ctx_pack_b7:
            st.session_state["b7_show_ctx_pack"] = True
        if st.session_state.get("b7_show_ctx_pack"):
            with st.spinner("Preparando datos para Copilot..."):
                try:
                    from reporte_tecnico import collect_kpis_para_reporte, generar_copilot_context_pack
                    _ctx_cp = collect_kpis_para_reporte(
                        ev_path=ev_path, n_evento=n_evento, unit_key=_sel_unit_b7,
                        delta_t=_b7_dt, t_falla=_b7_t_falla,
                        loc_gen_path=LOC_NAMES_GEN_PATH,
                    )
                    _ctx_cp.semestre = st.session_state.semestre_global or ""
                    _pack_txt = generar_copilot_context_pack(_ctx_cp)
                except Exception as _e_cp:
                    st.error(f"Error al preparar contexto: {_e_cp}")
                    _pack_txt = None
            if _pack_txt:
                st.markdown("---")
                _v4_section_head(
                    "Copilot Context Pack",
                    "Selecciona todo el bloque y pégalo directamente en Teams Copilot Chat o en Word Copilot",
                    "report",
                )
                st.info("💡 **Cómo usar:** Copia el texto de abajo → abre Teams → Copilot Chat → pega y envía el prompt que necesites.")
                st.code(_pack_txt, language="markdown")
                if st.button("✕ Cerrar Context Pack", key="b7_close_ctx"):
                    st.session_state["b7_show_ctx_pack"] = False
                    st.rerun()

        # ── Generación del Word ──────────────────────────────────────────────
        if _btn_generar_b7:
            with st.spinner(f"Generando reporte para {_unit_clean_b7}... (puede tardar 15-30 s)"):
                try:
                    from reporte_tecnico import (
                        collect_kpis_para_reporte,
                        generar_reporte_word,
                        subir_reporte_a_sharepoint,
                    )
                    _ctx_b7 = collect_kpis_para_reporte(
                        ev_path=ev_path, n_evento=n_evento, unit_key=_sel_unit_b7,
                        delta_t=_b7_dt, t_falla=_b7_t_falla,
                        loc_gen_path=LOC_NAMES_GEN_PATH,
                    )
                    _ctx_b7.semestre = st.session_state.semestre_global or ""
                    _docx_b7 = generar_reporte_word(
                        _ctx_b7,
                        incluir_historico=_b7_incluir_hist,
                        incluir_apendice=_b7_incluir_datos,
                    )
                    st.session_state["b7_docx_bytes"] = _docx_b7
                    st.session_state["b7_docx_ctx"]   = _ctx_b7
                    st.success("✅ Reporte generado correctamente.")

                    # Subida a SharePoint opcional
                    if _b7_subir_sp and _SP_OK:
                        import sharepoint_client as _sp_mod
                        _ok_sp, _msg_sp = subir_reporte_a_sharepoint(
                            _docx_b7, ev_path, n_evento, _unit_clean_b7,
                            st.session_state.semestre_global or "",
                            upload_fn=_sp.upload_file,
                            local_to_sp_fn=getattr(_sp, "local_path_to_sp_folder",
                                                    getattr(_sp, "_local_to_sp_folder", None)),
                            raiz_local=RAIZ,
                        )
                        if _ok_sp:
                            st.toast("✅ Reporte subido a SharePoint", icon="✅")
                        else:
                            st.warning(f"SharePoint: {_msg_sp}")
                    elif _b7_subir_sp and not _SP_OK:
                        st.warning("SharePoint no está conectado. Descarga el archivo manualmente.")
                except Exception as _e_gen:
                    st.error(f"Error al generar el reporte: {_e_gen}")
                    with st.expander("Detalle del error"):
                        import traceback
                        st.code(traceback.format_exc())

        # ── Botón de descarga (persiste entre reruns) ────────────────────────
        if st.session_state.get("b7_docx_bytes"):
            _ctx_dl = st.session_state.get("b7_docx_ctx")
            _unit_dl = (_ctx_dl.unit_name if _ctx_dl else _unit_clean_b7).replace(" ", "_")
            _sem_dl  = st.session_state.semestre_global or "SEM"
            _fname_dl = f"RPF_{_sem_dl}_Ev{n_evento}_{_unit_dl}.docx"
            st.download_button(
                label="⬇️ Descargar Reporte Word",
                data=st.session_state["b7_docx_bytes"],
                file_name=_fname_dl,
                mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                key="b7_dl_btn",
            )
            # Resumen de fuentes incluidas en el reporte
            if _ctx_dl:
                st.markdown("**Fuentes incluidas en el reporte:**")
                _audit_cols = st.columns(4)
                _audit_info = [
                    ("SCADA (1SEG)",           _ctx_dl.kpi_scada  is not None),
                    ("EMF CNDC",               _ctx_dl.kpi_emf    is not None),
                    (f"Sim E{n_evento}.0",     _ctx_dl.kpi_sim_e0 is not None),
                    (f"Sim E{n_evento}.1",     _ctx_dl.kpi_sim_e1 is not None),
                ]
                for (label_a, ok_a), col_a in zip(_audit_info, _audit_cols):
                    col_a.metric(label_a, "✅ Incluida" if ok_a else "⬜ N/D")

    # Botón para abrir la carpeta del evento (Solo Windows)
    if st.button("📂 Abrir Carpeta del Evento en Explorador"):
        if os.path.isdir(ev_path):
            os.startfile(ev_path)
        else:
            st.error("La ruta del evento no es válida.")
