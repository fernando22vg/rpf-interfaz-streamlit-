#!/usr/bin/env python3
"""
diagnostico_verificacion_eventos.py

Auditoria post-fix del bug de indices en ExtFLujos2daO_run.py (orden alfabetico
de carpetas "Evento N" vs orden numerico de Tabla_Eventos_*.xlsx).

Para cada carpeta "Evento N" de cada semestre en RAIZ, verifica:
  1. datos_simulacion_*_2daopcion.xlsx (Tab 1):
     - el numero de evento embebido en la hoja Info_Evento coincide con N
     - la fecha del nombre de archivo coincide con la fecha de la fila N en
       Tabla_Eventos_*.xlsx
  2. condiciones_iniciales_*_EvM.xlsx (Tab 2):
     - el sufijo "EvM" del nombre de archivo coincide con N
       (CondInicialesPF.py copia el numero de evento desde Info_Evento del
       archivo de Tab 1, asi que un mismatch aqui confirma que el archivo de
       Tab 1 en esa carpeta tenia datos de OTRO evento)
  3. carga_Ev{N}.log (Tab 3, generado por CargaCondIniciales_PF.py):
     - si existe y es mas antiguo que el condiciones_iniciales actual, la
       carga a PowerFactory se hizo con una version de datos distinta a la
       que hay ahora en disco -> debe re-cargarse

Resultado: texto en consola, marcando cada carpeta como OK / MISMATCH / FALTA
/ DESACTUALIZADO.
"""

import os
import re
import glob
import sys

import openpyxl

RAIZ = r"C:\Datos del CNDC\01_INFO CNDC_RPF"


def leer_tabla_eventos(semestre_path):
    """Devuelve {num_evento: {'fecha_hora': str, 'fecha_str': 'DDMMYY', 'disparo': str}}"""
    tabla_glob = glob.glob(os.path.join(semestre_path, "Tabla_Eventos_*.xlsx"))
    if not tabla_glob:
        return None
    wb = openpyxl.load_workbook(tabla_glob[0], data_only=True)
    sh = wb.active
    tabla = {}
    for fila in sh.iter_rows(min_row=3, values_only=True):
        if fila[0] is None:
            continue
        try:
            num = int(fila[0])
        except (TypeError, ValueError):
            continue
        fecha_hora = str(fila[1]).strip()
        d, m, y = fecha_hora.split(" ")[0].split("/")
        tabla[num] = {
            "fecha_hora": fecha_hora,
            "fecha_str": f"{d}{m}{y[2:]}",
            "disparo": str(fila[2]).strip() if fila[2] else "",
        }
    return tabla


def leer_info_evento(path_xlsx):
    try:
        wb = openpyxl.load_workbook(path_xlsx, data_only=True)
        sh = wb["Info_Evento"]
        info = {}
        for fila in sh.iter_rows(min_row=2, values_only=True):
            if fila[0] is None:
                continue
            info[str(fila[0]).strip()] = fila[1]
        return info
    except Exception as e:
        return {"_error": str(e)}


def auditar_semestre(semestre, semestre_path):
    print(f"\n{'=' * 78}")
    print(f"  SEMESTRE: {semestre}")
    print(f"{'=' * 78}")

    tabla = leer_tabla_eventos(semestre_path)
    if tabla is None:
        print("  [AVISO] No se encontro Tabla_Eventos_*.xlsx, se omite semestre.")
        return []

    eventos_dir = os.path.join(semestre_path, "Análisis_todos_los_eventos")
    if not os.path.isdir(eventos_dir):
        eventos_dir = os.path.join(semestre_path, "Analisis_todos_los_eventos")
    if not os.path.isdir(eventos_dir):
        print("  [AVISO] No se encontro carpeta de eventos.")
        return []

    carpetas = sorted(
        (d for d in os.listdir(eventos_dir) if os.path.isdir(os.path.join(eventos_dir, d))),
        key=lambda s: int(re.search(r"(\d+)$", s).group(1)) if re.search(r"(\d+)$", s) else 0,
    )

    problemas = []

    for carpeta in carpetas:
        m = re.search(r"(\d+)$", carpeta)
        if not m:
            continue
        n = int(m.group(1))
        ev_path = os.path.join(eventos_dir, carpeta)
        esperado = tabla.get(n)
        estado_linea = [f"  Evento {n:>2} ({carpeta})"]

        if esperado is None:
            print(f"  Evento {n:>2}: [AVISO] no esta en Tabla_Eventos (¿evento fuera de rango?)")
            continue

        # ---- Tab 1: datos_simulacion_*_2daopcion.xlsx ----
        dsim = glob.glob(os.path.join(ev_path, "datos_simulacion_*_2daopcion.xlsx"))
        if not dsim:
            print(f"  Evento {n:>2}: TAB1 FALTA       (no hay datos_simulacion_*_2daopcion.xlsx)")
            problemas.append((n, carpeta, "TAB1_FALTA"))
            tab1_ok = False
        else:
            fn = os.path.basename(dsim[0])
            mf = re.search(r"datos_simulacion_(\d+)_", fn)
            fecha_archivo = mf.group(1) if mf else "?"
            info = leer_info_evento(dsim[0])
            n_evento_info = str(info.get("Evento N°", "")).strip()
            fecha_ok = fecha_archivo == esperado["fecha_str"]
            num_ok = n_evento_info == str(n)
            if fecha_ok and num_ok:
                print(f"  Evento {n:>2}: TAB1 OK          ({fn})")
                tab1_ok = True
            else:
                print(f"  Evento {n:>2}: TAB1 MISMATCH    ({fn})"
                      f"  esperado fecha={esperado['fecha_str']} num={n}"
                      f"  -> archivo fecha={fecha_archivo} num={n_evento_info}"
                      f"  [contiene datos de OTRO evento]")
                problemas.append((n, carpeta, "TAB1_MISMATCH"))
                tab1_ok = False

        # ---- Tab 2: condiciones_iniciales_*_EvM.xlsx ----
        cini = glob.glob(os.path.join(ev_path, "condiciones_iniciales_*.xlsx"))
        if not cini:
            print(f"  Evento {n:>2}: TAB2 FALTA       (no hay condiciones_iniciales_*.xlsx)")
            problemas.append((n, carpeta, "TAB2_FALTA"))
            ci_path = None
        else:
            fn2 = os.path.basename(cini[0])
            m2 = re.search(r"_Ev(\d+)\.xlsx$", fn2)
            ev_suffix = int(m2.group(1)) if m2 else None
            if ev_suffix == n:
                print(f"  Evento {n:>2}: TAB2 OK          ({fn2})")
            else:
                print(f"  Evento {n:>2}: TAB2 MISMATCH    ({fn2})"
                      f"  carpeta=Evento {n} pero archivo dice Ev{ev_suffix}"
                      f"  [confirma que Tab1 tenia datos de otro evento al generar este]")
                problemas.append((n, carpeta, "TAB2_MISMATCH"))
            ci_path = cini[0]

        # ---- Tab 3: carga_Ev{N}.log (generado por CargaCondIniciales_PF.py) ----
        log_path = os.path.join(ev_path, f"carga_Ev{n}.log")
        if os.path.isfile(log_path):
            if ci_path is None:
                print(f"  Evento {n:>2}: TAB3 SOSPECHOSO  (existe carga_Ev{n}.log pero no hay"
                      f" condiciones_iniciales_*.xlsx actual -> no se puede verificar con que"
                      f" datos se cargo a PowerFactory, revisar manualmente)")
                problemas.append((n, carpeta, "TAB3_SIN_FUENTE_ACTUAL"))
            elif os.path.getmtime(log_path) < os.path.getmtime(ci_path):
                print(f"  Evento {n:>2}: TAB3 DESACTUALIZADO  (carga_Ev{n}.log es mas antiguo que"
                      f" {os.path.basename(ci_path)} -> la carga a PowerFactory NO usa"
                      f" la version actual de los datos, re-cargar)")
                problemas.append((n, carpeta, "TAB3_DESACTUALIZADO"))
            else:
                print(f"  Evento {n:>2}: TAB3 OK          (carga_Ev{n}.log al dia)")
        else:
            print(f"  Evento {n:>2}: TAB3 sin cargar  (no hay carga_Ev{n}.log, evento aun no cargado a PF)")

    return problemas


def main():
    if not os.path.isdir(RAIZ):
        print(f"[ERROR] No existe RAIZ: {RAIZ}")
        sys.exit(1)

    semestres = sorted(d for d in os.listdir(RAIZ) if os.path.isdir(os.path.join(RAIZ, d)))
    todos_problemas = []
    for semestre in semestres:
        problemas = auditar_semestre(semestre, os.path.join(RAIZ, semestre))
        for n, carpeta, tipo in problemas:
            todos_problemas.append((semestre, n, carpeta, tipo))

    print(f"\n{'=' * 78}")
    print("  RESUMEN")
    print(f"{'=' * 78}")
    if not todos_problemas:
        print("  Sin problemas detectados. Todos los archivos Tab1/Tab2/Tab3 son consistentes.")
    else:
        print(f"  {len(todos_problemas)} problema(s) detectado(s):\n")
        for semestre, n, carpeta, tipo in todos_problemas:
            print(f"    [{tipo:<20}] {semestre} / {carpeta}")
        print("\n  Accion recomendada: para cada TAB1_MISMATCH o TAB1_FALTA, re-ejecutar Tab 1")
        print("  (ahora con el fix aplicado), luego Tab 2, y si corresponde, re-cargar Tab 3.")


if __name__ == "__main__":
    main()
