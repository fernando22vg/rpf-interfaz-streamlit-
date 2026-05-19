# =================================================================
# SCRIPT GENERALIZADO v3.0 - EXTRACTOR DE TABLAS PES DESDE PDF
# =================================================================
# Extrae automaticamente tablas de cualquier PDF de Tablas PES
# REQUISITOS: pip install pypdfium2 pdfplumber pandas openpyxl
# USO: python tabla_DSL_pdf.py
# =================================================================

import pdfplumber
import pypdfium2 as pdfium
import pandas as pd
import re, sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# =================================================================
# CONFIGURACION - DESCRIPCIONES DE TABLAS
# =================================================================
TABLE_KEYWORDS = {
    "control de velocidad": {"desc": "Control de velocidad",      "model": "GOV_REIVAX"},
    "array k":              {"desc": "Array K",                   "model": "Array"},
    "arrayk":               {"desc": "Array K",                   "model": "Array"},
    "actuador":             {"desc": "Datos del actuador",         "model": "PELTON_ATUADOR_SIMP"},
    "control del actuador": {"desc": "Control del actuador",      "model": "PELTON_MALHA_POS_SIMP"},
    "conducto":             {"desc": "Conducto y turbina",        "model": "PELTON_CONDUTO_TURBINA"},
    "array px":             {"desc": "Array Px",                  "model": "Array"},
    "array yayd":           {"desc": "Array YAYD",                "model": "Array"},
    "yayd":                 {"desc": "Array YAYD",                "model": "Array"},
    "control de tension":   {"desc": "Control de Tensión (AVR)",  "model": "AVR"},
    "avr":                  {"desc": "Control de Tensión (AVR)",  "model": "AVR"},
    "drive":                {"desc": "Drive",                     "model": "DRIVE"},
    "excitatriz":           {"desc": "Excitatriz",               "model": "EXCITATRIZ"},
    "pss":                  {"desc": "PSS",                       "model": "PSS_COMP"},
    "limitador vhz":        {"desc": "Limitador VHZ",            "model": "VHZL"},
    "vhz":                  {"desc": "Limitador VHZ",            "model": "VHZL"},
    "limitador uel":        {"desc": "Limitador UEL",            "model": "UEL"},
    "uel":                  {"desc": "Limitador UEL",            "model": "UEL"},
    "limitador oel":        {"desc": "Limitador OEL",            "model": "OEL"},
    "oel":                  {"desc": "Limitador OEL",            "model": "OEL"},
    "limitador scl":        {"desc": "Limitador SCL",            "model": "SCL"},
    "scl":                  {"desc": "Limitador SCL",            "model": "SCL"},
    "limitador mel":        {"desc": "Limitador MEL",            "model": "MEL"},
    "mel":                  {"desc": "Limitador MEL",            "model": "MEL"},
}

# Numero de tablas validas esperadas (para diagnostico)
TABLA_RANGE = range(60, 100)


# =================================================================
# EXTRACCION DE TEXTO CON PYPDFIUM2
# =================================================================

def extract_text_pypdfium(pdf_path, page_index):
    """
    Extrae texto de una pagina usando pypdfium2 (motor PDFium de Chrome).
    Retorna el texto completo de la pagina.
    """
    try:
        doc = pdfium.PdfDocument(pdf_path)
        page = doc[page_index]
        textpage = page.get_textpage()
        text = textpage.get_text_range()
        doc.close()
        return text or ""
    except Exception as e:
        print(f"      [!] pypdfium2 error en pag {page_index+1}: {e}")
        return ""


def get_page_count_pypdfium(pdf_path):
    """Retorna el numero de paginas del PDF."""
    try:
        doc = pdfium.PdfDocument(pdf_path)
        n = len(doc)
        doc.close()
        return n
    except Exception as e:
        print(f"[!] Error al abrir PDF: {e}")
        return 0


# =================================================================
# PARSEO DE TABLAS DESDE TEXTO EXTRAIDO
# =================================================================

def normalize_number(s):
    """Convierte '0,020' -> '0.020' para homogenizar decimales."""
    s = s.strip()
    # Si tiene coma como decimal europeo (no separador de miles): 0,020 -> 0.020
    # Patron: digito-coma-digito
    if re.match(r'^-?\d+,\d+$', s):
        return s.replace(',', '.')
    return s


def identify_table_description(text_context):
    """Identifica descripcion y modelo segun el texto de la pagina."""
    text_lower = text_context.lower()
    for keyword, info in TABLE_KEYWORDS.items():
        if keyword in text_lower:
            return info["desc"], info["model"]
    return "Datos", "Desconocido"


def parse_tables_from_text(page_text):
    """
    Parsea el texto de una pagina y extrae tablas de parametros.

    Detecta dos tipos:
      1. Tabla de parametros: columnas Parametro | Valor
      2. Tabla de arrays:     columnas X | Y
      3. Tabla de 4 columnas (PSS): Parametro|Valor|Parametro|Valor

    Retorna lista de dicts: [{"num": int, "header": [...], "rows": [[...]]}]
    """
    lines = [l.strip() for l in page_text.splitlines()]
    lines = [l for l in lines if l]

    found_tables = []
    i = 0

    # Patron para "Tabla 63 - Datos del control de velocidad" o "Tabla 63"
    pat_tabla = re.compile(r'Tabla\s+(\d+)', re.IGNORECASE)
    # Patron para numero/valor (acepta negativos, comas, puntos)
    pat_num = re.compile(r'^-?\d+[,.]?\d*$')
    # Patron para linea de datos param+valor: "TmedP 0,020"
    pat_param_val = re.compile(
        r'^([A-Za-z_][A-Za-z0-9_\s\-\.]*?)\s{2,}(-?\d[\d,\.]*)\s*$'
        r'|^([A-Za-z_][A-Za-z0-9_\s\-\.]*?)\s+(-?\d[\d,\.]+)$'
    )

    while i < len(lines):
        line = lines[i]

        # --- Detectar encabezado de tabla ---
        m_tabla = pat_tabla.search(line)
        if not m_tabla:
            i += 1
            continue

        table_num = int(m_tabla.group(1))
        if table_num not in TABLA_RANGE:
            i += 1
            continue

        # Buscar la fila de encabezado de columnas en las siguientes lineas
        header = None
        header_idx = i + 1
        while header_idx < min(i + 6, len(lines)):
            hl = lines[header_idx].strip().lower()
            # Encabezado de tabla parametro/valor
            if re.search(r'par[aá]metro', hl) and 'valor' in hl:
                header = ['Parametro', 'Valor']
                # Detectar si tiene 4 columnas (doble par parametro-valor)
                if hl.count('valor') >= 2:
                    header = ['Parametro', 'Valor', 'Parametro2', 'Valor2']
                header_idx += 1
                break
            # Encabezado de array X/Y
            if re.match(r'^x\s+y$', hl) or hl in ('x', 'y'):
                header = ['X', 'Y']
                header_idx += 1
                break
            header_idx += 1

        if not header:
            i += 1
            continue

        # Extraer filas de datos
        rows = []
        j = header_idx

        if header[0] == 'X':
            # --- ARRAY X, Y: pares de numeros ---
            while j < len(lines):
                row_line = lines[j].strip()
                if not row_line:
                    j += 1
                    continue
                if pat_tabla.search(row_line):
                    break  # Inicio de otra tabla

                # Intentar extraer par X Y
                parts = row_line.split()
                if len(parts) == 2 and all(
                    re.match(r'^-?\d+[,.]?\d*$', p) for p in parts
                ):
                    rows.append([normalize_number(parts[0]),
                                  normalize_number(parts[1])])
                elif len(parts) == 1 and re.match(r'^-?\d+[,.]?\d*$', parts[0]):
                    # X e Y en lineas separadas (layout vertical)
                    if rows and len(rows[-1]) == 1:
                        rows[-1].append(normalize_number(parts[0]))
                    else:
                        rows.append([normalize_number(parts[0])])
                else:
                    # Linea no numerica: posiblemente fin de tabla
                    if rows:
                        break
                j += 1

        elif len(header) == 4:
            # --- TABLA 4 COLUMNAS (PSS) ---
            while j < len(lines):
                row_line = lines[j].strip()
                if not row_line:
                    j += 1
                    continue
                if pat_tabla.search(row_line):
                    break

                # Intentar partir en 4 columnas con multiples espacios
                parts = re.split(r'\s{3,}', row_line)
                if len(parts) == 4:
                    rows.append(parts)
                elif len(parts) == 2:
                    # Cada columna ocupa toda la linea - acumular en pares
                    rows.append(parts + ['', ''])
                else:
                    # Intentar split generico por tabulacion o doble espacio
                    parts2 = re.split(r'\t|\s{2,}', row_line)
                    if len(parts2) >= 2:
                        while len(parts2) < 4:
                            parts2.append('')
                        rows.append(parts2[:4])
                j += 1

        else:
            # --- TABLA PARAMETRO / VALOR (2 columnas) ---
            while j < len(lines):
                row_line = lines[j].strip()
                if not row_line:
                    j += 1
                    continue
                if pat_tabla.search(row_line):
                    break

                # Intentar split por multiples espacios: "TmedP   0,020"
                parts = re.split(r'\s{2,}|\t', row_line)
                if len(parts) >= 2:
                    param = parts[0].strip()
                    valor = parts[-1].strip()
                    # Filtrar si el param es solo un numero (fila no deseada)
                    if param and not re.match(r'^-?\d', param):
                        rows.append([param, normalize_number(valor)])
                elif len(parts) == 1:
                    # Caso: "TmedP 0,020" separado solo por un espacio
                    m = re.match(
                        r'^([A-Za-z_][A-Za-z0-9_\s]*?)\s+(-?\d[\d,\.]*)$',
                        row_line
                    )
                    if m:
                        rows.append([m.group(1).strip(),
                                      normalize_number(m.group(2))])
                j += 1

        if rows:
            found_tables.append({
                "num": table_num,
                "header": header,
                "rows": rows,
            })

        i = j  # continuar desde donde quedo el parser

    return found_tables


# =================================================================
# EXTRACCION PRINCIPAL
# =================================================================

def extract_metadata_from_text(text):
    """Extrae numero de documento y unidad del texto."""
    doc_match = re.search(r'(F\d+[-\d\w]+)', text)
    unit_match = re.search(r'(BOT\d{2}|UGH\d{2})', text, re.IGNORECASE)
    return {
        "documento": doc_match.group(1) if doc_match else "",
        "unidad": unit_match.group(1).upper() if unit_match else "",
        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


def clean_dataframe(df):
    """Limpia el DataFrame eliminando filas/columnas completamente vacias."""
    df = df.apply(lambda x: x.astype(str).str.strip() if x.dtype == "object" else x)
    df = df.replace("", pd.NA).dropna(how="all").fillna("")
    df = df.loc[:, (df != "").any(axis=0)]
    return df.reset_index(drop=True)


def extract_tables_from_pdf(pdf_path):
    """
    Extrae todas las tablas del PDF usando pypdfium2 para texto
    y pdfplumber como respaldo para tablas con bordes.
    Retorna (tables_dict, metadata).
    """
    tables = {}
    metadata = {"documento": "", "unidad": "", "fecha": datetime.now().strftime("%Y-%m-%d %H:%M")}

    print(f"\n📄 Procesando: {pdf_path}")
    print("=" * 70)

    num_pages = get_page_count_pypdfium(pdf_path)
    if num_pages == 0:
        print("⚠️  El PDF no contiene paginas legibles.")
        return {}, metadata

    print(f"   Total paginas: {num_pages}")
    print(f"\n🔍 Extrayendo texto con pypdfium2...")

    all_text_found = False

    for page_idx in range(num_pages):
        # 1. Extraer texto con pypdfium2
        page_text = extract_text_pypdfium(pdf_path, page_idx)

        # 2. Metadatos (solo primera vez)
        if not metadata["documento"] and page_text:
            meta = extract_metadata_from_text(page_text)
            if meta["documento"]:
                metadata.update(meta)

        # 3. Si pypdfium2 no extrae nada, intentar pdfplumber
        if not page_text.strip():
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    page_text = pdf.pages[page_idx].extract_text() or ""
            except Exception:
                pass

        if not page_text.strip():
            print(f"   ⚠️  Pag {page_idx+1}: sin texto (PDF escaneado). Ver nota al final.")
            continue

        all_text_found = True

        # 4. Parsear tablas desde el texto extraido
        parsed = parse_tables_from_text(page_text)

        if not parsed:
            # Intentar pdfplumber como fallback para tablas con bordes
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    plumber_page = pdf.pages[page_idx]
                    raw_tables = plumber_page.extract_tables({
                        "vertical_strategy": "lines",
                        "horizontal_strategy": "lines",
                        "intersection_tolerance": 8,
                        "snap_tolerance": 5,
                    })
                    if not raw_tables:
                        raw_tables = plumber_page.extract_tables({
                            "vertical_strategy": "text",
                            "horizontal_strategy": "text",
                            "intersection_tolerance": 5,
                        })
                    for ti, rt in enumerate(raw_tables or []):
                        if rt and len(rt) >= 2:
                            df = pd.DataFrame(rt)
                            df = clean_dataframe(df)
                            if len(df) >= 2:
                                table_key = f"Tabla P{page_idx+1}T{ti+1}"
                                desc, model = identify_table_description(page_text)
                                tables[table_key] = {
                                    "data": df,
                                    "num": f"P{page_idx+1}T{ti+1}",
                                    "description": desc,
                                    "model": model,
                                    "page": page_idx + 1,
                                    "is_array": False,
                                }
                                print(f"   ✅ {table_key} (pdfplumber): {desc} ({len(df)} filas)")
            except Exception as e:
                print(f"      [!] pdfplumber fallback error pag {page_idx+1}: {e}")
            continue

        # 5. Registrar tablas parseadas
        desc, model = identify_table_description(page_text)

        for tbl in parsed:
            table_num = tbl["num"]
            table_key = f"Tabla {table_num}"

            # Evitar duplicados (mantener el que tenga mas filas)
            if table_key in tables:
                if len(tbl["rows"]) <= len(tables[table_key]["data"]) - 1:
                    continue

            header = tbl["header"]
            rows = tbl["rows"]

            # Normalizar largo de filas al numero de columnas del header
            n_cols = len(header)
            rows_norm = [r + [''] * (n_cols - len(r)) if len(r) < n_cols else r[:n_cols]
                         for r in rows]

            df = pd.DataFrame(rows_norm, columns=header)
            df = clean_dataframe(df)

            is_array = header[0] in ('X', 'Y')
            # Refinar descripcion si la pagina menciona keywords especificos
            desc_page, model_page = identify_table_description(page_text)

            tables[table_key] = {
                "data": df,
                "num": table_num,
                "description": desc_page,
                "model": model_page,
                "page": page_idx + 1,
                "is_array": is_array,
            }
            print(f"   ✅ {table_key}: {desc_page} ({len(df)} filas) - Pag {page_idx+1}")

    if not all_text_found:
        print("\n⚠️  NOTA: El PDF parece ser escaneado (sin texto embebido).")
        print("   Para procesar PDFs escaneados instale Tesseract OCR:")
        print("   https://github.com/UB-Mannheim/tesseract/wiki")

    print(f"\n📊 Total tablas extraidas: {len(tables)}")
    return tables, metadata


# =================================================================
# GENERACION EXCEL
# =================================================================

def create_sheet_name(table_key, description, max_length=31):
    """Crea nombre de hoja valido para Excel."""
    num_match = re.search(r'\d+', str(table_key))
    num = num_match.group() if num_match else "XX"

    desc_short = (description
                  .replace("Datos del ", "").replace("Datos de la ", "")
                  .replace("Limitador ", "Lim ").replace("Control de ", "Ctrl ")
                  .replace("Control del ", "Ctrl "))

    name = f"Tabla {num} - {desc_short}"
    for ch in [":", "/", "\\", "?", "*", "[", "]"]:
        name = name.replace(ch, "-")
    return name[:max_length - 3] + "..." if len(name) > max_length else name


def generate_excel(tables, metadata, output_path):
    """Genera el archivo Excel con todas las tablas extraidas."""
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font  = Font(bold=True, size=14)
    sub_font    = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    # --- Hoja INDICE ---
    ws = wb.active
    ws.title = "INDICE"
    ws["A1"] = f"INDICE DE TABLAS - {metadata['unidad']}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")
    ws["A2"] = f"Documento: {metadata['documento']}"
    ws["A3"] = f"Fecha extraccion: {metadata['fecha']}"
    ws["A4"] = f"Total tablas: {len(tables)}"

    for col, h in enumerate(["Tabla (PDF)", "Descripcion", "Modelo", "Pagina"], 1):
        c = ws.cell(row=6, column=col, value=h)
        c.font = header_font; c.fill = header_fill; c.border = thin_border

    row = 7
    for name, info in sorted(tables.items(), key=lambda x: str(x[1]["num"])):
        ws.cell(row=row, column=1, value=name).border = thin_border
        ws.cell(row=row, column=2, value=info["description"]).border = thin_border
        ws.cell(row=row, column=3, value=info["model"]).border = thin_border
        ws.cell(row=row, column=4, value=info["page"]).border = thin_border
        row += 1

    for col, w in zip(["A","B","C","D"], [15, 35, 25, 10]):
        ws.column_dimensions[col].width = w

    # --- Hoja Datos_Modelado (solo tablas de parametros, no arrays) ---
    ws_data = wb.create_sheet("Datos_Modelado")
    for col, h in enumerate(["Tabla_PDF", "Parametro", "Valor", "Clave"], 1):
        c = ws_data.cell(row=1, column=col, value=h)
        c.font = header_font; c.fill = header_fill; c.border = thin_border

    row = 2
    for table_name, info in sorted(tables.items(), key=lambda x: str(x[1]["num"])):
        if info["is_array"]:
            continue
        df = info["data"]
        skip_words = {"parametro", "valor", "x", "y", ""}
        for idx in range(len(df)):
            param = str(df.iloc[idx, 0]) if df.shape[1] > 0 else ""
            valor = str(df.iloc[idx, 1]) if df.shape[1] > 1 else ""
            if param.lower() in skip_words:
                continue
            ws_data.cell(row=row, column=1, value=table_name).border = thin_border
            ws_data.cell(row=row, column=2, value=param).border = thin_border
            ws_data.cell(row=row, column=3, value=valor).border = thin_border
            ws_data.cell(row=row, column=4, value=f"{table_name}|{param}").border = thin_border
            row += 1

    for col, w in zip(["A","B","C","D"], [15, 25, 15, 35]):
        ws_data.column_dimensions[col].width = w

    # --- Hojas individuales por tabla ---
    used_names = set()
    for table_name, info in sorted(tables.items(), key=lambda x: str(x[1]["num"])):
        sheet_name = create_sheet_name(table_name, info["description"])
        if sheet_name in used_names:
            sheet_name = sheet_name[:27] + f"_{len(used_names)}"
        used_names.add(sheet_name)

        ws_t = wb.create_sheet(sheet_name)
        ws_t["A1"] = f"{table_name} - {info['description']}"
        ws_t["A1"].font = sub_font
        n_cols = len(info["data"].columns)
        if n_cols > 1:
            ws_t.merge_cells(f"A1:{chr(64+n_cols)}1")
        ws_t["A2"] = f"Modelo: {info['model']} | Pagina PDF: {info['page']}"

        df = info["data"]
        start_row = 4
        for r_idx, data_row in enumerate(dataframe_to_rows(df, index=False, header=False)):
            for c_idx, val in enumerate(data_row, 1):
                cell = ws_t.cell(row=start_row + r_idx, column=c_idx, value=val)
                cell.border = thin_border
                if r_idx == 0:
                    cell.font = header_font
                    cell.fill = header_fill

        for col_idx in range(1, n_cols + 1):
            ws_t.column_dimensions[chr(64 + col_idx)].width = 20

    wb.save(output_path)
    print(f"\n💾 Archivo guardado: {output_path}")
    print(f"   📄 Hojas creadas: {len(wb.sheetnames)}")


# =================================================================
# MENU INTERACTIVO
# =================================================================

def main():
    print("=" * 70)
    print("  EXTRACTOR DE TABLAS PES - Version 3.0 (pypdfium2)")
    print("=" * 70)

    BASE_PATH_ROOT = r"C:\Datos Cobee\03_DATOS GEN"
    MODELADO_SUBDIR = "02_MODELADO"
    PDF_PREFIX = "Tablas PES"

    if not os.path.isdir(BASE_PATH_ROOT):
        print(f"\n❌ Error: Ruta base no existe: {BASE_PATH_ROOT}")
        sys.exit(1)

    # Buscar centrales con unidades BOTxx
    centrales = sorted([
        f for f in os.listdir(BASE_PATH_ROOT)
        if os.path.isdir(os.path.join(BASE_PATH_ROOT, f))
        and any(re.match(r'BOT\d{2}', u, re.I)
                for u in os.listdir(os.path.join(BASE_PATH_ROOT, f))
                if os.path.isdir(os.path.join(BASE_PATH_ROOT, f, u)))
    ])

    if not centrales:
        print(f"\n❌ No se encontraron centrales con unidades BOTxx en '{BASE_PATH_ROOT}'")
        sys.exit(1)

    print("\nCentral:")
    for i, c in enumerate(centrales, 1):
        print(f"  {i}. {c}")

    while True:
        try:
            idx = int(input("Seleccione: ").strip()) - 1
            if 0 <= idx < len(centrales):
                central = centrales[idx]; break
        except ValueError:
            pass
        print("Seleccion invalida.")

    path_central = os.path.join(BASE_PATH_ROOT, central)
    unidades = sorted([
        u for u in os.listdir(path_central)
        if re.match(r'BOT\d{2}', u, re.I)
        and os.path.isfile(os.path.join(path_central, u, MODELADO_SUBDIR,
                                        f"{PDF_PREFIX} {u}.pdf"))
    ])

    if not unidades:
        print(f"\n❌ No se encontraron PDFs en '{path_central}'")
        sys.exit(1)

    print(f"\nUnidad en '{central}':")
    for i, u in enumerate(unidades, 1):
        print(f"  {i}. {u}")

    while True:
        try:
            idx = int(input("Seleccione: ").strip()) - 1
            if 0 <= idx < len(unidades):
                unit = unidades[idx]; break
        except ValueError:
            pass
        print("Seleccion invalida.")

    pdf_path = os.path.join(path_central, unit, MODELADO_SUBDIR,
                            f"{PDF_PREFIX} {unit}.pdf")
    output_path = os.path.join(path_central, unit, MODELADO_SUBDIR,
                               f"Tablas PES {unit}_EXTRAIDO.xlsx")

    if not os.path.exists(pdf_path):
        print(f"\n❌ No existe: {pdf_path}")
        sys.exit(1)

    try:
        tables, metadata = extract_tables_from_pdf(pdf_path)

        if not tables:
            print("\n⚠️  No se encontraron tablas en el PDF.")
            print("   Posibles causas:")
            print("   1. El PDF es escaneado (sin texto embebido) -> instale Tesseract")
            print("   2. El formato de las tablas no coincide con los patrones esperados")
            sys.exit(1)

        generate_excel(tables, metadata, output_path)

        print("\n" + "=" * 70)
        print("  ✅ EXTRACCION COMPLETADA")
        print("=" * 70)
        print(f"  PDF:     {pdf_path}")
        print(f"  Unidad:  {metadata['unidad']}")
        print(f"  Tablas:  {len(tables)}")
        print(f"  Excel:   {output_path}")

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback; traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
