# =============================================================================
# SCRIPT DE EXTRACCIÓN DE TABLAS - PDF COMPOSITE MODEL REIVAX
# =============================================================================

# Propósito: Extraer tablas del documento F19044-08-02-03-02-03-R2
# Dependencias: pip install pdfplumber pandas openpyxl
# =============================================================================

import pdfplumber
import pandas as pd
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURACIÓN - MODIFICAR SEGÚN SU UBICACIÓN
# =============================================================================
RUTA_PDF = r"C:\Datos Cobee\03_DATOS GEN\03_BOT\BOT03\02_MODELADO\TABLAS COMPOSITE MODEL.pdf"
RUTA_EXCEL_SALIDA = r"C:\Datos Cobee\03_DATOS GEN\03_BOT\BOT03\Tablas_Extraidas.xlsx"

# Definición de las tablas a extraer
TABLAS_CONFIG = {
    "Tabla 81": {"nombre": "Control Velocidad", "descripcion": "Datos del control de velocidad (GOV_REIVAX)", "tipo": "parametro_valor"},
    "Tabla 82": {"nombre": "Array K", "descripcion": "Array K - Control de velocidad", "tipo": "array_xy"},
    "Tabla 83": {"nombre": "Actuador", "descripcion": "Datos del actuador (PELTON_ATUADOR_SIMP)", "tipo": "parametro_valor"},
    "Tabla 84": {"nombre": "Control Actuador", "descripcion": "Datos del control del actuador", "tipo": "parametro_valor"},
    "Tabla 85": {"nombre": "Array K (2)", "descripcion": "Array K - Control del actuador", "tipo": "array_xy"},
    "Tabla 86": {"nombre": "Conducto Turbina", "descripcion": "Datos del conducto y turbina", "tipo": "parametro_valor"},
    "Tabla 87": {"nombre": "Array Px", "descripcion": "Array Px - Conducto y turbina", "tipo": "array_xy"},
    "Tabla 88": {"nombre": "Array YAYD", "descripcion": "Array YAYD - Conducto y turbina", "tipo": "array_xy"},
    "Tabla 89": {"nombre": "Control Tensión", "descripcion": "Datos del control de Tensión (AVR)", "tipo": "parametro_valor"},
    "Tabla 90": {"nombre": "Drive", "descripcion": "Datos del Drive", "tipo": "parametro_valor"},
    "Tabla 91": {"nombre": "Excitatriz", "descripcion": "Datos de la Excitatriz", "tipo": "parametro_valor"},
    "Tabla 92": {"nombre": "PSS", "descripcion": "Datos del PSS (PSS_COMP)", "tipo": "parametro_valor_multi"},
    "Tabla 93": {"nombre": "Limitador VHZ", "descripcion": "Datos del Limitador VHZ (VHZL)", "tipo": "parametro_valor"},
    "Tabla 94": {"nombre": "Limitador UEL", "descripcion": "Datos del Limitador UEL", "tipo": "parametro_valor"},
    "Tabla 95": {"nombre": "Array YAYD UEL", "descripcion": "Array YAYD - Limitador UEL", "tipo": "array_xy"},
    "Tabla 96": {"nombre": "Limitador OEL", "descripcion": "Datos del Limitador OEL", "tipo": "parametro_valor"},
    "Tabla 97": {"nombre": "Limitador SCL", "descripcion": "Datos del Limitador SCL", "tipo": "parametro_valor_multi"},
    "Tabla 98": {"nombre": "Limitador MEL", "descripcion": "Datos del Limitador MEL", "tipo": "parametro_valor"}
}

# Estilos para Excel
FILL_AZUL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_AZUL_CLARO = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
FILL_VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FONT_BLANCO = Font(color="FFFFFF", bold=True)
FONT_NEGRO_BOLD = Font(bold=True)

# =============================================================================
# FUNCIONES DE EXTRACCIÓN
# =============================================================================

def extraer_tablas_pdf(ruta_pdf):
    """Extrae todas las tablas del PDF usando pdfplumber"""
    tablas = []
    with pdfplumber.open(ruta_pdf) as pdf:
        for num_pagina, pagina in enumerate(pdf.pages, 1):
            tablas_pagina = pagina.extract_tables()
            for tabla in tablas_pagina:
                if tabla:  # Si la tabla no está vacía
                    tablas.append({
                        "pagina": num_pagina,
                        "datos": tabla
                    })
    return tablas


def limpiar_valor(valor):
    """Limpia y convierte un valor a número si es posible"""
    if valor is None:
        return None
    valor_str = str(valor).strip()
    if not valor_str:
        return None
    try:
        # Reemplazar coma por punto para decimales
        valor_limpio = valor_str.replace(",", ".")
        return float(valor_limpio)
    except ValueError:
        return valor_str


def procesar_tabla_parametro_valor(tabla_raw):
    """Procesa una tabla de tipo Parámetro | Valor"""
    datos = []
    for fila in tabla_raw:
        if fila and len(fila) >= 2:
            param = fila[0]
            valor = limpiar_valor(fila[1])
            if param and param not in ["Parámetro", "Parameter"]:
                datos.append({"Parámetro": param, "Valor": valor})
    return pd.DataFrame(datos)


def procesar_tabla_array_xy(tabla_raw):
    """Procesa una tabla de tipo Array X-Y"""
    datos = []
    for fila in tabla_raw:
        if fila and len(fila) >= 2:
            x = limpiar_valor(fila[0])
            y = limpiar_valor(fila[1])
            if x is not None and y is not None and str(fila[0]) not in ["X", "x"]:
                datos.append({"X": x, "Y": y})
    return pd.DataFrame(datos)


def procesar_tabla_multicolumna(tabla_raw):
    """Procesa una tabla con múltiples columnas parámetro-valor"""
    datos = []
    for fila in tabla_raw:
        if fila:
            # Procesar pares de columnas (param, valor, param, valor, ...)
            for i in range(0, len(fila) - 1, 2):
                param = fila[i]
                valor = limpiar_valor(fila[i + 1]) if i + 1 < len(fila) else None
                if param and param not in ["Parámetro", "Parameter"] and valor is not None:
                    datos.append({"Parámetro": param, "Valor": valor})
    return pd.DataFrame(datos)


def identificar_tabla_por_contenido(tabla_raw, tablas_config):
    """Identifica a qué tabla pertenece basándose en el contenido"""
    # Buscar identificadores en la primera fila o título
    texto_tabla = str(tabla_raw).lower()
    
    for tabla_id, config in tablas_config.items():
        numero = tabla_id.split()[1]
        if f"tabla {numero}" in texto_tabla or f"table {numero}" in texto_tabla:
            return tabla_id, config
    return None, None


# =============================================================================
# FUNCIONES DE ESCRITURA A EXCEL
# =============================================================================

def aplicar_formato_encabezado(ws, fila, col_inicio, col_fin):
    """Aplica formato de encabezado azul"""
    for col in range(col_inicio, col_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.fill = FILL_AZUL
        celda.font = FONT_BLANCO
        celda.alignment = Alignment(horizontal="center")


def crear_hoja_indice(wb, tablas_config):
    """Crea la hoja de índice"""
    ws = wb.create_sheet(title="ÍNDICE", index=0)
    
    ws.merge_cells("A1:B1")
    ws["A1"] = "ÍNDICE DE TABLAS - COMPOSITE MODEL (REIVAX)"
    aplicar_formato_encabezado(ws, 1, 1, 2)
    
    ws.merge_cells("A2:B2")
    ws["A2"] = "Documento: F19044-08-02-03-02-03-R2"
    
    ws["A4"] = "Tabla"
    ws["B4"] = "Descripción"
    aplicar_formato_encabezado(ws, 4, 1, 2)
    
    fila = 5
    for tabla_id, config in tablas_config.items():
        ws.cell(row=fila, column=1, value=tabla_id)
        ws.cell(row=fila, column=2, value=config["descripcion"])
        fila += 1
    
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 55
    return ws


def crear_hoja_datos_modelado(wb, tablas_procesadas):
    """Crea la hoja consolidada Datos_Modelado"""
    ws = wb.create_sheet(title="Datos_Modelado", index=1)
    
    encabezados = ["Tabla", "Parámetro", "Valor_Fuente", "Clave"]
    for col_idx, enc in enumerate(encabezados, 1):
        ws.cell(row=1, column=col_idx, value=enc)
    aplicar_formato_encabezado(ws, 1, 1, len(encabezados))
    
    fila_actual = 2
    for nombre_tabla, df in tablas_procesadas.items():
        if "Parámetro" in df.columns and "Valor" in df.columns:
            for _, row in df.iterrows():
                ws.cell(row=fila_actual, column=1, value=nombre_tabla)
                ws.cell(row=fila_actual, column=2, value=row["Parámetro"])
                ws.cell(row=fila_actual, column=3, value=row["Valor"])
                clave = f"{nombre_tabla}|{row['Parámetro']}"
                ws.cell(row=fila_actual, column=4, value=clave)
                fila_actual += 1
    
    # Ajustar columnas
    for col in ["A", "B", "C", "D"]:
        ws.column_dimensions[col].width = 20
    return ws


def crear_hoja_tabla(wb, tabla_id, config, df):
    """Crea una hoja individual para cada tabla"""
    nombre_hoja = f"{tabla_id} - {config['nombre']}"[:31]
    ws = wb.create_sheet(title=nombre_hoja)
    
    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    ws["A1"] = f"{tabla_id} - {config['descripcion']}"
    aplicar_formato_encabezado(ws, 1, 1, len(df.columns))
    
    # Encabezados de columnas
    for col_idx, columna in enumerate(df.columns, 1):
        ws.cell(row=2, column=col_idx, value=columna)
    aplicar_formato_encabezado(ws, 2, 1, len(df.columns))
    
    # Datos
    for row_idx, fila in enumerate(df.values, 3):
        for col_idx, valor in enumerate(fila, 1):
            ws.cell(row=row_idx, column=col_idx, value=valor)
    
    # Ajustar columnas
    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    return ws


# =============================================================================
# FUNCIÓN PRINCIPAL
# =============================================================================

def main():
    """Función principal de extracción"""
    print("=" * 70)
    print("       EXTRACCIÓN DE TABLAS - COMPOSITE MODEL REIVAX")
    print("=" * 70)
    
    # 1. Extraer tablas del PDF
    print("\n[1/4] Extrayendo tablas del PDF...")
    tablas_raw = extraer_tablas_pdf(RUTA_PDF)
    print(f"      ✓ Encontradas {len(tablas_raw)} tablas en el PDF")
    
    # 2. Procesar cada tabla
    print("\n[2/4] Procesando y clasificando tablas...")
    tablas_procesadas = {}
    
    for tabla_info in tablas_raw:
        tabla_datos = tabla_info["datos"]
        tabla_id, config = identificar_tabla_por_contenido(tabla_datos, TABLAS_CONFIG)
        
        if tabla_id and config:
            tipo = config["tipo"]
            
            if tipo == "parametro_valor":
                df = procesar_tabla_parametro_valor(tabla_datos)
            elif tipo == "array_xy":
                df = procesar_tabla_array_xy(tabla_datos)
            elif tipo == "parametro_valor_multi":
                df = procesar_tabla_multicolumna(tabla_datos)
            else:
                df = pd.DataFrame()
            
            if not df.empty:
                tablas_procesadas[tabla_id] = df
                print(f"      ✓ {tabla_id}: {len(df)} registros")
    
    # 3. Crear archivo Excel
    print("\n[3/4] Creando archivo Excel con formato...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    crear_hoja_indice(wb, TABLAS_CONFIG)
    crear_hoja_datos_modelado(wb, tablas_procesadas)
    
    for tabla_id, df in tablas_procesadas.items():
        config = TABLAS_CONFIG.get(tabla_id, {})
        if config:
            crear_hoja_tabla(wb, tabla_id, config, df)
    
    # 4. Guardar
    print("\n[4/4] Guardando archivo...")
    wb.save(RUTA_EXCEL_SALIDA)
    print(f"      ✓ Guardado en: {RUTA_EXCEL_SALIDA}")
    
    print("\n" + "=" * 70)
    print("       ✅ EXTRACCIÓN COMPLETADA EXITOSAMENTE")
    print("=" * 70)
    print(f"\n📊 Total de tablas procesadas: {len(tablas_procesadas)}")
    total_params = sum(len(df) for df in tablas_procesadas.values())
    print(f"📝 Total de parámetros extraídos: {total_params}")
    
    return tablas_procesadas


# =============================================================================
# EJECUCIÓN DEL SCRIPT
# =============================================================================

if __name__ == "__main__":
    tablas = main()
