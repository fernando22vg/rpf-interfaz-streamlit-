#!/usr/bin/env python3
"""
llenar_excel_DSL.py
════════════════════
Rellena la columna G (Valor) del archivo BOT0X_DSL_Diseño_Final.xlsx
con valores leídos desde VARSPRS.DAT, usando VARSPRS_DAT_vs_DSL.csv
como tabla de equivalencias.

Uso:
    python llenar_excel_DSL.py --unidad BOT03
    python llenar_excel_DSL.py --unidad BOT03 --dry-run
    python llenar_excel_DSL.py --unidad BOT03 --backup

Requiere: openpyxl  (pip install openpyxl)
"""

import re
import csv
import sys
import shutil
import argparse
from pathlib import Path
from datetime import datetime

# ── Rutas base ────────────────────────────────────────────────────────────────
BASE_DAT   = Path(r"C:\Users\jose.lozano\Downloads\05_PARAMETROS REGULADORES")
BASE_EXCEL = Path(r"C:\Datos Cobee\03_DATOS GEN\03_BOT")
MAPPING_CSV = Path(r"C:\Users\jose.lozano\Downloads\VARSPRS_DAT_vs_DSL.csv")
COL_VALOR = 7   # columna G (1-based en openpyxl)
COL_BLOQUE = 1  # A
COL_SUBSYS = 2  # B
COL_SIMBOL = 4  # D


# ════════════════════════════════════════════════════════════════════════════
# 1. PARSER VARSPRS.DAT
# ════════════════════════════════════════════════════════════════════════════

def parse_dat(filepath: Path) -> dict:
    """Retorna {nombre@modulo: valor} desde VARSPRS.DAT."""
    content = None
    for enc in ('utf-8', 'latin-1', 'cp1252'):
        try:
            content = filepath.read_text(encoding=enc)
            break
        except UnicodeDecodeError:
            continue
    if content is None:
        raise ValueError(f"No se pudo leer: {filepath}")

    pattern = re.compile(
        r'\[Var\s+\d+\]\s*\n'
        r'NS=(.+?)@(.+?)\s*\n'
        r'Tp=(\d+)\s*\n'
        r'V=(.+)',
        re.MULTILINE
    )
    result = {}
    for m in pattern.finditer(content):
        name, module, tp, raw = (
            m.group(1).strip(), m.group(2).strip(),
            int(m.group(3)), m.group(4).strip()
        )
        try:
            result[f"{name}@{module}"] = int(float(raw)) if int(tp) in (8, 9) else float(raw)
        except (ValueError, TypeError):
            result[f"{name}@{module}"] = raw
    return result


# ════════════════════════════════════════════════════════════════════════════
# 2. TABLA DE EQUIVALENCIAS CSV
# ════════════════════════════════════════════════════════════════════════════

_NO_DSL = {'', '—', '-', 'â', '→'}

def load_mapping(csv_path: Path) -> list:
    for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
        try:
            with open(csv_path, encoding=enc) as f:
                return list(csv.DictReader(f, delimiter=';'))
        except UnicodeDecodeError:
            continue
    raise ValueError(f"No se pudo leer: {csv_path}")


def extract_tabla_key(s: str) -> str:
    """'Tabla 81 — Control ...' → 'Tabla 81';  'Generador' → 'Generador'."""
    if not s:
        return ''
    m = re.search(r'Tabla\s+\d+', s)
    if m:
        return m.group(0)
    if 'generador' in s.lower():
        return 'Generador'
    return ''


def build_lookup(csv_rows: list, dat_values: dict):
    """
    Construye:
      lookup[(symbol_lower, tabla_key)] = dat_nombre
      sym_unique[symbol_lower] = dat_nombre   (solo si único candidato)
    Último en el CSV gana → prioriza módulos específicos sobre SYS_LOCAL.
    """
    lookup: dict[tuple, str] = {}
    sym_all: dict[str, list] = {}

    for row in csv_rows:
        nombre_dat = row.get('Nombre DAT', '').strip()
        simbolo    = row.get('Simbolo DSL equivalente', '').strip()
        tabla_raw  = row.get('Tabla DSL', '').strip()

        if simbolo in _NO_DSL or not simbolo:
            continue
        if nombre_dat not in dat_values:
            continue

        sym_lo = simbolo.lower()

        # "Tabla 96/98" → dos entradas
        for parte in re.split(r'[/,]', tabla_raw):
            tk = extract_tabla_key(parte.strip())
            lookup[(sym_lo, tk)] = nombre_dat

        sym_all.setdefault(sym_lo, [])
        if nombre_dat not in sym_all[sym_lo]:
            sym_all[sym_lo].append(nombre_dat)

    sym_unique = {s: lst[0] for s, lst in sym_all.items() if len(lst) == 1}
    return lookup, sym_unique


# ════════════════════════════════════════════════════════════════════════════
# 3. LLENADO DEL EXCEL
# ════════════════════════════════════════════════════════════════════════════

def fill_excel(ws, dat_values: dict, lookup: dict, sym_unique: dict,
               dry_run: bool) -> tuple:
    """Itera filas y rellena col G. Retorna (actualizados, sin_match)."""
    actualizados = []
    sin_match = []
    current_tabla = ''

    for row_idx in range(1, ws.max_row + 1):
        col_b = ws.cell(row_idx, COL_SUBSYS).value
        col_d = ws.cell(row_idx, COL_SIMBOL).value
        cell_g = ws.cell(row_idx, COL_VALOR)

        # Actualizar contexto de tabla cuando col B cambia
        if col_b and isinstance(col_b, str) and col_b.strip():
            current_tabla = col_b.strip()

        if not col_d or not isinstance(col_d, str) or not col_d.strip():
            continue

        simbolo = col_d.strip()
        sym_lo  = simbolo.lower()
        tk      = extract_tabla_key(current_tabla)

        # Buscar equivalencia: tabla exacta → tabla vacía → símbolo único
        dat_nombre = (
            lookup.get((sym_lo, tk))
            or lookup.get((sym_lo, ''))
            or sym_unique.get(sym_lo)
        )

        if dat_nombre is None:
            sin_match.append((row_idx, simbolo, current_tabla))
            continue

        new_val  = dat_values[dat_nombre]
        old_val  = cell_g.value

        if dry_run:
            marca = '≠' if str(old_val) != str(new_val) else '='
            print(f"  F{row_idx:3d} {marca}  {simbolo:28s}  {str(old_val):>12} → {new_val}  [{dat_nombre}]")
        else:
            cell_g.value = new_val

        actualizados.append((row_idx, simbolo, dat_nombre, old_val, new_val))

    return actualizados, sin_match


# ════════════════════════════════════════════════════════════════════════════
# 4. MAIN
# ════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='Rellena col G del Excel DSL desde VARSPRS.DAT',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  python llenar_excel_DSL.py --unidad BOT03
  python llenar_excel_DSL.py --unidad BOT03 --dry-run
  python llenar_excel_DSL.py --unidad BOT03 --backup
        """
    )
    parser.add_argument('--unidad',  required=True, help='Nombre de la unidad (ej: BOT03)')
    parser.add_argument('--dry-run', action='store_true',
                        help='Muestra cambios sin modificar el archivo')
    parser.add_argument('--backup',  action='store_true',
                        help='Crea copia .bak del Excel antes de modificar')
    args = parser.parse_args()

    unidad     = args.unidad.upper()
    dat_path   = BASE_DAT / unidad / "VARSPRS.DAT"
    excel_path = BASE_EXCEL / unidad / "02_MODELADO" / f"{unidad}_DSL_Diseño_Final.xlsx"

    print(f"\n{'═'*65}")
    print(f"  Unidad  : {unidad}")
    print(f"  DAT     : {dat_path}")
    print(f"  Excel   : {excel_path}")
    print(f"  Mapeo   : {MAPPING_CSV}")
    print(f"{'═'*65}\n")

    # Validar existencia
    for p in (dat_path, excel_path, MAPPING_CSV):
        if not p.exists():
            sys.exit(f"  ERROR: No encontrado: {p}")

    # 1) Leer VARSPRS.DAT
    print("  [1/4] Leyendo VARSPRS.DAT...")
    dat_values = parse_dat(dat_path)
    print(f"        {len(dat_values)} variables cargadas")

    # 2) Cargar CSV de equivalencias
    print("  [2/4] Cargando tabla de equivalencias...")
    csv_rows = load_mapping(MAPPING_CSV)
    print(f"        {len(csv_rows)} filas en CSV")

    # 3) Construir lookup
    print("  [3/4] Construyendo índice de equivalencias...")
    lookup, sym_unique = build_lookup(csv_rows, dat_values)
    print(f"        {len(lookup)} entradas en lookup ({len(sym_unique)} símbolos únicos)")

    # 4) Abrir Excel y rellenar
    try:
        import openpyxl
    except ImportError:
        sys.exit("  ERROR: Falta openpyxl. Instalar con: pip install openpyxl")

    print(f"\n  [4/4] {'Simulando cambios' if args.dry_run else 'Modificando Excel'}...\n")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    actualizados, sin_match = fill_excel(ws, dat_values, lookup, sym_unique, args.dry_run)

    # Guardar
    if not args.dry_run:
        if args.backup:
            ts  = datetime.now().strftime('%Y%m%d_%H%M%S')
            bak = excel_path.with_name(excel_path.stem + f'_bak_{ts}.xlsx')
            shutil.copy2(excel_path, bak)
            print(f"\n  Backup guardado: {bak.name}")

        wb.save(excel_path)
        print(f"\n  Excel guardado: {excel_path}")

    # ── Resumen ────────────────────────────────────────────────────────────
    cambios = sum(1 for *_, old, new in actualizados if str(old) != str(new))
    iguales = len(actualizados) - cambios

    print(f"\n{'═'*65}")
    print(f"  RESUMEN")
    print(f"{'─'*65}")
    print(f"  Parámetros con equivalencia DAT : {len(actualizados)}")
    print(f"    └─ Valores modificados         : {cambios}")
    print(f"    └─ Valores ya coincidían       : {iguales}")
    print(f"  Sin equivalencia en CSV          : {len(sin_match)}")

    if actualizados and not args.dry_run:
        print(f"\n  Parámetros actualizados:")
        for row_idx, sim, dat_n, old, new in actualizados:
            if str(old) != str(new):
                print(f"    F{row_idx:3d}  {sim:28s}  {str(old):>12} → {new}  [{dat_n}]")

    if sin_match:
        print(f"\n  Símbolos sin equivalencia DAT (informativo):")
        for r, s, t in sin_match[:20]:
            print(f"    F{r:3d}  {s:28s}  [{extract_tabla_key(t) or t[:30]}]")
        if len(sin_match) > 20:
            print(f"    ... y {len(sin_match)-20} más")

    print(f"{'═'*65}\n")


if __name__ == '__main__':
    main()
