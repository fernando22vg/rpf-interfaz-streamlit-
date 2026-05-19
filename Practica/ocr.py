# OCR_Extractor.py
# Extrae parámetros de tablas en PDF y escribe en Excel.
# Soporta BOT02 (tablas 63-80) y BOT03 (tablas 81-98).
# Tolerante a cambios de numeración en el PDF — mapea por orden de aparición.

import fitz
import easyocr
from PIL import Image, ImageEnhance  # ImageEnhance: mejora contraste antes de OCR
import pandas as pd
import openpyxl
import os
import re
import io
import difflib  # stdlib — fuzzy matching para errores de OCR
import numpy as np

# ============================================================================
# CONFIGURACIÓN
# ============================================================================
# NOTA: Ajusta estas rutas según el bot que estés procesando.
# Tu actualización indica:
#  - BOT03: C:\Datos Cobee\03_DATOS GEN\03_BOT\BOT03\02_MODELADO
#  - BOT02: C:\Datos Cobee\03_DATOS GEN\03_BOT\BOT02\02_MODELADO

PROYECTO = 'BOT02'  # 'BOT03' para procesar BOT03

BASE_DATOS_COBEE = r"C:\Datos Cobee\03_DATOS GEN\03_BOT"
MODELADO_SUBDIR = "02_MODELADO"

BOT02_DIR = os.path.join(BASE_DATOS_COBEE, "BOT02", MODELADO_SUBDIR)
BOT03_DIR = os.path.join(BASE_DATOS_COBEE, "BOT03", MODELADO_SUBDIR)

if PROYECTO == 'BOT03':
    DIR_MODELADO = BOT03_DIR
    PDF_PATH = os.path.join(DIR_MODELADO, "Tablas PES BOT03.pdf")
    # Si tu Excel template/plantilla está en otra ubicación, cámbialo aquí.
    EXCEL_TEMPLATE = os.path.join(DIR_MODELADO, "BOT03_DSL_Diseño_Final.xlsx")
    EXCEL_OUTPUT  = os.path.join(DIR_MODELADO, "BOT03_DSL_Diseño_Final.xlsx")
elif PROYECTO == 'BOT02':
    DIR_MODELADO = BOT02_DIR
    PDF_PATH = os.path.join(DIR_MODELADO, "Tablas PES BOT02.pdf")
    EXCEL_TEMPLATE = os.path.join(DIR_MODELADO, "BOT02_DSL_Diseño_Final.xlsx")
    EXCEL_OUTPUT  = os.path.join(DIR_MODELADO, "BOT02_DSL_Diseño_Final.xlsx")
else:
    raise ValueError(f"Proyecto '{PROYECTO}' no reconocido. Usa 'BOT02' o 'BOT03'")

OUTPUT_TXT = r"C:\Users\jose-\Downloads\extracted_text.txt"


# Parámetros por tabla en el orden exacto que aparecen en el Excel.
# El índice de la lista = posición de la tabla en el Excel (0-based).
# Si el PDF tiene distinta numeración, igual se mapea por posición.
TABLE_PARAMS_ORDERED = [
    # 0 - Control Velocidad (Tabla 63/81)
    ['TmedP', 'TmedW', 'bp', 'bt', 'Td', 'Tn', 'Kw', 'Bmuerta', 'YLim'],
    # 1 - (Tabla 64/82) — sin parámetros OCR conocidos
    [],
    # 2 - Actuador (Tabla 65/83)
    ['KH3', 'TVP2', 'KH1', 'TVP1', 'KH2', 'Quebra', 'TXFEXY2', 'TXFEXY1', 'TXFEXY12',
     'POSMIN2', 'POSMIN1', 'TXABY2', 'TXABY1', 'POSABY2', 'POSABY1', 'POSMAX2', 'POSMAXI', 'INCOPOS'],
    # 3 - Control Actuador (Tabla 66/84)
    ['KP2', 'KI2', 'KP1', 'KI1', 'TVP2', 'TVP1'],
    # 4 - (Tabla 67/85) — sin parámetros OCR conocidos
    [],
    # 5 - (Tabla 68/86) — sin parámetros OCR conocidos
    [],
    # 6 - (Tabla 69/87) — sin parámetros OCR conocidos
    [],
    # 7 - (Tabla 70/88) — sin parámetros OCR conocidos
    [],
    # 8 - Control Tensión (Tabla 71/89)
    ['Ka', 'Ti', 'Tf', 'Kpr', 'Kpa', 'Kia', 'Th', 'Tpass', 'Vmax', 'Vmin', 'Kvi', 'Kai'],
    # 9 - Drive (Tabla 72/90)
    ['Tg', 'Kr', 'Ka', 'T1', 'T2', 'T3', 'T4', 'Vmax', 'Vmin'],
    # 10 - Excitatriz (Tabla 73/91)
    ['Te', 'Ka', 'Kf', 'Tf', 'Efmax', 'Efmin', 'T1', 'T2'],
    # 11 - PSS (Tabla 74/92)
    ['Ks', 'T1', 'T2', 'T3', 'T4', 'T5', 'T6', 'Vs_max', 'Vs_min'],
    # 12 - Limitador VHZ (Tabla 75/93)
    ['Khz', 'Fk', 'T1', 'T2', 'VHzmax', 'VHzmin', 'T_vmin', 'Vmin'],
    # 13 - Limitador UEL (Tabla 76/94)
    ['REF_UEL', 'BETA', 'DT', 'HAB_CURVA', 'REF_IFD_PICO', 'Kp', 'IFHOFF', 'HAB', 'Tm', 'TIF',
     'TO_5', 'C_1', 'TO_4', 'REF_SCL_PICO', 'T_VT', 'VT_MIN', 'REF_SCL_TERM', 'C_09'],
    # 14 - Array YAYD UEL (Tabla 77/95)
    ['UVD', 'SOS', 'TDelay', 'To', 'Ktp', 'FILTER', 'Tfilter', 'DT', 'REF_HZ'],
    # 15 - Limitador OEL (Tabla 78/96)
    ['REF_IFD_TERM', 'BETA', 'DT', 'HAB_CURVA', 'REF_IFD_PICO', 'C_09', 'Kp', 'IFHOFF',
     'HAB', 'Tm', 'TIF', 'TO_5', 'C_1'],
    # 16 - Limitador SCL (Tabla 79/97)
    ['Tm', 'T_VT', 'VT_MIN', 'BETA', 'DT', 'HAB_CURVA', 'REF_SCL_PICO', 'REF_SCL_TERM',
     'T_IR', 'C_09', 'IX_MIN', 'T_IX', 'Kp'],
    # 17 - Limitador MEL (Tabla 80/98)
    ['Tm', 'TIF', 'REF_MEL', 'Kp', 'M1', 'HAB', 'IFHOFF', 'Set', 'TOFF', 'TEN',
     'LimMinMEL', 'LimMaxMEL'],
]

# Set único de todos los parámetros para fuzzy matching global
ALL_PARAMS_UNIQUE = list(dict.fromkeys(
    p for group in TABLE_PARAMS_ORDERED for p in group
))


# ============================================================================
# EXTRACCIÓN DE TEXTO
# ============================================================================

def extract_text_direct(pdf_path):
    """Intenta extracción directa de texto del PDF sin OCR.
    Retorna lista de strings por página, o None si el PDF parece escaneado."""
    doc = fitz.open(pdf_path)
    blocks = []
    for page in doc:
        blocks.append(page.get_text("text"))
    doc.close()
    total_chars = sum(len(b) for b in blocks)
    if total_chars < 200:
        return None
    return blocks


def preprocess_image(img):
    """Aumenta contraste de la imagen antes de pasarla a OCR."""
    img_gray = img.convert('L')
    return ImageEnhance.Contrast(img_gray).enhance(2.0)


def pdf_page_to_image(pdf_path, page_num, dpi=400):
    """Convierte una página del PDF a imagen PIL preprocesada para OCR."""
    doc = fitz.open(pdf_path)
    page = doc[page_num]
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    pix = page.get_pixmap(matrix=mat)
    img_bytes = pix.tobytes("png")
    doc.close()
    img = Image.open(io.BytesIO(img_bytes))
    return preprocess_image(img)


def extract_text_with_ocr(pdf_path, reader):
    """Extrae texto de todas las páginas del PDF usando EasyOCR."""
    doc = fitz.open(pdf_path)
    num_pages = len(doc)
    doc.close()
    print(f"   {num_pages} páginas encontradas")
    blocks = []
    for page_num in range(num_pages):
        print(f"   OCR página {page_num + 1}/{num_pages}...")
        img = pdf_page_to_image(pdf_path, page_num, dpi=400)
        img_array = np.array(img)
        results = reader.readtext(img_array)
        text = "\n".join(r[1] for r in results)
        blocks.append(text)
    return blocks


# ============================================================================
# PARSEO DE TABLAS
# ============================================================================

def fuzzy_match_param(text, candidates, threshold=0.82):
    """Coincidencia exacta primero, luego fuzzy. Retorna el param o None."""
    text_clean = text.strip()
    for param in candidates:
        if text_clean.lower() == param.lower():
            return param
    matches = difflib.get_close_matches(text_clean, candidates, n=1, cutoff=threshold)
    if matches:
        print(f"    [fuzzy] '{text_clean}' → '{matches[0]}'")
        return matches[0]
    return None


def preprocess_lines(raw_lines):
    """Une 'Tabla' separado de su número cuando el OCR los divide en dos líneas."""
    processed = []
    i = 0
    while i < len(raw_lines):
        line = raw_lines[i].strip()
        if re.match(r'^[Tt]abla$', line) and i + 1 < len(raw_lines):
            next_line = raw_lines[i + 1].strip()
            m = re.match(r'^(\d+)', next_line)
            if m:
                processed.append(f"Tabla {m.group(1)}")
                i += 2
                continue
        processed.append(line)
        i += 1
    return processed


def parse_tables_from_text(text_blocks):
    """Detecta CUALQUIER 'Tabla N' en el texto y agrupa parámetros bajo ella.

    - El número de tabla del PDF puede diferir del Excel: el mapeo es por posición.
    - Retorna lista ORDENADA: [(label_pdf, {params: [...]}), ...]
    """
    full_text = "\n".join(text_blocks)
    raw_lines = full_text.split('\n')
    lines = preprocess_lines(raw_lines)

    ordered_tables = []
    current_label = None
    current_params = []
    pending_param = None

    for i, line in enumerate(lines):
        line_clean = re.sub(r'[{}=|\\]', '', line).strip()

        # Detectar CUALQUIER "Tabla N"
        tabla_match = re.search(r'[Tt]abla\s*(\d+)', line_clean)
        if tabla_match:
            if current_label is not None:
                ordered_tables.append((current_label, {'params': list(current_params)}))
            current_label = f"Tabla {tabla_match.group(1)}"
            current_params = []
            pending_param = None
            print(f"\n  Detectada {current_label}...")
            continue

        if current_label is None:
            continue

        # Intentar reconocer parámetro con fuzzy matching
        matched = fuzzy_match_param(line_clean, ALL_PARAMS_UNIQUE)
        if matched:
            pending_param = matched
            continue

        # Si hay parámetro pendiente, intentar leer valor en esta línea
        if pending_param:
            value_line = line_clean.replace(',', '.')
            vm = re.fullmatch(r'[-+]?(\d+\.?\d*|\.\d+)([eE][+-]?\d+)?', value_line)
            if vm:
                try:
                    float(value_line)
                    current_params.append({'Parametro': pending_param, 'Valor': value_line})
                    print(f"    {pending_param} = {value_line}")
                    pending_param = None
                except ValueError:
                    pending_param = None
            elif line_clean and not re.search(r'[Tt]abla', line_clean):
                pending_param = None

    # Guardar la última tabla
    if current_label and current_params:
        ordered_tables.append((current_label, {'params': list(current_params)}))

    return ordered_tables


# ============================================================================
# ESCRITURA EN EXCEL
# ============================================================================

def build_excel_index(ws):
    """Escanea la hoja 'Parametros' y retorna:

    excel_sections: [(row_start, row_end, label), ...] ordenadas
    symbol_all_rows: {symbol: [row_idx, ...]} — TODOS los rows de cada símbolo
    """
    table_header_rows = []
    symbol_all_rows = {}

    for row_idx in range(1, ws.max_row + 1):
        # Buscar header de tabla en columnas A-E
        for col in range(1, 6):
            cell_val = ws.cell(row=row_idx, column=col).value
            if cell_val and re.search(r'[Tt]abla\s*\d+', str(cell_val)):
                table_header_rows.append((row_idx, str(cell_val).strip()))
                break

        # Acumular TODAS las filas de cada símbolo (col D)
        sym_val = ws.cell(row=row_idx, column=4).value
        if sym_val:
            sym = str(sym_val).strip()
            if sym:
                symbol_all_rows.setdefault(sym, []).append(row_idx)

    # Construir rangos de sección
    excel_sections = []
    for i, (row_start, label) in enumerate(table_header_rows):
        row_end = table_header_rows[i + 1][0] - 1 if i + 1 < len(table_header_rows) else ws.max_row
        excel_sections.append((row_start, row_end, label))

    return excel_sections, symbol_all_rows


def find_row_for_param(symbol, excel_section_idx, excel_sections, symbol_all_rows):
    """Elige el row correcto para un símbolo dado el índice de sección Excel.

    1. Si aparece en una sola fila → esa fila.
    2. Si aparece en varias → busca dentro del rango de la sección correspondiente.
    3. Fallback: primera fila (con aviso).
    """
    rows = symbol_all_rows.get(symbol, [])
    if not rows:
        return None
    if len(rows) == 1:
        return rows[0]

    if excel_sections and excel_section_idx < len(excel_sections):
        row_start, row_end, _ = excel_sections[excel_section_idx]
        for r in rows:
            if row_start <= r <= row_end:
                return r

    print(f"    [warn] '{symbol}' en {len(rows)} filas, usando primera (fila {rows[0]})")
    return rows[0]


def fill_excel_with_data(excel_template, excel_output, ordered_pdf_tables, excel_sections, symbol_all_rows):
    """Escribe los valores en el Excel y verifica que quedaron guardados.

    Mapeo: 1ª tabla PDF → 1ª sección Excel, 2ª → 2ª, etc. (por orden, no por número).
    """
    wb = openpyxl.load_workbook(excel_template)
    ws = wb['Parametros']

    print(f"\n  Secciones en Excel : {len(excel_sections)}")
    print(f"  Tablas en PDF      : {len(ordered_pdf_tables)}")
    if excel_sections and len(ordered_pdf_tables) != len(excel_sections):
        print("  [warn] Cantidad de tablas PDF ≠ secciones Excel. Mapeando por posición.")

    filled_count = 0
    not_found = 0
    reporte = []

    for pdf_order, (pdf_label, table_data) in enumerate(ordered_pdf_tables):
        excel_label = excel_sections[pdf_order][2] if pdf_order < len(excel_sections) else "???"
        print(f"\n  {pdf_label} (PDF) → {excel_label} (Excel), {len(table_data['params'])} parámetros")

        for item in table_data['params']:
            symbol = item['Parametro']
            value_pdf = float(item['Valor'])

            row_idx = find_row_for_param(symbol, pdf_order, excel_sections, symbol_all_rows)

            if row_idx is None:
                print(f"    ✗ NO ENCONTRADO: {symbol}")
                reporte.append({
                    'simbolo': symbol, 'tabla_pdf': pdf_label, 'tabla_excel': excel_label,
                    'valor_pdf': value_pdf, 'fila': None, 'valor_escrito': None,
                    'estado': 'NO_ENCONTRADO'
                })
                not_found += 1
                continue

            ws.cell(row=row_idx, column=7).value = value_pdf
            written = ws.cell(row=row_idx, column=7).value
            estado = 'OK' if written == value_pdf else 'ERROR_ESCRITURA'
            mark = '✓' if estado == 'OK' else '✗'
            print(f"    {mark} {symbol} = {value_pdf}  (fila {row_idx})")

            reporte.append({
                'simbolo': symbol, 'tabla_pdf': pdf_label, 'tabla_excel': excel_label,
                'valor_pdf': value_pdf, 'fila': row_idx, 'valor_escrito': written,
                'estado': estado
            })
            filled_count += 1

    wb.save(excel_output)

    # Verificación post-guardado: releer el archivo y comparar valor por valor
    print(f"\n  Verificando valores en disco...")
    wb_check = openpyxl.load_workbook(excel_output)
    ws_check = wb_check['Parametros']
    errores_disco = 0
    for r in reporte:
        if r['fila'] and r['estado'] == 'OK':
            val_disco = ws_check.cell(row=r['fila'], column=7).value
            try:
                if abs(float(val_disco) - r['valor_pdf']) > 1e-9:
                    r['estado'] = 'ERROR_DISCO'
                    errores_disco += 1
            except (TypeError, ValueError):
                r['estado'] = 'ERROR_DISCO'
                errores_disco += 1

    if errores_disco == 0:
        print("  Verificación OK — todos los valores están correctamente guardados.")
    else:
        print(f"  [ERROR] {errores_disco} valores no coinciden al releer el archivo.")

    # Reporte final
    print("\n" + "=" * 90)
    print("REPORTE DE VERIFICACIÓN")
    print("=" * 90)
    print(f"{'':2}{'Símbolo':<16}{'Tabla PDF':<12}{'Tabla Excel':<14}{'Valor':<14}{'Fila':<7}{'Estado'}")
    print("-" * 90)
    errores_total = 0
    for r in reporte:
        mark = '✓' if r['estado'] == 'OK' else '✗'
        fila_str = str(r['fila']) if r['fila'] else '-'
        print(f"{mark} {r['simbolo']:<15} {r['tabla_pdf']:<11} {r['tabla_excel']:<13} "
              f"{str(r['valor_pdf']):<13} {fila_str:<6} {r['estado']}")
        if r['estado'] != 'OK':
            errores_total += 1
    print("-" * 90)
    ok_count = len(reporte) - errores_total
    print(f"Total: {len(reporte)} | OK: {ok_count} | Errores: {errores_total} | No encontrados: {not_found}")
    print("=" * 90)
    print(f"\nExcel guardado: {excel_output}")


# ============================================================================
# MAIN
# ============================================================================

def main():
    print("=" * 65)
    print(f"  OCR Extractor — Tablas DISEÑO {PROYECTO}")
    print("=" * 65)
    print(f"  PDF  : {PDF_PATH}")
    print(f"  Excel: {EXCEL_OUTPUT}")
    print()

    # 1. Intentar extracción directa (sin OCR) — mucho más rápida y precisa
    print("1. Intentando extracción directa de texto del PDF...")
    text_blocks = extract_text_direct(PDF_PATH)
    if text_blocks:
        total = sum(len(b) for b in text_blocks)
        print(f"   PDF con texto seleccionable ({total} caracteres). Sin necesidad de OCR.")
        method = "Extracción directa"
    else:
        print("   PDF escaneado. Iniciando EasyOCR...")
        reader = easyocr.Reader(['es', 'en'], verbose=False)
        print("   EasyOCR listo. Procesando páginas...")
        text_blocks = extract_text_with_ocr(PDF_PATH, reader)
        method = "EasyOCR"

    # 2. Guardar texto extraído para diagnóstico
    with open(OUTPUT_TXT, "w", encoding="utf-8") as f:
        for i, blk in enumerate(text_blocks):
            f.write(f"\n{'=' * 60}\nPÁGINA {i + 1}\n{'=' * 60}\n{blk}\n")
    print(f"   Texto guardado para revisión: {OUTPUT_TXT}")

    # 3. Parsear tablas con detección flexible de numeración
    print(f"\n2. Parseando tablas ({method})...")
    ordered_pdf_tables = parse_tables_from_text(text_blocks)
    print(f"\n   Tablas detectadas en PDF: {len(ordered_pdf_tables)}")
    for label, data in ordered_pdf_tables:
        print(f"   {label}: {len(data['params'])} parámetros")

    if not ordered_pdf_tables:
        print("\n[ERROR] No se detectó ninguna tabla. Revisa el texto extraído:")
        print(f"  {OUTPUT_TXT}")
        return

    # 4. Leer estructura del Excel
    print("\n3. Leyendo estructura del Excel...")
    wb_tmp = openpyxl.load_workbook(EXCEL_TEMPLATE)
    ws_tmp = wb_tmp['Parametros']
    excel_sections, symbol_all_rows = build_excel_index(ws_tmp)
    wb_tmp.close()
    print(f"   Secciones de tabla en Excel : {len(excel_sections)}")
    print(f"   Símbolos únicos en columna D: {len(symbol_all_rows)}")

    # 5. Escribir en Excel y verificar
    print("\n4. Escribiendo en Excel...")
    fill_excel_with_data(EXCEL_TEMPLATE, EXCEL_OUTPUT, ordered_pdf_tables, excel_sections, symbol_all_rows)

    print("\n" + "=" * 65)
    print("  ¡Proceso completado!")
    print("=" * 65)


if __name__ == "__main__":
    main()