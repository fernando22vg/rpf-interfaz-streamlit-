# =============================================================================
# CargaCondIniciales_PF.py
# Carga condiciones iniciales (pgini/plini) en PowerFactory desde Excel
# y ejecuta Load Flow AC e inicializacion RMS.
#
# Flujo:
#   1. Seleccion interactiva de semestre y evento
#   2. Lee condiciones_iniciales_{fecha}_Ev{N}.xlsx (generado por CondInicialesPF.py)
#   3. Conecta a PowerFactory, activa proyecto y crea escenario de operacion
#   4. Restaura conectividad de red (cierra acopladores abiertos)
#   5. Asigna pgini a generadores; unidades sin despacho -> outserv=1
#      Para CNDC_proporcional: redistribuye proporcional a Pmax*ngnum del modelo PF
#   6. Reescalado de generacion (opcional): escala unidades CNDC_proporcional para
#      lograr Pgen = Pdem + Perdidas, respetando Pmin y Pmax de cada unidad
#   7. Selecciona maquina slack (ip_ctrl=1): prioridad GCH -> CAR -> WAR -> ERI
#   8. Asigna plini a cargas proporcionalmente desde el Excel; si hay que escalar
#      para igualar la demanda del evento, aplica restriccion de transformadores
#      (capacidad maxima por barra, sin sobrecargar)
#   9. Ejecuta Load Flow AC con regulacion de taps y limites reactivos activos
#  10. Si converge, ejecuta ComInc para inicializacion de simulacion RMS
#
# Entradas:
#   condiciones_iniciales_{fecha}_Ev{N}.xlsx  (generado por CondInicialesPF.py)
#   loc_names_xfo.xlsx  (para restriccion de capacidad de transformadores en cargas)
#
# Dependencias: pandas, openpyxl, PowerFactory Python API
# =============================================================================

import os, sys, glob, re
import pandas as pd

# ── Rutas fijas ───────────────────────────────────────────────────────────────
RAIZ        = r"C:\Datos del CNDC\01_INFO CNDC_RPF"
PF_BASE     = r"C:\Program Files\DIgSILENT\PowerFactory 2025 SP2"
LOC_XFO_PATH = (r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT"
                r"\Designacion de loc_name\loc_names_xfo.xlsx")

PF_PROYECTO    = "PMP_NOV25_OCT29_31102025(1)"   # nombre exacto del proyecto en PF
CASO_BASE      = "CNDC"                           # caso con la red configurada
EXCLUIR_SLACK  = {"sym_AGU02"}                    # generadores que NUNCA seran slack
XFO_PF         = 0.90    # factor Pnom(MVA) -> limite activo(MW)
# Para preservar el punto inicial asignado, la correccion post-LF queda deshabilitada por defecto.
# La slack absorbe las perdidas del Load Flow sin redistribuir potencia a otras unidades.
AJUSTAR_POST_LF = False

# Prefijos de plantas marginales (orden = prioridad de slack)
PREFIJOS_MARGINALES = ["GCH", "CAR", "WAR", "ERI"]

# ngnum manual: rellenar si la deteccion automatica falla para alguna unidad
# Clave = loc_name exacto en PF,  Valor = numero de maquinas en paralelo
NGNUM_MANUAL: dict[str, int] = {
    # "WT_QOL01_EQ": 17,
    # "PV_YVB01_EQ": 5,
}

# =============================================================================
# HELPERS
# =============================================================================

def elegir(opciones, titulo):
    print(f"\n{titulo}:")
    for i, op in enumerate(opciones, 1):
        print(f"  {i}. {op}")
    while True:
        try:
            sel = int(input("  Seleccionar numero: "))
            if 1 <= sel <= len(opciones):
                return opciones[sel - 1]
        except ValueError:
            pass
        print("  Opcion invalida, intente de nuevo.")


def _float(val, default=0.0):
    try:
        v = float(val)
        return v if not pd.isna(v) else default
    except Exception:
        return default


def separador(titulo="", ancho=60):
    if titulo:
        print(f"\n{'='*ancho}")
        print(f"  {titulo}")
        print(f"{'='*ancho}")
    else:
        print(f"{'='*ancho}")


def _agregar_pf_path(base):
    py_ver = f"3.{sys.version_info.minor}"
    py_dir = os.path.join(base, "Python")
    candidatos = [os.path.join(py_dir, py_ver)]
    if os.path.isdir(py_dir):
        for v in sorted(os.listdir(py_dir), reverse=True):
            c = os.path.join(py_dir, v)
            if c not in candidatos:
                candidatos.append(c)
    candidatos.append(base)
    for c in candidatos:
        if os.path.isfile(os.path.join(c, "powerfactory.pyd")):
            if c not in sys.path:
                sys.path.insert(0, c)
            return c
    raise FileNotFoundError(
        f"No se encontro powerfactory.pyd bajo:\n  {base}\n"
        "Verifique PF_BASE en el script.")


_agregar_pf_path(PF_BASE)

# =============================================================================
# [1] SELECCION DE SEMESTRE Y EVENTO
# =============================================================================

separador("SELECCION DE SEMESTRE Y EVENTO")

semestres = sorted(d for d in os.listdir(RAIZ)
                   if os.path.isdir(os.path.join(RAIZ, d)))
semestre  = elegir(semestres, "Semestre de estudio")

base_ev = os.path.join(RAIZ, semestre, "Análisis_todos_los_eventos")
eventos = sorted(d for d in os.listdir(base_ev)
                 if os.path.isdir(os.path.join(base_ev, d)))
evento  = elegir(eventos, "Evento")
ev_path = os.path.join(base_ev, evento)

m_ev     = re.search(r"(\d+)$", evento.strip())
n_evento = m_ev.group(1) if m_ev else evento.split()[-1]

# ── Leer p_desc y Pdem_evento desde Tabla_Eventos del semestre ────────────────
p_desc      = 0.0
Pdem_evento = 0.0
tabla_glob  = glob.glob(os.path.join(RAIZ, semestre, "Tabla_Eventos_*.xlsx"))
if tabla_glob:
    try:
        import openpyxl as _opx
        _wb = _opx.load_workbook(tabla_glob[0], data_only=True)
        _sh = _wb.active
        for fila in _sh.iter_rows(min_row=3, values_only=True):
            if fila[0] is None:
                continue
            try:
                if int(fila[0]) == int(n_evento):
                    p_desc = _float(fila[3])   # columna D = Potencia desc. [MW]
                    break
            except (ValueError, TypeError):
                pass
        print(f"  p_desc leido de Tabla_Eventos: {p_desc:.2f} MW  (Evento {n_evento})")
    except Exception as e:
        print(f"  [AVISO] No se pudo leer Tabla_Eventos: {e}")
else:
    print(f"  [AVISO] No se encontro Tabla_Eventos_*.xlsx en {os.path.join(RAIZ, semestre)}")

# ── Leer Pdem_evento desde datos_simulacion_*_2daopcion.xlsx  celda B9 ────────
# ev_path ya esta definido; B9 = fila 9, columna 2 (openpyxl es 1-indexado)
_dsim_glob = glob.glob(os.path.join(ev_path, "datos_simulacion_*_2daopcion.xlsx"))
if _dsim_glob:
    try:
        import openpyxl as _opx
        _wb_sim = _opx.load_workbook(_dsim_glob[0], data_only=True)
        _sh_sim = _wb_sim.active
        _b8_raw = _sh_sim.cell(row=8, column=2).value
        print(f"  Archivo: {os.path.basename(_dsim_glob[0])}")
        print(f"  Celda B8 (valor raw): {_b8_raw}")
        Pdem_evento = _float(_b8_raw)
        print(f"  Pdem_evento = {Pdem_evento:.2f} MW")
    except Exception as e:
        print(f"  [AVISO] No se pudo leer datos_simulacion 2daopcion: {e}")
else:
    print(f"  [AVISO] No se encontro datos_simulacion_*_2daopcion.xlsx en {ev_path}")

# =============================================================================
# [2] LEER condiciones_iniciales_*.xlsx
# =============================================================================

separador("LECTURA DE CONDICIONES INICIALES")

ci_files = glob.glob(os.path.join(ev_path, "condiciones_iniciales_*.xlsx"))
if not ci_files:
    raise FileNotFoundError(
        f"No se encontro condiciones_iniciales_*.xlsx en:\n  {ev_path}")
ci_path = ci_files[0]
print(f"Archivo fuente: {os.path.basename(ci_path)}")

# ── Resumen ───────────────────────────────────────────────────────────────────
df_res = pd.read_excel(ci_path, sheet_name="Resumen")
info   = dict(zip(df_res.iloc[:, 0].astype(str).str.strip(),
                  df_res.iloc[:, 1].astype(str).str.strip()))

fecha_h  = info.get("Fecha y hora", "")
disparo  = info.get("Disparo",      "")
hora_po  = info.get("Hora Po (cargas)", "")
hora_ev  = info.get("Hora evento (gen)", hora_po)

# Extraer STI codes del campo Disparo para identificar unidades desconectadas
_disp_str   = re.sub(r"\by\b", ",", disparo, flags=re.IGNORECASE)
sti_disparo = {x.strip() for x in _disp_str.split(",") if x.strip() and x.strip() != "nan"}

# ── pgini_GEN ─────────────────────────────────────────────────────────────────
df_pgini = pd.read_excel(ci_path, sheet_name="pgini_GEN")
df_pgini["loc_name PF"] = df_pgini["loc_name PF"].astype(str).str.strip()
df_pgini["pgini_MW"]    = df_pgini["pgini_MW"].apply(_float)
df_pgini["Fuente"]      = df_pgini.get("Fuente", pd.Series([""] * len(df_pgini))).fillna("").astype(str)
if "Generador_CNDC" in df_pgini.columns:
    df_pgini["Generador_CNDC"] = df_pgini["Generador_CNDC"].astype(str).str.strip()

print(f"  Evento N°    : {n_evento}")
print(f"  Fecha y hora : {fecha_h}")
print(f"  Disparo      : {disparo}")
print(f"  Hora gen.    : {hora_ev}  |  Hora cargas: {hora_po}")
print(f"  pgini_GEN    : {len(df_pgini)} unidades cargadas")

# ── Helper: extrae codigo STI de un loc_name PF ───────────────────────────────
def _sti_de(loc_name):
    s = re.sub(r"\(\d+\)$", "", str(loc_name).strip())
    for pref in ("sym_", "WT_", "PV-", "PV_", "sta_"):
        if s.lower().startswith(pref.lower()):
            s = s[len(pref):]
            break
    s = re.sub(r"_EQ$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"_II$", "", s, flags=re.IGNORECASE)
    return s.strip()

# Loc_names PF que corresponden a unidades del disparo
disparo_locs_pf = set(
    df_pgini.loc[
        df_pgini["loc_name PF"].apply(lambda lp: _sti_de(lp) in sti_disparo),
        "loc_name PF"
    ]
)

# ── Verificacion de potencia desconectada ─────────────────────────────────────
print(f"\n  Unidades del evento (disparo) — p_desc registrado: {p_desc:.2f} MW:")
if sti_disparo:
    filas_disp = df_pgini[df_pgini["loc_name PF"].isin(disparo_locs_pf)]
    if filas_disp.empty:
        print(f"    [AVISO] STI {sorted(sti_disparo)} no encontrados en pgini_GEN.")
    else:
        for _, r in filas_disp.iterrows():
            print(f"    {r['loc_name PF']:<22} {r['pgini_MW']:>7.2f} MW  ({r['Fuente']})  <- se sobreescribira con p_desc")
        print(f"    [NOTA] Los pgini de estas unidades se ajustaran a p_desc={p_desc:.2f} MW")
        print(f"           (proporcional a Pmax PF, se aplica en seccion [6] tras conectar a PF)")
else:
    print("    (sin unidades de disparo registradas)")

# ── plini_CAR + Perfil_MW_Dist ────────────────────────────────────────────────
df_plini  = pd.read_excel(ci_path, sheet_name="plini_CAR")
df_perfil = pd.read_excel(ci_path, sheet_name="Perfil_MW_Dist")

df_plini["loc_name PF"]  = df_plini["loc_name PF"].astype(str).str.strip()
df_plini["Distribuidor"] = df_plini["Distribuidor"].astype(str).str.strip()
df_plini["P_nom_MW"]     = df_plini["P_nom_MW"].apply(_float)
df_plini["plini_MW"]     = df_plini["plini_MW"].apply(_float)
if "Barra PF" in df_plini.columns:
    df_plini["Barra PF"] = df_plini["Barra PF"].astype(str).str.strip()

df_perfil["Distribuidor"] = df_perfil["Distribuidor"].astype(str).str.strip()
cols_hora = [c for c in df_perfil.columns if re.match(r"^\d{2}:\d{2}$", str(c))]
col_car   = hora_po if hora_po in cols_hora else (cols_hora[0] if cols_hora else None)

dict_dist_mw = {}
if col_car:
    for _, r in df_perfil.iterrows():
        dict_dist_mw[str(r["Distribuidor"]).strip()] = _float(r.get(col_car, 0))

# Demanda de referencia: usar suma de plini_MW del Excel (datos originales CNDC)
pdem_excel   = df_plini["plini_MW"].sum()
pgen_excel   = df_pgini["pgini_MW"].sum()

# Leer Pdem bloque del Resumen para mostrar
pdem_bloque  = _float(next((v for k, v in info.items() if "Pdem bloque" in k), pdem_excel))

# Fallback de Pdem_evento si no se pudo leer de datos_simulacion
if Pdem_evento <= 0.0:
    if pdem_bloque > 0.0:
        Pdem_evento = pdem_bloque
        _fuente_pdem = "pdem_bloque (Resumen condiciones_iniciales)"
    else:
        Pdem_evento = pdem_excel
        _fuente_pdem = "pdem_excel (suma plini_CAR)"
    print(f"  Pdem_evento = {Pdem_evento:.2f} MW  (fallback: {_fuente_pdem})")

separador("BALANCE DEMANDA vs DESPACHO")
print(f"  {'Concepto':<40} {'MW':>8}")
print(f"  {'-'*40} {'-'*8}")
print(f"  {'Demanda bloque  ' + hora_po + '  (CNDC original)':<40} {pdem_bloque:>8.1f}")
print(f"  {'Demanda asignada en Excel (plini_CAR)':<40} {pdem_excel:>8.1f}")
print(f"  {'-'*40} {'-'*8}")
print(f"  {'Despacho generadores (pgini_GEN)':<40} {pgen_excel:>8.1f}")
print(f"  {'Balance  (Pgen - Pdem)':<40} {pgen_excel - pdem_excel:>+8.1f}")

print(f"  plini_CAR: {len(df_plini)} cargas definidas")

# =============================================================================
# [3] CONECTAR A POWERFACTORY
# =============================================================================

separador("CONEXION A POWERFACTORY")

try:
    import powerfactory as pf
except ImportError:
    print("  [ERROR] No se pudo importar powerfactory.")
    sys.exit(1)

app = pf.GetApplication()
app.ClearOutputWindow()
app.Show()
print("  Conexion exitosa a PowerFactory")

print(f"  Activando proyecto '{PF_PROYECTO}'...")
res = app.ActivateProject(PF_PROYECTO)
if res != 0:
    raise RuntimeError(
        f"No se pudo activar '{PF_PROYECTO}' (codigo={res}).\n"
        f"Verifique PF_PROYECTO en el script.")
proyecto = app.GetActiveProject()
print(f"  Proyecto activo: {proyecto.loc_name}")

# =============================================================================
# [4] ACTIVAR CASO BASE + CREAR ESCENARIO DE OPERACION
# =============================================================================

separador("ESCENARIO DE OPERACION")

study_folder = app.GetProjectFolder("study")
todos_casos  = study_folder.GetContents("*.IntCase") or []
caso_base    = next((c for c in todos_casos if c.loc_name.strip() == CASO_BASE), None)
if caso_base is None:
    raise RuntimeError(f"No se encontro el caso base '{CASO_BASE}'.")
caso_base.Activate()
print(f"  Caso base activado: '{caso_base.loc_name}'")

scen_root = app.GetProjectFolder("scen")
if scen_root is None:
    raise RuntimeError("No se encontro la carpeta de escenarios en el proyecto.")

sem_folder = None
for obj in (scen_root.GetContents("*.IntFolder") or []):
    if obj.loc_name.strip() == semestre:
        sem_folder = obj
        break
if sem_folder is None:
    sem_folder = scen_root.CreateObject("IntFolder", semestre)
    print(f"  Carpeta creada: '{semestre}'")
else:
    print(f"  Carpeta encontrada: '{semestre}'")

nombre_escenario_base = f"Evento {n_evento}"
nombres_existentes = {obj.loc_name.strip()
                      for obj in (sem_folder.GetContents("*.IntScenario") or [])}
if nombre_escenario_base not in nombres_existentes:
    nombre_escenario = nombre_escenario_base
else:
    v = 1
    while f"{nombre_escenario_base}.v{v}" in nombres_existentes:
        v += 1
    nombre_escenario = f"{nombre_escenario_base}.v{v}"
escenario = sem_folder.CreateObject("IntScenario", nombre_escenario)
print(f"  Escenario creado: '{nombre_escenario}'")

escenario.Activate()
print(f"  Escenario activado: '{semestre} / {nombre_escenario}'")

# =============================================================================
# [4b] VERIFICAR Y RESTAURAR CONECTIVIDAD DE RED
# =============================================================================

separador("VERIFICACION Y RESTAURACION DE CONECTIVIDAD DE RED")

# IMPORTANTE: GetCalcRelevantObjects solo devuelve elementos en servicio.
# Para encontrar los que estan fuera de servicio hay que buscar en netdat.
_netdat_4b = None
try:
    _netdat_4b = app.GetProjectFolder("netdat")
except Exception:
    pass

def _get_todos(cls_filter):
    """Busca TODOS los elementos (incl. outserv=1) en la carpeta de datos de red."""
    objs = []
    try:
        if _netdat_4b:
            objs = _netdat_4b.GetContents(cls_filter, 1) or []
    except Exception:
        pass
    if not objs:
        objs = app.GetCalcRelevantObjects(cls_filter) or []
    return objs

couplers  = _get_todos("*.ElmCoup")
lineas_sw = _get_todos("*.ElmLne")
trafos_sw = app.GetCalcRelevantObjects("*.ElmTr2") or []

n_coup_act = 0
n_lne_act  = 0
n_lod_act  = 0
elementos_cerrados = []

# --- Conectores: poner en servicio y cerrar ---
for el in couplers:
    try:
        nombre = el.loc_name.strip()
        if getattr(el, "outserv", 0) == 1:
            el.outserv = 0
            n_coup_act += 1
            elementos_cerrados.append(("Acoplador (outserv→0)", nombre))
        elif getattr(el, "on_off", 1) == 0:
            el.on_off = 1
            elementos_cerrados.append(("Acoplador (on_off→1)", nombre))
    except Exception:
        pass

# --- Lineas: poner en servicio ---
for el in lineas_sw:
    try:
        if getattr(el, "outserv", 0) == 1:
            el.outserv = 0
            n_lne_act += 1
            elementos_cerrados.append(("Linea (outserv→0)", el.loc_name.strip()))
    except Exception:
        pass

# --- Cargas: poner TODAS en servicio ---
_lods_4b = _get_todos("*.ElmLod")
for el in _lods_4b:
    try:
        if getattr(el, "outserv", 0) == 1:
            el.outserv = 0
            n_lod_act += 1
    except Exception:
        pass

# --- Trafos: solo reporte ---
elementos_fuera_trafos = []
for el in trafos_sw:
    try:
        if getattr(el, "outserv", 0) == 1:
            elementos_fuera_trafos.append(el.loc_name.strip())
    except Exception:
        pass

print(f"  Elementos revisados:")
print(f"    Acopladores/Interruptores : {len(couplers)}")
print(f"    Lineas                    : {len(lineas_sw)}")
print(f"    Transformadores (2dev)    : {len(trafos_sw)}")
print(f"    Cargas                    : {len(_lods_4b)}")
print(f"  Conectores puestos en servicio : {n_coup_act}")
print(f"  Lineas puestas en servicio     : {n_lne_act}")
print(f"  Cargas puestas en servicio     : {n_lod_act}")

if elementos_cerrados:
    print(f"\n  Elementos restaurados — {len(elementos_cerrados)}:")
    for tipo, nombre in elementos_cerrados[:20]:
        print(f"    [{tipo}] {nombre}")
    if len(elementos_cerrados) > 20:
        print(f"    ... y {len(elementos_cerrados)-20} mas")
else:
    print("  Todos los elementos ya estaban en servicio.")

if elementos_fuera_trafos:
    print(f"\n  [AVISO] Transformadores FUERA DE SERVICIO (no se modifican): {len(elementos_fuera_trafos)}")
    for nombre in elementos_fuera_trafos[:10]:
        print(f"    {nombre}")
    if len(elementos_fuera_trafos) > 10:
        print(f"    ... y {len(elementos_fuera_trafos)-10} mas")

# =============================================================================
# [5] HELPERS DE POWERFACTORY (Pmax, Pmin, ngnum)
# =============================================================================

def _pf_attr(gen, attr, default=None):
    """Lee un atributo PF via COM de forma segura (GetAttribute primero)."""
    try:
        v = gen.GetAttribute(attr)
        if v is not None:
            return v
    except Exception:
        pass
    try:
        return getattr(gen, attr, default)
    except Exception:
        return default


def _get_ngnum(gen):
    """Numero de maquinas en paralelo.

    Orden de busqueda:
      1. Tabla NGNUM_MANUAL (override manual del usuario)
      2. Atributo 'ngnum' del objeto PF (ElmGenstat: parques eolicos/solares)
      3. Atributo 'c:ngnum' (valor calculado en algunos modelos)
      Para ElmSym/ElmAsm siempre retorna 1 (maquina unica).
    """
    if gen is None:
        return 1
    loc = gen.loc_name.strip()
    if loc in NGNUM_MANUAL:
        return NGNUM_MANUAL[loc]
    try:
        tipo = gen.GetClassName()
    except Exception:
        return 1
    if tipo not in ("ElmGenstat",):
        return 1
    for attr in ("ngnum", "c:ngnum"):
        raw = _pf_attr(gen, attr)
        try:
            n = int(float(raw))
            if n >= 1:
                return n
        except Exception:
            pass
    return 1


def _get_pmax_pf(gen):
    """Pmax total del generador en MW (respeta ngnum para ElmGenstat)."""
    if gen is None:
        return float("inf")
    try:
        tipo = gen.GetClassName()
    except Exception:
        tipo = ""
    p_unit = 0.0
    for attr in ("Pmax", "Pnom", "pgini"):
        raw = _pf_attr(gen, attr)
        try:
            v = float(raw)
            if v > 0:
                p_unit = v
                break
        except Exception:
            pass
    if p_unit <= 0:
        return float("inf")
    if tipo == "ElmGenstat":
        return p_unit * _get_ngnum(gen)
    return p_unit


def _get_pmin_pf(gen):
    """Pmin efectivo del generador en MW (0 para renovables/estaticos)."""
    if gen is None:
        return 0.0
    try:
        tipo = gen.GetClassName()
    except Exception:
        tipo = ""
    # Para ElmGenstat (eolico/solar) Pmin = 0 (pueden bajar a cero)
    if tipo == "ElmGenstat":
        return 0.0
    for attr in ("Pmin",):
        raw = _pf_attr(gen, attr)
        try:
            v = float(raw)
            if v >= 0:
                return v
        except Exception:
            pass
    return 0.0


def _set_outserv(gen, valor):
    """Pone en/fuera de servicio un generador y su composite model / plant model.
    Busca el composite model por c_pmod (ElmSym), c_pmod2 y c_stagen (ElmGenstat)."""
    gen.outserv = valor
    comp = None
    for attr in ("c_pmod", "c_pmod2", "c_stagen"):
        try:
            c = gen.GetAttribute(attr)
            if c is not None:
                comp = c
                break
        except Exception:
            pass
    if comp is not None:
        try:
            comp.outserv = valor
        except Exception:
            pass

# =============================================================================
# [6] ASIGNAR pgini A GENERADORES
# =============================================================================

separador("ASIGNANDO pgini A GENERADORES")

gen_dict = {}
gens = (app.GetCalcRelevantObjects("*.ElmSym")    or []) + \
       (app.GetCalcRelevantObjects("*.ElmGenstat") or [])
for gen in gens:
    gen_dict[gen.loc_name.strip()] = gen
print(f"  Generadores encontrados en PF: {len(gen_dict)}")

# Capturar pgini actual en PF ANTES de cualquier modificacion.
# Se usa en el diagnostico de mantenimiento para detectar unidades
# que estaban generando en el caso base y pasan a mantenimiento en este evento.
_pgini_previo_pf = {
    loc: float(getattr(gen, "pgini", 0) or 0)
    for loc, gen in gen_dict.items()
}

# Poner outserv=1 a generadores de PF no listados en el Excel del evento
_locs_gen_excel = set(df_pgini["loc_name PF"].str.strip())
_n_gen_fuera    = 0
_n_pmod_fuera   = 0
for _loc, _gen in gen_dict.items():
    if _loc not in _locs_gen_excel:
        try:
            _gen.outserv = 1
            _n_gen_fuera += 1
            try:
                _pmod = _gen.GetAttribute("c_pmod")
                if _pmod is not None:
                    _pmod.outserv = 1
                    _n_pmod_fuera += 1
            except Exception:
                pass
        except Exception:
            pass
print(f"  Generadores no listados -> outserv=1  : {_n_gen_fuera}")
print(f"  Plant models asociados  -> outserv=1  : {_n_pmod_fuera}")
print(f"  Generadores activos (en Excel)        : {len(_locs_gen_excel)}")

# =============================================================================
# [6_DISPARO] ASIGNACION DE POTENCIA — UNIDADES DEL DISPARO
# =============================================================================
# Los valores pgini_MW de las unidades del disparo ya fueron calculados por [6]
# (proporcional a Pmax). Esta seccion permite mantenerlos o reemplazarlos.
# Modo por defecto (Enter): mantener valores existentes, comportamiento identico
# al script original.

def _disp_verificar(idx_disp, p_desc, modo):
    """Imprime tabla de verificacion.
    Retorna (suma, dif, ok_continuar).
    En modo 'proporcional' nunca bloquea el flujo (solo informativo).
    """
    suma = df_pgini.loc[idx_disp, "pgini_MW"].sum()
    dif  = suma - p_desc
    _L   = "  " + "─" * 46
    separador("VERIFICACION POTENCIA DESCONECTADA")
    print(f"  {'p_desc registrado (evento)':<40}: {p_desc:>8.2f} MW")
    print(_L)
    for ix in idx_disp:
        print(f"  {df_pgini.loc[ix, 'loc_name PF']:<40}: "
              f"{df_pgini.loc[ix, 'pgini_MW']:>8.2f} MW")
    print(_L)
    print(f"  {'SUMA pgini unidades disparo':<40}: {suma:>8.2f} MW")
    print(f"  {'Diferencia (suma - p_desc)':<40}: {dif:>+8.2f} MW")
    print(_L)

    ok = True
    if abs(dif) >= 5.0:
        print(f"  [ERROR diferencia >= 5 MW]")
        if modo != "proporcional":
            resp = input("  ¿Continuar de todas formas? [s/N]: ").strip().lower()
            if resp != "s":
                ok = False
    elif abs(dif) >= 1.0:
        print(f"  [ADVERTENCIA diferencia >= 1 MW]")
        if modo != "proporcional":
            print("    La tesis puede justificar esta diferencia si proviene de redondeo")
            print("    en los datos fuente (DCDR/postot). Continuando...")
    else:
        print(f"  Estado: [OK]")
    return suma, dif, ok


def _disp_manual(idx_disp):
    """Modo [2]: ingreso manual por unidad. Modifica df_pgini en sitio."""
    for ix in idx_disp:
        loc      = df_pgini.loc[ix, "loc_name PF"]
        gen      = gen_dict.get(loc)
        pmax     = _get_pmax_pf(gen)
        actual   = df_pgini.loc[ix, "pgini_MW"]
        pmax_str = f"{pmax:.2f}" if pmax < float("inf") else "inf"
        while True:
            raw = input(f"  {loc}  (Pmax={pmax_str} MW)  pgini actual={actual:.2f} MW"
                        f"  ->  Nuevo valor [MW] (Enter=mantener): ").strip()
            if raw == "":
                break
            try:
                val = float(raw)
                if val < 0:
                    print(f"    [AVISO] Valor negativo no permitido.")
                    continue
                if pmax < float("inf") and val > pmax:
                    print(f"    [AVISO] {val:.2f} MW excede Pmax={pmax_str} MW. "
                          f"Ingrese un valor en [0, {pmax_str}]: ", end="")
                    continue
                df_pgini.loc[ix, "pgini_MW"] = round(val, 4)
                break
            except ValueError:
                print("    Valor invalido — ingrese un numero.")
        df_pgini.loc[ix, "Fuente"] = "disparo_manual"


def _disp_proporcional_pdesc(idx_disp, p_desc):
    """Modo [4]: escala pgini del disparo proporcional a sus valores actuales
    hasta que la suma == p_desc.  Respeta Pmax por unidad; el exceso se
    redistribuye equitativamente entre las unidades que aun tienen margen.
    Cambia Fuente a 'disparo_p_desc'."""
    suma_base = df_pgini.loc[idx_disp, "pgini_MW"].sum()

    if suma_base <= 0:
        # Sin base proporcional: reparto equitativo
        n     = len(idx_disp) or 1
        cuota = p_desc / n
        for ix in idx_disp:
            gen  = gen_dict.get(df_pgini.loc[ix, "loc_name PF"])
            pmax = _get_pmax_pf(gen)
            df_pgini.loc[ix, "pgini_MW"] = round(
                min(cuota, pmax) if pmax < float("inf") else cuota, 4)
    else:
        libres       = set(idx_disp)
        saturados_mw = 0.0
        for _ in range(50):
            idx_lib  = list(libres)
            sum_lib  = df_pgini.loc[idx_lib, "pgini_MW"].sum()
            restante = p_desc - saturados_mw
            if not idx_lib or sum_lib <= 1e-9:
                break
            factor    = restante / sum_lib
            nuevo_sat = False
            for ix in idx_lib:
                gen  = gen_dict.get(df_pgini.loc[ix, "loc_name PF"])
                pmax = _get_pmax_pf(gen)
                nuevo = round(df_pgini.loc[ix, "pgini_MW"] * factor, 4)
                if pmax < float("inf") and nuevo > pmax:
                    df_pgini.loc[ix, "pgini_MW"] = round(pmax, 4)
                    saturados_mw += pmax
                    libres.discard(ix)
                    nuevo_sat = True
                else:
                    df_pgini.loc[ix, "pgini_MW"] = nuevo
            if not nuevo_sat:
                break

    for ix in idx_disp:
        df_pgini.loc[ix, "Fuente"] = "disparo_p_desc"

    suma_final = df_pgini.loc[idx_disp, "pgini_MW"].sum()
    print(f"  Distribucion proporcional -> p_desc:")
    for ix in idx_disp:
        print(f"    {df_pgini.loc[ix, 'loc_name PF']:<26}->  "
              f"{df_pgini.loc[ix, 'pgini_MW']:.2f} MW")
    print(f"    {'SUMA':<26}->  {suma_final:.2f} MW  (objetivo={p_desc:.2f} MW, "
          f"dif={suma_final - p_desc:+.2f} MW)")


def _disp_excel(idx_disp):
    """Modo [3]: leer columna pgini_disparo_MW de la hoja pgini_GEN del Excel CI.
    Retorna el modo final aplicado ('disparo_excel' o 'disparo_manual').
    """
    col = "pgini_disparo_MW"
    if col not in df_pgini.columns:
        print(f"  [AVISO] Columna \"{col}\" no encontrada en hoja pgini_GEN.")
        print(f"          Para usarla, agréguela manualmente al Excel con los valores")
        print(f"          reales de despacho de cada unidad del disparo.")
        print(f"          Cambiando a modo [2] (ingreso manual)...")
        _disp_manual(idx_disp)
        return "disparo_manual"
    for ix in idx_disp:
        loc = df_pgini.loc[ix, "loc_name PF"]
        val = _float(df_pgini.loc[ix, col], default=float("nan"))
        if not pd.isna(val) and val >= 0:
            df_pgini.loc[ix, "pgini_MW"] = round(val, 4)
            df_pgini.loc[ix, "Fuente"]   = "disparo_excel"
            print(f"  {loc:<30} pgini_disparo_MW = {val:.4f} MW")
        else:
            fallback = df_pgini.loc[ix, "pgini_MW"]
            print(f"  [AVISO] {loc}: pgini_disparo_MW vacío — "
                  f"usando pgini_MW={fallback:.2f} MW (proporcional)")
            df_pgini.loc[ix, "Fuente"] = "disparo_excel"
    return "disparo_excel"


# ── Ejecucion del menu ────────────────────────────────────────────────────────
_modo_disparo  = "disparo_p_desc"   # valor si no hay unidades de disparo
_suma_disparo  = 0.0

if disparo_locs_pf:
    idx_disp  = df_pgini[df_pgini["loc_name PF"].isin(disparo_locs_pf)].index.tolist()
    locs_disp = [df_pgini.loc[ix, "loc_name PF"] for ix in idx_disp]
    _suma_act = df_pgini.loc[idx_disp, "pgini_MW"].sum()
    _dif_act  = _suma_act - p_desc

    # La unidad del disparo puede llegar con pgini=0 si CondInicialesPF la leyo
    # con P=0 en la hora del evento (ya habia disparado) y la marco 'mantenimiento'.
    # Pero en las condiciones iniciales (pre-evento) SI estaba generando p_desc MW.
    # Se corrige automaticamente: distribuir p_desc proporcional a Pmax como base.
    if _suma_act == 0 and p_desc > 0:
        separador("AUTO-CORRECCION UNIDADES DEL DISPARO")
        print(f"  Unidades del disparo tienen pgini=0 (marcadas como mantenimiento")
        print(f"  en la hora del evento). Corrigiendo a distribucion proporcional")
        print(f"  a Pmax para sumar p_desc = {p_desc:.2f} MW...")
        _disp_proporcional_pdesc(idx_disp, p_desc)
        for ix in idx_disp:
            loc = df_pgini.loc[ix, "loc_name PF"]
            gen = gen_dict.get(loc)
            if gen:
                gen.outserv = 0   # asegurar en servicio (no mantenimiento)
        _suma_act = df_pgini.loc[idx_disp, "pgini_MW"].sum()
        _dif_act  = _suma_act - p_desc
        print(f"  Correccion aplicada. El menu permite ajustar si es necesario.")

    while True:
        separador("ASIGNACION DE POTENCIA — UNIDADES DEL DISPARO")
        print(f"  Evento N°: {n_evento}  |  p_desc registrado: {p_desc:.2f} MW")
        print(f"  Unidades del disparo: {', '.join(locs_disp)}")
        print()
        print("  Valores actuales:")
        for ix in idx_disp:
            print(f"    {df_pgini.loc[ix, 'loc_name PF']:<26}->  "
                  f"{df_pgini.loc[ix, 'pgini_MW']:.2f} MW")
        _est = "[OK]" if abs(_dif_act) < 1.0 else (
               "[ADVERTENCIA]" if abs(_dif_act) < 5.0 else "[ERROR]")
        print(f"    {'SUMA':<26}->  {_suma_act:.2f} MW  |  "
              f"p_desc = {p_desc:.2f} MW  |  dif = {_dif_act:+.2f} MW  {_est}")
        print()
        print("  [1] Mantener valores actuales (proporcional)  <- DEFAULT")
        print("  [2] Ingreso manual por unidad")
        print('  [3] Leer desde columna "pgini_disparo_MW" del Excel')
        print("  [4] Distribuir p_desc proporcional a pgini actual (respeta Pmax)")
        separador()

        opcion = input("  Seleccionar [1/2/3/4] o Enter para mantener: ").strip()

        if opcion in ("", "1"):
            # Modo 1: no modifica nada, solo verificacion informativa
            _modo_disparo = "disparo_p_desc"
            print("  [OK] Modo disparo: proporcional (valores actuales mantenidos)")
            _suma_disparo, _, _ = _disp_verificar(idx_disp, p_desc, "proporcional")
            break

        elif opcion == "2":
            _disp_manual(idx_disp)
            _modo_disparo = "disparo_manual"

        elif opcion == "3":
            _modo_disparo = _disp_excel(idx_disp)

        elif opcion == "4":
            _disp_proporcional_pdesc(idx_disp, p_desc)
            _modo_disparo = "disparo_p_desc"

        else:
            print("  Opcion invalida — ingrese 1, 2, 3, 4 o Enter.")
            continue

        # Modos [2], [3] y [4]: actualizar PF y verificar con posibilidad de bloqueo
        for ix in idx_disp:
            loc = df_pgini.loc[ix, "loc_name PF"]
            gen = gen_dict.get(loc)
            if gen:
                gen.pgini   = df_pgini.loc[ix, "pgini_MW"]
                gen.outserv = 0

        _suma_disparo, _, _ok = _disp_verificar(idx_disp, p_desc, _modo_disparo)
        if _ok:
            break
        # Si no ok (dif >= 5 MW y usuario rechazo), volver al menu

elif p_desc > 0:
    print(f"  [AVISO] p_desc={p_desc:.2f} MW pero no se identificaron "
          f"unidades del disparo en df_pgini.")
else:
    # Sin disparo: calcular suma para el export
    _suma_disparo = df_pgini.loc[
        df_pgini["Fuente"] == "disparo_p_desc", "pgini_MW"
    ].sum()

# ── Distribuir CNDC_proporcional usando Pmax*ngnum real del modelo PF ─────────
# CondInicialesPF distribuyo proporcional a P_nom del Excel; aqui refinamos
# la distribucion usando los Pmax reales del modelo PF para las unidades con
# Fuente="CNDC_proporcional" del mismo Generador_CNDC.
mask_prop_ci = df_pgini["Fuente"] == "CNDC_proporcional"
if mask_prop_ci.any() and "Generador_CNDC" in df_pgini.columns:
    separador("REFINANDO DISTRIBUCION CNDC_proporcional CON Pmax PF")
    for cndc_grp, grupo in df_pgini[mask_prop_ci].groupby("Generador_CNDC"):
        # Suma del grupo = total CNDC ya distribuido proporcionalmente en Excel
        total_grp = grupo["pgini_MW"].sum()
        idx_list  = grupo.index.tolist()
        pmax_list = [_get_pmax_pf(gen_dict.get(df_pgini.loc[ix, "loc_name PF"]))
                     for ix in idx_list]
        sum_pmax  = sum(p for p in pmax_list if p < float("inf"))
        if sum_pmax <= 0:
            n = len(idx_list) or 1
            for ix in idx_list:
                df_pgini.loc[ix, "pgini_MW"] = round(total_grp / n, 4)
            print(f"  {cndc_grp:<28} {total_grp:.2f} MW -> igualado ({n} uds, sin Pmax PF)")
        else:
            for ix, pmax in zip(idx_list, pmax_list):
                peso  = pmax if pmax < float("inf") else sum_pmax / len(idx_list)
                parte = round(total_grp * (peso / sum_pmax), 4)
                df_pgini.loc[ix, "pgini_MW"] = parte
                loc = df_pgini.loc[ix, "loc_name PF"]
                print(f"  {cndc_grp:<28} {loc:<22} Pmax={pmax:.2f} -> pgini={parte:.4f} MW")

# ── Asignar pgini y outserv a cada generador ──────────────────────────────────
ok_gen        = 0
miss_gen      = []
mant_gen      = []
asignados_set = set()

for _idx_row, row in df_pgini.iterrows():
    loc       = row["loc_name PF"]
    pgini_val = row["pgini_MW"]
    fuente    = row["Fuente"]
    en_mant   = "mantenimiento" in fuente.lower()

    gen = gen_dict.get(loc)
    if gen is None:
        miss_gen.append(loc)
        continue

    # Clampar a [Pmin, Pmax] antes de asignar
    pmax = _get_pmax_pf(gen)
    pmin = _get_pmin_pf(gen) if pgini_val > 0 else 0.0
    if pgini_val > 0 and pmax < float("inf"):
        pgini_val = min(pgini_val, pmax)
    if pgini_val > 0:
        pgini_val = max(pgini_val, pmin)

    gen.pgini = pgini_val
    # Sincronizar df_pgini con el valor clampado para que [7b], [9d] y la
    # exportacion trabajen siempre con el mismo valor que tiene PF
    df_pgini.loc[_idx_row, "pgini_MW"] = pgini_val
    asignados_set.add(loc)

    if en_mant or pgini_val == 0.0:
        _set_outserv(gen, 1)
        if en_mant:
            mant_gen.append(loc)
    else:
        _set_outserv(gen, 0)

    ok_gen += 1

# ── Diagnostico: composite models de unidades en mantenimiento ──────────────────
# Compara el pgini que tenia cada unidad en PF ANTES de este script (caso base)
# con su estado de mantenimiento en este evento. Si pgini_previo > 0, la unidad
# estaba activa en el caso base → mayor riesgo de desbalance dinamico si su
# composite model no queda correctamente deshabilitado.
_mant_idx = df_pgini[df_pgini["Fuente"].str.contains("mantenimiento", case=False, na=False)].index
if len(_mant_idx) > 0:
    _n_mant_comp_ok = 0
    _n_mant_comp_nf = 0
    _locs_sin_comp  = []
    _n_activas      = 0

    separador("DIAGNOSTICO UNIDADES EN MANTENIMIENTO")
    _H = f"  {'loc_name PF':<26} {'pgini previo':>12}  {'Comp.model':>16}  Nota"
    print(_H)
    print("  " + "─" * 74)

    for _ix in _mant_idx:
        _loc_m      = df_pgini.loc[_ix, "loc_name PF"]
        _pgini_prev = _pgini_previo_pf.get(_loc_m, 0.0)   # pgini en PF antes del script
        _gen_m      = gen_dict.get(_loc_m)

        if _gen_m is None:
            print(f"  {_loc_m:<26} {_pgini_prev:>10.2f} MW  {'—':>16}  [no encontrado en PF]")
            continue

        _comp_found = False
        for _attr_m in ("c_pmod", "c_pmod2", "c_stagen"):
            try:
                _c = _gen_m.GetAttribute(_attr_m)
                if _c is not None:
                    _comp_found = True
                    break
            except Exception:
                pass

        _comp_str = "OK deshabilitado" if _comp_found else "NO ENCONTRADO  *"
        _nota     = "[ACTIVA en caso base]" if _pgini_prev > 0 else ""
        print(f"  {_loc_m:<26} {_pgini_prev:>10.2f} MW  {_comp_str:>16}  {_nota}")

        if _comp_found:
            _n_mant_comp_ok += 1
        else:
            _n_mant_comp_nf += 1
            _locs_sin_comp.append(_loc_m)
        if _pgini_prev > 0:
            _n_activas += 1

    print("  " + "─" * 74)
    print(f"  Total en mantenimiento                : {len(_mant_idx)}")
    print(f"  Activas en caso base (pgini > 0)      : {_n_activas}"
          + ("  <- riesgo si comp.model no deshabilitado" if _n_activas else ""))
    print(f"  Composite model encontrado y OK       : {_n_mant_comp_ok}")
    if _n_mant_comp_nf:
        print(f"  Sin composite model (*)               : {_n_mant_comp_nf}")
        for _lsc in _locs_sin_comp:
            print(f"    - {_lsc}")
        print(f"  -> Verificar en PF: modelo dinamico puede seguir activo.")

# Generadores en PF sin datos de despacho -> outserv=1
PREFIJOS_RENOVABLES = ("WT_", "PV_", "PV-")

def _es_marginal(gen):
    ln = gen.loc_name.strip().upper()
    for pref in ("SYM_", "WT_", "PV_", "PV-", "STA_"):
        if ln.startswith(pref):
            ln = ln[len(pref):]
            break
    return any(ln.startswith(p) for p in PREFIJOS_MARGINALES)

gen_no_asignados = [g for g in gens if g.loc_name.strip() not in asignados_set]
gen_no_marg      = [g for g in gen_no_asignados if _es_marginal(g)]
gen_no_fuera     = [g for g in gen_no_asignados if not _es_marginal(g)]

print(f"  Asignados            : {ok_gen} generadores")
print(f"  Mantenimiento        : {len(mant_gen)}")
if mant_gen:
    for m in mant_gen:
        print(f"    - {m}")
if miss_gen:
    print(f"  No encontrados en PF : {len(miss_gen)}")
    for m in miss_gen[:10]:
        print(f"    - {m}")
    if len(miss_gen) > 10:
        print(f"    ... y {len(miss_gen)-10} mas")

if gen_no_marg:
    print(f"\n  Marginales sin despacho ({len(gen_no_marg)}): outserv=1 (candidatos a slack)")
    for g in gen_no_marg:
        _set_outserv(g, 1)
        print(f"    {g.loc_name}")

if gen_no_fuera:
    print(f"\n  No marginales sin despacho ({len(gen_no_fuera)}): outserv=1")
    for g in gen_no_fuera:
        _set_outserv(g, 1)
    for g in gen_no_fuera[:10]:
        print(f"    {g.loc_name}")
    if len(gen_no_fuera) > 10:
        print(f"    ... y {len(gen_no_fuera)-10} mas")

# =============================================================================
# [6b] VERIFICACION Pmax vs pgini ASIGNADO
# =============================================================================

separador("VERIFICACION Pmax vs pgini ASIGNADO")

mask_en_serv = df_pgini["Fuente"].isin({"P0_medido", "CNDC_proporcional"})
print(f"\n  {'loc_name':<24} {'Tipo':<12} {'ngnum':>5} {'Pmax_tot':>9} "
      f"{'pgini_MW':>9} {'Pmin':>6}  {'':>6}")
print(f"  {'-'*24} {'-'*12} {'-'*5} {'-'*9} {'-'*9} {'-'*6}  {'-'*6}")

for _, row in df_pgini[mask_en_serv].iterrows():
    loc = row["loc_name PF"]
    gen = gen_dict.get(loc)
    if gen is None:
        continue
    try:
        tipo = gen.GetClassName()
    except Exception:
        tipo = ""
    ng     = _get_ngnum(gen)
    pmax   = _get_pmax_pf(gen)
    pmin   = _get_pmin_pf(gen)
    excel  = row["pgini_MW"]
    pmax_s = f"{pmax:.2f}" if pmax < float("inf") else "inf"
    warn   = "[WARN>Pmax]" if pmax < float("inf") and excel > pmax + 1e-4 else ""
    warn   = warn or ("[WARN<Pmin]" if excel > 0 and excel < pmin - 1e-4 else "")
    ng_s   = str(ng) if tipo == "ElmGenstat" else "—"
    print(f"  {loc:<24} {tipo:<12} {ng_s:>5} {pmax_s:>9} {excel:>9.2f} "
          f"{pmin:>6.2f}  {warn}")

mask_p0   = df_pgini["Fuente"] == "P0_medido"
mask_prop = df_pgini["Fuente"] == "CNDC_proporcional"
mask_mant = df_pgini["Fuente"].str.contains("mantenimiento", case=False)
pgen_total = df_pgini["pgini_MW"].sum()

print(f"\n  Fuente                           MW       Uds")
print(f"  {'─'*32} {'─'*8}  {'─'*4}")
print(f"  {'P0_medido':<32} {df_pgini.loc[mask_p0,'pgini_MW'].sum():>8.2f}  {mask_p0.sum():>4}")
print(f"  {'CNDC_proporcional':<32} {df_pgini.loc[mask_prop,'pgini_MW'].sum():>8.2f}  {mask_prop.sum():>4}")
print(f"  {'Mantenimiento':<32} {0.0:>8.2f}  {mask_mant.sum():>4}")
print(f"  {'─'*32} {'─'*8}  {'─'*4}")
print(f"  {'TOTAL GENERACION CNDC':<32} {pgen_total:>8.2f}")
print(f"  {'Demanda Excel (plini_CAR)':<32} {pdem_excel:>8.2f}")
print(f"  {'Diferencia (Pgen - Pdem)':<32} {pgen_total - pdem_excel:>+8.2f}  MW")

# =============================================================================
# [6c] IGUALAR GENERACION A LA DEMANDA REGISTRADA (respeta Pmin y Pmax)
# =============================================================================
# Objetivo: Pgen_total = Pdem_evento (demanda oficial del evento).
# Solo se ajustan unidades CNDC_proporcional que NO son del disparo.
# P0_medido (incluye slack), mantenimiento y disparo_p_desc son INTOCABLES.
# Algoritmo iterativo:
#   1. p_prop_obj = Pdem_evento - P_fijo  (P_fijo incluye P0_medido con slack)
#   2. Escalar proporcionalmente (manteniendo la distribucion relativa)
#   3. Clampar a [Pmin, Pmax]; redistribuir sobrante/deficit entre libres
#   4. Repetir hasta convergencia

separador("IGUALANDO GENERACION A DEMANDA REGISTRADA")

# Fuentes fijas (no se tocan — incluye P0_medido del slack)
fuentes_fijas = {"P0_medido", "sin_despacho", "mantenimiento",
                 "disparo_p_desc", "disparo_manual", "disparo_excel"}

idx_prop = df_pgini[
    ~df_pgini["Fuente"].isin(fuentes_fijas)
].index.tolist()

p_fijo   = df_pgini.loc[
    df_pgini["Fuente"].isin(fuentes_fijas),
    "pgini_MW"
].sum()
p_prop_obj = Pdem_evento - p_fijo   # MW que deben aportar las proporcionales

print(f"  Pdem_evento (objetivo oficial): {Pdem_evento:.2f} MW")
print(f"  P_fijo (P0_medido)            : {p_fijo:.2f} MW")
print(f"  Proporcionales objetivo       : {p_prop_obj:.2f} MW  ({len(idx_prop)} unidades)")

if not idx_prop:
    print("  Sin unidades proporcionales — no se ajusta generacion.")
    pgen_total = df_pgini["pgini_MW"].sum()
elif p_prop_obj <= 0:
    print("  [INFO] P_fijo >= Pdem — proporcionales se dejan en cero.")
    for ix in idx_prop:
        df_pgini.loc[ix, "pgini_MW"] = 0.0
    pgen_total = p_fijo
else:
    libres       = set(idx_prop)
    saturados_mw = 0.0
    convergio    = False

    for _ in range(50):
        idx_libres   = list(libres)
        sum_libre    = df_pgini.loc[idx_libres, "pgini_MW"].sum()
        restante_obj = p_prop_obj - saturados_mw

        if not idx_libres or restante_obj <= 1e-6:
            convergio = True
            break

        factor = (restante_obj / sum_libre) if sum_libre > 0 else 1.0
        nuevos_sat = False

        for ix in idx_libres:
            loc   = df_pgini.loc[ix, "loc_name PF"]
            gen   = gen_dict.get(loc)
            p_old = df_pgini.loc[ix, "pgini_MW"]
            p_new = (p_old * factor) if sum_libre > 0 else (restante_obj / len(idx_libres))

            pmax = _get_pmax_pf(gen)
            pmin = _get_pmin_pf(gen) if p_old > 0 else 0.0

            if pmax < float("inf") and p_new > pmax:
                df_pgini.loc[ix, "pgini_MW"] = pmax
                libres.discard(ix)
                saturados_mw += pmax
                nuevos_sat = True
            elif p_old > 0 and p_new < pmin:
                df_pgini.loc[ix, "pgini_MW"] = pmin
                libres.discard(ix)
                saturados_mw += pmin
                nuevos_sat = True
            else:
                df_pgini.loc[ix, "pgini_MW"] = round(p_new, 4)

        if not nuevos_sat:
            convergio = True
            break

    pgen_total = df_pgini["pgini_MW"].sum()
    print(f"  Resultado ({'OK' if convergio else 'no convergio'}): "
          f"Pgen_total={pgen_total:.2f} MW  |  "
          f"Diferencia con Pdem_evento={pgen_total - Pdem_evento:+.2f} MW")

    # Actualizar PF con valores reescalados
    for ix in idx_prop:
        loc       = df_pgini.loc[ix, "loc_name PF"]
        pgini_new = df_pgini.loc[ix, "pgini_MW"]
        gen = gen_dict.get(loc)
        if gen is None:
            continue
        gen.pgini = pgini_new
        if pgini_new == 0.0:
            _set_outserv(gen, 1)
        else:
            _set_outserv(gen, 0)

# =============================================================================
# [7] SELECCIONAR MAQUINA SLACK (ip_ctrl)
# =============================================================================

separador("MAQUINA DE REFERENCIA (SLACK)")

def _es_renovable(gen):
    ln = gen.loc_name.strip()
    return any(ln.upper().startswith(p.upper()) for p in PREFIJOS_RENOVABLES)

def _es_excluido(gen):
    return gen.loc_name.strip() in EXCLUIR_SLACK or _es_renovable(gen)

def _margen(gen):
    try:
        pgini_actual = float(getattr(gen, "pgini", 0) or 0)
    except Exception:
        pgini_actual = 0.0
    return max(_get_pmax_pf(gen) - pgini_actual, 0.0)

for g in gens:
    try:
        g.ip_ctrl = 0
    except Exception:
        pass

candidatos_validos = [
    g for g in gens
    if not _es_excluido(g)
    and getattr(g, "outserv", 1) == 0
    and float(getattr(g, "pgini", 0) or 0) > 0
]

print(f"  Plantas marginales (prioridad): {' -> '.join(PREFIJOS_MARGINALES)}")
print(f"  Candidatos en servicio con pgini>0: {len(candidatos_validos)}")

slack_gen  = None
slack_nota = ""

def _ln_base(g):
    ln = g.loc_name.strip().upper()
    for p in ("SYM_", "WT_", "PV_", "PV-", "STA_"):
        if ln.startswith(p):
            return ln[len(p):]
    return ln

def _es_unidad_gas(g):
    """Retorna True si el loc_name termina en '1' (gas de ciclo combinado).
    Las unidades terminadas en '0' son vapor — no aptas como slack porque
    no pueden regular frecuencia de forma independiente en el ciclo combinado."""
    ln = g.loc_name.strip()
    return ln[-1] == "1" if ln else False

def _es_unidad_vapor(g):
    """Retorna True si el loc_name termina en '0' (vapor de ciclo combinado)."""
    ln = g.loc_name.strip()
    return ln[-1] == "0" if ln else False

def _prioridad_slack(g):
    """Clave de ordenamiento para seleccion de slack dentro de una planta.
    Primero unidades gas (terminan en '1'), luego otras, luego vapor (terminan en '0').
    Desempate por mayor margen disponible."""
    if _es_unidad_gas(g):
        orden = 0
    elif _es_unidad_vapor(g):
        orden = 2
    else:
        orden = 1
    return (orden, -_margen(g))   # orden ASC, margen DESC

for prefijo in PREFIJOS_MARGINALES:
    en_planta = [g for g in candidatos_validos if _ln_base(g).startswith(prefijo)]
    if en_planta:
        # Ordenar: gas (terminan en 1) > otros > vapor (terminan en 0); desempate por margen
        en_planta.sort(key=_prioridad_slack)
        slack_gen = en_planta[0]
        fuente_u  = ("P0_medido" if slack_gen.loc_name.strip() in
                     set(df_pgini.loc[df_pgini["Fuente"] == "P0_medido", "loc_name PF"].str.strip())
                     else "proporcional")
        tipo_u    = ("gas"   if _es_unidad_gas(slack_gen)   else
                     "vapor" if _es_unidad_vapor(slack_gen) else "otro")
        slack_nota = f"planta {prefijo} | {fuente_u} | {tipo_u} | mayor margen"
        break

if slack_gen is None and candidatos_validos:
    # Fallback global: mismo criterio de prioridad
    candidatos_validos.sort(key=_prioridad_slack)
    slack_gen  = candidatos_validos[0]
    slack_nota = "fallback — mayor margen global (ninguna planta marginal disponible)"

if slack_gen is not None:
    slack_gen.ip_ctrl = 1
    pg  = float(getattr(slack_gen, "pgini", 0) or 0)
    pmx = _get_pmax_pf(slack_gen)
    ng  = _get_ngnum(slack_gen)
    ng_s = f"  ngnum={ng}" if ng > 1 else ""
    print(f"  Slack asignado : {slack_gen.loc_name}")
    print(f"    {slack_nota}")
    print(f"    pgini={pg:.2f} MW  |  Pmax_total={pmx:.2f} MW{ng_s}  |  "
          f"margen={pmx - pg:.2f} MW")
else:
    print("  [ERROR CRITICO] No se encontro ningun generador apto para slack.")

# =============================================================================
# [7b] AJUSTE FINO DE GENERACION PROPORCIONAL (post-slack, usa Pdem_evento)
# =============================================================================
# Una vez conocida la slack, excluirla explicitamente de p_fijo_7b para
# evitar doble resta. Solo se tocan unidades CNDC_proporcional != slack.
# Ecuacion: Pgen_prop = Pdem_evento - p_fijo_7b_sin_slack - slack_pgini_fijo
# Algoritmo iterativo identico al de [6c] (max 50 iteraciones, respeta Pmax/Pmin).

separador("AJUSTE FINO GENERACION PROPORCIONAL (post-slack)")

if slack_gen is not None:
    slack_loc      = slack_gen.loc_name.strip()
    slack_pgini_7b = float(getattr(slack_gen, "pgini", 0) or 0)

    fuentes_fijas_7b = {"P0_medido", "sin_despacho", "mantenimiento",
                        "disparo_p_desc", "disparo_manual", "disparo_excel"}

    # CORRECCION: excluir el slack de p_fijo_7b para evitar doble resta
    p_fijo_7b = df_pgini.loc[
        df_pgini["Fuente"].isin(fuentes_fijas_7b) &
        (df_pgini["loc_name PF"].str.strip() != slack_loc),
        "pgini_MW"
    ].sum()
    obj_prop_7b = Pdem_evento - p_fijo_7b - slack_pgini_7b

    # Indices de unidades proporcionales ajustables (excluye slack y fijas)
    idx_prop_7b = df_pgini[
        (~df_pgini["Fuente"].isin(fuentes_fijas_7b)) &
        (df_pgini["loc_name PF"].str.strip() != slack_loc)
    ].index.tolist()

    sum_prop_antes = df_pgini.loc[idx_prop_7b, "pgini_MW"].sum()
    print(f"  Pdem_evento                      : {Pdem_evento:.2f} MW")
    print(f"  P_fijo no-slack (P0_medido)      : {p_fijo_7b:.2f} MW")
    print(f"  Slack '{slack_loc}' pgini (FIJO) : {slack_pgini_7b:.2f} MW")
    print(f"  Objetivo proporcionales [7b]     : {obj_prop_7b:.2f} MW  ({len(idx_prop_7b)} unidades)")
    print(f"  Pgen proporcionales antes [7b]   : {sum_prop_antes:.2f} MW")

    if not idx_prop_7b:
        print("  Sin unidades proporcionales — no se realiza ajuste [7b].")
    elif obj_prop_7b <= 0:
        print("  [INFO] obj_prop_7b <= 0 — proporcionales se dejan en cero.")
        for ix in idx_prop_7b:
            df_pgini.loc[ix, "pgini_MW"] = 0.0
            loc = df_pgini.loc[ix, "loc_name PF"]
            gen = gen_dict.get(loc)
            if gen:
                gen.pgini = 0.0
                _set_outserv(gen, 1)
    else:
        libres_7b       = set(idx_prop_7b)
        saturados_mw_7b = 0.0
        convergio_7b    = False

        for _ in range(50):
            idx_lib_7b   = list(libres_7b)
            sum_lib_7b   = df_pgini.loc[idx_lib_7b, "pgini_MW"].sum()
            restante_7b  = obj_prop_7b - saturados_mw_7b

            if not idx_lib_7b or restante_7b <= 1e-6:
                convergio_7b = True
                break

            factor_7b  = (restante_7b / sum_lib_7b) if sum_lib_7b > 0 else 1.0
            nuevos_7b  = False

            for ix in idx_lib_7b:
                loc   = df_pgini.loc[ix, "loc_name PF"]
                gen   = gen_dict.get(loc)
                p_old = df_pgini.loc[ix, "pgini_MW"]
                p_new = (p_old * factor_7b) if sum_lib_7b > 0 \
                        else (restante_7b / len(idx_lib_7b))

                pmax = _get_pmax_pf(gen)
                pmin = _get_pmin_pf(gen) if p_old > 0 else 0.0

                if pmax < float("inf") and p_new > pmax:
                    df_pgini.loc[ix, "pgini_MW"] = pmax
                    libres_7b.discard(ix)
                    saturados_mw_7b += pmax
                    nuevos_7b = True
                elif p_old > 0 and p_new < pmin:
                    df_pgini.loc[ix, "pgini_MW"] = pmin
                    libres_7b.discard(ix)
                    saturados_mw_7b += pmin
                    nuevos_7b = True
                else:
                    df_pgini.loc[ix, "pgini_MW"] = round(p_new, 4)

            if not nuevos_7b:
                convergio_7b = True
                break

        # Aplicar a PF
        for ix in idx_prop_7b:
            loc       = df_pgini.loc[ix, "loc_name PF"]
            pgini_new = df_pgini.loc[ix, "pgini_MW"]
            gen = gen_dict.get(loc)
            if gen is None:
                continue
            gen.pgini = pgini_new
            if pgini_new == 0.0:
                _set_outserv(gen, 1)
            else:
                _set_outserv(gen, 0)

        sum_prop_tras = df_pgini.loc[idx_prop_7b, "pgini_MW"].sum()
        pgen_total_7b = p_fijo_7b + sum_prop_tras + slack_pgini_7b
        print(f"  Resultado [7b] ({'OK' if convergio_7b else 'no convergio'}):")
        print(f"  Slack '{slack_loc}' pgini = {slack_pgini_7b:.2f} MW  "
              f"(FIJO — P0_medido, no modificado)")
        print(f"  Pgen_no_slack post-[7b]   = {p_fijo_7b + sum_prop_tras:.2f} MW")
        print(f"  Pgen_total pre-LF         = {pgen_total_7b:.2f} MW  "
              f"(debe = Pdem_evento {Pdem_evento:.2f} MW)")
else:
    print("  Sin slack definido — no se realiza ajuste [7b].")

# =============================================================================
# [8] CARGAR CAPACIDAD DE TRANSFORMADORES (para restriccion en cargas)
# =============================================================================

cap_xfo = {}   # {barra_lv: MW_maximo}
if os.path.isfile(LOC_XFO_PATH):
    for sheet in ("Transformadores_2dev", "Transformadores_3dev"):
        try:
            df_xfo = pd.read_excel(LOC_XFO_PATH, sheet_name=sheet)
            for _, rx in df_xfo.iterrows():
                bus  = str(rx.get("Barra LV", "")).strip()
                pnom = _float(rx.get("Potencia nom. (MVA)", 0))
                if bus and bus != "nan" and pnom > 0:
                    cap_xfo[bus] = cap_xfo.get(bus, 0.0) + pnom * XFO_PF
        except Exception:
            pass
    print(f"\n  Capacidades de transformadores cargadas: {len(cap_xfo)} barras LV")
else:
    print(f"\n  [AVISO] No se encontro loc_names_xfo.xlsx — sin restriccion de transformadores")

# =============================================================================
# [9] ASIGNAR plini A CARGAS (con restriccion de transformadores si aplica)
# =============================================================================

separador("ASIGNANDO plini A CARGAS")

load_dict = {}
# Buscar en netdat para incluir cargas que estaban fuera de servicio
# (GetCalcRelevantObjects solo devuelve elementos en servicio)
lods = []
try:
    _nd9 = app.GetProjectFolder("netdat")
    if _nd9:
        lods = _nd9.GetContents("*.ElmLod", 1) or []
except Exception:
    pass
if not lods:
    lods = app.GetCalcRelevantObjects("*.ElmLod") or []
print(f"  Cargas encontradas en PF: {len(lods)}")
for load in lods:
    load_dict[load.loc_name.strip()] = load

# Todas las cargas deben permanecer en servicio (ya activadas en [4b]).
_locs_excel = set(df_plini["loc_name PF"].str.strip())
_cargas_sin_mapeo = sorted(_loc for _loc in load_dict if _loc not in _locs_excel)
print(f"  Cargas en Excel (mapeadas)                       : {len(_locs_excel)}")
print(f"  Cargas en PF sin mapeo en Excel                  : {len(_cargas_sin_mapeo)}")

# Mostrar diagnostico de cargas no mapeadas para identificar el problema
# CAUSA: estas cargas no estan en loc_name_cargas.xlsx o quedaron como
#        "Sin asignar" en MapeoRetirosSTI → no llegaron a condiciones_iniciales
if _cargas_sin_mapeo:
    print(f"\n  [DIAG] Cargas del modelo PF AUSENTES en el Excel condiciones_iniciales:")
    print(f"  {'loc_name':<30} {'P_nom MW':>10}")
    print(f"  {'-'*30} {'-'*10}")
    for _loc in _cargas_sin_mapeo:
        _ld_obj = load_dict[_loc]
        _pnom = 0.0
        for _attr in ("plini", "Pnom", "slini"):
            try:
                _v = float(getattr(_ld_obj, _attr, 0) or 0)
                if _v > 0:
                    _pnom = _v
                    break
            except Exception:
                pass
        print(f"  {_loc:<30} {_pnom:>10.4f}")
    print(f"\n  -> Actualizar loc_name_cargas.xlsx con MapeoRetirosSTI para incluirlas.")

# Identificar cargas encontradas y no encontradas en PF
miss_load  = []
found_rows = []   # (loc, plini_MW_excel, dist, barra_pf)

barra_col = "Barra PF" if "Barra PF" in df_plini.columns else None
for _, row in df_plini.iterrows():
    loc     = row["loc_name PF"]
    plini   = row["plini_MW"]
    dist    = row["Distribuidor"]
    barra   = str(row.get(barra_col, "")).strip() if barra_col else ""
    if load_dict.get(loc) is None:
        miss_load.append((loc, plini, dist))
    else:
        found_rows.append((loc, plini, dist, barra))

p_miss       = sum(mw for _, mw, _ in miss_load)
p_encontrada = sum(mw for _, mw, _, _ in found_rows)

# ── Diagnóstico: cargas del Excel no encontradas en PF ────────────────────────
if miss_load:
    _pf_names_lower = {n.lower(): n for n in load_dict}   # lower → nombre real PF
    print(f"\n  [DIAG] Cargas en Excel SIN coincidencia exacta en PF ({len(miss_load)}):")
    print(f"  {'loc_name Excel':<28} {'Candidato en PF (similar)':<28} {'Diferencia'}")
    print(f"  {'-'*28} {'-'*28} {'-'*30}")
    for miss_loc, miss_mw, miss_dist in miss_load:
        candidato = ""
        detalle   = ""
        # 1) coincidencia ignorando mayúsculas
        pf_match = _pf_names_lower.get(miss_loc.lower())
        if pf_match and pf_match != miss_loc:
            candidato = pf_match
            detalle   = "mayúsculas/minúsculas"
        else:
            # 2) coincidencia ignorando espacios internos y guiones
            _norm = lambda s: re.sub(r"[\s\-_]+", "", s).lower()
            for pf_low, pf_real in _pf_names_lower.items():
                if _norm(pf_low) == _norm(miss_loc):
                    candidato = pf_real
                    detalle   = "espacios/guiones/subguiones"
                    break
        if not candidato:
            # 3) buscar si el nombre Excel es substring del nombre PF o viceversa
            for pf_low, pf_real in _pf_names_lower.items():
                if miss_loc.lower() in pf_low or pf_low in miss_loc.lower():
                    candidato = pf_real
                    detalle   = "nombre parcialmente contenido"
                    break
        if not candidato:
            candidato = "(sin candidato)"
            detalle   = "nombre no existe en PF"
        print(f"  {miss_loc:<28} {candidato:<28} {detalle}")
    print()
# ─────────────────────────────────────────────────────────────────────────────

# factor_dem: escalar plini del Excel a Pdem_evento (referencia oficial).
# Esta correccion se aplica siempre: la demanda del evento es aproximada
# respecto del P0 medido y del despacho por bloque horario, por lo que
# incluso diferencias pequenas deben propagarse al modelo PF.
if pdem_excel > 0:
    factor_dem = Pdem_evento / pdem_excel
    print(f"  factor_dem = {factor_dem:.6f}  "
          f"(Pdem_evento={Pdem_evento:.1f} MW / pdem_excel={pdem_excel:.1f} MW)")
else:
    factor_dem = 1.0
    print("  [AVISO] pdem_excel <= 0; se omite el escalado de demanda por falta de base.")
    print(f"  factor_dem = {factor_dem:.6f}")
# factor_global: adicionalmente redistribuir MW de cargas no encontradas en PF

if p_encontrada > 0:
    factor_global = (pdem_excel * factor_dem) / p_encontrada
else:
    factor_global = 1.0

if abs(factor_global - 1.0) > 0.001:
    print(f"  Factor escala global (inc. redistrib.): {factor_global:.4f}  "
          f"(objetivo={pdem_excel * factor_dem:.1f} MW / encontrada={p_encontrada:.1f} MW)")
    if miss_load:
        print(f"  Cargas no encontradas en PF: {len(miss_load)} ({p_miss:.1f} MW redistribuidos)")

# Aplicar factor global y restriccion de transformadores por distribuidor
# Algoritmo iterativo por distribuidor:
#   1. Escalar plini proporcional al factor_global
#   2. Verificar si algun bus supera cap_xfo
#   3. Si supera -> saturar bus en cap_xfo, redistribuir resto entre buses libres
#   4. Repetir hasta convergencia
plini_final = {}   # {loc: MW_final}

# Agrupar found_rows por distribuidor
from collections import defaultdict
dist_rows = defaultdict(list)   # {dist: [(loc, plini_scaled, barra)]}
for loc, plini, dist, barra in found_rows:
    dist_rows[dist].append((loc, plini * factor_global, barra))

for dist, filas_d in dist_rows.items():
    p_dem_d = sum(mw for _, mw, _ in filas_d)
    if p_dem_d <= 0:
        for loc, _, _ in filas_d:
            plini_final[loc] = 0.0
        continue

    # Estructuras de trabajo
    idx_d    = list(range(len(filas_d)))
    vals     = [mw for _, mw, _ in filas_d]
    locs     = [loc for loc, _, _ in filas_d]
    barras   = [b   for _, _, b   in filas_d]

    libres       = set(idx_d)
    saturados_mw = 0.0

    for _ in range(50):
        idx_libres   = list(libres)
        sum_libre    = sum(vals[i] for i in idx_libres)
        restante     = p_dem_d - saturados_mw

        if not idx_libres or restante <= 1e-6:
            break

        factor_d = restante / sum_libre if sum_libre > 0 else 1.0
        for i in idx_libres:
            vals[i] = vals[i] * factor_d

        # Verificar capacidad de transformadores por barra
        nuevos_saturados = False
        barras_libres = {barras[i] for i in idx_libres if barras[i] and barras[i] != "nan"}
        for bus in barras_libres:
            cap_bus = cap_xfo.get(bus)
            if cap_bus is None:
                continue
            idx_bus = [i for i in idx_libres if barras[i] == bus]
            sum_bus = sum(vals[i] for i in idx_bus)
            if sum_bus > cap_bus + 1e-6:
                factor_sat = cap_bus / sum_bus
                for i in idx_bus:
                    vals[i]    *= factor_sat
                    libres.discard(i)
                saturados_mw  += cap_bus
                nuevos_saturados = True

        if not nuevos_saturados:
            break

    for i in idx_d:
        plini_final[locs[i]] = round(vals[i], 4)

# ── Redistribucion global de MW no asignados por limite de transformador ───────
_p_objetivo = pdem_excel * factor_dem        # MW totales objetivo
_p_asig_v1  = sum(plini_final.values())      # MW realmente asignados
_deficit    = _p_objetivo - _p_asig_v1       # MW perdidos por cap_xfo

if _deficit > 0.5:
    print(f"\n  [9-REDIST] Deficit por limites xfo: {_deficit:.2f} MW — redistribuyendo...")

    # Para cada barra calcular cuanto ya fue asignado (suma de sus cargas)
    _suma_por_barra = {}
    for loc, _, _, barra in found_rows:
        if barra and barra != "nan":
            _suma_por_barra[barra] = _suma_por_barra.get(barra, 0.0) \
                                     + plini_final.get(loc, 0.0)

    # Candidatas: cargas cuya barra tiene margen disponible o sin restriccion xfo
    _candidatas = []   # (loc, valor_actual, margen_disponible)
    for loc, _, _, barra in found_rows:
        val = plini_final.get(loc, 0.0)
        if val <= 0:
            continue
        if barra and barra != "nan" and barra in cap_xfo:
            margen = cap_xfo[barra] - _suma_por_barra.get(barra, 0.0)
        else:
            margen = float("inf")   # sin restriccion de xfo
        if margen > 1e-6:
            _candidatas.append((loc, val, margen))

    _suma_cand = sum(v for _, v, _ in _candidatas)
    if _suma_cand > 0:
        _redistribuido = 0.0
        for loc, val, margen in _candidatas:
            incremento = _deficit * (val / _suma_cand)
            if margen < float("inf"):
                incremento = min(incremento, margen)
            plini_final[loc] = round(plini_final.get(loc, 0.0) + incremento, 4)
            _redistribuido += incremento
        print(f"  [9-REDIST] Redistribuidos: {_redistribuido:.2f} MW "
              f"entre {len(_candidatas)} cargas")
        print(f"  [9-REDIST] Pdem_total post-redist: "
              f"{sum(plini_final.values()):.2f} MW  "
              f"(objetivo={_p_objetivo:.2f} MW)")
    else:
        print(f"  [9-REDIST] Sin cargas candidatas — deficit no redistribuible.")

# Asignar a objetos PF
ok_load      = 0
p_asig_total = 0.0
locs_asignadas = set()
for loc, plini, dist, barra in found_rows:
    ld = load_dict[loc]
    pf_val       = plini_final.get(loc, 0.0)
    ld.plini     = pf_val
    ld.outserv   = 0
    p_asig_total += pf_val
    ok_load      += 1
    locs_asignadas.add(loc)

dif_asig = p_asig_total - pdem_excel
print(f"  Asignadas              : {ok_load} cargas")
print(f"  Pdem Excel (referencia): {pdem_excel:>8.1f} MW")
print(f"  Pdem asignada a PF     : {p_asig_total:>8.1f} MW")
print(f"  Diferencia             : {dif_asig:>+8.1f} MW"
      + ("  [OK]" if abs(dif_asig) < 0.5 else "  [REVISAR]"))
if miss_load:
    print(f"\n  Cargas no encontradas en PF ({len(miss_load)}):")
    for m, mw, d in miss_load[:10]:
        print(f"    - {m:<28} {mw:.3f} MW  ({d})")
    if len(miss_load) > 10:
        print(f"    ... y {len(miss_load)-10} mas")

# Resumen por distribuidor
print(f"\n  {'Distribuidor':<35} {'P_dem MW':>9}  {'P_asig MW':>9}  {'N cargas':>8}")
print(f"  {'-'*35} {'-'*9}  {'-'*9}  {'-'*8}")
for dist in sorted(dict_dist_mw.keys()):
    p_dem  = sum(plini * factor_global
                 for loc, plini, d, _ in found_rows if d == dist)
    p_asig = sum(plini_final.get(loc, 0.0)
                 for loc, _, d, _ in found_rows if d == dist)
    n_c    = sum(1 for loc, _, d, _ in found_rows if d == dist)
    marca  = " !" if abs(p_dem - p_asig) > 0.5 else ""
    print(f"  {dist:<35} {p_dem:>9.2f}  {p_asig:>9.2f}  {n_c:>8}{marca}")

# =============================================================================
# [9d] REBALANCEAR GENERACION PROPORCIONAL A DEMANDA REAL CARGADA
# =============================================================================
# Sin pérdidas en PF: Pgen debe = Pdem_PF exacto.
# p_asig_total puede diferir de Pdem_evento por límites de transformadores.
# Reescalar solo las proporcionales para cerrar Pgen = p_asig_total,
# manteniendo P0_medido, disparo y slack intactos.

separador("REBALANCE GENERACION A DEMANDA REAL [9d]")

_brecha = p_asig_total - df_pgini["pgini_MW"].sum()
print(f"  p_asig_total (Pdem PF real) : {p_asig_total:.2f} MW")
print(f"  Pgen pre-[9d]               : {df_pgini['pgini_MW'].sum():.2f} MW")
print(f"  Brecha                      : {_brecha:+.2f} MW")

if abs(_brecha) > 0.5 and slack_gen is not None:
    _slack_loc_9d   = slack_gen.loc_name.strip()
    _slack_pgini_9d = float(getattr(slack_gen, "pgini", 0) or 0)
    _fuentes_fijas_9d = {"P0_medido", "sin_despacho", "mantenimiento",
                         "disparo_p_desc", "disparo_manual", "disparo_excel"}

    _p_fijo_9d = df_pgini.loc[
        df_pgini["Fuente"].isin(_fuentes_fijas_9d) &
        (df_pgini["loc_name PF"].str.strip() != _slack_loc_9d),
        "pgini_MW"
    ].sum()

    _obj_9d = p_asig_total - _p_fijo_9d - _slack_pgini_9d

    _idx_9d = df_pgini[
        (~df_pgini["Fuente"].isin(_fuentes_fijas_9d)) &
        (df_pgini["loc_name PF"].str.strip() != _slack_loc_9d)
    ].index.tolist()

    print(f"  p_fijo no-slack             : {_p_fijo_9d:.2f} MW")
    print(f"  Slack '{_slack_loc_9d}'     : {_slack_pgini_9d:.2f} MW  (FIJO)")
    print(f"  Obj proporcionales [9d]     : {_obj_9d:.2f} MW  ({len(_idx_9d)} uds)")

    if _obj_9d > 0 and _idx_9d:
        _libres_9d  = set(_idx_9d)
        _sat_mw_9d  = 0.0
        _conv_9d    = False

        for _ in range(50):
            _ilib   = list(_libres_9d)
            _slib   = df_pgini.loc[_ilib, "pgini_MW"].sum()
            _rest   = _obj_9d - _sat_mw_9d

            if not _ilib or _rest <= 1e-6:
                _conv_9d = True
                break

            _fact = (_rest / _slib) if _slib > 0 else 1.0
            _nsat = False

            for ix in _ilib:
                _loc  = df_pgini.loc[ix, "loc_name PF"]
                _gen  = gen_dict.get(_loc)
                _pold = df_pgini.loc[ix, "pgini_MW"]
                _pnew = _pold * _fact if _slib > 0 else _rest / len(_ilib)
                _pmax = _get_pmax_pf(_gen)
                _pmin = _get_pmin_pf(_gen) if _pold > 0 else 0.0

                if _pmax < float("inf") and _pnew > _pmax:
                    df_pgini.loc[ix, "pgini_MW"] = _pmax
                    _libres_9d.discard(ix); _sat_mw_9d += _pmax; _nsat = True
                elif _pold > 0 and _pnew < _pmin:
                    df_pgini.loc[ix, "pgini_MW"] = _pmin
                    _libres_9d.discard(ix); _sat_mw_9d += _pmin; _nsat = True
                else:
                    df_pgini.loc[ix, "pgini_MW"] = round(_pnew, 4)

                if _gen is not None:
                    _gen.pgini = df_pgini.loc[ix, "pgini_MW"]
                    if df_pgini.loc[ix, "pgini_MW"] == 0.0:
                        _set_outserv(_gen, 1)
                    else:
                        _set_outserv(_gen, 0)

            if not _nsat:
                _conv_9d = True
                break

        _pgen_9d = df_pgini["pgini_MW"].sum()
        print(f"  Resultado ({'OK' if _conv_9d else 'no convergio'}):")
        print(f"  Pgen_total post-[9d]        : {_pgen_9d:.2f} MW")
        print(f"  Balance Pgen - Pdem_PF      : {_pgen_9d - p_asig_total:+.2f} MW")
        print(f"  -> P_LF_slack esperado      : ~{_slack_pgini_9d:.2f} MW  (solo P0_medido)")
    else:
        print("  [INFO] Sin proporcionales ajustables o obj<=0 — sin cambio.")
else:
    print("  Brecha < 0.5 MW — sin ajuste necesario.")

# Guardar pgini original de la slack antes del Load Flow (para export en [9b])
pgini_slack_original = float(getattr(slack_gen, "pgini", 0) or 0) \
                       if slack_gen is not None else 0.0

# =============================================================================
# [9b] EXPORTAR DATOS CARGADOS A EXCEL
# =============================================================================

separador("EXPORTANDO DATOS CARGADOS")

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_HDR_FILL = PatternFill("solid", start_color="2E4057", end_color="2E4057")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_THIN     = Side(style="thin", color="CCCCCC")
_BORDER   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CTR      = Alignment(horizontal="center", vertical="center")

def _formato_ws(ws):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.fill = _HDR_FILL; cell.font = _HDR_FONT; cell.alignment = _CTR
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.border    = _BORDER
            cell.alignment = _CTR
    ws.freeze_panes = "A2"
    for c in range(1, ws.max_column + 1):
        col_let = get_column_letter(c)
        max_w   = max(len(str(ws.cell(r, c).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[col_let].width = min(max_w + 3, 40)

# ── Hoja pgini_GEN_FINAL ──────────────────────────────────────────────────────
_FILL_P0   = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
_FILL_PROP = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
_FILL_MANT = PatternFill("solid", start_color="FFCCCC", end_color="FFCCCC")

# Construir DataFrame con valores finales asignados a PF
filas_gen_final = []
for _, row in df_pgini.iterrows():
    loc    = row["loc_name PF"]
    fuente = row["Fuente"]
    gen    = gen_dict.get(loc)
    # Leer el pgini que quedo en PF (puede diferir de df_pgini si fue clampado)
    # Para la slack, usar el P0_medido original (antes del Load Flow)
    if gen is not None and slack_gen is not None \
            and gen.loc_name.strip() == slack_gen.loc_name.strip():
        pgini_pf = pgini_slack_original
    else:
        pgini_pf = float(getattr(gen, "pgini", row["pgini_MW"]) or 0) if gen else row["pgini_MW"]
    pmax     = _get_pmax_pf(gen)
    pmin_v   = _get_pmin_pf(gen)
    ng       = _get_ngnum(gen)
    try:
        tipo = gen.GetClassName() if gen else ""
    except Exception:
        tipo = ""
    filas_gen_final.append({
        "loc_name PF":    loc,
        "Generador_CNDC": row.get("Generador_CNDC", ""),
        "pgini_MW":       round(pgini_pf, 4),
        "Pmax_MW":        round(pmax, 2) if pmax < float("inf") else "",
        "Pmin_MW":        round(pmin_v, 2),
        "ngnum":          ng if tipo == "ElmGenstat" else "",
        "Fuente":         fuente,
        "Estado":         row.get("Estado", ""),
    })

df_gen_final = pd.DataFrame(filas_gen_final)

# ── Hoja plini_CAR_FINAL ──────────────────────────────────────────────────────
filas_car_final = []
for loc, plini_orig, dist, barra in found_rows:
    pf_val = plini_final.get(loc, 0.0)
    filas_car_final.append({
        "loc_name PF":  loc,
        "Distribuidor": dist,
        "Barra PF":     barra,
        "plini_MW":     round(pf_val, 4),
        "plini_orig_MW": round(plini_orig, 4),
    })
for loc, plini_orig, dist in miss_load:
    filas_car_final.append({
        "loc_name PF":  loc,
        "Distribuidor": dist,
        "Barra PF":     "",
        "plini_MW":     0.0,
        "plini_orig_MW": round(plini_orig, 4),
    })

df_car_final = pd.DataFrame(filas_car_final)

# ── Hoja Resumen_Cargado ──────────────────────────────────────────────────────
pgen_final_total = df_gen_final["pgini_MW"].sum()
pdem_final_total = df_car_final["plini_MW"].sum()

df_res_final = pd.DataFrame([
    {"Parametro": "Semestre",                "Valor": semestre},
    {"Parametro": "Evento",                  "Valor": f"Evento {n_evento}"},
    {"Parametro": "Fecha y hora",            "Valor": fecha_h},
    {"Parametro": "Disparo",                 "Valor": disparo},
    {"Parametro": "Hora gen.",               "Valor": hora_ev},
    {"Parametro": "Hora cargas",             "Valor": hora_po},
    {"Parametro": "Generadores asignados",   "Valor": ok_gen},
    {"Parametro": "  P0_medido",             "Valor": int(mask_p0.sum())},
    {"Parametro": "  CNDC_proporcional",     "Valor": int(mask_prop.sum())},
    {"Parametro": "  Mantenimiento",         "Valor": int(mask_mant.sum())},
    {"Parametro": "Pgen total asignada (MW)","Valor": round(pgen_final_total, 2)},
    {"Parametro": "Cargas asignadas",        "Valor": ok_load},
    {"Parametro": "Cargas no encontradas",   "Valor": len(miss_load)},
    {"Parametro": "Pdem total asignada (MW)","Valor": round(pdem_final_total, 2)},
    {"Parametro": "Balance Pgen-Pdem (MW)",  "Valor": round(pgen_final_total - pdem_final_total, 2)},
    {"Parametro": "Slack",                   "Valor": slack_gen.loc_name if slack_gen else ""},
    {"Parametro": "Pdem_evento (MW)",         "Valor": round(Pdem_evento, 2)},
    {"Parametro": "Balance Pgen-Pdem_evento (MW)", "Valor": round(pgen_total - Pdem_evento, 2)},
    # ── Disparo ──────────────────────────────────────────────────────────────
    {"Parametro": "Modo disparo",             "Valor": _modo_disparo},
    {"Parametro": "p_desc registrado (MW)",   "Valor": round(p_desc, 2)},
    {"Parametro": "Suma pgini disparo (MW)",  "Valor": round(_suma_disparo, 2)},
    {"Parametro": "Diferencia disparo (MW)",  "Valor": round(_suma_disparo - p_desc, 2)},


 ])

df_diag_control = pd.DataFrame([
    {"Parametro": "Pdem_evento (MW)",               "Valor": round(Pdem_evento, 2)},
    {"Parametro": "Pdem_excel (MW)",                "Valor": round(pdem_excel, 2)},
    {"Parametro": "factor_dem",                     "Valor": round(factor_dem, 6)},
    {"Parametro": "factor_global",                  "Valor": round(factor_global, 6)},
    {"Parametro": "Pgen pre-LF (MW)",               "Valor": round(pgen_total, 2)},
    {"Parametro": "Pdem PF real (MW)",              "Valor": round(p_asig_total, 2)},
    {"Parametro": "Balance Pgen-Pdem_evento (MW)",  "Valor": round(pgen_total - Pdem_evento, 2)},
    {"Parametro": "Balance Pgen-Pdem PF (MW)",      "Valor": round(pgen_total - p_asig_total, 2)},
    {"Parametro": "Generadores asignados",          "Valor": ok_gen},
    {"Parametro": "Cargas asignadas",               "Valor": ok_load},
    {"Parametro": "Cargas no encontradas",          "Valor": len(miss_load)},
    {"Parametro": "Slack",                          "Valor": slack_gen.loc_name if slack_gen else ""},
    {"Parametro": "Estado validacion",              "Valor": "REVISAR" if (len(miss_load) > 0 or abs(pgen_total - Pdem_evento) > 1.0) else "OK"},
])
# ── Exportar ──────────────────────────────────────────────────────────────────
datos_path = os.path.join(ev_path, f"datos_cargados_Ev{n_evento}.xlsx")

with pd.ExcelWriter(datos_path, engine="openpyxl") as writer:
    df_res_final.to_excel(writer, sheet_name="Resumen_Cargado",  index=False)
    df_diag_control.to_excel(writer, sheet_name="Diagnostico_Carga", index=False)
    df_gen_final.to_excel(writer, sheet_name="pgini_GEN_FINAL",  index=False)
    df_car_final.to_excel(writer, sheet_name="plini_CAR_FINAL",  index=False)

# Aplicar formato y colores
from openpyxl import load_workbook as _lw
wb_out = _lw(datos_path)

for sh in wb_out.sheetnames:
    _formato_ws(wb_out[sh])

# Color por fuente en pgini_GEN_FINAL
ws_g    = wb_out["pgini_GEN_FINAL"]
col_fue = next((c for c in range(1, ws_g.max_column + 1)
                if str(ws_g.cell(1, c).value) == "Fuente"), None)
if col_fue:
    for r in range(2, ws_g.max_row + 1):
        fuente_v = str(ws_g.cell(r, col_fue).value or "")
        if "mantenimiento" in fuente_v.lower():
            fill = _FILL_MANT
        elif "P0_medido" in fuente_v:
            fill = _FILL_P0
        else:
            fill = _FILL_PROP
        for c in range(1, ws_g.max_column + 1):
            ws_g.cell(r, c).fill = fill

wb_out.save(datos_path)
print(f"  Exportado: {os.path.basename(datos_path)}")
print(f"  Ruta     : {ev_path}")
print(f"    pgini_GEN_FINAL : {len(df_gen_final)} generadores  |  Pgen={pgen_final_total:.2f} MW")
print(f"    plini_CAR_FINAL : {len(df_car_final)} cargas       |  Pdem={pdem_final_total:.2f} MW")

# =============================================================================
# BALANCE PRE-LOAD FLOW
# =============================================================================

separador("BALANCE PRE-LOAD FLOW")

_p0_no_slack = df_pgini.loc[
    (df_pgini["Fuente"] == "P0_medido") &
    (df_pgini["loc_name PF"].str.strip() != (slack_gen.loc_name.strip() if slack_gen else "")),
    "pgini_MW"
].sum()
_p_disp  = df_pgini.loc[df_pgini["loc_name PF"].isin(disparo_locs_pf), "pgini_MW"].sum()
_modo_disp_lbl = _modo_disparo if "_modo_disparo" in dir() else "disparo"
_p_slack = float(getattr(slack_gen, "pgini", 0) or 0) if slack_gen else 0.0
_p_prop  = df_pgini.loc[df_pgini["Fuente"] == "CNDC_proporcional", "pgini_MW"].sum()
_pgen_total_pre = df_pgini["pgini_MW"].sum()
_plini_total    = sum(plini_final.get(loc, 0.0) for loc, *_ in found_rows)
_slack_name     = slack_gen.loc_name if slack_gen else "—"
_LINE = "  " + "─" * 50

print(f"  {'Pdem_evento (referencia oficial)':<40}: {Pdem_evento:>8.2f} MW")
print(f"  {'pdem_excel  (suma plini original)':<40}: {pdem_excel:>8.2f} MW")
print(f"  {'factor_dem  (escalado cargas)':<40}: {factor_dem:>10.6f}")
print(_LINE)
print(f"  {'P0_medido no-slack    (fijo)':<40}: {_p0_no_slack:>8.2f} MW")
print(f"  {(_modo_disp_lbl + '  (fijo)'):<40}: {_p_disp:>8.2f} MW")
print(f"  {'Slack (' + _slack_name + ')  (FIJO P0)':<40}: {_p_slack:>8.2f} MW")
print(f"  {'CNDC_proporcional     (escalado)':<40}: {_p_prop:>8.2f} MW")
print(_LINE)
print(f"  {'Pgen_total':<40}: {_pgen_total_pre:>8.2f} MW")
print(f"  {'plini total escalado':<40}: {_plini_total:>8.2f} MW")
print(_LINE)
_balance = _pgen_total_pre - Pdem_evento
_ok = "[OK]" if abs(_balance) < 1.0 else "[REVISAR]"
print(f"  {'Balance Pgen - Pdem_evento':<40}: {_balance:>+8.2f} MW  {_ok}")
print("  [El Load Flow ajustara la slack solo para cubrir perdidas de red]")
separador()

# =============================================================================
# RESUMEN FINAL
# =============================================================================

separador("RESUMEN FINAL")
print(f"  Semestre      : {semestre}")
print(f"  Evento        : Evento {n_evento}")
print(f"  Escenario PF  : {semestre} / {nombre_escenario}")
print(f"  Fecha y hora  : {fecha_h}")
print(f"  Disparo       : {disparo}")
print(f"  Generadores   : {ok_gen} asignados  |  {len(miss_gen)} no encontrados")
print(f"  Cargas        : {ok_load} asignadas  |  {len(miss_load)} no encontradas")
separador()

# =============================================================================
# [10b] MODELO ZIP — Potencia constante en todas las cargas en servicio
# =============================================================================
# Coeficientes ZIP activos (kpu_low, equiv_cur1) y reactivos (kqu_low, equiv_cur2):
#   a=1, b=0, c=0  → potencia constante pura
# Esto garantiza que plini se consuma exactamente sin corrección por tensión.
# =============================================================================

separador("MODELO ZIP CARGAS (potencia constante)")

_ZIP_ATTRS_P = ("kpu_low",  "aP",  "zip_aP",  "pload_a")
_ZIP_ATTRS_Q = ("kqu_low",  "aQ",  "zip_aQ",  "qload_a")
_ZIP_B_P     = ("bP",  "zip_bP",  "pload_b",  "equiv_cur1")
_ZIP_B_Q     = ("bQ",  "zip_bQ",  "qload_b")
_ZIP_C_P     = ("cP",  "zip_cP",  "pload_c",  "equiv_cur2")
_ZIP_C_Q     = ("cQ",  "zip_cQ",  "qload_c")

def _set_zip_attr(obj, candidates, value):
    """Intenta fijar 'value' en el primer atributo de 'candidates' que exista."""
    for attr in candidates:
        try:
            obj.SetAttribute(attr, value)
            return attr
        except Exception:
            pass
    return None

_n_zip_ok   = 0
_n_zip_warn = 0
for _loc, _ld in load_dict.items():
    if getattr(_ld, "outserv", 1) != 0:
        continue
    try:
        _ok_aP = _set_zip_attr(_ld, _ZIP_ATTRS_P, 1.0)
        _ok_aQ = _set_zip_attr(_ld, _ZIP_ATTRS_Q, 1.0)
        _ok_bP = _set_zip_attr(_ld, _ZIP_B_P,     0.0)
        _ok_bQ = _set_zip_attr(_ld, _ZIP_B_Q,     0.0)
        _ok_cP = _set_zip_attr(_ld, _ZIP_C_P,     0.0)
        _ok_cQ = _set_zip_attr(_ld, _ZIP_C_Q,     0.0)
        if _ok_aP:
            _n_zip_ok += 1
        else:
            _n_zip_warn += 1
    except Exception as _e:
        _n_zip_warn += 1

print(f"  Cargas configuradas ZIP (a=1,b=0,c=0) : {_n_zip_ok}")
if _n_zip_warn:
    print(f"  [AVISO] Cargas sin atributos ZIP reconocidos: {_n_zip_warn}")
    print("    -> Verificar atributos ZIP en ElmLod con dir() o GetAttribute")
separador()

# =============================================================================
# [10] LOAD FLOW
# =============================================================================

separador("ACTIVACION DE CONTROL DE TENSION EN SHUNTS")

# iswitch=1 activa el checkbox "Switchable" (Load Flow tab del ElmShnt),
# que permite al shunt ajustar sus pasos para controlar la tension en barra.
# Combinado con iopt_asht=1 en el Load Flow, habilita el control automatico.
_shunts_ci = []
try:
    _nd_sh = app.GetProjectFolder("netdat")
    if _nd_sh:
        _shunts_ci = _nd_sh.GetContents("*.ElmShnt", 1) or []
except Exception:
    pass
if not _shunts_ci:
    _shunts_ci = app.GetCalcRelevantObjects("*.ElmShnt") or []

_n_iswitch = 0
for _sh in _shunts_ci:
    try:
        if getattr(_sh, "outserv", 0) == 1:
            continue
        try:
            _sh.iswitch = 1
        except Exception:
            _sh.SetAttribute("iswitch", 1)
        _n_iswitch += 1
    except Exception as _e:
        print(f"  [AVISO] No se pudo activar Switchable en "
              f"{getattr(_sh, 'loc_name', '?')}: {_e}")
print(f"  Shunts con Switchable activo: {_n_iswitch} (de {len(_shunts_ci)} totales)")

separador("LOAD FLOW")

lf = app.GetFromStudyCase("ComLdf")
lf.iopt_net    = 0   # AC balanceado
lf.iopt_at     = 1   # Ajuste automatico de taps
lf.iopt_asht   = 1   # Ajuste automatico de shunts
lf.iopt_lim    = 1   # Limites reactivos activos
lf.iopt_island = 0   # Red unica
lf.iopt_pq     = 1   # Consider Voltage Dependency of Loads

resultado_lf = lf.Execute()

if resultado_lf == 0:
    print("  ✓ Load Flow CONVERGIDO correctamente")
else:
    print("  ✗ Load Flow NO convergio")
    print("    Causas frecuentes:")
    print("    1. Maquina slack con outserv=1 o pgini=0")
    print("    2. Areas aisladas sin generador de referencia")
    print("    3. Desequilibrio generacion-carga excesivo")
    print("    Verifique Output Window en PowerFactory")

# =============================================================================
# [10] CORRECCION POST-LF: MANTENER SLACK EN SU PGINI ASIGNADO
# =============================================================================
# Despues del LF el slack absorbe desequilibrios residuales de la topologia,
# por lo que su potencia real puede diferir de pgini_slack_original.
# Se pregunta al usuario si desea realizar el ajuste.
# Si acepta:
#   1. Lee la potencia real del slack post-LF (m:Psum:bus1)
#   2. Calcula delta = p_slack_lf - pgini_slack_original
#   3. Distribuye delta entre las unidades CNDC_proporcional (proporcional a pgini)
#   4. Restaura pgini del slack al valor original
#   5. Re-corre el LF
#   6. Itera hasta |delta| < tolerancia (max _MAX_ITER_10 veces)
#   7. Exporta datos finales ajustados a un Excel separado

def _read_p_gen_lf(gen):
    """Lee potencia activa real post-LF en MW (positivo = generando).
    Atributo PF: m:Psum:bus1 = Total Active Power, convenio generador,
    devuelve positivo cuando la maquina inyecta potencia a la red."""
    try:
        v = gen.GetAttribute("m:Psum:bus1")
        if v is not None:
            return float(v)
    except Exception:
        pass
    # Fallback: pgini es el setpoint, no el resultado real del LF
    return float(getattr(gen, "pgini", 0) or 0)


_TOL_SLACK_10  = 0.1    # MW — tolerancia aceptable en potencia del slack
_MAX_ITER_10   = 15     # iteraciones maximas del ciclo de correccion
_ajuste_10     = False  # flag: el usuario acepto realizar el ajuste

if resultado_lf == 0 and slack_gen is not None:
    # ── Consulta al usuario ───────────────────────────────────────────────────
    separador("CORRECCION POST-LF: SLACK A PGINI ASIGNADO")
    _p_slack_previo = _read_p_gen_lf(slack_gen)
    _delta_previo   = _p_slack_previo - pgini_slack_original
    print(f"  Slack '{slack_gen.loc_name.strip()}'")
    print(f"  pgini asignado (P0_medido)   : {pgini_slack_original:.2f} MW")
    print(f"  Potencia real post-LF        : {_p_slack_previo:.2f} MW")
    print(f"  Diferencia (delta)           : {_delta_previo:+.2f} MW")
    print()
    if not AJUSTAR_POST_LF:
        print("  [OMITIDO] Correccion post-LF deshabilitada para preservar las potencias iniciales asignadas.")
        print("  La slack conserva su punto de operacion y ComInc usara este estado sin redistribuir potencia.")
    elif abs(_delta_previo) <= _TOL_SLACK_10:
        print(f"  [OK] Delta <= {_TOL_SLACK_10} MW — no se requiere ajuste.")
    else:
        print(f"  El slack genera {abs(_delta_previo):.2f} MW "
              f"{'de mas' if _delta_previo > 0 else 'de menos'} respecto a su P0_medido.")
        print(f"  El ajuste redistribuye ese exceso entre las unidades proporcionales")
        print(f"  y exporta los datos finales a un Excel separado (_ajustado).")
        print()
        _resp_10 = input("  ¿Realizar ajuste? [s/N]: ").strip().lower()
        if _resp_10 == "s":
            _ajuste_10 = True
        else:
            print("  Ajuste omitido — ComInc se ejecutara con el estado actual del LF.")
    print()

if _ajuste_10 and resultado_lf == 0 and slack_gen is not None:
    _slack_loc_10    = slack_gen.loc_name.strip()
    _slack_target_10 = pgini_slack_original   # pgini asignado antes del LF

    # Indices proporcionales ajustables (identico criterio a [9d])
    _fuentes_fijas_10 = {"P0_medido", "sin_despacho", "mantenimiento",
                         "disparo_p_desc", "disparo_manual", "disparo_excel"}
    _idx_prop_10 = df_pgini[
        (~df_pgini["Fuente"].isin(_fuentes_fijas_10)) &
        (df_pgini["loc_name PF"].str.strip() != _slack_loc_10)
    ].index.tolist()

    print(f"  Tolerancia                 : {_TOL_SLACK_10:.1f} MW")
    print(f"  Proporcionales ajustables  : {len(_idx_prop_10)} unidades")
    print()

    # Estructura del loop: ajusta → corre LF → lee resultado → verifica convergencia
    # Asi el chequeo siempre refleja el LF recien ejecutado (sin desfase de iteracion)
    _conv_10     = False
    _p_slack_lf  = _read_p_gen_lf(slack_gen)   # lectura inicial (LF ya convergido arriba)
    _delta_10    = _p_slack_lf - _slack_target_10

    for _iter_10 in range(_MAX_ITER_10):
        _estado_10 = ("[OK]"            if abs(_delta_10) <= _TOL_SLACK_10 else
                      "[AJUSTANDO]"     if abs(_delta_10) <  5.0            else
                      "[AJUSTE GRANDE]")
        print(f"  Iter {_iter_10 + 1:2d}: p_slack_LF = {_p_slack_lf:8.2f} MW  "
              f"objetivo = {_slack_target_10:8.2f} MW  "
              f"delta = {_delta_10:+7.2f} MW  {_estado_10}")

        if abs(_delta_10) <= _TOL_SLACK_10:
            _conv_10 = True
            break

        # Restaurar pgini del slack al valor objetivo
        slack_gen.pgini = round(_slack_target_10, 4)

        # Distribuir delta entre proporcionales (proporcional a su pgini actual)
        # delta > 0: slack hizo MAS → proporcionales absorben el exceso (suben)
        # delta < 0: slack hizo MENOS → proporcionales liberan el deficit (bajan)
        if _idx_prop_10:
            _sum_prop_10 = df_pgini.loc[_idx_prop_10, "pgini_MW"].sum()
            if _sum_prop_10 > 1e-6:
                _libres_10 = set(_idx_prop_10)
                _sat_mw_10 = 0.0
                for _ in range(20):
                    _ilib_10 = list(_libres_10)
                    if not _ilib_10:
                        break
                    _slib_10 = df_pgini.loc[_ilib_10, "pgini_MW"].sum()
                    _rest_10 = _delta_10 - _sat_mw_10
                    _nsat_10 = False
                    for ix in _ilib_10:
                        _p_old = df_pgini.loc[ix, "pgini_MW"]
                        _inc   = (_p_old / _slib_10) * _rest_10 if _slib_10 > 0 else _rest_10 / len(_ilib_10)
                        _p_new = max(0.0, _p_old + _inc)
                        _gen_p = gen_dict.get(df_pgini.loc[ix, "loc_name PF"])
                        _pmax  = _get_pmax_pf(_gen_p)
                        _pmin  = _get_pmin_pf(_gen_p) if _p_old > 0 else 0.0
                        if _pmax < float("inf") and _p_new > _pmax:
                            df_pgini.loc[ix, "pgini_MW"] = round(_pmax, 4)
                            _sat_mw_10 += (_pmax - _p_old)
                            _libres_10.discard(ix)
                            _nsat_10 = True
                        elif _p_old > 0 and _p_new < _pmin:
                            df_pgini.loc[ix, "pgini_MW"] = round(_pmin, 4)
                            _sat_mw_10 += (_pmin - _p_old)
                            _libres_10.discard(ix)
                            _nsat_10 = True
                        else:
                            df_pgini.loc[ix, "pgini_MW"] = round(_p_new, 4)
                        if _gen_p is not None:
                            _gen_p.pgini = df_pgini.loc[ix, "pgini_MW"]
                    if not _nsat_10:
                        break

        # Re-correr LF y leer resultado inmediatamente
        resultado_lf = lf.Execute()
        if resultado_lf != 0:
            print(f"  [ERROR] LF no convergió en iteración {_iter_10 + 1} — abortando corrección")
            break
        # Actualizar delta con el resultado del LF recien ejecutado
        _p_slack_lf = _read_p_gen_lf(slack_gen)
        _delta_10   = _p_slack_lf - _slack_target_10

    # Resumen final — _p_slack_lf y _delta_10 ya tienen el ultimo valor post-LF
    _p_slack_final = _p_slack_lf
    _delta_final   = _delta_10
    if resultado_lf == 0:
        if _conv_10:
            print(f"\n  ✓ Slack estabilizado: p_slack = {_p_slack_final:.2f} MW  "
                  f"(error = {_delta_final:+.2f} MW)")
        else:
            print(f"\n  [AVISO] No convergió en {_MAX_ITER_10} iter.  "
                  f"p_slack = {_p_slack_final:.2f} MW  (error = {_delta_final:+.2f} MW)")
            print(f"  -> ComInc se ejecutará con este estado.")

        _pgen_no_slack = df_pgini.loc[
            df_pgini["loc_name PF"].str.strip() != _slack_loc_10, "pgini_MW"].sum()
        print(f"\n  Balance post-ajuste:")
        print(f"    Pgen proporcionales + P0   : {_pgen_no_slack:.2f} MW")
        print(f"    Pgen slack (real LF)       : {_p_slack_final:.2f} MW")
        print(f"    Pgen total                 : {_pgen_no_slack + _p_slack_final:.2f} MW")
        print(f"    Pdem evento (objetivo)     : {Pdem_evento:.2f} MW")

        # ── Exportar datos ajustados a Excel separado ─────────────────────────
        separador("EXPORTANDO DATOS AJUSTADOS")
        _filas_aj = []
        for _, _row_aj in df_pgini.iterrows():
            _loc_aj  = _row_aj["loc_name PF"]
            _gen_aj  = gen_dict.get(_loc_aj)
            # Para el slack usar el valor real del LF; para el resto usar pgini en PF
            if _loc_aj.strip() == _slack_loc_10:
                _pgini_aj = _p_slack_final
            else:
                _pgini_aj = float(getattr(_gen_aj, "pgini", _row_aj["pgini_MW"]) or 0) \
                            if _gen_aj else _row_aj["pgini_MW"]
            _pmax_aj = _get_pmax_pf(_gen_aj)
            _pmin_aj = _get_pmin_pf(_gen_aj)
            _ng_aj   = _get_ngnum(_gen_aj)
            try:
                _tipo_aj = _gen_aj.GetClassName() if _gen_aj else ""
            except Exception:
                _tipo_aj = ""
            _filas_aj.append({
                "loc_name PF":    _loc_aj,
                "Generador_CNDC": _row_aj.get("Generador_CNDC", ""),
                "pgini_MW":       round(_pgini_aj, 4),
                "Pmax_MW":        round(_pmax_aj, 2) if _pmax_aj < float("inf") else "",
                "Pmin_MW":        round(_pmin_aj, 2),
                "ngnum":          _ng_aj if _tipo_aj == "ElmGenstat" else "",
                "Fuente":         _row_aj["Fuente"],
                "Estado":         _row_aj.get("Estado", ""),
            })

        _df_gen_aj  = pd.DataFrame(_filas_aj)
        _pgen_aj_tot = _df_gen_aj["pgini_MW"].sum()
        _pdem_aj_tot = df_car_final["plini_MW"].sum()

        _df_res_aj = pd.DataFrame([
            {"Parametro": "Semestre",                     "Valor": semestre},
            {"Parametro": "Evento",                       "Valor": f"Evento {n_evento}"},
            {"Parametro": "Fecha y hora",                 "Valor": fecha_h},
            {"Parametro": "Slack",                        "Valor": _slack_loc_10},
            {"Parametro": "pgini slack P0_medido (MW)",   "Valor": round(_slack_target_10, 2)},
            {"Parametro": "pgini slack real LF (MW)",     "Valor": round(_p_slack_final, 2)},
            {"Parametro": "Delta slack corregido (MW)",   "Valor": round(_delta_final, 2)},
            {"Parametro": "Ajuste convergido",            "Valor": str(_conv_10)},
            {"Parametro": "Pgen total ajustada (MW)",     "Valor": round(_pgen_aj_tot, 2)},
            {"Parametro": "Pdem total (MW)",              "Valor": round(_pdem_aj_tot, 2)},
            {"Parametro": "Balance Pgen-Pdem (MW)",       "Valor": round(_pgen_aj_tot - _pdem_aj_tot, 2)},
            {"Parametro": "Pdem_evento (MW)",             "Valor": round(Pdem_evento, 2)},
        ])

        _df_diag_aj = pd.DataFrame([
            {"Parametro": "Pdem_evento (MW)",            "Valor": round(Pdem_evento, 2)},
            {"Parametro": "Pgen total ajustada (MW)",     "Valor": round(_pgen_aj_tot, 2)},
            {"Parametro": "Pdem total (MW)",             "Valor": round(_pdem_aj_tot, 2)},
            {"Parametro": "Balance Pgen-Pdem (MW)",      "Valor": round(_pgen_aj_tot - _pdem_aj_tot, 2)},
            {"Parametro": "Slack real LF (MW)",          "Valor": round(_p_slack_final, 2)},
            {"Parametro": "Delta slack corregido (MW)",   "Valor": round(_delta_final, 2)},
            {"Parametro": "Ajuste convergido",            "Valor": str(_conv_10)},
            {"Parametro": "Estado validacion",            "Valor": "REVISAR" if abs(_pgen_aj_tot - _pdem_aj_tot) > 1.0 else "OK"},
        ])

        _path_aj = os.path.join(ev_path, f"datos_cargados_Ev{n_evento}_ajustado.xlsx")
        with pd.ExcelWriter(_path_aj, engine="openpyxl") as _wr_aj:
            _df_res_aj.to_excel(_wr_aj, sheet_name="Resumen_Ajustado", index=False)
            _df_diag_aj.to_excel(_wr_aj, sheet_name="Diagnostico_Ajuste", index=False)
            _df_gen_aj.to_excel(_wr_aj, sheet_name="pgini_GEN_AJUSTADO", index=False)
            df_car_final.to_excel(_wr_aj, sheet_name="plini_CAR_FINAL", index=False)

        # Aplicar formato y colores
        _wb_aj = _lw(_path_aj)
        for _sh_aj in _wb_aj.sheetnames:
            _formato_ws(_wb_aj[_sh_aj])
        _ws_g_aj   = _wb_aj["pgini_GEN_AJUSTADO"]
        _col_fue_aj = next((c for c in range(1, _ws_g_aj.max_column + 1)
                            if str(_ws_g_aj.cell(1, c).value) == "Fuente"), None)
        if _col_fue_aj:
            for r in range(2, _ws_g_aj.max_row + 1):
                _fv = str(_ws_g_aj.cell(r, _col_fue_aj).value or "")
                _fill_aj = (_FILL_MANT if "mantenimiento" in _fv.lower() else
                            _FILL_P0   if "P0_medido"     in _fv          else
                            _FILL_PROP)
                for c in range(1, _ws_g_aj.max_column + 1):
                    _ws_g_aj.cell(r, c).fill = _fill_aj
        _wb_aj.save(_path_aj)
        print(f"  Exportado: {os.path.basename(_path_aj)}")
        print(f"  Ruta     : {ev_path}")
        print(f"    pgini_GEN_AJUSTADO : {len(_df_gen_aj)} generadores  |  Pgen={_pgen_aj_tot:.2f} MW")
        print(f"    plini_CAR_FINAL    : {len(df_car_final)} cargas      |  Pdem={_pdem_aj_tot:.2f} MW")
    print()

# =============================================================================
# [11] INICIALIZACION RMS (ComInc)
# =============================================================================

separador("INICIALIZACION RMS (ComInc)")

if resultado_lf == 0:
    inc = app.GetFromStudyCase("ComInc")
    try:
        inc.iopt_island = 0
    except Exception:
        pass
    resultado_inc = inc.Execute()
    if resultado_inc == 0:
        print("  ✓ ComInc ejecutado correctamente")
        print("    -> Modelos dinamicos inicializados en el punto de operacion")
        print("    -> Revise Output Window: WARNINGs indican modelos con oscilaciones desde t=0")
    else:
        print(f"  ✗ ComInc FALLO (codigo={resultado_inc})")
        print("    Causas frecuentes:")
        print("    1. pgini fuera de [Pmin, Pmax] del modelo dinamico")
        print("    2. Parametros de governor/AVR inconsistentes")
        print("    3. Tensiones nodales fuera de rango")
else:
    print("  [OMITIDO] Load Flow no convergido — ComInc requiere LF previo.")

print()
print("  PowerFactory permanece abierto para verificacion.")
print()
input("  Presione Enter para finalizar el script (PF permanece abierto)...")
