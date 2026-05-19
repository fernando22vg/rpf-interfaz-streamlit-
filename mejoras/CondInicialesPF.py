
# =============================================================================
# CondInicialesPF.py
# Genera condiciones iniciales pgini/plini para PowerFactory desde datos CNDC.
# Version mejorada:
#   - conserva la logica base original
#   - cierra el balance de demanda por distribuidor con correccion del residuo
#   - agrega hoja de diagnostico de balance plini
#   - deja trazabilidad de asignaciones P0 / proporcional / mantenimiento
# =============================================================================
#
# Flujo:
#   1. Seleccion interactiva de semestre y evento
#   2. Lee datos_simulacion_{fecha}_2daopcion.xlsx (generacion y demanda por distribuidor)
#   3. Asigna pgini por loc_name a cada generador usando los datos originales CNDC
#   4. Usa la demanda del bloque horario (hora_po) directamente
#   5. Distribuye plini proporcionalmente segun potencia nominal (P_nom_MW)
#      para cada carga dentro de cada distribuidor
#   6. Corrige el residuo de redondeo para que la suma por distribuidor sea exacta
#   7. Exporta condiciones_iniciales_{fecha}_Ev{N}.xlsx listo para cargar en PowerFactory
#
# Entradas:
#   datos_simulacion_{fecha}_2daopcion.xlsx   (carpeta del evento)
#   loc_names_gen.xlsx     -> Mapeo_Generadores, Detalle_PF
#   loc_name_cargas.xlsx   -> LocNames_por_Dist, Mapeo_Cargas, Curvas_LocNames
#   loc_names_xfo.xlsx     -> potencia nominal de transformadores
# Salidas:
#   condiciones_iniciales_{fecha}_Ev{N}.xlsx  (misma carpeta del evento)
#   Hoja pgini_GEN : una fila por loc_name PF generador
#   Hoja plini_CAR : una fila por loc_name PF carga
#   Hoja Perfil_MW_Dist : perfil horario por distribuidor
#   Hoja Balance_Plini : diagnostico por distribuidor
#   Hoja Resumen   : totales y metadata del evento
# =============================================================================

import os
import glob
import re
import unicodedata
import difflib
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Rutas fijas ───────────────────────────────────────────────────────────────
RAIZ = r"C:\Datos del CNDC\01_INFO CNDC_RPF"
LOC_DIR = r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\Designacion de loc_name"
LOC_GEN_PATH = os.path.join(LOC_DIR, "loc_names_gen.xlsx")
LOC_CAR_PATH = os.path.join(LOC_DIR, "loc_name_cargas.xlsx")
LOC_XFO_PATH = os.path.join(LOC_DIR, "loc_names_xfo.xlsx")

# ── Colores Excel ─────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", start_color="2E4057", end_color="2E4057")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
FILL_P0 = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")      # verde: P0 medido
FILL_PROP = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")    # amarillo: proporcional
FILL_MANT = PatternFill("solid", start_color="FFCCCC", end_color="FFCCCC")     # rojo claro: mantenimiento
FILL_SD = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")      # gris: sin despacho
FILL_BAL_OK = PatternFill("solid", start_color="E2F0D9", end_color="E2F0D9")  # verde suave
FILL_BAL_WARN = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PRECISION_PLINI = 4
SCALE_PLINI = 10 ** PRECISION_PLINI


# =============================================================================
# HELPERS
# =============================================================================

def elegir(opciones, titulo):
    print(f"\n{titulo}:")
    for i, op in enumerate(opciones, 1):
        print(f"  {i}. {op}")
    while True:
        try:
            sel = int(input("  Seleccionar numero: "))
            if 1 <= sel <= len(opciones):
                return opciones[sel - 1]
        except ValueError:
            pass
        print("  Opcion invalida, intente de nuevo.")


def extraer_sti(loc_name):
    """sym_ZON01 -> ZON01 | WT_QOL01_EQ -> QOL01 | sym_ZON01(1) -> ZON01"""
    s = re.sub(r"\(\d+\)$", "", str(loc_name).strip())
    for pref in ("sym_", "WT_", "PV-", "PV_", "sta_"):
        if s.lower().startswith(pref.lower()):
            s = s[len(pref):]
            break
    s = re.sub(r"_EQ$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"_II$", "", s, flags=re.IGNORECASE)
    return s.strip()


def buscar_p0(sti_code, dict_p0):
    """Busca P_0 con fallback: TIQ01 -> TIQ (codigo sin digitos finales)."""
    if sti_code in dict_p0:
        return dict_p0[sti_code]
    base = re.sub(r"\d+$", "", sti_code)
    if base and base in dict_p0:
        return dict_p0[base]
    return None


def _float(val, default=0.0):
    try:
        v = float(val)
        return v if not pd.isna(v) else default
    except Exception:
        return default


def _norm(s):
    """Quita tildes y pasa a minusculas para comparacion flexible."""
    return unicodedata.normalize("NFD", str(s)).encode("ascii", "ignore").decode().lower().strip()


def _norm_hora(s):
    """Normaliza hora a formato HH:MM de dos digitos."""
    s = str(s).strip().replace("::", ":")
    parts = s.split(":")
    if len(parts) >= 2:
        try:
            return f"{int(parts[0]):02d}:{parts[1]}"
        except ValueError:
            pass
    return s


def _formato_hoja(ws):
    """Encabezado oscuro + bordes + autowidth + freeze row 1."""
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).border = BORDER
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for c in range(1, ws.max_column + 1):
        col_let = get_column_letter(c)
        max_w = max(len(str(ws.cell(r, c).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[col_let].width = min(max_w + 3, 45)


def _ajustar_plini_por_distribuidor(df_plini, dict_plini_dist):
    """
    Ajusta plini por distribuidor con un metodo determinista de mayor residuo.

    Estrategia:
      1. Calcula la asignacion exacta por carga.
      2. Convierte a unidades de 1e-4 MW.
      3. Asigna las unidades faltantes a las cargas con mayor parte fraccionaria.
      4. Si hubiera un exceso residual por redondeo, lo retira de las menores partes
         fraccionarias que aun tengan unidades asignadas.

    Esto evita varias rondas de redistribucion y deja la suma exacta por distribuidor
    en una sola pasada.
    """
    balance_rows = []

    if df_plini.empty:
        return pd.DataFrame(columns=[
            "Distribuidor",
            "Pdem_objetivo_MW",
            "Pdem_base_MW",
            "Pdem_final_MW",
            "Residuo_MW",
            "N_cargas",
        ]), 0.0

    for dist, grupo in df_plini.groupby("Distribuidor", sort=False):
        idxs = grupo.index.tolist()
        objetivo = round(float(dict_plini_dist.get(dist, 0.0)), PRECISION_PLINI)
        base = round(float(grupo["plini_base_MW"].sum()), PRECISION_PLINI) if "plini_base_MW" in grupo.columns else 0.0
        pnom_total = float(grupo["P_nom_MW"].sum())
        n = len(idxs)

        if n == 0:
            continue

        if objetivo <= 0.0:
            df_plini.loc[idxs, "plini_MW"] = 0.0
            balance_rows.append({
                "Distribuidor": dist,
                "Pdem_objetivo_MW": objetivo,
                "Pdem_base_MW": base,
                "Pdem_final_MW": 0.0,
                "Residuo_MW": round(objetivo, PRECISION_PLINI),
                "N_cargas": n,
            })
            continue

        if pnom_total > 0:
            exactas = [objetivo * (float(grupo.loc[ix, "P_nom_MW"]) / pnom_total) for ix in idxs]
        else:
            cuota = objetivo / n
            exactas = [cuota for _ in idxs]

        exact_units = [val * SCALE_PLINI for val in exactas]
        base_units = [int(val // 1) for val in exact_units]
        fracciones = [exact_units[i] - base_units[i] for i in range(n)]
        objetivo_units = int(round(objetivo * SCALE_PLINI))
        faltantes = objetivo_units - sum(base_units)

        if faltantes > 0:
            orden = sorted(range(n), key=lambda i: (fracciones[i], float(grupo.loc[idxs[i], "P_nom_MW"]), -i), reverse=True)
            for i in orden[:faltantes]:
                base_units[i] += 1
        elif faltantes < 0:
            orden = sorted(range(n), key=lambda i: (fracciones[i], float(grupo.loc[idxs[i], "P_nom_MW"]), i))
            retirados = 0
            for i in orden:
                if retirados >= abs(faltantes):
                    break
                if base_units[i] > 0:
                    base_units[i] -= 1
                    retirados += 1

        finales = [round(units / SCALE_PLINI, PRECISION_PLINI) for units in base_units]
        df_plini.loc[idxs, "plini_MW"] = finales

        final = round(float(sum(finales)), PRECISION_PLINI)
        residuo_final = round(objetivo - final, PRECISION_PLINI)

        balance_rows.append({
            "Distribuidor": dist,
            "Pdem_objetivo_MW": objetivo,
            "Pdem_base_MW": base,
            "Pdem_final_MW": final,
            "Residuo_MW": residuo_final,
            "N_cargas": n,
        })

    df_balance = pd.DataFrame(balance_rows)
    max_residuo = float(df_balance["Residuo_MW"].abs().max()) if not df_balance.empty else 0.0
    return df_balance, max_residuo


# =============================================================================
# SISTEMAS COBEE — regla "sin_despacho"
# =============================================================================
_COBEE_ZONGO_STI = {
    "ZON", "TIQ", "BOT", "CUT", "SRO", "SAI", "CHU", "HAR", "CAH", "HUA",
}
# Para Miguillas se puede extender en el futuro:
# _COBEE_MIGUILLAS_STI = {"MIG", "ANG", "CHO", "CRB"}
# _COBEE_ZONGO_STI = _COBEE_ZONGO_STI | _COBEE_MIGUILLAS_STI


# =============================================================================
# SELECCION SEMESTRE / EVENTO
# =============================================================================

semestres = sorted(d for d in os.listdir(RAIZ) if os.path.isdir(os.path.join(RAIZ, d)))
semestre = elegir(semestres, "Semestre de estudio")

base_ev = os.path.join(RAIZ, semestre, "Análisis_todos_los_eventos")
eventos = sorted(d for d in os.listdir(base_ev) if os.path.isdir(os.path.join(base_ev, d)))
evento = elegir(eventos, "Evento")
ev_path = os.path.join(base_ev, evento)

sim_files = glob.glob(os.path.join(ev_path, "datos_simulacion_*_2daopcion.xlsx"))
if not sim_files:
    raise FileNotFoundError(f"No se encontro datos_simulacion_*_2daopcion.xlsx en:\n  {ev_path}")
sim_path = sim_files[0]
m = re.search(r"datos_simulacion_(\d+)_", os.path.basename(sim_path))
fecha_str = m.group(1) if m else "000000"
print(f"\nArchivo fuente: {os.path.basename(sim_path)}")


# =============================================================================
# [1] LEER INFO_EVENTO
# =============================================================================

print("\n[1/5] Leyendo Info_Evento...")
df_info = pd.read_excel(sim_path, sheet_name="Info_Evento")
info = dict(zip(df_info.iloc[:, 0].astype(str), df_info.iloc[:, 1].astype(str)))

hora_po = _norm_hora(info.get("Hora Po (inicio)", "").strip())
fecha_h = info.get("Fecha y hora", "").strip()
n_evento = info.get("Evento N°", evento.split()[-1]).strip()
disparo = info.get("Disparo", "").strip()
p_desc = _float(info.get("Potencia desc. [MW]", "0"))

_disp_str = re.sub(r"\by\b", ",", disparo, flags=re.IGNORECASE)
sti_disparo = {x.strip() for x in _disp_str.split(",") if x.strip()}
print(f"    Disparo          : {disparo}")
print(f"    Potencia desc.   : {p_desc} MW  |  Unidades STI: {sorted(sti_disparo)}")

df_gen_raw = pd.read_excel(sim_path, sheet_name="Generadores_pgini")
hora_evento = None
for col in df_gen_raw.columns:
    s = _norm_hora(str(col))
    if re.match(r"^\d{2}:\d{2}$", s) and not s.endswith(":00"):
        hora_evento = s
        break
if hora_evento is None:
    hora_evento = hora_po

print(f"    Hora Po (cargas): {hora_po}")
print(f"    Hora evento (gen): {hora_evento}")


# =============================================================================
# [2] LEER POTENCIAS DEL EVENTO
# =============================================================================

print("\n[2/5] Leyendo potencias de despacho y P0_inicial...")

try:
    df_p0 = pd.read_excel(sim_path, sheet_name="P0_inicial")
    dict_p0 = {
        str(r["Unidad"]).strip(): _float(r["P_0_MW"])
        for _, r in df_p0.iterrows()
        if pd.notna(r.get("P_0_MW"))
    }
except Exception:
    dict_p0 = {}
print(f"    P0 medido: {len(dict_p0)} unidades")

_norm_gen_cols = {_norm_hora(str(c)): c for c in df_gen_raw.columns}
_raw_gen_ev = _norm_gen_cols.get(hora_evento) or _norm_gen_cols.get(hora_po)
col_gen_ev = _raw_gen_ev if _raw_gen_ev is not None else (
    hora_evento if hora_evento in df_gen_raw.columns else hora_po
)

dict_pgini_cndc = {}
dict_estado = {}
for _, r in df_gen_raw.iterrows():
    cndc = str(r.get("Generador_CNDC", "")).strip()
    if not cndc:
        continue
    dict_pgini_cndc[cndc] = _float(r.get(col_gen_ev, 0))
    if "Estado" in df_gen_raw.columns:
        dict_estado[cndc] = str(r.get("Estado", "")).strip()

df_car_raw = pd.read_excel(sim_path, sheet_name="Cargas_plini")
cols_hora_c = [c for c in df_car_raw.columns if re.match(r"^\d{2}:\d{2}$", str(c))]

_df_dn = pd.read_excel(LOC_CAR_PATH, sheet_name="Deener_Nodos")
if not _df_dn.empty and "nombre" in _df_dn.columns and "subdist" in _df_dn.columns:
    _deener_map = {}
    for _, _r in _df_dn.iterrows():
        _n = str(_r["nombre"]).strip()
        _s = str(_r["subdist"]).strip()
        if _n and _s and _n != "nan" and _s != "nan":
            _deener_map.setdefault(_n, []).append(_s)
    print(f"    Mapeo Deener_Nodos cargado: {len(_deener_map)} nodos")
else:
    _deener_map = {}

perfil_dist = {}
_seen_count = {}
for _, r in df_car_raw.iterrows():
    nodo = str(r.get("Nodo_CNDC", "")).strip()
    if not nodo:
        continue
    if "aislados" in nodo.lower():
        continue

    if nodo in _deener_map:
        occ = _seen_count.get(nodo, 0)
        lista = _deener_map[nodo]
        dist = lista[occ] if occ < len(lista) else lista[-1]
        _seen_count[nodo] = occ + 1
    else:
        dist = nodo

    if dist not in perfil_dist:
        perfil_dist[dist] = {h: 0.0 for h in cols_hora_c}
    for h in cols_hora_c:
        perfil_dist[dist][h] += _float(r.get(h, 0))

col_car_ev = hora_po if hora_po in cols_hora_c else (cols_hora_c[0] if cols_hora_c else None)
dict_plini_dist = {dist: vals[col_car_ev] for dist, vals in perfil_dist.items()} if col_car_ev else {}
p_dem_bloque = sum(dict_plini_dist.values())
print(f"    Distribuidores activos (sin aislados): {len(perfil_dist)}")
if col_car_ev:
    print(f"    Pdem bloque {col_car_ev}: {p_dem_bloque:.1f} MW  (datos originales CNDC, sin modificar)")
    if p_dem_bloque > 0:
        print("    [INFO] La demanda de entrada se mantiene como referencia bloque horario.")
        print("           El escalado por distribuidor se conservara con trazabilidad en el Excel.")


# =============================================================================
# [3] LEER LOC_NAMES GENERADORES
# =============================================================================

print("\n[3/5] Leyendo loc_names generadores...")
df_mapeo_gen = pd.read_excel(LOC_GEN_PATH, sheet_name="Mapeo_Generadores")
df_detalle_pf = pd.read_excel(LOC_GEN_PATH, sheet_name="Detalle_PF")

dict_pnom_gen = {
    str(r["loc_name PF"]).strip(): _float(r.get("P nom. (MW)", 0))
    for _, r in df_detalle_pf.iterrows()
}
print(f"    {len(dict_pnom_gen)} loc_names PF de generadores cargados")


# =============================================================================
# [4] CALCULAR pgini POR LOC_NAME PF
# =============================================================================

print("\n[4/5] Calculando pgini por unidad PF...")

filas_gen = []
_mant_info = []   # (cndc, loc_names, total_p_cndc)

for _, row in df_mapeo_gen.iterrows():
    cndc = str(row.get("Generador_CNDC", "")).strip()
    loc_str = str(row.get("loc_names PF", "-")).strip()
    if not cndc or loc_str in ("-", "nan", ""):
        continue

    loc_names = [x.strip() for x in loc_str.split(",") if x.strip()]
    total_p = dict_pgini_cndc.get(cndc, 0.0)
    estado = dict_estado.get(cndc, "")
    mant = "mantenimiento" in estado.lower()

    if mant:
        _mant_info.append((cndc, loc_names, total_p))
        for lp in loc_names:
            filas_gen.append({
                "loc_name PF": lp,
                "Generador_CNDC": cndc,
                "pgini_MW": 0.0,
                "Fuente": "mantenimiento",
                "Estado": estado,
            })
        continue

    p0_ind = {}
    for lp in loc_names:
        sti = extraer_sti(lp)
        p0 = buscar_p0(sti, dict_p0)
        if p0 is not None:
            p0_ind[lp] = p0

    _sti_bases = {re.sub(r"\d+$", "", extraer_sti(lp)) for lp in loc_names}
    solo_p0 = bool(_sti_bases & _COBEE_ZONGO_STI)

    if len(p0_ind) == len(loc_names):
        for lp in loc_names:
            filas_gen.append({
                "loc_name PF": lp,
                "Generador_CNDC": cndc,
                "pgini_MW": round(p0_ind[lp], 4),
                "Fuente": "P0_medido",
                "Estado": estado,
            })
    elif solo_p0:
        for lp in loc_names:
            if lp in p0_ind:
                filas_gen.append({
                    "loc_name PF": lp,
                    "Generador_CNDC": cndc,
                    "pgini_MW": round(p0_ind[lp], 4),
                    "Fuente": "P0_medido",
                    "Estado": estado,
                })
            else:
                filas_gen.append({
                    "loc_name PF": lp,
                    "Generador_CNDC": cndc,
                    "pgini_MW": 0.0,
                    "Fuente": "sin_despacho",
                    "Estado": estado,
                })
    else:
        sum_p0_ind = sum(p0_ind.values())
        restante = max(0.0, total_p - sum_p0_ind)
        lp_sin_p0 = [lp for lp in loc_names if lp not in p0_ind]
        pnom_sin_p0 = sum(dict_pnom_gen.get(lp, 0.0) for lp in lp_sin_p0)

        for lp in loc_names:
            if lp in p0_ind:
                filas_gen.append({
                    "loc_name PF": lp,
                    "Generador_CNDC": cndc,
                    "pgini_MW": round(p0_ind[lp], 4),
                    "Fuente": "P0_medido",
                    "Estado": estado,
                })
            else:
                p_nom_lp = dict_pnom_gen.get(lp, 0.0)
                if pnom_sin_p0 > 0:
                    share = restante * (p_nom_lp / pnom_sin_p0)
                else:
                    share = restante / len(lp_sin_p0) if lp_sin_p0 else 0.0
                filas_gen.append({
                    "loc_name PF": lp,
                    "Generador_CNDC": cndc,
                    "pgini_MW": round(share, 4),
                    "Fuente": "CNDC_proporcional",
                    "Estado": estado,
                })

df_pgini = pd.DataFrame(filas_gen)
if df_pgini.empty:
    df_pgini = pd.DataFrame(columns=["loc_name PF", "Generador_CNDC", "pgini_MW", "Fuente", "Estado"])

n_p0real = (df_pgini["Fuente"] == "P0_medido").sum() if not df_pgini.empty else 0
n_mant = (df_pgini["Fuente"] == "mantenimiento").sum() if not df_pgini.empty else 0
n_pend = (df_pgini["Fuente"] == "CNDC_proporcional").sum() if not df_pgini.empty else 0
n_sd = (df_pgini["Fuente"] == "sin_despacho").sum() if not df_pgini.empty else 0

print(f"    {len(df_pgini)} unidades PF")
print(f"      Verde    (P0 medido):          {n_p0real}")
print(f"      Amarillo (CNDC proporcional):  {n_pend}")
print(f"      Gris     (sin despacho):       {n_sd}")
print(f"      Rojo     (mantenimiento):       {n_mant}")

if _mant_info:
    _activas_mant = [(c, lps, tp) for c, lps, tp in _mant_info if tp > 0]
    print(f"\n  Unidades en mantenimiento: {len(_mant_info)} generadores CNDC")
    print(f"  {'Generador CNDC':<28} {'Despacho CNDC':>13}  loc_names PF")
    print(f"  {'─'*28} {'─'*13}  {'─'*30}")
    for _c, _lps, _tp in _mant_info:
        _nota = "  <- ACTIVA antes del evento" if _tp > 0 else ""
        print(f"  {_c:<28} {_tp:>11.2f} MW  {', '.join(_lps)}{_nota}")
    print(f"  {'─'*28} {'─'*13}")
    if _activas_mant:
        _sum_act = sum(tp for _, _, tp in _activas_mant)
        print(f"  Activas con despacho > 0: {len(_activas_mant)}  |  total = {_sum_act:.2f} MW")
        print(f"  -> Al cargar en PF verificar que sus composite models queden deshabilitados.")


# =============================================================================
# [5] CALCULAR plini Y CURVA CARACTERISTICA DE CARGAS
# =============================================================================

print("\n[5/5] Calculando plini y curva caracteristica de cargas...")

# Compatible con v5 (Curvas_LocNames, una fila por carga) y v4 (LocNames_por_Dist)
_ALIAS_DIST = {}
_norm_perfil = {_norm(k): k for k in perfil_dist}

def _resolver_dist(dist_raw):
    """Devuelve la clave exacta en perfil_dist."""
    if dist_raw in perfil_dist:
        return dist_raw
    n = _norm(dist_raw)
    if n in _ALIAS_DIST:
        return _ALIAS_DIST[n]
    if n in _norm_perfil:
        return _norm_perfil[n]
    matches = difflib.get_close_matches(n, _norm_perfil.keys(), n=1, cutoff=0.85)
    if matches:
        return _norm_perfil[matches[0]]
    return None

import openpyxl as _opx_check
_wb_check = _opx_check.load_workbook(LOC_CAR_PATH, read_only=True)
_hojas_car = _wb_check.sheetnames
_wb_check.close()

if "Curvas_LocNames" in _hojas_car:
    df_curvas = pd.read_excel(LOC_CAR_PATH, sheet_name="Curvas_LocNames")
    df_mapeo_car = pd.read_excel(LOC_CAR_PATH, sheet_name="Mapeo_Cargas")

    filas_car = []
    filas_perfil = []
    _dists_vistos = set()
    _omitidas = []

    for _, row in df_curvas.iterrows():
        dist = str(row.get("Distribuidor / C.N.R.", "")).strip()
        lp = str(row.get("loc_name (PF)", "")).strip()
        pnom = _float(row.get("P nom. (MW)", 0))

        if lp in ("nan", ""):
            _omitidas.append(f"  loc_name vacío  (dist={dist!r})")
            continue
        dist_key = _resolver_dist(dist)
        if dist_key is None:
            _omitidas.append(f"  dist no encontrado: {dist!r}  ->  loc_name={lp!r}")
            continue
        if dist_key != dist:
            print(f"    [aprox] {dist!r} -> {dist_key!r}  ({lp})")

        if dist_key not in _dists_vistos:
            fila_p = {"Distribuidor": dist_key}
            for h in cols_hora_c:
                fila_p[h] = round(perfil_dist[dist_key][h], 3)
            filas_perfil.append(fila_p)
            _dists_vistos.add(dist_key)

        filas_car.append({
            "loc_name PF": lp,
            "Distribuidor": dist_key,
            "P_nom_MW": round(pnom, 4),
        })

    df_plini = pd.DataFrame(filas_car)
    df_perfil = pd.DataFrame(filas_perfil)

    if _omitidas:
        print(f"    AVISO: {len(_omitidas)} cargas omitidas del mapeo:")
        for msg in _omitidas:
            print(msg)
        dists_faltantes = sorted({
            m.split("dist no encontrado:")[1].split("->")[0].strip().strip("'")
            for m in _omitidas if "dist no encontrado" in m
        })
        if dists_faltantes:
            print(f"\n    Distribuidores presentes en perfil_dist (CNDC):")
            for d in sorted(perfil_dist):
                print(f"      {d!r}")
            print(f"\n    Distribuidores faltantes (no coinciden con CNDC):")
            for d in dists_faltantes:
                print(f"      {d!r}  <- revisar nombre exacto en loc_name_cargas.xlsx")

    bus_map = {
        str(r["Nombre carga (PF)"]).strip(): str(r["Barra PF"]).strip()
        for _, r in df_mapeo_car.iterrows()
        if "Barra PF" in df_mapeo_car.columns
    }
    df_plini["Barra PF"] = df_plini["loc_name PF"].map(bus_map)

else:
    df_dist_map = pd.read_excel(LOC_CAR_PATH, sheet_name="LocNames_por_Dist")
    df_mapeo_car = pd.read_excel(LOC_CAR_PATH, sheet_name="Mapeo_Cargas")

    dict_pnom_car = {
        str(r["Nombre carga (PF)"]): _float(r.get("P nom. (MW)", 0))
        for _, r in df_mapeo_car.iterrows()
    }

    filas_car = []
    filas_perfil = []
    _omitidas = []

    for _, row in df_dist_map.iterrows():
        dist = str(row.get("Distribuidor / C.N.R.", "")).strip()
        loc_str = str(row.get("loc_names (en servicio)", "")).strip()
        p_nom_t = _float(row.get("P nom. en servicio (MW)", 0))

        if loc_str in ("nan", ""):
            _omitidas.append(f"  loc_names vacío  (dist={dist!r})")
            continue
        if p_nom_t == 0:
            _omitidas.append(f"  P_nom=0  (dist={dist!r})")
            continue
        dist_key = _resolver_dist(dist)
        if dist_key is None:
            _omitidas.append(f"  dist no encontrado: {dist!r}")
            continue
        if dist_key != dist:
            print(f"    [aprox] {dist!r} -> {dist_key!r}")

        fila_p = {"Distribuidor": dist_key}
        for h in cols_hora_c:
            fila_p[h] = round(perfil_dist[dist_key][h], 3)
        filas_perfil.append(fila_p)

        loc_names = [x.strip() for x in loc_str.split(",") if x.strip()]
        for lp in loc_names:
            pnom = dict_pnom_car.get(lp, 0.0)
            filas_car.append({
                "loc_name PF": lp,
                "Distribuidor": dist_key,
                "P_nom_MW": round(pnom, 4),
            })

    df_plini = pd.DataFrame(filas_car)
    df_perfil = pd.DataFrame(filas_perfil)

    if _omitidas:
        print(f"    AVISO: {len(_omitidas)} distribuidores/cargas omitidos del mapeo:")
        for msg in _omitidas:
            print(msg)
        dists_faltantes = sorted({
            m.split("dist no encontrado:")[1].strip().strip("'")
            for m in _omitidas if "dist no encontrado" in m
        })
        if dists_faltantes:
            print(f"\n    Distribuidores presentes en perfil_dist (CNDC):")
            for d in sorted(perfil_dist):
                print(f"      {d!r}")
            print(f"\n    Distribuidores faltantes (no coinciden con CNDC):")
            for d in dists_faltantes:
                print(f"      {d!r}  <- revisar nombre en loc_name_cargas.xlsx")

    bus_map = {
        str(r["Nombre carga (PF)"]).strip(): str(r["Barra PF"]).strip()
        for _, r in df_mapeo_car.iterrows()
        if "Barra PF" in df_mapeo_car.columns
    }
    df_plini["Barra PF"] = df_plini["loc_name PF"].map(bus_map)

if df_plini.empty:
    df_plini = pd.DataFrame(columns=["loc_name PF", "Distribuidor", "P_nom_MW", "Barra PF", "plini_base_MW", "plini_MW"])

# plini_i = P_dem_dist × (P_nom_i / Σ P_nom_dist)
# Se conserva una version exacta por carga y luego se redondea una sola vez
# con el metodo de mayor residuo.
pnom_por_dist = df_plini.groupby("Distribuidor")["P_nom_MW"].sum().to_dict() if not df_plini.empty else {}
ratio_por_dist = {
    dist: (dict_plini_dist.get(dist, 0.0) / pnom_total) if pnom_total > 0 else 0.0
    for dist, pnom_total in pnom_por_dist.items()
}
df_plini["plini_exact_MW"] = (
    df_plini["Distribuidor"].map(ratio_por_dist).fillna(0.0) * df_plini["P_nom_MW"]
) if not df_plini.empty else []
df_plini["plini_base_MW"] = df_plini["plini_exact_MW"].round(PRECISION_PLINI) if not df_plini.empty else []
df_plini["plini_MW"] = df_plini["plini_base_MW"].copy() if not df_plini.empty else []

df_balance_plini, max_residuo_plini = _ajustar_plini_por_distribuidor(df_plini, dict_plini_dist)

print(f"    {len(df_plini)} cargas PF  |  {len(filas_perfil)} distribuidores")
if col_car_ev:
    print(f"    Demanda total en {col_car_ev}: {sum(dict_plini_dist.values()):.1f} MW  (datos originales CNDC, sin modificar)")
    print(f"    Pdem asignada (proporcional):  {df_plini['plini_MW'].sum():.1f} MW")
    print(f"    Residuo maximo por distribuidor: {max_residuo_plini:.4f} MW")
    if max_residuo_plini > 0.0001:
        print("    [AVISO] Existe residuo de redondeo en al menos un distribuidor.")
        print("           Revisar la hoja Balance_Plini para trazabilidad completa.")


# =============================================================================
# EXPORTAR EXCEL
# =============================================================================

output_path = os.path.join(ev_path, f"condiciones_iniciales_{fecha_str}_Ev{n_evento}.xlsx")

df_resumen = pd.DataFrame([
    {"Parametro": "Semestre", "Valor": semestre},
    {"Parametro": "Evento", "Valor": evento},
    {"Parametro": "Fecha y hora", "Valor": fecha_h},
    {"Parametro": "Disparo", "Valor": disparo},
    {"Parametro": "Hora Po (cargas)", "Valor": hora_po},
    {"Parametro": "Hora evento (gen)", "Valor": hora_evento},
    {"Parametro": "Unidades PF (total)", "Valor": len(df_pgini)},
    {"Parametro": "  P0 medido", "Valor": int(n_p0real)},
    {"Parametro": "  CNDC proporcional", "Valor": int(n_pend)},
    {"Parametro": "  sin_despacho", "Valor": int(n_sd)},
    {"Parametro": "  Mantenimiento", "Valor": int(n_mant)},
    {"Parametro": "Pgen CNDC (MW)", "Valor": round(df_pgini["pgini_MW"].sum(), 2)},
    {"Parametro": "Cargas PF (total)", "Valor": len(df_plini)},
    {"Parametro": "Distribuidores", "Valor": len(filas_perfil)},
    {"Parametro": f"Pdem bloque {col_car_ev} (MW)", "Valor": round(p_dem_bloque, 2) if col_car_ev else 0.0},
    {"Parametro": "Pdem asignada proporcional (MW)", "Valor": round(df_plini["plini_MW"].sum(), 2)},
    {"Parametro": "Residuo maximo plini por dist (MW)", "Valor": round(max_residuo_plini, 4)},
])

print(f"\nExportando a {os.path.basename(output_path)}...")
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df_resumen.to_excel(writer, sheet_name="Resumen", index=False)
    df_pgini.to_excel(writer, sheet_name="pgini_GEN", index=False)
    df_plini.to_excel(writer, sheet_name="plini_CAR", index=False)
    df_perfil.to_excel(writer, sheet_name="Perfil_MW_Dist", index=False)
    df_balance_plini.to_excel(writer, sheet_name="Balance_Plini", index=False)

wb = load_workbook(output_path)

for sh in wb.sheetnames:
    _formato_hoja(wb[sh])

# Color por fuente en pgini_GEN
ws_g = wb["pgini_GEN"]
fuente_col = next((c for c in range(1, ws_g.max_column + 1) if str(ws_g.cell(1, c).value) == "Fuente"), None)
if fuente_col:
    for r in range(2, ws_g.max_row + 1):
        fuente = str(ws_g.cell(r, fuente_col).value or "")
        if "mantenimiento" in fuente:
            fill = FILL_MANT
        elif "sin_despacho" in fuente:
            fill = FILL_SD
        elif "P0_medido" in fuente:
            fill = FILL_P0
        else:
            fill = FILL_PROP
        for c in range(1, ws_g.max_column + 1):
            ws_g.cell(r, c).fill = fill

# Color simple en Balance_Plini
if "Balance_Plini" in wb.sheetnames:
    ws_b = wb["Balance_Plini"]
    col_res = next((c for c in range(1, ws_b.max_column + 1) if str(ws_b.cell(1, c).value) == "Residuo_MW"), None)
    if col_res:
        for r in range(2, ws_b.max_row + 1):
            residuo = abs(float(ws_b.cell(r, col_res).value or 0))
            fill = FILL_BAL_OK if residuo < 0.0001 else FILL_BAL_WARN
            for c in range(1, ws_b.max_column + 1):
                ws_b.cell(r, c).fill = fill

wb.save(output_path)

# =============================================================================
# RESUMEN FINAL
# =============================================================================

print(f"\n{'='*60}")
print(f"  Archivo: {os.path.basename(output_path)}")
print(f"  Ruta   : {ev_path}")
print(f"{'='*60}")
print(f"  pgini_GEN: {len(df_pgini)} unidades PF")
print(f"    Verde    P0 medido:         {n_p0real}")
print(f"    Amarillo CNDC proporcional: {n_pend}")
print(f"    Gris     sin_despacho:      {n_sd}")
print(f"    Rojo     Mantenimiento:     {n_mant}")
print(f"    Pgen CNDC = {df_pgini['pgini_MW'].sum():.1f} MW")
print(f"  plini_CAR: {len(df_plini)} cargas PF")
if col_car_ev:
    print(f"    Pdem bloque  {col_car_ev}  : {p_dem_bloque:.1f} MW  (datos originales CNDC)")
    print(f"    Pdem asignada proporcional: {df_plini['plini_MW'].sum():.1f} MW")
    print(f"    Residuo maximo por dist    : {max_residuo_plini:.4f} MW")
print(f"  Balance_Plini hojas         : {len(df_balance_plini)} distribuidores")
print(f"{'='*60}")

input("\nPresiona Enter para cerrar...")
