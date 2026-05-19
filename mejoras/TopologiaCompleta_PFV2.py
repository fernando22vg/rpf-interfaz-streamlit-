# -*- coding: utf-8 -*-
"""
TopologiaCompleta_PF.py

Extraccion y verificacion de la topologia completa de un sistema electrico
modelado en DIgSILENT PowerFactory.

Salida principal:
- Excel con catalogo de barras
- Elementos de red
- Relaciones topologicas bus-bus / bus-elemento
- Componentes conectados
- Matriz de vecindad topologica
- Incidencias de conectividad
- Resumen de verificacion

El script esta pensado para ejecutarse dentro del entorno de PowerFactory
o desde un Python configurado con el runtime de PowerFactory.
"""

import json
import os
import sys
from itertools import combinations
from collections import defaultdict, deque

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURACION
# =============================================================================
PF_DIR = r"C:\Program Files\DIgSILENT\PowerFactory 2025 SP2"
PF_PY = r"C:\Program Files\DIgSILENT\PowerFactory 2025 SP2\Python\3.12"

# Proyecto PowerFactory usado por los scripts de extraccion.
# Si quieres usar el proyecto activo, deja este valor vacio.
PROJECT_NAME = "PMP_NOV25_OCT29_31102025(1)"

# Si True, ejecuta el flujo de carga base antes de extraer resultados (m:P, m:u, etc.)
# Poner False si el LF ya fue ejecutado manualmente en PowerFactory.
RUN_LF = True

OUTPUT_DIR = r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\Topologia"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "topologia_completa_pf.xlsx")
OUTPUT_GRAPH_JSON = os.path.join(OUTPUT_DIR, "topologia_grafo.json")

# Clases que se intentan leer para cubrir la red principal y la conexion de equipos.
ELEMENT_PATTERNS = [
    "*.ElmLne",
    "*.ElmTr2",
    "*.ElmTr3",
    "*.ElmLod",
    "*.ElmSym",
    "*.ElmGenstat",
    "*.ElmPvsys",
    "*.ElmWind",
    "*.ElmShnt",
    "*.ElmShunt",
    "*.ElmSvs",
    "*.ElmCoup",
    "*.ElmFuse",
    "*.StaSwitch",
]

# Colores por nivel de tension (hex sin #) — coincide con loc_namesLineas.py
_KV_COLORS = {
    500: "D9B3FF",
    230: "BDD7EE",
    115: "C5E0B4",
    69:  "FFE699",
    44:  "FCE4D6",
    25:  "F4CCCC",
    24:  "F4CCCC",
}


def _kv_hex(kv):
    """Devuelve color hex (#RRGGBB) para un nivel de tension dado."""
    if kv is None:
        return "#EDEDED"
    try:
        kv = float(kv)
    except Exception:
        return "#EDEDED"
    for nivel in sorted(_KV_COLORS.keys(), reverse=True):
        if kv >= nivel * 0.9:
            return "#" + _KV_COLORS[nivel]
    return "#EDEDED"

ZONE_ATTRS = ("zone", "pZone", "cpZone")
TERMINAL_ATTRS = ("bus1", "bus2", "bus3")

# Umbral de tension para ayudar a clasificar elementos si falta informacion
UMBRAL_STI_KV = 60.0

BUS_COLUMNS = [
    "Nombre",
    "Tension nom. (kV)",
    "Tension (pu)",
    "Tension (kV)",
    "Angulo (deg)",
    "P inyectada (MW)",
    "Q inyectada (Mvar)",
    "Zona",
    "En servicio",
    "Clase PF",
]

ELEMENT_COLUMNS = [
    "Nombre",
    "Clase PF",
    "Categoria",
    "En servicio",
    "N terminales",
    "Terminales",
    "Barra 1",
    "Barra 2",
    "Barra 3",
    "Barra conectada",
    "Longitud (km)",
    "Tension nom. (kV)",
    "Tension 1 nom. (kV)",
    "Tension 2 nom. (kV)",
    "Tension 3 nom. (kV)",
    "Potencia nom. (MVA)",
    "Corriente nom. (A)",
    "Carga (%)",
    "P from (MW)",
    "Q from (Mvar)",
    "P to (MW)",
    "Q to (Mvar)",
    "P lado 1 (MW)",
    "Q lado 1 (Mvar)",
    "P lado 2 (MW)",
    "Q lado 2 (Mvar)",
    "P lado 3 (MW)",
    "Q lado 3 (Mvar)",
    "P nom. (MW)",
    "Q nom. (Mvar)",
    "P result. (MW)",
    "Q result. (Mvar)",
    "Perdidas P (MW)",
    "Tap pos.",
    "Cos phi",
    "Estado cierre",
    "Posicion",
    "Tension (pu)",
    "Tipo",
    # Atributos de perdidas de transformadores (TypTr2)
    "uktr (%)",
    "curmg (%)",
    "pfe (kW)",
    "pcutr (kW)",
    # Atributos de diagnostico de generadores
    "pgini (MW)",
    "Pmax (MW)",
    "Pmin (MW)",
    "ip_ctrl",
    "ngnum",
]

RELATION_COLUMNS = [
    "TipoRelacion",
    "Elemento",
    "Clase PF",
    "Categoria",
    "Barra 1",
    "Barra 2",
    "Conexion",
    "Relacion",
    "En servicio",
]

ISSUE_COLUMNS = [
    "Elemento",
    "Clase PF",
    "Tipo incidencia",
    "Detalle",
]

COMPONENT_COLUMNS = [
    "Componente",
    "N barras",
    "Barra inicial",
    "Barras",
    "Aislada",
]

ADJACENCY_COLUMNS = [
    "Barra",
    "Componente",
    "N vecinos",
    "Vecinos",
    "N elementos asociados",
    "Elementos asociados",
    "Zona",
    "Tension nom. (kV)",
    "Tension (pu)",
    "En servicio",
    "Barra aislada",
]

# =============================================================================
# ESTILOS EXCEL
# =============================================================================
THIN = Side(border_style="thin", color="BFBFBF")
BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")

H_FILL = PatternFill("solid", start_color="1F3864")
H_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
N_FONT = Font(name="Arial", size=10)
B_FONT = Font(name="Arial", bold=True, size=10)


def _fill(color):
    return PatternFill("solid", start_color=color, end_color=color)


# =============================================================================
# HELPERS PowerFactory
# =============================================================================
def _setup_pf():
    sys.path.append(PF_PY)
    os.environ["PATH"] = PF_DIR + os.pathsep + os.environ.get("PATH", "")
    try:
        os.add_dll_directory(PF_DIR)
    except Exception:
        pass

    import powerfactory as pf  # noqa: E402

    app = None
    try:
        app = pf.GetApplication()
    except Exception:
        app = None

    if app is None:
        try:
            app = pf.GetApplicationExt()
        except Exception:
            app = None

    if app is None:
        raise RuntimeError(
            "No se pudo obtener la aplicacion de PowerFactory. "
            "Cierre PowerFactory si esta abierto y vuelva a ejecutar el script."
        )

    app.Show()

    if PROJECT_NAME:
        try:
            res = app.ActivateProject(PROJECT_NAME)
            print(f"Proyecto activo: {res}")
        except Exception as exc:
            print(f"[AVISO] No se pudo activar el proyecto '{PROJECT_NAME}': {exc}")

    try:
        active = app.GetActiveProject()
        if active is not None:
            print(f"Proyecto actual: {active.loc_name}")
        elif PROJECT_NAME:
            print("[AVISO] No hay proyecto activo; se intentara usar el proyecto configurado.")
    except Exception:
        pass

    # Ejecutar flujo de carga base (igual que DatsoGENBUSLNE.py)
    if RUN_LF:
        try:
            ldf = app.GetFromStudyCase("ComLdf")
            if ldf is not None:
                ldf.Execute()
                print("Flujo de carga base ejecutado.")
            else:
                print("[AVISO] No se encontro ComLdf en el caso de estudio activo.")
        except Exception as exc:
            print(f"[AVISO] No se pudo ejecutar el flujo de carga: {exc}")

    return app


def ga(obj, attr, default=None):
    try:
        if obj is None:
            return default
        val = obj.GetAttribute(attr)
        return default if val is None else val
    except Exception:
        return default


def ga_round(obj, attr, dec=4, default=None):
    val = ga(obj, attr, default=None)
    try:
        if val is None:
            return default
        return round(float(val), dec)
    except Exception:
        return default if default is not None else val


def _obj_class(obj):
    try:
        return obj.GetClassName()
    except Exception:
        return type(obj).__name__


def serv(obj):
    return "Si" if ga(obj, "outserv", 0) == 0 else "No"


def _first_non_empty(obj, attrs, default=None):
    for attr in attrs:
        val = ga(obj, attr, None)
        if val not in (None, "", "nan"):
            return val
    return default


def _resolve_terminal(obj, attr):
    """
    Resuelve un terminal PowerFactory a partir de un atributo tipo bus1/bus2/bus3.
    Patron DatsoGENBUSLNE: cubicle->cterm primero, luego notacion de dos puntos.
    """
    if obj is None:
        return None

    # Metodo primario: cubicle -> cterm (igual que barra_de en DatsoGENBUSLNE.py)
    try:
        cub = obj.GetAttribute(attr)
        if cub is not None:
            if _obj_class(cub) == "ElmTerm":
                return cub
            try:
                term = cub.GetAttribute("cterm")
                if term is not None and ga(term, "loc_name", None) is not None:
                    return term
            except Exception:
                pass
    except Exception:
        pass

    # Fallback: notacion directa attr:cterm
    direct = ga(obj, f"{attr}:cterm", None)
    if direct is not None and ga(direct, "loc_name", None) is not None:
        return direct

    return None


# Atributos alternativos de terminal para ElmTr3 (algunos modelos PF usan bushv/busmv/buslv)
_TR3_TERMINAL_ATTRS_ALT = ("bushv", "busmv", "buslv")


def _terminal_name(term):
    return str(ga(term, "loc_name", "") or "").strip()


def _terminal_kv(term, bus_kv_map=None):
    if term is None:
        return None
    kv = ga(term, "uknom", None)
    if kv is not None:
        try:
            return float(kv)
        except Exception:
            pass
    if bus_kv_map:
        return bus_kv_map.get(_terminal_name(term))
    return None


def _terminal_zone(term, bus_zone_map=None):
    if term is None:
        return None
    for attr in ZONE_ATTRS:
        z = ga(term, attr, None)
        if z is not None:
            name = ga(z, "loc_name", None)
            if name:
                return str(name).strip()

    if bus_zone_map:
        return bus_zone_map.get(_terminal_name(term))
    return None


def _netdat_folder(app):
    try:
        return app.GetProjectFolder("netdat")
    except Exception:
        return None


def _dedupe_objects(objs):
    result = []
    seen = set()
    for obj in objs or []:
        if obj is None:
            continue
        oid = id(obj)
        if oid in seen:
            continue
        seen.add(oid)
        result.append(obj)
    return result


def _collect_pf_objects(app, pattern):
    """
    Recupera objetos PowerFactory con estrategia de respaldo múltiple.

    Orden de búsqueda (igual que DatsoGENBUSLNE.py como fuente primaria):
      1. GetCalcRelevantObjects — necesario para atributos de resultado (m:P, m:u…)
      2. Árbol completo del proyecto activo (fallback topologico)
      3. Carpeta netdat (fallback adicional)
    """
    candidatos = []

    # 1) Caso de estudio activo — fuente primaria, igual que referencia DatsoGENBUSLNE
    try:
        candidatos.extend(app.GetCalcRelevantObjects(pattern) or [])
    except Exception:
        pass

    # 2) Proyecto activo completo (fallback si GetCalcRelevantObjects devuelve vacío)
    if not candidatos:
        try:
            active_project = app.GetActiveProject()
            if active_project is not None:
                candidatos.extend(active_project.GetContents(pattern, 1) or [])
        except Exception:
            pass

    # 3) Carpeta netdat (fallback final)
    if not candidatos:
        netdat = _netdat_folder(app)
        if netdat is not None:
            try:
                candidatos.extend(netdat.GetContents(pattern, 1) or [])
            except Exception:
                pass

    return _dedupe_objects(candidatos)


def _bus_zone_map(app):
    """
    Mapa de barras -> zona a partir de ElmZone y de atributos directos del bus.
    """
    result = {}

    zones = _collect_pf_objects(app, "*.ElmZone")
    for z in zones:
        zname = str(ga(z, "loc_name", "") or "").strip()
        if not zname:
            continue
        try:
            terms = z.GetContents("*.ElmTerm", 1) or []
            for t in terms:
                tname = _terminal_name(t)
                if tname:
                    result[tname] = zname
        except Exception:
            pass

    return result


def _bus_zone(term, bus_zone_map):
    if term is None:
        return None

    direct = _terminal_zone(term, None)
    if direct:
        return direct

    tname = _terminal_name(term)
    return bus_zone_map.get(tname)


def _bus_voltage(term):
    return ga(term, "uknom", None)


def _bus_voltage_pu(term):
    return ga(term, "m:u", None)


def _is_bus(obj):
    return _obj_class(obj) == "ElmTerm"


def _line_like_class(cls):
    return cls in {"ElmLne", "ElmCoup", "ElmFuse", "StaSwitch"}


def _transformer_like_class(cls):
    return cls in {"ElmTr2", "ElmTr3"}


def _attachment_like_class(cls):
    return cls in {
        "ElmLod",
        "ElmSym",
        "ElmGenstat",
        "ElmPvsys",
        "ElmWind",
        "ElmShnt",
        "ElmShunt",
        "ElmSvs",
    }


def _element_category(cls):
    if cls == "ElmTerm":
        return "Barra"
    if cls == "ElmLne":
        return "Linea"
    if cls == "ElmTr2":
        return "Transformador_2dev"
    if cls == "ElmTr3":
        return "Transformador_3dev"
    if cls == "ElmLod":
        return "Carga"
    if cls in {"ElmSym", "ElmGenstat", "ElmPvsys", "ElmWind"}:
        return "Generador"
    if cls in {"ElmShnt", "ElmShunt"}:
        return "Shunt"
    if cls == "ElmSvs":
        return "Compensador"
    if cls in {"ElmCoup", "ElmFuse", "StaSwitch"}:
        return "Interruptor"
    return "Otro"


def _ordered_terminals(obj):
    ordered = []
    for attr in TERMINAL_ATTRS:
        term = _resolve_terminal(obj, attr)
        if term is not None:
            name = _terminal_name(term)
            if name and name not in [n for _, n, _ in ordered]:
                ordered.append((attr, name, term))

    # Para ElmTr3: si no se resolvio nada con bus1/bus2/bus3, probar bushv/busmv/buslv
    if not ordered and _obj_class(obj) == "ElmTr3":
        for attr in _TR3_TERMINAL_ATTRS_ALT:
            term = _resolve_terminal(obj, attr)
            if term is not None:
                name = _terminal_name(term)
                if name and name not in [n for _, n, _ in ordered]:
                    ordered.append((attr, name, term))

    return ordered


def _pair_label(a1, a2):
    return f"{a1.upper()}-{a2.upper()}"


def _safe_float(v, default=None):
    try:
        if v is None:
            return default
        return float(v)
    except Exception:
        return default


def _empty_df(columns):
    return pd.DataFrame(columns=columns)


# =============================================================================
# EXTRACCION DE BARRAS
# =============================================================================
def extract_buses(app):
    bus_zone_map = _bus_zone_map(app)
    buses = _collect_pf_objects(app, "*.ElmTerm")

    rows = []
    bus_index = {}

    for bus in buses:
        name = _terminal_name(bus)
        if not name or name in ("nan", "LEYENDA"):
            continue

        kv = _bus_voltage(bus)
        upu = _bus_voltage_pu(bus)
        zone = _bus_zone(bus, bus_zone_map)

        row = {
            "Nombre": name,
            "Tension nom. (kV)": _safe_float(kv, None),
            "Tension (pu)": _safe_float(upu, None),
            "Tension (kV)": round(_safe_float(kv, 0.0) * _safe_float(upu, 0.0), 4)
            if kv is not None and upu is not None
            else None,
            "Angulo (deg)": ga_round(bus, "m:phiu", 4),
            "P inyectada (MW)": ga_round(bus, "m:P", 4),
            "Q inyectada (Mvar)": ga_round(bus, "m:Q", 4),
            "Zona": zone,
            "En servicio": serv(bus),
            "Clase PF": _obj_class(bus),
        }
        rows.append(row)
        bus_index[name] = row

    df_buses = pd.DataFrame(rows)
    if df_buses.empty:
        df_buses = _empty_df(BUS_COLUMNS)
    else:
        # Normaliza orden de columnas para salida estable
        for col in BUS_COLUMNS:
            if col not in df_buses.columns:
                df_buses[col] = None
        df_buses = df_buses.reindex(columns=BUS_COLUMNS)

    return df_buses, bus_index, bus_zone_map


# =============================================================================
# EXTRACCION DE ELEMENTOS Y RELACIONES TOPOLOGICAS
# =============================================================================
def extract_topology_elements(app, bus_index, bus_zone_map):
    element_rows = []
    relation_rows = []
    issues_rows = []

    bus_neighbors = defaultdict(set)
    bus_attached_elements = defaultdict(set)
    bus_attachment_counts = defaultdict(int)

    all_seen = set()

    for pattern in ELEMENT_PATTERNS:
        objs = _collect_pf_objects(app, pattern)
        for obj in objs:
            try:
                obj_id = id(obj)
                if obj_id in all_seen:
                    continue
                all_seen.add(obj_id)

                cls = _obj_class(obj)
                name = str(ga(obj, "loc_name", "") or "").strip()
                if not name:
                    name = f"({cls})_{obj_id}"

                cat = _element_category(cls)
                state = serv(obj)
                terminals = _ordered_terminals(obj)

                terminal_names = [t[1] for t in terminals]
                terminal_names_unique = []
                for tname in terminal_names:
                    if tname not in terminal_names_unique:
                        terminal_names_unique.append(tname)

                record = {
                    "Nombre": name,
                    "Clase PF": cls,
                    "Categoria": cat,
                    "En servicio": state,
                    "N terminales": len(terminal_names_unique),
                    "Terminales": " | ".join(terminal_names_unique),
                }

                # Campos especificos por tipo
                if cls == "ElmLne":
                    typ = ga(obj, "typ_id", None)
                    record.update(
                        {
                            "Barra 1": terminal_names_unique[0] if len(terminal_names_unique) > 0 else "",
                            "Barra 2": terminal_names_unique[1] if len(terminal_names_unique) > 1 else "",
                            "Longitud (km)": ga_round(obj, "dline", 4),
                            "Tension nom. (kV)": ga_round(typ, "uline", 4),
                            "Corriente nom. (A)": ga_round(typ, "InomAC", 4),
                            "Carga (%)": ga_round(obj, "c:loading", 4),
                            "P from (MW)": ga_round(obj, "m:P:bus1", 4),
                            "Q from (Mvar)": ga_round(obj, "m:Q:bus1", 4),
                            "P to (MW)": ga_round(obj, "m:P:bus2", 4),
                            "Q to (Mvar)": ga_round(obj, "m:Q:bus2", 4),
                            "Perdidas P (MW)": ga_round(obj, "m:Plosses", 4),
                        }
                    )

                elif cls == "ElmTr2":
                    typ = ga(obj, "typ_id", None)
                    record.update(
                        {
                            "Barra 1": terminal_names_unique[0] if len(terminal_names_unique) > 0 else "",
                            "Barra 2": terminal_names_unique[1] if len(terminal_names_unique) > 1 else "",
                            "Tension 1 nom. (kV)": ga_round(typ, "utrn_h", 4),
                            "Tension 2 nom. (kV)": ga_round(typ, "utrn_l", 4),
                            "Potencia nom. (MVA)": ga_round(typ, "strn", 4),
                            "Carga (%)": ga_round(obj, "c:loading", 4),
                            "P lado 1 (MW)": ga_round(obj, "m:P:bus1", 4),
                            "Q lado 1 (Mvar)": ga_round(obj, "m:Q:bus1", 4),
                            "P lado 2 (MW)": ga_round(obj, "m:P:bus2", 4),
                            "Q lado 2 (Mvar)": ga_round(obj, "m:Q:bus2", 4),
                            "Perdidas P (MW)": ga_round(obj, "m:Plosses", 4),
                            "Tap pos.": ga(obj, "nntap", None),
                            # Atributos de perdidas del tipo de transformador
                            "uktr (%)": ga_round(typ, "uktr", 4),
                            "curmg (%)": ga_round(typ, "curmg", 4),
                            "pfe (kW)": ga_round(typ, "pfe", 4),
                            "pcutr (kW)": ga_round(typ, "pcutr", 4),
                        }
                    )

                elif cls == "ElmTr3":
                    typ = ga(obj, "typ_id", None)
                    record.update(
                        {
                            "Barra 1": terminal_names_unique[0] if len(terminal_names_unique) > 0 else "",
                            "Barra 2": terminal_names_unique[1] if len(terminal_names_unique) > 1 else "",
                            "Barra 3": terminal_names_unique[2] if len(terminal_names_unique) > 2 else "",
                            "Tension 1 nom. (kV)": ga_round(typ, "utrn_h", 4),
                            "Tension 2 nom. (kV)": ga_round(typ, "utrn_m", 4),
                            "Tension 3 nom. (kV)": ga_round(typ, "utrn_l", 4),
                            "Potencia nom. (MVA)": ga_round(typ, "strn", 4),
                            "P lado 1 (MW)": ga_round(obj, "m:P:bus1", 4),
                            "Q lado 1 (Mvar)": ga_round(obj, "m:Q:bus1", 4),
                            "P lado 2 (MW)": ga_round(obj, "m:P:bus2", 4),
                            "Q lado 2 (Mvar)": ga_round(obj, "m:Q:bus2", 4),
                            "P lado 3 (MW)": ga_round(obj, "m:P:bus3", 4),
                            "Q lado 3 (Mvar)": ga_round(obj, "m:Q:bus3", 4),
                            "Perdidas P (MW)": ga_round(obj, "m:Plosses", 4),
                        }
                    )

                elif cls == "ElmLod":
                    term = terminal_names_unique[0] if terminal_names_unique else ""
                    record.update(
                        {
                            "Barra conectada": term,
                            "P nom. (MW)": ga_round(obj, "plini", 4),
                            "Q nom. (Mvar)": ga_round(obj, "qlini", 4),
                            "P result. (MW)": ga_round(obj, "m:P:bus1", 4),
                            "Q result. (Mvar)": ga_round(obj, "m:Q:bus1", 4),
                            "Cos phi": ga_round(obj, "coslini", 4),
                        }
                    )

                elif cls in {"ElmSym", "ElmGenstat", "ElmPvsys", "ElmWind"}:
                    term = terminal_names_unique[0] if terminal_names_unique else ""
                    typ = ga(obj, "typ_id", None)
                    pnom = None
                    for attr in ("Pnom", "Pmax"):
                        pnom = ga_round(obj, attr, 4)
                        if pnom not in (None, 0, 0.0):
                            break
                    record.update(
                        {
                            "Barra conectada": term,
                            "P nom. (MW)": pnom,
                            "Q nom. (Mvar)": ga_round(obj, "qgini", 4),
                            "P result. (MW)": ga_round(obj, "m:P:bus1", 4),
                            "Q result. (Mvar)": ga_round(obj, "m:Q:bus1", 4),
                            "Tension (pu)": ga_round(ga(_resolve_terminal(obj, "bus1"), "m:u", None), 4),
                            "Cos phi": ga_round(obj, "cosini", 4),
                            "Tipo": ga(typ, "loc_name", None),
                            # Campos de diagnostico de condiciones iniciales
                            "pgini (MW)": ga_round(obj, "pgini", 4),
                            "Pmax (MW)": ga_round(obj, "Pmax", 4),
                            "Pmin (MW)": ga_round(obj, "Pmin", 4),
                            "ip_ctrl": ga(obj, "ip_ctrl", None),
                            "ngnum": ga(obj, "ngnum", None),
                        }
                    )

                elif cls in {"ElmShnt", "ElmShunt"}:
                    term = terminal_names_unique[0] if terminal_names_unique else ""
                    record.update(
                        {
                            "Barra conectada": term,
                            "Q result. (Mvar)": ga_round(obj, "m:Q:bus1", 4),
                            "P result. (MW)": ga_round(obj, "m:P:bus1", 4),
                        }
                    )

                elif cls == "ElmSvs":
                    term = terminal_names_unique[0] if terminal_names_unique else ""
                    record.update(
                        {
                            "Barra conectada": term,
                            "Q result. (Mvar)": ga_round(obj, "m:Q:bus1", 4),
                            "P result. (MW)": ga_round(obj, "m:P:bus1", 4),
                            "Tension (pu)": ga_round(obj, "usetp", 4),
                        }
                    )

                elif cls in {"ElmCoup", "ElmFuse", "StaSwitch"}:
                    b1 = terminal_names_unique[0] if len(terminal_names_unique) > 0 else ""
                    b2 = terminal_names_unique[1] if len(terminal_names_unique) > 1 else ""
                    record.update(
                        {
                            "Barra 1": b1,
                            "Barra 2": b2,
                            "Estado cierre": ga(obj, "on_off", None),
                            "Posicion": ga(obj, "nstate", None),
                        }
                    )

                element_rows.append(record)

                # =========================
                # RELACIONES TOPOLOGICAS
                # =========================
                if len(terminal_names_unique) == 0:
                    issues_rows.append(
                        {
                            "Elemento": name,
                            "Clase PF": cls,
                            "Tipo incidencia": "Sin terminales",
                            "Detalle": "No se pudo resolver bus1/bus2/bus3",
                        }
                    )
                    continue

                if len(terminal_names_unique) == 1:
                    bus = terminal_names_unique[0]
                    relation_rows.append(
                        {
                            "TipoRelacion": "BUS_ELEM",
                            "Elemento": name,
                            "Clase PF": cls,
                            "Categoria": cat,
                            "Barra 1": bus,
                            "Barra 2": "",
                            "Relacion": "attachment",
                            "En servicio": state,
                        }
                    )
                    bus_attached_elements[bus].add(name)
                    bus_attachment_counts[bus] += 1

                    if bus not in bus_index:
                        issues_rows.append(
                            {
                                "Elemento": name,
                                "Clase PF": cls,
                                "Tipo incidencia": "Barra no encontrada",
                                "Detalle": f"Barra '{bus}' no existe en el catalogo de ElmTerm",
                            }
                        )

                else:
                    # Para conexiones de 3 devanados o elementos multi-terminal:
                    # se guardan pares para reconstruir la red como grafo.
                    pair_attrs = list(zip([t[0] for t in terminals], terminal_names_unique))
                    for (a1, b1), (a2, b2) in combinations(pair_attrs, 2):
                        relation_rows.append(
                            {
                                "TipoRelacion": "BUS_BUS",
                                "Elemento": name,
                                "Clase PF": cls,
                                "Categoria": cat,
                                "Barra 1": b1,
                                "Barra 2": b2,
                                "Conexion": _pair_label(a1, a2),
                                "En servicio": state,
                            }
                        )
                        if b1:
                            bus_neighbors[b1].add(b2)
                            bus_attached_elements[b1].add(name)
                        if b2:
                            bus_neighbors[b2].add(b1)
                            bus_attached_elements[b2].add(name)

                        if b1 not in bus_index:
                            issues_rows.append(
                                {
                                    "Elemento": name,
                                    "Clase PF": cls,
                                    "Tipo incidencia": "Barra no encontrada",
                                    "Detalle": f"Barra '{b1}' no existe en el catalogo de ElmTerm",
                                }
                            )
                        if b2 not in bus_index:
                            issues_rows.append(
                                {
                                    "Elemento": name,
                                    "Clase PF": cls,
                                    "Tipo incidencia": "Barra no encontrada",
                                    "Detalle": f"Barra '{b2}' no existe en el catalogo de ElmTerm",
                                }
                            )

                    if len(terminal_names_unique) == 2:
                        a, b = terminal_names_unique
                        bus_neighbors[a].add(b)
                        bus_neighbors[b].add(a)

            except Exception as exc:
                issues_rows.append(
                    {
                        "Elemento": str(ga(obj, "loc_name", "") or ""),
                        "Clase PF": _obj_class(obj),
                        "Tipo incidencia": "Error de extraccion",
                        "Detalle": str(exc),
                    }
                )

    # Diagnostico de resolucion de terminales para transformadores
    tr_rows = [r for r in element_rows if r.get("Clase PF") in ("ElmTr2", "ElmTr3")]
    tr_sin = sum(1 for r in tr_rows if r.get("N terminales", 0) == 0)
    tr_uno = sum(1 for r in tr_rows if r.get("N terminales", 0) == 1)
    tr_ok  = sum(1 for r in tr_rows if r.get("N terminales", 0) >= 2)
    print(f"  Transformadores: {len(tr_rows)} total  |  "
          f"con 2+ terminales={tr_ok}  |  con 1={tr_uno}  |  sin terminales={tr_sin}")
    if tr_sin > 0:
        ejemplos = [r["Nombre"] for r in tr_rows if r.get("N terminales", 0) == 0][:5]
        print(f"  [AVISO] Transformadores sin terminales (primeros 5): {ejemplos}")

    df_elements = pd.DataFrame(element_rows)
    if df_elements.empty:
        df_elements = _empty_df(ELEMENT_COLUMNS)
    else:
        for col in ELEMENT_COLUMNS:
            if col not in df_elements.columns:
                df_elements[col] = None
        df_elements = df_elements.reindex(columns=ELEMENT_COLUMNS + [c for c in df_elements.columns if c not in ELEMENT_COLUMNS])

    df_relations = pd.DataFrame(relation_rows)
    if df_relations.empty:
        df_relations = _empty_df(RELATION_COLUMNS)
    else:
        for col in RELATION_COLUMNS:
            if col not in df_relations.columns:
                df_relations[col] = None
        df_relations = df_relations.reindex(columns=RELATION_COLUMNS + [c for c in df_relations.columns if c not in RELATION_COLUMNS])

    df_issues = pd.DataFrame(issues_rows)
    if df_issues.empty:
        df_issues = _empty_df(ISSUE_COLUMNS)
    else:
        for col in ISSUE_COLUMNS:
            if col not in df_issues.columns:
                df_issues[col] = None
        df_issues = df_issues.reindex(columns=ISSUE_COLUMNS + [c for c in df_issues.columns if c not in ISSUE_COLUMNS])

    # Incidencias por barras aisladas
    isolated_buses = [b for b in bus_index if len(bus_neighbors.get(b, set())) == 0]
    for b in isolated_buses:
        df_issues.loc[len(df_issues)] = {
            "Elemento": b,
            "Clase PF": "ElmTerm",
            "Tipo incidencia": "Barra aislada",
            "Detalle": "No tiene conexiones a otras barras",
        }

    return (
        df_elements,
        df_relations,
        df_issues,
        bus_neighbors,
        bus_attached_elements,
        bus_attachment_counts,
    )


# =============================================================================
# COMPONENTES CONECTADOS
# =============================================================================
def build_components(bus_index, bus_neighbors):
    visited = set()
    comp_rows = []
    comp_map = {}

    buses = sorted(bus_index.keys())
    comp_id = 0

    for start in buses:
        if start in visited:
            continue
        comp_id += 1
        queue = deque([start])
        component = []

        visited.add(start)
        while queue:
            cur = queue.popleft()
            component.append(cur)
            for nxt in bus_neighbors.get(cur, set()):
                if nxt not in visited:
                    visited.add(nxt)
                    queue.append(nxt)

        for b in component:
            comp_map[b] = comp_id

        comp_rows.append(
            {
                "Componente": comp_id,
                "N barras": len(component),
                "Barra inicial": start,
                "Barras": " | ".join(component[:20]) + (" ..." if len(component) > 20 else ""),
                "Aislada": "Si" if len(component) == 1 else "No",
            }
        )

    df_components = pd.DataFrame(comp_rows)
    if df_components.empty:
        df_components = _empty_df(COMPONENT_COLUMNS)
    else:
        for col in COMPONENT_COLUMNS:
            if col not in df_components.columns:
                df_components[col] = None
        df_components = df_components.reindex(columns=COMPONENT_COLUMNS + [c for c in df_components.columns if c not in COMPONENT_COLUMNS])

    return df_components, comp_map


# =============================================================================
# MATRIZ DE VECINDAD
# =============================================================================
def build_adjacency(df_buses, bus_neighbors, bus_attached_elements):
    rows = []

    if df_buses.empty:
        return _empty_df(ADJACENCY_COLUMNS)

    for _, row in df_buses.iterrows():
        name = str(row.get("Nombre", "") or "")
        neighbors = sorted(bus_neighbors.get(name, set()))
        elems = sorted(bus_attached_elements.get(name, set()))

        rows.append(
            {
                "Barra": name,
                "Componente": row.get("Componente"),
                "N vecinos": len(neighbors),
                "Vecinos": " | ".join(neighbors),
                "N elementos asociados": len(elems),
                "Elementos asociados": " | ".join(elems[:50]) + (" ..." if len(elems) > 50 else ""),
                "Zona": row.get("Zona"),
                "Tension nom. (kV)": row.get("Tension nom. (kV)"),
                "Tension (pu)": row.get("Tension (pu)"),
                "En servicio": row.get("En servicio"),
                "Barra aislada": row.get("Barra aislada"),
            }
        )

    df_adj = pd.DataFrame(rows)
    if df_adj.empty:
        df_adj = _empty_df(ADJACENCY_COLUMNS)
    else:
        for col in ADJACENCY_COLUMNS:
            if col not in df_adj.columns:
                df_adj[col] = None
        df_adj = df_adj.reindex(columns=ADJACENCY_COLUMNS + [c for c in df_adj.columns if c not in ADJACENCY_COLUMNS])

    return df_adj


# =============================================================================
# RESUMEN
# =============================================================================
def build_summary(df_buses, df_elements, df_relations, df_issues, components_df, adjacency_df):
    n_buses = len(df_buses)
    n_elements = len(df_elements)
    n_busbus = int((df_relations["TipoRelacion"] == "BUS_BUS").sum()) if not df_relations.empty else 0
    n_buselem = int((df_relations["TipoRelacion"] == "BUS_ELEM").sum()) if not df_relations.empty else 0
    n_issues = len(df_issues)
    n_components = len(components_df)
    n_isolated = int((components_df["Aislada"] == "Si").sum()) if not components_df.empty else 0
    n_isolated_buses = int((components_df["N barras"] == 1).sum()) if not components_df.empty else 0
    n_adj_isolated = int((adjacency_df["Barra aislada"] == "Si").sum()) if not adjacency_df.empty else 0

    rows = [
        {"Campo": "Barras", "Valor": n_buses},
        {"Campo": "Elementos", "Valor": n_elements},
        {"Campo": "Relaciones BUS_BUS", "Valor": n_busbus},
        {"Campo": "Relaciones BUS_ELEM", "Valor": n_buselem},
        {"Campo": "Componentes conectados", "Valor": n_components},
        {"Campo": "Componentes aislados", "Valor": n_isolated},
        {"Campo": "Barras en componentes de 1 nodo", "Valor": n_isolated_buses},
        {"Campo": "Barras aisladas (vecindad)", "Valor": n_adj_isolated},
        {"Campo": "Incidencias detectadas", "Valor": n_issues},
        {
            "Campo": "Cobertura topologica",
            "Valor": (
                "Completa"
                if n_issues == 0 and n_isolated_buses == 0 and n_adj_isolated == 0
                else "Revisar incidencias"
            ),
        },
    ]
    return pd.DataFrame(rows)


# =============================================================================
# HOJAS ESPECIALIZADAS
# =============================================================================
TRANSFORMER_COLUMNS = [
    "Nombre", "En servicio", "Barra 1", "Barra 2",
    "Tension 1 nom. (kV)", "Tension 2 nom. (kV)", "Potencia nom. (MVA)",
    "Carga (%)", "Tap pos.",
    "P lado 1 (MW)", "Q lado 1 (Mvar)", "P lado 2 (MW)", "Q lado 2 (Mvar)",
    "Perdidas P (MW)",
    "uktr (%)", "curmg (%)", "pfe (kW)", "pcutr (kW)",
]

GENERATOR_COLUMNS = [
    "Nombre", "Clase PF", "En servicio", "Barra conectada",
    "pgini (MW)", "P result. (MW)", "Pmax (MW)", "Pmin (MW)",
    "Q nom. (Mvar)", "Q result. (Mvar)",
    "Tension (pu)", "Cos phi", "ip_ctrl", "ngnum", "Tipo",
]


def build_transformers_sheet(df_elements):
    mask = df_elements["Clase PF"] == "ElmTr2"
    df = df_elements[mask].copy()
    for col in TRANSFORMER_COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df.reindex(columns=TRANSFORMER_COLUMNS)


def build_generators_sheet(df_elements):
    gen_classes = {"ElmSym", "ElmGenstat", "ElmPvsys", "ElmWind"}
    mask = df_elements["Clase PF"].isin(gen_classes)
    df = df_elements[mask].copy()
    for col in GENERATOR_COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df.reindex(columns=GENERATOR_COLUMNS)


# =============================================================================
# EXPORTACION JSON PARA GRAFO INTERACTIVO (Streamlit / NetworkX / Plotly)
# =============================================================================
def _er_get(er, key, default=None):
    """Lee un valor de una fila pandas Series o dict evitando NaN."""
    if er is None:
        return default
    try:
        v = er[key] if hasattr(er, "__getitem__") else default
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return default
        return v
    except (KeyError, IndexError):
        return default


def _er_coalesce(er, *keys, default=None):
    """Primer valor no-None / no-NaN entre las claves dadas."""
    for k in keys:
        v = _er_get(er, k, None)
        if v is not None:
            return v
    return default


def export_graph_json(df_buses, df_elements, df_relations, filepath, proyecto=""):
    """
    Genera un JSON listo para visualizar con visualizar_red_pf.py (Streamlit).

    Estructura:
      {
        "metadata": { "proyecto", "timestamp", "n_barras", "n_elementos", "n_aristas" },
        "nodes": [ { "id", "label", "kv", "kv_color", "zone",
                     "v_pu", "v_kv", "angle_deg", "P_MW", "Q_Mvar",
                     "en_servicio", "componente", "aislada",
                     "attached": {"cargas", "generadores", "shunts", "compensadores"} } ],
        "edges": [ { "source", "target", "element", "tipo", "clase_pf",
                     "en_servicio", "loading_pct", "P_from_MW", "Q_from_Mvar",
                     "P_to_MW", "Q_to_Mvar", "perdidas_MW",
                     "kv_nom", "mva_nom", "km",
                     "uktr_pct", "curmg_pct", "pfe_kW", "pcutr_kW" } ]
      }
    """
    from datetime import datetime

    # ── Elementos adjuntos por barra (relaciones BUS_ELEM) ────────────────────
    _CAT_KEY = {
        "Carga":       "cargas",
        "Generador":   "generadores",
        "Shunt":       "shunts",
        "Compensador": "compensadores",
    }
    bus_attached: dict = {}
    if not df_relations.empty:
        for _, row in df_relations[df_relations["TipoRelacion"] == "BUS_ELEM"].iterrows():
            bus  = str(row.get("Barra 1", "") or "")
            elem = str(row.get("Elemento", "") or "")
            cat  = str(row.get("Categoria", "") or "")
            if not bus or not elem:
                continue
            if bus not in bus_attached:
                bus_attached[bus] = {"cargas": [], "generadores": [], "shunts": [], "compensadores": []}
            key = _CAT_KEY.get(cat)
            if key:
                bus_attached[bus][key].append(elem)

    # ── Nodos ─────────────────────────────────────────────────────────────────
    nodes = []
    for _, row in df_buses.iterrows():
        name = str(row.get("Nombre", "") or "")
        kv   = _safe_float(row.get("Tension nom. (kV)"))
        comp = row.get("Componente")
        nodes.append({
            "id":          name,
            "label":       name,
            "kv":          kv,
            "kv_color":    _kv_hex(kv),
            "zone":        str(row.get("Zona", "") or ""),
            "v_pu":        _safe_float(row.get("Tension (pu)")),
            "v_kv":        _safe_float(row.get("Tension (kV)")),
            "angle_deg":   _safe_float(row.get("Angulo (deg)")),
            "P_MW":        _safe_float(row.get("P inyectada (MW)")),
            "Q_Mvar":      _safe_float(row.get("Q inyectada (Mvar)")),
            "en_servicio": str(row.get("En servicio", "") or ""),
            "componente":  int(comp) if comp is not None and not (isinstance(comp, float) and pd.isna(comp)) else None,
            "aislada":     str(row.get("Barra aislada", "") or ""),
            "attached":    bus_attached.get(name, {"cargas": [], "generadores": [], "shunts": [], "compensadores": []}),
        })

    # ── Mapa de elementos para lookup O(1) ────────────────────────────────────
    elem_map: dict = {}
    if not df_elements.empty:
        for _, er in df_elements.iterrows():
            k = str(er.get("Nombre", "") or "")
            if k:
                elem_map[k] = er

    # ── Aristas (solo relaciones BUS_BUS) ─────────────────────────────────────
    edges = []
    bus_bus = (
        df_relations[df_relations["TipoRelacion"] == "BUS_BUS"]
        if not df_relations.empty
        else pd.DataFrame()
    )
    for _, row in bus_bus.iterrows():
        src       = str(row.get("Barra 1", "") or "")
        tgt       = str(row.get("Barra 2", "") or "")
        elem_name = str(row.get("Elemento", "") or "")
        er        = elem_map.get(elem_name)
        edges.append({
            "source":      src,
            "target":      tgt,
            "element":     elem_name,
            "tipo":        str(row.get("Categoria", "") or ""),
            "clase_pf":    str(row.get("Clase PF", "") or ""),
            "en_servicio": str(row.get("En servicio", "") or ""),
            "loading_pct": _safe_float(_er_get(er, "Carga (%)")),
            "P_from_MW":   _safe_float(_er_coalesce(er, "P from (MW)", "P lado 1 (MW)")),
            "Q_from_Mvar": _safe_float(_er_coalesce(er, "Q from (Mvar)", "Q lado 1 (Mvar)")),
            "P_to_MW":     _safe_float(_er_coalesce(er, "P to (MW)", "P lado 2 (MW)")),
            "Q_to_Mvar":   _safe_float(_er_coalesce(er, "Q to (Mvar)", "Q lado 2 (Mvar)")),
            "perdidas_MW": _safe_float(_er_get(er, "Perdidas P (MW)")),
            "kv_nom":      _safe_float(_er_coalesce(er, "Tension nom. (kV)", "Tension 1 nom. (kV)")),
            "mva_nom":     _safe_float(_er_get(er, "Potencia nom. (MVA)")),
            "km":          _safe_float(_er_get(er, "Longitud (km)")),
            "uktr_pct":    _safe_float(_er_get(er, "uktr (%)")),
            "curmg_pct":   _safe_float(_er_get(er, "curmg (%)")),
            "pfe_kW":      _safe_float(_er_get(er, "pfe (kW)")),
            "pcutr_kW":    _safe_float(_er_get(er, "pcutr (kW)")),
        })

    metadata = {
        "proyecto":    proyecto,
        "timestamp":   datetime.now().isoformat(timespec="seconds"),
        "n_barras":    len(nodes),
        "n_elementos": len(df_elements) if df_elements is not None else 0,
        "n_aristas":   len(edges),
    }
    graph = {"metadata": metadata, "nodes": nodes, "edges": edges}
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(graph, f, ensure_ascii=False, indent=2)
    print(f"  Grafo JSON exportado: {filepath}")
    print(f"  Nodos: {len(nodes)}  |  Aristas: {len(edges)}")
    return graph


# =============================================================================
# FORMATO EXCEL
# =============================================================================
def _auto_col(ws, max_w=45):
    for col in ws.columns:
        vals = []
        for c in col:
            try:
                vals.append(len(str(c.value)) if c.value is not None else 0)
            except Exception:
                vals.append(0)
        ancho = max(vals) if vals else 0
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ancho + 2, 10), max_w)


def _format_header(ws):
    mc = ws.max_column
    ws.row_dimensions[1].height = 24
    for c in range(1, mc + 1):
        cell = ws.cell(1, c)
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.alignment = CTR
        cell.border = BRD


def _format_sheet_generic(ws):
    _format_header(ws)
    mc = ws.max_column
    for r in range(2, ws.max_row + 1):
        for c in range(1, mc + 1):
            cell = ws.cell(r, c)
            cell.border = BRD
            cell.font = N_FONT
            cell.alignment = CTR if isinstance(cell.value, (int, float)) else LEFT
    _auto_col(ws)
    ws.freeze_panes = "A2"


def _format_buses(ws, col_service=None):
    _format_header(ws)
    mc = ws.max_column
    for r in range(2, ws.max_row + 1):
        grado = ws.cell(r, ws.max_column - 2).value if ws.max_column >= 3 else None
        servicio = ""
        if col_service is not None and 1 <= col_service <= ws.max_column:
            servicio = str(ws.cell(r, col_service).value or "").strip()

        if isinstance(grado, int) and grado == 0:
            fill = _fill("FFC7CE")
        else:
            fill = _fill("C6EFCE") if servicio == "Si" else _fill("FFC7CE") if servicio == "No" else _fill("FFFFFF")

        for c in range(1, mc + 1):
            cell = ws.cell(r, c)
            cell.border = BRD
            cell.font = N_FONT
            cell.fill = fill
            cell.alignment = CTR if isinstance(cell.value, (int, float)) else LEFT
    _auto_col(ws)
    ws.freeze_panes = "A2"


def _format_elements(ws):
    _format_sheet_generic(ws)


def _format_relations(ws):
    _format_header(ws)
    mc = ws.max_column
    for r in range(2, ws.max_row + 1):
        tipo = str(ws.cell(r, 1).value or "")
        en_serv = str(ws.cell(r, ws.max_column).value or "").strip()
        if tipo == "BUS_BUS":
            fill = _fill("D9EAF7")
        elif tipo == "BUS_ELEM":
            fill = _fill("EDEDED")
        else:
            fill = _fill("FFFFFF") if not en_serv else _fill("C6EFCE")
        for c in range(1, mc + 1):
            cell = ws.cell(r, c)
            cell.border = BRD
            cell.font = N_FONT
            cell.fill = fill
            cell.alignment = CTR if isinstance(cell.value, (int, float)) else LEFT
    _auto_col(ws)
    ws.freeze_panes = "A2"


def _format_issues(ws):
    _format_sheet_generic(ws)
    mc = ws.max_column
    for r in range(2, ws.max_row + 1):
        tipo = str(ws.cell(r, 3).value or "")
        fill = _fill("FFC7CE") if "aislada" in tipo.lower() or "error" in tipo.lower() else _fill("FFEB9C")
        for c in range(1, mc + 1):
            ws.cell(r, c).fill = fill


def _format_components(ws):
    _format_sheet_generic(ws)
    mc = ws.max_column
    for r in range(2, ws.max_row + 1):
        fill = _fill("FFC7CE") if str(ws.cell(r, 5).value or "") == "Si" else _fill("C6EFCE")
        for c in range(1, mc + 1):
            ws.cell(r, c).fill = fill


def _format_adjacency(ws):
    _format_sheet_generic(ws)
    mc = ws.max_column
    for r in range(2, ws.max_row + 1):
        fill = _fill("FFC7CE") if str(ws.cell(r, 11).value or "") == "Si" else _fill("FFFFFF")
        for c in range(1, mc + 1):
            ws.cell(r, c).fill = fill


def _format_summary(ws):
    _format_header(ws)
    mc = ws.max_column
    for r in range(2, ws.max_row + 1):
        for c in range(1, mc + 1):
            cell = ws.cell(r, c)
            cell.border = BRD
            cell.font = N_FONT
            cell.fill = _fill("D9EAF7") if c == 2 else _fill("FFFFFF")
            cell.alignment = LEFT if c == 1 else CTR
    _auto_col(ws)
    ws.freeze_panes = "A2"


# =============================================================================
# MAIN
# =============================================================================
def main():
    print("=" * 78)
    print("TopologiaCompleta_PF - Extraccion y verificacion de topologia completa")
    print("=" * 78)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    app = _setup_pf()

    print("\n[1/5] Leyendo barras...")
    df_buses, bus_index, bus_zone_map = extract_buses(app)
    print(f"  Barras extraidas: {len(df_buses)}")

    print("[2/5] Leyendo elementos y relaciones topologicas...")
    df_elements, df_relations, df_issues, bus_neighbors, bus_attached_elements, bus_attachment_counts = extract_topology_elements(
        app,
        bus_index,
        bus_zone_map,
    )
    print(f"  Elementos extraidos : {len(df_elements)}")
    print(f"  Relaciones topologicas: {len(df_relations)}")
    print(f"  Incidencias detectadas: {len(df_issues)}")

    print("[3/5] Construyendo componentes conectados...")
    df_components, comp_map = build_components(bus_index, bus_neighbors)
    print(f"  Componentes: {len(df_components)}")

    # Anadir componente y estado de aislamiento a barras
    if not df_buses.empty:
        df_buses["Componente"] = df_buses["Nombre"].map(comp_map)
        df_buses["Barra aislada"] = df_buses["Nombre"].apply(lambda x: "Si" if len(bus_neighbors.get(x, set())) == 0 else "No")

    print("[4/5] Construyendo matriz de vecindad...")
    df_adjacency = build_adjacency(df_buses, bus_neighbors, bus_attached_elements)
    print(f"  Barras en vecindad: {len(df_adjacency)}")

    print("[5/6] Generando resumen y exportando Excel...")
    df_summary = build_summary(df_buses, df_elements, df_relations, df_issues, df_components, df_adjacency)

    df_transformers = build_transformers_sheet(df_elements)
    df_generators = build_generators_sheet(df_elements)
    print(f"  Transformadores 2dev: {len(df_transformers)}")
    print(f"  Generadores: {len(df_generators)}")

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Resumen", index=False)
        df_buses.to_excel(writer, sheet_name="Barras", index=False)
        df_elements.to_excel(writer, sheet_name="Elementos", index=False)
        df_transformers.to_excel(writer, sheet_name="Transformadores", index=False)
        df_generators.to_excel(writer, sheet_name="Generadores", index=False)
        df_relations.to_excel(writer, sheet_name="Relaciones", index=False)
        df_components.to_excel(writer, sheet_name="Componentes", index=False)
        df_adjacency.to_excel(writer, sheet_name="Vecindad", index=False)
        df_issues.to_excel(writer, sheet_name="Incidencias", index=False)

    wb = load_workbook(OUTPUT_FILE)

    _format_summary(wb["Resumen"])
    col_service = list(df_buses.columns).index("En servicio") + 1 if "En servicio" in df_buses.columns else None
    _format_buses(wb["Barras"], col_service=col_service)
    _format_elements(wb["Elementos"])
    _format_sheet_generic(wb["Transformadores"])
    _format_sheet_generic(wb["Generadores"])
    _format_relations(wb["Relaciones"])
    _format_components(wb["Componentes"])
    _format_adjacency(wb["Vecindad"])
    _format_issues(wb["Incidencias"])

    # Nota final en Resumen
    ws = wb["Resumen"]
    row_note = ws.max_row + 2
    note_text = (
        "Cobertura topologica: la red se reconstruye a partir de barras (ElmTerm), "
        "enlaces entre barras (lineas, transformadores, interruptores/fusibles) y "
        "equipos conectados a barra (cargas, generadores, shunts). "
        "La hoja 'Transformadores' incluye atributos de perdidas del tipo (uktr, curmg, pfe, pcutr). "
        "La hoja 'Generadores' incluye pgini, Pmax, Pmin e ip_ctrl para diagnostico de condiciones iniciales. "
        "El archivo topologia_grafo.json contiene nodos y aristas para visualizacion interactiva con Streamlit/NetworkX/Plotly."
    )
    ws.cell(row_note, 1, "Nota").font = B_FONT
    ws.cell(row_note, 2, note_text)
    ws.merge_cells(start_row=row_note, start_column=2, end_row=row_note, end_column=max(2, ws.max_column))
    ws.cell(row_note, 2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row_note].height = 75

    wb.save(OUTPUT_FILE)

    print(f"\n[6/6] Exportando grafo JSON para visualizacion interactiva (Streamlit)...")
    proyecto_nombre = PROJECT_NAME or ""
    try:
        active = app.GetActiveProject()
        if active is not None:
            proyecto_nombre = str(ga(active, "loc_name") or PROJECT_NAME or "")
    except Exception:
        pass
    export_graph_json(df_buses, df_elements, df_relations, OUTPUT_GRAPH_JSON, proyecto=proyecto_nombre)

    isolated_count = int((df_components["Aislada"] == "Si").sum()) if not df_components.empty else 0
    isolated_adj_count = int((df_adjacency["Barra aislada"] == "Si").sum()) if not df_adjacency.empty else 0
    print(f"\nArchivo Excel generado: {OUTPUT_FILE}")
    print(f"Hojas: {', '.join(wb.sheetnames)}")
    print(f"Archivo JSON generado: {OUTPUT_GRAPH_JSON}")
    print(f"Barras aisladas: {isolated_count}")
    print(f"Barras aisladas en vecindad: {isolated_adj_count}")
    print(f"Transformadores con datos de perdidas: {len(df_transformers)}")
    print(f"Generadores con diagnostico pgini: {len(df_generators)}")
    print(f"Incidencias: {len(df_issues)}")
    print("Proceso finalizado.")


if __name__ == "__main__":
    main()
