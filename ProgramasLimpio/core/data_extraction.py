# core/data_extraction.py
"""
Refactor de ExtFLujos2daO.py.
Toda la ejecución a nivel de módulo se reemplaza por extraer_datos_evento().
Los input() originales son ahora parámetros: semestre, evento_num.
"""
from __future__ import annotations

import glob
import os
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Estilos Excel ─────────────────────────────────────────────────────────────
_THIN  = Side(border_style="thin", color="BFBFBF")
_BRD   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HFIL  = PatternFill("solid", start_color="1F3864", end_color="1F3864")
_HFNT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_NFNT  = Font(name="Arial", size=10)
_CTR   = Alignment(horizontal="center", vertical="center")
_LEFT  = Alignment(horizontal="left",   vertical="center")
_FILL_MANT = PatternFill("solid", start_color="D0D0D0", end_color="D0D0D0")
_FILL_PO   = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
_FILL_FAL  = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")

CARPETAS_RESULTADOS = [
    "Resultados_COBEE",
    "Resultados_ENDE_Andina",
    "Resultados_GUABIRA",
    "Resultados_SCADA",
    "Resultados_ENDE_Corani",
    "Resultados_ENDE_Guaracachi",
    "Resultados_AGUAI",
    "Resultados_HB",
    "Resultados_ENDE_Valle_Hermoso",
]


# ── Resultado tipado ──────────────────────────────────────────────────────────
@dataclass
class ExtraccionResult:
    archivo_salida: Path
    info_evento: pd.DataFrame
    generadores_pgini: pd.DataFrame
    cargas_plini: pd.DataFrame
    p0_inicial: pd.DataFrame
    warnings: list[str] = field(default_factory=list)
    n_con_po: int = 0
    n_fallback: int = 0
    n_mant: int = 0


# ── Lectores de archivos CNDC ─────────────────────────────────────────────────

def leer_tabla_eventos(semestre: str, raiz: str) -> tuple[list[dict], str]:
    """Lee Tabla_Eventos_*.xlsx. Retorna (lista_eventos, ruta_archivo)."""
    rutas = glob.glob(os.path.join(raiz, semestre, "Tabla_Eventos_*.xlsx"))
    if not rutas:
        raise FileNotFoundError(
            f"No se encontró Tabla_Eventos_*.xlsx en {os.path.join(raiz, semestre)}"
        )
    tabla_path = rutas[0]
    import openpyxl
    wb = openpyxl.load_workbook(tabla_path, data_only=True)
    sh = wb.active
    eventos: list[dict] = []
    for fila in sh.iter_rows(min_row=3, values_only=True):
        if fila[0] is None:
            continue
        try:
            num = int(fila[0])
        except (TypeError, ValueError):
            continue
        eventos.append({
            "num":        num,
            "fecha_hora": str(fila[1]).strip() if fila[1] else "",
            "desconexion": str(fila[2]).strip() if fila[2] else "",
            "pot_desc_MW": fila[3],
            "demanda_MW":  fila[4],
            "f0_Hz":       fila[5],
            "fmin_Hz":     fila[6],
        })
    return eventos, tabla_path


def _buscar_archivo(carpeta: str, patron: str) -> str:
    res = glob.glob(os.path.join(carpeta, patron))
    if not res:
        raise FileNotFoundError(f"No se encontró '{patron}' en: {carpeta}")
    if len(res) > 1:
        pass  # usa el primero
    return res[0]


def leer_dc_hidro_todas_horas(path: str) -> pd.DataFrame:
    import xlrd
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_name("HIDRO")
    cols_hora: dict[int, str] = {}
    for c in range(1, sh.ncols):
        v = sh.cell_value(8, c)
        if isinstance(v, float) and 0 < v <= 1:
            h = int(round(v * 24))
            if 1 <= h <= 24:
                cols_hora[c] = f"{h:02d}:00"
    ignorar = ("SISTEMA", "TOTAL", "DESPACHO", "Los valores", "HORA", "Comite", "RF:", "ND:")
    filas = []
    for r in range(9, sh.nrows):
        nombre = str(sh.cell_value(r, 0)).strip()
        if not nombre or any(nombre.startswith(p) for p in ignorar):
            continue
        fila: dict = {"Generador_CNDC": nombre}
        en_mant = False
        for c, etiq in sorted(cols_hora.items()):
            val = sh.cell_value(r, c)
            if isinstance(val, str) and val.strip().upper() == "M":
                fila[etiq] = 0.0
                en_mant = True
            else:
                fila[etiq] = round(float(val), 4) if isinstance(val, (int, float)) else None
        fila["_mant"] = en_mant
        filas.append(fila)
    return pd.DataFrame(filas)


def leer_dcdr_todas_horas(path: str) -> pd.DataFrame:
    import xlrd
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_name("POST")
    cols_hora: dict[int, str] = {}
    for c in range(1, sh.ncols):
        v = sh.cell_value(6, c)
        if isinstance(v, float) and v > 0:
            h = int(round(v * 24))
            if 1 <= h <= 24:
                cols_hora[c] = f"{h:02d}:00"
    subtotal_row = None
    for r in range(8, sh.nrows):
        if str(sh.cell_value(r, 0)).strip() == "SUBTOTAL HIDRO":
            subtotal_row = r
            break
    ignorar_set = {
        "SUBTOTAL HIDRO", "SUBTOTAL EOLICO", "SUBTOTAL SOLAR",
        "SUBTOTAL TERMO", "SUBTOTAL EXCEDENTES", "TOTAL",
        "RESERVA ROTANTE", "RESERVA PARADA", "SEGURIDAD DE AREAS",
    }
    ignorar_pref = ("RF:", "ND:", "Los valores")
    filas = []
    for r in range(8, sh.nrows):
        if subtotal_row is not None and r <= subtotal_row:
            continue
        nombre = str(sh.cell_value(r, 0)).strip()
        if not nombre or nombre in ignorar_set:
            continue
        if any(nombre.startswith(p) for p in ignorar_pref):
            continue
        fila: dict = {"Generador_CNDC": nombre}
        en_mant = False
        for c, etiq in sorted(cols_hora.items()):
            val = sh.cell_value(r, c)
            if isinstance(val, str) and val.strip().upper() == "M":
                fila[etiq] = 0.0
                en_mant = True
            else:
                fila[etiq] = round(float(val), 4) if isinstance(val, (int, float)) else None
        fila["_mant"] = en_mant
        filas.append(fila)
    return pd.DataFrame(filas)


def leer_deener_todas_horas(path: str) -> pd.DataFrame:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    sh = wb["Demanda"]
    header = list(sh.iter_rows(min_row=7, max_row=7, values_only=True))[0]
    cols_hora: dict[int, str] = {}
    for i, v in enumerate(header):
        if v is None:
            continue
        etiq = str(v).strip()
        if etiq.count(":") == 1:
            try:
                int(etiq[:2])
                int(etiq[3:])
                cols_hora[i] = etiq
            except ValueError:
                pass
    ignorar = ("TOTAL", "Los valores", "RETIROS")
    filas = []
    for row in sh.iter_rows(min_row=8, values_only=True):
        nombre = str(row[0]).strip() if row[0] is not None else ""
        if not nombre or any(nombre.startswith(p) for p in ignorar):
            continue
        fila: dict = {"Nodo_CNDC": nombre}
        for i, etiq in sorted(cols_hora.items()):
            val = row[i]
            try:
                fila[etiq] = round(float(val), 4) if val is not None else None
            except (ValueError, TypeError):
                fila[etiq] = None
        filas.append(fila)
    return pd.DataFrame(filas)


def leer_po_resultados(evento_path: str) -> pd.DataFrame:
    import openpyxl
    filas_po: list[dict] = []
    for carpeta in CARPETAS_RESULTADOS:
        carpeta_path = os.path.join(evento_path, carpeta)
        if not os.path.isdir(carpeta_path):
            continue
        nombre_archivo = f"tabla_resultados_{carpeta.replace('Resultados_', '')}.xlsx"
        archivo_path = os.path.join(carpeta_path, nombre_archivo)
        if not os.path.isfile(archivo_path):
            candidatos = glob.glob(os.path.join(carpeta_path, "tabla_resultados_*.xlsx"))
            if not candidatos:
                continue
            archivo_path = candidatos[0]
        try:
            wb = openpyxl.load_workbook(archivo_path, data_only=True)
            sh = wb.active
            encabezados = list(sh.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            fila_po = None
            for fila in sh.iter_rows(min_row=2, values_only=True):
                etiq = str(fila[0]).strip() if fila[0] else ""
                if "P_0" in etiq or "P0" in etiq.replace(" ", ""):
                    fila_po = fila
                    break
            if fila_po is None:
                continue
            for col_idx, unidad in enumerate(encabezados):
                if col_idx == 0 or unidad is None:
                    continue
                val = fila_po[col_idx]
                try:
                    po = round(float(val), 4) if val is not None else None
                except (ValueError, TypeError):
                    po = None
                filas_po.append({"Unidad": str(unidad).strip(), "P_0_MW": po, "Fuente": carpeta})
        except Exception:
            continue
    return (
        pd.DataFrame(filas_po) if filas_po
        else pd.DataFrame(columns=["Unidad", "P_0_MW", "Fuente"])
    )


def leer_codigos_sti_gen(path: str) -> dict[str, str]:
    if not os.path.isfile(path):
        return {}
    df = pd.read_excel(path, sheet_name="Mapeo_Generadores")
    resultado: dict[str, str] = {}
    for _, row in df.iterrows():
        cndc = str(row["Generador_CNDC"]).strip()
        locs = str(row["loc_names PF"]).strip()
        if not locs or locs == "-":
            resultado[cndc] = "-"
            continue
        first = locs.split(",")[0].strip()
        s = re.sub(r"^(sym_|WT_|PV-|PV_|sta_)", "", first, flags=re.IGNORECASE)
        s = re.sub(r"_EQ$", "", s, flags=re.IGNORECASE)
        s = re.sub(r"\(\d+\)$", "", s).strip()
        resultado[cndc] = s if s else "-"
    return resultado


def _limpiar_display(nombre: str) -> str:
    return re.sub(r"^CC([A-Z])", r"\1", str(nombre).strip())


def _limpiar_lookup(nombre: str) -> str:
    s = re.sub(r"^CC([A-Z])", r"\1", str(nombre).strip())
    return re.sub(r"\s*-\s*(RF|PPG)\.?\s*$", "", s, flags=re.IGNORECASE).strip()


def _insertar_columna_evento(
    df: pd.DataFrame, col_id: str, hora_evento: str, dict_po: dict
) -> pd.DataFrame:
    hora_h = int(str(hora_evento).strip().split(":", 1)[0])
    hora_ant = f"{hora_h:02d}:00"
    hora_sig = f"{(hora_h % 24) + 1:02d}:00"
    cols = list(df.columns)
    if hora_ant in cols:
        pos = cols.index(hora_ant) + 1
    elif hora_sig in cols:
        pos = cols.index(hora_sig)
    else:
        pos = len(cols)
    valores = []
    for _, row in df.iterrows():
        nombre = str(row[col_id]).strip()
        po = dict_po.get(nombre)
        if po is None:
            po = row.get(hora_ant)
        valores.append(po)
    df.insert(pos, hora_evento, valores)
    return df


def _aplicar_formato_xlsx(ws) -> None:
    mc = ws.max_column
    ws.row_dimensions[1].height = 22
    for c in range(1, mc + 1):
        cell = ws.cell(1, c)
        cell.fill = _HFIL
        cell.font = _HFNT
        cell.alignment = _CTR
        cell.border = _BRD
    for r in range(2, ws.max_row + 1):
        for c in range(1, mc + 1):
            cell = ws.cell(r, c)
            cell.font = _NFNT
            cell.border = _BRD
            cell.alignment = _CTR if isinstance(cell.value, (int, float)) else _LEFT
    for col in ws.columns:
        ancho = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ancho + 2, 10), 35)
    ws.freeze_panes = "A2"


# ── Función principal ─────────────────────────────────────────────────────────

def extraer_datos_evento(
    semestre: str,
    evento_num: int,
    raiz_cndc: str,
    loc_names_gen_path: str,
    progress_cb: Optional[Callable[[float, str], None]] = None,
) -> ExtraccionResult:
    """
    Reemplaza la ejecución interactiva de ExtFLujos2daO.py.
    Los input() originales son ahora parámetros semestre y evento_num.
    """
    warns: list[str] = []

    def _p(pct: float, msg: str) -> None:
        if progress_cb:
            progress_cb(pct, msg)

    _p(0.05, "Leyendo tabla de eventos…")
    eventos, _ = leer_tabla_eventos(semestre, raiz_cndc)
    ev_sel = next((e for e in eventos if e["num"] == evento_num), None)
    if ev_sel is None:
        raise ValueError(f"Evento {evento_num} no encontrado en semestre '{semestre}'")

    partes = ev_sel["fecha_hora"].split(" ")
    fecha_ddmmyyyy = partes[0]
    hora_evento = partes[1] if len(partes) > 1 else "00:00"
    hora_po = f"{int(hora_evento.split(':')[0]):02d}:00"
    d, m, y = fecha_ddmmyyyy.split("/")
    fecha_str = f"{d}{m}{y[2:]}"

    evento_path = os.path.join(
        raiz_cndc, semestre, "Análisis_todos_los_eventos", f"Evento {evento_num}"
    )
    if not os.path.isdir(evento_path):
        raise FileNotFoundError(f"Carpeta no encontrada: {evento_path}")

    despacho_dir = os.path.join(evento_path, "Despacho")
    demanda_dir  = os.path.join(evento_path, "Demanda de Energia y Potencia")
    dc_path    = _buscar_archivo(despacho_dir, "dc_*.xls*")
    dcdr_path  = _buscar_archivo(despacho_dir, "dcdr_*.xls*")
    deener_path = _buscar_archivo(demanda_dir, "deener_*.xlsx")

    _p(0.20, "Leyendo DC Hidro…")
    df_hidro = leer_dc_hidro_todas_horas(dc_path)

    _p(0.38, "Leyendo DCDR (no-hidro)…")
    df_gen_no_hidro = leer_dcdr_todas_horas(dcdr_path)

    df_gen = pd.concat([df_hidro, df_gen_no_hidro], ignore_index=True)

    _p(0.52, "Leyendo DEENER (demanda)…")
    df_carga = leer_deener_todas_horas(deener_path)

    _p(0.63, "Leyendo P_0 de tabla_resultados…")
    df_po = leer_po_resultados(evento_path)

    _p(0.72, "Cargando códigos STI desde loc_names_gen…")
    codigos_sti_map = leer_codigos_sti_gen(loc_names_gen_path)

    dict_po_gen: dict[str, float] = {}
    if not df_po.empty:
        dict_po_raw = {
            str(k).strip(): v
            for k, v in zip(df_po["Unidad"], df_po["P_0_MW"])
            if v is not None
        }
        dict_po_norm = dict(dict_po_raw)
        for k, v in dict_po_raw.items():
            if not re.search(r"\d+$", k):
                dict_po_norm.setdefault(k + "01", v)
        for nombre_cndc in df_gen["Generador_CNDC"]:
            orig    = str(nombre_cndc).strip()
            cleaned = _limpiar_lookup(orig)
            sti     = codigos_sti_map.get(cleaned)
            po = dict_po_norm.get(orig) or dict_po_norm.get(cleaned)
            if po is None and sti:
                po = dict_po_norm.get(sti)
            if po is not None:
                dict_po_gen[orig] = po

    _p(0.80, "Insertando columna hora del evento…")
    df_gen   = _insertar_columna_evento(df_gen,   "Generador_CNDC", hora_evento, dict_po_gen)
    df_carga = _insertar_columna_evento(df_carga, "Nodo_CNDC",      hora_evento, {})

    df_gen["Generador_CNDC"] = df_gen["Generador_CNDC"].apply(_limpiar_display)
    df_gen.insert(1, "Codigo STI",
                  df_gen["Generador_CNDC"].apply(
                      lambda n: codigos_sti_map.get(_limpiar_lookup(n), "-")))
    df_gen.insert(2, "Estado",
                  df_gen["_mant"].apply(lambda x: "Mantenimiento" if x else ""))
    df_gen.drop(columns=["_mant"], inplace=True, errors="ignore")

    n_mant     = int((df_gen["Estado"] == "Mantenimiento").sum())
    n_con_po   = sum(1 for n in df_gen["Generador_CNDC"] if str(n).strip() in dict_po_gen)
    n_fallback = len(df_gen) - n_mant - n_con_po

    df_evento = pd.DataFrame([
        {"Campo": "Semestre",            "Valor": semestre},
        {"Campo": "Evento N°",           "Valor": ev_sel["num"]},
        {"Campo": "Fecha y hora",        "Valor": ev_sel["fecha_hora"]},
        {"Campo": "Hora Po (inicio)",    "Valor": hora_po},
        {"Campo": "Disparo",             "Valor": ev_sel["desconexion"]},
        {"Campo": "Potencia desc. [MW]", "Valor": ev_sel["pot_desc_MW"]},
        {"Campo": "Demanda SIN [MW]",    "Valor": ev_sel["demanda_MW"]},
        {"Campo": "f0 [Hz]",             "Valor": ev_sel["f0_Hz"]},
        {"Campo": "fmin [Hz]",           "Valor": ev_sel["fmin_Hz"]},
    ])

    nombres_con_po  = {_limpiar_display(k) for k in dict_po_gen}
    nombres_en_mant = set(df_gen[df_gen["Estado"] == "Mantenimiento"]["Generador_CNDC"])

    df_po_out = pd.DataFrame(columns=["Unidad", "P_0_MW", "Fuente"])
    if not df_po.empty:
        df_po_out = (
            df_po.copy()
            .assign(Unidad=lambda d: d["Unidad"].apply(_limpiar_display))
            .dropna(subset=["P_0_MW"])
            .drop_duplicates(subset="Unidad", keep="first")
            .reset_index(drop=True)
        )

    _p(0.90, "Guardando Excel de salida…")
    output_path = Path(evento_path) / f"datos_simulacion_{fecha_str}_2daopcion.xlsx"
    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df_evento.to_excel(writer, sheet_name="Info_Evento",       index=False)
        df_gen.to_excel(   writer, sheet_name="Generadores_pgini", index=False)
        df_carga.to_excel( writer, sheet_name="Cargas_plini",      index=False)
        df_po_out.to_excel(writer, sheet_name="P0_inicial",        index=False)

    wb = load_workbook(str(output_path))
    for hoja in wb.sheetnames:
        _aplicar_formato_xlsx(wb[hoja])
    ws_gen = wb["Generadores_pgini"]
    for r in range(2, ws_gen.max_row + 1):
        nombre = str(ws_gen.cell(r, 1).value).strip()
        fill = (
            _FILL_MANT if nombre in nombres_en_mant
            else _FILL_PO if nombre in nombres_con_po
            else _FILL_FAL
        )
        for c in range(1, ws_gen.max_column + 1):
            ws_gen.cell(r, c).fill = fill
    wb.save(str(output_path))

    _p(1.0, "Completado.")
    return ExtraccionResult(
        archivo_salida=output_path,
        info_evento=df_evento,
        generadores_pgini=df_gen,
        cargas_plini=df_carga,
        p0_inicial=df_po_out,
        warnings=warns,
        n_con_po=n_con_po,
        n_fallback=max(n_fallback, 0),
        n_mant=n_mant,
    )


def cargar_datos_simulacion(path: str) -> Optional[ExtraccionResult]:
    """Lee un datos_simulacion_*.xlsx ya generado (sin re-procesar archivos CNDC)."""
    try:
        return ExtraccionResult(
            archivo_salida=Path(path),
            info_evento=pd.read_excel(path, sheet_name="Info_Evento"),
            generadores_pgini=pd.read_excel(path, sheet_name="Generadores_pgini"),
            cargas_plini=pd.read_excel(path, sheet_name="Cargas_plini"),
            p0_inicial=pd.read_excel(path, sheet_name="P0_inicial"),
        )
    except Exception:
        return None


def cargar_condiciones_iniciales(path: str) -> Optional[dict[str, pd.DataFrame]]:
    """Lee un condiciones_iniciales_*.xlsx ya generado."""
    try:
        xl = pd.ExcelFile(path)
        return {sheet: xl.parse(sheet) for sheet in xl.sheet_names}
    except Exception:
        return None
