#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
ExtractorDSL_Reivax.py
----------------------
Extrae parametros DSL del gobernador Reivax desde el composite model
SYM_FRAME2. Bloques objetivo: ATUADOR, POSICAO, Pot, VHZ, CND, lmc.

Solo extrae parametros configurables por usuario (seccion Parameters del
BlkDef DSL), NO variables de estado ni senales calculadas internamente.

Ejecucion: script interno de PowerFactory (Run as Script)
"""

import os
from datetime import datetime
import powerfactory as pf

# ══════════════════════════════════════════════════════════════
# Configuracion
# ══════════════════════════════════════════════════════════════

TARGET_LOC_NAMES = {
    'sym_BOT01', 'sym_BOT02', 'sym_BOT03',
    'sym_CUT01', 'sym_CUT02', 'sym_CUT03', 'sym_CUT04',
    'sym_ANG03', 'sym_CRB01',
}

# Slots DSL a extraer — nombres confirmados del composite model real.
# Se usa startswith(), asi que un prefijo cubre variantes del mismo bloque.
DSL_SLOTS = (
    'ATUADOR_SIMP',    # actuador / servomotor (gobernador Reivax)
    'POS_SIMP',        # posicion de apertura
    'CONDUTO_TURBINA', # conduto y turbina
    'VHZL',            # proteccion V/Hz
    'UEL',             # limitador de subexcitacion
    'OEL',             # limitador de sobreexcitacion
    'drpIEEEVC',       # droop regulacion de tension
    'PSS_COMP',        # estabilizador de sistema de potencia
    'RV',              # regulador de tension (parte del AVR)
    'AVR',             # bloque principal AVR
    'REG_TEN',         # regulador de tension
    'MED_TEN',         # medida de tension
    'MEL',             # minimum excitation limiter
    'SCL',             # stator current limiter
    'DRIVE',           # drive del excitador
    'EXCITATRIZ',      # excitatriz
)

# True: imprime estructura de composite model y BlkDef para depurar
# cuando no se encuentran parametros o los nombres de slot difieren.
MODO_DIAGNOSTICO = True

# ══════════════════════════════════════════════════════════════

# ─────────────────────────────────────────────────────────────
# Helpers generales
# ─────────────────────────────────────────────────────────────
def ga(obj, attr):
    try:
        return obj.GetAttribute(attr) if obj is not None else None
    except Exception:
        return None


def _cls(obj):
    try:
        return obj.GetClassName()
    except Exception:
        return ''


def _nom(obj, default='?'):
    if obj is None:
        return default
    return ga(obj, 'loc_name') or default


def _fullname(obj):
    """Ruta completa del objeto en la jerarquia PF. Usada para comparar identidad."""
    if obj is None:
        return None
    try:
        return obj.GetFullName()
    except Exception:
        return None



# ─────────────────────────────────────────────────────────────
# Atributos estandar de ElmDsl que NO son parametros de usuario
# ─────────────────────────────────────────────────────────────
# Cualquier atributo del ElmDsl que NO este en esta lista y sea
# numerico/string se considera parametro DSL de usuario.
_ATTRS_STD_ELMDSL = {
    # Comunes a todos los objetos PF
    'loc_name', 'fold_id', 'desc', 'for_name', 'chr_name', 'charact',
    'sernum', 'iSchemeStatus', 'GPSlat', 'GPSlon',
    # Referencia al tipo (BlkDef)
    'typ_id',
    # Estado y conexion
    'outserv', 'bus1', 'bus2', 'bus3', 'bus4', 'ibus', 'bbusbar',
    # Control
    'cpSite', 'cpCtrl', 'c_pmod', 'ip_ctrl', 'iOPslack', 'i_mot',
    'pFlicker', 'pStoch', 'pQPcurve',
    # Escalado (generadores)
    'scale0', 'scale0f', 'Pmax', 'Pmin',
    # Excluir variables internas/estado/input/output como solicitado
    'sInput', 'sOutput', 'sStates', 'sIntern',
}


# ─────────────────────────────────────────────────────────────
# Lectura de parametros desde ElmDsl
# ─────────────────────────────────────────────────────────────
def extraer_parametros_dsl(dsl_elm, app_logger=None):
    """
    Estrategia simplificada:
    Inspecciona las listas de definición sParams, sUpLimPar y sLowLimPar del BlkDef
    y extrae sus valores actuales desde el elemento ElmDsl.

    Retorna (dict{nombre: valor}, blk_def).
    """
    blk_def = ga(dsl_elm, 'typ_id')
    params = {}

    if blk_def is not None:
        listas_nombres = ('sParams', 'sUpLimPar', 'sLowLimPar')
        for attr_lista in listas_nombres:
            nombres_raw = ga(blk_def, attr_lista)
            if not nombres_raw:
                continue
            
            # Normalizar nombres_raw: puede ser una cadena o una lista de cadenas
            bloques_de_nombres = nombres_raw if isinstance(nombres_raw, (list, tuple)) else [nombres_raw]

            for bloque in bloques_de_nombres:
                if isinstance(bloque, str):
                    nombres_separados = [v.strip() for v in bloque.split(',') if v.strip()]
                    new_params = _leer_params_de_nombres(dsl_elm, nombres_separados, excluir=_ATTRS_STD_ELMDSL)
                    params.update(new_params)
                    
                    if app_logger and MODO_DIAGNOSTICO and new_params:
                        app_logger.PrintInfo(f"      [DEBUG] Atributo {attr_lista} encontró {len(new_params)} parámetros.")

    return params, blk_def


def _leer_params_de_nombres(dsl_elm, nombres, excluir):
    """
    Dado un ElmDsl y una lista de nombres de atributo, lee solo los que:
    - No estan en el set de exclusion
    - Son numericos (float/int) — parametros DSL tipicamente son numericos
    Retorna dict{nombre: valor}.
    """
    params = {}
    for name in nombres:
        if name in excluir:
            continue
        # Saltar prefijos de sistema PF ('c_', 'e_', 'i_' suelen ser internos)
        # EXCEPCION: muchos params Reivax empiezan con letras normales (Kp, Ti, etc.)
        val = ga(dsl_elm, name)
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            params[name] = val
        elif isinstance(val, str) and val and name not in excluir:
            params[name] = val
    return params


# ─────────────────────────────────────────────────────────────
# Busqueda de composite model
# ─────────────────────────────────────────────────────────────
def _pelm_lista(cm):
    """
    Devuelve todos los objetos referenciados en los slots pelm del ElmComp.
    Intenta obtenerlos como lista primero; si no, itera pelm[i].
    """
    # Intento 1: pelm como lista completa
    lista = ga(cm, 'pelm')
    if isinstance(lista, (list, tuple)) and len(lista) > 0:
        return [x for x in lista if x is not None]

    # Intento 2: acceso indexado pelm[i]
    resultado = []
    for i in range(60):
        ref = ga(cm, f'pelm[{i}]')
        if ref is None:
            break
        resultado.append(ref)
    return resultado


def _buscar_composite_model_por_c_pmod(sym_elm):
    """Atajo: muchos ElmSym tienen el atributo c_pmod que apunta directamente al ElmComp."""
    cm = ga(sym_elm, 'c_pmod')
    if cm is not None and _cls(cm) == 'ElmComp':
        return cm
    return None


def _buscar_composite_model_en_lista(sym_elm, todos_comp):
    """
    Busca en la lista de ElmComp el que referencie al ElmSym via sus slots pelm.
    Usa GetFullName() para la comparacion (la igualdad == no funciona en COM PF).
    """
    fn_sym = _fullname(sym_elm)
    if fn_sym is None:
        return None

    for cm in todos_comp:
        for ref in _pelm_lista(cm):
            if _fullname(ref) == fn_sym:
                return cm
    return None


def _buscar_composite_model_en_carpeta(sym_elm):
    """
    Fallback: busca ElmComp en la misma carpeta que el ElmSym
    (estructura tipica de unidades de generacion en PF).
    """
    try:
        parent = sym_elm.GetParent()
        if parent is None:
            return None
        comps = parent.GetContents('*.ElmComp', 0) or []
        if comps:
            return comps[0]
    except Exception:
        pass
    return None


def _buscar_composite_model(sym_elm, todos_comp):
    """
    Estrategia multicapa para encontrar el ElmComp de un ElmSym:
    1. Atributo directo c_pmod del ElmSym
    2. Busqueda por GetFullName() en la lista global de ElmComp
    3. Busqueda en la carpeta padre del ElmSym
    """
    cm = _buscar_composite_model_por_c_pmod(sym_elm)
    if cm is not None:
        return cm, 'c_pmod'

    cm = _buscar_composite_model_en_lista(sym_elm, todos_comp)
    if cm is not None:
        return cm, 'pelm_scan'

    cm = _buscar_composite_model_en_carpeta(sym_elm)
    if cm is not None:
        return cm, 'carpeta_padre'

    return None, None


def _dsl_slots_de_composite(cm):
    """Devuelve todos los ElmDsl referenciados como slots en el ElmComp."""
    return [ref for ref in _pelm_lista(cm) if _cls(ref) == 'ElmDsl']


# ─────────────────────────────────────────────────────────────
# Busqueda de todos los ElmComp del proyecto
# ─────────────────────────────────────────────────────────────
def _buscar_comps_en_proyecto(app, project):
    encontrados = {}

    try:
        for cm in (app.GetCalcRelevantObjects('*.ElmComp') or []):
            encontrados[_fullname(cm)] = cm
    except Exception:
        pass

    try:
        for cm in (project.GetContents('*.ElmComp', 1) or []):
            k = _fullname(cm)
            if k not in encontrados:
                encontrados[k] = cm
    except Exception:
        pass

    return list(encontrados.values())


# ─────────────────────────────────────────────────────────────
# Modo diagnostico
# ─────────────────────────────────────────────────────────────
def _diagnosticar_composite(app, cm, dsl_elms, metodo):
    app.PrintInfo(f"  Composite  : {_nom(cm)}  (via {metodo})")
    app.PrintInfo(f"  DSL slots objetivo: {len(dsl_elms)}")
    
    for dsl in dsl_elms:
        blk        = ga(dsl, 'typ_id')
        blk_nombre = _nom(blk, 'None')

        # Listas de variables de usuario en BlkDef
        listas_info = {}
        for attr in ('sParams', 'sUpLimPar', 'sLowLimPar'):
            val = ga(blk, attr)
            if val:
                listas_info[attr] = val[:70] + "..." if len(val) > 70 else val

        # Params leidos con la nueva estrategia simplificada
        params, _ = extraer_parametros_dsl(dsl, app)

        app.PrintInfo(f"    DSL: {_nom(dsl)}  |  BlkDef: {blk_nombre}")
        if listas_info:
            app.PrintInfo(f"      Definiciones en BlkDef: {listas_info}")
        app.PrintInfo(f"      Parámetros extraídos ({len(params)}): {list(params.keys())[:15]}")
        if len(params) > 15:
            app.PrintInfo(f"      ... y {len(params)-15} más.")


# ─────────────────────────────────────────────────────────────
# Seleccion de carpeta (tkinter)
# ─────────────────────────────────────────────────────────────
def _elegir_carpeta():
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    carpeta = filedialog.askdirectory(
        title='Carpeta de salida — Parametros DSL Reivax',
        mustexist=True,
    )
    root.destroy()
    if not carpeta:
        raise RuntimeError('Seleccion de carpeta cancelada.')
    return carpeta.replace('/', os.sep)


# ─────────────────────────────────────────────────────────────
# Helpers de color
# ─────────────────────────────────────────────────────────────
def _hex_claro(hex_color, factor):
    """Mezcla hex_color con blanco. factor=0 → color puro; factor=1 → blanco."""
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    return '{:02X}{:02X}{:02X}'.format(
        int(r + (255 - r) * factor),
        int(g + (255 - g) * factor),
        int(b + (255 - b) * factor),
    )


# Paleta de colores para secciones DSL (uno por bloque, cicla si hay mas)
_COLORES_BLOQUES = [
    '2E75B6', '70AD47', 'ED7D31', '7030A0',
    '00B0F0', 'FF0000', '44546A', 'FFC000',
    '00B050', 'C00000',
]


# ─────────────────────────────────────────────────────────────
# Exportacion a Excel (openpyxl)
# ─────────────────────────────────────────────────────────────
def _escribir_excel(filas_detalle, filas_ancho, carpeta):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None, 'openpyxl no disponible'

    ts     = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre = f"Parametros_DSL_Reivax_{ts}.xlsx"
    ruta   = os.path.join(carpeta, nombre)

    wb = openpyxl.Workbook()

    # ── Estilos base ──────────────────────────────────────────────────────
    fill_azul  = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    fill_verde = PatternFill(start_color='375623', end_color='375623', fill_type='solid')

    # Banding para hojas largas: alterna cada vez que cambia el bloque DSL
    fills_banda = [
        PatternFill(start_color='DEEAF1', end_color='DEEAF1', fill_type='solid'),
        PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    ]

    fnt_hdr   = Font(color='FFFFFF', bold=True)
    fnt_sec   = Font(color='FFFFFF', bold=True, size=11)
    fnt_datos = Font(size=10)

    alin_cen = Alignment(horizontal='center', vertical='center')
    alin_izq = Alignment(horizontal='left',   vertical='center', indent=1)
    alin_der = Alignment(horizontal='right',  vertical='center')

    _thin = Side(style='thin', color='BFBFBF')
    borde = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _cabecera(ws, fill, n_cols=None):
        row = ws[1]
        for cell in (row[:n_cols] if n_cols else row):
            cell.fill      = fill
            cell.font      = fnt_hdr
            cell.alignment = alin_cen
            cell.border    = borde

    def _autoajuste(ws):
        for col in ws.columns:
            ancho = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 50)

    def _fila_datos(ws, valores, fill, alin_ultima=None):
        ws.append(valores)
        for i, cell in enumerate(ws[ws.max_row], start=1):
            cell.fill      = fill
            cell.font      = fnt_datos
            cell.border    = borde
            cell.alignment = alin_der if (alin_ultima and i == len(valores)) else alin_izq

    # ── Hoja 1: Parametros_Largo ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Parametros_Largo'
    COLS_L = ['Unidad (loc_name)', 'Composite Model', 'DSL Bloque',
              'DSL Definicion', 'Parametro', 'Valor']
    ws1.append(COLS_L)
    _cabecera(ws1, fill_azul)
    ws1.freeze_panes = 'A2'

    prev_bloque = None
    band_idx    = 0
    for fila in filas_detalle:
        key = (fila.get('Unidad (loc_name)', ''), fila.get('DSL Bloque', ''))
        if key != prev_bloque:
            band_idx    = (band_idx + 1) % 2
            prev_bloque = key
        _fila_datos(ws1, [fila.get(c, '') for c in COLS_L], fills_banda[band_idx], alin_ultima=True)

    _autoajuste(ws1)

    # ── Hoja 2: Tabla_Pivot ───────────────────────────────────────────────
    ws2 = wb.create_sheet(title='Tabla_Pivot')
    if filas_ancho:
        FIJAS  = ['Unidad (loc_name)', 'Composite Model', 'DSL Bloque', 'DSL Definicion']
        params = []
        visto  = set()
        for fila in filas_ancho:
            for k in fila:
                if k not in FIJAS and k not in visto:
                    params.append(k)
                    visto.add(k)
        todas = FIJAS + params
        ws2.append(todas)
        _cabecera(ws2, fill_verde)
        ws2.freeze_panes = f'{get_column_letter(len(FIJAS) + 1)}2'

        prev_bloque = None
        band_idx    = 0
        for fila in filas_ancho:
            key = (fila.get('Unidad (loc_name)', ''), fila.get('DSL Bloque', ''))
            if key != prev_bloque:
                band_idx    = (band_idx + 1) % 2
                prev_bloque = key
            _fila_datos(ws2, [fila.get(c, '') for c in todas], fills_banda[band_idx])

        _autoajuste(ws2)

    # ── Hojas por Unidad ──────────────────────────────────────────────────
    # Diseño: fila encabezado de columnas + por cada bloque DSL una fila
    # de sección coloreada (merge) seguida de sus parámetros con zebra.
    from collections import defaultdict, OrderedDict

    unidades_filas = defaultdict(list)
    for fila in filas_detalle:
        unidades_filas[fila['Unidad (loc_name)']].append(fila)

    COLS_U   = ['Parametro', 'Valor']
    N_COLS_U = len(COLS_U)
    col_fin  = get_column_letter(N_COLS_U)

    for loc_name in sorted(unidades_filas.keys()):
        ws_u   = wb.create_sheet(title=str(loc_name)[:31])
        datos_u = unidades_filas[loc_name]

        # Cabecera de columnas
        ws_u.append(COLS_U)
        _cabecera(ws_u, fill_azul, N_COLS_U)
        ws_u.freeze_panes = 'A2'

        # Ordenar bloques respetando el orden de aparición
        bloques_ord  = list(OrderedDict.fromkeys(
            (f['DSL Bloque'], f['DSL Definicion']) for f in datos_u
        ))
        bloques_data = defaultdict(list)
        for f in datos_u:
            bloques_data[(f['DSL Bloque'], f['DSL Definicion'])].append(f)

        for idx_b, (dsl_bloque, dsl_def) in enumerate(bloques_ord):
            col_hex   = _COLORES_BLOQUES[idx_b % len(_COLORES_BLOQUES)]
            fill_sec  = PatternFill(start_color=col_hex, end_color=col_hex,  fill_type='solid')
            fill_par  = PatternFill(start_color=_hex_claro(col_hex, 0.82),
                                    end_color=_hex_claro(col_hex, 0.82),   fill_type='solid')
            fill_impar = PatternFill(start_color=_hex_claro(col_hex, 0.92),
                                     end_color=_hex_claro(col_hex, 0.92),  fill_type='solid')

            # Fila sección: celda A mergeada con label del bloque
            ws_u.append([''] * N_COLS_U)
            row_sec = ws_u.max_row
            ws_u.merge_cells(f'A{row_sec}:{col_fin}{row_sec}')
            cell_sec             = ws_u.cell(row=row_sec, column=1)
            cell_sec.value       = f'  {dsl_bloque}   —   {dsl_def}'
            cell_sec.fill        = fill_sec
            cell_sec.font        = fnt_sec
            cell_sec.alignment   = alin_izq
            ws_u.row_dimensions[row_sec].height = 18

            # Parámetros del bloque con zebra
            for i_p, f in enumerate(bloques_data[(dsl_bloque, dsl_def)]):
                fill_p = fill_par if i_p % 2 == 0 else fill_impar
                ws_u.append([f.get('Parametro', ''), f.get('Valor', '')])
                row_p = ws_u.max_row
                ws_u.cell(row_p, 1).fill      = fill_p
                ws_u.cell(row_p, 1).font      = fnt_datos
                ws_u.cell(row_p, 1).border    = borde
                ws_u.cell(row_p, 1).alignment = alin_izq
                ws_u.cell(row_p, 2).fill      = fill_p
                ws_u.cell(row_p, 2).font      = fnt_datos
                ws_u.cell(row_p, 2).border    = borde
                ws_u.cell(row_p, 2).alignment = alin_der

        _autoajuste(ws_u)

    wb.save(ruta)
    return ruta, None


# ─────────────────────────────────────────────────────────────
# Programa principal
# ─────────────────────────────────────────────────────────────
def main():
    app = pf.GetApplication()
    if app is None:
        return

    try:
        app.ClearOutputWindow()
    except Exception:
        pass

    SEP = '=' * 68
    app.PrintInfo(SEP)
    app.PrintInfo('  ExtractorDSL_Reivax  —  Parametros gobernador Reivax')
    app.PrintInfo(SEP)

    project = app.GetActiveProject()
    if project is None:
        app.PrintError('No hay proyecto activo.')
        return

    app.PrintInfo(f"Proyecto        : {_nom(project)}")
    app.PrintInfo(f"Modo diagnostico: {'Si' if MODO_DIAGNOSTICO else 'No'}")
    app.PrintInfo('')

    try:
        carpeta_salida = _elegir_carpeta()
    except RuntimeError as e:
        app.PrintError(str(e))
        return
    app.PrintInfo(f"Carpeta salida  : {carpeta_salida}")
    app.PrintInfo('')

    # ── Buscar ElmSym objetivo ───────────────────────────────────────────
    all_syms = app.GetCalcRelevantObjects('*.ElmSym') or []
    target_syms = {}
    for sym in all_syms:
        nombre = ga(sym, 'loc_name')
        if nombre in TARGET_LOC_NAMES:
            target_syms[nombre] = sym

    faltantes = TARGET_LOC_NAMES - set(target_syms.keys())
    app.PrintInfo(f"Unidades encontradas ({len(target_syms)}): {sorted(target_syms.keys())}")
    if faltantes:
        app.PrintWarn(f"No encontradas: {sorted(faltantes)}")
    app.PrintInfo('')

    # ── Buscar composite models ──────────────────────────────────────────
    all_comps = _buscar_comps_en_proyecto(app, project)
    app.PrintInfo(f"Composite models en proyecto: {len(all_comps)}")
    app.PrintInfo('')

    filas_detalle = []
    filas_ancho   = []

    for loc_name in sorted(target_syms.keys()):
        sym = target_syms[loc_name]
        app.PrintInfo(f"-> {loc_name}  [{_fullname(sym)}]")

        cm, metodo = _buscar_composite_model(sym, all_comps)
        if cm is None:
            app.PrintWarn(f"   Sin composite model. "
                          f"Ruta sym: {_fullname(sym)}")
            app.PrintInfo('')
            continue

        cm_name  = _nom(cm)
        dsl_elms = _dsl_slots_de_composite(cm)
        app.PrintInfo(f"   Composite: {cm_name}  (via {metodo})")
        app.PrintInfo(f"   DSL slots: {[_nom(d) for d in dsl_elms]}")

        if MODO_DIAGNOSTICO:
            _diagnosticar_composite(app, cm, dsl_elms, metodo)

        n_bloques = 0
        for dsl in dsl_elms:
            dsl_name = _nom(dsl, '')

            if not any(dsl_name.startswith(p) for p in DSL_SLOTS):
                continue

            params, blk_def = extraer_parametros_dsl(dsl, app)
            blk_name        = _nom(blk_def, 'N/A') if blk_def else 'N/A'

            if not params:
                app.PrintWarn(f"   Bloque {dsl_name} ({blk_name}): 0 params. "
                              f"Activar MODO_DIAGNOSTICO para inspeccionar BlkDef.")
            else:
                app.PrintInfo(f"   Bloque: {dsl_name} ({blk_name}) -> {len(params)} param.")

            for param_nombre, param_val in params.items():
                filas_detalle.append({
                    'Unidad (loc_name)': loc_name,
                    'Composite Model':   cm_name,
                    'DSL Bloque':        dsl_name,
                    'DSL Definicion':    blk_name,
                    'Parametro':         param_nombre,
                    'Valor':             param_val,
                })

            fila_ancho = {
                'Unidad (loc_name)': loc_name,
                'Composite Model':   cm_name,
                'DSL Bloque':        dsl_name,
                'DSL Definicion':    blk_name,
            }
            fila_ancho.update(params)
            filas_ancho.append(fila_ancho)
            n_bloques += 1

        if n_bloques == 0:
            app.PrintWarn(f"   Sin bloques Reivax. "
                          f"Verifica DSL_SLOTS con los nombres reales arriba.")
        app.PrintInfo('')

    # ── Exportar ─────────────────────────────────────────────────────────
    app.PrintInfo(f"Total params extraidos : {len(filas_detalle)}")
    app.PrintInfo(f"Total bloques DSL      : {len(filas_ancho)}")
    app.PrintInfo('')

    ruta_excel, err = _escribir_excel(filas_detalle, filas_ancho, carpeta_salida)

    if ruta_excel and os.path.isfile(ruta_excel):
        tam_kb = os.path.getsize(ruta_excel) / 1024.0
        app.PrintInfo(SEP)
        app.PrintInfo(f"  Exportado : {os.path.basename(ruta_excel)}")
        app.PrintInfo(f"  Tamano    : {tam_kb:.1f} kB")
        app.PrintInfo(f"  Carpeta   : {carpeta_salida}")
    else:
        app.PrintError(f"Error al exportar Excel: {err}")

    app.PrintInfo(SEP)


main()
