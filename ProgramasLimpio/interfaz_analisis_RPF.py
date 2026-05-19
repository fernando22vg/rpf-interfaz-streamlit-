r"""
interfaz_analisis_RPF.py (v2 - Integrada)
Interfaz Streamlit integrada para workflow completo RPF.

Módulos:
  · Extracción de datos CNDC (ExtFLujos2daO)
  · Generación de condiciones iniciales (CondInicialesPF)
  · Carga en PowerFactory (CargaCondIniciales_PF_run)

Ejecutar con:
    streamlit run ProgramasLimpio\interfaz_analisis_RPF.py
"""
import os
import sys
import glob
import re
import json
import subprocess
import threading
import time
from plotly.subplots import make_subplots

import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go

# Detecta si la app corre en Streamlit Cloud (sin acceso a rutas Windows locales)
IS_CLOUD = not os.path.isdir(r"C:\Datos del CNDC")

if IS_CLOUD:
    try:
        import sharepoint_client as _sp
        _SP_OK = True
    except Exception as _sp_err:
        _SP_OK = False
        _SP_ERR_MSG = str(_sp_err)
else:
    _SP_OK = False

# ─────────────────────────────────────────────────────────────────────────────
# MÓDULOS DE GRÁFICAS ESTÁNDARES
# ─────────────────────────────────────────────────────────────────────────────
from graph_config import DEFAULT_GRAPH_CONFIG
from graph_builders import (
    create_dual_axis_timeseries,
    create_comparison_chart,
    add_kpi_markers,
    add_reference_lines,
    apply_standard_layout,
)

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL STYLES (for formatted exports)
# ─────────────────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

_HDR_FILL = PatternFill("solid", start_color="2E4057", end_color="2E4057")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_THIN     = Side(style="thin", color="CCCCCC")
_BORDER   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CTR      = Alignment(horizontal="center", vertical="center")
_LEFT     = Alignment(horizontal="left", vertical="center")
_RIGHT    = Alignment(horizontal="right", vertical="center")

# Specific colors for KPI tables
_KPI_OK_FILL   = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE") # Light green
_KPI_WARN_FILL = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C") # Light yellow
_KPI_ERROR_FILL = PatternFill("solid", start_color="FFCCCC", end_color="FFCCCC") # Light red

def _formato_hoja_excel(ws, df, kpi_col=None, kpi_ok_val="✅ Sí", kpi_error_val="❌ No"):
    """Aplica formato visual a la hoja de Excel exportada."""
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.fill = _HDR_FILL; cell.font = _HDR_FONT; cell.alignment = _CTR; cell.border = _BORDER

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.border = _BORDER; cell.alignment = _CTR
            if kpi_col and str(ws.cell(1, c).value) == kpi_col:
                val = str(cell.value)
                if kpi_ok_val in val: cell.fill = _KPI_OK_FILL
                elif kpi_error_val in val: cell.fill = _KPI_ERROR_FILL

    for c in range(1, ws.max_column + 1):
        col_let = get_column_letter(c)
        max_w = 0
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, c).value
            if val: max_w = max(max_w, len(str(val)))
        ws.column_dimensions[col_let].width = min(max_w + 3, 50)
    ws.freeze_panes = "A2"

def _apply_excel_formatting(df, sheet_name="Sheet1", kpi_col=None, kpi_ok_val="✅ Sí", kpi_error_val="❌ No"):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]

        # Apply formatting (header, borders, auto-width)
        _formato_hoja_excel(worksheet, df, kpi_col, kpi_ok_val, kpi_error_val)

    output.seek(0)
    return output.getvalue()

def _buscar_archivo_unidad(unit_name, file_list):
    """Busca en una lista de archivos el que mejor coincida con el nombre de la unidad."""
    if not unit_name: return None
    u_norm = str(unit_name).upper().replace("SYM_", "").replace("SYM", "")
    # Coincidencia exacta (normalizada)
    for f in file_list:
        f_base = os.path.splitext(f)[0].upper().replace("SYM_", "").replace("SYM", "")
        if u_norm == f_base: return f
    # Coincidencia por sub-cadena
    for f in file_list:
        f_base = os.path.splitext(f)[0].upper().replace("SYM_", "").replace("SYM", "")
        if u_norm in f_base or f_base in u_norm: return f
    return None


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN DE PÁGINA
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Analisis RPF",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES GLOBALÉS
# ─────────────────────────────────────────────────────────────────────────────
CARPETA_COBEE_EMF = "Resultados_COBEE" # Output folder for ExtractorResultadosCNDC.py
CARPETA_DATOS_CURVAS = "Datos Curvas" # Output folder for DatosCurvas_v3.py
CARPETA_COSTO_MARGINAL = "Costo Marginal STI" # Subcarpeta para archivos postot/td_

# Heurísticas para identificar columnas de frecuencia
FREQ_MIN_HZ, FREQ_MAX_HZ, FREQ_RANGE_MAX_HZ = 45.0, 55.0, 10.0

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN DE RUTAS — cargada desde archivo JSON (persistente)
# ─────────────────────────────────────────────────────────────────────────────
_CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config_rutas.json")

_DEFAULTS_CONFIG = {
    "RAIZ":               r"C:\Datos del CNDC\01_INFO CNDC_RPF",
    "RAIZ_DATOS":         r"C:\Datos del CNDC\02_DATOS CNDC_RPF",
    "PF_BASE":            r"C:\Program Files\DIgSILENT\PowerFactory 2025 SP2",
    "LOC_NAMES_GEN_PATH": r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\Designacion de loc_name\loc_names_gen.xlsx",
    "LOC_CAR_PATH":       r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\Designacion de loc_name\loc_name_cargas.xlsx",
    "LOC_XFO_PATH":       r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\Designacion de loc_name\loc_names_xfo.xlsx",
    "PF_PROYECTO":        "PMP_NOV25_OCT29_31102025(1)",
    "CASO_BASE":          "CNDC",
    "EXCLUIR_SLACK":      "sym_AGU02",
    "XFO_PF":             0.90,
}

def _cargar_config():
    if os.path.isfile(_CONFIG_PATH):
        try:
            with open(_CONFIG_PATH, encoding="utf-8") as _f:
                data = json.load(_f)
            # Combinar con defaults para tolerar claves nuevas
            return {**_DEFAULTS_CONFIG, **data}
        except Exception:
            pass
    return dict(_DEFAULTS_CONFIG)

def _guardar_config(cfg: dict):
    with open(_CONFIG_PATH, "w", encoding="utf-8") as _f:
        json.dump(cfg, _f, ensure_ascii=False, indent=2)

_cfg = _cargar_config()

# ─────────────────────────────────────────────────────────────────────────────
# FUNCIONES DE ANÁLISIS RPF — compartidas entre bloques 2, 3 y 4
# ─────────────────────────────────────────────────────────────────────────────

def _to_plotly_time(t_val, show_hhmmss):
    """Convierte segundos a Datetime para que Plotly los maneje correctamente."""
    if not show_hhmmss:
        return t_val
    res = pd.to_datetime(t_val, unit='s')
    # Solución TypeError (Pandas 2.x): Plotly usa sum() internamente en anotaciones de add_vline/add_hline.
    # El objeto pd.Timestamp no permite sumas con el '0' inicial de sum(). Al devolver milisegundos 
    # desde la época (float), se permite el cálculo de promedios para posicionar etiquetas.
    if isinstance(res, pd.Timestamp):
        return res.timestamp() * 1000
    return res

@st.cache_data(ttl=60)
def _listar_archivos_cache(directorio, patron, recursivo=False):
    """Cache para evitar escaneos repetitivos del sistema de archivos."""
    if not os.path.isdir(directorio):
        return []
    if recursivo:
        import glob as _glob
        return sorted(_glob.glob(os.path.join(directorio, "**", patron), recursive=True))
    return sorted([
        f for f in os.listdir(directorio) 
        if f.lower().endswith(patron.replace('*', '').lower()) and not f.startswith('~$')
    ])

def _kill_powerfactory():
    """Mata cualquier proceso PowerFactory.exe activo en segundo plano para liberar licencias."""
    try:
        # /F fuerza el cierre, /IM nombre de imagen, /T cierra procesos hijos (como la API)
        subprocess.run(["taskkill", "/F", "/IM", "PowerFactory.exe", "/T"], capture_output=True, check=False)
        return True
    except Exception:
        return False

def _leer_log_generic(path):
    """Helper ligero para leer archivos de texto."""
    try:
        return open(path, encoding="utf-8", errors="replace").read()
    except OSError:
        return ""

@st.fragment(run_every=2)
def _monitor_process_fragment(log_path, status_file):
    """Fragmento optimizado: actualiza solo el log sin recargar toda la interfaz."""
    log_txt = _leer_log_generic(log_path)
    if log_txt:
        st.text_area("📋 Log en vivo (actualización local)", value=log_txt, height=300, disabled=True)
    if os.path.exists(status_file):
        st.rerun() # Rerun global solo cuando el proceso termina

def _parse_to_seconds(series):
    """Versión optimizada y vectorizada para convertir tiempo (HH:MM:SS o float) a segundos."""
    # Convertir a string y limpiar una sola vez
    s = series.astype(str).str.strip().str.replace(',', '.')
    
    # Inicializar resultado con NaNs
    result = pd.Series(np.nan, index=series.index)
    
    # Procesar formato con dos puntos (HH:MM:SS) de forma vectorizada
    has_colon = s.str.contains(':')
    if has_colon.any():
        parts = s[has_colon].str.split(':')
        h = pd.to_numeric(parts.str[0], errors='coerce').fillna(0)
        m = pd.to_numeric(parts.str[1], errors='coerce').fillna(0)
        sec = pd.to_numeric(parts.str[2], errors='coerce').fillna(0)
        result[has_colon] = h * 3600 + m * 60 + sec
        
    # Procesar valores que ya son numéricos (como floats de segundos)
    not_colon = ~has_colon
    if not_colon.any():
        result[not_colon] = pd.to_numeric(s[not_colon], errors='coerce')
        
    return result.fillna(0.0)

@st.cache_data
def _load_tech_map(path):
    """Carga P_max por loc_name desde Detalle_PF de loc_names_gen.xlsx."""
    try:
        df = pd.read_excel(path, sheet_name="Detalle_PF", engine="calamine")
        # Búsqueda flexible de la columna de potencia nominal/máxima
        pot_cols = [c for c in df.columns if any(kw in c.lower() for kw in ['p_max', 'p nom', 'potencia'])]
        col = pot_cols[0] if pot_cols else 'P nom. (MW)'
        return df.set_index('loc_name PF')[[col]].rename(columns={col: 'P_max (MW)'}).to_dict('index')
    except Exception:
        
        return {}


@st.cache_data
def _load_pmax_cargado(ev_path, n_evento):
    """Lee Pmax_MW de pgini_GEN_FINAL del Excel de cargado PF.

    Siempre usa pgini_GEN_FINAL (la Pmax no cambia con el ajuste post-LF).
    Devuelve dict {loc_name_pf: pmax_mw}.
    """
    import glob as _glob
    candidates = sorted(
        _glob.glob(os.path.join(ev_path, f"datos_cargados_Ev{n_evento}*.xlsx")),
        key=os.path.getmtime, reverse=True,
    )
    for path in candidates:
        try:
            xl = pd.ExcelFile(path, engine="calamine")
            # Pmax_MW es fija — leer siempre de pgini_GEN_FINAL (no ajustado)
            if "pgini_GEN_FINAL" not in xl.sheet_names:
                continue
            df = xl.parse("pgini_GEN_FINAL")
            if "loc_name PF" not in df.columns or "Pmax_MW" not in df.columns:
                continue
            df["loc_name PF"] = df["loc_name PF"].astype(str).str.strip()
            result = {}
            for _, row in df.iterrows():
                try:
                    v = float(row["Pmax_MW"])
                    if v > 0:
                        result[row["loc_name PF"]] = v
                except (ValueError, TypeError):
                    pass
            if result:
                return result
        except Exception:
            continue
    return {}

def _resolver_unit_key(name, lookup_dict):
    """Versión optimizada de resolución de claves."""
    # Normalizar nombre: quitar extensión y prefijo sym_
    bare_name = os.path.splitext(name)[0].replace("sym_", "").upper()
    
    # Búsqueda rápida por set de candidatos
    candidates = {bare_name, f"SYM_{bare_name}", bare_name.lower(), f"sym_{bare_name.lower()}"}
    for c in candidates:
        if c in lookup_dict: return c, True

    # Búsqueda por sub-cadena (TIQ -> sym_TIQ01)
    for key in lookup_dict:
        k_norm = key.replace("sym_", "").replace("SYM_", "").upper()
        if bare_name in k_norm or k_norm in bare_name:
            return key, True
            
    # Fallback fuzzy (solo si lo anterior falla)
    return bare_name, False


def _get_pmax(tdat):
    """Extrae P_max de un registro de tech_map."""
    v = tdat.get('P_max (MW)', 100.0)
    try:
        v = float(v)
        return v if v > 0 else 100.0
    except Exception:
        return 100.0


def _get_pmax_from_cargado(unit_name, pmax_cargado, tech_map, fallback=100.0):
    """Obtiene P_max buscando primero en datos_cargados y luego en tech_map."""
    tk, found = _resolver_unit_key(unit_name, pmax_cargado)
    if found:
        return pmax_cargado[tk], tk, "datos_cargados"
    tk, found = _resolver_unit_key(unit_name, tech_map)
    if found:
        return _get_pmax(tech_map[tk]), tk, "loc_names_gen"
    return fallback, os.path.splitext(unit_name)[0], None


@st.cache_data(ttl=30, show_spinner=False)
def get_event_units(ev_path=None, n_evento=None):
    """Obtiene la lista maestra de unidades disponibles para el evento seleccionado."""
    ev_path = ev_path or st.session_state.ev_path_global
    n_evento = n_evento or st.session_state.n_evento_global
    if not ev_path: return []
    
    # Patrones a ignorar (no son unidades de generación)
    BLACKLIST = [
        "Velocidades", "Ángulos", "Resumen", "Info", "tabla_resultados",
        "F.", "F.P.", "P.F", "Barras", "frecuencia", "slack", "Evento"
    ]

    def _is_valid_unit(name):
        name_up = name.upper()
        return not any(p.upper() in name_up for p in BLACKLIST)
    
    def _clean_list(d, p):
        files = _listar_archivos_cache(d, p)
        # Extraer nombre, quitar sym_ para unificar, y filtrar por blacklist
        names = {os.path.splitext(f)[0].replace("sym_", "").replace("SYM_", "") for f in files}
        return {n for n in names if _is_valid_unit(n)}

    # Buscar en SCADA
    u_scada = _clean_list(os.path.join(ev_path, "Graficas Registro 1SEG COBEE"), "*.xlsx")
    # Buscar en EMF
    u_emf   = _clean_list(os.path.join(ev_path, CARPETA_COBEE_EMF), "*.xlsx")
    # Buscar en Simulación (E0 y E1)
    u_sim0  = _clean_list(os.path.join(ev_path, f"E{n_evento}.0", CARPETA_DATOS_CURVAS), "*.xlsx")
    u_sim1  = _clean_list(os.path.join(ev_path, f"E{n_evento}.1", CARPETA_DATOS_CURVAS), "*.xlsx")
    
    all_raw = u_scada | u_emf | u_sim0 | u_sim1
    return sorted(list(all_raw))



def _rp_cfg_path(loc_gen_path):
    return os.path.join(os.path.dirname(loc_gen_path), "estatismo_config.json")

def _load_rp_cfg(loc_gen_path):
    p = _rp_cfg_path(loc_gen_path)
    try:
        if os.path.isfile(p):
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def _save_rp_cfg(loc_gen_path, loc_key, droop_pct):
    p = _rp_cfg_path(loc_gen_path)
    cfg = _load_rp_cfg(loc_gen_path)
    cfg[loc_key] = round(float(droop_pct), 3)
    with open(p, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def _get_rp_default(loc_key, loc_gen_path, fallback=5.0):
    cfg = _load_rp_cfg(loc_gen_path)
    v = cfg.get(loc_key, cfg.get(loc_key.replace("sym_", ""), None))
    return float(v) if v is not None else fallback

def _event_cfg_path(ev_path):
    return os.path.join(ev_path, "event_config.json")

def _load_event_cfg(ev_path):
    p = _event_cfg_path(ev_path)
    try:
        if os.path.isfile(p):
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def _save_event_cfg(ev_path, key, value):
    p = _event_cfg_path(ev_path)
    cfg = _load_event_cfg(ev_path)
    cfg[key] = value
    try:
        with open(p, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        st.error(f"Error al guardar configuración del evento: {e}")
        return False

def _get_unit_cfg(ev_path, unit, key, default):
    cfg = _load_event_cfg(ev_path)
    return cfg.get("units", {}).get(unit, {}).get(key, default)

def _save_unit_cfg(ev_path, unit, key, value):
    cfg = _load_event_cfg(ev_path)
    if "units" not in cfg:
        cfg["units"] = {}
    if unit not in cfg["units"]:
        cfg["units"][unit] = {}
    cfg["units"][unit][key] = value
    return _save_event_cfg(ev_path, "units", cfg["units"])

def _sync_rpf_y_axis(key_to_update, widget_key):
    """Sincroniza los límites de ejes Y en todas las pestañas de análisis (Bloques 3, 4 y 5)."""
    if not st.session_state.get("global_selected_unit"): return
    val = st.session_state.get(widget_key)
    if val is None: return
    
    st.session_state[f"b3_sync_{key_to_update}"] = val

    # Lista completa de prefijos de widgets en todos los bloques para sincronización cruzada
    # Bloque 3: sc (SCADA), emf (EMF), comp (Comparativa)
    # Bloque 4: sim0 (E0), sim1 (E1), simc (Comparativa Simu)
    # Bloque 5: val (Validación)
    sync_prefixes = ["b2_sc", "b2_emf", "b3_comp", "b3_sim0", "b3_sim1", "b3_simc", "b4_val"]
    for pfx in sync_prefixes:
        target_key = f"{pfx}_{key_to_update}"
        if target_key in st.session_state:
            st.session_state[target_key] = val

    _save_unit_cfg(st.session_state.ev_path_global, st.session_state.global_selected_unit, key_to_update, val)

def _sync_session_scale_config(ev_path, unit_name):
    """
    Sincroniza el session_state de Streamlit con los valores guardados en JSON.
    Esto asegura que los inputs numéricos siempre muestren los valores correctos
    al cambiar de unidad o evento.
    
    Parámetros:
        ev_path: ruta del evento
        unit_name: nombre de la unidad a sincronizar
    """
    if not unit_name or not ev_path:
        return
    
    # Cargar valores guardados con defaults
    y_f_min = _get_unit_cfg(ev_path, unit_name, "y_f_min", 49.0)
    y_f_max = _get_unit_cfg(ev_path, unit_name, "y_f_max", 51.0)
    y_p_min = _get_unit_cfg(ev_path, unit_name, "y_p_min", 0.0)
    y_p_max = _get_unit_cfg(ev_path, unit_name, "y_p_max", 200.0)
    y_auto = _get_unit_cfg(ev_path, unit_name, "y_auto", True)
    
    # 1. Variables sincronizadas para los gráficos
    st.session_state.b3_sync_y_f_min = float(y_f_min) if y_f_min is not None else 49.0
    st.session_state.b3_sync_y_f_max = float(y_f_max) if y_f_max is not None else 51.0
    st.session_state.b3_sync_y_p_min = float(y_p_min) if y_p_min is not None else 0.0
    st.session_state.b3_sync_y_p_max = float(y_p_max) if y_p_max is not None else 200.0
    st.session_state.b3_sync_y_auto = bool(y_auto) if y_auto is not None else True

    # 2. Sincronizar las llaves de los widgets de todas las pestañas para evitar valores obsoletos en la UI
    for pfx in ["b2_sc", "b2_emf", "b3_comp"]:
        st.session_state[f"{pfx}_y_f_min"] = st.session_state.b3_sync_y_f_min
        st.session_state[f"{pfx}_y_f_max"] = st.session_state.b3_sync_y_f_max
        st.session_state[f"{pfx}_y_p_min"] = st.session_state.b3_sync_y_p_min
        st.session_state[f"{pfx}_y_p_max"] = st.session_state.b3_sync_y_p_max
        st.session_state[f"{pfx}_y_auto"] = st.session_state.b3_sync_y_auto

    # 3. Sincronizar llaves de ejes X (aunque sean independientes entre pestañas)
    st.session_state.b2_sc_xaxis_min = float(_get_unit_cfg(ev_path, unit_name, "b3_tab_scada_xmin", 0.0))
    st.session_state.b2_sc_xaxis_max = float(_get_unit_cfg(ev_path, unit_name, "b3_tab_scada_xmax", 100.0))
    st.session_state.b2_emf_xaxis_min = float(_get_unit_cfg(ev_path, unit_name, "b3_tab_emf_xmin", 0.0))
    st.session_state.b2_emf_xaxis_max = float(_get_unit_cfg(ev_path, unit_name, "b3_tab_emf_xmax", 100.0))
    st.session_state.b3_comp_xmin = float(_get_unit_cfg(ev_path, unit_name, "b3_tab_comp_xmin", -10.0))
    st.session_state.b3_comp_xmax = float(_get_unit_cfg(ev_path, unit_name, "b3_tab_comp_xmax", 100.0))

    # 4. Sincronizar llaves de simulación (Bloque 4) - Persistencia por evento
    st.session_state.b3_sim0_xmin = float(_get_unit_cfg(ev_path, unit_name, "sim0_xmin", 0.0))
    st.session_state.b3_sim0_xmax = float(_get_unit_cfg(ev_path, unit_name, "sim0_xmax", 100.0))
    st.session_state.b3_sim1_xmin = float(_get_unit_cfg(ev_path, unit_name, "sim1_xmin", 0.0))
    st.session_state.b3_sim1_xmax = float(_get_unit_cfg(ev_path, unit_name, "sim1_xmax", 100.0))
    st.session_state.b4_simcomp_xmin = float(_get_unit_cfg(ev_path, unit_name, "simcomp_xmin", 0.0))
    st.session_state.b4_simcomp_xmax = float(_get_unit_cfg(ev_path, unit_name, "simcomp_xmax", 100.0))
    
    # Sincronizar selectores de auto-escala específicos
    st.session_state.b3_sim0_auto_toggle = bool(_get_unit_cfg(ev_path, unit_name, "y_auto", True))
    st.session_state.b3_sim1_auto_toggle = bool(_get_unit_cfg(ev_path, unit_name, "y_auto", True))
    st.session_state.b3_simcomp_auto_toggle = bool(_get_unit_cfg(ev_path, unit_name, "y_auto", True))

    # Sincronizar llaves de ejes Y para Bloque 4 (Simulación)
    for sim_pfx in ["b3_sim0", "b3_sim1", "b3_comp"]:
        st.session_state[f"{sim_pfx}_y1min"] = st.session_state.b3_sync_y_f_min
        st.session_state[f"{sim_pfx}_y1max"] = st.session_state.b3_sync_y_f_max
        st.session_state[f"{sim_pfx}_y2min"] = st.session_state.b3_sync_y_p_min
        st.session_state[f"{sim_pfx}_y2max"] = st.session_state.b3_sync_y_p_max

def _widget_pmax_rp(loc_key, loc_gen_path, extra_cols=None, key_prefix=""):
    """
    Renderiza una fila con P_max (desde Excel) y Estatismo editable + botón guardar.
    Devuelve (p_max_mw, rp_decimal).
    extra_cols: lista de (label, widget_fn) para columnas adicionales antes del botón.
    """
    tmap = _load_tech_map(loc_gen_path)
    td   = tmap.get(loc_key, tmap.get(f"sym_{loc_key}", {"P_max (MW)": 100.0}))
    pmax_val = _get_pmax(td)
    rp_val   = _get_rp_default(loc_key, loc_gen_path)

    n_extra = len(extra_cols) if extra_cols else 0
    cols = st.columns([2, 2] + ([1] * n_extra) + [1])

    p_max_out = cols[0].number_input(
        "P_max [MW]", value=float(pmax_val),
        min_value=0.1, step=0.1, key=f"{key_prefix}_pmax",
        help="Cargado automáticamente desde loc_names_gen.xlsx",
    )
    rp_pct = cols[1].number_input(
        "Estatismo [%]", value=float(rp_val),
        min_value=0.1, max_value=20.0, step=0.1, key=f"{key_prefix}_rp",
    )
    if extra_cols:
        for i, (lbl, fn) in enumerate(extra_cols):
            fn(cols[2 + i])

    cols[-1].markdown("&nbsp;", unsafe_allow_html=True)
    if cols[-1].button("💾", key=f"{key_prefix}_save_rp",
                       help="Guardar Estatismo para esta unidad"):
        _save_rp_cfg(loc_gen_path, loc_key, rp_pct)
        st.toast(f"Estatismo {loc_key}: {rp_pct:.1f}% guardado", icon="✅")

    return float(p_max_out), float(rp_pct) / 100.0


def _detectar_inicio_falla(freq_array, umbral_dfdt=-0.02, ventana_suavizado=5):
    """Detección robusta del inicio de falla por df/dt sostenido sobre señal suavizada."""
    n = len(freq_array)
    if n < ventana_suavizado + 2:
        return 0
    
    # Suavizado usando convolución de NumPy (más rápido que rolling de Pandas)
    kernel = np.ones(ventana_suavizado) / ventana_suavizado
    freq_smooth = np.convolve(freq_array.astype(float), kernel, mode='same')
    
    # Corregir bordes de la convolución
    freq_smooth[:ventana_suavizado//2] = freq_smooth[ventana_suavizado//2]
    freq_smooth[-(ventana_suavizado//2):] = freq_smooth[-(ventana_suavizado//2)-1]

    dfdt = np.diff(freq_smooth)
    
    # Detección vectorizada de caída sostenida
    condicion = (dfdt[:-1] < umbral_dfdt) & (dfdt[1:] < umbral_dfdt)
    indices = np.where(condicion)[0]
    
    if len(indices) > 0:
        return int(indices[0] + 1)

    candidatos = np.where(dfdt < umbral_dfdt)[0]
    return int(candidatos[0] + 1) if len(candidatos) > 0 else 0


def _cndc_kpis(t_arr, freq_arr, pot_arr, p_max, rp, delta_t, f_nom=50.0):
    """KPIs según metodología oficial CNDC RPF.

    Puntos: t₀ (f₀, P₀), nadir (f_min, t_min), t₀+Δt (f_Δt, P_Δt).
    ΔP = P_Δt − P₀ ; ΔP% = ΔP/P_max×100 ; Aporta si ΔP% ≥ 1.5 %.
    Droop = (Δf'/f_nom) / (ΔP/P_max) × 100, con banda muerta ±25 mHz.
    """
    if len(freq_arr) == 0 or p_max <= 0:
        return None
    idx_t0 = int(np.argmin(np.abs(t_arr)))
    f0 = float(freq_arr[idx_t0])
    p0 = float(pot_arr[idx_t0])
    mask_post = t_arr >= 0
    if not np.any(mask_post):
        return None
    t_post, f_post = t_arr[mask_post], freq_arr[mask_post]
    idx_nadir = np.argmin(f_post)
    f_min = float(f_post[idx_nadir])
    t_min = float(t_post[idx_nadir])
    delta_f = f0 - f_min
    # Paso 2: Puntos de evaluación (t0 + Delta_t)
    idx_dt = np.argmin(np.abs(t_arr - delta_t))
    f_dt = float(freq_arr[idx_dt])
    p_dt = float(pot_arr[idx_dt])

    # Paso 3: Cálculo del aporte
    dp = p_dt - p0
    dp_pct = (dp / p_max) * 100
    r_inic = p_max - p0
    r_inic_pct = (r_inic / p_max) * 100
    
    # Paso 4: Cálculo del Droop (Estatismo) con banda muerta de 25mHz
    # f_ref = 49.975 si f0 > 49.975 (considera banda muerta)
    f_ref = 49.975 if f0 > 49.975 else f0
    df_prime = f_ref - f_dt
    droop_calc = (df_prime / f_nom) / (dp / p_max) * 100 if abs(dp) > 0.001 else float('nan')

    return {
        'f0': round(f0, 4), 'p0': round(p0, 3),
        'f_min': round(f_min, 4), 't_min': round(t_min, 1), 'delta_f': round(delta_f, 4),
        'f_dt': round(f_dt, 4), 'p_dt': round(p_dt, 3), 't_dt': int(delta_t),
        'r_inic': round(r_inic, 3), 'r_inic_pct': round(r_inic_pct, 2),
        'dp': round(dp, 3), 'dp_pct': round(dp_pct, 2),
        'droop_calc': round(droop_calc, 2) if droop_calc == droop_calc else '—',
        'droop_nom': round(float(rp) * 100, 1),
        'aporta': dp_pct >= 1.5,
    }


def _calcular_rocof(t_arr, f_arr, ventana_s=3.0):
    """ROCOF [Hz/s] por regresión lineal en la ventana [0, ventana_s] post-falla."""
    mask = (t_arr >= 0) & (t_arr <= ventana_s)
    if np.sum(mask) < 2:
        return float('nan')
    t_w, f_w = t_arr[mask], f_arr[mask]
    valid = np.isfinite(t_w) & np.isfinite(f_w)
    t_w, f_w = t_w[valid], f_w[valid]
    if len(t_w) < 2 or np.ptp(t_w) == 0:
        return float('nan')
    try:
        coeffs = np.polyfit(t_w - t_w[0], f_w, 1)
        return round(float(coeffs[0]), 4)
    except np.linalg.LinAlgError:
        return float('nan')


def _add_marker(fig, t, y, label, symbol, color, yaxis='y', size=12):
    fig.add_trace(go.Scatter(
        x=[t], y=[y],
        mode='markers+text',
        marker=dict(symbol=symbol, size=size, color=color,
                    line=dict(width=1.5, color='white')),
        text=[label], textposition='top right',
        textfont=dict(size=10, color=color),
        yaxis=yaxis, showlegend=False, hoverinfo='skip',
    ))

def _get_kpi_table_data_lists(kpi, p_max, delta_t, rocof):
    """Helper para convertir el dict de KPI a listas para una Tabla de Plotly."""
    if not kpi: return [], []
    rows = [
        ("P_max [MW]", f"{p_max:.2f}"),
        ("f₀ [Hz]", f"{kpi['f0']:.4f}"),
        ("P₀ [MW]", f"{kpi['p0']:.3f}"),
        ("f_min [Hz]", f"{kpi['f_min']:.4f}"),
        ("t_min [s]", f"{kpi['t_min']:.1f}"),
        ("Δf [Hz]", f"{kpi['delta_f']:.4f}"),
        (f"f_Δt ({delta_t}s) [Hz]", f"{kpi['f_dt']:.4f}"),
        (f"P_Δt ({delta_t}s) [MW]", f"{kpi['p_dt']:.3f}"),
        ("ΔP [MW]", f"{kpi['dp']:.3f}"),
        ("ΔP% [%]", f"{kpi['dp_pct']:.2f}"),
        ("¿Aporta?", "Sí" if kpi['aporta'] else "No"),
        ("Droop Nom. [%]", f"{kpi['droop_nom']:.1f}"),
        ("Droop Calc. [%]", str(kpi['droop_calc'])),
    ]
    if rocof is not None and rocof == rocof:
        rows.append(("ROCOF [Hz/s]", f"{rocof:.4f}"))
    return [r[0] for r in rows], [r[1] for r in rows]

def _create_kpi_summary_table_fig(kpi, p_max, delta_t, rocof, unit_name, fuente):
    """Crea una figura de Plotly que contiene solo la tabla de KPIs para exportar como imagen compacta."""
    if not kpi: return None
    labels, values = _get_kpi_table_data_lists(kpi, p_max, delta_t, rocof)
    
    fig = go.Figure(data=[go.Table(
        header=dict(
            values=[['<b>Parámetro</b>'], [f'<b>{fuente}</b>']],
            fill_color='#2E4057',
            align='left',
            font=dict(color='white', size=12)
        ),
        cells=dict(
            values=[labels, values],
            fill_color='#F5F5F5',
            align='left',
            font=dict(color='#333333', size=11),
            height=22
        )
    )])
    
    fig.update_layout(
        title=dict(text=f"Resumen KPIs: {unit_name}", x=0.05, y=0.98),
        width=400,
        height=450,
        margin=dict(l=10, r=10, t=50, b=10)
    )
    return fig

def _load_sim_data_only(sim_type_suffix, sel_file, n_evento, ev_path):
    sim_dir = os.path.join(ev_path, f"E{n_evento}.{sim_type_suffix}", CARPETA_DATOS_CURVAS)
    if not sel_file or not os.path.isdir(sim_dir):
        return None, None
    df_sim = pd.read_excel(os.path.join(sim_dir, sel_file), engine="calamine").dropna()
    return df_sim, sel_file

def _mostrar_tabla_cndc(kpi, p_max, delta_t, fuente="Valor", rocof=None):
    """Renderiza la tabla CNDC siguiendo los pasos del manual."""
    if not kpi:
        st.warning("No se pudieron calcular KPIs.")
        return
    rows = [
        ("Potencia Efectiva (P_max) [MW]", f"{p_max:.2f}"),
        ("Paso 2: f₀ (Inicio evento) [Hz]", f"{kpi['f0']:.4f}"),
        ("Paso 2: P₀ (Inicio evento) [MW]", f"{kpi['p0']:.3f}"),
        ("Paso 2: f_min (Nadir) [Hz]",      f"{kpi['f_min']:.4f}"),
        ("Paso 2: t_min (Nadir) [s]",       f"{kpi['t_min']:.1f}"),
        ("Δf (f₀ − f_min) [Hz]",           f"{kpi['delta_f']:.4f}"),
        (f"Paso 2: f_Δt ({delta_t}s) [Hz]", f"{kpi['f_dt']:.4f}"),
        (f"Paso 2: P_Δt ({delta_t}s) [MW]", f"{kpi['p_dt']:.3f}"),
        ("Paso 3: Reserva Inicial (R_inic) [MW]",   f"{kpi['r_inic']:.3f}"),
        ("Paso 3: Reserva Inicial (R_inic) [%]",    f"{kpi['r_inic_pct']:.2f}"),
        ("Paso 3: Potencia entregada (ΔP) [MW]",     f"{kpi['dp']:.3f}"),
        ("Paso 3: Aporte Porcentual (ΔP%) [%]",      f"{kpi['dp_pct']:.2f}"),
        ("¿Aporta a la RPF? (ΔP% ≥ 1.5%)",           "✅ Sí" if kpi['aporta'] else "❌ No"),
        ("Estatismo (Droop) Nominal [%]",            f"{kpi['droop_nom']:.1f}"),
        ("Paso 4: Estatismo (Droop) Calculado [%]",  str(kpi['droop_calc'])),
    ]
    if rocof is not None and rocof == rocof:
        rows.append(("Paso 2: ROCOF (df/dt inicial) [Hz/s]", f"{rocof:.4f}"))
    _df_t = _df_safe(pd.DataFrame(rows, columns=["Parámetro", fuente]))
    def _col_ap(row):
        styles = [''] * len(row)
        if row["Parámetro"] == "Aporta a la RPF":
            for i in range(1, len(row)):
                v = str(row.iloc[i])
                styles[i] = 'background-color:#d4edda' if '✅' in v else 'background-color:#f8d7da'
        return styles
    st.dataframe(_df_t.style.apply(_col_ap, axis=1), use_container_width=True, hide_index=True)

    # Export button for KPI table
    if st.button(f"⬇️ Descargar KPIs a Excel ({fuente})", key=f"dl_kpis_{fuente.replace(' ', '_')}"):
        excel_data = _apply_excel_formatting(
            _df_t,
            sheet_name=f"KPIs_{fuente.replace(' ', '_')}",
            kpi_col=fuente,
            kpi_ok_val="✅ Sí",
            kpi_error_val="❌ No"
        )
        st.download_button(f"Descargar {fuente} KPIs", excel_data,
                           file_name=f"kpis_{fuente.replace(' ', '_')}_Ev{st.session_state.n_evento_global}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─────────────────────────────────────────────────────────────────────────────
# INICIALIZAR SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
if "global_selected_unit" not in st.session_state:
    st.session_state.global_selected_unit = None

# Inicialización robusta de configuración de gráficas
if "graph_config" not in st.session_state:
    st.session_state.graph_config = dict(DEFAULT_GRAPH_CONFIG)
else:
    # Asegurar que todas las claves por defecto existan (evita KeyErrors en v3)
    # preservando los valores personalizados del usuario.
    for key, value in DEFAULT_GRAPH_CONFIG.items():
        if key not in st.session_state.graph_config:
            st.session_state.graph_config[key] = value
if "semestre_global" not in st.session_state:
    st.session_state.semestre_global = None
if "evento_global" not in st.session_state:
    st.session_state.evento_global = None
if "ev_path_global" not in st.session_state:
    st.session_state.ev_path_global = None
if "n_evento_global" not in st.session_state:
    st.session_state.n_evento_global = None
if "pf_running" not in st.session_state:
    st.session_state.pf_running = False
if "pf_return_code" not in st.session_state:
    st.session_state.pf_return_code = None
if "pf_waiting_close" not in st.session_state:
    st.session_state.pf_waiting_close = False
if "pf_needs_rerun" not in st.session_state:
    st.session_state.pf_needs_rerun = False
if "pf_status_file" not in st.session_state:
    st.session_state.pf_status_file = None
if "pf_log_file" not in st.session_state:
    st.session_state.pf_log_file = None
if "pf_saved_log" not in st.session_state:
    st.session_state.pf_saved_log = None
# ── Tab 1: extracción CNDC ────────────────────────────────────────────────────
if "ext_running" not in st.session_state:
    st.session_state.ext_running = False
if "ext_status_file" not in st.session_state:
    st.session_state.ext_status_file = None
if "ext_return_code" not in st.session_state:
    st.session_state.ext_return_code = None
if "ext_log_file" not in st.session_state:
    st.session_state.ext_log_file = None
if "ext_saved_log" not in st.session_state:
    st.session_state.ext_saved_log = None
# ── Tab 1b: CondInicialesPF ───────────────────────────────────────────────────
if "ci_running" not in st.session_state:
    st.session_state.ci_running = False
if "ci_status_file" not in st.session_state:
    st.session_state.ci_status_file = None
if "ci_return_code" not in st.session_state:
    st.session_state.ci_return_code = None
if "ci_log_file" not in st.session_state:
    st.session_state.ci_log_file = None
if "ci_saved_log" not in st.session_state:
    st.session_state.ci_saved_log = None
# ── Bloque 0: DatsoGENBUSLNE (extracción modelo base) ────────────────────────
if "mod_running" not in st.session_state:
    st.session_state.mod_running = False
if "mod_status_file" not in st.session_state:
    st.session_state.mod_status_file = None
if "mod_return_code" not in st.session_state:
    st.session_state.mod_return_code = None
if "mod_log_file" not in st.session_state:
    st.session_state.mod_log_file = None
if "mod_saved_log" not in st.session_state:
    st.session_state.mod_saved_log = None
# ── Bloque 0: scripts adicionales de modelo base ─────────────────────────────
for _pfx in ("gen", "lne", "xfo", "sht", "car"):
    if f"{_pfx}_running"     not in st.session_state: st.session_state[f"{_pfx}_running"]     = False
    if f"{_pfx}_status_file" not in st.session_state: st.session_state[f"{_pfx}_status_file"] = None
    if f"{_pfx}_return_code" not in st.session_state: st.session_state[f"{_pfx}_return_code"] = None
    if f"{_pfx}_log_file"    not in st.session_state: st.session_state[f"{_pfx}_log_file"]    = None
    if f"{_pfx}_saved_log"   not in st.session_state: st.session_state[f"{_pfx}_saved_log"]   = None
# ── Tab 2: OrdenadorDatosEvento (SCADA) ──────────────────────────────────────
if "scada_running" not in st.session_state:
    st.session_state.scada_running = False
if "scada_status_file" not in st.session_state:
    st.session_state.scada_status_file = None
if "scada_return_code" not in st.session_state:
    st.session_state.scada_return_code = None
if "scada_log_file" not in st.session_state:
    st.session_state.scada_log_file = None
if "scada_saved_log" not in st.session_state:
    st.session_state.scada_saved_log = None
# ── Tab 3: ExtractorResultadosCNDC (EMF) ─────────────────────────────────────
if "emf_running" not in st.session_state:
    st.session_state.emf_running = False
if "emf_status_file" not in st.session_state:
    st.session_state.emf_status_file = None
if "emf_return_code" not in st.session_state:
    st.session_state.emf_return_code = None
if "emf_log_file" not in st.session_state:
    st.session_state.emf_log_file = None
if "emf_saved_log" not in st.session_state:
    st.session_state.emf_saved_log = None
# ── Bloque 3: Análisis de datos ──────────────────────────────────────────────
if "b2_scada_df" not in st.session_state:
    st.session_state.b2_scada_df = None
if "b2_emf_df" not in st.session_state:
    st.session_state.b2_emf_df = None
if "b2_selected_unit" not in st.session_state:
    st.session_state.b2_selected_unit = None
# Inicialización global de variables de sincronización para el Bloque 3
# Global defaults for Block 3 Y-axis synchronization
if "b3_sync_y_f_min" not in st.session_state: st.session_state.b3_sync_y_f_min = 49.0
if "b3_sync_y_f_max" not in st.session_state: st.session_state.b3_sync_y_f_max = 51.0
if "b3_sync_y_p_min" not in st.session_state: st.session_state.b3_sync_y_p_min = 0.0
if "b3_sync_y_p_max" not in st.session_state: st.session_state.b3_sync_y_p_max = 200.0
if "b3_sync_y_p_max" not in st.session_state: st.session_state.b3_sync_y_p_max = 200.0 # Generic default
if "b3_sync_y_auto" not in st.session_state: st.session_state.b3_sync_y_auto = True
if "b3_last_unit" not in st.session_state: st.session_state.b3_last_unit = None
if "b3_last_event_path" not in st.session_state: st.session_state.b3_last_event_path = None
if "b3_selected_unit" not in st.session_state:
    st.session_state.b3_selected_unit = None

# --- Variables para persistencia de descarga de KPIs compactos ---
if "b3_kpi_zip_bytes" not in st.session_state: st.session_state.b3_kpi_zip_bytes = None
if "b3_kpi_zip_name" not in st.session_state: st.session_state.b3_kpi_zip_name = ""
if "b3_kpi_zip_count" not in st.session_state: st.session_state.b3_kpi_zip_count = 0
if "b3_kpi_excel_bytes" not in st.session_state: st.session_state.b3_kpi_excel_bytes = None
if "b3_kpi_excel_name" not in st.session_state: st.session_state.b3_kpi_excel_name = ""

# --- Variables para persistencia de descarga de gráficos ZIP ---
if "b3_plots_zip_bytes" not in st.session_state: st.session_state.b3_plots_zip_bytes = None
if "b3_plots_zip_name" not in st.session_state: st.session_state.b3_plots_zip_name = ""
if "b4_sim_zip_bytes" not in st.session_state: st.session_state.b4_sim_zip_bytes = None
if "b4_sim_zip_name" not in st.session_state: st.session_state.b4_sim_zip_name = ""


# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR — CONFIGURACIÓN Y SELECTOR DE MÓDULO
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.title("⚙️ Configuración")
    
    # ─── NAVEGACIÓN DE BLOQUES (Prioridad Superior) ──────────────────────
    st.subheader(" Flujo de Trabajo")
    bloque_trabajo = st.radio(
        "Seleccione fase del análisis:",
        options=["modelo_base", "carga_datos", "config_unidades", "analisis_datos", "analisis_simulacion", "comparativa_real_simu", "reporte_tecnico", "config_global"],
        format_func=lambda x: {
            "modelo_base":          " 0: Datos del Modelo",
            "carga_datos":          "1: Carga de Datos",
            "config_unidades":      "2: Configuración Unidades",
            "analisis_datos":       "3: Análisis Registrados",
            "analisis_simulacion":  "4: Análisis Simulación",
            "comparativa_real_simu":"5: Real vs. Simulación",
            "reporte_tecnico":      "6: Reporte Técnico",
            "config_global":        "7: Configuración Gráficas",
        }[x],
        index=1, # Por defecto Carga de Datos
        key="bloque_radio_nav",
    )
    st.markdown("---")

    _any_running = (st.session_state.pf_running or st.session_state.mod_running
        or any(st.session_state.get(f"{_p}_running") for _p in ("gen","lne","xfo","sht","car"))
        or st.session_state.ext_running or st.session_state.ci_running
        or st.session_state.scada_running or st.session_state.emf_running)
    if _any_running:
        st.warning("Proceso en ejecución — navegación bloqueada.")

    # ─── SELECCIÓN DE SEMESTRE Y EVENTO (GLOBAL) ──────────────────────────
    st.subheader("📅 Evento Actual")

# Helper para identificar columnas de frecuencia
def _is_frequency_column(col_name, series_data):
    # Check by name
    if "frecuencia" in col_name.lower() or "freq" in col_name.lower() or "hz" in col_name.lower() or "m:f" in col_name.lower():
        return True
    # Check by value range (assuming 50Hz system)
    if len(series_data) > 1:
        numeric = pd.to_numeric(series_data, errors='coerce').dropna()
        if len(numeric) > 1:
            min_val = numeric.min()
            max_val = numeric.max()
            if FREQ_MIN_HZ <= min_val <= FREQ_MAX_HZ and FREQ_MIN_HZ <= max_val <= FREQ_MAX_HZ and (max_val - min_val) < FREQ_RANGE_MAX_HZ:
                return True
    return False


def _robust_col_detect(df):
    """Detecta columnas de tiempo, frecuencia y potencia en DataFrames de simulación."""
    cols = df.columns.tolist()
    tc = cols[0]  # Usualmente la primera es el tiempo

    # Buscar frecuencia
    fc_cands = [c for c in cols[1:] if _is_frequency_column(c, df[c])]
    fc_col = fc_cands[0] if fc_cands else cols[1]

    # Buscar potencia (lo que no sea tiempo ni frecuencia)
    pc_cands = [c for c in cols[1:] if c != fc_col]
    pc_col = pc_cands[0] if pc_cands else (cols[2] if len(cols) > 2 else fc_col)

    return tc, fc_col, pc_col

def _short_col_name(col):
    """Extrae el nombre corto de columnas con ruta completa de PowerFactory.
    Ejemplo: '\\user\\Prj\\SIN.ElmNet\\sym_ZON01.ElmSym' → 'sym_ZON01'
    """
    s = str(col).strip("': ")
    if '\\' in s:
        last = s.split('\\')[-1].strip("': ")
        if '.' in last:
            last = last.rsplit('.', 1)[0]
        return last or s
    return s


with st.sidebar:
    RAIZ = st.text_input(
        "Ruta base CNDC",
        value=_cfg["RAIZ"],
        help="Carpeta raíz donde están los semestres.",
        key="cfg_RAIZ",
    )

    RAIZ_DATOS = st.text_input(
        "Ruta origen de datos (SCADA/EMF)",
        value=_cfg["RAIZ_DATOS"],
        help="Ruta donde se encuentran los archivos fuente para procesar.",
        key="cfg_RAIZ_DATOS",
    )

    show_hhmmss = st.checkbox("Mostrar tiempo en HH:MM:SS", value=False, key="global_show_hhmmss")

    # ── Selector de semestre y evento ────────────────────────────────────────
    if IS_CLOUD:
        # ── Modo nube: datos desde SharePoint ────────────────────────────────
        if not _SP_OK:
            st.error(f"❌ No se pudo conectar a SharePoint: {_SP_ERR_MSG}")
            st.session_state.semestre_global = None
        else:
            try:
                semestres = _sp.listar_semestres()
            except Exception as _e:
                st.error(f"❌ Error listando semestres en SharePoint:\n{_e}")
                semestres = []

            if semestres:
                idx_sem = 0
                if st.session_state.semestre_global in semestres:
                    idx_sem = semestres.index(st.session_state.semestre_global)
                semestre_sel = st.selectbox(
                    "Seleccione Semestre", semestres, index=idx_sem, key="sel_semestre_global"
                )
                st.session_state.semestre_global = semestre_sel
            else:
                st.warning("❌ No se encontraron semestres en SharePoint")
                st.session_state.semestre_global = None

            if st.session_state.semestre_global:
                try:
                    eventos = _sp.listar_eventos(st.session_state.semestre_global)
                except Exception as _e:
                    st.error(f"❌ Error listando eventos:\n{_e}")
                    eventos = []

                if eventos:
                    idx_ev = 0
                    if st.session_state.evento_global in eventos:
                        idx_ev = eventos.index(st.session_state.evento_global)
                    evento_sel = st.selectbox(
                        "Seleccione Evento", eventos, index=idx_ev, key="sel_evento_global"
                    )
                    st.session_state.evento_global = evento_sel

                    # Descargar evento a /tmp/ si cambia la selección
                    _ev_key = f"{st.session_state.semestre_global}/{st.session_state.evento_global}"
                    if st.session_state.get("_sp_ev_key") != _ev_key:
                        with st.spinner("⬇️ Descargando datos del evento desde SharePoint..."):
                            try:
                                _local_ev = _sp.descargar_evento(
                                    st.session_state.semestre_global,
                                    st.session_state.evento_global,
                                )
                                st.session_state._sp_ev_local = str(_local_ev)
                                st.session_state._sp_ev_key = _ev_key
                            except Exception as _e:
                                st.error(f"❌ Error descargando evento:\n{_e}")
                                st.session_state._sp_ev_local = None

                    ev_path = st.session_state.get("_sp_ev_local")
                    if ev_path:
                        m_ev = re.search(r"(\d+)$", st.session_state.evento_global.strip())
                        n_evento = m_ev.group(1) if m_ev else st.session_state.evento_global.split()[-1]
                        st.session_state.ev_path_global = ev_path
                        st.session_state.n_evento_global = n_evento
                        if st.session_state.get("last_n_evento_global") != n_evento:
                            st.session_state.b3_kpi_zip_bytes = None
                            st.session_state.b3_kpi_excel_bytes = None
                            st.session_state.b3_plots_zip_bytes = None
                            st.session_state.b4_sim_zip_bytes = None
                            st.session_state.last_n_evento_global = n_evento
                        st.success(f"Evento {n_evento} listo")
                else:
                    st.warning("❌ No hay eventos en este semestre")
                    st.session_state.evento_global = None
            else:
                st.info("← Seleccione semestre primero")
                st.session_state.evento_global = None

    else:
        # ── Modo local: rutas de Windows ─────────────────────────────────────
        if os.path.isdir(RAIZ):
            semestres = sorted(
                d for d in os.listdir(RAIZ)
                if os.path.isdir(os.path.join(RAIZ, d))
            )
            if semestres:
                idx_sem = 0
                if st.session_state.semestre_global in semestres:
                    idx_sem = semestres.index(st.session_state.semestre_global)
                semestre_sel = st.selectbox(
                    "Seleccione Semestre", semestres, index=idx_sem, key="sel_semestre_global"
                )
                st.session_state.semestre_global = semestre_sel
            else:
                st.warning("❌ No se encontraron semestres")
                st.session_state.semestre_global = None
        else:
            st.error(f"❌ Ruta no encontrada:\n{RAIZ}")
            st.session_state.semestre_global = None

        if st.session_state.semestre_global:
            base_ev = os.path.join(RAIZ, st.session_state.semestre_global, "Análisis_todos_los_eventos")
            if os.path.isdir(base_ev):
                eventos = sorted(
                    d for d in os.listdir(base_ev)
                    if os.path.isdir(os.path.join(base_ev, d))
                )
                if eventos:
                    idx_ev = 0
                    if st.session_state.evento_global in eventos:
                        idx_ev = eventos.index(st.session_state.evento_global)
                    evento_sel = st.selectbox(
                        "Seleccione Evento", eventos, index=idx_ev, key="sel_evento_global"
                    )
                    st.session_state.evento_global = evento_sel
                    ev_path = os.path.join(RAIZ, st.session_state.semestre_global,
                                           "Análisis_todos_los_eventos", st.session_state.evento_global)
                    m_ev = re.search(r"(\d+)$", st.session_state.evento_global.strip())
                    n_evento = m_ev.group(1) if m_ev else st.session_state.evento_global.split()[-1]
                    st.session_state.ev_path_global = ev_path
                    st.session_state.n_evento_global = n_evento
                    if st.session_state.get("last_n_evento_global") != n_evento:
                        st.session_state.b3_kpi_zip_bytes = None
                        st.session_state.b3_kpi_excel_bytes = None
                        st.session_state.b3_plots_zip_bytes = None
                        st.session_state.b4_sim_zip_bytes = None
                        st.session_state.last_n_evento_global = n_evento
                    st.success(f"Evento {n_evento} seleccionado")
                else:
                    st.warning("❌ No hay eventos en este semestre")
                    st.session_state.evento_global = None
            else:
                st.error("❌ Carpeta de eventos no encontrada")
                st.session_state.evento_global = None
        else:
            st.info("← Seleccione semestre primero")
            st.session_state.evento_global = None
    
    # Validación de flujo: avisar si falta evento para bloques 1-5
    if bloque_trabajo != "modelo_base" and not st.session_state.evento_global:
        st.warning("⚠️ **Atención:** Para acceder a los bloques 1 al 5, primero debe seleccionar un evento arriba.")

    st.markdown("---")

    # ─── CONFIGURACIÓN DE RUTAS Y PARÁMETROS (Agrupados) ────────────────── # type: ignore
    with st.expander("🛠️ Rutas y Parámetros del Proyecto"):
        PF_BASE = st.text_input(
            "PowerFactory — directorio base",
            value=_cfg["PF_BASE"],
            key="cfg_PF_BASE",
        )
        LOC_NAMES_GEN_PATH = st.text_input(
            "loc_names_gen.xlsx",
            value=_cfg["LOC_NAMES_GEN_PATH"],
            key="cfg_LOC_NAMES_GEN_PATH",
        )
        LOC_CAR_PATH = st.text_input(
            "loc_name_cargas.xlsx",
            value=_cfg["LOC_CAR_PATH"],
            key="cfg_LOC_CAR_PATH",
        )
        LOC_XFO_PATH = st.text_input(
            "loc_names_xfo.xlsx",
            value=_cfg["LOC_XFO_PATH"],
            key="cfg_LOC_XFO_PATH",
        )
        PF_PROYECTO = st.text_input("Proyecto PowerFactory", value=_cfg["PF_PROYECTO"], key="cfg_PF_PROYECTO")
        CASO_BASE = st.text_input("Caso base", value=_cfg["CASO_BASE"], key="cfg_CASO_BASE")
        st.caption("Ajustes de flujo")
        EXCLUIR_SLACK = st.text_input("Generadores excluidos de slack", value=_cfg["EXCLUIR_SLACK"], key="cfg_EXCLUIR_SLACK")
        XFO_PF = st.number_input("Factor XFO_PF", value=float(_cfg["XFO_PF"]), key="cfg_XFO_PF")

    st.markdown("---")
    if st.button("Guardar configuración", help="Guarda las rutas actuales."):
        _guardar_config({
            "RAIZ":               RAIZ,
            "RAIZ_DATOS":         RAIZ_DATOS,
            "PF_BASE":            PF_BASE,
            "LOC_NAMES_GEN_PATH": LOC_NAMES_GEN_PATH,
            "LOC_CAR_PATH":       LOC_CAR_PATH,
            "LOC_XFO_PATH":       LOC_XFO_PATH,
            "PF_PROYECTO":        PF_PROYECTO,
            "CASO_BASE":          CASO_BASE,
            "EXCLUIR_SLACK":      EXCLUIR_SLACK,
            "XFO_PF":             XFO_PF,
        })
        st.success("Configuración guardada.")

# ─────────────────────────────────────────────────────────────────────────────
# TÍTULO PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────
st.title("Interfaz de Análisis RPF")

# ─── PANEL DE ESTADO DEL PROYECTO (Encabezado Dinámico) ───────────────────
if st.session_state.evento_global:
    _ev_p = st.session_state.ev_path_global
    _n_ev = st.session_state.n_evento_global
    
    # Verificación rápida de hitos
    has_sim = any(glob.glob(os.path.join(_ev_p, "datos_simulacion_*_2daopcion.xlsx"))) # type: ignore
    has_ci  = any(glob.glob(os.path.join(_ev_p, "condiciones_iniciales_*.xlsx"))) # type: ignore
    has_pf  = any(glob.glob(os.path.join(_ev_p, f"datos_cargados_Ev{_n_ev}*.xlsx"))) # type: ignore
    
    st.markdown(f"### 📍 Evento: `{st.session_state.evento_global}`")

    def _status_card(col, label, ready): # type: ignore
        icon = "OK" if ready else "Pendiente"
        color = "#28a745" if ready else "#6c757d"
        col.markdown(
            f"""<div style='border: 1px solid #eee; border-radius: 4px; padding: 6px; text-align: center; background-color: #fdfdfd; line-height: 1.1;'>
            <div style='font-size: 16px; margin-bottom: 2px;'>{icon}</div>
            <div style='font-size: 11px; color: {color}; font-weight: 700; text-transform: uppercase;'>{label}</div>
            </div>""",
            unsafe_allow_html=True
        )

    st_cols = st.columns(3)
    _status_card(st_cols[0], "1. Extracción CNDC", has_sim) # type: ignore
    _status_card(st_cols[1], "2. Condiciones Iniciales", has_ci) # type: ignore
    _status_card(st_cols[2], "3. Carga en PowerFactory", has_pf) # type: ignore

# ─── BARRA DE UNIDAD GLOBAL (BLOQUES 3, 4, 5) ───────────────────────────── # type: ignore
if bloque_trabajo in ["analisis_datos", "analisis_simulacion", "comparativa_real_simu"]:
    _available_units = get_event_units(st.session_state.ev_path_global, st.session_state.n_evento_global)
    if _available_units:
        st.markdown("### 🔌 Unidad de Análisis")
        _idx_current = 0
        if st.session_state.global_selected_unit in _available_units:
            _idx_current = _available_units.index(st.session_state.global_selected_unit)

        _sel_global = st.segmented_control(
            "Cambio rápido de unidad (Sincronizado):",
            options=_available_units,
            default=_available_units[_idx_current],
            key="global_unit_bar_ctrl"
        )
        if _sel_global:
            st.session_state.global_selected_unit = _sel_global

        if st.session_state.global_selected_unit and st.session_state.ev_path_global:
            if (st.session_state.get("b3_last_unit") != st.session_state.global_selected_unit or
                st.session_state.get("b3_last_event_path") != st.session_state.ev_path_global):
                _sync_session_scale_config(st.session_state.ev_path_global, st.session_state.global_selected_unit)
                st.session_state.b3_last_unit = st.session_state.global_selected_unit
                st.session_state.b3_last_event_path = st.session_state.ev_path_global

        st.divider()

def _df_safe(df):
    """Convierte columnas object con tipos mixtos a str para evitar ArrowTypeError."""
    out = df.copy()
    for col in out.columns:
        if out[col].dtype == object:
            out[col] = out[col].astype(str)
    return out

if bloque_trabajo == "modelo_base":
    st.header("Bloque 0: Obtención de Datos del Modelo")
    st.info(
        "🛠️ Este bloque se encarga de la extracción completa de parámetros técnicos "
        "y topológicos desde PowerFactory. Debe ejecutarse **únicamente** cuando el "
        "modelo base (.pfd) ha sido modificado (adición de barras, cambios de nombres, etc.)."
    )

    with st.expander("📝 Descripción de Tareas", expanded=True):
        st.markdown(f"""
        Los programas en `C:\\Programas Python\\ProgramasLimpio\\Programas_1_uso_modelo` realizan:

        *   **Barras y Líneas:** Extracción de tensiones nominales, longitudes y parámetros de carga.
        *   **Generadores y Cargas:** Mapeo de potencias nominales y conectividad a terminales.
        *   **Escenarios y Variaciones:** Búsqueda recursiva para identificar cambios realizados por cada escenario de operación y variación de red.
        *   **Casos de Estudio:** Indexación de Study Cases configurados en el árbol del proyecto.

        **Objetivo:** Actualizar los archivos Excel de mapeo base para que los simuladores operen con la última versión de la red.
        """)

    st.subheader("⚙️ Configuración de Escaneo")
    c1, c2 = st.columns(2)
    with c1:
        st.text_input("Proyecto PowerFactory", value=PF_PROYECTO, disabled=True)
        st.caption("Configurado en la barra lateral")
    with c2:
        st.text_input("Directorio de Scripts", value=r"...\Programas_1_uso_modelo", disabled=True)

    st.markdown("---")

    # Derivar rutas de salida desde LOC_NAMES_GEN_PATH # type: ignore
    _datos_extraidos_dir = os.path.dirname(os.path.dirname(LOC_NAMES_GEN_PATH))
    _mod_output_path     = os.path.join(_datos_extraidos_dir, "DatosSINdigsilent.xlsx")
    _mod_pf_py           = os.path.join(PF_BASE, "Python", "3.12")
    _mod_runner_path     = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "runners", "DatsoGENBUSLNE_run.py",
    )
    _mod_status_file = os.path.join(_datos_extraidos_dir, "_mod_status.txt")
    _mod_log_file    = os.path.join(_datos_extraidos_dir, "_mod_log.txt")
    _mod_params_path = os.path.join(_datos_extraidos_dir, "_mod_params.json")

    _can_mod = os.path.isfile(_mod_runner_path)
    if not _can_mod:
        st.error(f"No se encontró el runner: `{_mod_runner_path}`")

    if st.button(
        "Iniciar Extracción del Modelo Base",
        type="primary",
        use_container_width=True,
        disabled=not _can_mod or st.session_state.mod_running,
    ):
        # Limpiar archivos previos
        for _f in (_mod_status_file, _mod_log_file):
            if os.path.exists(_f):
                try:
                    os.remove(_f)
                except OSError:
                    pass

        _mod_params = {
            "PF_DIR":      PF_BASE,
            "PF_PY":       _mod_pf_py,
            "PF_PROYECTO": PF_PROYECTO,
            "output_path": _mod_output_path,
        }
        with open(_mod_params_path, "w", encoding="utf-8") as _fp:
            json.dump(_mod_params, _fp, ensure_ascii=False, indent=2)

        st.session_state.mod_running     = True
        st.session_state.mod_status_file = _mod_status_file
        st.session_state.mod_log_file    = _mod_log_file
        st.session_state.mod_return_code = None
        st.session_state.mod_saved_log   = None

        def _mod_thread_fn(runner, params_path, env_vars, status_file, log_file):
            rc = -1
            try:
                proc = subprocess.Popen(
                    [sys.executable, runner, params_path],
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    env=env_vars,
                )
                with open(log_file, "w", encoding="utf-8") as _lf:
                    for _line in proc.stdout:
                        _lf.write(_line)
                        _lf.flush()
                rc = proc.wait()
            except Exception as _exc:
                try:
                    with open(log_file, "a", encoding="utf-8") as _lf:
                        _lf.write(f"\n[ERROR] {_exc}\n")
                except OSError:
                    pass
            finally:
                with open(status_file, "w", encoding="utf-8") as _sf:
                    _sf.write(str(rc))

        _mod_env = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"}
        threading.Thread(
            target=_mod_thread_fn,
            args=(_mod_runner_path, _mod_params_path, _mod_env, _mod_status_file, _mod_log_file),
            daemon=True,
        ).start()
        st.rerun()

    def _leer_mod_log(path):
        try:
            return open(path, encoding="utf-8", errors="replace").read()
        except OSError:
            return ""

    if st.session_state.mod_running:
        _sf = st.session_state.get("mod_status_file") or _mod_status_file
        _lf = st.session_state.get("mod_log_file")    or _mod_log_file

        if _sf and os.path.exists(_sf):
            try:
                st.session_state.mod_return_code = int(open(_sf, encoding="utf-8").read().strip())
            except (OSError, ValueError):
                st.session_state.mod_return_code = -1
            st.session_state.mod_running = False
            st.rerun()
        else:
            st.info("⏳ PowerFactory extrayendo datos del modelo (Ejecución en fragmento)...")
            _monitor_process_fragment(_lf, _sf)

    elif st.session_state.mod_return_code is not None:
        _rc = st.session_state.mod_return_code
        if _rc == 0:
            st.success("Extracción completada. `DatosSINdigsilent.xlsx` actualizado.")
        else:
            st.error(f"❌ Error en la extracción (código {_rc}).")

        _saved = st.session_state.get("mod_saved_log")
        _lf    = st.session_state.get("mod_log_file") or _mod_log_file
        _log_content = _leer_mod_log(_saved) if _saved and os.path.isfile(_saved) else _leer_mod_log(_lf)
        if _log_content:
            with st.expander("📋 Log de extracción", expanded=(_rc != 0)):
                st.code(_log_content, language="")
                _col1, _col2 = st.columns(2)
                with _col1:
                    if not (_saved and os.path.isfile(_saved)):
                        if st.button("💾 Guardar log", key="mod_save_log"):
                            import datetime # type: ignore
                            _ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                            _dest = os.path.join(_datos_extraidos_dir, f"log_modelo_{_ts}.txt")
                            try:
                                open(_dest, "w", encoding="utf-8").write(_log_content)
                                st.session_state.mod_saved_log = _dest # type: ignore
                                st.success(f"Guardado en: `{_dest}`")
                            except OSError as _e:
                                st.error(str(_e))
                    else:
                        st.caption(f"💾 Guardado en: `{_saved}`")
                with _col2:
                    st.download_button(
                        "⬇️ Descargar log",
                        data=_log_content.encode("utf-8"),
                        file_name=os.path.basename(_saved) if _saved else "log_modelo.txt",
                        mime="text/plain",
                    )

    # Mostrar logs de ejecuciones anteriores si no hay una activa
    if st.session_state.mod_return_code is None and not st.session_state.mod_running:
        _prev_mod_logs = sorted(
            glob.glob(os.path.join(_datos_extraidos_dir, "log_modelo_*.txt")),
            reverse=True,
        )
        if _prev_mod_logs:
            with st.expander(f"📋 Logs anteriores ({len(_prev_mod_logs)})"):
                _sel = st.selectbox(
                    "Seleccionar log",
                    _prev_mod_logs,
                    format_func=os.path.basename,
                    key="sel_mod_log",
                )
                if _sel:
                    st.code(_leer_mod_log(_sel), language="")
                    st.download_button(
                        "⬇️ Descargar",
                        data=_leer_mod_log(_sel).encode("utf-8"),
                        file_name=os.path.basename(_sel),
                        mime="text/plain",
                        key="dl_mod_log",
                    )

    # ─────────────────────────────────────────────────────────────────────────
    # HELPER compartido para lanzar/monitorizar scripts de modelo base
    # ─────────────────────────────────────────────────────────────────────────
    def _bloque_script(pfx, runner_name, params_dict, log_name, any_other_running):
        """Lanza runner, muestra log en vivo y resultado final.
        pfx            : prefijo de session_state (gen, lne, xfo, sht, car)
        runner_name    : nombre del archivo .py en runners/
        params_dict    : dict con los parámetros JSON
        log_name       : prefijo para el archivo de log permanente
        any_other_running : True si otro script ya está corriendo
        """
        _runner = os.path.join( # type: ignore
            os.path.dirname(os.path.abspath(__file__)),
            "runners", runner_name,
        )
        _sf  = os.path.join(_datos_extraidos_dir, f"_{pfx}_status.txt")
        _lf  = os.path.join(_datos_extraidos_dir, f"_{pfx}_log.txt")
        _pf  = os.path.join(_datos_extraidos_dir, f"_{pfx}_params.json")

        _can = os.path.isfile(_runner)
        if not _can:
            st.error(f"Runner no encontrado: `{runner_name}`")

        _is_running = st.session_state.get(f"{pfx}_running", False)
        _rc_prev    = st.session_state.get(f"{pfx}_return_code") # type: ignore

        # Icono de estado en el título del botón de ejecución
        _lbl_icon = "⏳" if _is_running else ("✅" if _rc_prev == 0 else ("❌" if _rc_prev is not None else "▶️"))

        if st.button(
            f"{_lbl_icon} Ejecutar",
            key=f"btn_{pfx}",
            type="primary",
            disabled=not _can or _is_running or any_other_running,
        ):
            for _old in (_sf, _lf):
                if os.path.exists(_old):
                    try: os.remove(_old)
                    except OSError: pass
            with open(_pf, "w", encoding="utf-8") as _fp:
                json.dump(params_dict, _fp, ensure_ascii=False, indent=2)

            st.session_state[f"{pfx}_running"]     = True
            st.session_state[f"{pfx}_status_file"] = _sf
            st.session_state[f"{pfx}_log_file"]    = _lf
            st.session_state[f"{pfx}_return_code"] = None
            st.session_state[f"{pfx}_saved_log"]   = None

            def _thread(runner, params_path, status_file, log_file):
                rc = -1
                try:
                    proc = subprocess.Popen(
                        [sys.executable, runner, params_path],
                        stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                        text=True, encoding="utf-8", errors="replace",
                        env={**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"},
                    )
                    with open(log_file, "w", encoding="utf-8") as _lf2:
                        for line in proc.stdout:
                            _lf2.write(line); _lf2.flush()
                    rc = proc.wait()
                except Exception as exc:
                    try:
                        with open(log_file, "a", encoding="utf-8") as _lf2:
                            _lf2.write(f"\n[ERROR] {exc}\n")
                    except OSError: pass
                finally:
                    with open(status_file, "w", encoding="utf-8") as _sf2:
                        _sf2.write(str(rc))

            threading.Thread(target=_thread, args=(_runner, _pf, _sf, _lf), daemon=True).start()
            st.rerun()

        def _read(path):
            try: return open(path, encoding="utf-8", errors="replace").read()
            except OSError: return ""

        if _is_running:
            _sf_cur = st.session_state.get(f"{pfx}_status_file") or _sf
            _lf_cur = st.session_state.get(f"{pfx}_log_file")    or _lf
            if _sf_cur and os.path.exists(_sf_cur):
                try: st.session_state[f"{pfx}_return_code"] = int(open(_sf_cur).read().strip())
                except (OSError, ValueError): st.session_state[f"{pfx}_return_code"] = -1
                st.session_state[f"{pfx}_running"] = False
                st.rerun()
            else:
                st.info(f"⏳ Ejecutando {log_name}...")
                _monitor_process_fragment(_lf_cur, _sf_cur)

        elif _rc_prev is not None:
            if _rc_prev == 0:
                st.success("✅ Completado correctamente.")
            else:
                st.error(f"❌ Error (código {_rc_prev}).")
            _saved  = st.session_state.get(f"{pfx}_saved_log")
            _lf_cur = st.session_state.get(f"{pfx}_log_file") or _lf
            _txt    = _read(_saved) if _saved and os.path.isfile(_saved) else _read(_lf_cur)
            if _txt:
                with st.expander("📋 Log", expanded=(_rc_prev != 0)):
                    st.code(_txt, language="")
                    _c1, _c2 = st.columns(2)
                    with _c1:
                        if not (_saved and os.path.isfile(_saved)):
                            if st.button("💾 Guardar log", key=f"save_{pfx}_log"):
                                import datetime as _dt # type: ignore
                                _dest = os.path.join(_datos_extraidos_dir,
                                    f"log_{log_name}_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
                                try:
                                    open(_dest, "w", encoding="utf-8").write(_txt)
                                    st.session_state[f"{pfx}_saved_log"] = _dest # type: ignore
                                    st.success(f"Guardado en: `{_dest}`")
                                except OSError as _e:
                                    st.error(str(_e))
                        else:
                            st.caption(f"💾 Guardado en: `{_saved}`")
                    with _c2:
                        st.download_button("⬇️ Descargar log", data=_txt.encode(),
                                           file_name=os.path.basename(_saved) if _saved else f"log_{log_name}.txt",
                                           mime="text/plain", key=f"dl_{pfx}_log")

    # ─────────────────────────────────────────────────────────────────────────
    # SECCION: scripts adicionales del modelo base
    # ─────────────────────────────────────────────────────────────────────────
    _loc_designacion = os.path.join(_datos_extraidos_dir, "Designacion de loc_name")
    _any_mod_running = st.session_state.mod_running or any(
        st.session_state.get(f"{_p}_running") for _p in ("gen","lne","xfo","sht","car"))

    st.markdown("---") # type: ignore
    st.subheader("🗂️ Scripts de Mapeo y Catalogo")
    st.caption("Ejecutar en orden después de actualizar `DatosSINdigsilent.xlsx`.")

    # ── 2. loc_namesGEN ───────────────────────────────────────────────────────
    with st.expander("2️⃣ Mapeo Generadores CNDC → PF  (`loc_namesGEN`)", expanded=False):
        st.caption("Genera `loc_names_gen.xlsx` con el mapeo de unidades CNDC a loc_names de PowerFactory.")
        _gen_datos_sin = st.text_input(
            "Datos_SIN (*.xls)",
            value=r"C:\Datos del CNDC\Datos_SIN_20251210.xls",
            key="gen_datos_sin",
        )
        # Buscar automáticamente el archivo de referencia más reciente en RAIZ
        _sim_ref_default = ""
        try:
            import glob as _glob # type: ignore
            _candidates = sorted(
                _glob.glob(os.path.join(RAIZ, "**", "datos_simulacion_*.xlsx"), recursive=True),
                key=os.path.getmtime, reverse=True,
            )
            if _candidates:
                _sim_ref_default = _candidates[0]
        except Exception:
            pass
        _gen_sim_ref = st.text_input(
            "Archivo simulación de referencia (datos_simulacion_*.xlsx)",
            value=_sim_ref_default,
            key="gen_sim_ref",
            help="Cualquier evento del semestre activo — los nombres de generadores son consistentes.",
        )
        if _gen_sim_ref and not os.path.isfile(_gen_sim_ref): # type: ignore
            st.warning(f"⚠️ Archivo no encontrado: `{_gen_sim_ref}`")
        _bloque_script(
            pfx="gen",
            runner_name="loc_namesGEN_run.py",
            params_dict={
                "DATOS_PF":       _mod_output_path,
                "DATOS_SIN_PATH": _gen_datos_sin,
                "OUTPUT_DIR":     _loc_designacion,
                "SIM_REF_PATH":   _gen_sim_ref,
            },
            log_name="loc_namesGEN",
            any_other_running=_any_mod_running,
        )

    # ── 3. loc_namesLineas ────────────────────────────────────────────────────
    with st.expander("3️⃣ Catálogo de Líneas PF  (`loc_namesLineas`)", expanded=False):
        st.caption("Genera `loc_names_lineas.xlsx` con loc_names y nombre descriptivo de cada línea.")
        _bloque_script(
            pfx="lne",
            runner_name="loc_namesLineas_run.py",
            params_dict={
                "DATOS_PF":   _mod_output_path,
                "OUTPUT_DIR": _loc_designacion,
            },
            log_name="loc_namesLineas",
            any_other_running=_any_mod_running,
        )

    # ── 4. loc_names_xfo ─────────────────────────────────────────────────────
    with st.expander("4️⃣ Loc-names de Transformadores  (`loc_names_xfo`)", expanded=False):
        st.caption("Genera `loc_names_xfo.xlsx` con barras HV/LV de cada transformador.")
        _xfo_topologia = st.text_input(
            "Topología completa PF (topologia_completa_pf.xlsx)",
            value=os.path.join(_datos_extraidos_dir, "Topologia", "topologia_completa_pf.xlsx"),
            key="xfo_topologia",
            help="Opcional — si no existe se infieren las barras desde el nombre del transformador.",
        )
        _bloque_script(
            pfx="xfo",
            runner_name="loc_names_xfo_run.py",
            params_dict={
                "DATOS_PF":     _mod_output_path,
                "TOPOLOGIA_PF": _xfo_topologia,
                "OUTPUT_DIR":   _loc_designacion,
            },
            log_name="loc_names_xfo",
            any_other_running=_any_mod_running,
        )

    # ── 5. InventarioShunts ───────────────────────────────────────────────────
    with st.expander("5️⃣ Inventario de Shunts y Compensadores  (`InventarioShunts_PF`)", expanded=False):
        st.caption("Conecta a PowerFactory y genera el inventario de shunts del caso base.")
        _sht_tap = st.checkbox(
            "Activar 'Tap Adjustment of Shunts' en el Load Flow",
            value=False,
            key="sht_tap",
        )
        _bloque_script(
            pfx="sht",
            runner_name="InventarioShunts_PF_run.py",
            params_dict={
                "PF_BASE":     PF_BASE,
                "PF_PROYECTO": PF_PROYECTO,
                "CASO_BASE":   CASO_BASE,
                "OUTPUT_DIR":  _loc_designacion,
                "tap_shunts":  _sht_tap,
            },
            log_name="InventarioShunts",
            any_other_running=_any_mod_running,
        )

    # ── 6. MapeoRetirosSTI ────────────────────────────────────────────────────
    with st.expander("6️⃣ Mapeo de Cargas → Distribuidores  (`MapeoRetirosSTI_v6`)", expanded=False):
        st.caption("Genera `loc_name_cargas.xlsx` mapeando cargas PF a distribuidores/CNR usando el instructivo CNDC.")
        _car_deener = st.text_input(
            "Archivo deener_*.xlsx",
            value="",
            key="car_deener",
            help="Ruta completa al archivo de demanda de energía del evento de referencia.",
        )
        _car_postot = st.text_input(
            "Archivo postot_*.xlsx (opcional)",
            value="",
            key="car_postot",
            help="Dejar vacío si no está disponible.",
        )
        _car_hora = st.text_input(
            "Hora del evento (HH:MM)",
            value="18:45",
            key="car_hora",
        )
        _bloque_script(
            pfx="car",
            runner_name="MapeoRetirosSTI_run.py",
            params_dict={
                "DATOS_PATH":        _mod_output_path,
                "LOC_NAMES_XFO":     os.path.join(_loc_designacion, "loc_names_xfo.xlsx"),
                "DEENER_PATH":       _car_deener,
                "POSTOT_PATH":       _car_postot,
                "OUTPUT_DIR":        _loc_designacion,
                "HORA_EVENTO_LABEL": _car_hora,
            },
            log_name="MapeoRetirosSTI",
            any_other_running=_any_mod_running,
        )

elif bloque_trabajo == "config_unidades":
    st.header("Bloque 2: Configuración y Observación de Unidades")
    st.info("Visualice y ajuste los parámetros técnicos de las unidades detectadas en el evento.")
    
    if not st.session_state.evento_global:
        st.warning("👈 Seleccione un evento en la barra lateral.")
        st.stop()

    _units_cfg = get_event_units(st.session_state.ev_path_global, st.session_state.n_evento_global)
    if not _units_cfg: # type: ignore
        st.info("No se detectaron unidades procesadas para este evento.")
    else:
        _pmax_map = _load_pmax_cargado(st.session_state.ev_path_global, st.session_state.n_evento_global)
        _tmap = _load_tech_map(LOC_NAMES_GEN_PATH)
        
        st.subheader("📋 Inventario de Unidades del Evento")
        
        _cfg_rows = []
        for _u in _units_cfg:
            _pm_v, _tk, _src = _get_pmax_from_cargado(_u, _pmax_map, _tmap)
            # Solo incluir si resolvimos un ID de PowerFactory real (evita páginas de resumen)
            if "sym_" in _tk.lower() or _src == "datos_cargados" or _src == "loc_names_gen":
                _rp_v = _get_rp_default(_tk, LOC_NAMES_GEN_PATH)
                _cfg_rows.append({
                    "Unidad": _u,
                    "ID PowerFactory": _tk,
                    "P_max [MW]": _pm_v,
                    "Estatismo (Rp) [%]": _rp_v,
                    "Fuente Pmax": _src or "Default"
                })
        
        st.dataframe(_df_safe(pd.DataFrame(_cfg_rows)), use_container_width=True, hide_index=True)
        
        st.markdown("---") # type: ignore
        st.subheader("📥 Importar / Exportar Configuración")
        ci1, ci2 = st.columns(2)
        with ci1:
            st.markdown("**Cargar parámetros:**")
            # Opción 1: Subida de archivo (Recomendado para entornos web/remotos)
            _up_csv = st.file_uploader("Subir archivo CSV:", type=["csv"], key="config_uploader")
            
            # Opción 2: Ruta directa (Útil para ejecución local rápida)
            _csv_path_input = st.text_input("O ingresar ruta absoluta del archivo:", value=r"C:\Users\jose.lozano\Downloads\2026-05-07T00-20_export.csv")
            
            if st.button(" Procesar e Importar"):
                _source_df = None
                if _up_csv: _source_df = pd.read_csv(_up_csv)
                elif os.path.isfile(_csv_path_input): _source_df = pd.read_csv(_csv_path_input)
                
                if _source_df is not None:
                    if "ID PowerFactory" in _source_df.columns and "Estatismo (Rp) [%]" in _source_df.columns:
                        for _, row in _source_df.iterrows():
                            _save_rp_cfg(LOC_NAMES_GEN_PATH, str(row["ID PowerFactory"]), float(row["Estatismo (Rp) [%]"]))
                        st.success("✅ Parámetros de Estatismo importados correctamente. Los cambios se verán reflejados en los bloques de análisis.")
                        st.rerun() # type: ignore
                    else:
                        st.error("El archivo no tiene el formato correcto. Se requieren las columnas: 'ID PowerFactory' y 'Estatismo (Rp) [%]'.")
                else:
                    st.error("No se ha podido acceder al archivo. Verifique la ruta o suba el archivo manualmente.")

        with ci2:
            st.markdown("**Guardar estado actual:**")
            _export_csv = pd.DataFrame(_cfg_rows).to_csv(index=False).encode('utf-8')
            st.download_button("⬇️ Descargar Configuración Actual (CSV)", _export_csv, 
                               file_name=f"config_rpf_Ev{st.session_state.n_evento_global}.csv", mime="text/csv")

        st.markdown("---") # type: ignore
        st.subheader("✏️ Edición de Parámetros")
        _u_to_edit = st.selectbox("Seleccione unidad para modificar:", _units_cfg)
        
        if _u_to_edit:
            _pm_e, _tk_e, _ = _get_pmax_from_cargado(_u_to_edit, _pmax_map, _tmap)
            st.markdown(f"**Editando:** `{_tk_e}`")
            _widget_pmax_rp(_tk_e, LOC_NAMES_GEN_PATH, key_prefix="cfg_edit")

elif bloque_trabajo == "config_global": # type: ignore
    st.header("🎨 Bloque 7: Configuración Global de Gráficas")
    st.info("Personalice la apariencia de todas las gráficas generadas en el sistema.")
    
    c1, c2, c3 = st.columns(3)
    with c1:
        st.session_state.graph_config["freq_color_real"] = st.color_picker("Frecuencia Real", st.session_state.graph_config["freq_color_real"])
        st.session_state.graph_config["freq_color_sim0"] = st.color_picker("Frecuencia Sim E.0", st.session_state.graph_config["freq_color_sim0"])
        st.session_state.graph_config["freq_color_sim1"] = st.color_picker("Frecuencia Sim E.1", st.session_state.graph_config["freq_color_sim1"]) # type: ignore
    with c2:
        st.session_state.graph_config["pot_color_real"] = st.color_picker("Potencia Real", st.session_state.graph_config["pot_color_real"])
        st.session_state.graph_config["pot_color_sim0"] = st.color_picker("Potencia Sim E.0", st.session_state.graph_config["pot_color_sim0"])
        st.session_state.graph_config["pot_color_sim1"] = st.color_picker("Potencia Sim E.1", st.session_state.graph_config["pot_color_sim1"])
    with c3:
        st.session_state.graph_config["line_width"] = st.slider("Grosor de línea", 1.0, 5.0, float(st.session_state.graph_config["line_width"]), 0.5)
        st.session_state.graph_config["marker_size"] = st.slider("Tamaño de marcadores", 5, 25, int(st.session_state.graph_config["marker_size"]))
        st.session_state.graph_config["show_grid"] = st.checkbox("Mostrar cuadrícula", value=st.session_state.graph_config["show_grid"])
        st.session_state.graph_config["plot_height"] = st.slider("Altura del gráfico (px)", 400, 1000, int(st.session_state.graph_config["plot_height"]), 20)
        st.session_state.graph_config["template"] = st.selectbox("Plantilla de color", ["plotly_white", "plotly", "ggplot2", "seaborn", "simple_white", "none"],
                                                                  index=["plotly_white", "plotly", "ggplot2", "seaborn", "simple_white", "none"].index(st.session_state.graph_config["template"]),
                                                                  help="Plantilla de colores para los gráficos de Plotly.")

    st.caption("🔁 Nota: lo que ajustes en Bloque 7 (config_global) se aplica a todas las gráficas del sistema via st.session_state.graph_config.")
    st.markdown("---")
    st.subheader("Visibilidad de Marcadores CNDC")
    mc1, mc2 = st.columns(2)
    with mc1:
        st.session_state.graph_config["show_initial"] = st.toggle("Mostrar Iniciales (f₀, P₀)", value=st.session_state.graph_config["show_initial"])
        st.session_state.graph_config["show_nadir"] = st.toggle("Mostrar Nadir (f_min)", value=st.session_state.graph_config["show_nadir"])
    with mc2:
        st.session_state.graph_config["show_dt_eval"] = st.toggle("Mostrar t₀+35s (f_Δt, P_Δt)", value=st.session_state.graph_config["show_dt_eval"])
        st.session_state.graph_config["show_deadband"] = st.toggle("Mostrar Banda Muerta (±25mHz)", value=st.session_state.graph_config["show_deadband"])

elif bloque_trabajo == "carga_datos":
    st.subheader("Bloque de Trabajo: Carga de Datos")
    st.caption("Workflow lineal para la preparación y carga del modelo en PowerFactory.")

    tab_ext, tab_cond, tab_pf = st.tabs([
        "1. Extracción CNDC", # type: ignore
        "2. Condiciones Iniciales", # type: ignore
        "3. PowerFactory" # type: ignore
    ])

    semestre = st.session_state.semestre_global
    evento = st.session_state.evento_global
    ev_path = st.session_state.ev_path_global
    n_evento = st.session_state.n_evento_global

    # ═════════════════════════════════════════════════════════════════════════════
    # PASO 1: EXTRACCIÓN DE DATOS CNDC
    # ═════════════════════════════════════════════════════════════════════════════
    with tab_ext: # type: ignore
        st.header("1️⃣ Extracción de Datos CNDC")
        st.info(
            "📥 Este módulo extrae datos de despacho y demanda CNDC, combinando "
            "información de archivos DC, DCDR, DEENER y tabla_resultados para generar "
            "`datos_simulacion_*_2daopcion.xlsx`"
        )

        if semestre and evento:
            st.subheader(f"📍 Evento seleccionado: **{evento}**")
            st.caption(f"Semestre: **{semestre}**")

            st.markdown("---") # type: ignore
            st.subheader("✓ Archivos de entrada requeridos")

            dc_files       = glob.glob(os.path.join(ev_path, "Despacho", "dc_*.xls*"))
            dcdr_files     = glob.glob(os.path.join(ev_path, "Despacho", "dcdr_*.xls*"))
            deener_files   = glob.glob(os.path.join(ev_path, "Demanda de Energia y Potencia", "deener_*.xlsx"))
            tabla_files    = glob.glob(os.path.join(RAIZ, semestre, "Tabla_Eventos_*.xlsx"))
            result_files   = glob.glob(os.path.join(ev_path, "Resultados_*", "tabla_resultados_*.xlsx"))

            todos_ok = True
            status_archivos = []
            for nombre, archivos in [
                ("Despacho/dc_*.xls", dc_files),
                ("Despacho/dcdr_*.xls", dcdr_files),
                ("Demanda/deener_*.xlsx", deener_files),
                ("Tabla_Eventos_*.xlsx (semestre)", tabla_files),
                ("tabla_resultados_*.xlsx (evento)", result_files),
            ]:
                existe = len(archivos) > 0
                todos_ok = todos_ok and existe
                status_archivos.append({
                    "Archivo": nombre,
                    "Estado": "OK" if existe else "Falta",
                    "Cantidad": len(archivos),
                })

            st.dataframe(pd.DataFrame(status_archivos), use_container_width=True, hide_index=True)

            # Vista previa de Entradas: DC y DCDR
            if dc_files or dcdr_files:
                with st.expander("📋 Vista previa: Datos de Despacho (Entradas DC/DCDR)"):
                    col_a, col_b = st.columns(2)

                    with col_a:
                        if dc_files:
                            st.caption(f"Archivo DC: `{os.path.basename(dc_files[0])}`")
                            try:
                                xl_dc = pd.ExcelFile(dc_files[0], engine="calamine")
                                tabs_dc = st.tabs(xl_dc.sheet_names)
                                for i, sheet in enumerate(xl_dc.sheet_names):
                                    with tabs_dc[i]:
                                        st.dataframe(_df_safe(xl_dc.parse(sheet).head(20)), use_container_width=True)
                            except Exception as e:
                                st.error(f"Error al leer DC: {e}")

                    with col_b:
                        if dcdr_files:
                            st.caption(f"Archivo DCDR: `{os.path.basename(dcdr_files[0])}`")
                            try:
                                xl_dcdr = pd.ExcelFile(dcdr_files[0], engine="calamine")
                                tabs_dcdr = st.tabs(xl_dcdr.sheet_names)
                                for i, sheet in enumerate(xl_dcdr.sheet_names):
                                    with tabs_dcdr[i]:
                                        st.dataframe(_df_safe(xl_dcdr.parse(sheet).head(20)), use_container_width=True)
                            except Exception as e:
                                st.error(f"Error al leer DCDR: {e}")

            # Verificar si el archivo de salida ya existe y mostrar todas sus hojas
            output_sim = glob.glob(os.path.join(ev_path, "datos_simulacion_*_2daopcion.xlsx"))
            if output_sim:
                st.success(f"📂 **Archivo de salida detectado:** `{os.path.basename(output_sim[0])}`")
                with st.expander("📊 Vista previa completa del resultado (Todas las hojas)"):
                    try:
                        xl_out = pd.ExcelFile(output_sim[0], engine="calamine")
                        nombres_hojas = xl_out.sheet_names
                        tabs_out = st.tabs(nombres_hojas)

                        for i, nombre_hoja in enumerate(nombres_hojas):
                            with tabs_out[i]:
                                df_sheet = xl_out.parse(nombre_hoja)
                                st.caption(f"Mostrando primeras 100 filas de '{nombre_hoja}'")
                                st.dataframe(_df_safe(df_sheet.head(100)), use_container_width=True)
                    except Exception as e:
                        st.error(f"No se pudo leer la salida: {e}")
            else:
                st.info("ℹ️ No se detectó archivo de salida previo para este evento.")
            st.markdown("---") # type: ignore

            if todos_ok:
                _ext_runner = os.path.join(
                    os.path.dirname(os.path.abspath(__file__)),
                    "runners", "ExtFLujos2daO_run.py",
                )
                _can_ext = os.path.isfile(_ext_runner)
                if not _can_ext:
                    st.error(f"No se encontró el runner: `{_ext_runner}`")

                col_ext_btn, _ = st.columns([1, 3])
                with col_ext_btn:
                    ext_btn = st.button(
                        "Ejecutar Extracción de Datos",
                        type="primary",
                        use_container_width=True,
                        disabled=not _can_ext or st.session_state.ext_running,
                    )

                _ext_status_file = st.session_state.get("ext_status_file") or os.path.join(ev_path, "_ext_status.txt")

                if ext_btn and _can_ext:
                    _ext_status_f = os.path.join(ev_path, "_ext_status.txt")
                    if os.path.exists(_ext_status_f):
                        try:
                            os.remove(_ext_status_f)
                        except OSError:
                            pass

                    _ext_params = {
                        "semestre":           semestre,
                        "evento":             evento,
                        "RAIZ":               RAIZ,
                        "LOC_NAMES_GEN_PATH": LOC_NAMES_GEN_PATH,
                    }
                    _ext_params_path = os.path.join(ev_path, "_ext_params.json")
                    with open(_ext_params_path, "w", encoding="utf-8") as _fp:
                        json.dump(_ext_params, _fp, ensure_ascii=False, indent=2)

                    _ext_log_f = os.path.join(ev_path, "_ext_log.txt")
                    st.session_state.ext_running    = True
                    st.session_state.ext_status_file = _ext_status_f
                    st.session_state.ext_log_file    = _ext_log_f
                    st.session_state.ext_return_code = None
                    st.session_state.ext_saved_log   = None

                    def _ext_thread_fn(runner, params_path, env_vars, status_file, log_file):
                        rc = -1
                        try:
                            proc = subprocess.Popen(
                                [sys.executable, runner, params_path],
                                stdout=subprocess.PIPE,
                                stderr=subprocess.STDOUT,
                                text=True,
                                encoding="utf-8",
                                errors="replace",
                                env=env_vars,
                            )
                            with open(log_file, "w", encoding="utf-8") as _lf:
                                for _line in proc.stdout:
                                    _lf.write(_line)
                                    _lf.flush()
                            rc = proc.wait()
                        except Exception as _exc:
                            try:
                                with open(log_file, "a", encoding="utf-8") as _lf:
                                    _lf.write(f"\n[ERROR] {_exc}\n")
                            except OSError:
                                pass
                        finally:
                            with open(status_file, "w", encoding="utf-8") as _sf:
                                _sf.write(str(rc))

                    _ext_env = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"}
                    threading.Thread(
                        target=_ext_thread_fn,
                        args=(_ext_runner, _ext_params_path, _ext_env, _ext_status_f, _ext_log_f),
                        daemon=True,
                    ).start()
                    st.rerun()

                _ext_log_live = st.session_state.get("ext_log_file") or os.path.join(ev_path, "_ext_log.txt")

                def _leer_ext_log(path):
                    try:
                        return open(path, encoding="utf-8", errors="replace").read()
                    except OSError:
                        return ""

                def _guardar_ext_log(log_path, ev_path_, n_ev_):
                    contenido = _leer_ext_log(log_path)
                    if not contenido:
                        return None
                    import datetime
                    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    dest = os.path.join(ev_path_, f"log_EXT_Ev{n_ev_}_{ts}.txt")
                    try:
                        with open(dest, "w", encoding="utf-8") as _f:
                            _f.write(contenido)
                        return dest
                    except OSError:
                        return None

                if st.session_state.ext_running:
                    if _ext_status_file and os.path.exists(_ext_status_file):
                        try:
                            _ext_rc = int(open(_ext_status_file, encoding="utf-8").read().strip())
                        except (OSError, ValueError):
                            _ext_rc = -1
                        st.session_state.ext_return_code = _ext_rc
                        st.session_state.ext_running = False
                        st.rerun()
                    else:
                        st.info("⏳ Extracción CNDC en curso...")
                        _monitor_process_fragment(_ext_log_live, _ext_status_file)
                elif st.session_state.ext_return_code is not None:
                    _ext_rc = st.session_state.ext_return_code
                    if _ext_rc == 0:
                        st.success("✅ Extracción completada. Cambie a la pestaña **2. Condiciones Iniciales**.")
                    else:
                        st.error(f"❌ Error en extracción (código {_ext_rc}).")
                    _ext_saved = st.session_state.get("ext_saved_log")
                    _ext_log_content = _leer_ext_log(_ext_saved) if _ext_saved and os.path.isfile(_ext_saved) else _leer_ext_log(_ext_log_live)
                    if _ext_log_content:
                        with st.expander("📋 Log de ejecución Extracción", expanded=(_ext_rc != 0)):
                            st.code(_ext_log_content, language="")
                            _el1, _el2 = st.columns(2)
                            with _el1:
                                if not (_ext_saved and os.path.isfile(_ext_saved)):
                                    if st.button("💾 Guardar log", key="ext_save_log"):
                                        _dest = _guardar_ext_log(_ext_log_live, ev_path, n_evento) # type: ignore
                                        if _dest:
                                            st.session_state.ext_saved_log = _dest
                                            st.success(f"Guardado en: `{_dest}`")
                                        else:
                                            st.error("No se pudo guardar el log.")
                                else: # type: ignore
                                    st.caption(f"💾 Guardado en: `{_ext_saved}`")
                            with _el2:
                                st.download_button(
                                    "⬇️ Descargar log",
                                    data=_ext_log_content.encode("utf-8"),
                                    file_name=os.path.basename(_ext_saved) if _ext_saved else f"log_EXT_Ev{n_evento}.txt",
                                    mime="text/plain",
                                    key="ext_dl_log",
                                )
            else:
                st.warning("⚠️ Faltan archivos de entrada. Verifique la estructura de carpetas del evento.")
        else:
            st.warning("👈 Seleccione Semestre y Evento en la barra lateral.")

    # ═════════════════════════════════════════════════════════════════════════════
    # PASO 2: GENERACIÓN DE CONDICIONES INICIALES
    # ═════════════════════════════════════════════════════════════════════════════
    with tab_cond: # type: ignore
        st.header("2️⃣ Generación de Condiciones Iniciales")
        st.info(
            "📝 Este módulo genera condiciones iniciales (pgini para generadores y "
            "plini para cargas) desde archivos de datos de simulación."
        )

        if semestre and evento:
            st.subheader(f"📍 Evento seleccionado: **{evento}**")
            st.caption(f"Semestre: **{semestre}**")

            st.markdown("---") # type: ignore
            st.subheader("✓ Archivos de entrada requeridos")

            sim_files  = glob.glob(os.path.join(ev_path, "datos_simulacion_*_2daopcion.xlsx"))
            loc_gen_ok = os.path.isfile(LOC_NAMES_GEN_PATH)
            loc_car_ok = os.path.isfile(LOC_CAR_PATH)
            loc_xfo_ok = os.path.isfile(LOC_XFO_PATH)

            archivos_requeridos = [
                {
                    "Archivo": "datos_simulacion_*_2daopcion.xlsx",
                    "Estado": "OK" if sim_files else "Falta",
                    "Descripción": "Genera módulo Extracción",
                },
                {
                    "Archivo": os.path.basename(LOC_NAMES_GEN_PATH),
                    "Estado": "OK" if loc_gen_ok else "Falta",
                    "Descripción": "Mapeo generadores ↔ loc_names PF",
                },
                {
                    "Archivo": os.path.basename(LOC_CAR_PATH),
                    "Estado": "OK" if loc_car_ok else "Falta",
                    "Descripción": "Mapeo cargas ↔ loc_names PF",
                },
                {
                    "Archivo": os.path.basename(LOC_XFO_PATH),
                    "Estado": "OK" if loc_xfo_ok else "Falta",
                    "Descripción": "Parámetros transformadores",
                },
            ]

            st.dataframe(pd.DataFrame(archivos_requeridos), use_container_width=True, hide_index=True)

            # Vista previa de Entrada: Datos de simulación
            if sim_files:
                with st.expander("📋 Vista previa: Datos de Simulación (Entrada)"):
                    try:
                        df_sim_in = pd.read_excel(sim_files[0], engine="calamine")
                        st.dataframe(_df_safe(df_sim_in.head(20)), use_container_width=True)
                    except Exception as e:
                        st.error(f"Error al leer entrada de simulación: {e}")

            # Verificar si el archivo de salida ya existe
            output_ci = glob.glob(os.path.join(ev_path, "condiciones_iniciales_*.xlsx"))
            if output_ci:
                st.success(f"📂 **Archivo de salida detectado:** `{os.path.basename(output_ci[0])}`")
                with st.expander("📊 Vista previa: Condiciones Iniciales (Salida)"):
                    try:
                        xl_ci = pd.ExcelFile(output_ci[0], engine="calamine")
                        df_pg = xl_ci.parse("pgini_GEN")
                        df_pl = xl_ci.parse("plini_CAR")

                        c1, c2 = st.columns(2)
                        with c1:
                            st.caption("Generadores (pgini)")
                            st.dataframe(df_pg.head(15), use_container_width=True)
                        with c2:
                            st.caption("Cargas (plini)")
                            st.dataframe(df_pl.head(15), use_container_width=True)
                    except Exception as e:
                        st.info("El archivo existe pero no se pudieron leer las pestañas pgini/plini.")
                st.markdown("---")
            else:
                st.info("No se detectaron condiciones iniciales generadas para este evento.")
            st.markdown("---")

            todos_ok = len(sim_files) > 0 and loc_gen_ok and loc_car_ok and loc_xfo_ok

            if todos_ok:
                st.subheader("⚙️ Opciones de generación")

                col_o1, col_o2 = st.columns(2)
                with col_o1:
                    generar_balance = st.checkbox(
                        "Incluir hoja Balance_Plini",
                        value=True,
                        help="Diagnóstico de residuos de redondeo por distribuidor"
                    )
                with col_o2:
                    precision_decimales = st.number_input(
                        "Precisión (decimales)",
                        value=4,
                        min_value=2,
                        max_value=6,
                        help="Precisión para plini (potencia de cargas)"
                    )

                _ci_runner = os.path.join(
                    os.path.dirname(os.path.abspath(__file__)),
                    "CondInicialesPF_run.py",
                )
                _can_ci = os.path.isfile(_ci_runner)
                if not _can_ci:
                    st.error(f"No se encontró el runner: `{_ci_runner}`")

                col_ci_btn, _ = st.columns([1, 3])
                with col_ci_btn:
                    ci_btn = st.button(
                        "Generar Condiciones Iniciales",
                        type="primary",
                        use_container_width=True,
                        disabled=not _can_ci or st.session_state.ci_running,
                    )

                _ci_status_file = st.session_state.get("ci_status_file") or os.path.join(ev_path, "_ci_status.txt")

                if ci_btn and _can_ci:
                    _ci_status_f = os.path.join(ev_path, "_ci_status.txt")
                    if os.path.exists(_ci_status_f):
                        try:
                            os.remove(_ci_status_f)
                        except OSError:
                            pass

                    _ci_params = {
                        "semestre":     semestre,
                        "evento":       evento,
                        "RAIZ":         RAIZ,
                        "LOC_GEN_PATH": LOC_NAMES_GEN_PATH,
                        "LOC_CAR_PATH": LOC_CAR_PATH,
                        "LOC_XFO_PATH": LOC_XFO_PATH,
                    }
                    _ci_params_path = os.path.join(ev_path, "_ci_params.json")
                    with open(_ci_params_path, "w", encoding="utf-8") as _fp:
                        json.dump(_ci_params, _fp, ensure_ascii=False, indent=2)

                    st.session_state.ci_running    = True
                    st.session_state.ci_status_file = _ci_status_f
                    st.session_state.ci_return_code = None
                    st.session_state.ci_saved_log   = None

                    # CI log en vivo: no escribir a disco (evita crear _ci_log.txt)
                    if "ci_log_buffer" not in st.session_state:
                        st.session_state.ci_log_buffer = ""
                    st.session_state.ci_log_buffer = ""

                    def _ci_thread_fn(runner, params_path, env_vars, status_file):
                        rc = -1
                        try:
                            proc = subprocess.Popen(
                                [sys.executable, runner, params_path],
                                stdout=subprocess.PIPE,
                                stderr=subprocess.STDOUT,
                                text=True,
                                encoding="utf-8",
                                errors="replace",
                                env=env_vars,
                            )
                            with open(log_file, "w", encoding="utf-8") as _lf:
                                for _line in proc.stdout:
                                    _lf.write(_line)
                                    _lf.flush()
                            rc = proc.wait()
                        except Exception as exc:
                            try:
                                with open(log_file, "a", encoding="utf-8") as _lf:
                                    _lf.write(f"\n[ERROR] {exc}\n")
                            except OSError:
                                pass
                        finally:
                            with open(status_file, "w", encoding="utf-8") as _sf:
                                _sf.write(str(rc))

                    _ci_env = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"}
                    threading.Thread(
                        target=_ci_thread_fn,
                        args=(_ci_runner, _ci_params_path, _ci_env, _ci_status_f, _ci_log_f),
                        daemon=True,
                    ).start()
                    st.rerun()

                _ci_log_live = st.session_state.get("ci_log_file") or os.path.join(ev_path, "_ci_log.txt")

                def _leer_ci_log(path):
                    try:
                        return open(path, encoding="utf-8", errors="replace").read()
                    except OSError:
                        return ""

                def _guardar_ci_log(log_path, ev_path_, n_ev_):
                    contenido = _leer_ci_log(log_path)
                    if not contenido:
                        return None
                    import datetime
                    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    dest = os.path.join(ev_path_, f"log_CI_Ev{n_ev_}_{ts}.txt")
                    try:
                        with open(dest, "w", encoding="utf-8") as _f:
                            _f.write(contenido)
                        return dest
                    except OSError:
                        return None

                if st.session_state.ci_running:
                    if _ci_status_file and os.path.exists(_ci_status_file):
                        try:
                            _ci_rc = int(open(_ci_status_file, encoding="utf-8").read().strip())
                        except (OSError, ValueError):
                            _ci_rc = -1
                        st.session_state.ci_return_code = _ci_rc
                        st.session_state.ci_running = False
                        st.rerun()
                    else:
                        st.info("⏳ Generación de Condiciones Iniciales en curso...")
                        _monitor_process_fragment(_ci_log_live, _ci_status_file)
                elif st.session_state.ci_return_code is not None:
                    _ci_rc = st.session_state.ci_return_code
                    if _ci_rc == 0:
                        st.success("✅ Condiciones iniciales generadas correctamente.")
                    else:
                        st.error(f"❌ Error en la generación (código {_ci_rc}).")
                    _ci_saved = st.session_state.get("ci_saved_log")
                    _ci_log_content = _leer_ci_log(_ci_saved) if _ci_saved and os.path.isfile(_ci_saved) else _leer_ci_log(_ci_log_live)
                    if _ci_log_content:
                        with st.expander("📋 Log de ejecución Condiciones Iniciales", expanded=(_ci_rc != 0)):
                            st.code(_ci_log_content, language="")
                            _cl1, _cl2 = st.columns(2)
                            with _cl1:
                                if not (_ci_saved and os.path.isfile(_ci_saved)):
                                    if st.button("💾 Guardar log", key="ci_save_log"):
                                        _dest = _guardar_ci_log(_ci_log_live, ev_path, n_evento) # type: ignore
                                        if _dest:
                                            st.session_state.ci_saved_log = _dest
                                            st.success(f"Guardado en: `{_dest}`")
                                        else:
                                            st.error("No se pudo guardar el log.")
                                else: # type: ignore
                                    st.caption(f"💾 Guardado en: `{_ci_saved}`")
                            with _cl2:
                                st.download_button(
                                    "⬇️ Descargar log",
                                    data=_ci_log_content.encode("utf-8"),
                                    file_name=os.path.basename(_ci_saved) if _ci_saved else f"log_CI_Ev{n_evento}.txt",
                                    mime="text/plain",
                                    key="ci_dl_log",
                                )
            else:
                missing = []
                if not sim_files:
                    missing.append("• `datos_simulacion_*_2daopcion.xlsx` → ejecute **módulo Extracción**")
                if not loc_gen_ok:
                    missing.append(f"• `{os.path.basename(LOC_NAMES_GEN_PATH)}`")
                if not loc_car_ok:
                    missing.append(f"• `{os.path.basename(LOC_CAR_PATH)}`")
                if not loc_xfo_ok:
                    missing.append(f"• `{os.path.basename(LOC_XFO_PATH)}`")

                st.warning("⚠️ Faltan archivos:\n" + "\n".join(missing))
        else:
            st.warning("👈 Seleccione Semestre y Evento en la barra lateral.")

    # ═════════════════════════════════════════════════════════════════════════════
    # PASO 3: CARGA EN POWERFACTORY # type: ignore
    # ═════════════════════════════════════════════════════════════════════════════
    with tab_pf:
        st.header("3️⃣ Carga en PowerFactory")

        if semestre and evento:
            st.subheader(f"📍 Evento seleccionado: **{evento}**")
            st.caption(f"Semestre: **{semestre}**")
            st.markdown("---") # type: ignore

            # ─── SECCIÓN 2 — PANEL DE ARCHIVOS DE ENTRADA ───────────────────────────
            st.header("2 · Archivos de Entrada")

            ci_files    = glob.glob(os.path.join(ev_path, "condiciones_iniciales_*.xlsx"))
            dsim_files  = glob.glob(os.path.join(ev_path, "datos_simulacion_*_2daopcion.xlsx"))
            tabla_files = glob.glob(os.path.join(RAIZ, semestre, "Tabla_Eventos_*.xlsx"))
            xfo_ok      = os.path.isfile(LOC_XFO_PATH)

            # Check for Costo Marginal STI files
            costo_marginal_dir = os.path.join(ev_path, CARPETA_COSTO_MARGINAL)
            costo_marginal_files = []
            if os.path.isdir(costo_marginal_dir):
                costo_marginal_files = (glob.glob(os.path.join(costo_marginal_dir, "postot*.xlsx")) or
                                        glob.glob(os.path.join(costo_marginal_dir, "td_*.xlsx")))
            costo_marginal_found = bool(costo_marginal_files)
            def _estado(found):
                return "✅ OK" if found else "❌ Falta"
            
            tabla_archivos = pd.DataFrame([
                {
                    "Estado": _estado(ci_files),
                    "Archivo": os.path.basename(ci_files[0]) if ci_files else "condiciones_iniciales_*.xlsx",
                    "Descripción": "Condiciones iniciales (generadores + cargas)",
                    "Carpeta": os.path.dirname(ci_files[0]) if ci_files else ev_path, # type: ignore
                },
                {
                    "Estado": _estado(dsim_files),
                    "Archivo": os.path.basename(dsim_files[0]) if dsim_files else "datos_simulacion_*_2daopcion.xlsx",
                    "Descripción": "Pdem del evento (celda B8)",
                    "Carpeta": ev_path, # type: ignore
                },
                {
                    "Estado": _estado(tabla_files), # type: ignore
                    "Archivo": os.path.basename(tabla_files[0]) if tabla_files else "Tabla_Eventos_*.xlsx",
                    "Descripción": "p_desc del disparo (potencia desconectada)",
                    "Carpeta": os.path.join(RAIZ, semestre),
                },
                {
                    "Estado": _estado(xfo_ok),
                    "Archivo": os.path.basename(LOC_XFO_PATH),
                    "Descripción": "Capacidad de transformadores (restricción de cargas)", # type: ignore
                    "Carpeta": os.path.dirname(LOC_XFO_PATH),
                },
                {
                    "Estado": _estado(costo_marginal_found),
                    "Archivo": os.path.basename(costo_marginal_files[0]) if costo_marginal_files else "postot*.xlsx o td_*.xlsx",
                    "Descripción": "Costo Marginal STI (retiros CNDC nodal)",
                    "Carpeta": os.path.dirname(LOC_XFO_PATH),
                },
            ])
            st.dataframe(tabla_archivos, use_container_width=True, hide_index=True)

            # ─── SECCIÓN 3 — VISTA PREVIA DE CONDICIONES INICIALES ───────────────
            if ci_files:
                st.header("3 · Vista Previa de Condiciones Iniciales")
                ci_path = ci_files[0]

                try: # type: ignore
                    xl_ci_prev = pd.ExcelFile(ci_path, engine="calamine")
                    df_res_ci  = xl_ci_prev.parse("Resumen")
                    df_pgini   = xl_ci_prev.parse("pgini_GEN")
                    df_plini   = xl_ci_prev.parse("plini_CAR")

                    info_ci = dict(
                        zip(
                            df_res_ci.iloc[:, 0].astype(str).str.strip(),
                            df_res_ci.iloc[:, 1].astype(str).str.strip(),
                        )
                    )

                    pgen_ci = df_pgini["pgini_MW"].sum() if "pgini_MW" in df_pgini.columns else 0.0
                    pdem_ci = df_plini["plini_MW"].sum() if "plini_MW" in df_plini.columns else 0.0

                    k1, k2, k3, k4 = st.columns(4)
                    k1.metric("Generadores", len(df_pgini))
                    k2.metric("Cargas", len(df_plini))
                    k3.metric("Pgen Excel (MW)", f"{pgen_ci:.1f}")
                    k4.metric("Pdem Excel (MW)", f"{pdem_ci:.1f}") # type: ignore

                    ia, ib = st.columns(2)
                    with ia:
                        st.info(f"**Fecha y hora:** {info_ci.get('Fecha y hora', '—')}")
                        st.info(f"**Disparo:** {info_ci.get('Disparo', '—')}")
                    with ib:
                        st.info(f"**Hora generación:** {info_ci.get('Hora evento (gen)', '—')}")
                        st.info(f"**Hora cargas:** {info_ci.get('Hora Po (cargas)', '—')}")

                    with st.expander("📋 Generadores (pgini_GEN) — primeras 30 filas"):
                        st.dataframe(df_pgini.head(30), use_container_width=True, hide_index=True)

                    with st.expander("📋 Cargas (plini_CAR) — primeras 30 filas"):
                        st.dataframe(df_plini.head(30), use_container_width=True, hide_index=True)

                except Exception as _e:
                    st.error(f"Error al leer condiciones_iniciales: {_e}")
                    df_pgini = None
                    info_ci  = {}
            else: # type: ignore
                st.warning("No se encontró `condiciones_iniciales_*.xlsx`. Primero genere las condiciones iniciales.")
                df_pgini = None
                info_ci  = {}
                st.info("💡 Use el **módulo 2** para generar las condiciones iniciales.")

            # ─── SECCIÓN 4 — OPCIONES DE EJECUCIÓN ───────────────────────────────
            if ci_files:
                st.header("4 · Opciones de Ejecución")

                col_opt_a, col_opt_b = st.columns(2)

                with col_opt_a:
                    st.subheader("Potencia del disparo")

                    p_desc_ui = 0.0 # type: ignore
                    if tabla_files:
                        try:
                            import openpyxl as _opx
                            _wb = _opx.load_workbook(tabla_files[0], data_only=True)
                            _sh = _wb.active
                            for _fila in _sh.iter_rows(min_row=3, values_only=True):
                                if _fila[0] is None:
                                    continue
                                try:
                                    if int(_fila[0]) == int(n_evento):
                                        p_desc_ui = float(_fila[3]) if _fila[3] else 0.0
                                        break
                                except (ValueError, TypeError):
                                    pass
                        except Exception:
                            pass

                    if p_desc_ui > 0:
                        st.metric("p_desc registrado (MW)", f"{p_desc_ui:.2f}")
                    else:
                        st.caption("p_desc no encontrado en Tabla_Eventos")

                    # ── Identificar unidades del disparo desde el Excel CI ─────
                    _disp_units = [] # type: ignore
                    if df_pgini is not None and info_ci:
                        _disparo_str = info_ci.get("Disparo", "")
                        _disp_str_clean = re.sub(r"\by\b", ",", _disparo_str, flags=re.IGNORECASE)
                        _sti_disp = {x.strip() for x in _disp_str_clean.split(",") if x.strip() and x.strip() != "nan"}

                        def _sti_de_ui(loc_name):
                            s = re.sub(r"\(\d+\)$", "", str(loc_name).strip())
                            for _pref in ("sym_", "WT_", "PV-", "PV_", "sta_"):
                                if s.lower().startswith(_pref.lower()):
                                    s = s[len(_pref):]
                                    break
                            s = re.sub(r"_EQ$", "", s, flags=re.IGNORECASE)
                            s = re.sub(r"_II$", "", s, flags=re.IGNORECASE)
                            s = re.sub(r"^LOD_", "", s, flags=re.IGNORECASE)
                            return s.strip()

                        for _, _row in df_pgini.iterrows():
                            _loc = str(_row.get("loc_name PF", "")).strip()
                            if _sti_de_ui(_loc) in _sti_disp:
                                _disp_units.append({"loc": _loc, "pgini_actual": float(_row.get("pgini_MW", 0.0))})

                    def _preview_prop_pdesc(units, p_desc): # type: ignore
                        """Vista previa opción 3: proporcional a pgini actual → suma = p_desc."""
                        suma = sum(u["pgini_actual"] for u in units)
                        if suma <= 0 or p_desc <= 0:
                            n = len(units) or 1
                            return {u["loc"]: round(p_desc / n, 2) for u in units}
                        return {u["loc"]: round(u["pgini_actual"] * p_desc / suma, 2) for u in units}

                    modo_disparo = st.radio(
                        "Modo de asignación al disparo:",
                        options=["1", "2", "3"],
                        format_func=lambda x: { # type: ignore
                            "1": "Mantener valores actuales (proporcional)  <- DEFAULT",
                            "2": "Ingreso manual por unidad",
                            "3": "Distribuir p_desc proporcional a pgini actual (respeta Pmax)",
                        }[x],
                        index=0,
                        key="modo_disparo",
                    )

                    pgini_manual = {}

                    # ── Tabla de valores y entradas según opción ────────────── # type: ignore
                    if _disp_units:
                        def _dif_badge(dif):
                            if abs(dif) < 1.0:
                                st.success(f"Diferencia con p_desc: {dif:+.2f} MW ✓")
                            elif abs(dif) < 5.0:
                                st.warning(f"Diferencia con p_desc: {dif:+.2f} MW")
                            else:
                                st.error(f"Diferencia con p_desc: {dif:+.2f} MW")

                        if modo_disparo == "1":
                            _rows1 = [{"Unidad": u["loc"], "pgini asignado (MW)": round(u["pgini_actual"], 2)} for u in _disp_units]
                            _suma1 = sum(u["pgini_actual"] for u in _disp_units)
                            _rows1.append({"Unidad": "SUMA", "pgini asignado (MW)": round(_suma1, 2)})
                            st.dataframe(pd.DataFrame(_rows1), hide_index=True, use_container_width=True)
                            if p_desc_ui > 0:
                                _dif_badge(_suma1 - p_desc_ui)

                        elif modo_disparo == "2":
                            st.caption("Ingrese la potencia para cada unidad:")
                            _suma2 = 0.0
                            for _u in _disp_units:
                                # Intentar recuperar valor guardado previamente en este evento
                                _saved_manual = _get_unit_cfg(ev_path, _u['loc'], "manual_pgini", float(_u["pgini_actual"]))
                                _val2 = st.number_input(
                                    f"{_u['loc']}  (actual: {_u['pgini_actual']:.2f} MW)",
                                    value=float(_saved_manual),
                                    min_value=0.0,
                                    step=1.0,
                                    format="%.2f",
                                    key=f"pgini_manual_{_u['loc']}",
                                )
                                pgini_manual[_u["loc"]] = _val2
                                _suma2 += _val2

                            if st.button("💾 Guardar potencias manuales", key="save_manual_pgini_btn", use_container_width=True):
                                for _loc_m, _val_m in pgini_manual.items():
                                    _save_unit_cfg(ev_path, _loc_m, "manual_pgini", _val_m)
                                st.toast(f"Potencias manuales guardadas para el evento {n_evento}", icon="✅")

                            st.markdown(f"**Suma:** `{_suma2:.2f} MW`")
                            if p_desc_ui > 0:
                                _dif_badge(_suma2 - p_desc_ui)
                        elif modo_disparo == "3":
                            _prev3 = _preview_prop_pdesc(_disp_units, p_desc_ui) if p_desc_ui > 0 else {u["loc"]: u["pgini_actual"] for u in _disp_units}
                            _rows3 = [{"Unidad": loc, "pgini asignado (MW)": val} for loc, val in _prev3.items()]
                            _suma3 = sum(_prev3.values())
                            _rows3.append({"Unidad": "SUMA", "pgini asignado (MW)": round(_suma3, 2)})
                            st.dataframe(pd.DataFrame(_rows3), hide_index=True, use_container_width=True)
                            if p_desc_ui > 0:
                                _dif3 = _suma3 - p_desc_ui
                                if abs(_dif3) < 0.1:
                                    st.success(f"Diferencia con p_desc: {_dif3:+.2f} MW ✓")
                                else:
                                    st.caption(f"Diferencia: {_dif3:+.2f} MW (aproximado — sin restricción Pmax)")
                    else:
                        st.caption("No se identificaron unidades del disparo en las condiciones iniciales.")

                with col_opt_b:
                    st.subheader("Post Load Flow")

                    ajustar_post_lf = st.checkbox(
                        "Activar ajuste post-LF  (AJUSTAR_POST_LF)",
                        value=False, # type: ignore
                        help=(
                            "Si está activo, el script iterará para que la potencia real "
                            "de la slack coincida con su P0_medido, redistribuyendo el "
                            "delta entre unidades CNDC_proporcional."
                        ),
                    )

                    guardar_escenario = st.checkbox(
                        "Guardar escenario de operación al finalizar",
                        value=True,
                        help=(
                            "Llama a escenario.Save() en PowerFactory al terminar la carga, "
                            "guardando pgini/plini y el resultado del Load Flow en el "
                            "IntScenario creado para este evento."
                        ),
                    )

            # ─── SECCIÓN 5 — EJECUCIÓN ───────────────────────────────────────────
            st.header("5 · Ejecución")

            _can_run = bool(ci_files)
            if not _can_run:
                st.error("❌ No se puede ejecutar: falta `condiciones_iniciales_*.xlsx`.")

            col_btn, col_reset, col_nota = st.columns([1.2, 1.2, 2.6])
            with col_btn:
                run_btn = st.button(
                    "Ejecutar en PowerFactory",
                    disabled=not _can_run or st.session_state.pf_running,
                    type="primary",
                    use_container_width=True,
                )
            with col_reset:
                if st.button("🔄 Liberar Licencia", help="Cierra procesos colgados de PowerFactory y limpia el estado"):
                    with st.spinner("Limpiando procesos..."):
                        _kill_powerfactory()
                        st.session_state.pf_running = False
                        st.session_state.pf_return_code = None
                        st.session_state.pf_waiting_close = False
                        st.toast("Licencia liberada y procesos terminados", icon="🧹")
                        time.sleep(1)
                        st.rerun()
            with col_nota:
                st.info(
                    "PowerFactory debe estar instalado en esta máquina. "
                    "El proceso puede tardar varios minutos."
                )
            
            if run_btn and ci_files:
                _params = {
                    "semestre":        semestre,
                    "evento":          evento,
                    "RAIZ":            RAIZ,
                    "PF_BASE":         PF_BASE,
                    "LOC_XFO_PATH":    LOC_XFO_PATH,
                    "LOC_GEN_PATH":    LOC_NAMES_GEN_PATH,
                    "PF_PROYECTO":     PF_PROYECTO,
                    "CASO_BASE":       CASO_BASE,
                    "modo_disparo":    modo_disparo if 'modo_disparo' in locals() else "1",
                    "pgini_manual":    pgini_manual if 'pgini_manual' in locals() else {}, # type: ignore
                    "ajustar_post_lf":   ajustar_post_lf  if 'ajustar_post_lf'  in locals() else False,
                    "guardar_escenario": guardar_escenario if 'guardar_escenario' in locals() else True,
                    "excluir_slack":   [s.strip() for s in EXCLUIR_SLACK.split(",") if s.strip()],
                    "xfo_pf":          XFO_PF,
                    "keep_pf_open":    True,
                }

                _params_path = os.path.join(ev_path, "_streamlit_params.json")
                with open(_params_path, "w", encoding="utf-8") as _fp:
                    json.dump(_params, _fp, ensure_ascii=False, indent=2)

                _runner_path = os.path.join(
                    os.path.dirname(os.path.abspath(__file__)),
                    "runners", "CargaCondIniciales_PF_run.py",
                )
                if not os.path.isfile(_runner_path):
                    st.error(f"No se encontró el runner: `{_runner_path}`")
                    st.stop()

                # Limpiar flags residuales de ejecuciones anteriores
                for _old_flag in ("_pf_waiting.flag", "_pf_continue.flag"):
                    _fp_flag = os.path.join(ev_path, _old_flag)
                    if os.path.exists(_fp_flag):
                        os.remove(_fp_flag)

                _status_file = os.path.join(ev_path, "_pf_status.txt")
                if os.path.exists(_status_file):
                    try:
                        os.remove(_status_file)
                    except OSError:
                        pass

                _log_file = os.path.join(ev_path, "_pf_log.txt")
                if os.path.exists(_log_file):
                    try:
                        os.remove(_log_file)
                    except OSError:
                        pass

                st.session_state.pf_return_code   = None
                st.session_state.pf_waiting_close  = False
                st.session_state.pf_running        = True
                st.session_state.pf_status_file    = _status_file
                st.session_state.pf_log_file       = _log_file

                def _pf_thread_fn(runner, params_path, env_vars, status_file, log_file):
                    """Ejecuta el subprocess capturando stdout+stderr al log."""
                    rc = -1
                    try:
                        proc = subprocess.Popen(
                            [sys.executable, runner, params_path],
                            stdout=subprocess.PIPE,
                            stderr=subprocess.STDOUT,
                            text=True,
                            encoding="utf-8",
                            errors="replace",
                            env=env_vars,
                        )
                        with open(log_file, "w", encoding="utf-8") as _lf:
                            for _line in proc.stdout:
                                _lf.write(_line)
                                _lf.flush()
                        rc = proc.wait()
                    except Exception as _exc:
                        try:
                            with open(log_file, "a", encoding="utf-8") as _lf:
                                _lf.write(f"\n[ERROR] {_exc}\n")
                        except OSError:
                            pass
                    finally:
                        with open(status_file, "w", encoding="utf-8") as _sf:
                            _sf.write(str(rc))

                _env = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"}
                threading.Thread(
                    target=_pf_thread_fn,
                    args=(_runner_path, _params_path, _env, _status_file, _log_file),
                    daemon=True,
                ).start()
                st.rerun()

            # ── Estado de ejecución + botón "Cerrar PF" ──────────────────────────
            _flag_waiting  = os.path.join(ev_path, "_pf_waiting.flag")
            _flag_continue = os.path.join(ev_path, "_pf_continue.flag")
            _status_file   = st.session_state.get("pf_status_file") or os.path.join(ev_path, "_pf_status.txt")
            _log_live      = st.session_state.get("pf_log_file") or os.path.join(ev_path, "_pf_log.txt")

            def _leer_log(path):
                try:
                    return open(path, encoding="utf-8", errors="replace").read()
                except OSError:
                    return ""

            def _guardar_log_final(log_path, ev_path_, n_evento_):
                """Copia el log temporal a un archivo permanente con timestamp."""
                contenido = _leer_log(log_path)
                if not contenido:
                    return None
                import datetime
                ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                dest = os.path.join(ev_path_, f"log_PF_Ev{n_evento_}_{ts}.txt")
                try:
                    with open(dest, "w", encoding="utf-8") as _f:
                        _f.write(contenido)
                    return dest
                except OSError:
                    return None

            if st.session_state.pf_running:
                # ¿El hilo terminó? → existe el archivo de estado
                if _status_file and os.path.exists(_status_file):
                    try:
                        _rc_text = open(_status_file, encoding="utf-8").read().strip()
                        st.session_state.pf_return_code = int(_rc_text)
                    except (OSError, ValueError):
                        st.session_state.pf_return_code = -1
                    st.session_state.pf_running = False
                    st.session_state.pf_waiting_close = False
                    st.rerun()
                else:
                    # Detectar si el script está esperando que el usuario cierre PF
                    if os.path.exists(_flag_waiting):
                        st.session_state.pf_waiting_close = True

                    if st.session_state.pf_waiting_close:
                        st.success("Datos cargados en PowerFactory. DIgSILENT permanece abierto.")
                        if st.button("🔒 Cerrar PowerFactory", type="primary", use_container_width=True):
                            with open(_flag_continue, "w") as _fc:
                                _fc.write("continue")
                            st.info("Señal enviada — cerrando PowerFactory...")
                            st.session_state.pf_waiting_close = False
                            time.sleep(1)
                            st.rerun()
                    else:
                        st.info("⏳ Carga en PowerFactory en curso...")
                        _monitor_process_fragment(_log_live, _status_file)

            elif st.session_state.pf_return_code is not None:
                _rc = st.session_state.pf_return_code
                if _rc == 0:
                    st.success("Ejecución completada correctamente.")
                elif _rc in (-1073741819, 3221225477):
                    st.warning(
                        "⚠️ PowerFactory se cerró abruptamente (ACCESS_VIOLATION 0xC0000005). "
                        "Revise la **Sección 6** — si el archivo `datos_cargados_Ev*.xlsx` existe, "
                        "los datos **sí fueron cargados** antes del crash."
                    )
                else:
                    st.error(f"❌ Código de error {_rc} (0x{_rc & 0xFFFFFFFF:08X}).")

                _post_files = sorted(
                    glob.glob(os.path.join(ev_path, f"datos_cargados_Ev{n_evento}*.xlsx"))
                )
                if _post_files:
                    st.success(
                        f"`{os.path.basename(_post_files[0])}` encontrado — "
                        "datos cargados correctamente. Ver **Sección 6**."
                    )
                else:
                    st.info("ℹ️ No se encontró archivo de resultados.")

                # ── Log de la ejecución ───────────────────────────────────────
                _saved_log   = st.session_state.get("pf_saved_log")
                _log_content = _leer_log(_saved_log) if _saved_log and os.path.isfile(_saved_log) else _leer_log(_log_live)
                if _log_content:
                    with st.expander("📋 Log de ejecución PowerFactory", expanded=(_rc != 0)):
                        st.code(_log_content, language="")
                        _lc1, _lc2 = st.columns(2)
                        with _lc1:
                            if not (_saved_log and os.path.isfile(_saved_log)):
                                if st.button("💾 Guardar log", key="pf_save_log"):
                                    _dest = _guardar_log_final(_log_live, ev_path, n_evento) # type: ignore
                                    if _dest:
                                        st.session_state.pf_saved_log = _dest
                                        st.success(f"Guardado en: `{_dest}`")
                                    else:
                                        st.error("No se pudo guardar el log.")
                            else: # type: ignore
                                st.caption(f"💾 Guardado en: `{_saved_log}`")
                        with _lc2:
                            st.download_button(
                                "⬇️ Descargar log",
                                data=_log_content.encode("utf-8"),
                                file_name=os.path.basename(_saved_log) if _saved_log else f"log_PF_Ev{n_evento}.txt",
                                mime="text/plain",
                            )

            # Mostrar logs de ejecuciones previas si no hay una activa
            if st.session_state.pf_return_code is None and not st.session_state.pf_running:
                _prev_logs = sorted(
                    glob.glob(os.path.join(ev_path, f"log_PF_Ev{n_evento}_*.txt")),
                    reverse=True,
                )
                if _prev_logs:
                    with st.expander(f"📋 Logs guardados de ejecuciones anteriores ({len(_prev_logs)})"):
                        _sel_log = st.selectbox(
                            "Seleccionar log",
                            _prev_logs,
                            format_func=os.path.basename,
                            key="sel_prev_log_pf",
                        )
                        if _sel_log:
                            st.code(_leer_log(_sel_log), language="")
                            st.download_button(
                                "⬇️ Descargar",
                                data=_leer_log(_sel_log).encode("utf-8"),
                                file_name=os.path.basename(_sel_log),
                                mime="text/plain",
                                key="dl_prev_log_pf",
                            )

            # ─── SECCIÓN 6 — RESULTADOS ──────────────────────────────────────────
            st.header("6 · Resultados")

            _result_files = sorted(
                glob.glob(os.path.join(ev_path, f"datos_cargados_Ev{n_evento}*.xlsx"))
            )

            if not _result_files:
                st.info("Aún no hay archivos de resultados. Ejecute el programa primero.")
            else:
                for _rf in _result_files:
                    _rf_name = os.path.basename(_rf)
                    st.subheader(f"📊 {_rf_name}")

                    try:
                        _is_ajustado = "ajustado" in _rf_name
                        _sheet_res = "Resumen_Ajustado" if _is_ajustado else "Resumen_Cargado"
                        _sheet_gen = "pgini_GEN_AJUSTADO" if _is_ajustado else "pgini_GEN_FINAL"
                        _sheet_car = "plini_CAR_FINAL"

                        _xl = pd.ExcelFile(_rf, engine="calamine")
                        _df_gen_out = _xl.parse(_sheet_gen)

                        if _is_ajustado and "plini_CAR_AJUSTADO" in _xl.sheet_names:
                            _sheet_car = "plini_CAR_AJUSTADO"
                        elif "plini_CAR_FINAL" not in _xl.sheet_names and "plini_CAR" in _xl.sheet_names:
                            _sheet_car = "plini_CAR"
                        _df_car_out = _xl.parse(_sheet_car)
                        _df_res_out = _xl.parse(_sheet_res)

                        _res_dict = {
                            str(k).strip(): v
                            for k, v in zip(_df_res_out.iloc[:, 0], _df_res_out.iloc[:, 1])
                        }

                        def _v(*keys, default=0.0):
                            for k in keys:
                                raw = _res_dict.get(k)
                                if raw is not None:
                                    try:
                                        return float(raw)
                                    except (ValueError, TypeError):
                                        pass
                            return default

                        _pgen_out  = _df_gen_out["pgini_MW"].sum() if "pgini_MW" in _df_gen_out.columns else _v("Pgen total asignada (MW)", "Pgen total ajustada (MW)")
                        _pdem_out  = _df_car_out["plini_MW"].sum() if "plini_MW" in _df_car_out.columns else _v("Pdem total asignada (MW)", "Pdem total (MW)")
                        _bal_out   = _v("Balance Pgen-Pdem (MW)", default=_pgen_out - _pdem_out)
                        _pdem_ev   = _v("Pdem_evento (MW)")
                        _slack_out = str(_res_dict.get("Slack", "—"))

                        # Potencia asignada a la slack
                        _slack_pgini: float | None = None
                        _raw_sp = _res_dict.get("pgini slack P0_medido (MW)") \
                                  or _res_dict.get("Slack real LF (MW)")
                        if _raw_sp is not None:
                            try:
                                _slack_pgini = float(_raw_sp)
                            except (ValueError, TypeError):
                                pass
                        if _slack_pgini is None and "pgini_MW" in _df_gen_out.columns:
                            _col_loc = next(
                                (c for c in _df_gen_out.columns if "loc_name" in c.lower()), None
                            )
                            if _col_loc:
                                _mask_sl = (
                                    _df_gen_out[_col_loc].astype(str).str.strip()
                                    == _slack_out.strip()
                                )
                                if _mask_sl.any():
                                    _slack_pgini = float(
                                        _df_gen_out.loc[_mask_sl, "pgini_MW"].iloc[0]
                                    )

                        _gens_asig = len(_df_gen_out)
                        _cars_asig = len(_df_car_out)
                        _cars_miss = _v("Cargas no encontradas", "Cargas no encontradas (ajuste)")

                        kc1, kc2, kc3, kc4 = st.columns(4)
                        kc1.metric("Pgen total (MW)", f"{_pgen_out:.2f}")
                        kc2.metric("Pdem total (MW)", f"{_pdem_out:.2f}")
                        kc3.metric(
                            "Balance Pgen−Pdem (MW)",
                            f"{_bal_out:+.2f}",
                            delta=None,
                            help="Diferencia entre generación y demanda asignadas a PF.",
                        )
                        kc4.metric("Pdem evento (MW)", f"{_pdem_ev:.2f}")

                        kc5, kc6, kc7, kc8 = st.columns(4)
                        kc5.metric(
                            f"P slack — {_slack_out}",
                            f"{_slack_pgini:.2f} MW" if _slack_pgini is not None else "—",
                            help="Potencia asignada (pgini) a la unidad slack.",
                        )
                        kc6.metric("Generadores asignados", int(_gens_asig))
                        kc7.metric("Cargas asignadas", int(_cars_asig))
                        kc8.metric("Cargas no encontradas", int(_cars_miss))

                        _estado_val = str(_res_dict.get("Estado validacion", ""))
                        if _estado_val == "OK":
                            st.success("✅ Estado de validación: OK")
                        elif _estado_val:
                            st.warning(f"⚠️ Estado de validación: {_estado_val}")

                        _col_ta, _col_tb = st.columns(2)

                        with _col_ta:
                            with st.expander("📋 Generadores (pgini final)"):
                                st.dataframe(_df_safe(_df_gen_out), use_container_width=True, hide_index=True)

                        with _col_tb:
                            with st.expander(f"📋 Cargas ({_sheet_car})"):
                                st.dataframe(_df_safe(_df_car_out), use_container_width=True, hide_index=True)

                        with st.expander("📋 Resumen completo"):
                            st.dataframe(_df_safe(_df_res_out), use_container_width=True, hide_index=True)

                        with open(_rf, "rb") as _fdown:
                            st.download_button(
                                label=f"⬇️  Descargar {_rf_name}",
                                data=_fdown.read(),
                                file_name=_rf_name,
                                mime=(
                                    "application/vnd.openxmlformats-officedocument"
                                    ".spreadsheetml.sheet"
                                ),
                                key=f"dl_{_rf_name}",
                            )

                    except Exception as _e:
                        st.error(f"Error al leer `{_rf_name}`: {_e}")
        else:
            st.warning("👈 Seleccione Semestre y Evento en la barra lateral para habilitar el bloque de carga.")
    
elif bloque_trabajo == "analisis_datos":
    st.header("📊 Bloque 3: Análisis de Datos Registrados")
    st.info(
        "Este bloque permite procesar y visualizar los datos reales capturados durante el evento, "
        "comparando los registros del SCADA de COBEE con las gráficas oficiales del CNDC."
    )

    if not st.session_state.semestre_global or not st.session_state.evento_global:
        st.warning("👈 Seleccione Semestre y Evento en la barra lateral.")
        st.stop()

    ev_path = st.session_state.ev_path_global
    n_evento = st.session_state.n_evento_global
    _sel_unit = st.session_state.global_selected_unit

    # ── 1. DEFINICIÓN DEL CALLBACK (Global para bloques 3, 4, 5) ──
    def _sync_rpf_y_axis(key_to_update, widget_key):
        if not st.session_state.global_selected_unit: return
        val = st.session_state.get(widget_key)
        st.session_state[f"b3_sync_{key_to_update}"] = val

        # Sincronizar las llaves de los widgets en las otras pestañas de forma inmediata
        for pfx in ["b2_sc", "b2_emf", "b3_comp"]:
            target_key = f"{pfx}_{key_to_update}"
            if target_key in st.session_state:
                st.session_state[target_key] = val

        _save_unit_cfg(st.session_state.ev_path_global, st.session_state.global_selected_unit, key_to_update, val)

    # ── 2. LÓGICA DE CARGA Y SINCRONIZACIÓN POR UNIDAD ──
    if _sel_unit:
        _pmax_map_b3 = _load_pmax_cargado(ev_path, n_evento)
        _tmap_b3 = _load_tech_map(LOC_NAMES_GEN_PATH)
        _pm_v_b3, _, _ = _get_pmax_from_cargado(_sel_unit, _pmax_map_b3, _tmap_b3)
        _pmax_ref = float(_pm_v_b3)
    else:
        _pmax_ref = 200.0

    semestre = st.session_state.semestre_global
    evento = st.session_state.evento_global
    _sel_unit = st.session_state.global_selected_unit

    tab_scada, tab_emf, tab_comp = st.tabs([
        "Registro SCADA COBEE (1SEG)", # type: ignore
        "Extracción Gráficos EMF CNDC", # type: ignore
        "Comparativa SCADA vs CNDC" # type: ignore
    ])

    # ═════════════════════════════════════════════════════════════════════════
    # SUB-TAB 1: SCADA COBEE (OrdenadorDatosEvento)
    # ═════════════════════════════════════════════════════════════════════════
    with tab_scada:
        st.subheader("📡 Procesamiento de Registros SCADA (1 Segundo)")
        st.markdown(
            "Busca el archivo '1 seg' en la carpeta de FALLA del CNDC y organiza "
            "la potencia y frecuencia en archivos CSV individuales por unidad."
        )

        scada_dir = os.path.join(ev_path, "Graficas Registro 1SEG COBEE")
        
        col1, col2 = st.columns([1, 2])
        _scada_runner = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "runners", "OrdenadorDatosEvento_run.py",
        )
        _can_scada = os.path.isfile(_scada_runner)
        if not _can_scada:
            st.error(f"No se encontró el runner: `{_scada_runner}`")

        with col1:
            scada_btn = st.button(
                "Ejecutar Ordenador de Datos SCADA",
                type="primary",
                use_container_width=True,
                disabled=not _can_scada or st.session_state.scada_running,
            )

        _scada_status_file = st.session_state.get("scada_status_file") or os.path.join(ev_path, "_scada_status.txt")

        if scada_btn and _can_scada:
            _scada_status_f = os.path.join(ev_path, "_scada_status.txt")
            if os.path.exists(_scada_status_f):
                try:
                    os.remove(_scada_status_f)
                except OSError:
                    pass

            _scada_params = {
                "semestre":   semestre,
                "evento":     evento,
                "RAIZ_RPF":   RAIZ,
                "RAIZ_DATOS": RAIZ_DATOS,
            }
            _scada_params_path = os.path.join(ev_path, "_scada_params.json")
            with open(_scada_params_path, "w", encoding="utf-8") as _fp:
                json.dump(_scada_params, _fp, ensure_ascii=False, indent=2)

            _scada_log_f = os.path.join(ev_path, "_scada_log.txt")
            st.session_state.scada_running    = True
            st.session_state.scada_status_file = _scada_status_f
            st.session_state.scada_log_file    = _scada_log_f
            st.session_state.scada_return_code = None
            st.session_state.scada_saved_log   = None

            def _scada_thread_fn(runner, params_path, env_vars, status_file, log_file):
                rc = -1
                try:
                    proc = subprocess.Popen(
                        [sys.executable, runner, params_path],
                        stdout=subprocess.PIPE,
                        stderr=subprocess.STDOUT,
                        text=True,
                        encoding="utf-8",
                        errors="replace",
                        env=env_vars,
                    )
                    with open(log_file, "w", encoding="utf-8") as _lf:
                        for _line in proc.stdout:
                            _lf.write(_line)
                            _lf.flush()
                    rc = proc.wait()
                except Exception as _exc:
                    try:
                        with open(log_file, "a", encoding="utf-8") as _lf:
                            _lf.write(f"\n[ERROR] {_exc}\n")
                    except OSError:
                        pass
                finally:
                    with open(status_file, "w", encoding="utf-8") as _sf:
                        _sf.write(str(rc))

            _scada_env = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"}
            threading.Thread(
                target=_scada_thread_fn,
                args=(_scada_runner, _scada_params_path, _scada_env, _scada_status_f, _scada_log_f),
                daemon=True,
            ).start()
            st.rerun()

        if os.path.isdir(scada_dir):
            xlsx_files_scada = _listar_archivos_cache(scada_dir, "*.xlsx")
            if xlsx_files_scada:
                st.success(f"Se encontraron {len(xlsx_files_scada)} unidades procesadas.")

                _scada_file = _buscar_archivo_unidad(_sel_unit, xlsx_files_scada)
                if _scada_file:
                    df_scada = pd.read_excel(os.path.join(scada_dir, _scada_file), engine="calamine").dropna()
                    t_raw    = _parse_to_seconds(df_scada.iloc[:, 0])
                    # Si se pide HH:MM:SS, usamos el tiempo original (t_raw) para ver la hora del día
                    # Si no, usamos el tiempo relativo (t_norm) que empieza en 0
                    t_base   = t_raw if show_hhmmss else (t_raw - t_raw.min())
                    t_norm   = t_raw - t_raw.min() # Para cálculos internos de KPIs e índices
                    unit_col = df_scada.columns[2]
                    _freq_b2_arr = pd.to_numeric(df_scada['Frecuencia_Hz'], errors='coerce').ffill().values
                    _pot_b2_arr  = pd.to_numeric(df_scada[unit_col], errors='coerce').ffill().values

                    # Inicialización de límites de ejes X para SCADA (evita NameError)
                    xaxis_min_sc = _get_unit_cfg(ev_path, _sel_unit, "b3_tab_scada_xmin", float(t_norm.min()))
                    xaxis_max_sc = _get_unit_cfg(ev_path, _sel_unit, "b3_tab_scada_xmax", float(t_norm.max()))

                    # ── Parámetros CNDC (antes del gráfico) ───────────────────
                    st.markdown("#### ⚙️ Parámetros CNDC")
                    _pmax_cargado_b2 = _load_pmax_cargado(ev_path, n_evento) # type: ignore
                    _tmap_b2         = _load_tech_map(LOC_NAMES_GEN_PATH)
                    _b2_pm_val, _tk_b2, _b2_pm_fuente = _get_pmax_from_cargado(
                        _sel_unit, _pmax_cargado_b2, _tmap_b2 # type: ignore
                    )
                    if _b2_pm_fuente:
                        st.caption(f"✅ Unidad: **{_tk_b2}** — P_max desde `{_b2_pm_fuente}`")
                    else:
                        st.warning(f"⚠️ No se encontró **{_sel_unit}** en datos_cargados ni loc_names_gen.")

                    _b2_pmax = float(_b2_pm_val)
                    _b2_rp_pct = float(_get_rp_default(_tk_b2, LOC_NAMES_GEN_PATH))
                    _cx_opt = st.columns([1])[0]
                    with _cx_opt:
                        _b2_dt       = st.number_input("Δt CNDC [s]", value=35,
                                                        min_value=20, max_value=60, step=1, key="b2_sc_dt")
                        _b2_umbral_k = st.number_input("Umbral df/dt [Hz/s]", value=-0.04,
                                                        min_value=-2.0, max_value=-0.001, step=0.005,
                                                        format="%.3f", key="b2_sc_umbral")
                        _b2_vent_suav = st.number_input("Ventana suavizado", value=5,
                                                         min_value=2, max_value=20, step=1, key="b2_sc_vsuav")

                    _idx_auto_b2  = _detectar_inicio_falla(_freq_b2_arr, float(_b2_umbral_k), int(_b2_vent_suav))
                    _idx_saved_b2 = _get_unit_cfg(ev_path, _sel_unit, "scada_idx_falla", None)
                    _idx_auto_b2  = _detectar_inicio_falla(_freq_b2_arr, float(_b2_umbral_k), int(_b2_vent_suav)) if _idx_saved_b2 is None else int(_idx_saved_b2) # type: ignore
                    
                    _csl, _cbt, _cmt    = st.columns([3, 0.5, 1])
                    _idx_falla_b2 = _csl.slider(
                        "Índice inicio de falla (ajuste si auto-detección falla):",
                        min_value=0, max_value=len(_freq_b2_arr) - 1,
                        value=_idx_auto_b2, key="b2_sc_idx_falla",
                        help=f"Auto-detección: índice {_idx_auto_b2}  "
                             f"(t = {float(t_norm.iloc[_idx_auto_b2]):.0f} s del registro).",
                    )
                    _cbt.markdown("&nbsp;", unsafe_allow_html=True)
                    if _cbt.button("💾", key="save_idx_scada", help="Guardar inicio de falla para esta unidad"):
                        if _save_unit_cfg(ev_path, _sel_unit, "scada_idx_falla", _idx_falla_b2):
                            st.toast(f"Inicio de falla guardado para {_sel_unit}", icon="✅")

                    with _cmt:
                        st.metric("t falla", f"{float(t_norm.iloc[_idx_falla_b2]):.0f} s")
                        st.metric("f₀", f"{_freq_b2_arr[_idx_falla_b2]:.4f} Hz")
                        st.metric("P₀", f"{_pot_b2_arr[_idx_falla_b2]:.3f} MW")

                    _t_falla_abs = float(t_norm.iloc[_idx_falla_b2])
                    _t_al_b2     = (t_norm - t_norm.iloc[_idx_falla_b2]).values
                    _kpi_b2      = _cndc_kpis(_t_al_b2, _freq_b2_arr, _pot_b2_arr,
                                              _b2_pmax, _b2_rp_pct / 100.0, int(_b2_dt))
                    _rocof_b2    = _calcular_rocof(_t_al_b2, _freq_b2_arr, 3.0)

                    # Ejes X e Y - Inicialización de límites desde configuración
                    xaxis_min_sc = _get_unit_cfg(ev_path, _sel_unit, "b3_tab_scada_xmin", float(t_norm.min()))
                    xaxis_max_sc = _get_unit_cfg(ev_path, _sel_unit, "b3_tab_scada_xmax", float(t_norm.max()))

                    auto_scale_sc = st.toggle("Auto-escala (Plotly)", value=st.session_state.b3_sync_y_auto, 
                                              key="b2_sc_y_auto", on_change=_sync_rpf_y_axis, args=("y_auto", "b2_sc_y_auto"))

                    # Opciones de ejes
                    with st.expander("Opciones de Ejes"):
                        col_ax1, col_ax2, col_ax3, col_ax4 = st.columns([1,1,1,0.5])
                        xaxis_min = col_ax1.number_input("X Min (s)", value=xaxis_min_sc, key="b2_sc_xaxis_min")
                        xaxis_max = col_ax1.number_input("X Max (s)", value=xaxis_max_sc, key="b2_sc_xaxis_max")
                        
                        yaxis1_min = col_ax2.number_input("Y1 Min (Hz)", value=st.session_state.b3_sync_y_f_min, key="b2_sc_y_f_min", on_change=_sync_rpf_y_axis, args=("y_f_min", "b2_sc_y_f_min"))
                        yaxis1_max = col_ax2.number_input("Y1 Max (Hz)", value=st.session_state.b3_sync_y_f_max, key="b2_sc_y_f_max", on_change=_sync_rpf_y_axis, args=("y_f_max", "b2_sc_y_f_max"))
                        yaxis2_min = col_ax3.number_input("Y2 Min (MW)", value=st.session_state.b3_sync_y_p_min, key="b2_sc_y_p_min", on_change=_sync_rpf_y_axis, args=("y_p_min", "b2_sc_y_p_min"))
                        yaxis2_max = col_ax3.number_input("Y2 Max (MW)", value=st.session_state.b3_sync_y_p_max, key="b2_sc_y_p_max", on_change=_sync_rpf_y_axis, args=("y_p_max", "b2_sc_y_p_max"))
                        
                        c_btn1, c_btn2 = col_ax4.columns(2)
                        if c_btn1.button("🔄", key="reset_scada", help="Resetear límites"):
                            _save_unit_cfg(ev_path, _sel_unit, "b3_tab_scada_xmin", float(t_norm.min()))
                            _save_unit_cfg(ev_path, _sel_unit, "b3_tab_scada_xmax", float(t_norm.max()))
                            _save_unit_cfg(ev_path, _sel_unit, "y_f_min", 49.0)
                            _save_unit_cfg(ev_path, _sel_unit, "y_f_max", 51.0)
                            _save_unit_cfg(ev_path, _sel_unit, "y_p_min", 0.0)
                            _save_unit_cfg(ev_path, _sel_unit, "y_p_max", float(_b2_pmax * 1.1))
                            _save_unit_cfg(ev_path, _sel_unit, "y_auto", True)
                            _sync_session_scale_config(ev_path, _sel_unit)
                            st.rerun()

                        if c_btn2.button("💾", key="save_scale_scada", help="Guardar escalado"):
                                _save_unit_cfg(ev_path, _sel_unit, "b3_tab_scada_xmin", xaxis_min)
                                _save_unit_cfg(ev_path, _sel_unit, "b3_tab_scada_xmax", xaxis_max)
                                st.toast("Escalado guardado para la unidad")

                    # ── Gráfico con marcadores CNDC (usando funciones estándares) ─────── # type: ignore
                    _gcfg = st.session_state.graph_config
                    
                    # Crear gráfica dual eje usando función estándar
                    fig = create_dual_axis_timeseries(
                        t_data=t_base,
                        freq_data=_freq_b2_arr,
                        pot_data=_pot_b2_arr,
                        title=f"Registro SCADA con puntos CNDC — {_scada_file}",
                        freq_label="Frecuencia SCADA (Hz)",
                        pot_label=f"Potencia {unit_col} (MW)",
                        show_hhmmss=show_hhmmss,
                        freq_color=_gcfg["freq_color_real"],
                        pot_color=_gcfg["pot_color_real"],
                        line_width=_gcfg["line_width"],
                        template=_gcfg["template"],
                        height=_gcfg["plot_height"],
                        legend_position="bottom_center",
                        x_range=None if auto_scale_sc else [xaxis_min if not show_hhmmss else t_raw.min(), xaxis_max if not show_hhmmss else t_raw.max()],
                        y1_range=None if auto_scale_sc else [st.session_state.b3_sync_y_f_min, st.session_state.b3_sync_y_f_max],
                        y2_range=None if auto_scale_sc else [st.session_state.b3_sync_y_p_min, st.session_state.b3_sync_y_p_max],
                    )
                    
                    # Añadir líneas de referencia (banda muerta, t₀, t₀+Δt)
                    if _kpi_b2:
                        # El tiempo de falla debe ser absoluto si la base es t_raw
                        _t_falla_plot = t_raw.iloc[_idx_falla_b2] if show_hhmmss else _t_falla_abs
                        _t_dt_abs  = _t_falla_plot + int(_b2_dt)
                        
                        # Líneas de referencia
                        fig = add_reference_lines(
                            fig,
                            t_fault_abs=_t_falla_abs,
                            t_eval_abs=_t_dt_abs,
                            show_hhmmss=show_hhmmss,
                            show_deadband=_gcfg["show_deadband"],
                            show_fault_line=True,
                            show_eval_line=True,
                            eval_line_label=f"t₀+Δt ({int(_b2_dt)} s)",
                        )
                        
                        # Marcadores KPI CNDC
                        # Tiempos plotables en la MISMA escala del eje X del gráfico
                        # (create_dual_axis_timeseries usa x=t_base y show_hhmmss)
                        _t0_plot  = _t_falla_plot
                        _tmin_plot = _t0_plot + float(_kpi_b2['t_min'])
                        _tdt_plot  = _t0_plot + int(_b2_dt)

                        fig = add_kpi_markers(
                            fig,
                            t_fault_abs=_t_falla_abs,
                            kpi_dict=_kpi_b2,
                            show_hhmmss=show_hhmmss,
                            dt_seconds=int(_b2_dt),
                            marker_size=_gcfg["marker_size"],
                            freq_color=_gcfg["freq_color_real"],
                            pot_color=_gcfg["pot_color_real"],
                            t0_plot=_t0_plot,
                            tmin_plot=_tmin_plot,
                            tdt_plot=_tdt_plot,
                        )
                    else:
                        # Banda muerta sin KPI
                        fig = add_reference_lines(
                            fig,
                            show_hhmmss=show_hhmmss,
                            show_deadband=_gcfg["show_deadband"],
                            show_fault_line=False,
                            show_eval_line=False,
                        )
                    
                    st.plotly_chart(fig, use_container_width=True)

                    with st.expander("📄 Ver tabla de datos"):
                        st.dataframe(_df_safe(df_scada), use_container_width=True)
                        if st.button("⬇️ Descargar datos SCADA a Excel", key=f"dl_scada_data_{_sel_unit}"):
                            excel_data = _apply_excel_formatting(
                                df_scada,
                                sheet_name=f"SCADA_{_sel_unit}",
                            )
                            st.download_button(
                                f"Descargar SCADA {_sel_unit}",
                                excel_data,
                                file_name=f"scada_data_{_sel_unit}_Ev{n_evento}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                    # ── Tabla KPIs CNDC ───────────────────────────────────────
                    st.markdown("---") # type: ignore
                    st.markdown("#### 📋 KPIs CNDC — Criterio RPF (Registro Real)")
                    _mostrar_tabla_cndc(_kpi_b2, _b2_pmax, int(_b2_dt),
                                        fuente="SCADA COBEE (1SEG)", rocof=_rocof_b2)

                elif _sel_unit:
                    st.warning(f"La unidad **{_sel_unit}** no tiene datos SCADA procesados para este evento.")
            else:
                st.warning("No se encontraron archivos Excel en 'Graficas Registro 1SEG COBEE'.")
        else:
            st.info("ℹ️ Presione el botón para organizar los datos del SCADA.")

        _scada_log_live = st.session_state.get("scada_log_file") or os.path.join(ev_path, "_scada_log.txt")

        def _leer_scada_log(path):
            try:
                return open(path, encoding="utf-8", errors="replace").read()
            except OSError:
                return ""

        def _guardar_scada_log(log_path, ev_path_, n_ev_):
            contenido = _leer_scada_log(log_path)
            if not contenido:
                return None
            import datetime
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            dest = os.path.join(ev_path_, f"log_SCADA_Ev{n_ev_}_{ts}.txt")
            try:
                with open(dest, "w", encoding="utf-8") as _f:
                    _f.write(contenido)
                return dest
            except OSError:
                return None

        if st.session_state.scada_running:
            if _scada_status_file and os.path.exists(_scada_status_file):
                try:
                    _scada_rc = int(open(_scada_status_file, encoding="utf-8").read().strip())
                except (OSError, ValueError):
                    _scada_rc = -1
                st.session_state.scada_return_code = _scada_rc
                st.session_state.scada_running = False
                st.rerun()
            else:
                st.info("⏳ Procesamiento SCADA en curso...")
                _monitor_process_fragment(_scada_log_live, _scada_status_file)
        elif st.session_state.scada_return_code is not None:
            _scada_rc = st.session_state.scada_return_code
            if _scada_rc == 0:
                st.success("✅ Procesamiento SCADA completado.")
            else:
                st.error(f"❌ Error en procesamiento SCADA (código {_scada_rc}).")
            _scada_saved = st.session_state.get("scada_saved_log")
            _scada_log_content = _leer_scada_log(_scada_saved) if _scada_saved and os.path.isfile(_scada_saved) else _leer_scada_log(_scada_log_live)
            if _scada_log_content:
                with st.expander("📋 Log de ejecución SCADA", expanded=(_scada_rc != 0)):
                    st.code(_scada_log_content, language="")
                    _sl1, _sl2 = st.columns(2)
                    with _sl1:
                        if not (_scada_saved and os.path.isfile(_scada_saved)):
                            if st.button("💾 Guardar log", key="scada_save_log"):
                                _dest = _guardar_scada_log(_scada_log_live, ev_path, n_evento)
                                if _dest:
                                    st.session_state.scada_saved_log = _dest
                                    st.success(f"Guardado en: `{_dest}`")
                                else:
                                    st.error("No se pudo guardar el log.")
                        else:
                            st.caption(f"💾 Guardado en: `{_scada_saved}`")
                    with _sl2:
                        st.download_button(
                            "⬇️ Descargar log",
                            data=_scada_log_content.encode("utf-8"),
                            file_name=os.path.basename(_scada_saved) if _scada_saved else f"log_SCADA_Ev{n_evento}.txt",
                            mime="text/plain",
                            key="scada_dl_log",
                        )

    # ═════════════════════════════════════════════════════════════════════════
    # SUB-TAB 2: GRÁFICOS EMF CNDC (ExtractorResultadosCNDC)
    # ═════════════════════════════════════════════════════════════════════════
    with tab_emf:
        st.subheader("📉 Extracción de Datos desde Gráficos EMF")
        st.markdown("Digitaliza archivos EMF (Enhanced Metafile) para extraer los puntos "
                    "exactos de frecuencia y potencia graficados por el CNDC.")

        emf_dir = os.path.join(ev_path, CARPETA_COBEE_EMF)
        
        col1, col2 = st.columns([1, 2])
        _emf_runner = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "runners", "ExtractorResultadosCNDC_run.py",
        )
        _can_emf = os.path.isfile(_emf_runner)
        if not _can_emf:
            st.error(f"No se encontró el runner: `{_emf_runner}`")

        with col1:
            emf_btn = st.button(
                "Ejecutar Extractor de Gráficos EMF",
                type="primary",
                use_container_width=True,
                disabled=not _can_emf or st.session_state.emf_running,
            )

        _emf_status_file = st.session_state.get("emf_status_file") or os.path.join(ev_path, "_emf_status.txt")

        if emf_btn and _can_emf:
            _emf_status_f = os.path.join(ev_path, "_emf_status.txt")
            if os.path.exists(_emf_status_f):
                try:
                    os.remove(_emf_status_f)
                except OSError:
                    pass

            _emf_params = {
                "semestre":      semestre,
                "evento":        evento,
                "RAIZ":          RAIZ,
                "CARPETA_COBEE": CARPETA_COBEE_EMF,
            }
            _emf_params_path = os.path.join(ev_path, "_emf_params.json")
            with open(_emf_params_path, "w", encoding="utf-8") as _fp:
                json.dump(_emf_params, _fp, ensure_ascii=False, indent=2)

            _emf_log_f = os.path.join(ev_path, "_emf_log.txt")
            st.session_state.emf_running    = True
            st.session_state.emf_status_file = _emf_status_f
            st.session_state.emf_log_file    = _emf_log_f
            st.session_state.emf_return_code = None
            st.session_state.emf_saved_log   = None

            def _emf_thread_fn(runner, params_path, env_vars, status_file, log_file):
                rc = -1
                try:
                    proc = subprocess.Popen(
                        [sys.executable, runner, params_path],
                        stdout=subprocess.PIPE,
                        stderr=subprocess.STDOUT,
                        text=True,
                        encoding="utf-8",
                        errors="replace",
                        env=env_vars,
                    )
                    with open(log_file, "w", encoding="utf-8") as _lf:
                        for _line in proc.stdout:
                            _lf.write(_line)
                            _lf.flush()
                    rc = proc.wait()
                except Exception as _exc:
                    try:
                        with open(log_file, "a", encoding="utf-8") as _lf:
                            _lf.write(f"\n[ERROR] {_exc}\n")
                    except OSError:
                        pass
                finally:
                    with open(status_file, "w", encoding="utf-8") as _sf:
                        _sf.write(str(rc))

            _emf_env = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"}
            threading.Thread(
                target=_emf_thread_fn,
                args=(_emf_runner, _emf_params_path, _emf_env, _emf_status_f, _emf_log_f),
                daemon=True,
            ).start()
            st.rerun()

        if os.path.isdir(emf_dir):
            # Buscar los Excels generados por el extractor
            xlsx_files = _listar_archivos_cache(emf_dir, "*.xlsx")
            if xlsx_files:
                st.success(f"Se encontraron {len(xlsx_files)} gráficos digitalizados.")

                _emf_file = _buscar_archivo_unidad(_sel_unit, xlsx_files)
                if _emf_file:
                    df_emf = pd.read_excel(os.path.join(emf_dir, _emf_file), engine="calamine").dropna()

                    # Detectar columnas
                    col_tiempo = 'tiempo_s' if 'tiempo_s' in df_emf.columns else df_emf.columns[0]
                    col_freq = 'frecuencia_hz' if 'frecuencia_hz' in df_emf.columns else None

                    # Normalizar tiempo (hh:mm:ss -> segundos -> relativo a 0)
                    t_raw = _parse_to_seconds(df_emf[col_tiempo])
                    t_norm = t_raw - t_raw.min()

                    # La columna de potencia es la que no es tiempo ni frecuencia
                    cols_pot = [c for c in df_emf.columns if c not in [col_tiempo, col_freq, 'hora']]

                    fig_emf = go.Figure()

                    # Eje X: puede ser segundos o la columna de hora formateada
                    x_axis = _to_plotly_time(t_norm, show_hhmmss)

                    if col_freq is not None:
                        fig_emf.add_trace(go.Scatter(
                            x=x_axis, y=df_emf[col_freq],
                            name="Frecuencia CNDC (Hz)", line=dict(color='darkblue', width=2.5),
                            yaxis="y1"
                        ))

                    for cp in cols_pot:
                        fig_emf.add_trace(go.Scatter(
                            x=x_axis, y=df_emf[cp],
                            name=f"Potencia CNDC {cp} (MW)", line=dict(color=st.session_state.graph_config["pot_color_real"], width=st.session_state.graph_config["line_width"]),
                            yaxis="y2"
                        ))

                    # ── Integración de Metodología CNDC en pestaña EMF ────────
                    st.markdown("#### ⚙️ Parámetros de Análisis (Metodología CNDC)")
                    _pmax_cargado_emf = _load_pmax_cargado(ev_path, n_evento)
                    _tmap_emf         = _load_tech_map(LOC_NAMES_GEN_PATH)
                    _emf_pm_val, _tk_emf, _emf_pm_fuente = _get_pmax_from_cargado(
                        _sel_unit, _pmax_cargado_emf, _tmap_emf
                    )

                    _gcfg = st.session_state.graph_config
                    _emf_pmax = float(_emf_pm_val)
                    _emf_rp_pct = float(_get_rp_default(_tk_emf, LOC_NAMES_GEN_PATH))

                    _ecx_opt = st.columns([1])[0]
                    _emf_dt = _ecx_opt.number_input("Δt CNDC [s]", value=35, min_value=20, max_value=60, key="b2_emf_dt")
                    _emf_umbral_k = _ecx_opt.number_input("Umbral df/dt [Hz/s]", value=-0.04, format="%.3f", key="b2_emf_um")

                    # Detección y Análisis
                    _freq_emf_arr = pd.to_numeric(df_emf[col_freq], errors='coerce').ffill().values
                    _pot_emf_arr = pd.to_numeric(df_emf[cols_pot[0]], errors='coerce').ffill().values
                    _initial_auto_idx_emf = _detectar_inicio_falla(_freq_emf_arr, _emf_umbral_k)

                    _idx_saved_emf = _get_unit_cfg(ev_path, _sel_unit, "emf_idx_falla", None)
                    _default_idx_for_slider_emf = int(_idx_saved_emf) if _idx_saved_emf is not None else _initial_auto_idx_emf
                    
                    c_emf_1, c_emf_btn = st.columns([3, 1])
                    _idx_falla_emf = c_emf_1.slider("Ajuste t₀ (falla):", 0, len(_freq_emf_arr)-1, _default_idx_for_slider_emf, key="b2_emf_idx")
                    if c_emf_btn.button("💾 Guardar t₀ EMF", key="save_idx_emf"):
                        if _save_unit_cfg(ev_path, _sel_unit, "emf_idx_falla", _idx_falla_emf): # type: ignore
                            st.toast(f"Inicio falla EMF guardado", icon="✅") # type: ignore
                    
                    # Opciones de ejes
                    with st.expander("Opciones de Ejes"):
                        auto_scale_emf = st.toggle("Auto-escala (Plotly)", value=st.session_state.b3_sync_y_auto, 
                                                   key="b2_emf_y_auto", on_change=_sync_rpf_y_axis, args=("y_auto", "b2_emf_y_auto"))

                        col_ax1, col_ax2, col_ax3, col_ax4 = st.columns([1,1,1,0.5])
                        xaxis_min = col_ax1.number_input("X Min (s)", value=_get_unit_cfg(ev_path, _sel_unit, "b3_tab_emf_xmin", float(t_norm.min())), key="b2_emf_xaxis_min")
                        xaxis_max = col_ax1.number_input("X Max (s)", value=_get_unit_cfg(ev_path, _sel_unit, "b3_tab_emf_xmax", float(t_norm.max())), key="b2_emf_xaxis_max")
                        
                        yaxis1_min = col_ax2.number_input("Y1 Min (Hz)", value=st.session_state.b3_sync_y_f_min, key="b2_emf_y_f_min", on_change=_sync_rpf_y_axis, args=("y_f_min", "b2_emf_y_f_min"))
                        yaxis1_max = col_ax2.number_input("Y1 Max (Hz)", value=st.session_state.b3_sync_y_f_max, key="b2_emf_y_f_max", on_change=_sync_rpf_y_axis, args=("y_f_max", "b2_emf_y_f_max"))
                        yaxis2_min = col_ax3.number_input("Y2 Min (MW)", value=st.session_state.b3_sync_y_p_min, key="b2_emf_y_p_min", on_change=_sync_rpf_y_axis, args=("y_p_min", "b2_emf_y_p_min"))
                        yaxis2_max = col_ax3.number_input("Y2 Max (MW)", value=st.session_state.b3_sync_y_p_max, key="b2_emf_y_p_max", on_change=_sync_rpf_y_axis, args=("y_p_max", "b2_emf_y_p_max"))
                        
                        c_btn1, c_btn2 = col_ax4.columns(2)
                        if c_btn1.button("Reset", key="reset_emf", help="Auto-detectar límites de datos y guardar"):
                            _save_unit_cfg(ev_path, _sel_unit, "b3_tab_emf_xmin", float(t_norm.min()))
                            _save_unit_cfg(ev_path, _sel_unit, "b3_tab_emf_xmax", float(t_norm.max()))
                            _sync_session_scale_config(ev_path, _sel_unit)
                            st.rerun() # type: ignore
                            
                        if c_btn2.button("Guardar", key="save_scale_emf", help="Guardar escala manual"):
                            _save_unit_cfg(ev_path, _sel_unit, "b3_tab_emf_xmin", xaxis_min); _save_unit_cfg(ev_path, _sel_unit, "b3_tab_emf_xmax", xaxis_max)
                            st.toast("Escalado EMF guardado")

                    _t_falla_emf = float(t_norm.iloc[_idx_falla_emf])
                    _t_al_emf = (t_norm - _t_falla_emf).values
                    
                    _kpi_emf = _cndc_kpis(_t_al_emf, _freq_emf_arr, _pot_emf_arr, _emf_pmax, _emf_rp_pct/100.0, _emf_dt)
                    
                    # ── Gráfico EMF con metodología CNDC (usando funciones estándares) ───
                    fig_emf = create_dual_axis_timeseries(
                        t_data=t_norm,
                        freq_data=_freq_emf_arr,
                        pot_data=_pot_emf_arr,
                        title=f"Análisis Metodológico EMF - {_emf_file}",
                        freq_label="Frecuencia CNDC (Hz)",
                        pot_label=f"Potencia {cols_pot[0]} (MW)",
                        show_hhmmss=show_hhmmss,
                        freq_color="cyan",
                        pot_color=_gcfg["pot_color_sim0"],
                        line_width=_gcfg["line_width"],
                        template=_gcfg["template"],
                        height=_gcfg["plot_height"],
                        legend_position="bottom_center",
                        x_range=None if auto_scale_emf else [xaxis_min if not show_hhmmss else t_raw.min(), xaxis_max if not show_hhmmss else t_raw.max()],
                        y1_range=None if auto_scale_emf else [st.session_state.b3_sync_y_f_min, st.session_state.b3_sync_y_f_max],
                        y2_range=None if auto_scale_emf else [st.session_state.b3_sync_y_p_min, st.session_state.b3_sync_y_p_max],
                    )
                    
                    # Añadir líneas de referencia y marcadores KPI
                    if _kpi_emf:
                        fig_emf = add_reference_lines(
                            fig_emf,
                            t_fault_abs=_t_falla_emf,
                            t_eval_abs=_t_falla_emf + _emf_dt,
                            show_hhmmss=show_hhmmss,
                            show_deadband=_gcfg["show_deadband"],
                            show_fault_line=True,
                            show_eval_line=True,
                            eval_line_label=f"t₀+Δt ({_emf_dt}s)",
                        )
                        
                        # Tiempos plotables en la MISMA escala del eje X del gráfico
                        _t0_plot_emf  = _t_falla_emf
                        _tmin_plot_emf = _t0_plot_emf + float(_kpi_emf['t_min'])
                        _tdt_plot_emf  = _t0_plot_emf + int(_emf_dt)

                        fig_emf = add_kpi_markers(
                            fig_emf,
                            t_fault_abs=_t_falla_emf,
                            kpi_dict=_kpi_emf,
                            show_hhmmss=show_hhmmss,
                            dt_seconds=_emf_dt,
                            marker_size=_gcfg["marker_size"],
                            freq_color="cyan",
                            pot_color=_gcfg["pot_color_sim0"],
                            t0_plot=_t0_plot_emf,
                            tmin_plot=_tmin_plot_emf,
                            tdt_plot=_tdt_plot_emf,
                        )
                    else:
                        fig_emf = add_reference_lines(
                            fig_emf,
                            show_hhmmss=show_hhmmss,
                            show_deadband=_gcfg["show_deadband"],
                            show_fault_line=False,
                            show_eval_line=False,
                        )
                    
                    st.plotly_chart(fig_emf, use_container_width=True)

                    st.markdown("#### 📋 KPIs CNDC — Criterio RPF (Registro EMF)")
                    if st.button("Descargar datos EMF a Excel", key=f"dl_emf_data_{_sel_unit}"): # type: ignore
                        excel_data = _apply_excel_formatting(
                            df_emf,
                            sheet_name=f"EMF_{_sel_unit}",
                        )
                        st.download_button(
                            f"Descargar EMF {_sel_unit}",
                            excel_data,
                            file_name=f"emf_data_{_sel_unit}_Ev{n_evento}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    _mostrar_tabla_cndc(_kpi_emf, _emf_pmax, _emf_dt, fuente="Gráfico EMF CNDC")
                elif _sel_unit:
                    st.warning(f"La unidad **{_sel_unit}** no tiene datos EMF procesados para este evento.")
            else:
                # Mostrar archivos EMF disponibles pero no procesados # type: ignore
                emfs = [f for f in os.listdir(emf_dir) if f.lower().endswith('.emf')]
                if emfs:
                    st.info(f"Se detectaron {len(emfs)} archivos .emf listos para extraer. Presione el botón superior.")
                    for e in emfs[:5]:
                        st.caption(f"• {e}")
                    if len(emfs) > 5: st.caption(f"... y {len(emfs)-5} más.")
                else:
                    st.warning("No se encontraron archivos .emf en la carpeta 'Resultados_COBEE'.")
        else:
            st.info(f"ℹ️ La carpeta '{CARPETA_COBEE_EMF}' no existe en este evento.")

        _emf_log_live = st.session_state.get("emf_log_file") or os.path.join(ev_path, "_emf_log.txt")

        def _leer_emf_log(path):
            try:
                return open(path, encoding="utf-8", errors="replace").read()
            except OSError:
                return ""

        def _guardar_emf_log(log_path, ev_path_, n_ev_):
            contenido = _leer_emf_log(log_path)
            if not contenido:
                return None
            import datetime
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            dest = os.path.join(ev_path_, f"log_EMF_Ev{n_ev_}_{ts}.txt")
            try:
                with open(dest, "w", encoding="utf-8") as _f:
                    _f.write(contenido)
                return dest
            except OSError:
                return None

        if st.session_state.emf_running:
            if _emf_status_file and os.path.exists(_emf_status_file):
                try:
                    _emf_rc = int(open(_emf_status_file, encoding="utf-8").read().strip())
                except (OSError, ValueError):
                    _emf_rc = -1
                st.session_state.emf_return_code = _emf_rc
                st.session_state.emf_running = False
                st.rerun()
            else:
                st.info("⏳ Extracción EMF en curso...")
                _monitor_process_fragment(_emf_log_live, _emf_status_file)
        elif st.session_state.emf_return_code is not None:
            _emf_rc = st.session_state.emf_return_code
            if _emf_rc == 0:
                st.success("✅ Extracción EMF completada.")
            else:
                st.error(f"❌ Error en extracción EMF (código {_emf_rc}).")
            _emf_saved = st.session_state.get("emf_saved_log")
            _emf_log_content = _leer_emf_log(_emf_saved) if _emf_saved and os.path.isfile(_emf_saved) else _leer_emf_log(_emf_log_live)
            if _emf_log_content:
                with st.expander("📋 Log de ejecución EMF", expanded=(_emf_rc != 0)):
                    st.code(_emf_log_content, language="")
                    _eml1, _eml2 = st.columns(2)
                    with _eml1:
                        if not (_emf_saved and os.path.isfile(_emf_saved)):
                            if st.button("💾 Guardar log", key="emf_save_log"):
                                _dest = _guardar_emf_log(_emf_log_live, ev_path, n_evento)
                                if _dest:
                                    st.session_state.emf_saved_log = _dest
                                    st.success(f"Guardado en: `{_dest}`")
                                else:
                                    st.error("No se pudo guardar el log.")
                        else:
                            st.caption(f"💾 Guardado en: `{_emf_saved}`")
                    with _eml2:
                        st.download_button(
                            "⬇️ Descargar log",
                            data=_emf_log_content.encode("utf-8"),
                            file_name=os.path.basename(_emf_saved) if _emf_saved else f"log_EMF_Ev{n_evento}.txt",
                            mime="text/plain",
                            key="emf_dl_log",
                        )

    # ═════════════════════════════════════════════════════════════════════════
    # SUB-TAB 3: COMPARATIVA (Súper Gráfico)
    # ═════════════════════════════════════════════════════════════════════════
    with tab_comp:
        st.subheader("📈 Comparativa Dinámica: SCADA vs CNDC")
        
        scada_dir = os.path.join(ev_path, "Graficas Registro 1SEG COBEE")
        emf_dir = os.path.join(ev_path, "Resultados_COBEE") # type: ignore

        if not _sel_unit:
            st.info("ℹ️ Seleccione una unidad en el selector superior para ver la comparativa.")
        elif os.path.isdir(scada_dir) and os.path.isdir(emf_dir):
            _s_file = _buscar_archivo_unidad(_sel_unit, _listar_archivos_cache(scada_dir, "*.xlsx"))
            _e_file = _buscar_archivo_unidad(_sel_unit, _listar_archivos_cache(emf_dir, "*.xlsx"))
            _has_s = _s_file is not None
            _has_e = _e_file is not None

            # Cargar índices de falla guardados por el usuario
            scada_idx_falla = _get_unit_cfg(ev_path, _sel_unit, "scada_idx_falla", 0) # type: ignore
            emf_idx_falla = _get_unit_cfg(ev_path, _sel_unit, "emf_idx_falla", 0)

            if _has_s and _has_e:
                # --- CARGA Y ALINEACIÓN DE DATOS ---
                df_s = pd.read_excel(os.path.join(scada_dir, _s_file), engine="calamine").dropna() # type: ignore
                df_e = pd.read_excel(os.path.join(emf_dir, _e_file), engine="calamine").dropna()

                ts_raw = _parse_to_seconds(df_s.iloc[:, 0])
                df_s['t_norm'] = ts_raw - ts_raw.min() 
                t_falla_scada = float(df_s['t_norm'].iloc[scada_idx_falla]) if scada_idx_falla < len(df_s) else 0.0
                t_scada_aligned = df_s['t_norm'] - t_falla_scada

                te_raw = _parse_to_seconds(df_e['tiempo_s'])
                df_e['t_norm'] = te_raw - te_raw.min() 
                t_falla_emf = float(df_e['t_norm'].iloc[emf_idx_falla]) if emf_idx_falla < len(df_e) else 0.0
                t_emf_aligned = df_e['t_norm'] - t_falla_emf

                # --- CONSTRUCCIÓN DEL GRÁFICO ESTANDARIZADO ---
                _gcfg = st.session_state.graph_config
                p_col_s = df_s.columns[2]
                p_col_e = [c for c in df_e.columns if c not in ['tiempo_s', 'frecuencia_hz', 'hora', 't_norm']][0]

                # 1. Crear base con SCADA (igual que Tab 1)
                fig_c = create_dual_axis_timeseries(
                    t_data=t_scada_aligned,
                    freq_data=df_s['Frecuencia_Hz'],
                    pot_data=df_s[p_col_s],
                    title=f"Comparativa Registro Real vs CNDC — {_sel_unit}",
                    freq_label="Frec. SCADA",
                    pot_label="Pot. SCADA",
                    show_hhmmss=False,
                    freq_color=_gcfg["freq_color_real"],
                    pot_color=_gcfg["pot_color_real"],
                    line_width=_gcfg["line_width"],
                    template=_gcfg["template"],
                    height=_gcfg["plot_height"]
                )

                # 2. Añadir capas de CNDC (EMF) usando colores distintos para diferenciar fuente
                #    (SCADA = tonos "real", EMF = tonos "sim" para que se vean claramente)
                fig_c.add_trace(go.Scatter(
                    x=t_emf_aligned,
                    y=df_e['frecuencia_hz'],
                    name="Frec. CNDC (EMF)",
                    line=dict(color="cyan", width=_gcfg["line_width"]),
                    yaxis="y",
                ))
                fig_c.add_trace(go.Scatter(
                    x=t_emf_aligned,
                    y=df_e[p_col_e],
                    name="Pot. CNDC (EMF)",
                    line=dict(color=_gcfg["pot_color_sim0"], width=_gcfg["line_width"]),
                    yaxis="y2",
                ))

                # 3. Aplicar líneas de referencia segmentadas (t0 y t0+35s)
                fig_c = add_reference_lines(
                    fig_c,
                    t_fault_abs=0.0,   # Ya alineado
                    t_eval_abs=35.0,  # Marcador solicitado a t=35
                    show_hhmmss=False,
                    show_deadband=_gcfg["show_deadband"],
                    show_fault_line=True,
                    show_eval_line=True,
                    eval_line_label="t₀+35s (CNDC)"
                )

                # 4. Finalizar con el layout estándar de la aplicación
                fig_c = apply_standard_layout(
                    fig_c,
                    title=f"Comparativa Registro Real vs CNDC — {_sel_unit}",
                    xaxis_title="Segundos desde inicio falla (Alineado)",
                    yaxis_title="Frecuencia (Hz)",
                    yaxis2_title="Potencia (MW)",
                    template=_gcfg["template"],
                    height=_gcfg["plot_height"]
                )

                # Mini análisis de error
                with st.expander("Análisis de desviación"):
                    st.write("Diferencia promedio en Frecuencia: " +
                             f"{abs(df_s['Frecuencia_Hz'].mean() - df_e['frecuencia_hz'].mean()):.4f} Hz")
                    st.write(f"Diferencia promedio en Potencia: " +
                             f"{abs(df_s[p_col_s].mean() - df_e[p_col_e].mean()):.2f} MW") # type: ignore

                # Opciones de ejes (se aplican al primer y único gráfico)
                with st.expander("Opciones de Ejes"):
                    col_ax1, col_ax2, col_ax3, col_ax4 = st.columns([1, 1, 1, 0.5])
                    # Defaults based on aligned data
                    _t_comb_aligned = pd.concat([t_scada_aligned, t_emf_aligned]).dropna()
                    _f_comb = pd.concat([df_s['Frecuencia_Hz'], df_e['frecuencia_hz']]).dropna()
                    _p_comb = pd.concat([df_s[p_col_s], df_e[p_col_e]]).dropna()

                    xaxis_min = col_ax1.number_input("X Min (s)", value=_get_unit_cfg(ev_path, _sel_unit, "b3_tab_comp_xmin", float(_t_comb_aligned.min()) if not _t_comb_aligned.empty else -10.0), key="b3_comp_xmin")
                    xaxis_max = col_ax1.number_input("X Max (s)", value=_get_unit_cfg(ev_path, _sel_unit, "b3_tab_comp_xmax", float(_t_comb_aligned.max()) if not _t_comb_aligned.empty else 100.0), key="b3_comp_xmax")
                    
                    yaxis1_min = col_ax2.number_input("Y1 Min (Hz)", value=st.session_state.b3_sync_y_f_min, key="b3_comp_y_f_min", on_change=_sync_rpf_y_axis, args=("y_f_min", "b3_comp_y_f_min"))
                    yaxis1_max = col_ax2.number_input("Y1 Max (Hz)", value=st.session_state.b3_sync_y_f_max, key="b3_comp_y_f_max", on_change=_sync_rpf_y_axis, args=("y_f_max", "b3_comp_y_f_max"))
                    yaxis2_min = col_ax3.number_input("Y2 Min (MW)", value=st.session_state.b3_sync_y_p_min, key="b3_comp_y_p_min", on_change=_sync_rpf_y_axis, args=("y_p_min", "b3_comp_y_p_min"))
                    yaxis2_max = col_ax3.number_input("Y2 Max (MW)", value=st.session_state.b3_sync_y_p_max, key="b3_comp_y_p_max", on_change=_sync_rpf_y_axis, args=("y_p_max", "b3_comp_y_p_max"))

                    auto_scale_comp = st.toggle("Auto-escala (Plotly)", value=st.session_state.b3_sync_y_auto, 
                                                key="b3_comp_y_auto", on_change=_sync_rpf_y_axis, args=("y_auto", "b3_comp_y_auto"))

                    c_btn1, c_btn2 = col_ax4.columns(2)
                    if c_btn1.button("🔄", key="reset_comp", help="Auto-detectar límites de datos y guardar"):
                        _save_unit_cfg(ev_path, _sel_unit, "b3_tab_comp_xmin", float(_t_comb_aligned.min()) if not _t_comb_aligned.empty else -10.0)
                        _save_unit_cfg(ev_path, _sel_unit, "b3_tab_comp_xmax", float(_t_comb_aligned.max()) if not _t_comb_aligned.empty else 100.0) # type: ignore
                        _sync_session_scale_config(ev_path, _sel_unit)
                        st.rerun()

                    if c_btn2.button("Guardar", key="save_scale_comp", help="Guardar escala manual"):
                        _save_unit_cfg(ev_path, _sel_unit, "b3_tab_comp_xmin", xaxis_min); _save_unit_cfg(ev_path, _sel_unit, "b3_tab_comp_xmax", xaxis_max)
                        st.toast("Escala Comparativa guardada")

                fig_c.update_layout(
                    xaxis=dict(range=None if auto_scale_comp else [_to_plotly_time(xaxis_min, show_hhmmss), _to_plotly_time(xaxis_max, show_hhmmss)]),
                    yaxis=dict(range=None if auto_scale_comp else [st.session_state.b3_sync_y_f_min, st.session_state.b3_sync_y_f_max]),
                    yaxis2=dict(range=None if auto_scale_comp else [st.session_state.b3_sync_y_p_min, st.session_state.b3_sync_y_p_max]),
                )
                st.plotly_chart(fig_c, use_container_width=True)

            else:
                _missing = []
                if not _has_s: _missing.append("SCADA")
                if not _has_e: _missing.append("EMF CNDC")
                st.warning(f"La unidad **{_sel_unit}** no tiene datos procesados en: {', '.join(_missing)}.")
        else:
            st.info("Asegúrese de haber ejecutado los procesadores en las pestañas anteriores.")

    # ── Exportación masiva Bloque 3 ──────────────────────────────────────────
    st.markdown("---") # type: ignore
    st.subheader("📥 Exportar todos los gráficos de Bloque 3")
    st.caption("Genera capturas PNG de SCADA y EMF para todas las unidades disponibles.")

    if st.button("🗂️ Generar ZIP de gráficos registrados (SCADA/EMF)", key="btn_zip_b2"):
        import io, zipfile, datetime
        from plotly.io import to_image

        # Resetear descarga previa
        st.session_state.b3_plots_zip_bytes = None

        _zip_buf = io.BytesIO()
        _n_ok = 0
        _available = get_event_units(ev_path, n_evento)
        _pmax_map_exp = _load_pmax_cargado(ev_path, n_evento)
        _tmap_exp = _load_tech_map(LOC_NAMES_GEN_PATH)
        _prog = st.progress(0, text="Iniciando exportación masiva...")
        _gcfg = st.session_state.graph_config
        
        with zipfile.ZipFile(_zip_buf, 'w', zipfile.ZIP_DEFLATED) as _zf:
            for _idx, _uname in enumerate(_available):
                _prog.progress((_idx + 1) / len(_available), text=f"Procesando {_uname}...")
                
                # --- EXPORTAR SCADA ---
                _s_file = os.path.join(ev_path, "Graficas Registro 1SEG COBEE", f"{_uname}.xlsx")
                if not os.path.isfile(_s_file):
                    _s_file = _buscar_archivo_unidad(_uname, _listar_archivos_cache(os.path.join(ev_path, "Graficas Registro 1SEG COBEE"), "*.xlsx"))
                    if _s_file: _s_file = os.path.join(ev_path, "Graficas Registro 1SEG COBEE", _s_file)

                if os.path.isfile(_s_file):
                    try: # type: ignore
                        _df_s = pd.read_excel(_s_file, engine="calamine").dropna()
                        _tr_s = _parse_to_seconds(_df_s.iloc[:, 0])
                        _t_norm_s = _tr_s - _tr_s.min()
                        _fr_s = pd.to_numeric(_df_s['Frecuencia_Hz'], errors='coerce').ffill().values
                        _pt_s = pd.to_numeric(_df_s.iloc[:, 2], errors='coerce').ffill().values
                        
                        _idx_f = int(_get_unit_cfg(ev_path, _uname, "scada_idx_falla", _detectar_inicio_falla(_fr_s)))
                        _t_f_abs = float(_t_norm_s.iloc[_idx_f])
                        
                        _pm_v, _tk, _ = _get_pmax_from_cargado(_uname, _pmax_map_exp, _tmap_exp)
                        _rp_v = _get_rp_default(_tk, LOC_NAMES_GEN_PATH) / 100.0
                        _dt_v = int(_get_unit_cfg(ev_path, _uname, "b2_sc_dt", 35))
                        _t_al_s = (_t_norm_s - _t_f_abs).values
                        _kpi = _cndc_kpis(_t_al_s, _fr_s, _pt_s, _pm_v, _rp_v, _dt_v)
                        _rocof_s = _calcular_rocof(_t_al_s, _fr_s, 3.0)

                        _y_auto = _get_unit_cfg(ev_path, _uname, "y_auto", True)
                        
                        _fig = create_dual_axis_timeseries(
                            t_data=_t_norm_s if not show_hhmmss else _tr_s,
                            freq_data=_fr_s, pot_data=_pt_s,
                            title=f"Registro SCADA - {_uname}",
                            show_hhmmss=show_hhmmss,
                            x_range=None if _y_auto else [_get_unit_cfg(ev_path, _uname, "b3_tab_scada_xmin", None), _get_unit_cfg(ev_path, _uname, "b3_tab_scada_xmax", None)],
                            y1_range=None if _y_auto else [_get_unit_cfg(ev_path, _uname, "y_f_min", None), _get_unit_cfg(ev_path, _uname, "y_f_max", None)],
                            y2_range=None if _y_auto else [_get_unit_cfg(ev_path, _uname, "y_p_min", None), _get_unit_cfg(ev_path, _uname, "y_p_max", None)],
                        )
                        _fig = add_reference_lines(_fig, t_fault_abs=_t_f_abs if not show_hhmmss else _tr_s.iloc[_idx_f], 
                                                  t_eval_abs=(_t_f_abs + _dt_v) if not show_hhmmss else (_tr_s.iloc[_idx_f] + _dt_v),
                                                  show_hhmmss=show_hhmmss)
                        _fig = add_kpi_markers(_fig, t_fault_abs=_t_f_abs if not show_hhmmss else _tr_s.iloc[_idx_f], 
                                              kpi_dict=_kpi, show_hhmmss=show_hhmmss, dt_seconds=_dt_v)
                        
                        _img = to_image(_fig, format="png", width=1200, height=600, scale=2)
                        _zf.writestr(f"SCADA_{_uname}_Ev{n_evento}.png", _img)
                        _n_ok += 1
                    except: pass

                # --- EXPORTAR EMF ---
                _e_file = os.path.join(ev_path, CARPETA_COBEE_EMF, f"{_uname}.xlsx")
                if not os.path.isfile(_e_file):
                    _e_file = _buscar_archivo_unidad(_uname, _listar_archivos_cache(os.path.join(ev_path, CARPETA_COBEE_EMF), "*.xlsx"))
                    if _e_file: _e_file = os.path.join(ev_path, CARPETA_COBEE_EMF, _e_file)

                if os.path.isfile(_e_file):
                    try: # type: ignore
                        _df_e = pd.read_excel(_e_file, engine="calamine").dropna()
                        _tr_e = _parse_to_seconds(_df_e['tiempo_s'])
                        _t_norm_e = _tr_e - _tr_e.min()
                        _fr_e = pd.to_numeric(_df_e['frecuencia_hz'], errors='coerce').ffill().values
                        _pcol = [c for c in _df_e.columns if c not in ['tiempo_s', 'frecuencia_hz', 'hora', 't_norm']][0]
                        _pt_e = pd.to_numeric(_df_e[_pcol], errors='coerce').ffill().values
                        
                        _idx_e = int(_get_unit_cfg(ev_path, _uname, "emf_idx_falla", _detectar_inicio_falla(_fr_e)))
                        _t_fe_abs = float(_t_norm_e.iloc[_idx_e])
                        
                        _pm_v, _tk, _ = _get_pmax_from_cargado(_uname, _pmax_map_exp, _tmap_exp)
                        _rp_v = _get_rp_default(_tk, LOC_NAMES_GEN_PATH) / 100.0
                        _dt_v = int(_get_unit_cfg(ev_path, _uname, "b2_emf_dt", 35))
                        _t_al_e = (_t_norm_e - _t_fe_abs).values
                        _kpi_e = _cndc_kpis(_t_al_e, _fr_e, _pt_e, _pm_v, _rp_v, _dt_v)
                        _rocof_e = _calcular_rocof(_t_al_e, _fr_e, 3.0)

                        _y_auto_e = _get_unit_cfg(ev_path, _uname, "y_auto", True)
                        
                        _fig_e = create_dual_axis_timeseries(
                            t_data=_t_norm_e if not show_hhmmss else _tr_e,
                            freq_data=_fr_e, pot_data=_pt_e,
                            title=f"Gráfico CNDC (EMF) - {_uname}",
                            show_hhmmss=show_hhmmss,
                            freq_color="cyan",
                            pot_color=_gcfg["pot_color_sim0"],
                            x_range=None if _y_auto_e else [_get_unit_cfg(ev_path, _uname, "b3_tab_emf_xmin", None), _get_unit_cfg(ev_path, _uname, "b3_tab_emf_xmax", None)],
                            y1_range=None if _y_auto_e else [_get_unit_cfg(ev_path, _uname, "y_f_min", None), _get_unit_cfg(ev_path, _uname, "y_f_max", None)],
                            y2_range=None if _y_auto_e else [_get_unit_cfg(ev_path, _uname, "y_p_min", None), _get_unit_cfg(ev_path, _uname, "y_p_max", None)],
                        )
                        _fig_e = add_reference_lines(_fig_e, t_fault_abs=_t_fe_abs if not show_hhmmss else _tr_e.iloc[_idx_e], 
                                                   t_eval_abs=(_t_fe_abs + _dt_v) if not show_hhmmss else (_tr_e.iloc[_idx_e] + _dt_v),
                                                   show_hhmmss=show_hhmmss)
                        _fig_e = add_kpi_markers(_fig_e, t_fault_abs=_t_fe_abs if not show_hhmmss else _tr_e.iloc[_idx_e], 
                                               kpi_dict=_kpi_e, show_hhmmss=show_hhmmss, dt_seconds=_dt_v,
                                               freq_color="cyan",
                                               pot_color=_gcfg["pot_color_sim0"])
                        
                        _img_e = to_image(_fig_e, format="png", width=1200, height=600, scale=2)
                        _zf.writestr(f"EMF_{_uname}_Ev{n_evento}.png", _img_e)
                        _n_ok += 1
                    except: pass

        _prog.empty()
        if _n_ok > 0:
            st.session_state.b3_plots_zip_bytes = _zip_buf.getvalue()
            st.session_state.b3_plots_zip_name = f"graficos_registrados_Ev{n_evento}_{datetime.datetime.now().strftime('%H%M%S')}.zip"
            st.success(f"✅ Se generaron {_n_ok} gráficos con éxito.")
        else:
            st.error("No se pudieron generar imágenes. Verifique que existan archivos procesados.")

    if st.session_state.get("b3_plots_zip_bytes"):
        st.download_button(
            label=f"⬇️ Descargar ZIP de Gráficos Registrados",
            data=st.session_state.b3_plots_zip_bytes,
            file_name=st.session_state.b3_plots_zip_name,
            mime="application/zip",
            type="primary"
        )

elif bloque_trabajo == "analisis_simulacion":
    st.header("Bloque 4: Análisis de Resultados de Simulación")
    st.info(
        "Este bloque permite visualizar y comparar los resultados de las simulaciones RMS generados por PowerFactory."
    )
    st.warning(
        "⚠️ Para usar este bloque, primero debe ejecutar el script `DatosCurvas_v3.py` "
        "**dentro de PowerFactory** para generar los archivos Excel en las carpetas "
        "`E{N}.0/Datos Curvas/` y `E{N}.1/Datos Curvas/`."
    )

    if not st.session_state.semestre_global or not st.session_state.evento_global:
        st.warning("👈 Seleccione Semestre y Evento en la barra lateral.")
        st.stop()

    ev_path = st.session_state.ev_path_global
    n_evento = st.session_state.n_evento_global
    _sel_unit = st.session_state.global_selected_unit
    _event_cfg = _load_event_cfg(ev_path)

    _dir0_b3 = os.path.join(ev_path, f"E{n_evento}.0", CARPETA_DATOS_CURVAS)
    _dir1_b3 = os.path.join(ev_path, f"E{n_evento}.1", CARPETA_DATOS_CURVAS) # type: ignore

    # Resolver archivo de simulación basado en la unidad global seleccionada
    _sel_file_b3 = None
    if st.session_state.global_selected_unit:
        _target = st.session_state.global_selected_unit.replace("sym_", "")
        # Buscar en las carpetas de simulación
        for _d in [_dir0_b3, _dir1_b3]:
            if os.path.isdir(_d):
                for _f in os.listdir(_d):
                    if _target in _f and _f.endswith('.xlsx'):
                        _sel_file_b3 = _f
                        break # type: ignore
                if _sel_file_b3: break

    # ── Parámetros de análisis CNDC (compartidos entre pestañas) ────────────────
    st.markdown("---") # type: ignore
    st.markdown("### ⚙️ Parámetros de Análisis CNDC")
    _bp1, _bp2 = st.columns(2)
    _b3_t_falla = _bp1.number_input(
        "Tiempo de falla en simulación (t₀) [s]",
        value=float(_event_cfg.get("t_sim_falla", 5.0)),
        min_value=0.0, max_value=300.0, step=0.5,
        help="Instante t en PowerFactory donde ocurre el evento (t=0 en la comparativa).",
        key="b3_t_falla",
    )
    _b3_dt = int(_bp2.number_input( # type: ignore
        "Δt CNDC [s]",
        value=int(_event_cfg.get("delta_t_cndc", 35)),
        min_value=20, max_value=60, step=1,
        help="Tiempo desde t₀ para evaluar f_Δt y P_Δt. CNDC usa 30–50 s (típicamente 35 s).",
        key="b3_dt",
    ))
    
    tab_sim_cndc, tab_sim_cobee, tab_sim_comp = st.tabs([ # type: ignore
        f"Simulación E{n_evento}.0 (CNDC)",
        f"Simulación E{n_evento}.1 (COBEE)",
        "Comparativa de Simulaciones"
    ])

    # Helper: gráfico anotado + tabla CNDC para un DataFrame de simulación
    def _create_sim_figure(df_sim, sim_label, t_falla, delta_t, unit_name, ev_path, n_evento, show_hhmmss, xaxis_range=None, yaxis1_range=None, yaxis2_range=None):
        if df_sim is None or _sel_file_b3 is None: # type: ignore
            return
        _pmax_cargado_b3 = _load_pmax_cargado(ev_path, n_evento)
        _gcfg = st.session_state.graph_config
        _tech_b3         = _load_tech_map(LOC_NAMES_GEN_PATH)
        _pm_val, _tk_b3, _pm_fuente = _get_pmax_from_cargado(
            _sel_file_b3, _pmax_cargado_b3, _tech_b3
        )
        _tk_encontrado = _pm_fuente is not None # type: ignore

        # st.markdown("---") # Removed for batch export
        st.markdown(f"#### 📊 Análisis RPF — Puntos CNDC ({sim_label})")

        _tc = df_sim.columns[0]
        _t_raw = pd.to_numeric(df_sim[_tc], errors='coerce').values

        if _tk_encontrado:
            _fuente_label = "datos_cargados" if _pm_fuente == "datos_cargados" else "loc_names_gen.xlsx"
            st.caption(f"✅ Unidad: **{_tk_b3}** — P_max desde `{_fuente_label}`")
        else:
            st.warning(
                f"⚠️ No se encontró **{os.path.splitext(_sel_file_b3)[0]}** en datos_cargados ni loc_names_gen. "
                "Ingrese P_max manualmente."
            )
        _pm = float(_pm_val)
        _rp = float(_get_rp_default(_tk_b3, LOC_NAMES_GEN_PATH)) # type: ignore

        _tc = df_sim.columns[0]
        _t_raw = pd.to_numeric(df_sim[_tc], errors='coerce').values
        _dcols = [c for c in df_sim.columns if c != _tc] # type: ignore
        
        # Selección robusta: priorizar columnas con variación (dinámicas) sobre referencias fijas
        _fc_cands = [c for c in _dcols if _is_frequency_column(c, df_sim[c])]
        if len(_fc_cands) > 1:
            _fc_cands = sorted(_fc_cands, key=lambda c: pd.to_numeric(df_sim[c], errors='coerce').std(), reverse=True)
        
        _fc = _fc_cands[:1]
        _pc = [c for c in _dcols if c not in _fc_cands]
        
        if not _fc or not _pc:
            st.warning("No se pudieron identificar columnas de frecuencia/potencia.")
            return
            
        _freq_raw = pd.to_numeric(df_sim[_fc[0]], errors='coerce').ffill().bfill().values
        # Conversión automática p.u. -> Hz para análisis CNDC (requerido para metodología RPF)
        _freq_s = _freq_raw * 50.0 if np.nanmax(_freq_raw) < 2.0 else _freq_raw
        
        _pot_s  = pd.to_numeric(df_sim[_pc[0]], errors='coerce').ffill().bfill().values
        # KPIs usan tiempo alineado a t0=0 (como Bloque 3)
        _t_al = _t_raw - t_falla
        _kpi  = _cndc_kpis(_t_al, _freq_s, _pot_s, _pm, _rp / 100.0, delta_t)

        # Para ubicar marcadores/referencias en la curva, t_fault_abs debe estar en el
        # mismo sistema de eje X que usa la gráfica: t_data=_t_raw (tiempo absoluto en el archivo).
        # Por eso usamos explícitamente t_fault_abs=t_falla (no t0 alineado).
        t_fault_abs_plot = t_falla
        _rocof  = _calcular_rocof(_t_al, _freq_s, 3.0)

        _c_f = _gcfg["freq_color_sim0"] if sim_label.endswith('.0') else _gcfg["freq_color_sim1"]
        _c_p = _gcfg["pot_color_sim0"]  if sim_label.endswith('.0') else _gcfg["pot_color_sim1"]

        # Usar funciones constructoras estándares (idéntico al Bloque 3)
        # Estandarización completa: mismas funciones y contrato visual que Bloque 3
        # (doble eje Y + referencias CNDC + marcadores KPI + layout estándar)
        _show_deadband = _gcfg.get("show_deadband", True)

        fig_ann = create_dual_axis_timeseries(
            t_data=_t_raw,
            freq_data=_freq_s,
            pot_data=_pot_s,
            title=f"Puntos de evaluación CNDC — {sim_label}",
            freq_label="Frecuencia (Hz)",
            pot_label="Potencia (MW)",
            show_hhmmss=show_hhmmss,
            freq_color=_c_f,
            pot_color=_c_p,
            line_width=_gcfg.get("line_width", None),
            template=_gcfg.get("template", None),
            height=_gcfg.get("plot_height", None),
            legend_position="bottom_center",
            x_range=xaxis_range,
            y1_range=yaxis1_range,
            y2_range=yaxis2_range,
        )

        # Referencias (banda muerta + t0 + t0+Δt) — mismas claves que Bloque 3
        fig_ann = add_reference_lines(
            fig_ann,
                            # Mantener compatibilidad: si `t_fault_abs_plot` no existe (algunos refactors),
                            # se usa el t_falla recibido como referencia.
t_fault_abs=(t_fault_abs_plot if 't_fault_abs_plot' in locals() else t_falla),
t_eval_abs=((t_fault_abs_plot if 't_fault_abs_plot' in locals() else t_falla) + delta_t),
            show_hhmmss=show_hhmmss,
            show_deadband=_show_deadband,
            show_fault_line=_gcfg.get("show_fault_line", True),
            show_eval_line=_gcfg.get("show_eval_line", True),
            eval_line_label=f"t₀+Δt ({int(delta_t)} s)",
        )

        # KPIs/marcadores CNDC (○/×/●) con mismas funciones
        # Nota: mantén exactamente los mismos marcadores/orden que Bloque 3.
        fig_ann = add_kpi_markers(
            fig_ann,
            t_fault_abs=t_fault_abs_plot,
            kpi_dict=_kpi,
            show_hhmmss=show_hhmmss,
            dt_seconds=int(delta_t),
            marker_size=_gcfg.get("marker_size", None),
            freq_color=_c_f,
            pot_color=_c_p,
        )

        # Layout estándar final (idéntico estilo general Bloque 3)
        fig_ann = apply_standard_layout(
            fig_ann,
            title=f"Puntos de evaluación CNDC — {sim_label}",
            legend_position="bottom_center",
            template=_gcfg.get("template", None),
            height=_gcfg.get("plot_height", None),
            show_grid=_gcfg.get("show_grid", True),
        )

        return fig_ann, _kpi, _pm, _rocof


    def _create_comp_figure(df0, df1, sim_label0, sim_label1, unit_name, ev_path, n_evento, show_hhmmss, xaxis_range=None, yaxis1_range=None, yaxis2_range=None):
        # Estandarización de comparativa de simulación (E0 vs E1) siguiendo el patrón del Bloque 3
        _tc0 = df0.columns[0]; _tc1 = df1.columns[0]
        _fc0 = [c for c in df0.columns[1:] if _is_frequency_column(c, df0[c])][0]
        _pc0 = [c for c in df0.columns[1:] if c != _fc0][0]
        _fc1 = [c for c in df1.columns[1:] if _is_frequency_column(c, df1[c])][0]
        _pc1 = [c for c in df1.columns[1:] if c != _fc1][0]

        _gcfg = st.session_state.graph_config
        
        # 1. Crear base con E0 (Simulación CNDC)
        fig_cmp = create_dual_axis_timeseries(
            t_data=df0[_tc0], freq_data=df0[_fc0], pot_data=df0[_pc0],
            title=f"Comparativa de Simulaciones — {unit_name}",
            freq_label=f"Frec. {sim_label0}", pot_label=f"Pot. {sim_label0}",
            show_hhmmss=show_hhmmss, freq_color=_gcfg["freq_color_sim0"], pot_color=_gcfg["pot_color_sim0"],
            line_width=_gcfg["line_width"], template=_gcfg["template"], height=_gcfg["plot_height"],
            x_range=xaxis_range, y1_range=yaxis1_range, y2_range=yaxis2_range
        )
        
        # 2. Añadir capas de E1 (Simulación COBEE) con estilo discontinuo (dash)
        fig_cmp.add_trace(go.Scatter(
            x=_to_plotly_time(df1[_tc1], show_hhmmss), y=df1[_fc1],
            name=f"Frec. {sim_label1}", 
            line=dict(color=_gcfg["freq_color_sim1"], width=_gcfg["line_width"], dash="dash"), 
            yaxis="y"
        ))
        fig_cmp.add_trace(go.Scatter(
            x=_to_plotly_time(df1[_tc1], show_hhmmss), y=df1[_pc1],
            name=f"Pot. {sim_label1}", 
            line=dict(color=_gcfg["pot_color_sim1"], width=_gcfg["line_width"], dash="dash"), 
            yaxis="y2"
        ))
        
        # 3. Aplicar líneas de referencia estándar (t0 y delta_t)
        _t_falla = st.session_state.b3_t_falla
        _dt = st.session_state.b3_dt
        fig_cmp = add_reference_lines(
            fig_cmp, t_fault_abs=_t_falla, t_eval_abs=_t_falla + _dt,
            show_hhmmss=show_hhmmss, show_deadband=_gcfg["show_deadband"],
            eval_line_label=f"t₀+{int(_dt)}s"
        )
        
        return fig_cmp

    def _display_sim_section(df_sim, sim_label, t_falla, delta_t, unit_name, ev_path, n_evento, show_hhmmss, xaxis_range=None, yaxis1_range=None, yaxis2_range=None):
        """Helper to display the figure and KPI table for a single simulation."""
        fig_ann, _kpi, _pm, _rocof = _create_sim_figure(df_sim, sim_label, t_falla, delta_t, unit_name, ev_path, n_evento, show_hhmmss, xaxis_range, yaxis1_range, yaxis2_range)
        st.plotly_chart(fig_ann, use_container_width=True)

        # ── Tabla KPIs CNDC ───────────────────────────────────────────────────
        if st.button(f"⬇️ Descargar datos Simulación {sim_label} a Excel", key=f"dl_sim_data_{sim_label}"):
            _sheet_n = f"Sim_{sim_label}_{os.path.splitext(unit_name)[0]}"[:31].replace(".", "_")
            excel_data = _apply_excel_formatting(
                df_sim,
                sheet_name=_sheet_n,
            )
            st.download_button(
                f"Descargar Simulación {sim_label}",
                excel_data,
                file_name=f"sim_data_{sim_label}_{os.path.splitext(unit_name)[0]}_Ev{n_evento}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.markdown("#### 📋 KPIs CNDC — Criterio RPF")
        _mostrar_tabla_cndc(_kpi, _pm, delta_t, fuente=sim_label, rocof=_rocof)

    def load_and_display_simulation_data(sim_type_suffix, sel_file):
        """Loads simulation data and displays basic info and dataframe."""
        sim_dir = os.path.join(st.session_state.ev_path_global, f"E{n_evento}.{sim_type_suffix}", CARPETA_DATOS_CURVAS) # type: ignore

        if not sel_file:
            st.info("ℹ️ Seleccione una unidad en el selector superior.")
            return None, None

        if not os.path.isdir(sim_dir):
            st.info(f"ℹ️ La carpeta '{sim_dir}' no existe. Asegúrese de haber ejecutado `DatosCurvas_v3.py` en PowerFactory.")
            return None, None

        xlsx_files = sorted([f for f in os.listdir(sim_dir) if f.endswith('.xlsx') and not f.startswith('~$')])

        if not xlsx_files:
            st.warning(f"No se encontraron archivos Excel en '{sim_dir}'.")
            return None, None

        if sel_file not in xlsx_files: # type: ignore
            st.warning(f"El archivo **{sel_file}** no existe en E{n_evento}.{sim_type_suffix}.")
            return None, None

        st.success(f"✅ {len(xlsx_files)} archivos disponibles — mostrando: **{sel_file}**")
        df_sim = pd.read_excel(os.path.join(sim_dir, sel_file), engine="calamine").dropna() # type: ignore

        time_col = df_sim.columns[0]
        data_cols = [col for col in df_sim.columns if col != time_col]

        with st.expander("📄 Ver tabla de datos"):
            st.dataframe(df_sim, use_container_width=True)
        return df_sim, sel_file

    # Pestaña 1: Simulación E{N}.0 (CNDC)
    with tab_sim_cndc: # type: ignore
        df_sim0_raw, _ = load_and_display_simulation_data("0", _sel_file_b3)

        # Opciones de ejes
        with st.expander("Opciones de Ejes"):
            if df_sim0_raw is not None:
                _auto_y = _get_unit_cfg(ev_path, _sel_unit, "y_auto", True)
                auto_scale_s0 = st.toggle("Auto-escala (Plotly)", value=_auto_y, key="b3_sim0_auto_toggle")

                col_ax1, col_ax2, col_ax3, col_ax4 = st.columns([1, 1, 1, 0.5])
                _t0_s = _parse_to_seconds(df_sim0_raw.iloc[:, 0]).dropna()
                _f0_s = pd.to_numeric(df_sim0_raw.iloc[:, 1], errors='coerce').dropna()
                _p0_s = pd.to_numeric(df_sim0_raw.iloc[:, 2], errors='coerce').dropna()

                xaxis_min = col_ax1.number_input("X Min (s)", value=_get_unit_cfg(ev_path, _sel_unit, "sim0_xmin", float(_t0_s.min()) if not _t0_s.empty else 0.0), key="b3_sim0_xmin")
                xaxis_max = col_ax1.number_input("X Max (s)", value=_get_unit_cfg(ev_path, _sel_unit, "sim0_xmax", float(_t0_s.max()) if not _t0_s.empty else 100.0), key="b3_sim0_xmax")
                yaxis1_min = col_ax2.number_input("Y1 Min (Hz)", value=_get_unit_cfg(ev_path, _sel_unit, "y_f_min", 49.0), key="b3_sim0_y1min")
                yaxis1_max = col_ax2.number_input("Y1 Max (Hz)", value=_get_unit_cfg(ev_path, _sel_unit, "y_f_max", 51.0), key="b3_sim0_y1max")
                yaxis2_min = col_ax3.number_input("Y2 Min (MW)", value=_get_unit_cfg(ev_path, _sel_unit, "y_p_min", 0.0), key="b3_sim0_y2min")
                yaxis2_max = col_ax3.number_input("Y2 Max (MW)", value=_get_unit_cfg(ev_path, _sel_unit, "y_p_max", 200.0), key="b3_sim0_y2max")

                c_btn1, c_btn2 = col_ax4.columns(2)
                if c_btn1.button("Reset", key="reset_sim0", help="Auto-detectar límites de datos y guardar"):
                    _save_unit_cfg(ev_path, _sel_unit, "sim0_xmin", float(_t0_s.min()) if not _t0_s.empty else 0.0)
                    _save_unit_cfg(ev_path, _sel_unit, "sim0_xmax", float(_t0_s.max()) if not _t0_s.empty else 100.0)
                    _save_unit_cfg(ev_path, _sel_unit, "y_f_min", 49.0)
                    _save_unit_cfg(ev_path, _sel_unit, "y_f_max", 51.0)
                    _save_unit_cfg(ev_path, _sel_unit, "y_p_min", 0.0)
                    _save_unit_cfg(ev_path, _sel_unit, "y_p_max", 200.0)
                    _save_unit_cfg(ev_path, _sel_unit, "y_auto", True) # type: ignore
                    _sync_session_scale_config(ev_path, _sel_unit)
                    st.rerun() # type: ignore

                if c_btn2.button("Guardar", key="save_scale_sim0", help="Guardar escala manual"):
                    _save_unit_cfg(ev_path, _sel_unit, "sim0_xmin", xaxis_min); _save_unit_cfg(ev_path, _sel_unit, "sim0_xmax", xaxis_max)
                    _save_unit_cfg(ev_path, _sel_unit, "y_f_min", yaxis1_min); _save_unit_cfg(ev_path, _sel_unit, "y_f_max", yaxis1_max)
                    _save_unit_cfg(ev_path, _sel_unit, "y_p_min", yaxis2_min); _save_unit_cfg(ev_path, _sel_unit, "y_p_max", yaxis2_max)
                    _save_unit_cfg(ev_path, _sel_unit, "y_auto", auto_scale_s0)
                    st.toast("Escala E0 guardada")
            else:
                st.info("Cargue los datos de simulación E0 para configurar los ejes.")

        st.subheader(f"Resultados de Simulación E{n_evento}.0 (Modelo CNDC)")
        df_sim0, file_sim0 = _load_sim_data_only("0", _sel_file_b3, n_evento, ev_path)

        if df_sim0 is not None:
            st.session_state[f'df_sim_E{n_evento}.0'] = df_sim0 # type: ignore
            st.session_state[f'file_sim_E{n_evento}.0'] = file_sim0
            _display_sim_section(df_sim0, f"E{n_evento}.0", _b3_t_falla, _b3_dt,
                            _sel_file_b3, ev_path, n_evento, show_hhmmss,
                            # Pass ranges for display
                            xaxis_range=None if auto_scale_s0 else [_to_plotly_time(xaxis_min, show_hhmmss), _to_plotly_time(xaxis_max, show_hhmmss)],
                            yaxis1_range=None if auto_scale_s0 else [yaxis1_min, yaxis1_max], yaxis2_range=None if auto_scale_s0 else [yaxis2_min, yaxis2_max])

    # Pestaña 2: Simulación E{N}.1 (COBEE)
    with tab_sim_cobee: # type: ignore
        df_sim1_raw, _ = load_and_display_simulation_data("1", _sel_file_b3)
        if df_sim1_raw is not None:
            with st.expander("Opciones de Ejes"):
                _auto_y = _get_unit_cfg(ev_path, _sel_unit, "y_auto", True)
                auto_scale_s1 = st.toggle("Auto-escala (Plotly)", value=_auto_y, key="b3_sim1_auto_toggle")

                col_ax1, col_ax2, col_ax3, col_ax4 = st.columns([1, 1, 1, 0.5])
                _t1_s = _parse_to_seconds(df_sim1_raw.iloc[:, 0]).dropna()
                _f1_s = pd.to_numeric(df_sim1_raw.iloc[:, 1], errors='coerce').dropna()
                _p1_s = pd.to_numeric(df_sim1_raw.iloc[:, 2], errors='coerce').dropna()

                xaxis_min = col_ax1.number_input("X Min (s)", value=_get_unit_cfg(ev_path, _sel_unit, "sim1_xmin", float(_t1_s.min()) if not _t1_s.empty else 0.0), key="b3_sim1_xmin")
                xaxis_max = col_ax1.number_input("X Max (s)", value=_get_unit_cfg(ev_path, _sel_unit, "sim1_xmax", float(_t1_s.max()) if not _t1_s.empty else 100.0), key="b3_sim1_xmax")
                yaxis1_min = col_ax2.number_input("Y1 Min (Hz)", value=_get_unit_cfg(ev_path, _sel_unit, "y_f_min", 49.0), key="b3_sim1_y1min")
                yaxis1_max = col_ax2.number_input("Y1 Max (Hz)", value=_get_unit_cfg(ev_path, _sel_unit, "y_f_max", 51.0), key="b3_sim1_y1max")
                yaxis2_min = col_ax3.number_input("Y2 Min (MW)", value=_get_unit_cfg(ev_path, _sel_unit, "y_p_min", 0.0), key="b3_sim1_y2min")
                yaxis2_max = col_ax3.number_input("Y2 Max (MW)", value=_get_unit_cfg(ev_path, _sel_unit, "y_p_max", 200.0), key="b3_sim1_y2max")

                c_btn1, c_btn2 = col_ax4.columns(2)
                if c_btn1.button("Reset", key="reset_sim1", help="Auto-detectar límites de datos y guardar"):
                    _save_unit_cfg(ev_path, _sel_unit, "sim1_xmin", float(_t1_s.min()) if not _t1_s.empty else 0.0)
                    _save_unit_cfg(ev_path, _sel_unit, "sim1_xmax", float(_t1_s.max()) if not _t1_s.empty else 100.0)
                    _save_unit_cfg(ev_path, _sel_unit, "y_f_min", 49.0)
                    _save_unit_cfg(ev_path, _sel_unit, "y_f_max", 51.0)
                    _save_unit_cfg(ev_path, _sel_unit, "y_p_min", 0.0)
                    _save_unit_cfg(ev_path, _sel_unit, "y_p_max", 200.0)
                    _save_unit_cfg(ev_path, _sel_unit, "y_auto", True)
                    _sync_session_scale_config(ev_path, _sel_unit)
                    st.rerun()

                if c_btn2.button("💾", key="save_scale_sim1", help="Guardar escala manual"):
                    _save_unit_cfg(ev_path, _sel_unit, "sim1_xmin", xaxis_min); _save_unit_cfg(ev_path, _sel_unit, "sim1_xmax", xaxis_max)
                    _save_unit_cfg(ev_path, _sel_unit, "y_f_min", yaxis1_min); _save_unit_cfg(ev_path, _sel_unit, "y_f_max", yaxis1_max)
                    _save_unit_cfg(ev_path, _sel_unit, "y_p_min", yaxis2_min); _save_unit_cfg(ev_path, _sel_unit, "y_p_max", yaxis2_max)
                    _save_unit_cfg(ev_path, _sel_unit, "y_auto", auto_scale_s1)
                    st.toast("Escala E1 guardada")

        st.subheader(f"Resultados de Simulación E{n_evento}.1 (Modelo COBEE Validado)") # type: ignore
        df_sim1, file_sim1 = _load_sim_data_only("1", _sel_file_b3, n_evento, ev_path)
        if df_sim1 is not None:
            st.session_state[f'df_sim_E{n_evento}.1'] = df_sim1
            st.session_state[f'file_sim_E{n_evento}.1'] = file_sim1
            _display_sim_section(df_sim1, f"E{n_evento}.1", _b3_t_falla, _b3_dt,
                            _sel_file_b3, ev_path, n_evento, show_hhmmss,
                            # Pass ranges for display
                            xaxis_range=None if auto_scale_s1 else [_to_plotly_time(xaxis_min, show_hhmmss), _to_plotly_time(xaxis_max, show_hhmmss)],
                            yaxis1_range=None if auto_scale_s1 else [yaxis1_min, yaxis1_max], yaxis2_range=None if auto_scale_s1 else [yaxis2_min, yaxis2_max])
    
    # Pestaña 3: Comparativa de Simulaciones
    with tab_sim_comp:
        st.subheader(f"Comparativa E{n_evento}.0 vs E{n_evento}.1")

        dir0 = _dir0_b3
        dir1 = _dir1_b3

        ok0, ok1 = os.path.isdir(dir0), os.path.isdir(dir1)
        if not ok0 or not ok1:
            missing = []
            if not ok0: missing.append(f"`E{n_evento}.0/Datos Curvas`")
            if not ok1: missing.append(f"`E{n_evento}.1/Datos Curvas`")
            st.warning(f"Faltan carpetas de datos: {', '.join(missing)}. Ejecute `DatosCurvas_v3.py` en PowerFactory.")
        elif not _sel_file_b3:
            st.info("ℹ️ Seleccione un archivo en el selector superior para ver la comparativa.")
        else:
            files0 = {f for f in os.listdir(dir0) if f.endswith('.xlsx') and not f.startswith('~$')}
            files1 = {f for f in os.listdir(dir1) if f.endswith('.xlsx') and not f.startswith('~$')}

            _has0 = _sel_file_b3 in files0
            _has1 = _sel_file_b3 in files1
            if not _has0 or not _has1:
                _miss = []
                if not _has0: _miss.append(f"E{n_evento}.0")
                if not _has1: _miss.append(f"E{n_evento}.1")
                st.warning(f"El archivo **{_sel_file_b3}** no existe en: {', '.join(_miss)}.")
            else:
                sel_page = _sel_file_b3

                try:
                    df0 = pd.read_excel(os.path.join(dir0, sel_page), engine="calamine")
                    df1 = pd.read_excel(os.path.join(dir1, sel_page), engine="calamine")
                except Exception as exc:
                    st.error(f"Error al cargar archivos: {exc}")
                    st.stop()

                # Opciones de ejes
                with st.expander("Opciones de Ejes"):
                    _auto_y = _get_unit_cfg(ev_path, _sel_unit, "y_auto", True)
                    auto_scale_scmp = st.toggle("Auto-escala (Plotly)", value=_auto_y, key="b3_simcomp_auto_toggle")

                    col_ax1, col_ax2, col_ax3, col_ax4 = st.columns([1, 1, 1, 0.5])
                    tc0, tc1 = df0.columns[0], df1.columns[0]
                    _t_comb = pd.concat([_parse_to_seconds(df0[tc0]), _parse_to_seconds(df1[tc1])]).dropna()
                    _f_comb = pd.concat([pd.to_numeric(df0.iloc[:, 1], errors='coerce'), pd.to_numeric(df1.iloc[:, 1], errors='coerce')]).dropna()
                    _p_comb = pd.concat([pd.to_numeric(df0.iloc[:, 2], errors='coerce'), pd.to_numeric(df1.iloc[:, 2], errors='coerce')]).dropna()

                    xaxis_min = col_ax1.number_input("X Min (s)", value=_get_unit_cfg(ev_path, _sel_unit, "simcomp_xmin", float(_t_comb.min()) if not _t_comb.empty else 0.0), key="b3_comp_xmin")
                    xaxis_max = col_ax1.number_input("X Max (s)", value=_get_unit_cfg(ev_path, _sel_unit, "simcomp_xmax", float(_t_comb.max()) if not _t_comb.empty else 100.0), key="b3_comp_xmax")
                    yaxis1_min = col_ax2.number_input("Y1 Min (Hz)", value=_get_unit_cfg(ev_path, _sel_unit, "y_f_min", 49.0), key="b3_comp_y1min")
                    yaxis1_max = col_ax2.number_input("Y1 Max (Hz)", value=_get_unit_cfg(ev_path, _sel_unit, "y_f_max", 51.0), key="b3_comp_y1max")
                    yaxis2_min = col_ax3.number_input("Y2 Min (MW)", value=_get_unit_cfg(ev_path, _sel_unit, "y_p_min", 0.0), key="b3_comp_y2min")
                    yaxis2_max = col_ax3.number_input("Y2 Max (MW)", value=_get_unit_cfg(ev_path, _sel_unit, "y_p_max", 200.0), key="b3_comp_y2max")

                    c_btn1, c_btn2 = col_ax4.columns(2)
                    if c_btn1.button("Reset", key="reset_simcomp", help="Auto-detectar límites de datos y guardar"):
                        _save_unit_cfg(ev_path, _sel_unit, "simcomp_xmin", float(_t_comb.min()) if not _t_comb.empty else 0.0)
                        _save_unit_cfg(ev_path, _sel_unit, "simcomp_xmax", float(_t_comb.max()) if not _t_comb.empty else 100.0)
                        _save_unit_cfg(ev_path, _sel_unit, "y_f_min", 49.0)
                        _save_unit_cfg(ev_path, _sel_unit, "y_f_max", 51.0)
                        _save_unit_cfg(ev_path, _sel_unit, "y_p_min", 0.0)
                        _save_unit_cfg(ev_path, _sel_unit, "y_p_max", 200.0)
                        _save_unit_cfg(ev_path, _sel_unit, "y_auto", True)
                        _sync_session_scale_config(ev_path, _sel_unit)
                        st.rerun() # type: ignore

                    if c_btn2.button("Guardar", key="save_scale_simcomp", help="Guardar escala manual"):
                        _save_unit_cfg(ev_path, _sel_unit, "simcomp_xmin", xaxis_min); _save_unit_cfg(ev_path, _sel_unit, "simcomp_xmax", xaxis_max) # type: ignore
                        _save_unit_cfg(ev_path, _sel_unit, "y_f_min", yaxis1_min); _save_unit_cfg(ev_path, _sel_unit, "y_f_max", yaxis1_max)
                        _save_unit_cfg(ev_path, _sel_unit, "y_p_min", yaxis2_min); _save_unit_cfg(ev_path, _sel_unit, "y_p_max", yaxis2_max)
                        _save_unit_cfg(ev_path, _sel_unit, "y_auto", auto_scale_scmp)
                        st.toast("Escala Comparativa Simulación guardada")

                # --- Generación del gráfico comparativo (estandarizado con Bloque 3) ---
                # Convertir columnas de datos a numérico
                for _df in (df0, df1):
                    for _c in _df.columns[1:]:
                        _df[_c] = pd.to_numeric(_df[_c], errors='coerce')

                tc0 = df0.columns[0]
                tc1 = df1.columns[0]

                # Selección automática de columnas: frecuencia vs potencia
                dc0 = [c for c in df0.columns if c != tc0]
                dc1 = [c for c in df1.columns if c != tc1]
                _fc0 = [c for c in dc0 if _is_frequency_column(c, df0[c])]
                _fc1 = [c for c in dc1 if _is_frequency_column(c, df1[c])]
                _fc0 = _fc0[:1] if _fc0 else dc0[:1]
                _fc1 = _fc1[:1] if _fc1 else dc1[:1]
                _pc0 = [c for c in dc0 if c not in _fc0][:1]
                _pc1 = [c for c in dc1 if c not in _fc1][:1]

                # Para compatibilidad con el flujo previo: eliminar referencias a `is_multi`.
                # Bloque 4 comparativa estandarizada: siempre usa la doble-curva (frecuencia + potencia) con referencias.
                _gcfg = st.session_state.graph_config

                if not _pc0 or not _pc1 or not _fc0 or not _fc1:
                    st.warning("No se pudieron identificar columnas de frecuencia/potencia para la comparativa.")
                else:
                    # Ranges para estandarización: si auto-escala, pasamos None
                    _x_range = None if auto_scale_scmp else [_to_plotly_time(xaxis_min, show_hhmmss), _to_plotly_time(xaxis_max, show_hhmmss)]
                    _y1_range = None if auto_scale_scmp else [yaxis1_min, yaxis1_max]
                    _y2_range = None if auto_scale_scmp else [yaxis2_min, yaxis2_max]

                    fig_cmp = create_dual_axis_timeseries(

                        t_data=df0[tc0],
                        freq_data=df0[_fc0[0]],
                        pot_data=df0[_pc0[0]],
                        title=f"Comparativa de Simulaciones — {os.path.splitext(sel_page)[0]}",
                        freq_label=f"Frec. E{n_evento}.0",
                        pot_label=f"Pot. E{n_evento}.0",
                        show_hhmmss=show_hhmmss,
                        freq_color=_gcfg["freq_color_sim0"],
                        pot_color=_gcfg["pot_color_sim0"],
                        line_width=_gcfg["line_width"],
                        template=_gcfg["template"],
                        height=_gcfg["plot_height"],
                        x_range=_x_range,
                        y1_range=_y1_range,
                        y2_range=_y2_range,
                    )

                    # Capa E1 (dash)
                    fig_cmp.add_trace(go.Scatter(
                        x=_to_plotly_time(df1[tc1], show_hhmmss),
                        y=df1[_fc1[0]],
                        name=f"Frec. E{n_evento}.1",
                        line=dict(color=_gcfg["freq_color_sim1"], width=_gcfg["line_width"], dash="dash"),
                        yaxis="y",
                    ))
                    fig_cmp.add_trace(go.Scatter(
                        x=_to_plotly_time(df1[tc1], show_hhmmss),
                        y=df1[_pc1[0]],
                        name=f"Pot. E{n_evento}.1",
                        line=dict(color=_gcfg["pot_color_sim1"], width=_gcfg["line_width"], dash="dash"),
                        yaxis="y2",
                    ))

                    # Referencias CNDC (mismos marcadores que Bloque 3)
                    _t_falla = st.session_state.b3_t_falla
                    _dt = st.session_state.b3_dt
                    fig_cmp = add_reference_lines(
                        fig_cmp,
                        t_fault_abs=_t_falla,
                        t_eval_abs=_t_falla + _dt,
                        show_hhmmss=show_hhmmss,
                        show_deadband=_gcfg["show_deadband"],
                        show_fault_line=_gcfg.get("show_fault_line", True),
                        show_eval_line=_gcfg.get("show_eval_line", True),
                        eval_line_label=f"t₀+{int(_dt)}s",
                    )

                    st.plotly_chart(fig_cmp, use_container_width=True)

                if st.button(f"Descargar datos Comparativa Simulación a Excel", key=f"dl_sim_comp_data_{_sel_file_b3}"): # type: ignore
                    excel_data = _apply_excel_formatting(
                        df0.merge(df1, on=df0.columns[0], suffixes=('_E0', '_E1')), # Simple merge for export
                        sheet_name=f"CompSim_{os.path.splitext(_sel_file_b3)[0]}",
                    )
                    st.download_button(
                        f"Descargar Comparativa Simulación {_sel_file_b3}",
                        excel_data,
                        file_name=f"sim_comp_data_{os.path.splitext(_sel_file_b3)[0]}_Ev{n_evento}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                # ── Análisis de desviación ────────────────────────────────────
                with st.expander("Análisis de desviación"):
                    if is_multi:
                        pares = list(zip(sel0, sel1))
                        if not pares:
                            st.info("Seleccione curvas en ambas simulaciones para calcular desviaciones.")
                        else:
                            rows = []
                            for c0, c1 in pares:
                                if c0 not in df0.columns or c1 not in df1.columns:
                                    continue
                                s0 = pd.to_numeric(df0[c0], errors='coerce').dropna()
                                s1 = pd.to_numeric(df1[c1], errors='coerce').dropna()
                                if s0.empty or s1.empty:
                                    continue
                                rows.append({
                                    "Curva E{}.0".format(n_evento): _short_col_name(c0),
                                    "Curva E{}.1".format(n_evento): _short_col_name(c1),
                                    "Media E.0": f"{s0.mean():.4f}",
                                    "Media E.1": f"{s1.mean():.4f}",
                                    "Dif. media": f"{abs(s0.mean() - s1.mean()):.4f}",
                                    "Dif. máx.": f"{abs(s0.max() - s1.max()):.4f}",
                                    "Dif. mín.": f"{abs(s0.min() - s1.min()):.4f}",
                                })
                            if rows:
                                st.dataframe(pd.DataFrame(rows), use_container_width=True)
                    else:
                        # Emparejar por tipo (freq↔freq, potencia↔potencia)
                        # para no depender del orden de columnas entre proyectos.
                        def _clasificar(df, cols):
                            freq, other = [], []
                            for c in cols:
                                num = pd.to_numeric(df[c], errors='coerce').dropna()
                                (freq if _is_frequency_column(c, num) else other).append(c)
                            return freq, other

                        freq0, other0 = _clasificar(df0, dc0)
                        freq1, other1 = _clasificar(df1, dc1)

                        rows = []
                        for pares, unidad in [(zip(freq0, freq1), "Hz"),
                                              (zip(other0, other1), "MW")]:
                            for col0, col1 in pares:
                                s0 = pd.to_numeric(df0[col0], errors='coerce').dropna()
                                s1 = pd.to_numeric(df1[col1], errors='coerce').dropna()
                                if s0.empty or s1.empty:
                                    continue
                                rows.append({
                                    "Variable": _short_col_name(col0),
                                    "Unidad": unidad,
                                    "Media E.0": f"{s0.mean():.4f}",
                                    "Media E.1": f"{s1.mean():.4f}",
                                    "Dif. media": f"{abs(s0.mean() - s1.mean()):.4f}",
                                    "Dif. máx.": f"{abs(s0.max() - s1.max()):.4f}",
                                    "Dif. mín.": f"{abs(s0.min() - s1.min()):.4f}",
                                })
                        if rows:
                            st.dataframe(pd.DataFrame(rows), use_container_width=True)

                # ── Comparativa KPIs CNDC entre E0 y E1 ──────────────────────
                st.markdown("---") # type: ignore
                st.markdown("#### 📋 Comparativa KPIs CNDC — E{}.0 vs E{}.1".format(n_evento, n_evento))
                _pmax_cargado_b3c = _load_pmax_cargado(ev_path, n_evento)
                _tech_b3c         = _load_tech_map(LOC_NAMES_GEN_PATH)
                _pm_b3c_val, _tk_b3c, _pm_b3c_fuente = _get_pmax_from_cargado(
                    sel_page, _pmax_cargado_b3c, _tech_b3c
                )
                if _pm_b3c_fuente:
                    st.caption(f"✅ Unidad: **{_tk_b3c}** — P_max desde `{_pm_b3c_fuente}`")

                _pm_b3c = float(_pm_b3c_val) # type: ignore
                _rp_pct_b3c = float(_get_rp_default(_tk_b3c, LOC_NAMES_GEN_PATH))
                _rp_b3c = _rp_pct_b3c / 100.0

                _kpi_comp_rows = []
                for _sfx, _dfc in [("0", df0), ("1", df1)]:
                    _tcx   = _dfc.columns[0]
                    _t_x   = pd.to_numeric(_dfc[_tcx], errors='coerce').values
                    _dcx   = [c for c in _dfc.columns if c != _tcx]
                    _fcx   = [c for c in _dcx if _is_frequency_column(c, pd.to_numeric(_dfc[c], errors='coerce').dropna())] # type: ignore
                    _pcx   = [c for c in _dcx if c not in _fcx]
                    if not _fcx or not _pcx:
                        continue
                    _fq_x  = pd.to_numeric(_dfc[_fcx[0]], errors='coerce').ffill().bfill().values
                    _pt_x  = pd.to_numeric(_dfc[_pcx[0]], errors='coerce').ffill().bfill().values
                    _t_al_x = _t_x - _b3_t_falla
                    _kx     = _cndc_kpis(_t_al_x, _fq_x, _pt_x, _pm_b3c, _rp_b3c, _b3_dt)
                    _roc_x  = _calcular_rocof(_t_al_x, _fq_x, 3.0)
                    if _kx:
                        _kpi_comp_rows.append({'Fuente': f"E{n_evento}.{_sfx}", **_kx,
                                               'rocof': _roc_x})

                if len(_kpi_comp_rows) == 2:
                    _param_labels_c = {
                        'f0': "f₀ [Hz]", 'p0': "P₀ [MW]",
                        'f_min': "f_min [Hz]", 't_min': "t_min [s]",
                        'delta_f': "Δf [Hz]",
                        'f_dt': f"f_Δt ({_b3_dt} s) [Hz]", 'p_dt': f"P_Δt ({_b3_dt} s) [MW]",
                        'dp': "ΔP [MW]", 'dp_pct': "ΔP [%]",
                        'aporta': "Aporta", 'droop_nom': "Droop nom. [%]",
                        'droop_calc': "Droop calc. [%]", 'rocof': "ROCOF [Hz/s]",
                    }
                    _r0c, _r1c = _kpi_comp_rows[0], _kpi_comp_rows[1]
                    _tbl_c = {'Parámetro': [], _r0c['Fuente']: [], _r1c['Fuente']: [], 'Δ (E1−E0)': []}
                    for _pk, _pl in _param_labels_c.items():
                        _v0 = _r0c.get(_pk, '—')
                        _v1 = _r1c.get(_pk, '—')
                        _tbl_c['Parámetro'].append(_pl)
                        if _pk == 'aporta':
                            _tbl_c[_r0c['Fuente']].append("✅ Sí" if _v0 else "❌ No")
                            _tbl_c[_r1c['Fuente']].append("✅ Sí" if _v1 else "❌ No") # type: ignore
                            _tbl_c['Δ (E1−E0)'].append('—')
                        else:
                            try:
                                _dv = round(float(_v1) - float(_v0), 4) if _v0 != '—' and _v1 != '—' else '—'
                            except (TypeError, ValueError):
                                _dv = '—'
                            _tbl_c[_r0c['Fuente']].append(f"{_v0:.4f}" if isinstance(_v0, float) else str(_v0))
                            _tbl_c[_r1c['Fuente']].append(f"{_v1:.4f}" if isinstance(_v1, float) else str(_v1))
                            _tbl_c['Δ (E1−E0)'].append(f"{_dv:.4f}" if isinstance(_dv, float) else str(_dv))
                    _df_kpi_c = _df_safe(pd.DataFrame(_tbl_c))

                    def _col_aporta_c(row):
                        styles = [''] * len(row)
                        if row['Parámetro'] == "Aporta":
                            for _i in range(1, len(row)):
                                styles[_i] = 'background-color:#d4edda' if '✅' in str(row.iloc[_i]) else 'background-color:#f8d7da'
                        return styles

                    st.dataframe(_df_kpi_c.style.apply(_col_aporta_c, axis=1), use_container_width=True, hide_index=True) # type: ignore

                    if st.button(f"⬇️ Descargar KPIs Comparativa Simulación a Excel", key=f"dl_kpis_sim_comp"):
                        excel_data = _apply_excel_formatting(
                            _df_kpi_c,
                            sheet_name=f"KPIs_CompSim",
                            kpi_col="Parámetro",
                            kpi_ok_val="✅ Sí",
                            kpi_error_val="❌ No"
                        )
                        st.download_button(f"Descargar KPIs Comparativa Simulación", excel_data,
                                           file_name=f"kpis_sim_comp_Ev{n_evento}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                elif _kpi_comp_rows:
                    st.info("Solo hay datos de una versión de simulación.")
                else:
                    st.warning("No se pudieron calcular KPIs — verifique columnas de frecuencia/potencia.")
    
    # ── Exportación masiva Bloque 4 ──────────────────────────────────────────
    st.markdown("---")
    st.subheader("📥 Exportar todos los gráficos de Simulación")
    st.caption("Genera capturas PNG de las simulaciones E.0 y E.1 para todas las unidades disponibles.")
    b4_incluir_kpi_zip = st.checkbox("Incluir tablas KPI al lado de las gráficas (Reporte Combinado)", value=False, key="b4_incluir_kpi_zip")

    if st.button("🗂️ Generar ZIP de gráficos de simulación (E.0 / E.1)", key="btn_zip_b4"):
        import io, zipfile, datetime
        from plotly.io import to_image

        _zip_buf = io.BytesIO()
        _n_ok = 0
        _available = get_event_units(ev_path, n_evento)
        _pmax_map_exp = _load_pmax_cargado(ev_path, n_evento)
        _tmap_exp = _load_tech_map(LOC_NAMES_GEN_PATH)
        _prog = st.progress(0, text="Iniciando exportación de simulaciones...")
        _gcfg = st.session_state.graph_config # type: ignore

        # Resetear descarga previa
        st.session_state.b4_sim_zip_bytes = None

        with zipfile.ZipFile(_zip_buf, 'w', zipfile.ZIP_DEFLATED) as _zf:
            for _idx, _uname in enumerate(_available):
                _prog.progress((_idx + 1) / len(_available), text=f"Procesando {_uname}...")
                
                # Intentar exportar E.0 y E.1
                for _sfx in ["0", "1"]:
                    _s_dir = os.path.join(ev_path, f"E{n_evento}.{_sfx}", CARPETA_DATOS_CURVAS)
                    if not os.path.isdir(_s_dir): continue

                    _match = glob.glob(os.path.join(_s_dir, f"*{_uname}*.xlsx"))
                    if not _match: _match = glob.glob(os.path.join(_s_dir, f"*{_uname.replace('sym_', '')}*.xlsx"))
                    
                    if _match:
                        try:
                            _df_s = pd.read_excel(_match[0], engine="calamine").dropna()
                            _sim_lbl = f"E{n_evento}.{_sfx}"
                            
                            # Usar los límites guardados y respetar y_auto
                            _y_auto = _get_unit_cfg(ev_path, _uname, "y_auto", True)
                            _xmin = _get_unit_cfg(ev_path, _uname, f"sim{_sfx}_xmin", None)
                            _xmax = _get_unit_cfg(ev_path, _uname, f"sim{_sfx}_xmax", None)
                            _y1min = _get_unit_cfg(ev_path, _uname, "y_f_min", None)
                            _y1max = _get_unit_cfg(ev_path, _uname, "y_f_max", None)
                            _y2min = _get_unit_cfg(ev_path, _uname, "y_p_min", None)
                            _y2max = _get_unit_cfg(ev_path, _uname, "y_p_max", None)
                            
                            # Generar figura usando la función existente (forzando segundos para el PNG) # type: ignore
                            _fig_data = _create_sim_figure(
                                _df_s, _sim_lbl, _b3_t_falla, _b3_dt, _uname, ev_path, n_evento, 
                                show_hhmmss=False,
                                xaxis_range=[_xmin, _xmax] if _xmin is not None else None,
                                yaxis1_range=None if _y_auto else [_y1min, _y1max],
                                yaxis2_range=None if _y_auto else [_y2min, _y2max]
                            )
                            
                            if _fig_data: # type: ignore
                                _fig, _, _, _ = _fig_data
                                _img = to_image(_fig, format="png", width=1000, height=800, scale=2)
                                _zf.writestr(f"SIM_{_sfx}_{_uname}_Ev{n_evento}.png", _img)
                                _n_ok += 1
                        except: pass

                # --- EXPORTAR COMPARATIVA DE SIMULACIONES ---
                _dir0 = os.path.join(ev_path, f"E{n_evento}.0", CARPETA_DATOS_CURVAS)
                _dir1 = os.path.join(ev_path, f"E{n_evento}.1", CARPETA_DATOS_CURVAS)

                if os.path.isdir(_dir0) and os.path.isdir(_dir1):
                    _m0 = glob.glob(os.path.join(_dir0, f"*{_uname}*.xlsx"))
                    _m1 = glob.glob(os.path.join(_dir1, f"*{_uname}*.xlsx"))
                    if _m0 and _m1:
                        try:
                            _df0 = pd.read_excel(_m0[0], engine="calamine").dropna()
                            _df1 = pd.read_excel(_m1[0], engine="calamine").dropna()
                            
                            _y_auto_c = _get_unit_cfg(ev_path, _uname, "y_auto", True)
                            _xmin = _get_unit_cfg(ev_path, _uname, "simcomp_xmin", None)
                            _xmax = _get_unit_cfg(ev_path, _uname, "simcomp_xmax", None)
                            _y1min = _get_unit_cfg(ev_path, _uname, "y_f_min", None)
                            _y1max = _get_unit_cfg(ev_path, _uname, "y_f_max", None)
                            _y2min = _get_unit_cfg(ev_path, _uname, "y_p_min", None)
                            _y2max = _get_unit_cfg(ev_path, _uname, "y_p_max", None)

                            _fig_c = _create_comp_figure(
                                _df0, _df1, f"E{n_evento}.0", f"E{n_evento}.1", _uname, ev_path, n_evento, 
                                show_hhmmss=False,
                                xaxis_range=[_xmin, _xmax] if _xmin is not None else None,
                                yaxis1_range=None if _y_auto_c else [_y1min, _y1max],
                                yaxis2_range=None if _y_auto_c else [_y2min, _y2max]
                            )
                            
                            _img_c = to_image(_fig_c, format="png", width=1000, height=800, scale=2)
                            _zf.writestr(f"COMP_SIM_{_uname}_Ev{n_evento}.png", _img_c)
                            _n_ok += 1
                        except: pass

        _prog.empty()
        if _n_ok > 0:
            _ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            st.session_state.b4_sim_zip_bytes = _zip_buf.getvalue()
            st.session_state.b4_sim_zip_name = f"graficos_simulacion_Ev{n_evento}_{_ts}.zip"
            st.success(f"✅ Se generaron {_n_ok} gráficos de simulación.")
        else:
            st.error("No se pudieron generar imágenes de simulación. Verifique los archivos en 'Datos Curvas'.")

    if st.session_state.get("b4_sim_zip_bytes"):
        st.download_button(
            label=f"⬇️ Descargar ZIP de Simulaciones",
            data=st.session_state.b4_sim_zip_bytes,
            file_name=st.session_state.b4_sim_zip_name,
            mime="application/zip",
            type="primary"
        )


    elif bloque_trabajo == "comparativa_real_simu":
        st.header("⚖️ Bloque 5: Validación Real vs. Simulación (Criterios CNDC)")
    st.info(
        "Compara registros reales SCADA con simulaciones RMS. "
        "Las curvas se alinean automáticamente usando el **inicio de la falla** como t = 0."
    )

    if not st.session_state.semestre_global or not st.session_state.evento_global:
        st.warning("👈 Seleccione Semestre y Evento en la barra lateral.")
        st.stop()

    ev_path = st.session_state.ev_path_global
    n_evento = st.session_state.n_evento_global

    _event_cfg = _load_event_cfg(ev_path)

    # ── Funciones auxiliares ────────────────────────────────────────────────────
    def _local_parse_sec(series):
        def _to_sec(val):
            if pd.isna(val): return 0.0
            s = str(val).strip()
            if ':' in s:
                parts = s.split(':')
                try: return int(parts[0]) * 3600 + int(parts[1]) * 60 + (int(parts[2]) if len(parts) > 2 else 0)
                except: pass
            try: return float(s.replace(',', '.'))
            except: return 0.0
        return series.apply(_to_sec)

    # ── Panel de configuración ─────────────────────────────────────────────────
    st.markdown("### Configuración")
    col_src, col_sim, col_tech = st.columns([1, 1, 1])
    with col_src:
        src_real = st.radio("Fuente Real:", ["SCADA COBEE (1SEG)", "EMF CNDC"], key="b4_src_real")
    with col_sim:
        src_sim = st.multiselect("Simulaciones:", [f"E{n_evento}.0", f"E{n_evento}.1"], default=[f"E{n_evento}.0", f"E{n_evento}.1"])

    with col_tech:
        try:
            df_tech = pd.read_excel(LOC_NAMES_GEN_PATH, sheet_name="Detalle_PF", engine="calamine")
            _pcol = 'P_max (MW)' if 'P_max (MW)' in df_tech.columns else 'P nom. (MW)'
            tech_map = (df_tech.set_index('loc_name PF')[[_pcol]]
                        .rename(columns={_pcol: 'P_max (MW)'})
                        .to_dict('index'))
            st.success(f"✅ P_max cargada ({len(tech_map)} unidades desde '{_pcol}').")
        except Exception as _e:
            tech_map = {}
            st.warning(f"⚠️ No se pudo cargar loc_names_gen.xlsx: {_e}")

    st.markdown("---") # type: ignore

    # ── Parámetros de alineación temporal y visualización ───────────────────── # type: ignore
    st.markdown("### Alineación Temporal y Visualización")
    col_p1, col_p2, col_p3, col_p4, col_p5 = st.columns(5)
    with col_p1:
        _event_cfg = _load_event_cfg(ev_path)
        t_sim_falla = st.number_input(
            "Tiempo de falla en simulación (s)",
            value=_event_cfg.get("t_sim_falla", 5.0), min_value=0.0, max_value=300.0, step=0.5,
            help="Instante t en la simulación RMS donde ocurre el evento (PowerFactory).",
            key="b4_t_sim_falla",
        )
        if st.button("Guardar", key="save_t_sim_falla", help="Guardar tiempo de falla de simulación"):
            if _save_event_cfg(ev_path, "t_sim_falla", t_sim_falla):
                st.toast("Tiempo de falla guardado.")

    with col_p2:
        umbral_dfdt = st.number_input(
            "Umbral detección df/dt (Hz/s)",
            value=-0.04, min_value=-2.0, max_value=-0.001, step=0.005, format="%.3f",
            help="Caída sostenida de frecuencia (suavizada) para detectar inicio de falla.",
            key="b4_umbral_dfdt",
        )
    with col_p3:
        ventana_suavizado = st.number_input(
            "Ventana suavizado detección (muestras)",
            value=5, min_value=2, max_value=20, step=1,
            help="Número de muestras para rolling mean antes de calcular df/dt.",
            key="b4_ventana_suav",
        )
    with col_p4:
        ventana_pre = st.number_input(
            "Ventana pre-falla a mostrar (s)",
            value=10, min_value=0, max_value=120, step=5,
            help="Segundos antes de t=0 a incluir en la gráfica.",
            key="b4_ventana_pre",
        )
    with col_p5:
        delta_t_cndc = st.number_input(
            "Δt CNDC (s)",
            value=35, min_value=20, max_value=60, step=1,
            help="Tiempo desde t₀ para leer f_Δt y P_Δt. CNDC usa entre 30 y 50 s (típicamente 35 s).",
            key="b4_delta_t_cndc",
        )

    st.markdown("---") # type: ignore
    st.markdown("### 📈 Opciones de Gráfico")

    # Opciones de ejes sincronizados
    with st.expander("🛠️ Opciones de Ejes de Validación"):
        auto_v = st.toggle(
            "Auto-escala",
            value=st.session_state.b3_sync_y_auto,
            key="b4_val_auto",
            on_change=None,
        )
        c1, c2, c3, c4 = st.columns([1, 1, 1, 0.5])
        _xmin_v = c1.number_input("X Min", value=-10.0, key="b4_val_xmin")
        _xmax_v = c1.number_input("X Max", value=100.0, key="b4_val_xmax")
        _y1min_v = c2.number_input("Y1 Min", value=st.session_state.b3_sync_y_f_min, key="b4_val_y1min", on_change=_sync_rpf_y_axis, args=("y_f_min", "b4_val_y1min"))
        _y1max_v = c2.number_input("Y1 Max", value=st.session_state.b3_sync_y_f_max, key="b4_val_y1max", on_change=_sync_rpf_y_axis, args=("y_f_max", "b4_val_y1max"))
        _y2min_v = c3.number_input("Y2 Min", value=st.session_state.b3_sync_y_p_min, key="b4_val_y2min", on_change=_sync_rpf_y_axis, args=("y_p_min", "b4_val_y2min"))
        _y2max_v = c3.number_input("Y2 Max", value=st.session_state.b3_sync_y_p_max, key="b4_val_y2max", on_change=_sync_rpf_y_axis, args=("y_p_max", "b4_val_y2max"))

    st.markdown("---")

    # ── Carga y Alineación Real ────────────────────────────────────────────────
    _sel_unit = st.session_state.global_selected_unit
    real_subdir = "Graficas Registro 1SEG COBEE" if "SCADA" in src_real else "Resultados_COBEE"
    _r_dir = os.path.join(ev_path, real_subdir)
    
    if _sel_unit and os.path.isdir(_r_dir):
        _rf_match = _buscar_archivo_unidad(_sel_unit, os.listdir(_r_dir))
        if _rf_match:
            df_r = pd.read_excel(os.path.join(_r_dir, _rf_match), engine="calamine").dropna()
            tc_r = df_r.columns[0]
            tr_raw = _parse_to_seconds(df_r[tc_r])
            tr_norm = tr_raw - tr_raw.min()
            
            # Identificar frecuencia real dinámicamente
            _fr_c = [c for c in df_r.columns if any(kw in c.lower() for kw in ['frec', 'hz'])]
            _fr_col = _fr_c[0] if _fr_c else df_r.columns[1]
            _pr_col = df_r.columns[2]
            
            _fr_arr = pd.to_numeric(df_r[_fr_col], errors='coerce').ffill().values
            _pr_arr = pd.to_numeric(df_r[_pr_col], errors='coerce').ffill().values
            
            # Detección Falla Real
            idx_f_r = _detectar_inicio_falla(_fr_arr, umbral_dfdt, int(ventana_suavizado))
            t_f_r = float(tr_norm.iloc[idx_f_r])
            tr_aligned = (tr_norm - t_f_r).values

            # --- Construcción del Gráfico de Validación ---
            _gcfg = st.session_state.graph_config
            fig = create_dual_axis_timeseries(
                t_data=tr_aligned, freq_data=_fr_arr, pot_data=_pr_arr,
                title=f"Validación Real vs Simulación — {_sel_unit}",
                freq_label=f"Frec. Real ({src_real})", pot_label=f"Pot. Real ({src_real})",
                freq_color=_gcfg["freq_color_real"], pot_color=_gcfg["pot_color_real"],
                show_hhmmss=False,
                x_range=None if auto_v else [_xmin_v, _xmax_v], # type: ignore
                y1_range=None if auto_v else [_y1min_v, _y1max_v], # type: ignore
                y2_range=None if auto_v else [_y2min_v, _y2max_v] # type: ignore
            )
            
            # Marcadores Reales
            _pmax_map_v = _load_pmax_cargado(ev_path, n_evento)
            _pm_v, _tk, _ = _get_pmax_from_cargado(_sel_unit, _pmax_map_v, _load_tech_map(LOC_NAMES_GEN_PATH))
            _rp_v = _get_rp_default(_tk, LOC_NAMES_GEN_PATH) / 100.0
            _kr = _cndc_kpis(tr_aligned, _fr_arr, _pr_arr, _pm_v, _rp_v, delta_t_cndc)
            fig = add_kpi_markers(fig, t_fault_abs=0.0, kpi_dict=_kr, show_hhmmss=False, dt_seconds=delta_t_cndc)
            fig = add_reference_lines(fig, t_fault_abs=0.0, t_eval_abs=delta_t_cndc, show_hhmmss=False)

            # ── Carga y Alineación Simulación ──────────────────────────────────
            _kpi_rows = [{'Fuente': 'REAL', **_kr}] if _kr else []
            _sim_for_error = []

            for s_ver in src_sim:
                _s_dir = os.path.join(ev_path, s_ver, CARPETA_DATOS_CURVAS)
                if not os.path.isdir(_s_dir): continue
                _sf_match = _buscar_archivo_unidad(_sel_unit, os.listdir(_s_dir))
                if _sf_match:
                    df_s = pd.read_excel(os.path.join(_s_dir, _sf_match), engine="calamine").dropna()
                    tc_s, fc_s, pc_s = _robust_col_detect(df_s)
                    
                    ts_raw = pd.to_numeric(df_s[tc_s], errors='coerce').values
                    ts_aligned = ts_raw - t_sim_falla
                    
                    fs_raw = pd.to_numeric(df_s[fc_s], errors='coerce').ffill().values
                    fs_hz = fs_raw * 50.0 if np.nanmax(fs_raw) < 2.0 else fs_raw
                    ps_mw = pd.to_numeric(df_s[pc_s], errors='coerce').ffill().values
                    
                    _color_f = _gcfg["freq_color_sim0"] if "0" in s_ver else _gcfg["freq_color_sim1"]
                    _color_p = _gcfg["pot_color_sim0"] if "0" in s_ver else _gcfg["pot_color_sim1"]
                    
                    fig.add_trace(go.Scatter(x=ts_aligned, y=fs_hz, name=f"Frec. {s_ver}", line=dict(color=_color_f, dash="dash", width=2), yaxis="y"))
                    fig.add_trace(go.Scatter(x=ts_aligned, y=ps_mw, name=f"Pot. {s_ver}", line=dict(color=_color_p, dash="dash", width=2), yaxis="y2"))
                    
                    _ks = _cndc_kpis(ts_aligned, fs_hz, ps_mw, _pm_v, _rp_v, delta_t_cndc)
                    if _ks: _kpi_rows.append({'Fuente': s_ver, **_ks})
                    _sim_for_error.append({'ver': s_ver, 't': ts_aligned, 'f': fs_hz, 'color': _color_f})

            st.plotly_chart(fig, use_container_width=True)
            
            # --- Tablas de Métricas (Igual a Bloque 3) ---
            if _kpi_rows:
                st.markdown("#### 📋 Comparativa de Desempeño CNDC")
                _df_kpi_v = _df_safe(pd.DataFrame(_kpi_rows))
                # Arrow/pyarrow falla si alguna columna object mezcla tipos (ej. int/bytes/str).
                # Forzamos columnas 'object' a string para evitar el error.
                for _c in _df_kpi_v.columns:
                    if _df_kpi_v[_c].dtype == object:
                        _df_kpi_v[_c] = _df_kpi_v[_c].astype(str)
                st.dataframe(_df_kpi_v, hide_index=True, use_container_width=True)


                # --- Curva de Error y Barras KPI ---
                if _sim_for_error:
                    ce_col, bar_col = st.columns([1, 1])
                    with ce_col:
                        fig_err = go.Figure()
                        for s in _sim_for_error:
                            # Interpolar real sobre simu para calcular error instantáneo
                            f_real_interp = np.interp(s['t'], tr_aligned, _fr_arr)
                            err = f_real_interp - s['f']
                            fig_err.add_trace(go.Scatter(x=s['t'], y=err, name=f"Err {s['ver']}", line=dict(color=s['color'])))
                        fig_err.update_layout(title="Error de Seguimiento (Hz)", height=350, template=_gcfg["template"])
                        st.plotly_chart(fig_err, use_container_width=True)
                    
                    with bar_col:
                        fig_bar = go.Figure()
                        _fuentes = [r['Fuente'] for r in _kpi_rows]
                        _dps = [r['dp_pct'] for r in _kpi_rows]
                        fig_bar.add_trace(go.Bar(x=_fuentes, y=_dps, marker_color='#2ca02c', text=_dps, textposition='outside'))
                        fig_bar.add_hline(y=1.5, line_dash="dash", line_color="red", annotation_text="Mínimo 1.5%")
                        fig_bar.update_layout(title="Aporte Porcentual ΔP (%)", height=350, template=_gcfg["template"])
                        st.plotly_chart(fig_bar, use_container_width=True)
else:
        st.info("ℹ️ Seleccione una unidad y verifique los archivos en las carpetas correspondientes.")

if bloque_trabajo == "reporte_tecnico":
    st.header("📝 Bloque 6: Gestión de Reporte y Auditoría")
    st.header("📝 Bloque 6: Gestión de Reporte y Auditoría")
    
    st.subheader("⚙️ Configuración del Reporte")
    if not st.session_state.semestre_global or not st.session_state.evento_global:
        st.warning("👈 Seleccione Semestre y Evento en la barra lateral para ver la auditoría del proyecto.")
        st.stop()

    ev_path = st.session_state.ev_path_global
    n_evento = st.session_state.n_evento_global

    # ─── SECCIÓN 1: AUDITORÍA DE ARCHIVOS DEL EVENTO ───────────────────────────
    st.subheader("Auditoría de Archivos del Evento")
    st.caption("Estado actual de los archivos generados y requeridos para el evento seleccionado.")
    
    def _check_file(path_glob):
        hits = glob.glob(path_glob)
        return (True, os.path.basename(hits[0])) if hits else (False, "Faltante")

    audit_data = [
        {"Bloque": "1. Carga", "Concepto": "Datos Simulación (Extracción)", "Estado": _check_file(os.path.join(ev_path, "datos_simulacion_*_2daopcion.xlsx"))},
        {"Bloque": "1. Carga", "Concepto": "Condiciones Iniciales", "Estado": _check_file(os.path.join(ev_path, "condiciones_iniciales_*.xlsx"))}, # type: ignore
        {"Bloque": "1. Carga", "Concepto": "Resultados Carga PowerFactory", "Estado": _check_file(os.path.join(ev_path, f"datos_cargados_Ev{n_evento}.xlsx"))}, # type: ignore
        {"Bloque": "2. Real", "Concepto": "Registros SCADA (1 SEG)", "Estado": (os.path.isdir(os.path.join(ev_path, "Graficas Registro 1SEG COBEE")), "Carpeta de Unidades")}, # type: ignore
        {"Bloque": "2. Real", "Concepto": "Digitalización EMF CNDC", "Estado": (os.path.isdir(os.path.join(ev_path, "Resultados_COBEE")), "Carpeta de Unidades")}, # type: ignore
        {"Bloque": "3. Simu", "Concepto": "Curvas Simulación CNDC (E.0)", "Estado": (os.path.isdir(os.path.join(ev_path, f"E{n_evento}.0", "Datos Curvas")), "Disponible")},
        {"Bloque": "3. Simu", "Concepto": "Curvas Simulación COBEE (E.1)", "Estado": (os.path.isdir(os.path.join(ev_path, f"E{n_evento}.1", "Datos Curvas")), "Disponible")},
    ]

    df_audit = pd.DataFrame([
        {
            "Bloque": d["Bloque"],
            "Documento/Proceso": d["Concepto"],
            "Estado": "✅" if d["Estado"][0] else "❌",
            "Detalle": d["Estado"][1]
        } for d in audit_data
    ])
    st.table(_df_safe(df_audit))

    if st.button("⬇️ Exportar Auditoría a Excel"):
        _df_audit_exp = df_audit.copy()
        _df_audit_exp.columns = ["Bloque", "Documento/Proceso", "Estado (Icono)", "Detalle"]
        excel_audit = _apply_excel_formatting(
            _df_audit_exp,
            sheet_name="Auditoria_Proyecto",
            kpi_col="Estado (Icono)",
            kpi_ok_val="✅",
            kpi_error_val="❌"
        )
        st.download_button("Descargar Reporte de Auditoría", excel_audit,
                           file_name=f"auditoria_Ev{n_evento}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # ─── SECCIÓN 2: DOCUMENTACIÓN DE MEMORIA Y PROCESOS ────────────────────────
    st.markdown("---") # type: ignore
    st.subheader("📚 Manual Técnico y Memoria del Sistema")

    with st.expander("🧠 Archivos de Memoria y Contexto", expanded=True):
        st.markdown(r"""
        La interfaz mantiene la continuidad del trabajo mediante archivos JSON que actúan como la 'memoria' del sistema:

        | Archivo | Función | Ubicación |
        | :--- | :--- | :--- |
        | `config_rutas.json` | Almacena las rutas raíz (CNDC, PowerFactory, Mapeos) definidas en el Sidebar. | Directorio del script principal |
        | `estatismo_config.json` | Guarda los valores de Estatismo (Rp) ajustados manualmente para cada unidad generadora. | `...\Designacion de loc_name\` |
        | `_streamlit_params.json` | Puente de parámetros entre la UI y los scripts de ejecución no interactiva. | Carpeta del Evento |
        """)

    with st.expander(" Bloque 0: Datos del Modelo (Gestión de Mapeo)", expanded=False):
        st.markdown(r"""
        **Objetivo:** Sincronizar los catálogos técnicos con el modelo `.pfd` de PowerFactory.
        
        | Script | Entrada | Ubicación Entrada | Salida | Ubicación Salida |
        | :--- | :--- | :--- | :--- | :--- |
        | `DatsoGENBUSLNE.py` | Red PowerFactory | Proyecto PF Activo | `DatosSINdigsilent.xlsx` | `...\DATOS EXTRAIDOS DE DIGSILENT\` |
        | `loc_namesGEN.py` | `DatosSINdigsilent.xlsx`, Despacho CNDC | `...\DATOS EXTRAIDOS...`, `RAIZ_DATOS` | `loc_names_gen.xlsx` | `...\Designacion de loc_name\` |
        | `loc_namesLineas.py` | `DatosSINdigsilent.xlsx` | `...\DATOS EXTRAIDOS...` | `loc_names_lineas.xlsx` | `...\Designacion de loc_name\` |
        | `loc_names_xfo.py` | `DatosSINdigsilent.xlsx` | `...\DATOS EXTRAIDOS...` | `loc_names_xfo.xlsx` | `...\Designacion de loc_name\` |
        | `MapeoRetirosSTI.py`| `deener_*.xlsx`, `DatosSINdigsilent.xlsx` | Carpeta Evento / `...\DATOS...` | `loc_name_cargas.xlsx` | `...\Designacion de loc_name\` |
        """)

    with st.expander("📦 Bloque 1: Carga de Datos y PowerFactory", expanded=False):
        st.markdown(fr"""
        **Objetivo:** Definir el despacho y demanda (Snapshot) para la simulación RMS.

        | Script | Entrada | Ubicación Entrada | Salida | Ubicación Salida |
        | :--- | :--- | :--- | :--- | :--- |
        | `ExtFLujos2daO.py` | `dc_*, dcdr_*, deener_*, Tabla_Eventos` | Carpeta Evento / `RAIZ` | `datos_simulacion_*.xlsx` | Carpeta del Evento |
        | `CondInicialesPF.py`| `datos_simulacion_*.xlsx`, Mapeos (GEN/CAR/XFO) | Carpeta Evento / `...\Designacion...` | `condiciones_iniciales_*.xlsx` | Carpeta del Evento |
        | `CargaCondIniciales_PF.py` | `condiciones_iniciales_*.xlsx`, `loc_names_xfo.xlsx` | Carpeta Evento / `...\Designacion...` | `datos_cargados_Ev{n_evento}.xlsx` | Carpeta del Evento |
        """)

    with st.expander("📊 Bloque 2: Análisis de Datos Registrados", expanded=False):
        st.markdown(r"""
        **Objetivo:** Procesar registros de campo (SCADA) y gráficas oficiales CNDC.

        | Script | Entrada | Ubicación Entrada | Salida | Ubicación Salida |
        | :--- | :--- | :--- | :--- | :--- |
        | `OrdenadorDatosEvento.py` | `1 seg.*.xls` (Falla) | `RAIZ_DATOS` / Año / Carpeta FALLA | `Unidad.xlsx` (Excel 1seg) | `...\Graficas Registro 1SEG COBEE\` |
        | `ExtractorResultadosCNDC.py`| Archivos `*.emf` del CNDC | Carpeta del Evento | `Unidad.xlsx` (Digitalizado) | `...\Resultados_COBEE\` |
        """)

    with st.expander("📈 Bloque 3: Análisis de Simulación", expanded=False):
        st.markdown(r"""
        **Objetivo:** Exportar los resultados de las simulaciones RMS de PowerFactory.

        | Script | Ejecución | Salida | Ubicación Salida |
        | :--- | :--- | :--- | :--- |
        | `DatosCurvas_v3.py` | Script Python en PowerFactory | Archivos `.xlsx` por página de gráficos | `...\E{N}.x\Datos Curvas\` |
        """)

    with st.expander("⚖️ Bloque 4: Validación y Comparativa", expanded=False):
        st.markdown("""
        **Integración de Datos:**
        
        | Comparación | Fuente A | Fuente B | t = 0 (Referencia) |
        | :--- | :--- | :--- | :--- |
        | **Alineación** | Registro Real (SCADA/EMF) | Simulación (PowerFactory) | Instante de falla (df/dt < Umbral) |

        **Salidas:**
        *   **KPIs:** f₀, f_min, ΔP, Droop calculado, ROCOF, RMSE.
        *   **Exportación:** Imágenes `.png` individuales y archivos `.zip` consolidados por evento.
        """)

    # ─── SECCIÓN 3: RESUMEN DE RUTAS ──────────────────────────────────────────
    st.markdown("---")
    st.subheader("⚙️ Rutas de Memoria Activas")
    
    c_r1, c_r2 = st.columns(2)
    with c_r1:
        st.write("**📁 Almacenamiento CNDC:**")
        st.code(f"RAIZ RPF: {RAIZ}\nRAIZ DATOS: {RAIZ_DATOS}")
        st.write("**⚙️ PowerFactory:**")
        st.code(f"Proyecto: {PF_PROYECTO}\nCaso Base: {CASO_BASE}")
    
    with c_r2:
        st.write("**📄 Archivos de Mapeo:**")
        st.caption(f"Generadores: `{os.path.basename(LOC_NAMES_GEN_PATH)}`")
        st.caption(f"Cargas: `{os.path.basename(LOC_CAR_PATH)}`")
        st.caption(f"Trafos: `{os.path.basename(LOC_XFO_PATH)}`")

    # ─── PRÓXIMAMENTE ──────────────────────────────────────────────────────────
    st.markdown("---") # type: ignore
    col_rep1, col_rep2 = st.columns([2,1])
    with col_rep1:
        st.subheader(" Generador de Reporte Final")
        st.write("Consolida los KPIs de todas las unidades y las gráficas comparativas en un documento PDF/Word.")
    with col_rep2:
        st.write("")
        st.button("📄 Próximamente: Exportar PDF", disabled=True)

    # Botón para abrir la carpeta del evento (Solo Windows)
    if st.button("📂 Abrir Carpeta del Evento en Explorador"):
        if os.path.isdir(ev_path):
            os.startfile(ev_path)
        else:
            st.error("La ruta del evento no es válida.")
