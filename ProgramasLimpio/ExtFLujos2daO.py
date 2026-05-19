# =============================================================================
# EXTRACCION DE DATOS CNDC PARA SIMULACION EN POWERFACTORY - 2DA OPCION
#1

# Lee Tabla_Eventos_*.xlsx   -> seleccion de evento, fecha/hora exacta del Po
# Lee dc_(fecha).xls         -> potencia activa unidades hidro (hoja HIDRO)
# Lee dcdr_(fecha).xls       -> potencia activa generadores no-hidro (hoja POST)
# Lee deener_(fecha).xlsx    -> demanda por nodo, todas las horas
# Lee tabla_resul1tados_*.xlsx -> P_0 [MW] por unidad (potencia inicial real)
#
# Salida: datos_simulacion_{fecha}_2daopcion.xlsx  en la carpeta del evento
# =============================================================================

import os, glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RAIZ = r"C:\Datos del CNDC\01_INFO CNDC_RPF"
LOC_NAMES_GEN_PATH = (r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT"
                      r"\Designacion de loc_name\loc_names_gen.xlsx")

CARPETAS_RESULTADOS = [
    "Resultados_COBEE",
    "Resultados_ENDE_Andina",
    "Resultados_GUABIRA",
    "Resultados_SCADA",
    "Resultados_ENDE_Corani",
    "Resultados_ENDE_Guaracachi",
    "Resultados_AGUAI",
    "Resultados_HB",
    "Resultados_ENDE_Valle_Hermoso",
]

# =============================================================================
# SELECCION INTERACTIVA
# =============================================================================

def elegir_idx(n, titulo):
    print(titulo)
    while True:
        try:
            idx = int(input("Selecciona numero: ")) - 1
            if 0 <= idx < n:
                return idx
        except ValueError:
            pass
        print("  Opcion invalida, intenta de nuevo.")


# 1. Semestre
semestres = sorted([d for d in os.listdir(RAIZ)
                    if os.path.isdir(os.path.join(RAIZ, d))])
print("\nSelecciona el semestre:")
for i, s in enumerate(semestres, 1):
    print(f"  {i}. {s}")
semestre = semestres[elegir_idx(len(semestres), "")]

# =============================================================================
# LEER TABLA DE EVENTOS DEL SEMESTRE
# =============================================================================

tabla_glob = glob.glob(os.path.join(RAIZ, semestre, "Tabla_Eventos_*.xlsx"))
if not tabla_glob:
    raise FileNotFoundError(
        f"No se encontro Tabla_Eventos_*.xlsx en:\n  {os.path.join(RAIZ, semestre)}")
tabla_path = tabla_glob[0]

import openpyxl as _opx
_wb = _opx.load_workbook(tabla_path, data_only=True)
_sh = _wb.active

# Leer todos los eventos (filas desde la 3ra, saltando titulo y encabezado)
eventos_tabla = []
for fila in _sh.iter_rows(min_row=3, values_only=True):
    if fila[0] is None:
        continue
    num         = int(fila[0])
    fecha_hora  = str(fila[1]).strip()   # "DD/MM/YYYY HH:MM"
    desconexion = str(fila[2]).strip() if fila[2] else ""
    pot_desc    = fila[3]
    demanda     = fila[4]
    f0          = fila[5]
    fmin        = fila[6]
    eventos_tabla.append({
        "num"        : num,
        "fecha_hora" : fecha_hora,
        "desconexion": desconexion,
        "pot_desc_MW": pot_desc,
        "demanda_MW" : demanda,
        "f0_Hz"      : f0,
        "fmin_Hz"    : fmin,
    })

# 2. Mostrar tabla de eventos para seleccion
print(f"\nEventos del semestre '{semestre}'  [{os.path.basename(tabla_path)}]")
print(f"  {'N°':>3}  {'Fecha y hora':<20}  {'Disparo':<35}  {'P[MW]':>7}  {'f0[Hz]':>7}  {'fmin[Hz]':>8}")
print(f"  {'-'*3}  {'-'*20}  {'-'*35}  {'-'*7}  {'-'*7}  {'-'*8}")
for ev in eventos_tabla:
    print(f"  {ev['num']:>3}  {ev['fecha_hora']:<20}  {ev['desconexion']:<35}  "
          f"{str(ev['pot_desc_MW']):>7}  {str(ev['f0_Hz']):>7}  {str(ev['fmin_Hz']):>8}")

sel_idx = elegir_idx(len(eventos_tabla), "\nSelecciona numero de evento:")
ev_sel  = eventos_tabla[sel_idx]

# Extraer fecha y hora del evento
partes      = ev_sel["fecha_hora"].split(" ")   # ["DD/MM/YYYY", "HH:MM"]
fecha_ddmmyyyy = partes[0]                      # "DD/MM/YYYY"
hora_evento    = partes[1] if len(partes) > 1 else "00:00"  # "HH:MM"

# Hora del Po = hora en punto del evento (ej. 06:21 -> "06:00", 1:58 -> "01:00")
hora_po = f"{int(hora_evento.split(':')[0]):02d}:00"

# Construir fecha_str en formato DDMMYY para buscar dcdr/deener
d, m, y = fecha_ddmmyyyy.split("/")
fecha_str = f"{d}{m}{y[2:]}"   # "290125"
fecha_fmt = f"{d}/{m}/{y}"

print(f"\n  Evento seleccionado : #{ev_sel['num']}  {fecha_fmt}  {hora_evento}")
print(f"  Disparo             : {ev_sel['desconexion']}")
print(f"  Potencia desconec.  : {ev_sel['pot_desc_MW']} MW")
print(f"  Demanda SIN         : {ev_sel['demanda_MW']} MW")
print(f"  f0 / fmin           : {ev_sel['f0_Hz']} Hz / {ev_sel['fmin_Hz']} Hz")
print(f"  Hora Po (inicio)    : {hora_po}")

# Carpeta del evento
eventos_raiz = os.path.join(RAIZ, semestre, "Análisis_todos_los_eventos")
evento_path  = os.path.join(eventos_raiz, f"Evento {ev_sel['num']}")
if not os.path.isdir(evento_path):
    raise FileNotFoundError(f"No se encontro la carpeta: {evento_path}")

# =============================================================================
# LOCALIZAR ARCHIVOS DCDR Y DEENER
# =============================================================================

def buscar_archivo(carpeta, patron):
    resultados = glob.glob(os.path.join(carpeta, patron))
    if not resultados:
        raise FileNotFoundError(f"No se encontro '{patron}' en:\n  {carpeta}")
    if len(resultados) > 1:
        print(f"  [AVISO] Multiples archivos, se usa: {os.path.basename(resultados[0])}")
    return resultados[0]

dc_path     = buscar_archivo(os.path.join(evento_path, "Despacho"),       "dc_*.xls*")
dcdr_path   = buscar_archivo(os.path.join(evento_path, "Despacho"),       "dcdr_*.xls*")
deener_path = buscar_archivo(os.path.join(evento_path, "Demanda de Energia y Potencia"), "deener_*.xlsx")

print(f"\n  DC     : {os.path.basename(dc_path)}")
print(f"  DCDR   : {os.path.basename(dcdr_path)}")
print(f"  DEENER : {os.path.basename(deener_path)}")

# =============================================================================
# LECTURA DC HIDRO — unidades hidroelectricas individuales, todas las horas
# =============================================================================

def leer_dc_hidro_todas_horas(path):
    """Lee la hoja HIDRO del archivo dc_ y devuelve unidades hidro individuales."""
    import xlrd
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_name('HIDRO')

    # Fila 8 (indice 8) contiene HORA y fracciones de dia
    cols_hora = {}
    for c in range(1, sh.ncols):
        v = sh.cell_value(8, c)
        if isinstance(v, float) and 0 < v <= 1:
            h = int(round(v * 24))
            if 1 <= h <= 24:
                cols_hora[c] = f"{h:02d}:00"

    horas_ord = [cols_hora[c] for c in sorted(cols_hora)]
    print(f"    DC_HIDRO -> {len(cols_hora)} horas: {horas_ord[0]} .. {horas_ord[-1]}")

    # Saltar encabezados de sistema (SISTEMA ...), totales (TOTAL ...) y notas
    ignorar_pref = ('SISTEMA', 'TOTAL', 'DESPACHO', 'Los valores', 'HORA',
                    'Comite', 'RF:', 'ND:')

    filas = []
    for r in range(9, sh.nrows):
        nombre = str(sh.cell_value(r, 0)).strip()
        if not nombre:
            continue
        if any(nombre.startswith(p) for p in ignorar_pref):
            continue
        fila = {"Generador_CNDC": nombre}
        en_mantenimiento = False
        for c, etiq in sorted(cols_hora.items()):
            val = sh.cell_value(r, c)
            if isinstance(val, str) and val.strip().upper() == 'M':
                fila[etiq] = 0.0
                en_mantenimiento = True
            else:
                fila[etiq] = round(float(val), 4) if isinstance(val, (int, float)) else None
        fila["_mant"] = en_mantenimiento
        filas.append(fila)

    return pd.DataFrame(filas)


# =============================================================================
# LECTURA DCDR — generadores NO-HIDRO (eolico, solar, termo), todas las horas
# =============================================================================

def leer_dcdr_todas_horas(path):
    """Lee la hoja POST del archivo dcdr_, saltando la seccion hidro."""
    import xlrd
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_name('POST')

    cols_hora = {}
    for c in range(1, sh.ncols):
        v = sh.cell_value(6, c)
        if isinstance(v, float) and v > 0:
            h = int(round(v * 24))
            if 1 <= h <= 24:
                cols_hora[c] = f"{h:02d}:00"

    horas_ord = [cols_hora[c] for c in sorted(cols_hora)]
    print(f"    DCDR   -> {len(cols_hora)} horas: {horas_ord[0]} .. {horas_ord[-1]}")

    # Encontrar la fila de SUBTOTAL HIDRO para saltarse toda la seccion hidro
    subtotal_hidro_row = None
    for r in range(8, sh.nrows):
        if str(sh.cell_value(r, 0)).strip() == 'SUBTOTAL HIDRO':
            subtotal_hidro_row = r
            break

    ignorar = {
        'SUBTOTAL HIDRO', 'SUBTOTAL EOLICO', 'SUBTOTAL SOLAR',
        'SUBTOTAL TERMO', 'SUBTOTAL EXCEDENTES', 'TOTAL',
        'RESERVA ROTANTE', 'RESERVA PARADA', 'SEGURIDAD DE AREAS',
    }
    ignorar_pref = ('RF:', 'ND:', 'Los valores')

    filas = []
    for r in range(8, sh.nrows):
        # Saltar toda la seccion hidro (hasta e incluyendo SUBTOTAL HIDRO)
        if subtotal_hidro_row is not None and r <= subtotal_hidro_row:
            continue
        nombre = str(sh.cell_value(r, 0)).strip()
        if not nombre or nombre in ignorar:
            continue
        if any(nombre.startswith(p) for p in ignorar_pref):
            continue
        fila = {"Generador_CNDC": nombre}
        en_mantenimiento = False
        for c, etiq in sorted(cols_hora.items()):
            val = sh.cell_value(r, c)
            if isinstance(val, str) and val.strip().upper() == 'M':
                fila[etiq] = 0.0
                en_mantenimiento = True
            else:
                fila[etiq] = round(float(val), 4) if isinstance(val, (int, float)) else None
        fila["_mant"] = en_mantenimiento
        filas.append(fila)

    return pd.DataFrame(filas)


# =============================================================================
# LECTURA DEENER — todas las horas
# =============================================================================

def leer_deener_todas_horas(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    sh = wb['Demanda']

    header = list(sh.iter_rows(min_row=7, max_row=7, values_only=True))[0]

    cols_hora = {}
    for i, v in enumerate(header):
        if v is not None and str(v).strip().count(':') == 1:
            etiq = str(v).strip()
            try:
                int(etiq[:2]); int(etiq[3:])
                cols_hora[i] = etiq
            except ValueError:
                pass

    horas_ord = [cols_hora[i] for i in sorted(cols_hora)]
    print(f"    DEENER -> {len(cols_hora)} horas: {horas_ord[0]} .. {horas_ord[-1]}")

    ignorar_pref = ('TOTAL', 'Los valores', 'RETIROS')

    filas = []
    for row in sh.iter_rows(min_row=8, values_only=True):
        nombre = str(row[0]).strip() if row[0] is not None else ''
        if not nombre or any(nombre.startswith(p) for p in ignorar_pref):
            continue
        fila = {"Nodo_CNDC": nombre}
        for i, etiq in sorted(cols_hora.items()):
            val = row[i]
            try:
                fila[etiq] = round(float(val), 4) if val is not None else None
            except (ValueError, TypeError):
                fila[etiq] = None
        filas.append(fila)

    return pd.DataFrame(filas)


# =============================================================================
# LECTURA P_0 DE ARCHIVOS tabla_resultados_*.xlsx
# =============================================================================

def leer_po_resultados(evento_path):
    import openpyxl

    filas_po       = []
    encontradas    = []
    no_encontradas = []

    for carpeta in CARPETAS_RESULTADOS:
        carpeta_path = os.path.join(evento_path, carpeta)
        if not os.path.isdir(carpeta_path):
            no_encontradas.append(carpeta)
            continue

        nombre_archivo = f"tabla_resultados_{carpeta.replace('Resultados_', '')}.xlsx"
        archivo_path   = os.path.join(carpeta_path, nombre_archivo)

        if not os.path.isfile(archivo_path):
            candidatos = glob.glob(os.path.join(carpeta_path, "tabla_resultados_*.xlsx"))
            if candidatos:
                archivo_path   = candidatos[0]
                nombre_archivo = os.path.basename(archivo_path)
            else:
                print(f"    [AVISO] Sin Excel en: {carpeta}")
                no_encontradas.append(carpeta)
                continue

        try:
            wb = openpyxl.load_workbook(archivo_path, data_only=True)
            sh = wb.active

            encabezados = list(sh.iter_rows(min_row=1, max_row=1, values_only=True))[0]

            fila_po = None
            for fila in sh.iter_rows(min_row=2, values_only=True):
                etiq = str(fila[0]).strip() if fila[0] else ''
                if 'P_0' in etiq or 'P0' in etiq.replace(' ', ''):
                    fila_po = fila
                    break

            if fila_po is None:
                print(f"    [AVISO] No se encontro fila P_0 en: {nombre_archivo}")
                no_encontradas.append(carpeta)
                continue

            n_unidades = 0
            for col_idx, unidad in enumerate(encabezados):
                if col_idx == 0 or unidad is None:
                    continue
                val = fila_po[col_idx]
                try:
                    po = round(float(val), 4) if val is not None else None
                except (ValueError, TypeError):
                    po = None
                filas_po.append({
                    "Unidad" : str(unidad).strip(),
                    "P_0_MW" : po,
                    "Fuente" : carpeta,
                })
                n_unidades += 1

            encontradas.append(carpeta)
            print(f"    OK  {carpeta:<38} -> {n_unidades} unidades")

        except Exception as e:
            print(f"    [ERROR] {carpeta}: {e}")
            no_encontradas.append(carpeta)

    print(f"\n    Carpetas leidas   : {len(encontradas)}")
    if no_encontradas:
        print(f"    Carpetas ausentes : {no_encontradas}")

    return pd.DataFrame(filas_po) if filas_po else pd.DataFrame(
        columns=["Unidad", "P_0_MW", "Fuente"])


# =============================================================================
# CODIGOS STI Y LIMPIEZA DE NOMBRES CNDC
# =============================================================================

import re as _re


def leer_codigos_sti_gen(path):
    """
    Lee Mapeo_Generadores de loc_names_gen.xlsx.
    Extrae el codigo STI del primer loc_name PF de cada unidad.
    Retorna {Generador_CNDC_limpio: codigo_sti}.
    En loc_names_gen los nombres ya estan limpios (sin CC, sin RF).
    """
    if not os.path.isfile(path):
        print(f"  [AVISO] No se encontro loc_names_gen.xlsx en:\n  {path}")
        return {}
    df = pd.read_excel(path, sheet_name="Mapeo_Generadores")
    resultado = {}
    for _, row in df.iterrows():
        cndc = str(row["Generador_CNDC"]).strip()
        locs = str(row["loc_names PF"]).strip()
        if not locs or locs == "-":
            resultado[cndc] = "-"
            continue
        first = locs.split(",")[0].strip()
        s = _re.sub(r'^(sym_|WT_|PV-|PV_|sta_)', '', first, flags=_re.IGNORECASE)
        s = _re.sub(r'_EQ$', '', s, flags=_re.IGNORECASE)
        s = _re.sub(r'\(\d+\)$', '', s).strip()
        resultado[cndc] = s if s else "-"
    return resultado


def _limpiar_display(nombre):
    """Quita el prefijo CC del nombre para mostrar. Conserva ' - RF'."""
    return _re.sub(r'^CC([A-Z])', r'\1', str(nombre).strip())


def _limpiar_lookup(nombre):
    """
    Limpia el nombre para buscar en el catalogo (igual a limpiar_codigo_cndc
    de loc_namesGEN): quita CC y quita ' - RF' / ' - PPG'.
    """
    s = _re.sub(r'^CC([A-Z])', r'\1', str(nombre).strip())
    return _re.sub(r'\s*-\s*(RF|PPG)\.?\s*$', '', s, flags=_re.IGNORECASE).strip()


# =============================================================================
# INSERCION DE COLUMNA CON HORA EXACTA DEL EVENTO
# Inserta la columna hora_evento (ej. '06:21') entre las columnas adyacentes.
# Usa P_0 de tabla_resultados; si una unidad no tiene P_0, usa el valor de
# la hora anterior al evento (hora en punto = hora_po).
# =============================================================================

def insertar_columna_evento(df, col_nombre_id, hora_evento, dict_po):
    hora_parte = str(hora_evento).strip().split(':', 1)[0]
    hora_h      = int(hora_parte)
    hora_ant    = f"{hora_h:02d}:00"
    hora_sig    = f"{(hora_h % 24) + 1:02d}:00"

    cols = list(df.columns)
    if hora_ant in cols:
        pos = cols.index(hora_ant) + 1
    elif hora_sig in cols:
        pos = cols.index(hora_sig)
    else:
        pos = len(cols)

    valores = []
    for _, row in df.iterrows():
        nombre = str(row[col_nombre_id]).strip()
        po = dict_po.get(nombre)
        if po is None:
            po = row.get(hora_ant)   # fallback: hora anterior al evento
        valores.append(po)

    df.insert(pos, hora_evento, valores)
    return df


# =============================================================================
# FORMATO SIMPLE PARA HOJAS EXCEL
# =============================================================================
_THIN = Side(border_style="thin", color="BFBFBF")
_BRD  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HFIL = PatternFill("solid", start_color="1F3864")
_HFNT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_NFNT = Font(name="Arial", size=10)
_CTR  = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left",   vertical="center")


def _aplicar_formato(ws):
    """Formato simple: encabezado azul, bordes, anchos automaticos, freeze fila 1."""
    mc = ws.max_column
    ws.row_dimensions[1].height = 22
    for c in range(1, mc + 1):
        cell = ws.cell(1, c)
        cell.fill = _HFIL; cell.font = _HFNT
        cell.alignment = _CTR; cell.border = _BRD
    for r in range(2, ws.max_row + 1):
        for c in range(1, mc + 1):
            cell = ws.cell(r, c)
            cell.font = _NFNT; cell.border = _BRD
            cell.alignment = _CTR if isinstance(cell.value, (int, float)) else _LEFT
    for col in ws.columns:
        ancho = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ancho + 2, 10), 35)
    ws.freeze_panes = "A2"


# =============================================================================
# LEER Y EXPORTAR
# =============================================================================

print(f"\n[1/5] Leyendo DC HIDRO (unidades hidroelectricas individuales)...")
df_hidro = leer_dc_hidro_todas_horas(dc_path)
print(f"    {len(df_hidro)} unidades hidroelectricas")

print(f"\n[2/5] Leyendo DCDR (generadores no-hidro: eolico, solar, termo)...")
df_gen_no_hidro = leer_dcdr_todas_horas(dcdr_path)
print(f"    {len(df_gen_no_hidro)} generadores (eolico, solar, termo)")

# Combinar: primero hidro individuales, luego el resto
df_gen = pd.concat([df_hidro, df_gen_no_hidro], ignore_index=True)
print(f"    Total combinado: {len(df_gen)} unidades generadoras")

print(f"\n[3/5] Leyendo DEENER (todas las horas)...")
df_carga = leer_deener_todas_horas(deener_path)
print(f"    {len(df_carga)} nodos de demanda")

print(f"\n[4/5] Leyendo P_0 de archivos de resultados...")
df_po = leer_po_resultados(evento_path)
print(f"    {len(df_po)} registros de P_0 leidos")

# Cargar mapa STI antes del lookup de P_0 (necesario como puente de nombres)
print(f"    Cargando codigos STI desde loc_names_gen.xlsx...")
codigos_sti_map = leer_codigos_sti_gen(LOC_NAMES_GEN_PATH)

# Diccionario P_0 indexado por nombre CNDC original, usando el codigo STI como puente.
# Normalizacion: codigo STI sin digitos finales (TIQ, ZON) -> anadir "01" (TIQ01, ZON01)
# para cubrir las distintas formas en que tabla_resultados registra el nombre.
dict_po_gen = {}
if not df_po.empty:
    dict_po_raw = {str(k).strip(): v
                   for k, v in zip(df_po["Unidad"], df_po["P_0_MW"])
                   if v is not None}
    # TIQ -> TIQ01, ZON -> ZON01, etc. (solo si el nombre aun no tiene digitos al final)
    dict_po_norm = dict(dict_po_raw)
    for k, v in dict_po_raw.items():
        if not _re.search(r'\d+$', k):
            dict_po_norm.setdefault(k + "01", v)

    for nombre_cndc in df_gen["Generador_CNDC"]:
        orig    = str(nombre_cndc).strip()
        cleaned = _limpiar_lookup(orig)
        sti     = codigos_sti_map.get(cleaned)
        # intento 1: nombre CNDC tal cual (ej. "BOT01")
        po = dict_po_norm.get(orig)
        # intento 2: nombre CNDC limpiado (sin CC, sin RF)
        if po is None:
            po = dict_po_norm.get(cleaned)
        # intento 3: codigo STI como puente (TIQUIMANI -> TIQ01)
        if po is None and sti:
            po = dict_po_norm.get(sti)
        if po is not None:
            dict_po_gen[orig] = po

    # Diagnostico: mostrar cuales unidades se resolvieron via puente STI
    print(f"\n  P_0 resueltos ({len(dict_po_gen)} unidades):")
    print(f"    {'Unidad CNDC':<22} {'STI key':<10} {'P_0 [MW]':>9}")
    print(f"    {'-'*22} {'-'*10} {'-'*9}")
    for nombre_cndc in df_gen["Generador_CNDC"]:
        orig    = str(nombre_cndc).strip()
        cleaned = _limpiar_lookup(orig)
        sti     = codigos_sti_map.get(cleaned, "-")
        po      = dict_po_gen.get(orig)
        if po is not None:
            via = "STI" if (dict_po_norm.get(orig) is None and dict_po_norm.get(cleaned) is None) else "dir"
            print(f"    {orig:<22} {sti:<10} {po:>9.3f}  [{via}]")
        else:
            print(f"    {orig:<22} {sti:<10}  (sin P_0 -> fallback hora anterior)")

# Insertar columna hora exacta del evento; si no hay P_0, cae al valor de hora anterior
df_gen   = insertar_columna_evento(df_gen,   "Generador_CNDC", hora_evento, dict_po_gen)
df_carga = insertar_columna_evento(df_carga, "Nodo_CNDC",      hora_evento, {})
n_con_po   = sum(1 for n in df_gen["Generador_CNDC"] if str(n).strip() in dict_po_gen)
n_fallback = len(df_gen) - n_con_po
print(f"\n  Columna '{hora_evento}': {n_con_po} unidades con P_0 real, "
      f"{n_fallback} con valor de hora {hora_po}")

# Limpiar nombres CNDC para display: quita prefijo CC, conserva ' - RF'
df_gen["Generador_CNDC"] = df_gen["Generador_CNDC"].apply(_limpiar_display)

# Agregar Codigo STI como segunda columna
df_gen.insert(1, "Codigo STI",
              df_gen["Generador_CNDC"].apply(
                  lambda n: codigos_sti_map.get(_limpiar_lookup(n), "-")))
n_mapeados = (df_gen["Codigo STI"] != "-").sum()
print(f"    {n_mapeados} de {len(df_gen)} unidades con Codigo STI asignado")

# Columna Estado: "Mantenimiento" si la unidad mostro 'M' en alguna hora del despacho
df_gen.insert(2, "Estado",
              df_gen["_mant"].apply(lambda x: "Mantenimiento" if x else ""))
df_gen.drop(columns=["_mant"], inplace=True)
n_mant = (df_gen["Estado"] == "Mantenimiento").sum()
if n_mant:
    print(f"    {n_mant} unidades en mantenimiento: "
          + ", ".join(df_gen[df_gen["Estado"] == "Mantenimiento"]["Generador_CNDC"].tolist()))

# Hoja de metadatos del evento
df_evento = pd.DataFrame([
    {"Campo": "Semestre",            "Valor": semestre},
    {"Campo": "Evento N°",           "Valor": ev_sel["num"]},
    {"Campo": "Fecha y hora",        "Valor": ev_sel["fecha_hora"]},
    {"Campo": "Hora Po (inicio)",    "Valor": hora_po},
    {"Campo": "Disparo",             "Valor": ev_sel["desconexion"]},
    {"Campo": "Potencia desc. [MW]", "Valor": ev_sel["pot_desc_MW"]},
    {"Campo": "Demanda SIN [MW]",    "Valor": ev_sel["demanda_MW"]},
    {"Campo": "f0 [Hz]",             "Valor": ev_sel["f0_Hz"]},
    {"Campo": "fmin [Hz]",           "Valor": ev_sel["fmin_Hz"]},
])

# Conjunto de nombres display que tienen P_0 real (para colorear en Excel)
nombres_con_po = {_limpiar_display(k) for k in dict_po_gen}

# Tabla P0_inicial: unidad (display), P_0_MW y fuente, sin duplicados
if not df_po.empty:
    df_po_out = df_po.copy()
    df_po_out["Unidad"] = df_po_out["Unidad"].apply(_limpiar_display)
    # Una fila por unidad (primera aparicion con P_0 no nulo)
    df_po_out = (df_po_out.dropna(subset=["P_0_MW"])
                 .drop_duplicates(subset="Unidad", keep="first")
                 .reset_index(drop=True))
else:
    df_po_out = pd.DataFrame(columns=["Unidad", "P_0_MW", "Fuente"])

output_path = os.path.join(evento_path, f"datos_simulacion_{fecha_str}_2daopcion.xlsx")

print(f"\n[5/5] Exportando a Excel...")
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df_evento.to_excel(writer, sheet_name='Info_Evento',       index=False)
    df_gen.to_excel(writer,    sheet_name='Generadores_pgini', index=False)
    df_carga.to_excel(writer,  sheet_name='Cargas_plini',      index=False)
    df_po_out.to_excel(writer, sheet_name='P0_inicial',        index=False)

wb = load_workbook(output_path)

# Formato general para todas las hojas
for nombre_hoja in wb.sheetnames:
    _aplicar_formato(wb[nombre_hoja])

# Color diferencial en Generadores_pgini:
#   gris   (D0D0D0) = unidad en mantenimiento (despacho con "M" en alguna hora)
#   verde  (C6EFCE) = unidad con P_0 real de tabla_resultados
#   amarillo (FFEB9C) = unidad sin P_0 (usa valor de hora anterior)
_FILL_MANT = PatternFill("solid", start_color="D0D0D0", end_color="D0D0D0")
_FILL_PO   = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
_FILL_FAL  = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")

# Conjunto de nombres en mantenimiento (ya con nombres display, col Generador_CNDC limpiada)
nombres_en_mant = set(df_gen[df_gen["Estado"] == "Mantenimiento"]["Generador_CNDC"])

ws_gen = wb["Generadores_pgini"]
mc = ws_gen.max_column
for r in range(2, ws_gen.max_row + 1):
    nombre_celda = str(ws_gen.cell(r, 1).value).strip()
    if nombre_celda in nombres_en_mant:
        fill = _FILL_MANT
    elif nombre_celda in nombres_con_po:
        fill = _FILL_PO
    else:
        fill = _FILL_FAL
    for c in range(1, mc + 1):
        ws_gen.cell(r, c).fill = fill

wb.save(output_path)

print(f"\n  Archivo creado en:")
print(f"  {output_path}")
print(f"\n  Hojas:")
print(f"    Info_Evento       -> datos del evento (fecha, disparo, f0, fmin)")
print(f"    Generadores_pgini -> {len(df_gen)} unidades x {len(df_gen.columns)-3} horas"
      f"  |  verde={n_con_po}, amarillo={n_fallback}, gris={n_mant} (mant.)")
print(f"      (hidro: {len(df_hidro)}, no-hidro: {len(df_gen_no_hidro)})")
print(f"    Cargas_plini      -> {len(df_carga)} nodos x {len(df_carga.columns)-1} horas")
print(f"    P0_inicial        -> {len(df_po_out)} unidades con P_0 real registrado")

input("\nPresiona Enter para cerrar...")
