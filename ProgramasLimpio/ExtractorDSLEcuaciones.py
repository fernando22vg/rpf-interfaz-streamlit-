#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
ExtractorDSLEcuaciones.py — v3
-------------------------------
Los modelos Reivax/Andritz en PF usan bloques GRAFICOS (BlkRef), no texto DSL.
Las "ecuaciones" son parametros numericos de bloques estandar (K, 1/(1+sT),
Limits, TDPU, Lookup, etc.) conectados en IntGrfnet.

MODO EXTRAER_PARAMS (default):
  Lee los parametros numericos (K, T, limites, tabla) de cada BlkRef en todos
  los generadores objetivo y los exporta a Excel con comparacion entre unidades.

MODO DIAGNOSTICO:
  Explora la estructura de sub-bloques de una unidad representativa.

Ejecucion: script interno de PowerFactory (Run as Script)
"""

import os
from datetime import datetime
import powerfactory as pf

# ══════════════════════════════════════════════════════════════
# Configuracion
# ══════════════════════════════════════════════════════════════

# 'EXTRAER_PARAMS' → extrae parametros de cada BlkRef para todas las unidades
# 'DIAGNOSTICO'   → inspecciona estructura de sub-bloques (una unidad)
MODO = 'EXTRAER_PARAMS'

TARGET_LOC_NAMES = {
    'sym_BOT01', 'sym_BOT02', 'sym_BOT03',
    'sym_CUT01', 'sym_CUT02', 'sym_CUT03', 'sym_CUT04',
    'sym_ANG03', 'sym_CRB01',
}

DSL_SLOTS = (
    'ATUADOR_SIMP', 'POS_SIMP', 'CONDUTO_TURBINA',
    'VHZL', 'UEL', 'OEL', 'drpIEEEVC',
    'PSS_COMP', 'RV', 'AVR', 'REG_TEN', 'MED_TEN',
    'MEL', 'SCL', 'DRIVE', 'EXCITATRIZ',
)

# Nombres de parametros numericos a buscar en cada BlkRef.
# Cubre ganancias, constantes de tiempo, limites y tablas de bloquess PF estandar.
_PARAM_NOMBRES = (
    # Ganancias
    'K', 'Ks', 'Ka', 'Kb', 'Kc', 'Kd', 'Ke', 'Kf', 'Kp', 'Ki', 'Kn', 'Km',
    'K1', 'K2', 'K3', 'K4', 'Kconst',
    # Constantes de tiempo
    'T', 'T1', 'T2', 'T3', 'T4', 'T5', 'T6',
    'Ta', 'Tb', 'Tc', 'Td', 'Te', 'Tf', 'Tm', 'Tn', 'Tp', 'Tr', 'Tu',
    'Tdo', 'Tpu', 'Tdu', 'Tw',
    # Limites
    'Tmax', 'Tmin', 'ulMax', 'ulMin', 'yupper', 'ylower',
    'Pmax', 'Pmin', 'Qmax', 'Qmin',
    'xmax', 'xmin', 'ymax', 'ymin',
    # Constante fija / umbral
    'C', 'C1', 'C2',
    # Lookup tables (hasta 12 puntos)
    'x1', 'y1', 'x2', 'y2', 'x3', 'y3', 'x4', 'y4',
    'x5', 'y5', 'x6', 'y6', 'x7', 'y7', 'x8', 'y8',
    'x9', 'y9', 'x10', 'y10', 'x11', 'y11', 'x12', 'y12',
    # Delay / exponencial
    'Tdelay', 'Td1',
    # Otros comunes en PF
    'n', 'alpha', 'beta', 'omega', 'f',
    # Atributos DSL de texto que aun podrian existir
    'sInput', 'sOutput', 'sIntern',
)

# ══════════════════════════════════════════════════════════════


# ─────────────────────────────────────────────────────────────
# Helpers basicos
# ─────────────────────────────────────────────────────────────
def ga(obj, attr):
    try:
        return obj.GetAttribute(attr) if obj is not None else None
    except Exception:
        return None


def _cls(obj):
    try:
        return obj.GetClassName()
    except Exception:
        return ''


def _nom(obj, default='?'):
    if obj is None:
        return default
    return ga(obj, 'loc_name') or default


def _fullname(obj):
    if obj is None:
        return None
    try:
        return obj.GetFullName()
    except Exception:
        return None


def _sanitize(s):
    for c in r'\/:*?"<>|':
        s = s.replace(c, '_')
    return s.strip()


def _to_str(v):
    """Convierte cualquier valor a cadena legible."""
    if v is None:
        return ''
    if isinstance(v, (list, tuple)):
        return ', '.join(str(x) for x in v)
    if isinstance(v, float):
        # Formato compacto: elimina ceros innecesarios
        return f"{v:.6g}"
    return str(v)


# ─────────────────────────────────────────────────────────────
# Lectura de parametros numericos de un BlkRef
# ─────────────────────────────────────────────────────────────
def leer_params_blkref(blkref):
    """
    Lee todos los parametros (numericos y texto) de una instancia BlkRef.
    Retorna dict {nombre_param: valor_str} solo para los que tienen valor.
    """
    params = {}
    for p in _PARAM_NOMBRES:
        try:
            v = blkref.GetAttribute(p)
        except Exception:
            continue
        if v is None:
            continue
        # Filtramos solo si es string vacio; 0.0 es un valor valido
        if isinstance(v, str) and not v.strip():
            continue
        params[p] = _to_str(v)
    return params


# ─────────────────────────────────────────────────────────────
# Composite model helpers
# ─────────────────────────────────────────────────────────────
def _pelm_lista(cm):
    lista = ga(cm, 'pelm')
    if isinstance(lista, (list, tuple)) and len(lista) > 0:
        return [x for x in lista if x is not None]
    resultado = []
    for i in range(60):
        ref = ga(cm, f'pelm[{i}]')
        if ref is None:
            break
        resultado.append(ref)
    return resultado


def _buscar_composite_model(sym_elm, todos_comp):
    cm = ga(sym_elm, 'c_pmod')
    if cm is not None and _cls(cm) == 'ElmComp':
        return cm, 'c_pmod'
    fn_sym = _fullname(sym_elm)
    for cm in todos_comp:
        for ref in _pelm_lista(cm):
            if _fullname(ref) == fn_sym:
                return cm, 'pelm_scan'
    try:
        parent = sym_elm.GetParent()
        if parent:
            comps = parent.GetContents('*.ElmComp', 0) or []
            if comps:
                return comps[0], 'carpeta_padre'
    except Exception:
        pass
    return None, None


def _dsl_slots_de_composite(cm):
    return [ref for ref in _pelm_lista(cm) if _cls(ref) == 'ElmDsl']


def _buscar_comps_en_proyecto(app, project):
    encontrados = {}
    try:
        for cm in (app.GetCalcRelevantObjects('*.ElmComp') or []):
            encontrados[_fullname(cm)] = cm
    except Exception:
        pass
    try:
        for cm in (project.GetContents('*.ElmComp', 1) or []):
            k = _fullname(cm)
            if k not in encontrados:
                encontrados[k] = cm
    except Exception:
        pass
    return list(encontrados.values())


# ─────────────────────────────────────────────────────────────
# Dialogo de carpeta
# ─────────────────────────────────────────────────────────────
def _elegir_carpeta(titulo='Carpeta de salida'):
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk(); root.withdraw(); root.attributes('-topmost', True)
    carpeta = filedialog.askdirectory(title=titulo, mustexist=True)
    root.destroy()
    if not carpeta:
        raise RuntimeError('Seleccion cancelada.')
    return carpeta.replace('/', os.sep)


# ─────────────────────────────────────────────────────────────
# MODO EXTRAER_PARAMS — extrae parametros numericos de cada BlkRef
# ─────────────────────────────────────────────────────────────
def _extraer_params_unidad(loc_name, sym, all_comps, app):
    """
    Para una unidad (ElmSym), devuelve lista de registros:
    {loc_name, dsl_nombre, blkdef_nombre, blkref_nombre,
     tipo_subbloque, sIntern_tipo, sIntern_inst, params{}}
    """
    registros = []
    cm, _ = _buscar_composite_model(sym, all_comps)
    if cm is None:
        app.PrintWarn(f"   {loc_name}: sin composite model.")
        return registros

    dsl_elms = _dsl_slots_de_composite(cm)
    for dsl in dsl_elms:
        dsl_nombre = _nom(dsl, '')
        if not any(dsl_nombre.startswith(p) for p in DSL_SLOTS):
            continue

        blkdef = ga(dsl, 'typ_id')
        blkdef_nombre = _nom(blkdef, 'N/A')

        try:
            hijos = blkdef.GetContents('*', 1) if blkdef else []
        except Exception:
            hijos = []

        for hijo in (hijos or []):
            if _cls(hijo) != 'BlkRef':
                continue

            blkref_nombre  = _nom(hijo, '?')
            sub_blkdef     = ga(hijo, 'typ_id')
            tipo_sub       = _nom(sub_blkdef, 'N/A') if sub_blkdef else 'N/A'
            sintern_tipo   = (_to_str(ga(sub_blkdef, 'sIntern'))
                              if sub_blkdef else '')
            sintern_inst   = _to_str(ga(hijo, 'sIntern'))

            params = leer_params_blkref(hijo)

            registros.append({
                'loc_name':      loc_name,
                'dsl_nombre':    dsl_nombre,
                'blkdef_nombre': blkdef_nombre,
                'blkref_nombre': blkref_nombre,
                'tipo_sub':      tipo_sub,
                'sintern_tipo':  sintern_tipo,
                'sintern_inst':  sintern_inst,
                'params':        params,
            })

    return registros


def _escribir_excel_params(todos_registros, todas_unidades, carpeta, ts):
    """
    Genera Excel con:
    Hoja 'Todos'    : tabla larga con todos los registros.
    Hoja por DSL    : comparacion de parametros entre unidades.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None, 'openpyxl no disponible'

    ruta = os.path.join(carpeta, f"DSL_Params_{ts}.xlsx")
    wb   = openpyxl.Workbook()

    azul   = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    verde  = PatternFill(start_color='375623', end_color='375623', fill_type='solid')
    blanco = Font(color='FFFFFF', bold=True)
    ctr    = Alignment(horizontal='center')
    izq    = Alignment(horizontal='left', wrap_text=False)

    def cab(ws, fill=None):
        f = fill or azul
        for cell in ws[1]:
            cell.fill = f; cell.font = blanco; cell.alignment = ctr

    def ajuste(ws, max_w=60):
        for col in ws.columns:
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w + 3, max_w)

    # ── Hoja 1: tabla completa ───────────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = 'Todos'

    # Recolectar todos los params usados en todo el proyecto
    todos_params = []
    for r in todos_registros:
        for p in r['params']:
            if p not in todos_params:
                todos_params.append(p)
    todos_params = sorted(todos_params)

    cols_all = (['Unidad', 'DSL Bloque', 'BlkDef', 'BlkRef Instancia',
                 'Tipo Sub-bloque', 'Estado (tipo)', 'Estado (inst)']
                + todos_params)
    ws_all.append(cols_all)
    cab(ws_all)

    for r in todos_registros:
        fila = [
            r['loc_name'], r['dsl_nombre'], r['blkdef_nombre'],
            r['blkref_nombre'], r['tipo_sub'],
            r['sintern_tipo'], r['sintern_inst'],
        ]
        for p in todos_params:
            fila.append(r['params'].get(p, ''))
        ws_all.append(fila)
        for cell in ws_all[ws_all.max_row]:
            cell.alignment = izq
    ajuste(ws_all, 50)

    # ── Hojas por DSL bloque: comparacion entre unidades ────────────────────
    from collections import defaultdict
    por_dsl = defaultdict(list)
    for r in todos_registros:
        por_dsl[r['dsl_nombre']].append(r)

    unidades_orden = sorted(todas_unidades)

    for dsl_nom in sorted(por_dsl.keys()):
        regs = por_dsl[dsl_nom]
        nombre_hoja = _sanitize(dsl_nom)[:31]
        ws = wb.create_sheet(title=nombre_hoja)

        # Clave de identificacion de cada fila: (blkref_nombre, tipo_sub)
        # Columnas: identificadores + una columna por unidad (sus params)
        claves_orden = []
        vistas = set()
        for r in regs:
            k = (r['blkref_nombre'], r['tipo_sub'], r['sintern_inst'])
            if k not in vistas:
                claves_orden.append(k)
                vistas.add(k)

        # Recolectar params usados en este DSL
        params_dsl = []
        for r in regs:
            for p in r['params']:
                if p not in params_dsl:
                    params_dsl.append(p)
        params_dsl = sorted(params_dsl)

        if not params_dsl:
            # Sin parametros numericos — muestra tabla basica de bloques
            ws.append(['BlkRef Instancia', 'Tipo Sub-bloque', 'Estado (inst)'])
            cab(ws)
            for k in claves_orden:
                ws.append(list(k))
            ajuste(ws, 40)
            continue

        # Encabezado: BlkRef | Tipo | sIntern | [param x unidad...]
        # Para comparar: agrupamos por (blkref, tipo, sintern_inst)
        # y ponemos cada param de cada unidad en columnas separadas

        # Construir: para cada clave, dict unidad→params
        datos_por_clave = {}
        for r in regs:
            k = (r['blkref_nombre'], r['tipo_sub'], r['sintern_inst'])
            datos_por_clave.setdefault(k, {})[r['loc_name']] = r['params']

        # Si un param tiene el mismo valor en todas las unidades → col "Comun"
        # Si difiere → col separada por unidad
        cab_row = ['BlkRef', 'Tipo Sub-bloque', 'sIntern']
        for p in params_dsl:
            # Verificar si el param varia entre unidades
            vals = set()
            for k in claves_orden:
                for u in unidades_orden:
                    v = datos_por_clave.get(k, {}).get(u, {}).get(p, '')
                    if v:
                        vals.add(v)
            if len(vals) <= 1:
                cab_row.append(p)          # un solo valor: columna comun
            else:
                for u in unidades_orden:
                    cab_row.append(f"{p}\n[{u}]")

        ws.append(cab_row)
        cab(ws, verde)
        ws.row_dimensions[1].height = 30

        for k in claves_orden:
            blkref_n, tipo_sub, sintern_i = k
            fila = [blkref_n, tipo_sub, sintern_i]
            for p in params_dsl:
                vals_u = {u: datos_por_clave.get(k, {}).get(u, {}).get(p, '')
                          for u in unidades_orden}
                vals_set = set(v for v in vals_u.values() if v)
                if len(vals_set) <= 1:
                    fila.append(next(iter(vals_set), ''))
                else:
                    for u in unidades_orden:
                        fila.append(vals_u.get(u, ''))
            ws.append(fila)
            for cell in ws[ws.max_row]:
                cell.alignment = izq

        ajuste(ws, 35)

    wb.save(ruta)
    return ruta, None


def modo_extraer_params(app, project):
    SEP = '=' * 68
    app.PrintInfo(SEP)
    app.PrintInfo('  EXTRAER_PARAMS — Parametros numericos de sub-bloques DSL')
    app.PrintInfo(SEP)

    carpeta  = _elegir_carpeta('Carpeta de salida — Parametros DSL')
    all_syms = app.GetCalcRelevantObjects('*.ElmSym') or []
    target   = {ga(s, 'loc_name'): s for s in all_syms
                if ga(s, 'loc_name') in TARGET_LOC_NAMES}

    faltantes = TARGET_LOC_NAMES - set(target.keys())
    app.PrintInfo(f"Unidades encontradas ({len(target)}): {sorted(target.keys())}")
    if faltantes:
        app.PrintWarn(f"No encontradas: {sorted(faltantes)}")
    app.PrintInfo('')

    all_comps     = _buscar_comps_en_proyecto(app, project)
    todos_registros = []

    for loc_name in sorted(target.keys()):
        app.PrintInfo(f"-> {loc_name}")
        regs = _extraer_params_unidad(loc_name, target[loc_name], all_comps, app)
        todos_registros.extend(regs)

        # Resumen por DSL
        from collections import Counter
        cnt = Counter(r['dsl_nombre'] for r in regs)
        for dsl, n in sorted(cnt.items()):
            n_con_params = sum(1 for r in regs
                               if r['dsl_nombre'] == dsl and r['params'])
            app.PrintInfo(f"   {dsl}: {n} sub-bloques, {n_con_params} con params")
        app.PrintInfo('')

    app.PrintInfo(f"Total registros: {len(todos_registros)}")
    n_con_params = sum(1 for r in todos_registros if r['params'])
    app.PrintInfo(f"Con parametros numericos: {n_con_params}")
    app.PrintInfo('')

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    ruta_xl, err_xl = _escribir_excel_params(
        todos_registros, set(target.keys()), carpeta, ts)

    if ruta_xl and os.path.isfile(ruta_xl):
        tam_kb = os.path.getsize(ruta_xl) / 1024.0
        app.PrintInfo(SEP)
        app.PrintInfo(f"  Excel : {os.path.basename(ruta_xl)}")
        app.PrintInfo(f"  Tam   : {tam_kb:.1f} kB")
        app.PrintInfo(f"  Hoja 'Todos'    : tabla completa de todos los sub-bloques")
        app.PrintInfo(f"  Hojas por DSL   : comparacion entre unidades")
        app.PrintInfo(f"                    params en VERDE = difieren entre unidades")
    else:
        app.PrintError(f"Error Excel: {err_xl}")
    app.PrintInfo(SEP)


# ─────────────────────────────────────────────────────────────
# MODO DIAGNOSTICO — estructura de sub-bloques (una unidad)
# ─────────────────────────────────────────────────────────────
def modo_diagnostico(app, project):
    SEP = '=' * 68
    app.PrintInfo(SEP)
    app.PrintInfo('  DIAGNOSTICO — Sub-bloques y parametros (unidad representativa)')
    app.PrintInfo(SEP)

    all_syms = app.GetCalcRelevantObjects('*.ElmSym') or []
    target   = {ga(s, 'loc_name'): s for s in all_syms
                if ga(s, 'loc_name') in TARGET_LOC_NAMES}

    if not target:
        app.PrintError('Sin unidades objetivo.')
        return

    all_comps = _buscar_comps_en_proyecto(app, project)
    loc_name  = sorted(target.keys())[0]
    app.PrintInfo(f"Unidad: {loc_name}")

    regs = _extraer_params_unidad(loc_name, target[loc_name], all_comps, app)

    for r in regs:
        if not r['params']:
            continue
        params_str = '  |  '.join(f"{k}={v}" for k, v in r['params'].items())
        app.PrintInfo(f"  {r['dsl_nombre']} / {r['blkref_nombre']}"
                      f" [{r['tipo_sub']}]  {r['sintern_inst'] or ''}")
        app.PrintInfo(f"    {params_str}")

    total_con = sum(1 for r in regs if r['params'])
    app.PrintInfo(f"\nSub-bloques con parametros: {total_con} / {len(regs)}")
    app.PrintInfo(SEP)


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────
def main():
    app = pf.GetApplication()
    if app is None:
        return
    try:
        app.ClearOutputWindow()
    except Exception:
        pass

    project = app.GetActiveProject()
    if project is None:
        app.PrintError('No hay proyecto activo.')
        return

    app.PrintInfo(f"Proyecto : {_nom(project)}")
    app.PrintInfo(f"Modo     : {MODO}")
    app.PrintInfo('')

    try:
        if MODO == 'EXTRAER_PARAMS':
            modo_extraer_params(app, project)
        elif MODO == 'DIAGNOSTICO':
            modo_diagnostico(app, project)
        else:
            app.PrintError(f"MODO invalido: {MODO!r}")
    except RuntimeError as e:
        app.PrintError(str(e))


main()
