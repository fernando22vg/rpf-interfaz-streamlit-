#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
ExtractorDSL_Reivax.py
----------------------
Extrae parametros DSL del gobernador Reivax desde el composite model
SYM_FRAME2. Bloques objetivo: ATUADOR, POSICAO, Pot, VHZ, CND, lmc.

Solo extrae parametros configurables por usuario (seccion Parameters del
BlkDef DSL), NO variables de estado ni senales calculadas internamente.

Ejecucion: script interno de PowerFactory (Run as Script)
"""

import os
from datetime import datetime
import powerfactory as pf

# ══════════════════════════════════════════════════════════════
# Configuracion
# ══════════════════════════════════════════════════════════════

TARGET_LOC_NAMES = {
    'sym_BOT01', 'sym_BOT02', 'sym_BOT03',
    'sym_CUT01', 'sym_CUT02', 'sym_CUT03', 'sym_CUT04',
    'sym_ANG03', 'sym_CRB01',
}

# Slots DSL a extraer — nombres confirmados del composite model real.
# Se usa startswith(), asi que un prefijo cubre variantes del mismo bloque.
DSL_SLOTS = (
    'ATUADOR_SIMP',    # actuador / servomotor (gobernador Reivax)
    'POS_SIMP',        # posicion de apertura
    'CONDUTO_TURBINA', # conduto y turbina
    'VHZL',            # proteccion V/Hz
    'UEL',             # limitador de subexcitacion
    'OEL',             # limitador de sobreexcitacion
    'drpIEEEVC',       # droop regulacion de tension
    'PSS_COMP',        # estabilizador de sistema de potencia
    'RV',              # regulador de tension (parte del AVR)
    'MEL',             # minimum excitation limiter
    'SCL',             # stator current limiter
    'DRIVE',           # drive del excitador
    'EXCITATRIZ',      # excitatriz
)

# True: imprime estructura de composite model y BlkDef para depurar
# cuando no se encuentran parametros o los nombres de slot difieren.
MODO_DIAGNOSTICO = True

# ══════════════════════════════════════════════════════════════


# ─────────────────────────────────────────────────────────────
# Helpers generales
# ─────────────────────────────────────────────────────────────
def ga(obj, attr):
    try:
        return obj.GetAttribute(attr) if obj is not None else None
    except Exception:
        return None


def _cls(obj):
    try:
        return obj.GetClassName()
    except Exception:
        return ''


def _nom(obj, default='?'):
    if obj is None:
        return default
    return ga(obj, 'loc_name') or default


def _fullname(obj):
    """Ruta completa del objeto en la jerarquia PF. Usada para comparar identidad."""
    if obj is None:
        return None
    try:
        return obj.GetFullName()
    except Exception:
        return None



# ─────────────────────────────────────────────────────────────
# Atributos estandar de ElmDsl que NO son parametros de usuario
# ─────────────────────────────────────────────────────────────
# Cualquier atributo del ElmDsl que NO este en esta lista y sea
# numerico/string se considera parametro DSL de usuario.
_ATTRS_STD_ELMDSL = {
    # Comunes a todos los objetos PF
    'loc_name', 'fold_id', 'desc', 'for_name', 'chr_name', 'charact',
    'sernum', 'iSchemeStatus', 'GPSlat', 'GPSlon',
    # Referencia al tipo (BlkDef)
    'typ_id',
    # Estado y conexion
    'outserv', 'bus1', 'bus2', 'bus3', 'bus4', 'ibus', 'bbusbar',
    # Control
    'cpSite', 'cpCtrl', 'c_pmod', 'ip_ctrl', 'iOPslack', 'i_mot',
    'pFlicker', 'pStoch', 'pQPcurve',
    # Escalado (generadores)
    'scale0', 'scale0f', 'Pmax', 'Pmin',
}


# ─────────────────────────────────────────────────────────────
# Lectura de parametros desde ElmDsl
# ─────────────────────────────────────────────────────────────
def extraer_parametros_dsl(dsl_elm):
    """
    Extrae parametros de usuario de un ElmDsl usando GetAttributeNames().
    En PF 2025 los valores estan en el ElmDsl, no en el BlkDef.

    Estrategias en orden:
    1. GetAttributeNames() en ElmDsl  — filtra atributos estandar
    2. GetAttributeNames() en BlkDef  — fallback si ElmDsl no responde
    3. GetContents('*') en BlkDef     — busca objetos hijo con params

    Retorna (dict{nombre: valor}, blk_def).
    """
    blk_def = ga(dsl_elm, 'typ_id')
    params  = {}

    # ── Estrategia 1: GetAttributeNames en ElmDsl ────────────────────────
    nombres = _attr_names_de(dsl_elm)
    if nombres:
        params = _leer_params_de_nombres(dsl_elm, nombres, excluir=_ATTRS_STD_ELMDSL)

    # ── Estrategia 2: GetAttributeNames en BlkDef ────────────────────────
    if not params and blk_def is not None:
        nombres_blk = _attr_names_de(blk_def)
        if nombres_blk:
            params = _leer_params_de_nombres(dsl_elm, nombres_blk, excluir=_ATTRS_STD_ELMDSL)

    # ── Estrategia 3: GetContents('*') en BlkDef ─────────────────────────
    if not params and blk_def is not None:
        try:
            hijos = blk_def.GetContents('*', 1) or []
            for hijo in hijos:
                nombre = ga(hijo, 'loc_name')
                if nombre and nombre not in _ATTRS_STD_ELMDSL:
                    val = ga(dsl_elm, nombre)
                    if isinstance(val, (int, float)) and not isinstance(val, bool):
                        params[nombre] = val
        except Exception:
            pass

    return params, blk_def


def _attr_names_de(obj):
    """Devuelve lista de nombres de atributo via GetAttributeNames(), o [] si falla."""
    if obj is None:
        return []
    try:
        names = obj.GetAttributeNames()
        if names and isinstance(names, (list, tuple)):
            return list(names)
    except Exception:
        pass
    return []


def _leer_params_de_nombres(dsl_elm, nombres, excluir):
    """
    Dado un ElmDsl y una lista de nombres de atributo, lee solo los que:
    - No estan en el set de exclusion
    - Son numericos (float/int) — parametros DSL tipicamente son numericos
    Retorna dict{nombre: valor}.
    """
    params = {}
    for name in nombres:
        if name in excluir:
            continue
        # Saltar prefijos de sistema PF ('c_', 'e_', 'i_' suelen ser internos)
        # EXCEPCION: muchos params Reivax empiezan con letras normales (Kp, Ti, etc.)
        val = ga(dsl_elm, name)
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            params[name] = val
        elif isinstance(val, str) and val and name not in excluir:
            params[name] = val
    return params


# ─────────────────────────────────────────────────────────────
# Busqueda de composite model
# ─────────────────────────────────────────────────────────────
def _pelm_lista(cm):
    """
    Devuelve todos los objetos referenciados en los slots pelm del ElmComp.
    Intenta obtenerlos como lista primero; si no, itera pelm[i].
    """
    # Intento 1: pelm como lista completa
    lista = ga(cm, 'pelm')
    if isinstance(lista, (list, tuple)) and len(lista) > 0:
        return [x for x in lista if x is not None]

    # Intento 2: acceso indexado pelm[i]
    resultado = []
    for i in range(60):
        ref = ga(cm, f'pelm[{i}]')
        if ref is None:
            break
        resultado.append(ref)
    return resultado


def _buscar_composite_model_por_c_pmod(sym_elm):
    """Atajo: muchos ElmSym tienen el atributo c_pmod que apunta directamente al ElmComp."""
    cm = ga(sym_elm, 'c_pmod')
    if cm is not None and _cls(cm) == 'ElmComp':
        return cm
    return None


def _buscar_composite_model_en_lista(sym_elm, todos_comp):
    """
    Busca en la lista de ElmComp el que referencie al ElmSym via sus slots pelm.
    Usa GetFullName() para la comparacion (la igualdad == no funciona en COM PF).
    """
    fn_sym = _fullname(sym_elm)
    if fn_sym is None:
        return None

    for cm in todos_comp:
        for ref in _pelm_lista(cm):
            if _fullname(ref) == fn_sym:
                return cm
    return None


def _buscar_composite_model_en_carpeta(sym_elm):
    """
    Fallback: busca ElmComp en la misma carpeta que el ElmSym
    (estructura tipica de unidades de generacion en PF).
    """
    try:
        parent = sym_elm.GetParent()
        if parent is None:
            return None
        comps = parent.GetContents('*.ElmComp', 0) or []
        if comps:
            return comps[0]
    except Exception:
        pass
    return None


def _buscar_composite_model(sym_elm, todos_comp):
    """
    Estrategia multicapa para encontrar el ElmComp de un ElmSym:
    1. Atributo directo c_pmod del ElmSym
    2. Busqueda por GetFullName() en la lista global de ElmComp
    3. Busqueda en la carpeta padre del ElmSym
    """
    cm = _buscar_composite_model_por_c_pmod(sym_elm)
    if cm is not None:
        return cm, 'c_pmod'

    cm = _buscar_composite_model_en_lista(sym_elm, todos_comp)
    if cm is not None:
        return cm, 'pelm_scan'

    cm = _buscar_composite_model_en_carpeta(sym_elm)
    if cm is not None:
        return cm, 'carpeta_padre'

    return None, None


def _dsl_slots_de_composite(cm):
    """Devuelve todos los ElmDsl referenciados como slots en el ElmComp."""
    return [ref for ref in _pelm_lista(cm) if _cls(ref) == 'ElmDsl']


# ─────────────────────────────────────────────────────────────
# Busqueda de todos los ElmComp del proyecto
# ─────────────────────────────────────────────────────────────
def _buscar_comps_en_proyecto(app, project):
    encontrados = {}

    try:
        for cm in (app.GetCalcRelevantObjects('*.ElmComp') or []):
            encontrados[_fullname(cm)] = cm
    except Exception:
        pass

    try:
        for cm in (project.GetContents('*.ElmComp', 1) or []):
            k = _fullname(cm)
            if k not in encontrados:
                encontrados[k] = cm
    except Exception:
        pass

    return list(encontrados.values())


# ─────────────────────────────────────────────────────────────
# Modo diagnostico
# ─────────────────────────────────────────────────────────────
def _diagnosticar_composite(app, cm, dsl_elms, metodo):
    app.PrintInfo(f"  Composite  : {_nom(cm)}  (via {metodo})")

    todos_slots = _pelm_lista(cm)
    app.PrintInfo(f"  Slots total: {len(todos_slots)}")
    for ref in todos_slots:
        app.PrintInfo(f"    [{_cls(ref):12s}] {_nom(ref)}")

    app.PrintInfo(f"  DSL slots  : {len(dsl_elms)}")
    for dsl in dsl_elms:
        blk        = ga(dsl, 'typ_id')
        blk_nombre = _nom(blk, 'None')
        blk_clase  = _cls(blk)

        # GetAttributeNames en ElmDsl y BlkDef
        names_dsl = _attr_names_de(dsl)
        names_blk = _attr_names_de(blk)

        # Params leidos con la nueva estrategia
        params, _ = extraer_parametros_dsl(dsl)

        # Contenidos del BlkDef
        try:
            hijos_blk = [_nom(h) for h in (blk.GetContents('*', 1) or [])]
        except Exception:
            hijos_blk = []

        app.PrintInfo(f"    DSL: {_nom(dsl)}  |  BlkDef: {blk_nombre}  [{blk_clase}]")
        app.PrintInfo(f"      GetAttributeNames(ElmDsl): {names_dsl[:30]}")
        app.PrintInfo(f"      GetAttributeNames(BlkDef): {names_blk[:30]}")
        app.PrintInfo(f"      GetContents(BlkDef)      : {hijos_blk[:20]}")
        app.PrintInfo(f"      params leidos            : {list(params.keys())[:20]}")


# ─────────────────────────────────────────────────────────────
# Seleccion de carpeta (tkinter)
# ─────────────────────────────────────────────────────────────
def _elegir_carpeta():
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    carpeta = filedialog.askdirectory(
        title='Carpeta de salida — Parametros DSL Reivax',
        mustexist=True,
    )
    root.destroy()
    if not carpeta:
        raise RuntimeError('Seleccion de carpeta cancelada.')
    return carpeta.replace('/', os.sep)


# ─────────────────────────────────────────────────────────────
# Exportacion a Excel (openpyxl)
# ─────────────────────────────────────────────────────────────
def _escribir_excel(filas_detalle, filas_ancho, carpeta):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None, 'openpyxl no disponible'

    ts     = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre = f"Parametros_DSL_Reivax_{ts}.xlsx"
    ruta   = os.path.join(carpeta, nombre)

    wb = openpyxl.Workbook()

    azul   = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    verde  = PatternFill(start_color='375623', end_color='375623', fill_type='solid')
    blanco = Font(color='FFFFFF', bold=True)
    centro = Alignment(horizontal='center')

    def estilo_cabecera(ws, fill):
        for cell in ws[1]:
            cell.fill = fill
            cell.font = blanco
            cell.alignment = centro

    def autoajuste(ws):
        for col in ws.columns:
            ancho = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(ancho + 4, 50)

    # ── Hoja 1: Formato largo ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Parametros_Largo'
    COLS = ['Unidad (loc_name)', 'Composite Model', 'DSL Bloque',
            'DSL Definicion', 'Parametro', 'Valor']
    ws1.append(COLS)
    estilo_cabecera(ws1, azul)
    for fila in filas_detalle:
        ws1.append([fila.get(c, '') for c in COLS])
    autoajuste(ws1)

    # ── Hoja 2: Tabla pivot ───────────────────────────────────────────────
    ws2 = wb.create_sheet(title='Tabla_Pivot')
    if filas_ancho:
        FIJAS  = ['Unidad (loc_name)', 'DSL Bloque', 'DSL Definicion']
        params = []
        visto  = set()
        for fila in filas_ancho:
            for k in fila:
                if k not in FIJAS and k not in visto:
                    params.append(k)
                    visto.add(k)
        todas = FIJAS + params
        ws2.append(todas)
        estilo_cabecera(ws2, verde)
        for fila in filas_ancho:
            ws2.append([fila.get(c, '') for c in todas])
        autoajuste(ws2)

    wb.save(ruta)
    return ruta, None


# ─────────────────────────────────────────────────────────────
# Programa principal
# ─────────────────────────────────────────────────────────────
def main():
    app = pf.GetApplication()
    if app is None:
        return

    try:
        app.ClearOutputWindow()
    except Exception:
        pass

    SEP = '=' * 68
    app.PrintInfo(SEP)
    app.PrintInfo('  ExtractorDSL_Reivax  —  Parametros gobernador Reivax')
    app.PrintInfo(SEP)

    project = app.GetActiveProject()
    if project is None:
        app.PrintError('No hay proyecto activo.')
        return

    app.PrintInfo(f"Proyecto        : {_nom(project)}")
    app.PrintInfo(f"Modo diagnostico: {'Si' if MODO_DIAGNOSTICO else 'No'}")
    app.PrintInfo('')

    try:
        carpeta_salida = _elegir_carpeta()
    except RuntimeError as e:
        app.PrintError(str(e))
        return
    app.PrintInfo(f"Carpeta salida  : {carpeta_salida}")
    app.PrintInfo('')

    # ── Buscar ElmSym objetivo ───────────────────────────────────────────
    all_syms = app.GetCalcRelevantObjects('*.ElmSym') or []
    target_syms = {}
    for sym in all_syms:
        nombre = ga(sym, 'loc_name')
        if nombre in TARGET_LOC_NAMES:
            target_syms[nombre] = sym

    faltantes = TARGET_LOC_NAMES - set(target_syms.keys())
    app.PrintInfo(f"Unidades encontradas ({len(target_syms)}): {sorted(target_syms.keys())}")
    if faltantes:
        app.PrintWarn(f"No encontradas: {sorted(faltantes)}")
    app.PrintInfo('')

    # ── Buscar composite models ──────────────────────────────────────────
    all_comps = _buscar_comps_en_proyecto(app, project)
    app.PrintInfo(f"Composite models en proyecto: {len(all_comps)}")
    app.PrintInfo('')

    filas_detalle = []
    filas_ancho   = []

    for loc_name in sorted(target_syms.keys()):
        sym = target_syms[loc_name]
        app.PrintInfo(f"-> {loc_name}  [{_fullname(sym)}]")

        cm, metodo = _buscar_composite_model(sym, all_comps)
        if cm is None:
            app.PrintWarn(f"   Sin composite model. "
                          f"Ruta sym: {_fullname(sym)}")
            app.PrintInfo('')
            continue

        cm_name  = _nom(cm)
        dsl_elms = _dsl_slots_de_composite(cm)
        app.PrintInfo(f"   Composite: {cm_name}  (via {metodo})")
        app.PrintInfo(f"   DSL slots: {[_nom(d) for d in dsl_elms]}")

        if MODO_DIAGNOSTICO:
            _diagnosticar_composite(app, cm, dsl_elms, metodo)

        n_bloques = 0
        for dsl in dsl_elms:
            dsl_name = _nom(dsl, '')

            if not any(dsl_name.startswith(p) for p in DSL_SLOTS):
                continue

            params, blk_def = extraer_parametros_dsl(dsl)
            blk_name        = _nom(blk_def, 'N/A') if blk_def else 'N/A'

            if not params:
                app.PrintWarn(f"   Bloque {dsl_name} ({blk_name}): 0 params. "
                              f"Activar MODO_DIAGNOSTICO para inspeccionar BlkDef.")
            else:
                app.PrintInfo(f"   Bloque: {dsl_name} ({blk_name}) -> {len(params)} param.")

            for param_nombre, param_val in params.items():
                filas_detalle.append({
                    'Unidad (loc_name)': loc_name,
                    'Composite Model':   cm_name,
                    'DSL Bloque':        dsl_name,
                    'DSL Definicion':    blk_name,
                    'Parametro':         param_nombre,
                    'Valor':             param_val,
                })

            fila_ancho = {
                'Unidad (loc_name)': loc_name,
                'DSL Bloque':        dsl_name,
                'DSL Definicion':    blk_name,
            }
            fila_ancho.update(params)
            filas_ancho.append(fila_ancho)
            n_bloques += 1

        if n_bloques == 0:
            app.PrintWarn(f"   Sin bloques Reivax. "
                          f"Verifica DSL_SLOTS con los nombres reales arriba.")
        app.PrintInfo('')

    # ── Exportar ─────────────────────────────────────────────────────────
    app.PrintInfo(f"Total params extraidos : {len(filas_detalle)}")
    app.PrintInfo(f"Total bloques DSL      : {len(filas_ancho)}")
    app.PrintInfo('')

    ruta_excel, err = _escribir_excel(filas_detalle, filas_ancho, carpeta_salida)

    if ruta_excel and os.path.isfile(ruta_excel):
        tam_kb = os.path.getsize(ruta_excel) / 1024.0
        app.PrintInfo(SEP)
        app.PrintInfo(f"  Exportado : {os.path.basename(ruta_excel)}")
        app.PrintInfo(f"  Tamano    : {tam_kb:.1f} kB")
        app.PrintInfo(f"  Carpeta   : {carpeta_salida}")
    else:
        app.PrintError(f"Error al exportar Excel: {err}")

    app.PrintInfo(SEP)


main()
