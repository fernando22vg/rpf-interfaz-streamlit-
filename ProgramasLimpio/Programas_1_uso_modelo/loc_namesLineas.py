# =============================================================================
# CATALOGO DE LINEAS — loc_name PowerFactory
#
# Ejecutar UNA SOLA VEZ: los loc_names del modelo PF son fijos.
# Re-ejecutar solo si se actualiza DatosSINdigsilent.
#
# Lee la hoja Lineas de DatosSINdigsilent.xlsx y genera un catalogo completo
# con nombre descriptivo derivado automaticamente de los nodos extremos.
#
# Nombre descriptivo: "{NODO1} - {NODO2} {kV} kV"
#   (prefijo alfabetico del bus, sin sufijos de tension ni circuito)
#   Ej.: lne_AAC_ACH115  -> "AAC - ACH 115 kV"
#        lne_LCA_URU230  -> "LCA - URU 230 kV"
#        lne_SAN_PCA23001(2) -> "SAN - PCA 230 kV"
#
# Salida: loc_names_lineas.xlsx  en OUTPUT_DIR
#   Hoja LocNames_Lineas  -> una fila por linea con loc_name y nombre descriptivo
#   Hoja Resumen_kV       -> conteo y lista de loc_names agrupados por nivel de tension
# =============================================================================

import os, re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATOS_PF    = r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\DatosSINdigsilent.xlsx"
OUTPUT_DIR  = r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\Designacion de loc_name"
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "loc_names_lineas.xlsx")

# Colores por nivel de tension
COLOR_KV = {
    500: "D9B3FF",   # violeta
    230: "BDD7EE",   # azul
    115: "C5E0B4",   # verde
     69: "FFE699",   # amarillo
     44: "FCE4D6",   # naranja claro
     25: "F4CCCC",   # rosa
     24: "F4CCCC",   # rosa (24.9 kV)
}
COLOR_DEFAULT    = "EDEDED"
COLOR_FUERA_SERV = "FFC7CE"

# =============================================================================
# DERIVAR NOMBRE DESCRIPTIVO
# =============================================================================

def prefijo_bus(nombre_bus):
    """
    Extrae el prefijo alfabetico de un nombre de barra PF.
    Ejemplos:
      AAC115    -> AAC
      EPO11501  -> EPO
      LOM23001  -> LOM
      SH_PCA230 -> PCA   (quita el prefijo SH_ de barra fantasma)
      CHQ11502  -> CHQ
    """
    s = str(nombre_bus).strip()
    s = re.sub(r'^SH_', '', s)         # quitar prefijo de barra fantasma
    m = re.match(r'^([A-Za-z_]+)', s)
    return m.group(1).rstrip('_') if m else s


def kv_desde_bus(nombre_bus):
    """
    Infiere el nivel de tension (kV) a partir del sufijo numerico de un bus.
    Ejemplos:
      TRN23001  -> 230    LOM23001 -> 230
      ORU11501  -> 115    ACH115   -> 115
      ANG069    -> 69     TAJ115   -> 115
      TRI02401  -> 24     (024 -> 24)
    Retorna float o None si no se puede inferir.
    """
    s = re.sub(r'^SH_', '', str(nombre_bus).strip())
    m = re.search(r'(\d+)', s)
    if not m:
        return None
    digits = m.group(1)
    # Interpretar los primeros 3 digits como el nivel nominal
    # 230xx -> 230, 115xx -> 115, 069xx -> 69, 024xx -> 24
    kv_map = {"500": 500, "230": 230, "115": 115, "069": 69,
              "044": 44, "025": 25, "024": 24, "024.9": 24.9}
    for prefix, val in kv_map.items():
        if digits.startswith(prefix.replace(".", "")):
            return float(val)
    # Fallback: si los 3 primeros digitos son exactamente el nivel
    try:
        return float(digits[:3])
    except ValueError:
        return None


def nombre_descriptivo(nodo_from, nodo_to, kv):
    """
    Genera el nombre legible de la linea: "NODO1 - NODO2 {kV} kV"
    Si kv es NaN, intenta inferirlo desde los nombres de barra.
    """
    n1 = prefijo_bus(nodo_from)
    n2 = prefijo_bus(nodo_to)
    if pd.isna(kv):
        kv = kv_desde_bus(nodo_from) or kv_desde_bus(nodo_to)
    kv_str = f"{int(kv)} kV" if kv is not None and not pd.isna(kv) else "? kV"
    return f"{n1} - {n2} {kv_str}"


def color_para_kv(kv):
    if pd.isna(kv):
        return COLOR_DEFAULT
    kv_int = int(round(kv))
    # Buscar el nivel mas cercano en la tabla
    for nivel in sorted(COLOR_KV.keys(), reverse=True):
        if kv_int >= nivel * 0.9:
            return COLOR_KV[nivel]
    return COLOR_DEFAULT

# =============================================================================
# LECTURA Y CONSTRUCCION DE TABLAS
# =============================================================================

def leer_lineas_pf(datos_pf_path):
    df = pd.read_excel(datos_pf_path, sheet_name="Lineas")
    df["Nombre"] = df["Nombre"].astype(str).str.strip()
    # Excluir filas de leyenda u otras no-lineas
    df = df[df["Nombre"].str.startswith("lne_")].copy()
    return df


def construir_catalogo(df_lne):
    """Una fila por linea con loc_name y nombre descriptivo."""
    filas = []
    for _, r in df_lne.iterrows():
        loc   = r["Nombre"]
        kv    = r.get("Tension nom. (kV)")
        from_ = str(r.get("Nodo From", "")).strip()
        to_   = str(r.get("Nodo To",   "")).strip()
        # Inferir tension cuando no esta disponible en la columna
        if pd.isna(kv):
            kv = kv_desde_bus(from_) or kv_desde_bus(to_)
        filas.append({
            "loc_name PF":          loc,
            "Nombre descriptivo":   nombre_descriptivo(from_, to_, kv),
            "Nodo From":            from_,
            "Nodo To":              to_,
            "Tension nom. (kV)":    kv,
            "Distancia (km)":       r.get("Distancia (km)"),
            "Corriente nom. (A)":   r.get("Corriente nom. (A)"),
            "Carga (%)":            r.get("Carga (%)"),
            "P from (MW)":          r.get("P from (MW)"),
            "En servicio":          r.get("En servicio", ""),
        })
    df = pd.DataFrame(filas)
    # Ordenar: por tension desc, luego nombre
    df = df.sort_values(
        ["Tension nom. (kV)", "Nombre descriptivo"],
        ascending=[False, True]
    ).reset_index(drop=True)
    return df


def construir_resumen_kv(df_cat):
    """Agrupa por nivel de tension con conteo y lista de loc_names."""
    filas = []
    for kv, grp in df_cat.groupby("Tension nom. (kV)", sort=False):
        kv_orden = kv if pd.notna(kv) else -1
        en_serv  = grp[grp["En servicio"].astype(str).str.lower() == "si"]
        filas.append({
            "Tension (kV)":          kv,
            "N lineas total":        len(grp),
            "N en servicio":         len(en_serv),
            "N fuera de servicio":   len(grp) - len(en_serv),
            "loc_names (todos)":     ", ".join(grp["loc_name PF"].tolist()),
            "loc_names (en servicio)": ", ".join(en_serv["loc_name PF"].tolist()),
        })
    df = pd.DataFrame(filas)
    df = df.sort_values("Tension (kV)", ascending=False).reset_index(drop=True)
    return df

# =============================================================================
# FORMATO EXCEL
# =============================================================================
_THIN = Side(border_style="thin", color="BFBFBF")
_BRD  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left",   vertical="center")
_HFIL = PatternFill("solid", start_color="1F3864")
_HFNT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_NFNT = Font(name="Arial", size=10)
_BFNT = Font(name="Arial", bold=True, size=10)


def _fill(h): return PatternFill("solid", start_color=h, end_color=h)


def _aplicar_catalogo(ws, col_kv, col_serv):
    ws.row_dimensions[1].height = 30
    mc = ws.max_column
    for c in range(1, mc + 1):
        cell = ws.cell(1, c)
        cell.fill = _HFIL; cell.font = _HFNT
        cell.alignment = _CTR; cell.border = _BRD

    for r in range(2, ws.max_row + 1):
        kv_val  = ws.cell(r, col_kv).value
        serv    = str(ws.cell(r, col_serv).value or "").strip().lower()
        if serv not in ("si", "true", "1"):
            color = COLOR_FUERA_SERV
        else:
            color = color_para_kv(kv_val)
        f = _fill(color)
        for c in range(1, mc + 1):
            cell = ws.cell(r, c)
            cell.font = _NFNT; cell.border = _BRD; cell.fill = f
            cell.alignment = _CTR if isinstance(cell.value, (int, float)) else _LEFT

    # Anchos de columna
    for col in ws.columns:
        hdr   = str(ws.cell(1, col[0].column).value or "").lower()
        ancho = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        if "nombre descriptivo" in hdr or "loc_names" in hdr:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 35
        else:
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ancho + 3, 12), 40)
    ws.freeze_panes = "C2"


def _aplicar_resumen(ws, col_kv):
    ws.row_dimensions[1].height = 30
    mc = ws.max_column
    for c in range(1, mc + 1):
        cell = ws.cell(1, c)
        cell.fill = _HFIL; cell.font = _HFNT
        cell.alignment = _CTR; cell.border = _BRD

    for r in range(2, ws.max_row + 1):
        kv_val = ws.cell(r, col_kv).value
        color  = color_para_kv(kv_val)
        f = _fill(color)
        for c in range(1, mc + 1):
            cell = ws.cell(r, c)
            val  = cell.value
            cell.font = _NFNT; cell.border = _BRD; cell.fill = f
            if "loc_names" in str(ws.cell(1, c).value or "").lower():
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True)
                ws.row_dimensions[r].height = min(
                    15 * (max(1, len(str(val or "")) // 100) + 1), 120)
            elif isinstance(val, (int, float)):
                cell.alignment = _CTR
            else:
                cell.alignment = _LEFT

    for col in ws.columns:
        hdr   = str(ws.cell(1, col[0].column).value or "").lower()
        ancho = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        if "loc_names" in hdr:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 100
        else:
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ancho + 3, 14), 30)
    ws.freeze_panes = "B2"


def _agregar_leyenda(ws, fila_inicio):
    fila = fila_inicio + 2
    t = ws.cell(fila, 1, "LEYENDA — Nivel de tension")
    t.fill = _HFIL; t.font = _HFNT; t.alignment = _CTR; t.border = _BRD
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)

    items = sorted(COLOR_KV.items(), reverse=True)
    items.append((None, COLOR_DEFAULT))
    items.append((None, COLOR_FUERA_SERV))
    etiquetas = {k: f"{k} kV" for k, _ in COLOR_KV.items()}
    etiquetas[None] = ["Tension desconocida", "Fuera de servicio"]

    for i, (kv, color) in enumerate(items):
        fila += 1
        etiq = (f"{kv} kV" if kv is not None
                else ("Tension desconocida" if i == len(items) - 2
                      else "Fuera de servicio"))
        c1 = ws.cell(fila, 1, ""); c1.fill = _fill(color); c1.border = _BRD
        c2 = ws.cell(fila, 2, etiq)
        c2.font = _NFNT; c2.border = _BRD; c2.fill = _fill(color)
        c2.alignment = _LEFT

# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 60)
    print("loc_namesLineas — Catalogo de Lineas PF")
    print("(ejecutar UNA SOLA VEZ; resultado valido para todos los eventos)")
    print("=" * 60)

    print(f"\n  DatosPF : {os.path.basename(DATOS_PF)}")
    print(f"  Salida  : {OUTPUT_PATH}")

    # -- Leer datos ----------------------------------------------------------
    print("\n[1/3] Leyendo lineas del modelo PF...")
    df_lne = leer_lineas_pf(DATOS_PF)
    print(f"      {len(df_lne)} lineas en modelo PF")
    en_serv = (df_lne["En servicio"].astype(str).str.lower() == "si").sum()
    print(f"      En servicio: {en_serv}  |  Fuera: {len(df_lne) - en_serv}")

    # -- Construir tablas ----------------------------------------------------
    print("[2/3] Construyendo catalogo...")
    df_cat    = construir_catalogo(df_lne)
    df_resumen = construir_resumen_kv(df_cat)

    print(f"\n      Por nivel de tension:")
    for _, r in df_resumen.iterrows():
        kv_str = f"{int(r['Tension (kV)'])} kV" if pd.notna(r["Tension (kV)"]) else "?"
        print(f"        {kv_str:>8}  :  {r['N lineas total']:3d} lineas  "
              f"({r['N en servicio']} en serv., {r['N fuera de servicio']} fuera)")

    # -- Exportar ------------------------------------------------------------
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print(f"\n[3/3] Exportando a Excel...")
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        df_cat.to_excel(    writer, sheet_name="LocNames_Lineas", index=False)
        df_resumen.to_excel(writer, sheet_name="Resumen_kV",      index=False)

    wb = load_workbook(OUTPUT_PATH)

    cols_c = list(df_cat.columns)
    _aplicar_catalogo(wb["LocNames_Lineas"],
                      col_kv  = cols_c.index("Tension nom. (kV)") + 1,
                      col_serv= cols_c.index("En servicio") + 1)
    _agregar_leyenda(wb["LocNames_Lineas"], wb["LocNames_Lineas"].max_row)

    cols_r = list(df_resumen.columns)
    _aplicar_resumen(wb["Resumen_kV"],
                     col_kv=cols_r.index("Tension (kV)") + 1)

    wb.save(OUTPUT_PATH)

    print(f"\n  Archivo creado en:")
    print(f"  {OUTPUT_PATH}")
    print(f"\n  Hojas:")
    for sh in wb.sheetnames:
        print(f"    {sh:<20} -> {wb[sh].max_row - 1} filas")


if __name__ == "__main__":
    main()
    input("\nPresiona Enter para cerrar...")
