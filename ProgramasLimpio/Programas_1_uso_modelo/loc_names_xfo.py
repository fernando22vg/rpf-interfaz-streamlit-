# =============================================================================
# LOC_NAMES DE TRANSFORMADORES — PowerFactory (DatosSINdigsilent)
#
# Ejecutar UNA SOLA VEZ: los loc_names del modelo PF son fijos y no cambian
# entre eventos ni semestres. Re-ejecutar solo si se actualiza DatosSINdigsilent.
#
# Fuente primaria de buses (si existe): topologia_completa_pf.xlsx generado por
# TopologiaCompleta_PFV2.py. Lee las barras reales directamente de PowerFactory.
#   - Hoja "Transformadores" : ElmTr2 con columnas Barra 1 / Barra 2
#   - Hoja "Elementos"       : ElmTr3 con columnas Barra 1 / Barra 2 / Barra 3
#
# Fuente de respaldo (si la topologia no esta disponible o falta algun xfo):
# Infiere las barras desde el nombre del transformador y la hoja Barras de
# DatosSINdigsilent.xlsx usando el patron original:
#   trf_AAC069  -> sub=AAC, HV_code=069 -> Barra HV = AAC069 (69 kV)
#   atr_CAR50001 -> sub=CAR, HV_code=500 -> Barra HV = CAR500 (500 kV)
#
# Salida: loc_names_xfo.xlsx  en OUTPUT_DIR
#   Hoja Transformadores_2dev -> transformadores de 2 devanados con sus datos
#   Hoja Transformadores_3dev -> transformadores de 3 devanados con sus datos
#   Hoja Resumen              -> estadisticas por nivel de tension
# =============================================================================

import os, re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATOS_PF      = r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\DatosSINdigsilent.xlsx"
TOPOLOGIA_PF  = r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\Topologia\topologia_completa_pf.xlsx"
OUTPUT_DIR    = r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\Designacion de loc_name"
OUTPUT_PATH   = os.path.join(OUTPUT_DIR, "loc_names_xfo.xlsx")

# =============================================================================
# LECTURA DE BUSES DESDE TOPOLOGIA (fuente primaria)
# =============================================================================

def leer_buses_de_topologia(path):
    """
    Lee los buses reales de transformadores desde topologia_completa_pf.xlsx.

    Retorna dos dicts:
      buses_2dev: {nombre: {"hv": barra1, "lv": barra2}}
      buses_3dev: {nombre: {"hv": barra1, "mv": barra2, "lv": barra3}}

    Si el archivo no existe o falla, retorna dicts vacios (se usa inferencia de respaldo).
    """
    buses_2dev = {}
    buses_3dev = {}

    if not os.path.isfile(path):
        return buses_2dev, buses_3dev

    def _limpia(v):
        s = str(v).strip() if v is not None else ""
        return "" if s in ("nan", "None", "-") else s

    try:
        # ElmTr2: hoja "Transformadores" generada por TopologiaCompleta_PFV2.py
        df2 = pd.read_excel(path, sheet_name="Transformadores")
        for _, r in df2.iterrows():
            nombre = _limpia(r.get("Nombre"))
            b1     = _limpia(r.get("Barra 1"))
            b2     = _limpia(r.get("Barra 2"))
            if nombre and (b1 or b2):
                buses_2dev[nombre] = {"hv": b1, "lv": b2}
    except Exception:
        pass

    try:
        # ElmTr3: hoja "Elementos" filtrada por clase ElmTr3
        df_elem = pd.read_excel(path, sheet_name="Elementos")
        df3 = df_elem[df_elem["Clase PF"].astype(str).str.strip() == "ElmTr3"].copy()
        for _, r in df3.iterrows():
            nombre = _limpia(r.get("Nombre"))
            b1     = _limpia(r.get("Barra 1"))
            b2     = _limpia(r.get("Barra 2"))
            b3     = _limpia(r.get("Barra 3"))
            if nombre and (b1 or b2 or b3):
                buses_3dev[nombre] = {"hv": b1, "mv": b2, "lv": b3}
    except Exception:
        pass

    return buses_2dev, buses_3dev


# =============================================================================
# COLORES POR NIVEL DE TENSION
# =============================================================================
COLOR_TENSION = {
    "500 kV":  "C00000",   # rojo oscuro (blanco de fuente)
    "230 kV":  "BDD7EE",   # azul claro
    "115 kV":  "C5E0B4",   # verde claro
    "69 kV":   "FFE699",   # amarillo
    "34.5 kV": "FCE4D6",   # salmon
    "Otro":    "E2EFDA",   # verde muy claro
}
FONT_BLANCA = {"500 kV"}   # niveles donde la fuente debe ser blanca


def nivel_tension(kv):
    """Clasifica la tension HV nominal en un nivel estandar."""
    try:
        v = float(kv)
    except (TypeError, ValueError):
        return "Otro"
    if v >= 450:
        return "500 kV"
    if v >= 200:
        return "230 kV"
    if v >= 100:
        return "115 kV"
    if v >= 60:
        return "69 kV"
    if v >= 30:
        return "34.5 kV"
    return "Otro"


# =============================================================================
# LECTURA Y PREPROCESO DE BARRAS
# =============================================================================
_EXCLUIR = {"LEYENDA", "nan", ""}


def leer_barras(path):
    """
    Retorna un dict: {(sub_code_upper, kv_float): [bus_name, ...]}
    sub_code se extrae del prefijo alfabetico del nombre de barra.
    """
    df = pd.read_excel(path, sheet_name="Barras")
    df["Nombre"] = df["Nombre"].astype(str).str.strip()
    df = df[~df["Nombre"].isin(_EXCLUIR)].copy()
    df["kv"] = pd.to_numeric(df["Tension nom. (kV)"], errors="coerce")

    sub_buses = {}   # (sub, kv) -> [bus_name, ...]
    for _, row in df.iterrows():
        bname = row["Nombre"]
        kv    = row["kv"]
        if pd.isna(kv):
            continue
        m = re.match(r"^([A-Za-zÀ-ÿ]{2,5})\d", bname, re.IGNORECASE)
        if m:
            key = (m.group(1).upper(), kv)
            sub_buses.setdefault(key, []).append(bname)
    return sub_buses


# =============================================================================
# HELPERS DE NOMBRE
# =============================================================================

def _parse_nombre(nombre):
    """
    Extrae (sub_code, hv_kv, unit_num) de un nombre como:
      trf_AAC069, trf_AAR11501, atr_CAR50001, trf_CAR23002, trf_CAÑ069
    Retorna (sub, hv_kv, unit) o (None, None, None).
    Acepta letras Unicode (Ñ, etc.) en el codigo de subestacion.
    """
    m = re.match(r"(?:trf|atr)_([\w]+?)(\d{3})(\d{2})?$", nombre, re.IGNORECASE)
    if not m:
        return None, None, None
    sub  = m.group(1).upper()
    code = m.group(2)          # '069', '115', '230', '500'
    unit = int(m.group(3)) if m.group(3) else None
    hv_kv = float(code)
    return sub, hv_kv, unit


def _buscar_bus_con_tolerancia(sub, kv_objetivo, sub_buses, tol=0.10):
    """
    Busca buses en la subestacion con kv_objetivo ± tol (10% por defecto).
    Cubre diferencias de diseño: 66/69 kV, 245/230 kV, 117.9/115 kV, etc.
    Retorna la lista de buses de la kv mas cercana dentro de la tolerancia.
    """
    buses_exactos = sub_buses.get((sub, kv_objetivo), [])
    if buses_exactos:
        return buses_exactos, kv_objetivo
    mejor_kv, mejor_buses = None, []
    mejor_diff = float("inf")
    for (s, k), v in sub_buses.items():
        if s != sub:
            continue
        diff = abs(k - kv_objetivo) / max(kv_objetivo, 0.001)
        if diff < tol and diff < mejor_diff:
            mejor_diff = diff
            mejor_kv   = k
            mejor_buses = v
    return mejor_buses, mejor_kv


def _elegir_bus(bus_lista, unit_num):
    """
    Dado una lista de buses en la misma subestacion/kV, elige el mas apropiado:
      - Si hay uno solo, lo retorna directamente.
      - Si hay unit_num, intenta bus_lista[unit_num - 1] (0-indexed).
      - Si no, retorna el primero o una cadena con todos.
    """
    if not bus_lista:
        return "-"
    if len(bus_lista) == 1:
        return bus_lista[0]
    if unit_num is not None and 1 <= unit_num <= len(bus_lista):
        return bus_lista[unit_num - 1]
    return bus_lista[0]   # por defecto el primero


def inferir_barras_2dev(nombre, ten_hv, ten_lv, sub_buses):
    """
    Deduce Barra HV y Barra LV para un transformador de 2 devanados.
    ten_hv, ten_lv: tensiones nominales en kV (pueden ser NaN).
    Usa tolerancia del 10% para cubrir diferencias de diseño (66/69 kV, 245/230 kV, etc.)
    """
    sub, hv_kv, unit = _parse_nombre(nombre)
    if sub is None:
        return "-", "-"

    # Barra HV
    if pd.notna(ten_hv):
        hv_kv = float(ten_hv)
    buses_hv, _ = _buscar_bus_con_tolerancia(sub, hv_kv, sub_buses)
    barra_hv = _elegir_bus(buses_hv, unit)

    # Barra LV
    barra_lv = "-"
    if pd.notna(ten_lv):
        buses_lv, _ = _buscar_bus_con_tolerancia(sub, float(ten_lv), sub_buses)
        barra_lv = _elegir_bus(buses_lv, unit)

    return barra_hv, barra_lv


def inferir_barras_3dev(nombre, ten_hv, ten_mv, ten_lv, sub_buses):
    """
    Deduce Barra HV, MV y LV para un autotransformador de 3 devanados.
    Si las tensiones nominales son NaN (comun en 3dev), se infieren a partir
    de los buses disponibles en la subestacion (mayor a menor kV).
    """
    sub, hv_kv, unit = _parse_nombre(nombre)
    if sub is None:
        return "-", "-", "-", None, None, None

    # Buses disponibles en la subestacion, ordenados de mayor a menor kV
    kvs_en_sub = sorted(
        {k for (s, k) in sub_buses if s == sub}, reverse=True
    )

    def bus_de(kv_nom, idx_fallback):
        """Busca bus con tolerancia o usa el idx-esimo de kvs_en_sub."""
        if pd.notna(kv_nom):
            bl, kv_real = _buscar_bus_con_tolerancia(sub, float(kv_nom), sub_buses)
            return _elegir_bus(bl, unit), kv_real if kv_real else kv_nom
        elif idx_fallback < len(kvs_en_sub):
            kv2 = kvs_en_sub[idx_fallback]
            return _elegir_bus(sub_buses.get((sub, kv2), []), unit), kv2
        return "-", None

    barra_hv, kv_hv = bus_de(ten_hv, 0)
    barra_mv, kv_mv = bus_de(ten_mv, 1)
    barra_lv, kv_lv = bus_de(ten_lv, 2)

    return barra_hv, barra_mv, barra_lv, kv_hv, kv_mv, kv_lv


# =============================================================================
# LECTURA DE DATOS
# =============================================================================

def leer_transformadores_2dev(path):
    df = pd.read_excel(path, sheet_name="Transformadores_2dev")
    df["Nombre"] = df["Nombre"].astype(str).str.strip()
    df = df[~df["Nombre"].isin(_EXCLUIR)].copy()
    df = df[df["Nombre"].str.len() > 2].copy()
    return df.reset_index(drop=True)


def leer_transformadores_3dev(path):
    df = pd.read_excel(path, sheet_name="Transformadores_3dev")
    df["Nombre"] = df["Nombre"].astype(str).str.strip()
    df = df[~df["Nombre"].isin(_EXCLUIR)].copy()
    df = df[df["Nombre"].str.len() > 2].copy()
    return df.reset_index(drop=True)


# =============================================================================
# CONSTRUCCION DE TABLAS DE SALIDA
# =============================================================================

def construir_tabla_2dev(df, sub_buses, buses_topologia=None):
    """
    buses_topologia: dict {nombre: {"hv": barra1, "lv": barra2}} leido de
    topologia_completa_pf.xlsx. Si None o vacio, se usa solo la inferencia.
    """
    if buses_topologia is None:
        buses_topologia = {}
    filas = []
    n_topo = 0
    n_inferido = 0
    for _, r in df.iterrows():
        nombre   = r["Nombre"]
        ten_hv   = r.get("Tension HV nom. (kV)")
        ten_lv   = r.get("Tension LV nom. (kV)")
        potencia = r.get("Potencia nom. (MVA)")
        en_serv  = str(r.get("En servicio", "")).strip()

        fuente = "inferido"

        # 1) Fuente primaria: topologia_completa_pf.xlsx (buses reales de PF)
        topo = buses_topologia.get(nombre, {})
        barra_hv = topo.get("hv", "")
        barra_lv = topo.get("lv", "")
        if barra_hv or barra_lv:
            fuente = "topologia"
            n_topo += 1
        else:
            # 2) Fuente de respaldo: inferencia desde nombre del transformador
            barra_hv, barra_lv = inferir_barras_2dev(nombre, ten_hv, ten_lv, sub_buses)
            n_inferido += 1

        # 3) Override explícito en la hoja de DatosSINdigsilent (si existe y no esta vacio)
        bh_excel = str(r.get("Barra HV", "")).strip()
        bl_excel = str(r.get("Barra LV", "")).strip()
        if bh_excel not in ("nan", "", "-"):
            barra_hv = bh_excel
        if bl_excel not in ("nan", "", "-"):
            barra_lv = bl_excel

        curmg = r.get("curmg")

        filas.append({
            "loc_name":              nombre,
            "Barra HV":              barra_hv,
            "Barra LV":              barra_lv,
            "Tension HV nom. (kV)":  ten_hv,
            "Tension LV nom. (kV)":  ten_lv,
            "Potencia nom. (MVA)":   potencia,
            "Perd. vacio curmg (%)": curmg if pd.notna(curmg) else None,
            "Nivel tension HV":      nivel_tension(ten_hv),
            "En servicio":           en_serv if en_serv not in ("nan", "") else "-",
            "Fuente buses":          fuente,
        })
    if n_topo or n_inferido:
        print(f"      2dev — desde topologia: {n_topo}  |  inferido: {n_inferido}")
    return pd.DataFrame(filas)


def construir_tabla_3dev(df, sub_buses, buses_topologia=None):
    """
    buses_topologia: dict {nombre: {"hv": barra1, "mv": barra2, "lv": barra3}} leido de
    topologia_completa_pf.xlsx. Si None o vacio, se usa solo la inferencia.
    """
    if buses_topologia is None:
        buses_topologia = {}
    filas = []
    n_topo = 0
    n_inferido = 0
    for _, r in df.iterrows():
        nombre   = r["Nombre"]
        ten_hv   = r.get("Tension HV nom. (kV)")
        ten_mv   = r.get("Tension MV nom. (kV)")
        ten_lv   = r.get("Tension LV nom. (kV)")
        potencia = r.get("Potencia nom. (MVA)")
        en_serv  = str(r.get("En servicio", "")).strip()

        fuente = "inferido"
        kv_hv, kv_mv, kv_lv = ten_hv, ten_mv, ten_lv

        # 1) Fuente primaria: topologia_completa_pf.xlsx (buses reales de PF)
        topo = buses_topologia.get(nombre, {})
        barra_hv = topo.get("hv", "")
        barra_mv = topo.get("mv", "")
        barra_lv = topo.get("lv", "")
        if barra_hv or barra_mv or barra_lv:
            fuente = "topologia"
            n_topo += 1
        else:
            # 2) Fuente de respaldo: inferencia desde nombre del transformador
            barra_hv, barra_mv, barra_lv, kv_hv, kv_mv, kv_lv = inferir_barras_3dev(
                nombre, ten_hv, ten_mv, ten_lv, sub_buses
            )
            n_inferido += 1

        # 3) Override explícito en la hoja de DatosSINdigsilent (si existe y no esta vacio)
        bh_ex = str(r.get("Barra HV", "")).strip()
        bm_ex = str(r.get("Barra MV", "")).strip()
        bl_ex = str(r.get("Barra LV", "")).strip()
        if bh_ex not in ("nan", "", "-"):
            barra_hv = bh_ex
        if bm_ex not in ("nan", "", "-"):
            barra_mv = bm_ex
        if bl_ex not in ("nan", "", "-"):
            barra_lv = bl_ex

        pot_val = potencia if pd.notna(potencia) else None

        curmg = r.get("curmg")

        filas.append({
            "loc_name":              nombre,
            "Barra HV":              barra_hv,
            "Barra MV":              barra_mv,
            "Barra LV":              barra_lv,
            "Tension HV nom. (kV)":  kv_hv if pd.isna(ten_hv) and kv_hv else ten_hv,
            "Tension MV nom. (kV)":  kv_mv if pd.isna(ten_mv) and kv_mv else ten_mv,
            "Tension LV nom. (kV)":  kv_lv if pd.isna(ten_lv) and kv_lv else ten_lv,
            "Potencia nom. (MVA)":   pot_val,
            "Perd. vacio curmg (%)": curmg if pd.notna(curmg) else None,
            "Nivel tension HV":      nivel_tension(kv_hv if pd.isna(ten_hv) else ten_hv),
            "En servicio":           en_serv if en_serv not in ("nan", "") else "-",
            "Fuente buses":          fuente,
        })
    if n_topo or n_inferido:
        print(f"      3dev — desde topologia: {n_topo}  |  inferido: {n_inferido}")
    return pd.DataFrame(filas)


def construir_resumen(df2, df3):
    """Resumen de transformadores por nivel de tension."""
    filas = []
    for nivel in ["500 kV", "230 kV", "115 kV", "69 kV", "34.5 kV", "Otro"]:
        # 2 devanados
        grp2 = df2[df2["Nivel tension HV"] == nivel]
        if not grp2.empty:
            p2 = grp2["Potencia nom. (MVA)"].sum(min_count=1)
            filas.append({
                "Tipo":                 f"2dev — {nivel}",
                "N transformadores":    len(grp2),
                "Potencia total (MVA)": round(p2, 2) if pd.notna(p2) else None,
                "loc_names":            ", ".join(grp2["loc_name"].tolist()),
            })
        # 3 devanados
        grp3 = df3[df3["Nivel tension HV"] == nivel]
        if not grp3.empty:
            p3 = grp3["Potencia nom. (MVA)"].sum(min_count=1)
            filas.append({
                "Tipo":                 f"3dev — {nivel}",
                "N transformadores":    len(grp3),
                "Potencia total (MVA)": round(p3, 2) if pd.notna(p3) else None,
                "loc_names":            ", ".join(grp3["loc_name"].tolist()),
            })
    return pd.DataFrame(filas)


# =============================================================================
# FORMATO EXCEL
# =============================================================================
_THIN = Side(border_style="thin", color="BFBFBF")
_BRD  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left",   vertical="center")
_HFIL = PatternFill("solid", start_color="1F3864")
_HFNT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_NFNT = Font(name="Arial", size=10)
_WFNT = Font(name="Arial", size=10, color="FFFFFF")   # fuente blanca para 500 kV


def _fill(h):
    return PatternFill("solid", start_color=h, end_color=h)


def _aplicar_hoja(ws, col_nivel_idx, col_serv_idx=None, freeze="A2"):
    """
    col_nivel_idx: indice de columna (1-based) con el nivel de tension.
    col_serv_idx:  indice de columna (1-based) con En servicio (o None).
    """
    ws.row_dimensions[1].height = 30
    mc = ws.max_column
    # Cabecera
    for c in range(1, mc + 1):
        cell = ws.cell(1, c)
        cell.fill = _HFIL; cell.font = _HFNT
        cell.alignment = _CTR; cell.border = _BRD
    # Filas de datos
    for r in range(2, ws.max_row + 1):
        nivel = str(ws.cell(r, col_nivel_idx).value or "")
        serv  = str(ws.cell(r, col_serv_idx).value or "Si") if col_serv_idx else "Si"
        if col_serv_idx and serv.strip().lower() not in ("si", "true", "1"):
            color = "FFC7CE"
            use_white = False
        else:
            color     = COLOR_TENSION.get(nivel, "FFFFFF")
            use_white = nivel in FONT_BLANCA
        f    = _fill(color)
        font = _WFNT if use_white else _NFNT
        for c in range(1, mc + 1):
            cell = ws.cell(r, c)
            cell.fill = f; cell.border = _BRD; cell.font = font
            cell.alignment = _CTR if isinstance(cell.value, (int, float)) else _LEFT
    # Anchos de columna
    for col in ws.columns:
        hdr   = str(ws.cell(1, col[0].column).value or "").lower()
        ancho = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        if "loc_names" in hdr:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 80
        else:
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ancho + 3, 12), 45)
    if freeze:
        ws.freeze_panes = freeze


def _agregar_leyenda(ws, fila_inicio):
    fila = fila_inicio + 2
    t = ws.cell(fila, 1, "LEYENDA — Nivel de tension HV")
    t.fill = _HFIL; t.font = _HFNT; t.alignment = _CTR; t.border = _BRD
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
    for nivel, color in COLOR_TENSION.items():
        fila += 1
        c1 = ws.cell(fila, 1, ""); c1.fill = _fill(color); c1.border = _BRD
        c2 = ws.cell(fila, 2, nivel)
        font = _WFNT if nivel in FONT_BLANCA else _NFNT
        c2.font = font; c2.border = _BRD; c2.fill = _fill(color)
        c2.alignment = _LEFT
    # Fuera de servicio
    fila += 1
    c1 = ws.cell(fila, 1, ""); c1.fill = _fill("FFC7CE"); c1.border = _BRD
    c2 = ws.cell(fila, 2, "Fuera de servicio")
    c2.font = _NFNT; c2.border = _BRD; c2.fill = _fill("FFC7CE")
    c2.alignment = _LEFT


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 62)
    print("loc_names_xfo — loc_names Transformadores PowerFactory")
    print("(ejecutar UNA SOLA VEZ; resultado valido para todos los eventos)")
    print("=" * 62)

    if not os.path.isfile(DATOS_PF):
        raise FileNotFoundError(
            f"No se encontro el archivo de datos PF:\n  {DATOS_PF}")

    print(f"\n  DatosPF    : {os.path.basename(DATOS_PF)}")
    print(f"  Topologia  : {os.path.basename(TOPOLOGIA_PF)}"
          f"{'  [OK]' if os.path.isfile(TOPOLOGIA_PF) else '  [no encontrado — se usara inferencia]'}")
    print(f"  Salida     : {OUTPUT_PATH}")

    # -- Leer buses reales desde topologia (fuente primaria) ------------------
    print("\n[1/4] Cargando buses desde topologia_completa_pf.xlsx...")
    buses_2dev_topo, buses_3dev_topo = leer_buses_de_topologia(TOPOLOGIA_PF)
    print(f"      ElmTr2 con buses en topologia: {len(buses_2dev_topo)}"
          f"  |  ElmTr3: {len(buses_3dev_topo)}")
    if not buses_2dev_topo and not buses_3dev_topo:
        print("      [AVISO] Topologia no disponible o sin datos de transformadores.")
        print("      Se usara inferencia desde nombre (fallback).")
        print("      Para obtener buses reales: ejecutar TopologiaCompleta_PFV2.py primero.")

    # -- Leer datos de respaldo -----------------------------------------------
    print("[2/4] Leyendo barras y transformadores de DatosSINdigsilent...")
    sub_buses = leer_barras(DATOS_PF)
    n_sub = len({s for (s, _) in sub_buses})
    print(f"      {sum(len(v) for v in sub_buses.values())} barras en {n_sub} subestaciones")

    df_raw2 = leer_transformadores_2dev(DATOS_PF)
    df_raw3 = leer_transformadores_3dev(DATOS_PF)
    print(f"      {len(df_raw2)} transformadores 2dev  |  {len(df_raw3)} transformadores 3dev")

    # -- Construir tablas -----------------------------------------------------
    print("\n[3/4] Construyendo tablas (topologia primaria, inferencia de respaldo)...")
    df2    = construir_tabla_2dev(df_raw2, sub_buses, buses_2dev_topo)
    df3    = construir_tabla_3dev(df_raw3, sub_buses, buses_3dev_topo)
    df_res = construir_resumen(df2, df3)

    # Resumen de calidad de buses
    if "Fuente buses" in df2.columns:
        n2_t = (df2["Fuente buses"] == "topologia").sum()
        n2_i = (df2["Fuente buses"] == "inferido").sum()
        n2_sin = (df2["Barra HV"].isin(["", "-"])).sum()
        print(f"\n      2dev — topologia:{n2_t}  inferido:{n2_i}  sin_barra_HV:{n2_sin}")
    if "Fuente buses" in df3.columns:
        n3_t = (df3["Fuente buses"] == "topologia").sum()
        n3_i = (df3["Fuente buses"] == "inferido").sum()
        n3_sin = (df3["Barra HV"].isin(["", "-"])).sum()
        print(f"      3dev — topologia:{n3_t}  inferido:{n3_i}  sin_barra_HV:{n3_sin}")

    # Estadisticas
    print(f"\n      Transformadores 2 devanados : {len(df2)}")
    for nivel in ["500 kV", "230 kV", "115 kV", "69 kV", "34.5 kV", "Otro"]:
        n = (df2["Nivel tension HV"] == nivel).sum()
        if n:
            print(f"        {nivel:<10}: {n}")
    print(f"\n      Transformadores 3 devanados : {len(df3)}")
    for nivel in ["500 kV", "230 kV", "115 kV", "69 kV", "34.5 kV", "Otro"]:
        n = (df3["Nivel tension HV"] == nivel).sum()
        if n:
            print(f"        {nivel:<10}: {n}")

    n_sin_barra = (df2["Barra HV"] == "-").sum() + (df3["Barra HV"] == "-").sum()
    if n_sin_barra:
        print(f"\n      [AVISO] {n_sin_barra} transformadores sin barra HV identificada")

    n_fuera = (
        (df2["En servicio"].str.lower().isin(["no", "false", "0"])).sum() +
        (df3["En servicio"].str.lower().isin(["no", "false", "0"])).sum()
    )
    if n_fuera:
        print(f"      [AVISO] {n_fuera} transformadores fuera de servicio")

    # -- Exportar -------------------------------------------------------------
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"\n[4/4] Exportando a Excel...")

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        df2.to_excel(   writer, sheet_name="Transformadores_2dev", index=False)
        df3.to_excel(   writer, sheet_name="Transformadores_3dev", index=False)
        df_res.to_excel(writer, sheet_name="Resumen",              index=False)

    wb = load_workbook(OUTPUT_PATH)

    # Hoja 2dev
    cols2      = list(df2.columns)
    idx_nivel2 = cols2.index("Nivel tension HV") + 1
    idx_serv2  = cols2.index("En servicio") + 1
    _aplicar_hoja(wb["Transformadores_2dev"], idx_nivel2, idx_serv2, freeze="B2")
    _agregar_leyenda(wb["Transformadores_2dev"], wb["Transformadores_2dev"].max_row)

    # Hoja 3dev
    cols3      = list(df3.columns)
    idx_nivel3 = cols3.index("Nivel tension HV") + 1
    idx_serv3  = cols3.index("En servicio") + 1
    _aplicar_hoja(wb["Transformadores_3dev"], idx_nivel3, idx_serv3, freeze="B2")
    _agregar_leyenda(wb["Transformadores_3dev"], wb["Transformadores_3dev"].max_row)

    # Hoja Resumen
    cols_r    = list(df_res.columns)
    idx_tipo  = cols_r.index("Tipo") + 1
    ws_r = wb["Resumen"]
    # Color por nivel en columna Tipo
    def _nivel_de_tipo(r):
        tipo = str(ws_r.cell(r, idx_tipo).value or "")
        for nivel in COLOR_TENSION:
            if nivel in tipo:
                return nivel
        return "Otro"
    ws_r.row_dimensions[1].height = 30
    mc = ws_r.max_column
    for c in range(1, mc + 1):
        cell = ws_r.cell(1, c)
        cell.fill = _HFIL; cell.font = _HFNT
        cell.alignment = _CTR; cell.border = _BRD
    for r in range(2, ws_r.max_row + 1):
        nivel = _nivel_de_tipo(r)
        color = COLOR_TENSION.get(nivel, "FFFFFF")
        use_w = nivel in FONT_BLANCA
        f = _fill(color)
        for c in range(1, mc + 1):
            cell = ws_r.cell(r, c)
            cell.fill = f; cell.border = _BRD
            cell.font = _WFNT if use_w else _NFNT
            cell.alignment = _CTR if isinstance(cell.value, (int, float)) else _LEFT
    for col in ws_r.columns:
        hdr   = str(ws_r.cell(1, col[0].column).value or "").lower()
        ancho = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws_r.column_dimensions[get_column_letter(col[0].column)].width = (
            80 if "loc_names" in hdr else min(max(ancho + 3, 12), 45)
        )
    ws_r.freeze_panes = "A2"

    wb.save(OUTPUT_PATH)

    print(f"\n  Archivo creado en:")
    print(f"  {OUTPUT_PATH}")
    print(f"\n  Hojas:")
    for sh in wb.sheetnames:
        print(f"    {sh:<25} -> {wb[sh].max_row - 1} filas")


if __name__ == "__main__":
    main()
    input("\nPresiona Enter para cerrar...")
