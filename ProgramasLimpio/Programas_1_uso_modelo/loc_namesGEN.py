# =============================================================================
# MAPEO DE GENERADORES CNDC → loc_name PowerFactory
#
# Ejecutar UNA SOLA VEZ: los loc_names del modelo PF son fijos y no cambian
# entre eventos ni semestres. Re-ejecutar solo si se actualiza DatosSINdigsilent.
#
# Lee la lista maestra de unidades CNDC desde DatosSINdigsilent (hoja Generadores)
# o desde un datos_simulacion de referencia (configurar SIM_REF_PATH).
# Lee loc_names, barras y potencias del modelo PF (hoja Generadores).
#
# Estrategia de mapeo:
#   1. Tabla manual MAPA_CNDC_PF para hidro y renovables
#   2. Auto-match para termicas: extrae codigo (quita " - RF", " - PPG"),
#      convierte "CCERI30" -> "ERI30", busca sym_{codigo} en PF
#
# Nombres en la salida: limpios (sin sufijos " - RF", " - PPG", " - PPG.")
#   Esto permite emparejar con otros archivos que u
# sen el mismo nombre base.
#
# Salida: loc_names_gen.xlsx  en OUTPUT_DIR
#   Hoja Mapeo_Generadores -> una fila por unidad CNDC con loc_name(s) PF
#   Hoja Detalle_PF        -> una fila por loc_name PF con su unidad CNDC
#   Hoja Sin_Asignar       -> unidades CNDC sin loc_name en el modelo
#   Hoja PF_Sin_Mapeo      -> loc_names PF que no aparecen en ningun mapeo
#   Hoja LocNames_resumen  -> loc_names agrupados por tipo de generador
# =============================================================================

import os, re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATOS_PF      = r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\DatosSINdigsilent.xlsx"
DATOS_SIN_PATH = r"C:\Datos del CNDC\Datos_SIN_20251210.xls"
OUTPUT_DIR    = r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\Designacion de loc_name"
OUTPUT_PATH   = os.path.join(OUTPUT_DIR, "loc_names_gen.xlsx")

# Archivo de simulacion de referencia para obtener la lista de unidades CNDC.
# Puede ser cualquier evento; los nombres de unidades son consistentes entre eventos.
SIM_REF_PATH = (r"C:\Datos del CNDC\01_INFO CNDC_RPF\2024 sem2"
                r"\Análisis_todos_los_eventos\Evento 1"
                r"\datos_simulacion_210724_2daopcion.xlsx")

# =============================================================================
# LECTURA DE CODIGOS STI (Datos_SIN)
# =============================================================================

def leer_codigos_sin(path):
    """
    Lee la hoja GEN de Datos_SIN_*.xls y retorna {codigo: nombre_unidad}.
    La fila de encabezado esta en el indice 6; los datos comienzan en el indice 8.
    Columna 4 = CODIGO, columna 2 = NOMBRE.
    """
    import xlrd
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_name("GEN")
    resultado = {}
    for r in range(8, sh.nrows):
        codigo = str(sh.cell_value(r, 4)).strip()
        nombre = str(sh.cell_value(r, 2)).strip()
        if codigo and codigo not in ("nan", ""):
            resultado[codigo] = nombre
    return resultado


def codigo_desde_loc_name(loc_name, codigo_dict):
    """
    Extrae el codigo STI y el nombre de unidad para un loc_name PF.
    Estrategia:
      1. Quita prefijos: sym_, WT_, PV-, PV_, sta_
      2. Quita sufijo _EQ y parentesis: (1), (2)
      3. Busqueda exacta en codigo_dict
      4. Si no, busca el codigo mas largo del dict que sea prefijo del codigo extraido
         (ej. "ZON" es prefijo de "ZON01")
    Retorna (codigo, nombre) o ("-", "-") si no hay coincidencia.
    """
    s = re.sub(r'\(\d+\)$', '', loc_name.strip()).strip()
    for prefix in ("sym_", "WT_", "PV-", "PV_", "sta_"):
        if s.lower().startswith(prefix.lower()):
            s = s[len(prefix):]
            break
    s = re.sub(r'_EQ$', '', s, flags=re.IGNORECASE).strip()

    # 1. Exacto
    if s in codigo_dict:
        return s, codigo_dict[s]

    # 2. El codigo del dict es prefijo del codigo extraido (min 3 chars para evitar falsos)
    mejor = ""
    for codigo in codigo_dict:
        if s.startswith(codigo) and len(codigo) > len(mejor) and len(codigo) >= 3:
            mejor = codigo
    if mejor:
        return mejor, codigo_dict[mejor]

    return "-", "-"


# =============================================================================
# TABLA DE MAPEO MANUAL: nombre CNDC -> lista de loc_names PF
# Hidro y renovables no se deducen por codigo; se mapean explicitamente.
# =============================================================================
MAPA_CNDC_PF = {
    # -- SISTEMA ZONGO --------------------------------------------------------
    "ZONGO":         ["sym_ZON01", "sym_ZON01(1)"],
    "TIQUIMANI":     ["sym_TIQ01"],
    "BOTIJLACA":     ["sym_BOT01", "sym_BOT02", "sym_BOT03"],
    "CUTICUCHO":     ["sym_CUT01", "sym_CUT02", "sym_CUT03", "sym_CUT04", "sym_CUT05"],
    "SANTA ROSA 1":  ["sym_SRO01"],
    "SANTA ROSA 2":  ["sym_SRO02"],
    "SAINANI":       ["sym_SAI01"],
    "CHURURAQUI":    ["sym_CHU01", "sym_CHU02"],
    "HARCA":         ["sym_HAR01", "sym_HAR02"],
    "CAHUA":         ["sym_CAH01", "sym_CAH02"],
    "HUAJI":         ["sym_HUA01", "sym_HUA02"],
    # -- SISTEMA TAQUESI ------------------------------------------------------
    "CHOJLLA":       ["sym_CHJ01"],
    "YANACACHI":     ["sym_YAN01"],
    # -- SISTEMA CORANI -------------------------------------------------------
    "CORANI":        ["sym_COR01", "sym_COR02", "sym_COR03", "sym_COR04", "sym_COR05"],
    "SANTA ISABEL":  ["sym_SIS01", "sym_SIS02", "sym_SIS03", "sym_SIS04", "sym_SIS05"],
    "SAN JOSE 1":    ["sym_SJE01", "sym_SJE02"],   # SJE = San Jose Este/1
    "SAN JOSE 2":    ["sym_SJS01", "sym_SJS02"],   # SJS = San Jose Sur/2
    # -- SISTEMA MISICUNI -----------------------------------------------------
    "MISICUNI":      ["sym_MIS01", "sym_MIS02", "sym_MIS03"],
    # -- SISTEMA MIGUILLA -----------------------------------------------------
    "MIGUILLA":      ["sym_MIG01", "sym_MIG02"],
    "ANGOSTURA":     ["sym_ANG01", "sym_ANG02", "sym_ANG03"],
    "CHOQUETANGA":   ["sym_CHO01", "sym_CHO02", "sym_CHO03"],
    "CARABUCO":      ["sym_CRB01"],
    # -- SISTEMA YURA (sin representacion en modelo PF) -----------------------
    "KILLPANI":      [],   # no modelado en PF
    "LANDARA":       [],   # no modelado en PF
    "PUNUTUMA":      [],   # no modelado en PF
    # -- HIDRO STANDALONE -----------------------------------------------------
    "KANATA":        ["sym_KAN01"],
    "QUEHATA":       ["sym_QUE01", "sym_QUE02"],
    "SAN JACINTO":   ["sym_SJA01", "sym_SJA02"],   # SJA = San Jacinto
    # -- EOLICO ---------------------------------------------------------------
    "QOLLPANA I":    ["WT_QOL01_EQ"],
    "QOLLPANA II":   ["WT_QOL02_EQ"],
    "WARNES":        ["WT_EWA01_EQ"],
    "SAN JULIAN":    ["WT_SJU01_EQ"],
    "EL DORADO":     ["WT_EDO01_EQ"],
    # -- SOLAR ----------------------------------------------------------------
    "UYUNI":         ["PV_UYU_EQ", "PV_UYU02_EQ"],
    "YUNCHARA":      ["PV-YUNCHA-EQ", "sta_YUN024"],
    "ORURO I":       ["PV_ORU_EQ"],
    "ORURO II":      ["PV_ORU_II_EQ"],
}

# Número de máquinas en paralelo (ngnum) por loc_name PF para objetos ElmGenStat.
# Estos valores son fijos del modelo PowerFactory y no cambian entre eventos.
# Actualizar si se modifica el modelo PF (agregar/quitar máquinas).
NGNUM_PF = {
    "WT_QOL01_EQ":   6,    # Qollpana I  — completar con valor real del modelo
    "WT_QOL02_EQ":   6,    # Qollpana II — completar con valor real del modelo
    "WT_EWA01_EQ":   9,    # Warnes      — completar con valor real del modelo
    "WT_SJU01_EQ":   6,    # San Julian  — completar con valor real del modelo
    "WT_EDO01_EQ":   6,    # El Dorado   — completar con valor real del modelo
    "PV_UYU_EQ":    21,    # Uyuni (obj 1)
    "PV_UYU02_EQ":   1,    # Uyuni (obj 2) — completar con valor real del modelo
    "PV-YUNCHA-EQ":  2,    # Yunchara (obj 1) — completar
    "sta_YUN024":    1,    # Yunchara (obj 2) — completar
    "PV_ORU_EQ":     1,    # Oruro I     — completar con valor real del modelo
    "PV_ORU_II_EQ":  1,    # Oruro II    — completar con valor real del modelo
}

_HIDRO_NOMBRES  = set(MAPA_CNDC_PF) - {
    "QOLLPANA I","QOLLPANA II","WARNES","SAN JULIAN","EL DORADO",
    "UYUNI","YUNCHARA","ORURO I","ORURO II",
}
_EOLICO_NOMBRES = {"QOLLPANA I","QOLLPANA II","WARNES","SAN JULIAN","EL DORADO"}
_SOLAR_NOMBRES  = {"UYUNI","YUNCHARA","ORURO I","ORURO II"}

COLOR_TIPO = {
    "HIDRO":       "BDD7EE",
    "EOLICO":      "C5E0B4",
    "SOLAR":       "FFE699",
    "TERMO":       "FCE4D6",
    "Sin modelo":  "D9D9D9",
    "Sin asignar": "FFC7CE",
}

# =============================================================================
# HELPERS
# =============================================================================

def tipo_generador(nombre_cndc, loc_names_pf):
    if nombre_cndc in _HIDRO_NOMBRES:
        return "HIDRO"
    if nombre_cndc in _EOLICO_NOMBRES or any(lp.startswith("WT_") for lp in loc_names_pf):
        return "EOLICO"
    if nombre_cndc in _SOLAR_NOMBRES or any(
            lp.startswith("PV") or lp.startswith("sta_YUN") for lp in loc_names_pf):
        return "SOLAR"
    return "TERMO"


def limpiar_codigo_cndc(nombre):
    """
    Extrae el codigo base de una unidad termica CNDC:
      - Quita sufijos ' - RF', '  -RF', ' - PPG', ' - PPG.'
      - Convierte 'CCERI30' -> 'ERI30'
    """
    s = str(nombre).strip()
    s = re.sub(r'\s*[-]\s*(RF|PPG)\.?\s*$', '', s, flags=re.IGNORECASE).strip()
    s = re.sub(r'^CC(ERI\d+)$', r'\1', s)
    return s


def auto_match_termica(nombre_cndc, pf_lower_map):
    """
    Busca sym_{codigo} en el modelo PF.
    pf_lower_map: dict {nombre_pf.lower(): nombre_pf_original}
    Retorna (lista, metodo).
    """
    codigo    = limpiar_codigo_cndc(nombre_cndc)
    candidato = f"sym_{codigo}".lower()
    if candidato in pf_lower_map:
        return [pf_lower_map[candidato]], "auto_sym"
    return [], "sin_match"


# =============================================================================
# LECTURA DE DATOS
# =============================================================================

def leer_generadores_cndc(sim_path):
    df = pd.read_excel(sim_path, sheet_name="Generadores_pgini")
    return [str(r).strip() for r in df["Generador_CNDC"]
            if pd.notna(r) and str(r).strip()]


def leer_generadores_pf(datos_pf_path):
    df = pd.read_excel(datos_pf_path, sheet_name="Generadores")
    excluir = {"LEYENDA", "SVC Plus", "Synchronous Machine", "nan"}
    df = df[~df["Nombre"].astype(str).str.strip().isin(excluir)].copy()
    df = df[df["Nombre"].astype(str).str.strip().str.len() > 2].copy()
    df["Nombre"] = df["Nombre"].astype(str).str.strip()
    # ngnum viene del diccionario manual NGNUM_PF (no está en DatosSINdigsilent)
    df["ngnum"] = df["Nombre"].map(NGNUM_PF).fillna(1).astype(int)
    return df


# =============================================================================
# LOGICA DE MAPEO
# =============================================================================

def construir_mapeo(nombres_cndc, df_pf):
    pf_lower_map = {n.lower(): n for n in df_pf["Nombre"].tolist()}
    pf_info      = df_pf.set_index("Nombre").to_dict("index")

    filas = []
    for cndc_orig in nombres_cndc:
        # Nombre limpio: sin sufijos " - RF", " - PPG" (Reserva Fria / PPG)
        # Esto permite emparejar con otros archivos que usen el nombre base.
        cndc = limpiar_codigo_cndc(cndc_orig)
        if cndc in MAPA_CNDC_PF:
            loc_names     = MAPA_CNDC_PF[cndc]
            loc_names_ok  = [lp for lp in loc_names if lp in pf_info]
            loc_names_fal = [lp for lp in loc_names if lp not in pf_info]
            metodo = "manual"
        else:
            loc_names_ok, metodo = auto_match_termica(cndc, pf_lower_map)
            loc_names_fal = []

        tipo = tipo_generador(cndc, loc_names_ok)

        p_total = (
            sum((pf_info[lp].get("P nom. (MW)") or 0)
                for lp in loc_names_ok
                if isinstance(pf_info[lp].get("P nom. (MW)"), (int, float)))
            if loc_names_ok else None
        )
        # N unidades = suma de ngnum de cada loc_name (maquinas en paralelo dentro del objeto PF)
        n_unidades = sum(int(pf_info[lp].get("ngnum", 1) or 1) for lp in loc_names_ok)
        barras   = [str(pf_info[lp].get("Barra conectada","")) for lp in loc_names_ok]
        en_servs = [str(pf_info[lp].get("En servicio","?"))    for lp in loc_names_ok]

        filas.append({
            "Generador_CNDC":         cndc,
            "Tipo":                   tipo,
            "N unidades PF":          n_unidades if n_unidades > 0 else len(loc_names_ok),
            "loc_names PF":           ", ".join(loc_names_ok) if loc_names_ok else "-",
            "Metodo":                 metodo,
            "P nom. total PF (MW)":   round(p_total, 4) if p_total is not None else None,
            "En servicio PF":         ", ".join(en_servs) if en_servs else "-",
            "Barras PF":              ", ".join(b for b in barras if b)[:100],
            "loc_names no en modelo": ", ".join(loc_names_fal) if loc_names_fal else "",
        })

    return pd.DataFrame(filas)


def construir_detalle_pf(df_mapeo, df_pf, codigo_dict=None):
    """Una fila por loc_name PF, indicando a que unidad CNDC pertenece."""
    if codigo_dict is None:
        codigo_dict = {}
    lp_a_cndc = {}
    for _, row in df_mapeo.iterrows():
        for lp in str(row["loc_names PF"]).split(", "):
            lp = lp.strip()
            if lp and lp != "-":
                lp_a_cndc[lp] = row["Generador_CNDC"]

    filas = []
    for _, r in df_pf.iterrows():
        lp   = str(r["Nombre"]).strip()
        cndc = lp_a_cndc.get(lp, "-")
        tipo = "-"
        if cndc != "-":
            m = df_mapeo[df_mapeo["Generador_CNDC"] == cndc]
            tipo = m.iloc[0]["Tipo"] if not m.empty else "-"
        cod_sti, nom_sti = codigo_desde_loc_name(lp, codigo_dict)
        ngnum   = int(r.get("ngnum", 1) or 1)
        p_max   = r.get("P_max (MW)")
        try:
            pmax_total = round(float(p_max) * ngnum, 4) if p_max is not None else None
        except Exception:
            pmax_total = None
        filas.append({
            "loc_name PF":       lp,
            "Codigo STI":        cod_sti,
            "Nombre STI":        nom_sti,
            "Clase PF":          r.get("Clase PF", ""),
            "Barra conectada":   r.get("Barra conectada", ""),
            "P nom. (MW)":       r.get("P nom. (MW)"),
            "P_max (MW)":        p_max,
            "ngnum":             ngnum,
            "Pmax_total (MW)":   pmax_total,
            "En servicio":       r.get("En servicio", ""),
            "Generador_CNDC":    cndc,
            "Tipo":              tipo,
        })
    return pd.DataFrame(filas)


def construir_locnames_resumen(df_mapeo):
    """loc_names agrupados por tipo, con P total y lista completa."""
    filas = []
    for tipo in ["HIDRO", "EOLICO", "SOLAR", "TERMO", "Sin modelo"]:
        grp = (df_mapeo[df_mapeo["loc_names PF"] == "-"]
               if tipo == "Sin modelo"
               else df_mapeo[df_mapeo["Tipo"] == tipo])
        if grp.empty:
            continue
        all_lps = []
        for lps_str in grp["loc_names PF"]:
            for lp in str(lps_str).split(", "):
                lp = lp.strip()
                if lp and lp != "-":
                    all_lps.append(lp)
        filas.append({
            "Tipo":                 tipo,
            "N generadores CNDC":   len(grp),
            "N loc_names PF":       len(all_lps),
            "P nom. total PF (MW)": round(grp["P nom. total PF (MW)"].sum(), 3),
            "Generadores CNDC":     ", ".join(grp["Generador_CNDC"].tolist()),
            "loc_names PF":         ", ".join(all_lps),
        })
    return pd.DataFrame(filas)


# =============================================================================
# FORMATO EXCEL
# =============================================================================
_THIN = Side(border_style="thin", color="BFBFBF")
_BRD  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left",   vertical="center")
_HFIL = PatternFill("solid", start_color="1F3864")
_HFNT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_NFNT = Font(name="Arial", size=10)


def _fill(h): return PatternFill("solid", start_color=h, end_color=h)


def _aplicar_hoja(ws, col_tipo=None, col_serv=None, freeze="A2"):
    ws.row_dimensions[1].height = 30
    mc = ws.max_column
    for c in range(1, mc + 1):
        cell = ws.cell(1, c)
        cell.fill = _HFIL; cell.font = _HFNT
        cell.alignment = _CTR; cell.border = _BRD
    for r in range(2, ws.max_row + 1):
        tipo = str(ws.cell(r, col_tipo).value or "") if col_tipo else ""
        serv = str(ws.cell(r, col_serv).value or "") if col_serv else "Si"
        if col_serv and serv.lower() not in ("si", "true", "1"):
            color = "FFC7CE"
        else:
            color = COLOR_TIPO.get(tipo, "FFFFFF")
        f = _fill(color)
        for c in range(1, mc + 1):
            cell = ws.cell(r, c)
            cell.font = _NFNT; cell.border = _BRD; cell.fill = f
            cell.alignment = _CTR if isinstance(cell.value, (int, float)) else _LEFT
    for col in ws.columns:
        hdr   = str(ws.cell(1, col[0].column).value or "").lower()
        ancho = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        if any(k in hdr for k in ("loc_names", "generadores cndc", "barras")):
            ws.column_dimensions[get_column_letter(col[0].column)].width = 80
        else:
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ancho + 3, 12), 45)
    if freeze:
        ws.freeze_panes = freeze


def _agregar_leyenda(ws, fila_inicio):
    fila = fila_inicio + 2
    t = ws.cell(fila, 1, "LEYENDA — Tipo de generador")
    t.fill = _HFIL; t.font = _HFNT; t.alignment = _CTR; t.border = _BRD
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
    for tipo, color in COLOR_TIPO.items():
        fila += 1
        c1 = ws.cell(fila, 1, ""); c1.fill = _fill(color); c1.border = _BRD
        c2 = ws.cell(fila, 2, tipo)
        c2.font = _NFNT; c2.border = _BRD; c2.fill = _fill(color)
        c2.alignment = _LEFT


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 60)
    print("loc_namesGEN — Mapeo Generadores CNDC -> PF")
    print("(ejecutar UNA SOLA VEZ; resultado valido para todos los eventos)")
    print("=" * 60)

    if not os.path.isfile(SIM_REF_PATH):
        raise FileNotFoundError(
            f"No se encontro el archivo de referencia:\n  {SIM_REF_PATH}\n"
            f"Ajusta SIM_REF_PATH en la seccion de rutas.")

    print(f"\n  Referencia  : {os.path.basename(SIM_REF_PATH)}")
    print(f"  DatosPF     : {os.path.basename(DATOS_PF)}")
    print(f"  Salida      : {OUTPUT_PATH}")

    # -- Leer datos ----------------------------------------------------------
    print("\n[1/4] Leyendo generadores CNDC desde simulacion...")
    nombres_cndc = leer_generadores_cndc(SIM_REF_PATH)
    print(f"      {len(nombres_cndc)} generadores CNDC")

    print("[2/4] Leyendo modelo PF (DatosSINdigsilent)...")
    df_pf = leer_generadores_pf(DATOS_PF)
    print(f"      {len(df_pf)} loc_names en modelo PF")

    print(f"      Leyendo codigos STI ({os.path.basename(DATOS_SIN_PATH)})...")
    codigo_dict = leer_codigos_sin(DATOS_SIN_PATH) if os.path.isfile(DATOS_SIN_PATH) else {}
    print(f"      {len(codigo_dict)} codigos STI cargados")

    # -- Construir mapeo -----------------------------------------------------
    print("[3/4] Construyendo mapeo CNDC -> PF...")
    df_mapeo   = construir_mapeo(nombres_cndc, df_pf)
    df_detalle = construir_detalle_pf(df_mapeo, df_pf, codigo_dict)
    df_locs    = construir_locnames_resumen(df_mapeo)
    df_sin     = df_mapeo[df_mapeo["loc_names PF"] == "-"].copy()
    # Agregar Codigo STI a Sin_Asignar buscando el nombre CNDC directamente en el dict
    df_sin["Codigo STI"] = df_sin["Generador_CNDC"].apply(
        lambda n: codigo_dict.get(str(n).strip(), "-"))
    df_sin["Nombre STI"] = df_sin["Codigo STI"].apply(
        lambda c: codigo_dict.get(c, "-") if c != "-" else "-")
    df_pf_sin  = df_detalle[df_detalle["Generador_CNDC"] == "-"].copy()

    # Estadisticas
    n_manual = (df_mapeo["Metodo"] == "manual").sum()
    n_auto   = (df_mapeo["Metodo"] == "auto_sym").sum()
    n_sin    = (df_mapeo["loc_names PF"] == "-").sum()
    print(f"\n      Total CNDC           : {len(df_mapeo)}")
    print(f"      Mapeados (manual)    : {n_manual}")
    print(f"      Mapeados (auto_sym)  : {n_auto}")
    print(f"      Sin asignar          : {n_sin}")
    print(f"\n      Por tipo:")
    for tipo in ["HIDRO", "EOLICO", "SOLAR", "TERMO"]:
        n = (df_mapeo["Tipo"] == tipo).sum()
        if n:
            print(f"        {tipo:<10}: {n}")
    if n_sin:
        print(f"\n      [AVISO] Generadores CNDC sin loc_name en modelo PF:")
        for c in df_sin["Generador_CNDC"].tolist():
            print(f"        - {c}")
    if len(df_pf_sin):
        print(f"\n      loc_names PF sin mapeo CNDC : {len(df_pf_sin)}")

    # -- Exportar ------------------------------------------------------------
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print(f"\n[4/4] Exportando a Excel...")
    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        df_mapeo.to_excel(  writer, sheet_name="Mapeo_Generadores", index=False)
        df_detalle.to_excel(writer, sheet_name="Detalle_PF",        index=False)
        df_sin.to_excel(    writer, sheet_name="Sin_Asignar",       index=False)
        df_pf_sin.to_excel( writer, sheet_name="PF_Sin_Mapeo",      index=False)
        df_locs.to_excel(   writer, sheet_name="LocNames_resumen",  index=False)

    wb = load_workbook(OUTPUT_PATH)
    cols_m = list(df_mapeo.columns)
    _aplicar_hoja(wb["Mapeo_Generadores"],
                  col_tipo=cols_m.index("Tipo") + 1, freeze="C2")
    _agregar_leyenda(wb["Mapeo_Generadores"], wb["Mapeo_Generadores"].max_row)

    cols_d = list(df_detalle.columns)
    _aplicar_hoja(wb["Detalle_PF"],
                  col_tipo=cols_d.index("Tipo") + 1,
                  col_serv=cols_d.index("En servicio") + 1,
                  freeze="B2")

    if wb["Sin_Asignar"].max_row > 1:
        _aplicar_hoja(wb["Sin_Asignar"],
                      col_tipo=cols_m.index("Tipo") + 1, freeze="A2")

    if wb["PF_Sin_Mapeo"].max_row > 1:
        cols_ps = list(df_pf_sin.columns)
        _aplicar_hoja(wb["PF_Sin_Mapeo"],
                      col_tipo=cols_ps.index("Tipo") + 1,
                      col_serv=cols_ps.index("En servicio") + 1, freeze="A2")

    cols_l = list(df_locs.columns)
    _aplicar_hoja(wb["LocNames_resumen"],
                  col_tipo=cols_l.index("Tipo") + 1, freeze="A2")

    wb.save(OUTPUT_PATH)

    print(f"\n  Archivo creado en:")
    print(f"  {OUTPUT_PATH}")
    print(f"\n  Hojas:")
    for sh in wb.sheetnames:
        print(f"    {sh:<22} -> {wb[sh].max_row - 1} filas")


if __name__ == "__main__":
    main()
    input("\nPresiona Enter para cerrar...")
