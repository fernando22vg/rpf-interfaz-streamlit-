# =============================================================================
# MAPEO DE CARGAS PF → DISTRIBUIDOR / CONSUMIDOR NO REGULADO (CNDC)
# Versión 6 — trf_propietario + confianza + instructivo CNDC512_25
#
# Mejoras sobre v5:
#   V6-1 - Diccionario TRF_PROPIETARIO construido desde el instructivo oficial
#           CNDC512_25_InstructivosRestitucionSIN.pdf (fuente autoritativa)
#           + fallback automático desde NODO_SUBCLASE via prefijo de Barra HV/LV
#   V6-2 - Nueva capa de resolución "trf_propietario" con prioridad sobre zona_directa:
#           Para cada carga identifica el transformador inmediato aguas arriba
#           (BFS ≤ 3 saltos solo por barras BT/MT) y busca su propietario
#   V6-3 - Nueva jerarquía de métodos:
#           1.trf_propietario  2.prefijo_bus  3.subclase_directo
#           4.zona+prefijo     5.bfs_subclase 6.zona_directa/bfs_zona  7.sin_ruta
#   V6-4 - Columnas nuevas en Mapeo_Cargas:
#           "loc_name trf inmediato", "Propietario trf", "Confianza mapeo"
#   V6-5 - Hoja nueva "TRF_Propietario": diccionario completo con revisión
#   V6-6 - Hoja Resumen_Dist con columnas de confianza por distribuidor
#
# Fuente autoritativa de propietarios:
#   CNDC512_25_InstructivosRestitucionSIN.pdf (oct-2025)
#   Diferencias respecto al instructivo anterior:
#   - AT02 Kenko: cambia de DELAPAZ a ENDE TRANSMISION
#   - AT01 Mazocruz tap: 11→5 (dato técnico, no afecta propietario)
#   - S/E WAD Warnes 2: nueva subestación 115 kV (no tenía transformadores listados)
#   - Velarde → Velarde II: solo renombre descriptivo
#
# Entradas:
#   DatosSINdigsilent.xlsx  → Barras, Transformadores_2dev, Transformadores_3dev,
#                             Lineas, Cargas
#   loc_names_xfo.xlsx      → Transformadores_2dev, Transformadores_3dev (con barras HV/LV)
#   deener_*.xlsx           → MW por nodo en hora del evento
#   postot_*.xlsx           → MW retiros STI (opcional)
#
# Salidas: loc_name_cargas.xlsx
#   Mapeo_Cargas       → una fila por carga con distribuidor, método y confianza
#   Resumen_Dist       → totales por distribuidor + columnas de confianza
#   Curvas_LocNames    → loc_names + peso proporcional (scale0 PF)
#   TRF_Propietario    → diccionario transformador→propietario con revisión
#   Deener_Nodos       → nodos deener parseados con sub-distribuidor
# =============================================================================

import os, re
from collections import deque
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# RUTAS — ajustar según entorno
# =============================================================================
DATOS_PATH       = r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\DatosSINdigsilent.xlsx"
LOC_NAMES_XFO    = r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\Designacion de loc_name\loc_names_xfo.xlsx"
DEENER_PATH      = r"C:\Datos del CNDC\01_INFO CNDC_RPF\2025 sem2\Análisis_todos_los_eventos\Evento 1\Demanda de Energia y Potencia\deener_190725.xlsx"
POSTOT_PATH      = r""
OUTPUT_DIR       = r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\Designacion de loc_name"
OUTPUT_PATH      = os.path.join(OUTPUT_DIR, "loc_name_cargas.xlsx")

HOJA_POSTOT       = "MWh y Costo Marginal en Nodos"
UMBRAL_STI_KV     = 60.0
HORA_EVENTO_LABEL = "18:45"
BFS_TRF_MAX_SALTOS = 3   # máximo de saltos BT/MT para buscar trf inmediato aguas arriba

# =============================================================================
# ORDEN DEENER — secuencia canónica
# =============================================================================
ORDEN_DEENER = [
    "CRE - Santa Cruz", "CRE(Sist. Aislados) - S. Cruz", "EMDEECRUZ",
    "DELAPAZ - La Paz", "DELAPAZ - San Buenaventura", "DELAPAZ - Cumbre",
    "ELFEC - Cochabamba", "ELFEC - Chimor - Carrasco",
    "ENDE DEORURO - Jeruyo", "ENDE DEORURO - Oruro", "ENDE DEORURO - Catavi", "ENDE DEORURO - Lucianita",
    "CESSA - Sucre", "CESSA - Mariaca",
    "SEPSA - Potosi", "SEPSA - Punutuma", "SEPSA - Torre Huayco", "SEPSA - Portugalete",
    "SEPSA - Chilcobija", "SEPSA - Telamayu", "SEPSA - Don Diego",
    "ENDE DELBENI (3)", "ENDE - Varios (2)",
    "SETAR - Tarija", "SETAR - Villamontes", "SETAR - Yacuiba", "SETAR - Bermejo",
    "SAN CRISTOBAL - C. No Reg.", "Otros - C. No Regulados", "Varios (1)", "Sin asignar",
]

# =============================================================================
# MAPEO ZONA PF → DISTRIBUIDOR GRUPO
# =============================================================================
ZONA_DISTRIBUIDOR = {
    "Norte":    "DELAPAZ - La Paz",
    "Central":  "ELFEC - Cochabamba",
    "Oruro":    "ENDE DEORURO - Oruro",
    "Oriental": "CRE - Santa Cruz",
    "Sur":      "SEPSA - Potosi",
    "Sucre":    "CESSA - Sucre",
    "Beni":     "ENDE DELBENI (3)",
    "Litio":    "ENDE - Varios (2)",
    "Tarija":   "SETAR - Tarija",
    "Lomas":    "Otros - C. No Regulados",
    "Mina":     "SAN CRISTOBAL - C. No Reg.",
}

# =============================================================================
# NODO_SUBCLASE — barra STI → sub-distribuidor
# =============================================================================
NODO_SUBCLASE = {
    "CHI230": "ELFEC - Chimor - Carrasco",
    "JER115": "ENDE DEORURO - Jeruyo",
    "PAG115": "ENDE DEORURO - Oruro", "ORU115": "ENDE DEORURO - Oruro",
    "CAT069": "ENDE DEORURO - Catavi", "CAT115": "ENDE DEORURO - Catavi",
    "LUC115": "ENDE DEORURO - Lucianita",
    "VIN069": "ENDE DEORURO - Jeruyo", "VIN115": "ENDE DEORURO - Jeruyo",
    "ARJ069": "CESSA - Mariaca",  "MAR069": "CESSA - Mariaca",
    "SUC069": "CESSA - Sucre",    "SUC115": "CESSA - Sucre",
    "PUN069": "SEPSA - Punutuma",
    "THU069": "SEPSA - Torre Huayco",
    "POR069": "SEPSA - Portugalete",
    "CHL069": "SEPSA - Chilcobija",
    "TEL069": "SEPSA - Telamayu",  "TEL044": "SEPSA - Telamayu",
    "DDI069": "SEPSA - Don Diego",
    "POT069": "SEPSA - Potosi",   "POT115": "SEPSA - Potosi",
    "SAC115": "SEPSA - Potosi",   "OCU115": "SEPSA - Potosi",
    "PLA115": "SEPSA - Potosi",   "EPO115": "SEPSA - Potosi",
    "LIT115": "SEPSA - Potosi",   "KAR069": "SEPSA - Potosi",
    "SBU115": "DELAPAZ - San Buenaventura",
    "CHS115": "DELAPAZ - Cumbre",
    "CRN115": "DELAPAZ - La Paz", "KEN115": "DELAPAZ - La Paz",
    "CUM115": "DELAPAZ - La Paz", "HUA115": "DELAPAZ - La Paz",
    "MAZ230": "DELAPAZ - La Paz", "PCA115": "DELAPAZ - La Paz",
    "TAJ115": "SETAR - Tarija",
    "YAG069": "SETAR - Villamontes",
    "YAG230": "SETAR - Yacuiba",
    "BER115": "SETAR - Bermejo",
    "SBO115": "ENDE DELBENI (3)", "MOX115": "ENDE DELBENI (3)",
    "TRI115": "ENDE DELBENI (3)", "YUC115": "ENDE DELBENI (3)",
    "PRA115": "ENDE DELBENI (3)",
    "SAL115": "ENDE - Varios (2)", "LCA230": "ENDE - Varios (2)",
    "UYU230": "Varios (1)",
    "LIT230": "SAN CRISTOBAL - C. No Reg.",
    "ARB230": "Otros - C. No Regulados",
    "IRP115": "ELFEC - Cochabamba",
}

BUS_PREFIJO_CLASE = {
    "SCR": "SAN CRISTOBAL - C. No Reg.",
}

CONEXIONES_MANUALES = {
    "TRI024": "TRI115", "TRI02402": "TRI115", "MOA024": "MOX115",
    "CAT NEUTRO": "CAT115", "VSA115": "YAP230",
}

KEYWORD_CLASE = [
    ("CRE (MISIONES)",  "CRE(Sist. Aislados) - S. Cruz"),
    ("CRE (GUARAYOS)",  "CRE(Sist. Aislados) - S. Cruz"),
    ("CRE MISIONES",    "CRE(Sist. Aislados) - S. Cruz"),
    ("CRE GUARAYOS",    "CRE(Sist. Aislados) - S. Cruz"),
    ("EMDEECRUZ",       "EMDEECRUZ"),
    ("CRE",             "CRE - Santa Cruz"),
    ("ELFEC",           "ELFEC - Cochabamba"),
    ("ENDE DEORURO",    "ENDE DEORURO - Oruro"),
    ("CESSA",           "CESSA - Sucre"),
    ("SEPSA",           "SEPSA - Potosi"),
    ("DELAPAZ",         "DELAPAZ - La Paz"),
    ("DE LA PAZ",       "DELAPAZ - La Paz"),
    ("ENDE DELBENI",    "ENDE DELBENI (3)"),
    ("SETAR (BERMEJO)", "SETAR - Bermejo"),
    ("SETAR BERMEJO",   "SETAR - Bermejo"),
    ("SETAR VILLA",     "SETAR - Villamontes"),
    ("SETAR YACUIBA",   "SETAR - Yacuiba"),
    ("SETAR",           "SETAR - Tarija"),
    ("SAN CRISTOBAL",   "SAN CRISTOBAL - C. No Reg."),
    ("ENDE",            "ENDE - Varios (2)"),
    ("LAS LOMAS",       "Otros - C. No Regulados"),
    ("COBOCE",          "Otros - C. No Regulados"),
    ("EMPACAR",         "Otros - C. No Regulados"),
    ("EM VINTO",        "Otros - C. No Regulados"),
    ("CERAM",           "Otros - C. No Regulados"),
    ("LUTUM",           "Otros - C. No Regulados"),
]

# =============================================================================
# TRF_PROPIETARIO — fuente: CNDC512_25_InstructivosRestitucionSIN.pdf (oct-2025)
#
# Conversión de códigos PDF → loc_name PF:
#   TRXXX... → trf_XXX...  (quitar TR, agregar trf_)
#   ATXXX... → atr_XXX...  (quitar AT, agregar atr_)
# Propietarios del instructivo → ORDEN_DEENER:
#   DELAPAZ → DELAPAZ - La Paz  (sub-clase refinada via NODO_SUBCLASE)
#   CRE     → CRE - Santa Cruz
#   ELFEC   → ELFEC - Cochabamba  (sub-clase refinada via Chimor)
#   SEPSA   → SEPSA - Potosi      (sub-clase refinada via NODO_SUBCLASE)
#   CESSA   → CESSA - Sucre
#   SETAR   → SETAR - Tarija
#   ENDE / ENDE DELBENI → ENDE DELBENI (3)
#   COBOCE  → Otros - C. No Regulados
#   EMSC    → SAN CRISTOBAL - C. No Reg.
#   CM VINTO / ENDE DEORURO → ENDE DEORURO - Oruro
#   COBEE   → "COBEE"  (fuera del ORDEN_DEENER — generación, no distribución)
# =============================================================================

# Mapa propietario PDF → sub-distribuidor CNDC
_PROP_A_DIST = {
    "DELAPAZ":          "DELAPAZ - La Paz",   # se refina via NODO_SUBCLASE
    "CRE":              "CRE - Santa Cruz",
    "ELFEC":            "ELFEC - Cochabamba",  # se refina via NODO_SUBCLASE (CHI230)
    "SEPSA":            "SEPSA - Potosi",       # se refina via NODO_SUBCLASE
    "CESSA":            "CESSA - Sucre",        # se refina via NODO_SUBCLASE (ARJ069)
    "SETAR":            "SETAR - Tarija",
    "SETAR YACUIBA":    "SETAR - Yacuiba",
    "ENDE":             "ENDE DELBENI (3)",
    "ENDE DELBENI":     "ENDE DELBENI (3)",
    "ENDE CORANI":      "ELFEC - Cochabamba",  # transformadores de Corani en zona ELFEC
    "ENDE DEORURO":     "ENDE DEORURO - Oruro",
    "ENDE TRANSMISION": "ENDE - Varios (2)",   # transmisión troncal → no es distribución
    "ENDE TRASMISION":  "ENDE - Varios (2)",
    "ISABOL":           "Otros - C. No Regulados",
    "COBOCE":           "Otros - C. No Regulados",
    "EMSC":             "SAN CRISTOBAL - C. No Reg.",
    "CM VINTO":         "Otros - C. No Regulados",
    "HIDROBOL":         "Otros - C. No Regulados",
    "COBEE":            "Otros - C. No Regulados",
    "EMDEECRUZ":        "EMDEECRUZ",
}

# Diccionario manual derivado del instructivo CNDC512_25 (oct-2025)
# Formato: loc_name PF exacto → sub-distribuidor CNDC
# Nota: los transformadores de TRANSMISION se excluyen del trf_propietario
# porque sus barras LV son STI (≥60kV), no distribución.
TRF_PROPIETARIO: dict[str, str] = {
    # ── AREA NORTE — DELAPAZ ─────────────────────────────────────────────────
    "trf_AAC069":   "DELAPAZ - La Paz",
    "trf_AAR11501": "DELAPAZ - La Paz",
    "trf_AAR11502": "DELAPAZ - La Paz",
    "trf_ACH069":   "DELAPAZ - La Paz",
    "trf_ACI06902": "DELAPAZ - La Paz",
    "trf_ALP11501": "DELAPAZ - La Paz",
    "trf_ALP11502": "DELAPAZ - La Paz",
    "trf_BOL11501": "DELAPAZ - La Paz",
    "trf_BOL11502": "DELAPAZ - La Paz",
    "trf_CBA11501": "DELAPAZ - La Paz",
    "trf_CBA11502": "DELAPAZ - La Paz",
    "trf_CHA11501": "DELAPAZ - La Paz",
    "trf_CHA11502": "DELAPAZ - La Paz",
    "trf_CHG069":   "DELAPAZ - La Paz",
    "trf_CHQ11501": "DELAPAZ - La Paz",
    "trf_CHQ11502": "DELAPAZ - La Paz",
    "trf_CHS11501": "DELAPAZ - Cumbre",
    "trf_COS11501": "DELAPAZ - La Paz",
    "trf_COS11502": "DELAPAZ - La Paz",
    "trf_COT11501": "DELAPAZ - La Paz",
    "trf_COT11502": "DELAPAZ - La Paz",
    "trf_CRN11501": "DELAPAZ - La Paz",
    "trf_CRN11502": "DELAPAZ - La Paz",
    "trf_CTC11501": "DELAPAZ - La Paz",
    "trf_CTC11502": "DELAPAZ - La Paz",
    "trf_GUN11501": "DELAPAZ - La Paz",
    "trf_HUR069":   "DELAPAZ - La Paz",
    "trf_KEN06901": "DELAPAZ - La Paz",
    "trf_KEN06902": "DELAPAZ - La Paz",
    "trf_KEN06905": "DELAPAZ - La Paz",
    "trf_KEN11501": "DELAPAZ - La Paz",
    "trf_MAL11501": "DELAPAZ - La Paz",
    "trf_MUN06901": "DELAPAZ - La Paz",
    "trf_MUN06902": "DELAPAZ - La Paz",
    "trf_PAM11501": "DELAPAZ - La Paz",
    "trf_PAM11502": "DELAPAZ - La Paz",
    "trf_PCA11501": "DELAPAZ - La Paz",
    "trf_PIC115":   "DELAPAZ - La Paz",
    "trf_ROS11501": "DELAPAZ - La Paz",
    "trf_ROS11502": "DELAPAZ - La Paz",
    "trf_RSE11501": "DELAPAZ - La Paz",
    "trf_RSE11502": "DELAPAZ - La Paz",
    "trf_TAR06901": "DELAPAZ - La Paz",
    "trf_TAR06902": "DELAPAZ - La Paz",
    "trf_TEM06901": "DELAPAZ - La Paz",
    "trf_TEM06902": "DELAPAZ - La Paz",
    "trf_VIA11501": "DELAPAZ - La Paz",
    "trf_VIA11502": "DELAPAZ - La Paz",
    "trf_VIP11501": "DELAPAZ - La Paz",
    "trf_VIP11502": "DELAPAZ - La Paz",
    "trf_SBU11501": "DELAPAZ - San Buenaventura",
    "trf_VIO11501": "DELAPAZ - La Paz",
    "trf_ALI11501": "DELAPAZ - La Paz",
    "trf_SRQ06901": "DELAPAZ - La Paz",
    # AREA NORTE — ENDE (Beni)
    "trf_MOX11501": "ENDE DELBENI (3)",
    "trf_SBO11501": "ENDE DELBENI (3)",
    "trf_TRI11501": "ENDE DELBENI (3)",
    "trf_TRI11502": "ENDE DELBENI (3)",
    "trf_YUC11501": "ENDE DELBENI (3)",
    "trf_PRA11502": "ENDE DELBENI (3)",
    # AREA NORTE — COBEE
    "trf_TIQ115":   "Otros - C. No Regulados",
    # AREA NORTE — HIDROBOL
    "trf_CJL024":   "Otros - C. No Regulados",
    # ── AREA CENTRAL-ORIENTAL — CRE ──────────────────────────────────────────
    "trf_ARB11501": "CRE - Santa Cruz",
    "trf_ARB11502": "CRE - Santa Cruz",
    "trf_BRE11501": "CRE - Santa Cruz",
    "trf_BRE11502": "CRE - Santa Cruz",
    "trf_CAN06901": "CRE - Santa Cruz",
    "trf_CAN06902": "CRE - Santa Cruz",
    "trf_CHN11501": "CRE - Santa Cruz",
    "trf_FER069":   "CRE - Santa Cruz",
    "trf_GCH06903": "CRE - Santa Cruz",
    "trf_GUP11501": "CRE - Santa Cruz",
    "trf_GUP11502": "CRE - Santa Cruz",
    "trf_MAP06901": "CRE - Santa Cruz",
    "trf_MAP06902": "CRE - Santa Cruz",
    "trf_MON11501": "CRE - Santa Cruz",
    "trf_MON11502": "CRE - Santa Cruz",
    "trf_MON11503": "CRE - Santa Cruz",
    "trf_NJE06901": "CRE - Santa Cruz",
    "trf_NJE06902": "CRE - Santa Cruz",
    "trf_PAL06901": "CRE - Santa Cruz",
    "trf_PAL06902": "CRE - Santa Cruz",
    "trf_PAR06901": "CRE - Santa Cruz",
    "trf_PAR06902": "CRE - Santa Cruz",
    "trf_PIN06901": "CRE - Santa Cruz",
    "trf_PIN06902": "CRE - Santa Cruz",
    "trf_PLM06901": "CRE - Santa Cruz",
    "trf_PMA06901": "CRE - Santa Cruz",
    "trf_PMA06902": "CRE - Santa Cruz",
    "trf_TRN11501": "CRE - Santa Cruz",
    "trf_TRN11502": "CRE - Santa Cruz",
    "trf_TRO06901": "CRE - Santa Cruz",
    "trf_TRO06902": "CRE - Santa Cruz",
    # AREA CENTRAL-ORIENTAL — EMDEECRUZ
    "trf_WAR11501": "EMDEECRUZ",
    # AREA CENTRAL-ORIENTAL — ELFEC
    "trf_ALA11501": "ELFEC - Cochabamba",
    "trf_ALA11502": "ELFEC - Cochabamba",
    "trf_ARO11501": "ELFEC - Cochabamba",
    "trf_CAL11501": "ELFEC - Cochabamba",
    "trf_CAL11502": "ELFEC - Cochabamba",
    "trf_CAR23004": "ELFEC - Cochabamba",
    "trf_CEN11501": "ELFEC - Cochabamba",
    "trf_CEN11502": "ELFEC - Cochabamba",
    "trf_CEN11503": "ELFEC - Cochabamba",
    "trf_CHI23001": "ELFEC - Chimor - Carrasco",
    "trf_CHI23002": "ELFEC - Chimor - Carrasco",
    "trf_COL11501": "ELFEC - Cochabamba",
    "trf_IRP115":   "ELFEC - Cochabamba",
    "trf_PAY11501": "ELFEC - Cochabamba",
    "trf_QOL11501": "ELFEC - Cochabamba",
    "trf_QOL11502": "ELFEC - Cochabamba",
    "trf_QOL11503": "ELFEC - Cochabamba",
    "trf_QUI11501": "ELFEC - Cochabamba",
    "trf_QUI11502": "ELFEC - Cochabamba",
    "trf_SAN11501": "ELFEC - Cochabamba",
    "trf_SJO11501": "ELFEC - Cochabamba",
    # AREA CENTRAL-ORIENTAL — ENDE CORANI
    "trf_COR11506": "ELFEC - Cochabamba",
    # ── AREA SUR — SEPSA ─────────────────────────────────────────────────────
    "trf_CHL069":   "SEPSA - Chilcobija",
    "trf_DDI06901": "SEPSA - Don Diego",
    "trf_DDI06902": "SEPSA - Don Diego",
    "trf_DDI06903": "SEPSA - Don Diego",
    "trf_EPO11501": "SEPSA - Potosi",
    "trf_EPO11502": "SEPSA - Potosi",
    "trf_KIL02401": "SEPSA - Potosi",
    "trf_KIL04401": "SEPSA - Potosi",
    "trf_KIL04402": "SEPSA - Potosi",
    "trf_KIL04403": "SEPSA - Potosi",
    "trf_KIL04404": "SEPSA - Potosi",
    "trf_LAG06901": "SEPSA - Potosi",
    "trf_LAN06903": "SEPSA - Potosi",
    "trf_LAN003":   "SEPSA - Potosi",
    "trf_PLA11501": "SEPSA - Potosi",
    "trf_POT06901": "SEPSA - Potosi",
    "trf_PUH069":   "SEPSA - Punutuma",
    "trf_PUN069":   "SEPSA - Punutuma",
    "trf_QLC069":   "SEPSA - Potosi",
    "trf_SAC115":   "SEPSA - Potosi",
    "trf_SOC069":   "SEPSA - Potosi",
    "trf_SUC06901": "SEPSA - Potosi",
    "trf_SUC11501": "SEPSA - Potosi",
    "trf_TAM06902": "SEPSA - Potosi",
    "trf_TUP06901": "SEPSA - Potosi",
    "trf_TUP06902": "SEPSA - Potosi",
    "trf_VEL11501": "SEPSA - Potosi",
    "trf_VEL11502": "SEPSA - Potosi",
    "trf_VEL11503": "SEPSA - Potosi",
    "trf_VEL11504": "SEPSA - Potosi",
    "trf_OCU115":   "SEPSA - Potosi",
    # AREA SUR — CESSA
    "trf_ARJ06904": "CESSA - Mariaca",
    "trf_LAN06901": "CESSA - Sucre",
    "trf_PAD11501": "CESSA - Sucre",
    "trf_SUD06902": "CESSA - Sucre",
    "trf_SUD06903": "CESSA - Sucre",
    # AREA SUR — ENDE DEORURO
    "trf_AVI06901": "ENDE DEORURO - Oruro",
    "trf_AVI06902": "ENDE DEORURO - Oruro",
    "trf_BLV06901": "ENDE DEORURO - Oruro",
    "trf_BLV06902": "ENDE DEORURO - Oruro",
    "trf_BOM06901": "ENDE DEORURO - Jeruyo",
    "trf_CRC069":   "ENDE DEORURO - Oruro",
    "trf_CLQ069":   "ENDE DEORURO - Oruro",
    "trf_CRQ06901": "ENDE DEORURO - Oruro",
    "trf_CRQ06902": "ENDE DEORURO - Oruro",
    "trf_CSG069":   "ENDE DEORURO - Catavi",
    "trf_EST06901": "ENDE DEORURO - Oruro",
    "trf_EST06902": "ENDE DEORURO - Oruro",
    "trf_HUN06901": "ENDE DEORURO - Catavi",
    "trf_HUN06902": "ENDE DEORURO - Catavi",
    "trf_HUY06901": "ENDE DEORURO - Oruro",
    "trf_INT11501": "ENDE DEORURO - Oruro",
    "trf_INT11502": "ENDE DEORURO - Oruro",
    "trf_MCH069":   "ENDE DEORURO - Catavi",
    "trf_NOR06901": "ENDE DEORURO - Oruro",
    "trf_NOR06902": "ENDE DEORURO - Oruro",
    "trf_PAI069":   "ENDE DEORURO - Oruro",
    "trf_PAG11501": "ENDE DEORURO - Oruro",
    "trf_POR06901": "SEPSA - Portugalete",
    "trf_SBA06901": "ENDE DEORURO - Oruro",
    "trf_SBA06902": "ENDE DEORURO - Oruro",
    "trf_SUD06901": "ENDE DEORURO - Oruro",
    "trf_TAB069":   "ENDE DEORURO - Oruro",
    "trf_TAM06901": "ENDE DEORURO - Oruro",
    "trf_TES069":   "ENDE DEORURO - Oruro",
    "trf_LUC11501": "ENDE DEORURO - Lucianita",
    "trf_LUC11502": "ENDE DEORURO - Lucianita",
    # AREA SUR — SETAR
    "trf_TAJ11501": "SETAR - Tarija",
    "trf_TAJ11502": "SETAR - Tarija",
    "trf_VAB11501": "SETAR - Yacuiba",
    "trf_VAB11502": "SETAR - Yacuiba",
    "trf_VAB11503": "SETAR - Yacuiba",
    "trf_BER11501": "SETAR - Bermejo",
    "trf_CAZ06901": "SETAR - Yacuiba",
    # AREA SUR — ENDE (Varios)
    "trf_LCA23001": "ENDE - Varios (2)",
    # AREA SUR — COBOCE
    "trf_COB11501": "Otros - C. No Regulados",
    "trf_COB11502": "Otros - C. No Regulados",
    "trf_COB11503": "Otros - C. No Regulados",
    # AREA SUR — EMSC (San Cristóbal)
    "trf_SCR23001": "SAN CRISTOBAL - C. No Reg.",
    "trf_SCR23002": "SAN CRISTOBAL - C. No Reg.",
    # AREA SUR — CM VINTO
    "trf_CMV06901": "Otros - C. No Regulados",
    "trf_CMV06902": "Otros - C. No Regulados",
    # AREA SUR — MARIACA
    "trf_MAR069":   "CESSA - Mariaca",
}
# Normalizar todas las claves a minúsculas para que el lookup via .lower() funcione
# (construir_grafo guarda loc_names con .lower(); sin esto TRF_PROPIETARIO.get() siempre falla)
TRF_PROPIETARIO = {k.lower(): v for k, v in TRF_PROPIETARIO.items()}

# =============================================================================
# DEENER NODO → SUB-DISTRIBUIDOR (para parseo Formato 2)
# =============================================================================
def _norm(s):
    s = str(s).lower().strip()
    for a, b in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),("ñ","n"),("ü","u")]:
        s = s.replace(a, b)
    s = re.sub(r'\s*\b(115|230|069|024|kv)\b', '', s).strip()
    return s

DEENER_NODO_A_SUBDIST = {
    "guaracachi": "CRE - Santa Cruz",
    "urubu":      "CRE - Santa Cruz",
    "urubo":      "CRE - Santa Cruz",
    "arboleda":   "CRE - Santa Cruz",
    "warnes":     "CRE - Santa Cruz",
    "brechas":    "CRE - Santa Cruz",
    "yapacani":   "CRE - Santa Cruz",
    "belgica":    "CRE - Santa Cruz",
    "san julian": "CRE - Santa Cruz",
    "camiri":     "CRE - Santa Cruz",
    "los troncos":              "CRE - Santa Cruz",
    "los troncos las misiones": "CRE(Sist. Aislados) - S. Cruz",
    "guarayos":                 "CRE(Sist. Aislados) - S. Cruz",
    "kenko":            "DELAPAZ - La Paz",
    "cumbre":           "DELAPAZ - Cumbre",
    "chuspipata":       "DELAPAZ - Cumbre",
    "caranavi":         "DELAPAZ - La Paz",
    "san buenaventura": "DELAPAZ - San Buenaventura",
    "palca":            "DELAPAZ - La Paz",
    "contorno bajo":    "DELAPAZ - La Paz",
    "choquetanga":      "DELAPAZ - La Paz",
    "arocagua":         "ELFEC - Cochabamba",
    "valle hermoso":    "ELFEC - Cochabamba",
    "irpa irpa":        "ELFEC - Cochabamba",
    "chimore":          "ELFEC - Chimor - Carrasco",
    "san jose":         "ELFEC - Cochabamba",
    "paracaya":         "ELFEC - Cochabamba",
    "carrasco":         "ELFEC - Chimor - Carrasco",
    "qollpana":         "ELFEC - Cochabamba",
    "villa tunari":     "ELFEC - Chimor - Carrasco",
    "santivanez":       "ELFEC - Cochabamba",
    "vinto":            "ENDE DEORURO - Jeruyo",
    "catavi":           "ENDE DEORURO - Catavi",
    "jeruyo":           "ENDE DEORURO - Jeruyo",
    "lucianita":        "ENDE DEORURO - Lucianita",
    "pagador":          "ENDE DEORURO - Oruro",
    "oruro":            "ENDE DEORURO - Oruro",
    "sacaca":           "SEPSA - Potosi",
    "ocuri":            "SEPSA - Potosi",
    "potosi":           "SEPSA - Potosi",
    "punutuma":         "SEPSA - Punutuma",
    "don diego":        "SEPSA - Don Diego",
    "cm karachipampa":  "SEPSA - Potosi",
    "karachipampa":     "SEPSA - Potosi",
    "litio":            "SEPSA - Potosi",
    "torre huayco":     "SEPSA - Torre Huayco",
    "portugalete":      "SEPSA - Portugalete",
    "chilcobija":       "SEPSA - Chilcobija",
    "telamayu":         "SEPSA - Telamayu",
    "la plata":         "SEPSA - Potosi",
    "ecebol - potosi":  "SEPSA - Potosi",
    "ecebol":           "SEPSA - Potosi",
    "sucre":            "CESSA - Sucre",
    "sucre - fancesa":  "CESSA - Sucre",
    "fancesa":          "CESSA - Sucre",
    "mariaca":          "CESSA - Mariaca",
    "tazna":            "ENDE - Varios (2)",
    "uyuni":            "Varios (1)",
    "carreras":         "ENDE - Varios (2)",
    "tarija":           "SETAR - Tarija",
    "villamontes":      "SETAR - Villamontes",
    "yacuiba":          "SETAR - Yacuiba",
    "bermejo":          "SETAR - Bermejo",
    "yucumo":           "ENDE DELBENI (3)",
    "san borja":        "ENDE DELBENI (3)",
    "san ignacio de moxos": "ENDE DELBENI (3)",
    "san ignacio":      "ENDE DELBENI (3)",
    "trinidad":         "ENDE DELBENI (3)",
    "paraiso":          "ENDE DELBENI (3)",
    "emdeecruz":        "EMDEECRUZ",
    "emvinto - comibol":"Otros - C. No Regulados",
    "emvinto":          "Otros - C. No Regulados",
    "coboce":           "Otros - C. No Regulados",
    "san cristobal":    "SAN CRISTOBAL - C. No Reg.",
    "retiros ende para ylb": "ENDE - Varios (2)",
    "las lomas":        "Otros - C. No Regulados",
    "ceramica guadalquivir": "Otros - C. No Regulados",
    "ceramica":         "Otros - C. No Regulados",
    "empacar":          "Otros - C. No Regulados",
}

# Colores por clase
COLOR_CLASE = {
    "CRE - Santa Cruz":              "BDD7EE",
    "CRE(Sist. Aislados) - S. Cruz": "DDEEFF",
    "EMDEECRUZ":                     "C5E0B4",
    "DELAPAZ - La Paz":              "FFE699",
    "DELAPAZ - San Buenaventura":    "FFF2CC",
    "DELAPAZ - Cumbre":              "FFEB9C",
    "ELFEC - Cochabamba":            "FCE4D6",
    "ELFEC - Chimor - Carrasco":     "F8CBAD",
    "ENDE DEORURO - Jeruyo":         "E2EFDA",
    "ENDE DEORURO - Oruro":          "C6EFCE",
    "ENDE DEORURO - Catavi":         "A9D18E",
    "ENDE DEORURO - Lucianita":      "D6E4F0",
    "CESSA - Sucre":                 "EAD1DC",
    "CESSA - Mariaca":               "F4CCCC",
    "SEPSA - Potosi":                "D9E1F2",
    "SEPSA - Punutuma":              "B4C6E7",
    "SEPSA - Torre Huayco":          "9DC3E6",
    "SEPSA - Portugalete":           "CFE2F3",
    "SEPSA - Chilcobija":            "BDD7EE",
    "SEPSA - Telamayu":              "DAEEF3",
    "SEPSA - Don Diego":             "C5DCF5",
    "ENDE DELBENI (3)":              "F4CCCC",
    "ENDE - Varios (2)":             "D9D9D9",
    "SETAR - Tarija":                "FDE9D9",
    "SETAR - Villamontes":           "FCE4D6",
    "SETAR - Yacuiba":               "F8CBAD",
    "SETAR - Bermejo":               "F4B8B8",
    "SAN CRISTOBAL - C. No Reg.":    "D9B3FF",
    "Otros - C. No Regulados":       "EDEDED",
    "Varios (1)":                    "BFBFBF",
    "Sin asignar":                   "FFC7CE",
}
COLOR_DEFAULT = "FFFFFF"
COLOR_CONFIANZA = {"Alta": "C6EFCE", "Media": "FFEB9C", "Baja": "FFC7CE"}

ANUARIO_GRUPO = {
    "CRE - Santa Cruz": 728.82, "CRE(Sist. Aislados) - S. Cruz": 728.82,
    "EMDEECRUZ": 4.83,
    "DELAPAZ - La Paz": 331.88, "DELAPAZ - San Buenaventura": 331.88, "DELAPAZ - Cumbre": 331.88,
    "ELFEC - Cochabamba": 258.83, "ELFEC - Chimor - Carrasco": 258.83,
    "ENDE DEORURO - Jeruyo": 95.16, "ENDE DEORURO - Oruro": 95.16,
    "ENDE DEORURO - Catavi": 95.16, "ENDE DEORURO - Lucianita": 95.16,
    "CESSA - Sucre": 51.06, "CESSA - Mariaca": 51.06,
    "SEPSA - Potosi": 103.92, "SEPSA - Punutuma": 103.92, "SEPSA - Torre Huayco": 103.92,
    "SEPSA - Portugalete": 103.92, "SEPSA - Chilcobija": 103.92,
    "SEPSA - Telamayu": 103.92, "SEPSA - Don Diego": 103.92,
    "ENDE DELBENI (3)": 43.11, "ENDE - Varios (2)": 8.19,
    "SETAR - Tarija": 66.99, "SETAR - Villamontes": 66.99,
    "SETAR - Yacuiba": 66.99, "SETAR - Bermejo": 66.99,
    "SAN CRISTOBAL - C. No Reg.": 59.23, "Otros - C. No Regulados": 59.23, "Varios (1)": 59.23,
}
ANUARIO_TOTAL = 1752.0

RE_PREFIX_BUS = re.compile(r'^([A-Za-zÑñÀ-ÿ_]+)')
RE_TRF_NAME   = re.compile(r'(?:trf|atr)_([A-Za-zÑñÀ-ÿ_]+)(\d{3})', re.IGNORECASE)

# =============================================================================
# CONSTRUIR TRF_PROPIETARIO AUTOMÁTICO desde loc_names_xfo
# Complementa el diccionario manual para transformadores no listados en el PDF
# =============================================================================
def _construir_trf_propietario_auto(df_xfo, trf_manual, nodo_subclase):
    """
    Para cada trf en df_xfo no presente en trf_manual:
      1. Busca prefijo de Barra HV en NODO_SUBCLASE
      2. Si no, busca prefijo de Barra LV en NODO_SUBCLASE
      3. Si no, usa ZONA_DISTRIBUIDOR via prefijo
    Retorna dict ampliado y lista de filas para hoja TRF_Propietario.
    """
    resultado = dict(trf_manual)
    filas_hoja = []

    for _, row in df_xfo.iterrows():
        loc = str(row.get("loc_name", "") or "").strip().lower()
        if not loc:
            continue
        barra_hv = str(row.get("Barra HV", "") or "").strip()
        barra_lv = str(row.get("Barra LV", "") or "").strip()
        kv_hv    = row.get("Tension HV nom. (kV)", 0) or 0
        kv_lv    = row.get("Tension LV nom. (kV)", 0) or 0
        pot      = row.get("Potencia nom. (MVA)", "")
        en_serv  = str(row.get("En servicio", "Si") or "Si").strip()

        ya_manual  = loc in trf_manual
        dist_auto  = None
        met_auto   = ""
        rev        = "No"

        if not ya_manual:
            # Buscar por prefijo de Barra HV
            m = RE_PREFIX_BUS.match(barra_hv)
            if m:
                px = m.group(1).upper()
                # Exacto en NODO_SUBCLASE
                if barra_hv in nodo_subclase:
                    dist_auto = nodo_subclase[barra_hv]
                    met_auto  = "nodo_subclase_HV"
                else:
                    # Por prefijo
                    for key, val in nodo_subclase.items():
                        if key.startswith(px):
                            dist_auto = val
                            met_auto  = "prefijo_HV"
                            rev = "Si"
                            break
            # Si no, buscar por prefijo de Barra LV
            if dist_auto is None:
                m2 = RE_PREFIX_BUS.match(barra_lv)
                if m2:
                    px2 = m2.group(1).upper()
                    if barra_lv in nodo_subclase:
                        dist_auto = nodo_subclase[barra_lv]
                        met_auto  = "nodo_subclase_LV"
                    else:
                        for key, val in nodo_subclase.items():
                            if key.startswith(px2):
                                dist_auto = val
                                met_auto  = "prefijo_LV"
                                rev = "Si"
                                break

            if dist_auto:
                resultado[loc] = dist_auto
            else:
                met_auto = "requiere_instructivo"
                rev = "Si"

        dist_final = resultado.get(loc, dist_auto or "Sin asignar")
        filas_hoja.append({
            "loc_name trf":           loc,
            "Barra HV":               barra_hv,
            "Barra LV":               barra_lv,
            "kV HV":                  kv_hv,
            "kV LV":                  kv_lv,
            "Potencia MVA":           pot,
            "En servicio":            en_serv,
            "Distribuidor asignado":  dist_final,
            "Metodo asignacion":      "instructivo_CNDC512" if ya_manual else met_auto,
            "Requiere revision":      "No" if ya_manual else rev,
        })

    return resultado, filas_hoja


# =============================================================================
# PARSER DEENER — auto-detecta Formato 1 (plano) vs Formato 2 (agrupado)
# =============================================================================
def parsear_deener(path, hora_label="18:45"):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    hdr_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and any(str(v).strip() in ("01:00","01:00:00") for v in row if v):
            hdr_row = i; break
    if hdr_row is None:
        print("  [ERROR] No se encontró fila de encabezado en deener.")
        return {}, "desconocido", [], []

    headers = [str(ws.cell(hdr_row, c).value).strip()
               for c in range(1, ws.max_column + 1)]

    col_hora = None
    hora_sin = hora_label[:5]
    for j, h in enumerate(headers):
        if h == hora_sin or h.startswith(hora_sin):
            col_hora = j + 1; break
    if col_hora is None:
        try:
            hora_h = int(hora_sin.split(":")[0])
            target = f"{hora_h:02d}:00"
            for j, h in enumerate(headers):
                if h == target:
                    col_hora = j + 1; break
        except Exception:
            pass
    if col_hora is None:
        print(f"  [AVISO] Hora '{hora_label}' no encontrada en deener.")

    formato = "F1_plano"
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if str(row[0] or "").strip().upper().startswith("TOTAL -"):
            formato = "F2_agrupado"; break

    mw_subdist = {}; nodos_raw = []; sin_mapa = []

    if formato == "F1_plano":
        for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            nombre = str(row[0] or "").strip()
            if not nombre or nombre.upper() in ("TOTAL","TOTAL COINCIDENTAL","NAN"):
                continue
            subdist = _match_orden_deener(nombre)
            if subdist is None:
                sin_mapa.append(nombre); subdist = nombre
            mw = _get_mw_row(row, col_hora)
            mw_subdist[subdist] = mw_subdist.get(subdist, 0.0) + mw
            nodos_raw.append({"nombre": nombre, "subdist": subdist, "mw": mw})
    else:
        grupo_actual = None
        for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            nombre = str(row[0] or "").strip()
            if not nombre or nombre.upper() == "NAN":
                continue
            if nombre.upper().startswith("TOTAL -"):
                grupo_actual = nombre[len("TOTAL -"):].strip(); continue
            if nombre.upper() == "TOTAL":
                continue
            subdist = DEENER_NODO_A_SUBDIST.get(_norm(nombre))
            if subdist is None:
                if grupo_actual:
                    subdist = _grupo_deener_a_clase(grupo_actual)
                if subdist is None:
                    sin_mapa.append(nombre); subdist = "Sin asignar"
            mw = _get_mw_row(row, col_hora)
            mw_subdist[subdist] = mw_subdist.get(subdist, 0.0) + mw
            nodos_raw.append({"nombre": nombre, "subdist": subdist, "mw": mw,
                              "grupo": grupo_actual or ""})

    return mw_subdist, formato, nodos_raw, sin_mapa


def _get_mw_row(row, col_hora):
    if col_hora and col_hora - 1 < len(row):
        v = row[col_hora - 1]
        if isinstance(v, (int, float)):
            return round(float(v), 4)
    for v in reversed(row[1:]):
        if isinstance(v, (int, float)) and v > 0:
            return round(float(v), 4)
    return 0.0

def _match_orden_deener(nombre):
    nb = nombre.strip()
    if nb in ORDEN_DEENER: return nb
    nb_up = nb.upper()
    for d in ORDEN_DEENER:
        if d.upper() == nb_up: return d
    return None

def _grupo_deener_a_clase(grupo):
    g = grupo.upper().strip()
    mapping = {
        "CRE": "CRE - Santa Cruz", "DELAPAZ": "DELAPAZ - La Paz",
        "ELFEC": "ELFEC - Cochabamba", "ENDE DEORURO": "ENDE DEORURO - Oruro",
        "SEPSA": "SEPSA - Potosi", "CESSA": "CESSA - Sucre",
        "ENDE": "ENDE - Varios (2)", "SETAR": "SETAR - Tarija",
        "ENDEDELBENI": "ENDE DELBENI (3)", "ENDE DELBENI": "ENDE DELBENI (3)",
    }
    for k, v in mapping.items():
        if g == k or g.startswith(k): return v
    return None


# =============================================================================
# POSTOT (opcional)
# =============================================================================
def parsear_postot(path, hoja, hora_label=None):
    if not path or not os.path.isfile(path): return {}, {}, None
    df = pd.read_excel(path, sheet_name=hoja, header=None)
    hdr_idx = None
    for i, row in df.iterrows():
        if str(row.iloc[0]).strip() == "NODO": hdr_idx = i; break
    if hdr_idx is None: return {}, {}, None
    headers  = list(df.iloc[hdr_idx])
    col_hora = next((j for j, h in enumerate(headers) if str(h).strip() == hora_label), None) if hora_label else None
    nodo_actual = None; nrm = {}; nmm = {}
    for i in range(hdr_idx + 1, len(df)):
        row  = df.iloc[i]
        col0 = str(row.iloc[0]).strip()
        col1 = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
        if col0 not in ("nan","","NODO","TOTALES") and not col0.startswith("Comit"):
            nodo_actual = col0; nrm.setdefault(col0, []); nmm.setdefault(col0, {})
        if nodo_actual and re.match(r'RETIRO', col1, re.IGNORECASE):
            e = re.sub(r'^RETIROS?\s*','',col1,flags=re.IGNORECASE)
            e = re.sub(r'\s*[-]\s*MW.*$','',e,flags=re.IGNORECASE).strip()
            if e and e not in nrm[nodo_actual]: nrm[nodo_actual].append(e)
            if e:
                if col_hora and col_hora < len(row):
                    v = row.iloc[col_hora]; mw = float(v) if isinstance(v,(int,float)) else 0.0
                else:
                    vals = [row.iloc[c] for c in range(2,min(98,len(row))) if isinstance(row.iloc[c],(int,float))]
                    mw = max(vals) if vals else 0.0
                nmm[nodo_actual][e] = round(mw, 4)
    return nrm, nmm, col_hora


# =============================================================================
# GRAFO TOPOLÓGICO
# Agrega campo trf_por_barra: {barra_LV: loc_name_trf} para el método trf_propietario
# =============================================================================
def construir_grafo(datos_path):
    df_bar = pd.read_excel(datos_path, sheet_name="Barras")
    df_tr2 = pd.read_excel(datos_path, sheet_name="Transformadores_2dev")
    df_tr3 = pd.read_excel(datos_path, sheet_name="Transformadores_3dev")
    df_lne = pd.read_excel(datos_path, sheet_name="Lineas")

    bus_kv = {}; bus_zona = {}; pb = {}
    for _, r in df_bar.iterrows():
        n = str(r["Nombre"]).strip()
        if n in ("LEYENDA","nan",""): continue
        try:   kv = float(r.get("Tension nom. (kV)") or 0)
        except: kv = 0.0
        bus_kv[n] = kv
        zona = str(r.get("Zona") or "").strip()
        if zona and zona != "nan": bus_zona[n] = zona
        m = RE_PREFIX_BUS.match(n)
        if m:
            p = m.group(1).upper(); pb.setdefault(p, []).append((n, kv))
    for p in pb: pb[p].sort(key=lambda x: x[1], reverse=True)

    gu = {}
    # trf_por_barra: barra LV (BT/MT) → loc_name del transformador que la conecta
    trf_por_barra: dict[str, str] = {}

    def add(s, d):
        gu.setdefault(s, [])
        if d not in gu[s]: gu[s].append(d)

    def trf_fallback(nt):
        m = RE_TRF_NAME.match(str(nt))
        if not m: return None, None
        px = m.group(1).upper(); kv_nom = int(m.group(2))
        bp = pb.get(px, [])
        if not bp: return None, None
        cands = [(n, kv) for n, kv in bp if kv >= kv_nom * 0.7]
        if kv_nom < 50:
            ba = [(n, kv) for n, kv in bp if kv >= 100]
            if ba: cands = ba
        if not cands: return None, None
        hv_n, hv_k = (max(cands, key=lambda x: x[1]) if kv_nom < 50
                      else min(cands, key=lambda x: abs(x[1] - kv_nom)))
        lv_buses = []
        for lv_n, lv_k in bp:
            if lv_k < hv_k - 1: add(lv_n, hv_n); lv_buses.append(lv_n)
        return hv_n, lv_buses

    for _, r in df_tr2.iterrows():
        loc = str(r.get("Nombre","") or "").strip().lower()
        hv  = str(r.get("Barra HV","") or "").strip()
        lv  = str(r.get("Barra LV","") or "").strip()
        if hv and lv and hv not in ("-","nan") and lv not in ("-","nan"):
            add(lv, hv)
            if loc: trf_por_barra[lv] = loc
        else:
            _, lv_buses = trf_fallback(r.get("Nombre",""))
            if lv_buses and loc:
                for lb in lv_buses: trf_por_barra[lb] = loc

    for _, r in df_tr3.iterrows():
        loc = str(r.get("Nombre","") or "").strip().lower()
        hv  = str(r.get("Barra HV","") or "").strip()
        mv  = str(r.get("Barra MV","") or "").strip()
        lv  = str(r.get("Barra LV","") or "").strip()
        buses_ok = [x for x in [hv, mv, lv] if x and x not in ("-","nan")]
        if len(buses_ok) >= 2:
            for b, a in [(mv, hv), (lv, hv), (lv, mv)]:
                if b not in ("-","nan") and a not in ("-","nan") and bus_kv.get(a,0) >= bus_kv.get(b,0):
                    add(b, a)
                    if loc and b not in ("-","nan"): trf_por_barra[b] = loc
        else:
            _, lv_buses = trf_fallback(r.get("Nombre",""))
            if lv_buses and loc:
                for lb in lv_buses: trf_por_barra[lb] = loc

    for _, r in df_lne.iterrows():
        n1 = str(r.get("Nodo From","") or "").strip()
        n2 = str(r.get("Nodo To","")   or "").strip()
        if n1 and n2 and n1 != "nan" and n2 != "nan": add(n1,n2); add(n2,n1)
    for s, d in CONEXIONES_MANUALES.items():
        add(s, d)
        if d not in bus_kv: bus_kv[d] = UMBRAL_STI_KV

    return gu, bus_kv, bus_zona, trf_por_barra


# =============================================================================
# CLASIFICACIÓN — nueva jerarquía con trf_propietario como prioridad 1
# =============================================================================
def _buscar_trf_aguas_arriba(bus, gu, bus_kv, trf_por_barra, max_saltos=BFS_TRF_MAX_SALTOS):
    """
    BFS limitado a barras BT/MT (< UMBRAL_STI_KV).
    Retorna (loc_name_trf, dist_propietario) del primer transformador encontrado
    cuya barra LV sea alcanzable desde 'bus' en ≤ max_saltos.
    """
    if bus in trf_por_barra:
        loc_trf = trf_por_barra[bus]
        dist    = TRF_PROPIETARIO.get(loc_trf)
        if dist: return loc_trf, dist

    visited = {bus}; queue = deque([(bus, 0)])
    while queue:
        cur, saltos = queue.popleft()
        if saltos >= max_saltos: continue
        for v in gu.get(cur, []):
            if v in visited: continue
            visited.add(v)
            if v in trf_por_barra:
                loc_trf = trf_por_barra[v]
                dist    = TRF_PROPIETARIO.get(loc_trf)
                if dist: return loc_trf, dist
            # Solo seguir por barras BT/MT
            if bus_kv.get(v, 0) < UMBRAL_STI_KV:
                queue.append((v, saltos + 1))
    return "", ""


def clasificar_carga(bus, gu, bus_kv, bus_zona, trf_por_barra):
    """
    Retorna (clase, metodo, barra_ref, loc_trf_inmediato, prop_trf).
    """
    loc_trf = ""; prop_trf = ""

    if not bus or bus == "nan":
        return "Sin asignar", "bus_nulo", "", "", ""

    # P1: Prefijo de barra
    m = RE_PREFIX_BUS.match(bus)
    if m and m.group(1).upper() in BUS_PREFIJO_CLASE:
        return BUS_PREFIJO_CLASE[m.group(1).upper()], "prefijo_bus", bus, "", ""

    # P2: Transformador inmediato aguas arriba (trf_propietario)
    loc_trf, prop_trf = _buscar_trf_aguas_arriba(bus, gu, bus_kv, trf_por_barra)
    if prop_trf:
        return prop_trf, "trf_propietario", bus, loc_trf, prop_trf

    # P3: Barra directa en NODO_SUBCLASE
    if bus in NODO_SUBCLASE:
        return NODO_SUBCLASE[bus], "subclase_directo", bus, "", ""

    # P4: Zona + prefijo de la barra directa
    if bus in bus_zona:
        zona = bus_zona[bus]
        dist = ZONA_DISTRIBUIDOR.get(zona)
        if dist:
            m2 = RE_PREFIX_BUS.match(bus)
            if m2:
                px = m2.group(1).upper()
                for key, val in NODO_SUBCLASE.items():
                    if key.startswith(px):
                        return val, "zona+prefijo", bus, "", ""
            return dist, "zona_directa", bus, "", ""

    # P5: BFS — buscar NODO_SUBCLASE o Zona en la red
    visited = {bus}; queue = deque([bus])
    fallback_zona = None; fallback_bus = None

    while queue:
        cur = queue.popleft()
        for v in gu.get(cur, []):
            if v in visited: continue
            visited.add(v)
            if v in NODO_SUBCLASE:
                return NODO_SUBCLASE[v], "bfs_subclase", v, "", ""
            if v in bus_zona and fallback_zona is None:
                dist = ZONA_DISTRIBUIDOR.get(bus_zona[v])
                if dist: fallback_zona = dist; fallback_bus = v
            if bus_kv.get(v, 0) >= UMBRAL_STI_KV and fallback_bus is None:
                fallback_bus = v
            queue.append(v)

    if fallback_zona:
        return fallback_zona, "bfs_zona", fallback_bus or "", "", ""

    return "Sin asignar", "sin_ruta", "", "", ""


def _confianza(metodo):
    if metodo in ("trf_lv_directo", "trf_propietario", "prefijo_bus", "subclase_directo"):
        return "Alta"
    if metodo in ("zona+prefijo", "bfs_subclase"):
        return "Media"
    return "Baja"


# =============================================================================
# MAPEO DIRECTO TRANSFORMADOR → BARRA LV → CARGAS
# Lógica inversa: usa los loc_names del PDF para encontrar la Barra LV
# y luego busca cargas directamente conectadas a esa barra.
# =============================================================================
def construir_barra_lv_a_dist(df_xfo_all, trf_prop):
    """
    Construye dict {barra_LV: (loc_name_trf, distribuidor)} usando loc_names_xfo.xlsx.
    Solo incluye transformadores cuyos loc_names están en trf_prop (fuente PDF).
    """
    resultado = {}
    for _, row in df_xfo_all.iterrows():
        loc      = str(row.get("loc_name", "") or "").strip().lower()
        barra_lv = str(row.get("Barra LV", "") or "").strip()
        if not loc or not barra_lv or barra_lv in ("-", "nan"):
            continue
        dist = trf_prop.get(loc)
        if dist:
            resultado[barra_lv] = (loc, dist)
    return resultado


def construir_mapeo_trf_cargas(df_xfo_all, df_c, trf_prop):
    """
    Para cada transformador en trf_prop (fuente PDF):
      1. Obtiene su Barra LV desde loc_names_xfo
      2. Busca cargas (lod_*) directamente conectadas a esa barra
      3. Retorna DataFrame: loc_name_trf | distribuidor | Barra LV | loc_name carga | MW | En servicio
    Incluye una fila "(sin carga directa)" para barras LV sin cargas.
    """
    filas = []
    for _, row in df_xfo_all.iterrows():
        loc      = str(row.get("loc_name", "") or "").strip().lower()
        barra_lv = str(row.get("Barra LV", "") or "").strip()
        kv_lv    = row.get("Tension LV nom. (kV)", None)
        if not loc or not barra_lv or barra_lv in ("-", "nan"):
            continue
        dist = trf_prop.get(loc)
        if not dist:
            continue  # trf no está en el instructivo PDF

        cargas_en_bus = df_c[df_c["Barra conectada"].astype(str).str.strip() == barra_lv]
        if not cargas_en_bus.empty:
            for _, carga in cargas_en_bus.iterrows():
                filas.append({
                    "loc_name trf":     loc,
                    "Distribuidor PDF": dist,
                    "Barra LV":         barra_lv,
                    "kV LV":            kv_lv,
                    "loc_name carga":   carga.get("Nombre", ""),
                    "Barra carga PF":   carga.get("Barra conectada", ""),
                    "P nom. (MW)":      carga.get("P nom. (MW)"),
                    "En servicio":      carga.get("En servicio", ""),
                    "Tiene carga":      "Sí",
                })
        else:
            filas.append({
                "loc_name trf":     loc,
                "Distribuidor PDF": dist,
                "Barra LV":         barra_lv,
                "kV LV":            kv_lv,
                "loc_name carga":   "(sin carga directa)",
                "Barra carga PF":   "",
                "P nom. (MW)":      None,
                "En servicio":      "",
                "Tiene carga":      "No",
            })
    return pd.DataFrame(filas)


# =============================================================================
# CONSTRUIR MAPEO COMPLETO
# =============================================================================
def construir_mapeo(df_c, gu, bus_kv, bus_zona, trf_por_barra, barra_lv_a_dist=None):
    res = []
    for _, row in df_c.iterrows():
        bp = str(row.get("Barra conectada", "")).strip()
        zp = str(row.get("Zona", "")).strip()

        # P0: Barra LV directa desde loc_names_xfo (fuente PDF) — máxima confianza
        if barra_lv_a_dist and bp in barra_lv_a_dist:
            loc_trf_p0, dist_p0 = barra_lv_a_dist[bp]
            cl, met, bref, loc_trf, prop_trf = dist_p0, "trf_lv_directo", bp, loc_trf_p0, dist_p0
        else:
            cl, met, bref, loc_trf, prop_trf = clasificar_carga(bp, gu, bus_kv, bus_zona, trf_por_barra)

        if cl == "Sin asignar" and zp in ZONA_DISTRIBUIDOR:
            cl  = ZONA_DISTRIBUIDOR[zp]
            met = met + "→zona_pf"

        res.append({
            "Nombre carga (PF)":       row.get("Nombre", ""),
            "Barra PF":                bp,
            "Zona PF":                 zp,
            "Barra ref. mapeo":        bref,
            "loc_name trf inmediato":  loc_trf,
            "Propietario trf":         prop_trf,
            "Metodo":                  met,
            "Confianza mapeo":         _confianza(met),
            "Distribuidor / C.N.R.":   cl,
            "P nom. (MW)":             row.get("P nom. (MW)"),
            "P result. (MW)":          row.get("P result. (MW)"),
            "Q result. (Mvar)":        row.get("Q result. (Mvar)"),
            "Cos phi":                 row.get("Cos phi"),
            "En servicio":             row.get("En servicio", ""),
        })
    return pd.DataFrame(res)


# =============================================================================
# RESUMEN POR DISTRIBUIDOR (con columnas de confianza)
# =============================================================================
def construir_resumen(df_mapeo, mw_deener):
    filas = []
    for dist in ORDEN_DEENER:
        grp = df_mapeo[df_mapeo["Distribuidor / C.N.R."] == dist]
        es  = grp[grp["En servicio"].astype(str).str.lower().isin(["si","true","1"])] if not grp.empty else grp
        n_alta  = (grp["Confianza mapeo"] == "Alta").sum()
        n_media = (grp["Confianza mapeo"] == "Media").sum()
        n_baja  = (grp["Confianza mapeo"] == "Baja").sum()
        n_tot   = len(grp)
        pct_alta = round((n_alta + n_media) / n_tot * 100, 1) if n_tot > 0 else None
        pnom_tot  = round(grp["P nom. (MW)"].sum(), 3) if not grp.empty else 0.0
        pnom_serv = round(es["P nom. (MW)"].sum(), 3)  if not grp.empty else 0.0
        mw_dee    = round(mw_deener.get(dist, 0.0), 3)
        factor    = round(mw_dee / pnom_tot, 4) if pnom_tot > 0 else None
        cob_dee   = round(pnom_tot / mw_dee * 100, 1) if mw_dee > 0 else None
        mw_anuar  = ANUARIO_GRUPO.get(dist, 0)
        cob_anuar = round(pnom_tot / mw_anuar * 100, 1) if mw_anuar > 0 else None
        estado    = ("✓ Bueno" if cob_dee and cob_dee >= 95 else
                     "~ Aceptable" if cob_dee and cob_dee >= 80 else
                     "✗ Bajo" if cob_dee else "⚠ Sin deener")
        filas.append({
            "Distribuidor / C.N.R.":    dist,
            "N° cargas":                n_tot,
            "N° en servicio":           len(es),
            "P nom. total (MW)":        pnom_tot,
            "P nom. en servicio (MW)":  pnom_serv,
            f"MW deener {HORA_EVENTO_LABEL}": mw_dee,
            "Factor calibración":       factor,
            "Cobertura vs deener (%)":  cob_dee,
            "MW Anuario 2024 (grupo)":  mw_anuar if mw_anuar > 0 else None,
            "Cobertura vs Anuario (%)": cob_anuar,
            "N Alta confianza":         int(n_alta),
            "N Media confianza":        int(n_media),
            "N Baja confianza":         int(n_baja),
            "% Alta+Media":             pct_alta,
            "Estado":                   estado,
        })
    tot_pnom = round(df_mapeo["P nom. (MW)"].sum(), 3)
    tot_dee  = round(sum(mw_deener.values()), 3)
    n_alta_t  = (df_mapeo["Confianza mapeo"] == "Alta").sum()
    n_media_t = (df_mapeo["Confianza mapeo"] == "Media").sum()
    n_baja_t  = (df_mapeo["Confianza mapeo"] == "Baja").sum()
    filas.append({
        "Distribuidor / C.N.R.":    "TOTAL",
        "N° cargas":                len(df_mapeo),
        "N° en servicio":           len(df_mapeo[df_mapeo["En servicio"].astype(str).str.lower().isin(["si","true","1"])]),
        "P nom. total (MW)":        tot_pnom,
        "P nom. en servicio (MW)":  round(df_mapeo[df_mapeo["En servicio"].astype(str).str.lower().isin(["si","true","1"])]["P nom. (MW)"].sum(), 3),
        f"MW deener {HORA_EVENTO_LABEL}": tot_dee,
        "Factor calibración":       round(tot_dee / tot_pnom, 4) if tot_pnom > 0 else None,
        "Cobertura vs deener (%)":  round(tot_pnom / tot_dee * 100, 1) if tot_dee > 0 else None,
        "MW Anuario 2024 (grupo)":  ANUARIO_TOTAL,
        "Cobertura vs Anuario (%)": round(tot_pnom / ANUARIO_TOTAL * 100, 1),
        "N Alta confianza":         int(n_alta_t),
        "N Media confianza":        int(n_media_t),
        "N Baja confianza":         int(n_baja_t),
        "% Alta+Media":             round((n_alta_t + n_media_t) / len(df_mapeo) * 100, 1) if len(df_mapeo) > 0 else None,
        "Estado":                   "",
    })
    return pd.DataFrame(filas)


def construir_locnames_detalle(df_mapeo):
    detalle = []
    for dist in ORDEN_DEENER:
        grp = df_mapeo[df_mapeo["Distribuidor / C.N.R."] == dist]
        if grp.empty: continue
        pt = grp["P nom. (MW)"].sum()
        for _, row in grp.sort_values("P nom. (MW)", ascending=False).iterrows():
            pnom = row.get("P nom. (MW)") or 0
            detalle.append({
                "Distribuidor / C.N.R.":  dist,
                "loc_name (PF)":          row.get("Nombre carga (PF)",""),
                "Barra PF":               row.get("Barra PF",""),
                "Barra ref. mapeo":        row.get("Barra ref. mapeo",""),
                "loc_name trf":            row.get("loc_name trf inmediato",""),
                "Zona PF":                 row.get("Zona PF",""),
                "Metodo":                  row.get("Metodo",""),
                "Confianza mapeo":         row.get("Confianza mapeo",""),
                "P nom. (MW)":             pnom,
                "P result. (MW)":          row.get("P result. (MW)"),
                "En servicio":             row.get("En servicio",""),
                "Peso / dist. total":      round(pnom / pt, 6) if pt > 0 else 0,
                "Nota curva PF":           "Escalar curva_dist × Peso/total",
            })
    return pd.DataFrame(detalle)


# =============================================================================
# FORMATO EXCEL
# =============================================================================
THIN = Side(border_style="thin", color="BFBFBF")
BRD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left",   vertical="center")
H_FILL = PatternFill("solid", start_color="1F3864")
H_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
N_FONT = Font(name="Arial", size=10)
B_FONT = Font(name="Arial", bold=True, size=10)
def _fill(h): return PatternFill("solid", start_color=h, end_color=h)

def _auto_col(ws, max_w=50):
    for col in ws.columns:
        a = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(a + 3, 12), max_w)

def _fmt_mapeo(ws, col_clase, col_serv, col_conf):
    ws.row_dimensions[1].height = 30; mc = ws.max_column
    for c in range(1, mc + 1):
        cell = ws.cell(1, c); cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = CTR; cell.border = BRD
    for r in range(2, ws.max_row + 1):
        clase = ws.cell(r, col_clase).value
        serv  = str(ws.cell(r, col_serv).value).strip()
        color = ("FFC7CE" if serv.lower() not in ("si","true","1")
                 else COLOR_CLASE.get(str(clase or ""), COLOR_DEFAULT))
        f = _fill(color)
        for c in range(1, mc + 1):
            cell = ws.cell(r, c); cell.font = N_FONT; cell.border = BRD; cell.fill = f
            cell.alignment = CTR if isinstance(cell.value, (int, float)) else LEFT
        # Columna confianza: color semáforo
        cconf = ws.cell(r, col_conf)
        cconf.fill = _fill(COLOR_CONFIANZA.get(str(cconf.value or ""), COLOR_DEFAULT))
    _auto_col(ws); ws.freeze_panes = "E2"

def _fmt_resumen(ws):
    ESTADO_COLOR = {"✓ Bueno":"C6EFCE","~ Aceptable":"FFEB9C","✗ Bajo":"FFC7CE","⚠ Sin deener":"F2F2F2"}
    ws.row_dimensions[1].height = 30; mc = ws.max_column
    for c in range(1, mc + 1):
        cell = ws.cell(1, c); cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = CTR; cell.border = BRD
    for r in range(2, ws.max_row + 1):
        dist = str(ws.cell(r, 1).value or "")
        es_tot = dist.upper() == "TOTAL"
        es_sin = dist == "Sin asignar"
        estado = str(ws.cell(r, mc).value or "")
        if es_tot:   color = "2F5496"; ff = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        elif es_sin: color = "FFC7CE"; ff = N_FONT
        else:        color = COLOR_CLASE.get(dist, COLOR_DEFAULT); ff = N_FONT
        f = _fill(color)
        for c in range(1, mc + 1):
            cell = ws.cell(r, c); cell.border = BRD; cell.fill = f; cell.font = ff
            cell.alignment = CTR if isinstance(cell.value, (int, float)) else LEFT
        ec = ws.cell(r, mc)
        if not es_tot and not es_sin:
            ec.fill = _fill(ESTADO_COLOR.get(estado, COLOR_DEFAULT))
    _auto_col(ws); ws.freeze_panes = "B2"

def _fmt_locnames(ws):
    ws.row_dimensions[1].height = 30; mc = ws.max_column
    for c in range(1, mc + 1):
        cell = ws.cell(1, c); cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = CTR; cell.border = BRD
    for r in range(2, ws.max_row + 1):
        dist = str(ws.cell(r, 1).value or ""); nota = str(ws.cell(r, mc).value or "")
        f  = _fill(COLOR_CLASE.get(dist, COLOR_DEFAULT))
        fn = _fill("FFC7CE" if "FUERA" in nota else "C6EFCE")
        for c in range(1, mc + 1):
            cell = ws.cell(r, c); cell.font = N_FONT; cell.border = BRD
            cell.fill = fn if c == mc else f
            cell.alignment = CTR if isinstance(cell.value, (int, float)) else LEFT
    _auto_col(ws); ws.freeze_panes = "B2"

def _fmt_trf(ws, col_dist, col_rev):
    ws.row_dimensions[1].height = 30; mc = ws.max_column
    for c in range(1, mc + 1):
        cell = ws.cell(1, c); cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = CTR; cell.border = BRD
    for r in range(2, ws.max_row + 1):
        dist = str(ws.cell(r, col_dist).value or "")
        rev  = str(ws.cell(r, col_rev).value  or "")
        color = "FFEB9C" if rev == "Si" else COLOR_CLASE.get(dist, COLOR_DEFAULT)
        f = _fill(color)
        for c in range(1, mc + 1):
            cell = ws.cell(r, c); cell.font = N_FONT; cell.border = BRD; cell.fill = f
            cell.alignment = CTR if isinstance(cell.value, (int, float)) else LEFT
    _auto_col(ws); ws.freeze_panes = "B2"

def _fmt_deener(ws):
    ws.row_dimensions[1].height = 30; mc = ws.max_column
    for c in range(1, mc + 1):
        cell = ws.cell(1, c); cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = CTR; cell.border = BRD
    for r in range(2, ws.max_row + 1):
        dist  = str(ws.cell(r, 1).value or "")
        color = "FFC7CE" if dist == "Sin asignar" else COLOR_CLASE.get(dist, COLOR_DEFAULT)
        f = _fill(color)
        for c in range(1, mc + 1):
            cell = ws.cell(r, c); cell.font = N_FONT; cell.border = BRD; cell.fill = f
            cell.alignment = CTR if isinstance(cell.value, (int, float)) else LEFT
    _auto_col(ws)


# =============================================================================
# MAIN
# =============================================================================
print("=" * 60)
print("MapeoRetirosSTI v6 — trf_propietario + instructivo CNDC512_25")
print("=" * 60)

print("\n[1/6] Cargando cargas PF...")
df_c = pd.read_excel(DATOS_PATH, sheet_name="Cargas")
df_c = df_c[df_c["Nombre"].astype(str).str.startswith("lod_")].reset_index(drop=True)
print(f"      {len(df_c)} cargas")

print("[2/6] Parseando deener...")
if not DEENER_PATH or not os.path.isfile(DEENER_PATH):
    print("  [AVISO] DEENER_PATH no configurado o no existe.")
    mw_deener = {}; fmt_dee = "sin_deener"; nodos_raw = []; sin_mapa_dee = []
else:
    mw_deener, fmt_dee, nodos_raw, sin_mapa_dee = parsear_deener(DEENER_PATH, HORA_EVENTO_LABEL)
    print(f"      Formato: {fmt_dee}  |  {len(nodos_raw)} nodos  |  {len(sin_mapa_dee)} sin mapeo")
    if sin_mapa_dee:
        print(f"      Sin mapa: {sin_mapa_dee[:10]}{'...' if len(sin_mapa_dee)>10 else ''}")
    print(f"      MW total deener: {sum(mw_deener.values()):.1f} MW")

print("[3/6] Grafo topológico + transformadores...")
gu, bus_kv, bus_zona, trf_por_barra = construir_grafo(DATOS_PATH)
print(f"      {len(bus_kv)} buses  |  {len(bus_zona)} con Zona  |  "
      f"{sum(len(v) for v in gu.values())} aristas  |  {len(trf_por_barra)} barras LV mapeadas a trf")

print("[4/6] Construyendo TRF_PROPIETARIO ampliado...")
trf_prop_final = dict(TRF_PROPIETARIO)
filas_trf_hoja = []
if LOC_NAMES_XFO and os.path.isfile(LOC_NAMES_XFO):
    df_xfo2 = pd.read_excel(LOC_NAMES_XFO, sheet_name="Transformadores_2dev")
    df_xfo3 = pd.read_excel(LOC_NAMES_XFO, sheet_name="Transformadores_3dev")
    df_xfo_all = pd.concat([df_xfo2, df_xfo3], ignore_index=True)
    df_xfo_all = df_xfo_all[df_xfo_all["En servicio"].astype(str).str.strip().str.lower().isin(["si","sí","yes","true","1"])]
    trf_prop_final, filas_trf_hoja = _construir_trf_propietario_auto(df_xfo_all, TRF_PROPIETARIO, NODO_SUBCLASE)
    n_manual = sum(1 for f in filas_trf_hoja if f["Metodo asignacion"] == "instructivo_CNDC512")
    n_auto   = sum(1 for f in filas_trf_hoja if f["Metodo asignacion"] != "instructivo_CNDC512" and f["Distribuidor asignado"] != "Sin asignar")
    n_rev    = sum(1 for f in filas_trf_hoja if f["Requiere revision"] == "Si")
    print(f"      {len(trf_prop_final)} transformadores  |  "
          f"{n_manual} desde instructivo  |  {n_auto} auto  |  {n_rev} requieren revisión")
else:
    print("  [AVISO] LOC_NAMES_XFO no encontrado — usando solo diccionario manual.")
    for loc, dist in TRF_PROPIETARIO.items():
        filas_trf_hoja.append({
            "loc_name trf": loc, "Barra HV": "", "Barra LV": "",
            "kV HV": "", "kV LV": "", "Potencia MVA": "", "En servicio": "Si",
            "Distribuidor asignado": dist, "Metodo asignacion": "instructivo_CNDC512",
            "Requiere revision": "No",
        })

# Actualizar TRF_PROPIETARIO con el diccionario ampliado
TRF_PROPIETARIO.update(trf_prop_final)

# Construir mapeo directo Barra LV → (loc_name_trf, distribuidor) desde loc_names_xfo
barra_lv_a_dist: dict = {}
df_trf_cargas = pd.DataFrame()
if LOC_NAMES_XFO and os.path.isfile(LOC_NAMES_XFO):
    barra_lv_a_dist = construir_barra_lv_a_dist(df_xfo_all, trf_prop_final)
    df_trf_cargas   = construir_mapeo_trf_cargas(df_xfo_all, df_c, trf_prop_final)
    n_con_carga = (df_trf_cargas["Tiene carga"] == "Sí").sum() if not df_trf_cargas.empty else 0
    n_sin_carga = (df_trf_cargas["Tiene carga"] == "No").sum() if not df_trf_cargas.empty else 0
    print(f"      Mapeo directo trf→LV: {len(barra_lv_a_dist)} barras LV | "
          f"{n_con_carga} con carga directa | {n_sin_carga} sin carga directa")

print("[5/6] Mapeando cargas...")
df_map = construir_mapeo(df_c, gu, bus_kv, bus_zona, trf_por_barra, barra_lv_a_dist)
n_sin = (df_map["Distribuidor / C.N.R."] == "Sin asignar").sum()
print(f"\n      {'Metodo':<40} {'N':>5}  {'MW':>8}")
print(f"      {'-'*56}")
for met, grp in df_map.groupby("Metodo"):
    print(f"      {met:<40} {len(grp):>5}  {grp['P nom. (MW)'].sum():>8.1f}")
print(f"      {'-'*56}")
for conf in ["Alta", "Media", "Baja"]:
    g = df_map[df_map["Confianza mapeo"] == conf]
    print(f"      Confianza {conf:<6}: {len(g):>4} cargas  ({len(g)/len(df_map)*100:.1f}%)")
print(f"      → SIN ASIGNAR: {n_sin} cargas")

print("\n[6/6] Exportando...")
df_res = construir_resumen(df_map, mw_deener)
df_det = construir_locnames_detalle(df_map)
df_dee_cob = pd.DataFrame(nodos_raw) if nodos_raw else pd.DataFrame(columns=["nombre","subdist","mw"])
df_trf_hoja = pd.DataFrame(filas_trf_hoja)

with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as w:
    df_map.to_excel(     w, sheet_name="Mapeo_Cargas",    index=False)
    df_res.to_excel(     w, sheet_name="Resumen_Dist",    index=False)
    df_det.to_excel(     w, sheet_name="Curvas_LocNames", index=False)
    df_trf_hoja.to_excel(w, sheet_name="TRF_Propietario",index=False)
    df_dee_cob.to_excel( w, sheet_name="Deener_Nodos",   index=False)
    if not df_trf_cargas.empty:
        df_trf_cargas.to_excel(w, sheet_name="TRF_Cargas", index=False)

wb = load_workbook(OUTPUT_PATH)

cols = list(df_map.columns)
_fmt_mapeo(wb["Mapeo_Cargas"],
           cols.index("Distribuidor / C.N.R.") + 1,
           cols.index("En servicio") + 1,
           cols.index("Confianza mapeo") + 1)
_fmt_resumen(wb["Resumen_Dist"])
_fmt_locnames(wb["Curvas_LocNames"])
cols_trf = list(df_trf_hoja.columns)
_fmt_trf(wb["TRF_Propietario"],
         cols_trf.index("Distribuidor asignado") + 1,
         cols_trf.index("Requiere revision") + 1)
_fmt_deener(wb["Deener_Nodos"])

if "TRF_Cargas" in wb.sheetnames:
    ws_tc = wb["TRF_Cargas"]
    ws_tc.row_dimensions[1].height = 30
    for c in range(1, ws_tc.max_column + 1):
        cell = ws_tc.cell(1, c); cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = CTR; cell.border = BRD
    cols_tc = list(df_trf_cargas.columns)
    col_dist_tc = cols_tc.index("Distribuidor PDF") + 1
    col_carga_tc = cols_tc.index("Tiene carga") + 1
    for r in range(2, ws_tc.max_row + 1):
        dist_v  = str(ws_tc.cell(r, col_dist_tc).value or "")
        carga_v = str(ws_tc.cell(r, col_carga_tc).value or "")
        color = ("FFC7CE" if carga_v == "No"
                 else COLOR_CLASE.get(dist_v, COLOR_DEFAULT))
        f = _fill(color)
        for c in range(1, ws_tc.max_column + 1):
            cell = ws_tc.cell(r, c); cell.font = N_FONT; cell.border = BRD; cell.fill = f
            cell.alignment = CTR if isinstance(cell.value, (int, float)) else LEFT
    _auto_col(ws_tc); ws_tc.freeze_panes = "D2"

# Nota al pie en Resumen
ws_res = wb["Resumen_Dist"]
fn = ws_res.max_row + 2
nota = ws_res.cell(fn, 1,
    f"MW deener {HORA_EVENTO_LABEL}: demanda por nodo del CNDC ({os.path.basename(DEENER_PATH)}). "
    f"Formato detectado: {fmt_dee}. "
    "Propietarios de transformadores: CNDC512_25_InstructivosRestitucionSIN.pdf (oct-2025). "
    "Factor calibración = MW_deener / P_nom_total. "
    "Confianza Alta: trf_lv_directo/trf_propietario/prefijo_bus/subclase_directo. "
    "Media: zona+prefijo/bfs_subclase. Baja: zona_directa/bfs_zona. "
    "trf_lv_directo: carga directamente en Barra LV del trf (fuente PDF loc_names_xfo).")
nota.font = Font(name="Arial", italic=True, size=9, color="595959")
nota.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws_res.row_dimensions[fn].height = 55
ws_res.merge_cells(start_row=fn, start_column=1, end_row=fn, end_column=ws_res.max_column)

wb.save(OUTPUT_PATH)

print(f"\n  ✓ {OUTPUT_PATH}")
print(f"  Hojas: {', '.join(wb.sheetnames)}")
print(f"\n  Resumen final:")
print(f"    Cargas totales     : {len(df_map)}")
print(f"    Sin asignar        : {n_sin}")
_tot = df_res[df_res["Distribuidor / C.N.R."] == "TOTAL"].iloc[0] if "TOTAL" in df_res["Distribuidor / C.N.R."].values else {}
if _tot is not None and len(_tot):
    print(f"    % Alta+Media conf. : {_tot.get('% Alta+Media', '?')}%")
print(f"    MW deener total    : {sum(mw_deener.values()):.1f} MW")
print(f"    P_nom PF total     : {df_map['P nom. (MW)'].sum():.1f} MW")
if sin_mapa_dee:
    print(f"\n  Nodos deener sin mapeo ({len(sin_mapa_dee)}):")
    for n in sin_mapa_dee:
        print(f"    → {n}")
    print("  Agregar a DEENER_NODO_A_SUBDIST si corresponde.")
