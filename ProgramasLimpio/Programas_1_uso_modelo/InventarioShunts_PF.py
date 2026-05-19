# =============================================================================
# INVENTARIO DE SHUNTS Y COMPENSADORES — PowerFactory (SIN Boliviano)
#
# Ejecutar con PowerFactory ya abierto y un escenario activo con Load Flow
# previo ejecutado. No modifica ningun parametro del modelo PF.
#
# Recopila todos los elementos shunt del caso activo:
#   ElmShnt — condensadores, reactores, resistivos
#   ElmSvs  — compensadores estaticos de tension (SVC)
#   ElmVsc  — convertidores VSC (si existen)
#
# Clasifica cada elemento por tipo segun resultado del LF:
#   Condensador — Q_LF > +0.1 MVAr
#   Reactor     — Q_LF < -0.1 MVAr
#   Resistivo   — |P_LF| > 0.01 MW  y  |Q_LF| <= 0.1
#   Inactivo    — en caso contrario (o fuera de servicio)
#
# Salida: inventario_shunts_{escenario}.xlsx  en OUTPUT_DIR
#   Hoja Inventario  -> una fila por shunt con datos nominales y resultados LF
#   Hoja Resumen     -> totales por tipo + balance activo en slack
# =============================================================================

import os, sys, re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# RUTAS
# =============================================================================

PF_BASE     = r"C:\Program Files\DIgSILENT\PowerFactory 2025 SP2"
PF_PROYECTO = "PMP_NOV25_OCT29_31102025(1)"
CASO_BASE   = "CNDC"
OUTPUT_DIR  = r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\Designacion de loc_name"

# =============================================================================
# COLORES POR TIPO
# =============================================================================

COLOR_TIPO = {
    "Condensador": "C6EFCE",   # verde claro
    "Reactor":     "DDEBF7",   # azul claro
    "Resistivo":   "FFEB9C",   # amarillo/naranja
    "Inactivo":    "F2F2F2",   # gris claro
    "fuera":       "FFC7CE",   # rojo claro — fuera de servicio
}

# =============================================================================
# HELPERS  (mismo estilo que loc_namesGEN.py y CargaCondIniciales_PF.py)
# =============================================================================

def _float(val, default=0.0):
    try:
        v = float(val)
        return v if not pd.isna(v) else default
    except Exception:
        return default


def separador(titulo="", ancho=60):
    if titulo:
        print(f"\n{'='*ancho}")
        print(f"  {titulo}")
        print(f"{'='*ancho}")
    else:
        print(f"{'='*ancho}")


def _pf_attr(obj, attr, default=None):
    """Lee un atributo PF via COM de forma segura (GetAttribute primero)."""
    try:
        v = obj.GetAttribute(attr)
        if v is not None:
            return v
    except Exception:
        pass
    try:
        return getattr(obj, attr, default)
    except Exception:
        return default


def _agregar_pf_path(base):
    py_ver = f"3.{sys.version_info.minor}"
    py_dir = os.path.join(base, "Python")
    candidatos = [os.path.join(py_dir, py_ver)]
    if os.path.isdir(py_dir):
        for v in sorted(os.listdir(py_dir), reverse=True):
            c = os.path.join(py_dir, v)
            if c not in candidatos:
                candidatos.append(c)
    candidatos.append(base)
    for c in candidatos:
        if os.path.isfile(os.path.join(c, "powerfactory.pyd")):
            if c not in sys.path:
                sys.path.insert(0, c)
            return c
    raise FileNotFoundError(
        f"No se encontro powerfactory.pyd bajo:\n  {base}\n"
        "Verifique PF_BASE en el script.")


# =============================================================================
# FORMATO EXCEL  (mismo esquema que loc_namesGEN.py y loc_names_xfo.py)
# =============================================================================

_THIN = Side(border_style="thin", color="BFBFBF")
_BRD  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left",   vertical="center")
_HFIL = PatternFill("solid", start_color="1F3864")
_HFNT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_NFNT = Font(name="Arial", size=10)
_BFNT = Font(name="Arial", bold=True, size=10)


def _fill(h): return PatternFill("solid", start_color=h, end_color=h)


def _aplicar_hoja(ws, col_tipo=None, col_estado=None, freeze="A2"):
    ws.row_dimensions[1].height = 30
    mc = ws.max_column
    for c in range(1, mc + 1):
        cell = ws.cell(1, c)
        cell.fill = _HFIL; cell.font = _HFNT
        cell.alignment = _CTR; cell.border = _BRD
    for r in range(2, ws.max_row + 1):
        tipo   = str(ws.cell(r, col_tipo).value   or "") if col_tipo   else ""
        estado = str(ws.cell(r, col_estado).value or "") if col_estado else "en_servicio"
        if estado == "fuera_servicio":
            color = COLOR_TIPO["fuera"]
        elif tipo == "TOTAL EN SERVICIO":
            color = "2F5496"
        else:
            color = COLOR_TIPO.get(tipo, "FFFFFF")
        f = _fill(color)
        for c in range(1, mc + 1):
            cell = ws.cell(r, c)
            cell.border    = _BRD
            cell.fill      = f
            cell.alignment = _CTR if isinstance(cell.value, (int, float)) else _LEFT
            cell.font      = (_BFNT if tipo == "TOTAL EN SERVICIO"
                              else Font(name="Arial", size=10,
                                        color=("FFFFFF" if tipo == "TOTAL EN SERVICIO" else "000000")))
    for col in ws.columns:
        ancho = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ancho + 3, 12), 45)
    ws.freeze_panes = freeze


def _agregar_leyenda(ws, fila_inicio):
    fila = fila_inicio + 2
    t = ws.cell(fila, 1, "LEYENDA — Tipo de elemento shunt")
    t.fill = _HFIL; t.font = _HFNT; t.alignment = _CTR; t.border = _BRD
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
    leyenda = [
        ("Condensador", COLOR_TIPO["Condensador"], "Q_LF > +0.1 MVAr"),
        ("Reactor",     COLOR_TIPO["Reactor"],     "Q_LF < -0.1 MVAr"),
        ("Resistivo",   COLOR_TIPO["Resistivo"],   "|P_LF| > 0.01 MW y |Q_LF| <= 0.1"),
        ("Inactivo",    COLOR_TIPO["Inactivo"],     "Sin efecto activo/reactivo"),
        ("Fuera serv.", COLOR_TIPO["fuera"],        "outserv = 1"),
    ]
    for etiq, color, desc in leyenda:
        fila += 1
        c1 = ws.cell(fila, 1, "");   c1.fill = _fill(color); c1.border = _BRD
        c2 = ws.cell(fila, 2, etiq); c2.fill = _fill(color); c2.border = _BRD
        c2.font = _NFNT; c2.alignment = _LEFT
        c3 = ws.cell(fila, 3, desc); c3.fill = _fill(color); c3.border = _BRD
        c3.font = _NFNT; c3.alignment = _LEFT


# =============================================================================
# EXTRACCION DE SHUNTS
# =============================================================================

def extraer_shunts(app):
    """
    Recopila ElmShnt, ElmSvs y ElmVsc del caso activo.
    Retorna lista de dicts con datos nominales y resultados LF.
    """
    CLASES = ["*.ElmShnt", "*.ElmSvs", "*.ElmVsc"]
    filas  = []
    n_por_clase = {}

    for cls in CLASES:
        objs = app.GetCalcRelevantObjects(cls) or []
        n_por_clase[cls] = len(objs)
        for el in objs:
            try:
                loc_name = str(_pf_attr(el, "loc_name") or "").strip()

                # Barra conectada
                try:
                    terminal = el.bus1.cterm
                    barra    = str(terminal.loc_name).strip()
                    kv_nom   = _float(_pf_attr(terminal, "uknom"))
                except Exception:
                    barra  = ""
                    kv_nom = 0.0

                # Estado
                outserv = int(_float(_pf_attr(el, "outserv"), 0))
                estado  = "fuera_servicio" if outserv else "en_servicio"

                # Potencia reactiva nominal
                q_nom = 0.0
                for _a in ("qtotn", "bcap", "Qnom", "q_nom"):
                    v = _pf_attr(el, _a)
                    if v is not None:
                        q_nom = _float(v)
                        if q_nom != 0.0:
                            break

                # Potencia activa nominal
                p_nom = 0.0
                for _a in ("ptotn", "gcap", "Pnom", "p_nom"):
                    v = _pf_attr(el, _a)
                    if v is not None:
                        p_nom = _float(v)
                        if p_nom != 0.0:
                            break

                # Pasos conectados
                pasos = max(1, int(_float(_pf_attr(el, "nshph"), 1)))

                # Resultados Load Flow
                # Primero intenta leer P/Q del elemento; si da 0 lee del terminal
                p_lf = 0.0
                for _a in ("m:P:bus1", "m:P1:bus1", "m:Psum:bus1", "m:P"):
                    v = _pf_attr(el, _a)
                    if v is not None and _float(v) != 0.0:
                        p_lf = _float(v); break

                q_lf = 0.0
                for _a in ("m:Q:bus1", "m:Q1:bus1", "m:Qsum:bus1", "m:Q"):
                    v = _pf_attr(el, _a)
                    if v is not None and _float(v) != 0.0:
                        q_lf = _float(v); break

                # Tension: leer directamente del terminal (igual que DatsoGENBUSLNE)
                v_lf = 0.0
                try:
                    v_lf = _float(_pf_attr(terminal, "m:u"))
                except Exception:
                    pass
                if v_lf == 0.0:
                    for _a in ("m:u1:bus1", "m:u:bus1", "m:U1:bus1", "m:u1"):
                        v = _pf_attr(el, _a)
                        if v is not None:
                            v_lf = _float(v); break

                # Tipo
                if outserv:
                    tipo = "Inactivo"
                elif q_lf > 0.1:
                    tipo = "Condensador"
                elif q_lf < -0.1:
                    tipo = "Reactor"
                elif abs(p_lf) > 0.01 and abs(q_lf) <= 0.1:
                    tipo = "Resistivo"
                else:
                    tipo = "Inactivo"

                # Marca de tension fuera de rango [!]
                v_alerta = "[!]" if (v_lf > 1.05 or (0 < v_lf < 0.95)) else ""

                filas.append({
                    "loc_name":    loc_name,
                    "Barra":       barra,
                    "kV_nominal":  round(kv_nom, 3),
                    "Estado":      estado,
                    "Tipo":        tipo,
                    "Q_nom_MVAr":  round(q_nom, 3),
                    "P_nom_MW":    round(p_nom, 3),
                    "Pasos":       pasos,
                    "P_LF_MW":     round(p_lf, 4),
                    "Q_LF_MVAr":   round(q_lf, 4),
                    "V_LF_pu":     round(v_lf, 4),
                    "Alerta_V":    v_alerta,
                })
            except Exception as e:
                loc = str(getattr(el, "loc_name", "?"))
                print(f"  [AVISO] Error procesando {loc}: {e}")

    return filas, n_por_clase


def construir_resumen(df):
    """Tabla de totales por tipo + fila TOTAL EN SERVICIO."""
    df_serv = df[df["Estado"] == "en_servicio"]
    filas = []
    for tipo in ["Condensador", "Reactor", "Resistivo", "Inactivo"]:
        g = df_serv[df_serv["Tipo"] == tipo]
        filas.append({
            "Tipo":             tipo,
            "N elementos":      len(g),
            "Q_nom_MVAr":       round(g["Q_nom_MVAr"].sum(), 3),
            "P_nom_MW":         round(g["P_nom_MW"].sum(),   3),
            "P_LF_MW":          round(g["P_LF_MW"].sum(),   4),
            "Q_LF_MVAr":        round(g["Q_LF_MVAr"].sum(), 4),
        })
    filas.append({
        "Tipo":         "TOTAL EN SERVICIO",
        "N elementos":  len(df_serv),
        "Q_nom_MVAr":   round(df_serv["Q_nom_MVAr"].sum(), 3),
        "P_nom_MW":     round(df_serv["P_nom_MW"].sum(),   3),
        "P_LF_MW":      round(df_serv["P_LF_MW"].sum(),   4),
        "Q_LF_MVAr":    round(df_serv["Q_LF_MVAr"].sum(), 4),
    })
    return pd.DataFrame(filas)


# =============================================================================
# MAIN
# =============================================================================

def main():
    separador("INVENTARIO DE SHUNTS Y COMPENSADORES — SIN")

    # [1] Conexion a PowerFactory
    import os as _os
    _os.system("taskkill /f /im PowerFactory.exe >nul 2>&1")

    _agregar_pf_path(PF_BASE)
    import powerfactory as pf

    app = pf.GetApplication()
    if app is None:
        raise RuntimeError(
            "No se pudo obtener la aplicacion de PowerFactory.")

    app.Show()
    app.ActivateProject(PF_PROYECTO)

    proyecto = app.GetActiveProject()
    if proyecto is None:
        raise RuntimeError(
            f"No se pudo activar el proyecto '{PF_PROYECTO}'. "
            "Verifique que PF_PROYECTO sea el nombre exacto del proyecto.")

    # Desactivar eventos activos (IntEvt) para trabajar sobre el caso base
    try:
        for _evt in (app.GetFromStudyCase("*.ComInc") or []):
            try: _evt.Deactivate()
            except Exception: pass
    except Exception:
        pass
    try:
        study_case = app.GetActiveStudyCase()
        if study_case:
            for _evt in (study_case.GetContents("*.IntEvt") or []):
                try: _evt.Deactivate()
                except Exception: pass
    except Exception:
        pass

    # Activar escenario de operacion base CNDC
    _scen_activado = False
    try:
        _scenarios = app.GetProjectFolder("scen")
        if _scenarios:
            for _s in (_scenarios.GetContents("*.IntScenario") or []):
                if str(getattr(_s, "loc_name", "")).strip() == CASO_BASE:
                    _s.Activate()
                    _scen_activado = True
                    break
    except Exception:
        pass
    if not _scen_activado:
        print(f"  [AVISO] No se encontro escenario '{CASO_BASE}' — usando escenario activo.")

    # [1b] Preguntar si activar Tap Adjustment of Shunts
    separador("OPCIONES DE LOAD FLOW")
    print("  ¿Activar 'Tap Adjustment of Shunts' en el Load Flow?")
    print("    [1] SI — activar ajuste automatico de shunts")
    print("    [2] NO — ejecutar sin ajuste automatico (defecto)")
    while True:
        _resp = input("  Seleccion [1/2]: ").strip()
        if _resp in ("1", "2"):
            break
        print("  Entrada no valida. Ingrese 1 o 2.")
    tap_shunt_on = (_resp == "1")
    tap_sufijo   = "tapON" if tap_shunt_on else "tapOFF"
    print(f"  Tap Adjustment of Shunts : {'ACTIVADO' if tap_shunt_on else 'DESACTIVADO'}")

    # Ejecutar Load Flow sobre el caso base
    ldf = app.GetFromStudyCase("ComLdf")
    # iopt_asht: 1 = activar ajuste automatico de shunts, 0 = desactivado
    try:
        ldf.iopt_asht = 1 if tap_shunt_on else 0
    except Exception as _e:
        print(f"  [AVISO] No se pudo setear iopt_asht: {_e}")
    ldf.Execute()
    print("  [OK] Load Flow ejecutado.")

    _esc     = app.GetActiveScenario()
    esc_name = _esc.loc_name if _esc else "sin_escenario"
    _esc_safe = re.sub(r'[\\/:*?"<>|]', "_", esc_name).strip()

    print(f"  Proyecto  : {proyecto.loc_name}")
    print(f"  Escenario : {esc_name}")
    print(f"  Salida    : {OUTPUT_DIR}")

    # [2] Extraccion
    separador("EXTRAYENDO ELEMENTOS SHUNT")

    filas, n_por_clase = extraer_shunts(app)
    for cls, n in n_por_clase.items():
        print(f"  {cls:<15}: {n} encontrados")
    print(f"  Total     : {len(filas)}")

    if not filas:
        print("  [AVISO] No se encontraron elementos shunt en el caso activo.")
        return

    df = pd.DataFrame(filas)

    # Ordenar: Estado (en_servicio primero), Tipo, kV descendente
    _ord_est  = {"en_servicio": 0, "fuera_servicio": 1}
    _ord_tipo = {"Condensador": 0, "Reactor": 1, "Resistivo": 2, "Inactivo": 3}
    df["_se"] = df["Estado"].map(_ord_est).fillna(9)
    df["_st"] = df["Tipo"].map(_ord_tipo).fillna(9)
    df = (df.sort_values(["_se", "_st", "kV_nominal"], ascending=[True, True, False])
            .drop(columns=["_se", "_st"])
            .reset_index(drop=True))

    df_resumen = construir_resumen(df)

    # [3] Resumen en consola
    separador("RESUMEN")

    _en_serv = (df["Estado"] == "en_servicio").sum()
    _fuera   = len(df) - _en_serv
    print(f"  Total encontrados        : {len(df)}")
    print(f"  En servicio              : {_en_serv}")
    print(f"  Fuera de servicio        : {_fuera}")
    print(f"  {'-'*56}")

    df_serv = df[df["Estado"] == "en_servicio"]
    for tipo in ["Condensador", "Reactor", "Resistivo", "Inactivo"]:
        g = df_serv[df_serv["Tipo"] == tipo]
        n = len(g)
        if tipo == "Inactivo":
            print(f"  {tipo:<14} ({n:3d})        —")
            continue
        print(f"  {tipo:<14} ({n:3d})        "
              f"P_total= {g['P_LF_MW'].sum():+8.2f} MW   "
              f"Q_total= {g['Q_LF_MVAr'].sum():+9.2f} MVAr")

    print(f"  {'-'*56}")

    # Detalle por tipo
    for tipo in ["Condensador", "Reactor", "Resistivo"]:
        g = df_serv[df_serv["Tipo"] == tipo]
        if g.empty:
            continue
        print(f"\n  --- {tipo.upper()} ---")
        print(f"  {'loc_name':<22} {'Barra':<12} {'kV':>6}  "
              f"{'V_LF':>7}  {'Q_LF':>9}  {'P_LF':>8}")
        for _, r in g.iterrows():
            alerta = r["Alerta_V"]
            marca  = f"[!] {r['loc_name']:<18}" if alerta else f"    {r['loc_name']:<18}"
            print(f"  {marca} {r['Barra']:<12} "
                  f"{r['kV_nominal']:>6.1f}  {r['V_LF_pu']:>7.4f}  "
                  f"{r['Q_LF_MVAr']:>+9.3f}  {r['P_LF_MW']:>+8.3f}")

    _p_tot = df_serv["P_LF_MW"].sum()
    print(f"\n  {'-'*56}")
    print(f"  EFECTO EN BALANCE ACTIVO:")
    print(f"  P activa total shunts en servicio : {_p_tot:+.2f} MW")
    print(f"    -> Cubierta por la maquina slack")
    print(f"    -> Desbalance esperado en slack  : pgini_slack + {_p_tot:+.2f} MW")

    _alertas = df_serv[df_serv["Alerta_V"] == "[!]"]
    if not _alertas.empty:
        print(f"\n  [AVISO] Elementos con tension fuera de rango [0.95–1.05 pu]: "
              f"{len(_alertas)}")
        for _, r in _alertas.iterrows():
            print(f"    [!] {r['loc_name']:<22} V={r['V_LF_pu']:.4f} pu  "
                  f"Barra={r['Barra']}")

    separador()

    # [4] Exportar Excel
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, f"inventario_shunts_{_esc_safe}_{tap_sufijo}.xlsx")

    COL_INV = ["loc_name", "Barra", "kV_nominal", "Estado", "Tipo",
               "Q_nom_MVAr", "P_nom_MW", "Pasos",
               "P_LF_MW", "Q_LF_MVAr", "V_LF_pu", "Alerta_V"]

    print(f"[4] Exportando a Excel...")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df[COL_INV].to_excel(writer, sheet_name="Inventario", index=False)
        df_resumen.to_excel( writer, sheet_name="Resumen",    index=False)

    wb = load_workbook(output_path)

    cols_inv = list(df[COL_INV].columns)
    _aplicar_hoja(wb["Inventario"],
                  col_tipo   = cols_inv.index("Tipo")   + 1,
                  col_estado = cols_inv.index("Estado") + 1,
                  freeze     = "D2")
    # Marca [!] en rojo negrita
    _col_alerta = cols_inv.index("Alerta_V") + 1
    for r in range(2, wb["Inventario"].max_row + 1):
        if str(wb["Inventario"].cell(r, _col_alerta).value or "") == "[!]":
            for c in range(1, wb["Inventario"].max_column + 1):
                wb["Inventario"].cell(r, c).font = Font(
                    name="Arial", bold=True, color="C00000", size=10)

    cols_res = list(df_resumen.columns)
    _aplicar_hoja(wb["Resumen"],
                  col_tipo = cols_res.index("Tipo") + 1,
                  freeze   = "A2")
    _agregar_leyenda(wb["Inventario"], wb["Inventario"].max_row)

    wb.save(output_path)

    print(f"\n  Archivo creado en:")
    print(f"  {output_path}")
    print(f"\n  Hojas:")
    for sh in wb.sheetnames:
        print(f"    {sh:<16} -> {wb[sh].max_row - 1} filas")


if __name__ == "__main__":
    main()
    input("\nPresiona Enter para cerrar...")
