# Extraccion completa de datos de red PowerFactory
# Barras, Lineas, Generadores, Cargas, Transformadores

import sys
import os

PF_DIR = r"C:\Program Files\DIgSILENT\PowerFactory 2025 SP2"
PF_PY  = r"C:\Program Files\DIgSILENT\PowerFactory 2025 SP2\Python\3.12"

sys.path.append(PF_PY)
os.environ["PATH"] = PF_DIR + os.pathsep + os.environ["PATH"]
os.add_dll_directory(PF_DIR)

import pandas as pd
import powerfactory as pf
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Conexion a PowerFactory
# ---------------------------------------------------------------------------
os.system("taskkill /f /im PowerFactory.exe >nul 2>&1")

app = pf.GetApplication()
if app is None:
    raise RuntimeError(
        "No se pudo obtener la aplicacion de PowerFactory. "
        "Cierre PowerFactory si esta abierto y vuelva a ejecutar el script."
    )

app.Show()
user = app.GetCurrentUser()
prj  = app.ActivateProject("PMP_NOV25_OCT29_31102025(1)")
print(f"Proyecto activo: {prj}")

# ---------------------------------------------------------------------------
# Helpers de atributos
# ---------------------------------------------------------------------------
def ga(obj, attr):
    try:
        return obj.GetAttribute(attr) if obj is not None else None
    except Exception:
        return None

def ga_round(obj, attr, dec=2):
    val = ga(obj, attr)
    try:
        return round(val, dec) if val is not None else None
    except Exception:
        return val

def terminal_de(elm):
    cubicle = ga(elm, 'bus1')
    return ga(cubicle, 'cterm') if cubicle is not None else None

def barra_de(elm, lado):
    try:
        cub = elm.GetAttribute(lado)
        if cub is not None:
            term = cub.GetAttribute('cterm')
            if term is not None:
                return term
        return elm.GetAttribute(f'{lado}:cterm')
    except Exception:
        return None

def serv(obj):
    """Devuelve 'Si' o 'No' segun el atributo outserv."""
    return "Si" if ga(obj, 'outserv') == 0 else "No"

# ---------------------------------------------------------------------------
# Ejecutar flujo de carga BASE
# ---------------------------------------------------------------------------
ldf = app.GetFromStudyCase('ComLdf')
ldf.Execute()
print("Flujo de carga base ejecutado.")

# ---------------------------------------------------------------------------
# Mapa zona: construido una vez, usado en Barras y Cargas
# ---------------------------------------------------------------------------
_ZONE_ATTRS = ('zone', 'pZone', 'cpZone')

_bus_to_zona = {}
_zonas_pf = app.GetCalcRelevantObjects('*.ElmZone')
print(f"Zonas ElmZone encontradas: {len(_zonas_pf)}")
for _z in _zonas_pf:
    _zname = ga(_z, 'loc_name')
    try:
        _contenido = _z.GetContents('*.ElmTerm', 1)  # 1 = recursivo
        for _t in (_contenido or []):
            _bname = ga(_t, 'loc_name')
            if _bname:
                _bus_to_zona[_bname] = _zname
    except Exception:
        pass
print(f"Barras mapeadas a zona via ElmZone: {len(_bus_to_zona)}")

# ---------------------------------------------------------------------------
# 1. BARRAS
# ---------------------------------------------------------------------------
buses = app.GetCalcRelevantObjects('*.ElmTerm')
bus_data = []
for bus in buses:
    uknom = ga(bus, 'uknom')
    u_pu  = ga(bus, 'm:u')
    u_kv  = round(u_pu * uknom, 4) if u_pu is not None and uknom is not None else None
    bus_data.append({
        "Nombre":             ga(bus, 'loc_name'),
        "Tension nom. (kV)":  uknom,
        "Tension (pu)":       u_pu,
        "Tension (kV)":       u_kv,
        "Angulo (deg)":       ga_round(bus, 'm:phiu', 4),
        "P inyectada (MW)":   ga_round(bus, 'm:P', 4),
        "Q inyectada (Mvar)": ga_round(bus, 'm:Q', 4),
        "Zona":               next(
                                  (ga(ga(bus, a), 'loc_name') for a in _ZONE_ATTRS if ga(bus, a) is not None),
                                  _bus_to_zona.get(ga(bus, 'loc_name'))
                              ),
        "En servicio":        serv(bus),
    })
df_buses = pd.DataFrame(bus_data)
print(f"Barras: {len(df_buses)}")

# ---------------------------------------------------------------------------
# 2. LINEAS
# ---------------------------------------------------------------------------
lines = app.GetCalcRelevantObjects('*.ElmLne')
line_data = []
for ln in lines:
    typ     = ga(ln, 'typ_id')
    loading = ga(ln, 'c:loading')
    term1   = barra_de(ln, 'bus1')
    term2   = barra_de(ln, 'bus2')
    line_data.append({
        "Nombre":             ga(ln, 'loc_name'),
        "Nodo From":          ga(term1, 'loc_name'),
        "Nodo To":            ga(term2, 'loc_name'),
        "Distancia (km)":     ga_round(ln, 'dline', 2),
        "Tension nom. (kV)":  ga(typ, 'uline')  if typ else None,
        "Corriente nom. (A)": ga(typ, 'InomAC') if typ else None,
        "Carga (%)":          round(loading, 2) if loading is not None else None,
        "P from (MW)":        ga_round(ln, 'm:P:bus1', 4),
        "Q from (Mvar)":      ga_round(ln, 'm:Q:bus1', 4),
        "P to (MW)":          ga_round(ln, 'm:P:bus2', 4),
        "Q to (Mvar)":        ga_round(ln, 'm:Q:bus2', 4),
        "Perdidas P (MW)":    ga_round(ln, 'm:Plosses', 4),
        "En servicio":        serv(ln),
    })
df_lines = pd.DataFrame(line_data)
print(f"Lineas: {len(df_lines)}")

# ---------------------------------------------------------------------------
# 3. GENERADORES
# ---------------------------------------------------------------------------

# Códigos STI de unidades COBEE (Zongo + Taquesi).
# Para estas unidades se usa el atributo P_max (Active Power Rating Max).
# Para el resto se usa Pnom = Pr(rated).
_COBEE_STI = {
    "ZON", "TIQ", "BOT", "CUT", "SRO", "SAI", "CHU", "HAR", "CAH", "HUA",  # Zongo
    "CHJ", "YAN",                                                              # Taquesi
    "MIG", "ANG", "CHO", "CRB",                                               # Miguilla
}

def _es_cobee(loc_name):
    """True si el loc_name corresponde a una unidad COBEE."""
    import re as _re
    s = _re.sub(r'\(\d+\)$', '', str(loc_name).strip())
    for pref in ("sym_", "WT_", "PV-", "PV_", "sta_"):
        if s.lower().startswith(pref.lower()):
            s = s[len(pref):]
            break
    return s[:3].upper() in _COBEE_STI

gen_data = []
for cls in ('*.ElmSym', '*.ElmGenstat', '*.ElmPvsys', '*.ElmWind'):
    objs = app.GetCalcRelevantObjects(cls)
    if not objs:
        continue
    for gen in objs:
        term = terminal_de(gen)
        u_pu = ga(term, 'm:u') if term is not None else None
        try:
            clase_pf = gen.GetClassName()
        except Exception:
            clase_pf = "?"
        _loc = ga(gen, 'loc_name') or ""
        # Potencia nominal instalada: Pnom (ElmSym) o Pmax (ElmGenStat/ElmPvsys/ElmWind)
        _pnom = None
        for _attr in ("Pnom", "Pmax"):
            _v = ga(gen, _attr)
            if _v is not None:
                try:
                    _f = float(_v)
                    if _f > 0:
                        _pnom = round(_f, 4)
                        break
                except Exception:
                    pass
        # P_max (MW): para COBEE -> atributo P_max (Active Power Rating Max)
        #             para otros -> Pnom = Pr(rated); fallback a Pmax si Pnom no existe
        if _es_cobee(_loc):
            _p_max_col = ga_round(gen, 'P_max', 4)
        else:
            _p_max_col = None
            for _attr in ("Pnom", "Pmax"):
                _v = ga(gen, _attr)
                try:
                    _f = float(_v)
                    if _f > 0:
                        _p_max_col = round(_f, 4)
                        break
                except Exception:
                    pass
        gen_data.append({
            "Nombre":           _loc,
            "Clase PF":         clase_pf,
            "Barra conectada":  ga(term, 'loc_name'),
            "P nom. (MW)":      _pnom,
            "P_max (MW)":       _p_max_col,
            "Q nom. (Mvar)":    ga_round(gen, 'qgini', 4),
            "P result. (MW)":   ga_round(gen, 'm:P:bus1', 4),
            "Q result. (Mvar)": ga_round(gen, 'm:Q:bus1', 4),
            "Tension (pu)":     round(u_pu, 4) if u_pu is not None else None,
            "Cos phi":          ga_round(gen, 'cosini', 4),
            "En servicio":      serv(gen),
        })
df_gens = pd.DataFrame(gen_data)
print(f"Generadores: {len(df_gens)}")

# ---------------------------------------------------------------------------
# 4. CARGAS
# ---------------------------------------------------------------------------

def zona_de(elm):
    """Obtiene el nombre de la zona (ElmZone) a la que pertenece un elemento."""
    # 1) Atributo directo del elemento (varios nombres posibles)
    for attr in _ZONE_ATTRS:
        zone = ga(elm, attr)
        if zone is not None:
            name = ga(zone, 'loc_name')
            if name:
                return name
    # 2) Via la barra conectada (atributos directos)
    cub = ga(elm, 'bus1')
    if cub is not None:
        term = ga(cub, 'cterm')
        if term is not None:
            for attr in _ZONE_ATTRS:
                zone = ga(term, attr)
                if zone is not None:
                    name = ga(zone, 'loc_name')
                    if name:
                        return name
            # 3) Lookup desde mapa ElmZone
            bname = ga(term, 'loc_name')
            if bname and bname in _bus_to_zona:
                return _bus_to_zona[bname]
    return None

loads = app.GetCalcRelevantObjects('*.ElmLod')
load_data = []
for ld in loads:
    term = terminal_de(ld)
    load_data.append({
        "Nombre":           ga(ld, 'loc_name'),
        "Barra conectada":  ga(term, 'loc_name'),
        "Zona":             zona_de(ld),
        "P nom. (MW)":      ga_round(ld, 'plini', 4),
        "Q nom. (Mvar)":    ga_round(ld, 'qlini', 4),
        "P result. (MW)":   ga_round(ld, 'm:P:bus1', 4),
        "Q result. (Mvar)": ga_round(ld, 'm:Q:bus1', 4),
        "Cos phi":          ga_round(ld, 'coslini', 4),
        "En servicio":      serv(ld),
    })
df_loads = pd.DataFrame(load_data)
print(f"Cargas: {len(df_loads)}")

# ---------------------------------------------------------------------------
# 5. TRANSFORMADORES DE 2 DEVANADOS (ElmTr2)
# ---------------------------------------------------------------------------
tr2_data = []
for tr in app.GetCalcRelevantObjects('*.ElmTr2'):
    typ     = ga(tr, 'typ_id')
    loading = ga(tr, 'c:loading')
    term1   = barra_de(tr, 'bus1')
    term2   = barra_de(tr, 'bus2')
    tr2_data.append({
        "Nombre":               ga(tr, 'loc_name'),
        "Barra HV":             ga(term1, 'loc_name'),
        "Barra LV":             ga(term2, 'loc_name'),
        "Tension HV nom. (kV)": ga(typ, 'utrn_h') if typ else None,
        "Tension LV nom. (kV)": ga(typ, 'utrn_l') if typ else None,
        "Potencia nom. (MVA)":  ga(typ, 'strn')   if typ else None,
        "curmg":                ga(typ, 'curmg')  if typ else None,
        "Carga (%)":            round(loading, 2) if loading is not None else None,
        "P HV (MW)":            ga_round(tr, 'm:P:bus1', 4),
        "Q HV (Mvar)":          ga_round(tr, 'm:Q:bus1', 4),
        "P LV (MW)":            ga_round(tr, 'm:P:bus2', 4),
        "Q LV (Mvar)":          ga_round(tr, 'm:Q:bus2', 4),
        "Perdidas P (MW)":      ga_round(tr, 'm:Plosses', 4),
        "Tap pos.":             ga(tr, 'nntap'),
        "En servicio":          serv(tr),
    })
df_tr2 = pd.DataFrame(tr2_data)
print(f"Transformadores 2-dev: {len(df_tr2)}")

# ---------------------------------------------------------------------------
# 6. TRANSFORMADORES DE 3 DEVANADOS (ElmTr3)
# ---------------------------------------------------------------------------
tr3_data = []
for tr in app.GetCalcRelevantObjects('*.ElmTr3'):
    typ   = ga(tr, 'typ_id')
    term1 = barra_de(tr, 'bus1')
    term2 = barra_de(tr, 'bus2')
    term3 = barra_de(tr, 'bus3')
    tr3_data.append({
        "Nombre":               ga(tr, 'loc_name'),
        "Barra HV":             ga(term1, 'loc_name'),
        "Barra MV":             ga(term2, 'loc_name'),
        "Barra LV":             ga(term3, 'loc_name'),
        "Tension HV nom. (kV)": ga(typ, 'utrn_h') if typ else None,
        "Tension MV nom. (kV)": ga(typ, 'utrn_m') if typ else None,
        "Tension LV nom. (kV)": ga(typ, 'utrn_l') if typ else None,
        "Potencia nom. (MVA)":  ga(typ, 'strn')   if typ else None,
        "curmg":                ga(typ, 'curmg')  if typ else None,
        "P HV (MW)":            ga_round(tr, 'm:P:bus1', 4),
        "Q HV (Mvar)":          ga_round(tr, 'm:Q:bus1', 4),
        "P MV (MW)":            ga_round(tr, 'm:P:bus2', 4),
        "Q MV (Mvar)":          ga_round(tr, 'm:Q:bus2', 4),
        "P LV (MW)":            ga_round(tr, 'm:P:bus3', 4),
        "Q LV (Mvar)":          ga_round(tr, 'm:Q:bus3', 4),
        "En servicio":          serv(tr),
    })
df_tr3 = pd.DataFrame(tr3_data)
print(f"Transformadores 3-dev: {len(df_tr3)}")

# ===========================================================================
# FORMATO EXCEL
# ===========================================================================
def xfill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def xborder():
    s = Side(style='thin', color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

# Colores de encabezado por hoja
HDR = {
    'Barras':               "1F4E79",  # azul oscuro
    'Lineas':               "375623",  # verde oscuro
    'Generadores':          "7B2C2C",  # rojo oscuro
    'Cargas':               "7F6000",  # dorado oscuro
    'Transformadores_2dev': "4B1B6B",  # violeta oscuro
    'Transformadores_3dev': "1C4B4B",  # verde azulado
}

# Colores de filas por estado
C_SI = "C6EFCE"   # verde claro  — En servicio
C_NO = "FFC7CE"   # rojo claro   — Fuera de servicio

def formato_hoja(ws, nombre_hoja, col_serv):
    """Aplica encabezado, colorea filas y ajusta anchos."""
    # --- Encabezado ---
    hdr_color = HDR.get(nombre_hoja, "404040")
    for cell in ws[1]:
        cell.fill      = xfill(hdr_color)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = xborder()
    ws.row_dimensions[1].height = 30

    # --- Filas de datos ---
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        en_serv = str(row[col_serv - 1].value).strip().lower()
        color = C_SI if en_serv in ("si", "true", "1") else C_NO
        f = xfill(color)
        for cell in row:
            cell.fill   = f
            cell.border = xborder()

    # --- Ancho de columnas ---
    for col in ws.columns:
        ancho = max(len(str(c.value) or "") for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ancho + 3, 12), 40)

def agregar_leyenda(ws, fila_inicio, items):
    """Bloque de leyenda debajo de los datos."""
    fila_inicio += 1  # fila en blanco de separacion
    titulo = ws.cell(row=fila_inicio, column=1, value="LEYENDA")
    titulo.fill      = xfill("404040")
    titulo.font      = Font(bold=True, color="FFFFFF", size=10)
    titulo.alignment = Alignment(horizontal="center")
    titulo.border    = xborder()
    ws.cell(row=fila_inicio, column=2, value="Descripcion").font = Font(bold=True)
    ws.cell(row=fila_inicio, column=2).border = xborder()

    for offset, (color, texto) in enumerate(items, start=1):
        fila = fila_inicio + offset
        c1 = ws.cell(row=fila, column=1, value="")
        c1.fill = xfill(color); c1.border = xborder()
        c2 = ws.cell(row=fila, column=2, value=texto)
        c2.border = xborder()
        c2.alignment = Alignment(vertical="center")

LEY_SERV = [
    (C_SI, "En servicio"),
    (C_NO, "Fuera de servicio"),
]

# ---------------------------------------------------------------------------
# Exportar a Excel con formato
# ---------------------------------------------------------------------------
output_path = r"C:\Datos del CNDC\DATOS EXTRAIDOS DE DIGSILENT\DatosSINdigsilent.xlsx"

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df_buses.to_excel(writer, sheet_name='Barras',              index=False)
    df_lines.to_excel(writer, sheet_name='Lineas',              index=False)
    df_gens.to_excel( writer, sheet_name='Generadores',         index=False)
    df_loads.to_excel(writer, sheet_name='Cargas',              index=False)
    df_tr2.to_excel(  writer, sheet_name='Transformadores_2dev',index=False)
    if not df_tr3.empty:
        df_tr3.to_excel(writer, sheet_name='Transformadores_3dev', index=False)

    wb = writer.book

    # Barras
    ws = wb['Barras']
    col_serv = list(df_buses.columns).index("En servicio") + 1
    formato_hoja(ws, 'Barras', col_serv)
    agregar_leyenda(ws, ws.max_row, LEY_SERV)

    # Lineas
    ws = wb['Lineas']
    col_serv = list(df_lines.columns).index("En servicio") + 1
    formato_hoja(ws, 'Lineas', col_serv)
    agregar_leyenda(ws, ws.max_row, LEY_SERV)

    # Generadores
    ws = wb['Generadores']
    col_serv = list(df_gens.columns).index("En servicio") + 1
    formato_hoja(ws, 'Generadores', col_serv)
    agregar_leyenda(ws, ws.max_row, LEY_SERV)

    # Cargas
    ws = wb['Cargas']
    col_serv = list(df_loads.columns).index("En servicio") + 1
    formato_hoja(ws, 'Cargas', col_serv)
    agregar_leyenda(ws, ws.max_row, LEY_SERV)

    # Transformadores 2-dev
    ws = wb['Transformadores_2dev']
    col_serv = list(df_tr2.columns).index("En servicio") + 1
    formato_hoja(ws, 'Transformadores_2dev', col_serv)
    agregar_leyenda(ws, ws.max_row, LEY_SERV)

    # Transformadores 3-dev (si existe)
    if 'Transformadores_3dev' in wb.sheetnames:
        ws = wb['Transformadores_3dev']
        col_serv = list(df_tr3.columns).index("En servicio") + 1
        formato_hoja(ws, 'Transformadores_3dev', col_serv)
        agregar_leyenda(ws, ws.max_row, LEY_SERV)

print(f"\nDatos exportados a: {output_path}")
input("\nScript finalizado. PowerFactory sigue abierto. Presiona Enter para cerrar...")
os._exit(0)
